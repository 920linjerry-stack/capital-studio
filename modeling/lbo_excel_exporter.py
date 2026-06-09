"""V4.5 Professional LBO Workbook Packaging.

The workbook presents values produced by run_lbo() / build_lbo_attribution() /
assess_lbo_suitability() / the multi-tranche capital-structure helpers. It does
NOT recalculate IRR / MOIC / waterfall / covenants / attribution. The workbook
is values-only: no formulas, no external links, no circular references, and no
second hidden calculation truth. Python output is the single source of truth.

V4.5 upgrades the prior 4-sheet "results table" into a professional, reviewable
multi-sheet workbook (Cover, Model Map, Assumptions, Sources & Uses, Operating
Forecast, Debt Schedule, Covenant Check, Returns Summary, Returns Attribution,
Maturity Wall, Audit & Disclosures). It is packaging / presentation only and
adds no engine math, scenario store, PIK, refinancing, covenant cure, dividend
recap, or M&A.
"""

from __future__ import annotations

import math
from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modeling.lbo_calculator import AUDIT_DISCLOSURES
from modeling.lbo_attribution import MOIC_CONTRIBUTION_NOTE_EN, MOIC_CONTRIBUTION_NOTE_CN
from modeling.lbo_return_context import build_return_context
from modeling.lbo_scenarios import (
    DISCLOSURE_EXCEL_DETERMINISTIC,
    DISCLOSURE_EXCEL_MULTI_FACTOR,
    SCENARIO_LABELS,
)
from modeling.templates.goldman_dcf import (
    border_thin,
    fill_header,
    fill_section,
    font_formula,
    font_header,
    font_label,
    COLOR_HEADER_BG,
)


MODEL_VERSION = "LBO Engine Core v4.0"
WORKBOOK_VERSION = "V4.5 Professional Workbook Packaging"

SHEET_NAMES = [
    "Cover",
    "Model Map",
    "Assumptions",
    "Sources & Uses",
    "Operating Forecast",
    "Debt Schedule",
    "Covenant Check",
    "Returns Summary",
    "Returns Attribution",
    "Maturity Wall",
    "Audit & Disclosures",
]

# V4.6 optional sheet, inserted at position 3 (after Cover, Model Map) only when
# scenario comparison data is attached to the export payload.
SCENARIO_SUMMARY_SHEET = "Scenario Summary"
SENSITIVITY_SHEET = "Sensitivity Analysis"
_SCENARIO_INSERT_INDEX = 2

# ── Number formats (unified) ─────────────────────────────────────────────────
FMT_MONEY = "#,##0.0"   # amount / balance / EV / debt / equity
FMT_MULT1 = "0.0x"      # multiple
FMT_MULT2 = "0.00x"     # MOIC / DSCR / coverage
FMT_PCT = "0.0%"        # interest / margin / IRR / DSCR threshold
FMT_YEAR = "0"          # year

_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_CENTER = Alignment(horizontal="center", vertical="center")
_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Light blue fill marks user-editable inputs / assumptions.
_FILL_INPUT = PatternFill("solid", fgColor="DDEBF7")


def _yn(value: Any) -> Any:
    if value is True:
        return "Yes"
    if value is False:
        return "No"
    return value


def _title(ws, text: str, span: int = 2) -> None:
    ws.sheet_view.showGridLines = False
    c = ws.cell(1, 1, text)
    c.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    ws.row_dimensions[1].height = 22
    band = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    for col in range(2, span + 1):
        ws.cell(1, col).fill = band


def _section(ws, row: int, text: str, span: int = 5) -> None:
    c = ws.cell(row, 1, text)
    c.font = font_label()
    c.fill = fill_section()
    c.border = border_thin()
    for col in range(2, span + 1):
        cc = ws.cell(row, col)
        cc.fill = fill_section()
        cc.border = border_thin()


def _header_row(ws, row: int, labels: list[str], start_col: int = 1) -> None:
    for i, label in enumerate(labels):
        c = ws.cell(row, start_col + i, label)
        c.font = font_header()
        c.fill = fill_header()
        c.alignment = _CENTER
        c.border = border_thin()


def _cell(ws, row: int, col: int, value: Any, fmt: str | None = None,
          align: str | None = None, bold: bool = False, wrap: bool = False):
    c = ws.cell(row, col, value)
    c.border = border_thin()
    c.font = font_label() if bold else font_formula()
    if fmt is not None:
        c.number_format = fmt
    if wrap:
        c.alignment = _WRAP
    elif align == "right" or (
        align is None and isinstance(value, (int, float)) and not isinstance(value, bool)
    ):
        c.alignment = _RIGHT
    else:
        c.alignment = _LEFT
    return c


def _write_pairs(ws, start_row: int, rows: list[tuple]) -> int:
    """rows: (label, value[, fmt[, kind]]). kind='input' adds light-blue fill."""
    r = start_row
    for item in rows:
        label, value = item[0], item[1]
        fmt = item[2] if len(item) > 2 else None
        kind = item[3] if len(item) > 3 else "output"
        _cell(ws, r, 1, label, bold=True, align="left")
        c = _cell(ws, r, 2, value, fmt=fmt)
        if kind == "input":
            c.fill = _FILL_INPUT
        r += 1
    return r


def _set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# Lighter "footnote" styling so notes read as supporting captions, not large
# warning blocks. Soft pale yellow + small font, compact row heights.
_NOTE_FILL = PatternFill("solid", fgColor="FFF8E1")
_NOTE_FONT = Font(name="Arial", size=9, color="000000")
_NOTE_MIN_HEIGHT = 28.0
_NOTE_MAX_HEIGHT = 90.0


def _effective_len(text: str) -> int:
    """Count CJK characters as 2 so row-height estimation is realistic."""
    total = 0
    for ch in text:
        total += 2 if "　" <= ch <= "鿿" or "＀" <= ch <= "￯" else 1
    return total


def _estimate_height(text: Any, total_width: float) -> float:
    if not isinstance(text, str) or not text:
        return _NOTE_MIN_HEIGHT
    capacity = max(20.0, total_width)
    lines = max(1, math.ceil(_effective_len(text) / capacity))
    height = lines * 13.5 + 6
    return max(_NOTE_MIN_HEIGHT, min(_NOTE_MAX_HEIGHT, height))


def _note(ws, row: int, col: int, text: Any, end_col: int = 6) -> None:
    """Pale-yellow footnote/disclosure block: wrap, top-left, merged across
    columns. Row height is sized to the text so it stays readable but compact."""
    end_col = max(end_col, col)
    if end_col > col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    total_width = 0.0
    for c in range(col, end_col + 1):
        cell = ws.cell(row, c)
        cell.fill = _NOTE_FILL
        cell.border = border_thin()
        total_width += ws.column_dimensions[get_column_letter(c)].width or 8.43
    top = ws.cell(row, col, text)
    top.font = _NOTE_FONT
    top.alignment = _WRAP
    ws.row_dimensions[row].height = _estimate_height(text, total_width)


# ── Sheet builders ───────────────────────────────────────────────────────────


_SUPPORTED_CURRENCIES = ("USD", "HKD", "CNY")


def _ccy(result) -> str:
    """Defensive currency normalization for display/unit labels only."""
    code = str((result or {}).get("currency") or "").strip().upper()
    return code if code in _SUPPORTED_CURRENCIES else "USD"


def _append_units_note(ws, result) -> None:
    """Append a compact 'Amounts in {ccy} mm' note below existing content."""
    _note(ws, ws.max_row + 2, 1, f"Amounts in {_ccy(result)} mm.")


def _forecast_source_label(inputs) -> str:
    """User-facing forecast source. 'Flat Forecast' is never surfaced; legacy /
    missing modes are reported as the default Growth Forecast."""
    mode = str((inputs or {}).get("forecast_mode") or "").strip().lower()
    return "Manual Year-by-Year" if mode == "manual" else "Growth Forecast"


def _build_cover(ws, inputs, result, is_multi, *, formula_native: bool = False,
                 version: str | None = None) -> None:
    _title(ws, "LBO Model Workbook", span=2)
    _set_widths(ws, {"A": 30, "B": 40})
    ret = result["returns"]
    suitability = inputs.get("suitability") or {}
    rows = [
        ("Symbol", result.get("symbol")),
        ("Currency", _ccy(result)),
        ("Amounts In", f"{_ccy(result)} mm"),
        ("Forecast Source", _forecast_source_label(inputs)),
        ("Model Version", version or WORKBOOK_VERSION),
        ("Engine Output Mode", "multi_tranche" if is_multi else "single_tranche"),
        ("Generated", date.today().isoformat()),
        ("Holding Period (years)", result["transaction_summary"]["exit_year"], FMT_YEAR),
        ("Headline IRR", ret["irr"], FMT_PCT),
        ("Headline MOIC", ret["moic"], FMT_MULT2),
    ]
    if suitability.get("suitability"):
        rows.append(("Suitability Status", suitability.get("suitability")))
    row = _write_pairs(ws, 3, rows)
    row += 1
    _section(ws, row, "Disclosure", span=2)
    row += 1
    if formula_native:
        _note(ws, row, 1,
              "This workbook supports LBO framework analysis. It is not an "
              "investment or acquisition recommendation. The Python run_lbo() "
              "engine remains the calculation oracle for the exported case; the "
              "core transaction, operating, debt-schedule and returns sections "
              "include live Excel formulas — in both single-tranche and "
              "multi-tranche mode, including the tranche-level debt waterfall and "
              "cash sweep — so key assumptions can be edited and recalculated in "
              "Excel. Scenario Summary and Sensitivity Analysis remain "
              "Python-precomputed, oracle-backed grids (not live Excel Data "
              "Tables); V4.1.3 adds assumption mapping, an output bridge, live "
              "formula tie-out checks and an IC review structure over them. "
              "V4.1.4 is the final release-candidate formatting pass: editable "
              "inputs use the light-blue fill, cross-sheet references the green "
              "link font, headline outputs (MOIC / IRR) a peach fill, and the "
              "pass/fail tie-out regions their own check sections, with frozen "
              "panes and fit-to-width print setup on the core sheets.")
    else:
        _note(ws, row, 1,
              "This workbook supports LBO framework analysis. It is not an investment "
              "or acquisition recommendation. The Python engine is the single source "
              "of truth; Excel displays model output as values only (no formulas, no "
              "external links).")


def _build_model_map(ws, has_scenarios: bool = False, has_sensitivity: bool = False,
                     *, formula_native: bool = False) -> None:
    _title(ws, "Model Map", span=4)
    _set_widths(ws, {"A": 22, "B": 42, "C": 26, "D": 46})
    _header_row(ws, 3, ["Sheet", "Purpose", "Source", "Notes"])
    if formula_native:
        scenario_rows = [(
            SCENARIO_SUMMARY_SHEET,
            "Base / Upside / Downside case comparison",
            "Python build_lbo_scenarios + Excel formula checks",
            "Oracle-backed cases; V4.1.3 adds assumption map, output bridge, "
            "formula tie-out checks and IC review",
        )] if has_scenarios else []
        sensitivity_rows = [(
            SENSITIVITY_SHEET,
            "Two-factor return sensitivity tables",
            "Python build_lbo_sensitivity + Excel formula checks",
            "Deterministic precomputed grid; V4.1.3 adds axis map and formula "
            "dimension / best-base-worst checks",
        )] if has_sensitivity else []
    else:
        scenario_rows = [(
            SCENARIO_SUMMARY_SHEET,
            "Base / Upside / Downside case comparison",
            "Python build_lbo_scenarios",
            "Deterministic modeling cases, not forecasts",
        )] if has_scenarios else []
        sensitivity_rows = [(
            SENSITIVITY_SHEET,
            "Two-factor return sensitivity tables",
            "Python build_lbo_sensitivity",
            "Deterministic sensitivity, not probability-weighted",
        )] if has_sensitivity else []
    # V4.1.1.1: the formula-native workbook recalculates the core model in Excel,
    # so its Model Map describes the formula chain; the legacy values-only
    # workbook keeps the Python-precomputed / values-only wording.
    if formula_native:
        core_rows = [
            ("Assumptions", "Key transaction / operating assumptions", "User input / Assumptions anchors", "Editable inputs drive the formula chain"),
            ("Sources & Uses", "Deal funding sources vs uses", "Excel formulas / Assumptions anchors", "Live formulas; Python run_lbo() remains export oracle. Total Sources ties to Total Uses"),
            ("Operating Forecast", "Operating drivers feeding debt capacity", "Excel formulas / input drivers", "Live formulas (margins, tax shield, cash flow before debt service). Years across columns"),
            ("Debt Schedule", "Debt waterfall and cash sweep", "Excel formulas / Assumptions anchors / Python oracle tie-out", "Live Excel formulas for single-tranche and multi-tranche debt waterfall / cash sweep"),
            ("Covenant Check", "Covenant detection", "Excel formulas / Operating Forecast + Debt Schedule / Python oracle tie-out", "Detection only, not cure"),
            ("Returns Summary", "Headline returns and exit summary", "Excel formulas / named ranges", "Live formulas; Python run_lbo() remains export oracle for IRR / MOIC"),
            ("Returns Attribution", "Equity value bridge", "Python build_lbo_attribution", "MOIC contribution sums to MOIC - 1.0"),
            ("Maturity Wall", "Debt outstanding at exit by maturity", "Excel formulas / Debt Schedule + tranche assumptions / Python oracle tie-out", "Outstanding at exit, not face value"),
            ("Audit & Disclosures", "Source-of-truth and limitations", "Exporter metadata", "Not investment advice"),
        ]
    else:
        core_rows = [
            ("Assumptions", "Key transaction / operating assumptions", "User input / default builder", "Editable in UI, exported as values"),
            ("Sources & Uses", "Deal funding sources vs uses", "Python run_lbo", "Total Sources ties to Total Uses"),
            ("Operating Forecast", "Operating inputs feeding debt capacity", "User input / forecast", "Years across columns"),
            ("Debt Schedule", "Debt waterfall output", "Python run_lbo", "No Excel recalculation"),
            ("Covenant Check", "Covenant detection", "Python capital structure helper", "Detection only, not cure"),
            ("Returns Summary", "Headline returns and exit summary", "Python run_lbo", "IRR / MOIC source of truth"),
            ("Returns Attribution", "Equity value bridge", "Python build_lbo_attribution", "MOIC contribution sums to MOIC - 1.0"),
            ("Maturity Wall", "Debt outstanding at exit by maturity", "Python capital structure helper", "Outstanding at exit, not face value"),
            ("Audit & Disclosures", "Source-of-truth and limitations", "Exporter metadata", "Not investment advice"),
        ]
    rows = [
        ("Cover", "Workbook identity and headline returns", "Exporter metadata", "Read first"),
        ("Model Map", "What each sheet contains", "Exporter metadata", "This sheet"),
        *scenario_rows,
        *sensitivity_rows,
        *core_rows,
    ]
    r = 4
    for sheet, purpose, source, notes in rows:
        _cell(ws, r, 1, sheet, bold=True, align="left")
        _cell(ws, r, 2, purpose, wrap=True)
        _cell(ws, r, 3, source, wrap=True)
        _cell(ws, r, 4, notes, wrap=True)
        r += 1
    ws.freeze_panes = "A4"


def _build_assumptions(ws, inputs, result, is_multi, cap_summary) -> None:
    _title(ws, "Assumptions", span=5)
    _set_widths(ws, {"A": 30, "B": 18, "C": 22, "D": 16, "E": 46})
    ts = result["transaction_summary"]
    row = 3
    _section(ws, row, "Transaction Assumptions", span=5)
    row += 1
    row = _write_pairs(ws, row, [
        ("Entry EBITDA", ts["entry_ebitda"], FMT_MONEY, "input"),
        ("Entry Multiple", ts["entry_multiple"], FMT_MULT1, "input"),
        ("Entry EV", ts["entry_ev"], FMT_MONEY),
        ("Transaction Fees", ts["transaction_fees"], FMT_MONEY),
        ("Exit Multiple", ts["exit_multiple"], FMT_MULT1, "input"),
        ("Exit Year", ts["exit_year"], FMT_YEAR, "input"),
    ])
    row += 1
    _section(ws, row, "Debt / Capital Structure Assumptions", span=5)
    row += 1
    if is_multi:
        _header_row(ws, row, ["Tranche", "Opening Balance", "Interest Rate", "Maturity Yr", "Type"])
        row += 1
        for t in cap_summary.get("tranches") or []:
            _cell(ws, row, 1, t.get("name"), align="left")
            _cell(ws, row, 2, t.get("opening_balance"), fmt=FMT_MONEY)
            _cell(ws, row, 3, t.get("interest_rate"), fmt=FMT_PCT)
            _cell(ws, row, 4, t.get("maturity_year"), fmt=FMT_YEAR)
            _cell(ws, row, 5, t.get("type"), align="left")
            row += 1
        row = _write_pairs(ws, row, [
            ("Total Opening Debt", cap_summary.get("total_opening_debt"), FMT_MONEY),
            ("Cash to Balance Sheet", ts.get("cash_to_balance_sheet", 0.0), FMT_MONEY),
            ("Wtd Avg Cash Interest (Y1)", cap_summary.get("weighted_avg_cash_interest_rate_year_1"), FMT_PCT),
        ])
    else:
        debt = inputs.get("debt") or {}
        row = _write_pairs(ws, row, [
            ("Debt Amount", ts["debt_amount"], FMT_MONEY, "input"),
            ("Interest Rate", debt.get("interest_rate"), FMT_PCT, "input"),
            ("Mandatory Amortization %", debt.get("mandatory_amortization_pct"), FMT_PCT, "input"),
            ("Cash Sweep %", debt.get("cash_sweep_pct"), FMT_PCT, "input"),
            ("Cash to Balance Sheet", ts.get("cash_to_balance_sheet", 0.0), FMT_MONEY, "input"),
        ])
    row += 1

    row += 1
    _section(ws, row, "Tax Shield Assumptions", span=5)
    row += 1
    of = result.get("operating_forecast") or {}
    row = _write_pairs(ws, row, [
        ("Tax Rate", of.get("tax_rate", inputs.get("tax_rate", 0.25)), FMT_PCT, "input"),
        ("Tax Shield Enabled", _yn(of.get("tax_shield_enabled", inputs.get("tax_shield_enabled", True))), None, "input"),
    ])
    row += 1

    row = _assumptions_extra_sections(ws, inputs, row)
    ws.freeze_panes = "A3"


def _assumptions_extra_sections(ws, inputs, row: int) -> int:
    """Render the Default Builder Provenance and Suitability Status sections.

    Shared by the legacy values-only Assumptions sheet and the formula-native
    Assumptions sheet so both surface identical provenance / suitability
    disclosures from the same payload fields.
    """
    builder = inputs.get("default_builder") or inputs.get("v41_defaults_result") or {}
    provenance = builder.get("provenance") or {}
    if provenance:
        _section(ws, row, "Default Builder Provenance", span=5)
        row += 1
        _note(ws, row, 1,
              "Describes the system default structure only. If assumptions were "
              "edited before export, provenance may not reflect every manual change.")
        row += 1
        _header_row(ws, row, ["Assumption", "Value", "Source", "Confidence", "Rationale"])
        row += 1
        assumptions = builder.get("assumptions") or {}
        for key in [
            "entry_ebitda", "entry_multiple", "exit_multiple", "exit_year",
            "transaction_fees_pct_ev", "entry_structure", "leverage_multiple",
            "debt_amount", "interest_rate", "mandatory_amortization_pct",
            "cash_sweep_pct", "cash_to_balance_sheet", "tax_shield_serviceability",
        ]:
            meta = provenance.get(key)
            if not meta:
                continue
            _cell(ws, row, 1, key, align="left")
            _cell(ws, row, 2, assumptions.get(key, ""), align="left")
            _cell(ws, row, 3, meta.get("source"), align="left")
            _cell(ws, row, 4, meta.get("confidence"), align="left")
            _cell(ws, row, 5, meta.get("rationale_cn"), wrap=True)
            row += 1
        row += 1

    suitability = inputs.get("suitability") or {}
    if suitability:
        _section(ws, row, "Suitability Status", span=5)
        row += 1
        row = _write_pairs(ws, row, [
            ("Suitability", suitability.get("suitability")),
            ("Veto Triggered", _yn(suitability.get("veto_triggered"))),
            ("Score", suitability.get("score")),
        ])
        reasons = (suitability.get("veto_reasons") or []) + (suitability.get("penalty_reasons") or [])
        for entry in reasons[:5]:
            _cell(ws, row, 1, entry.get("code", ""), align="left")
            _note(ws, row, 2, entry.get("message_cn") or entry.get("message") or "", end_col=5)
            row += 1
    return row


def _build_sources_uses(ws, inputs, result, is_multi, cap_summary) -> None:
    _title(ws, "Sources & Uses", span=2)
    _set_widths(ws, {"A": 32, "B": 20})
    ts = result["transaction_summary"]
    entry_ev = ts["entry_ev"]
    fees = ts["transaction_fees"]
    total_uses = ts["total_uses"]
    debt_amount = ts["debt_amount"]
    sponsor_equity = ts["sponsor_equity"]
    entry_ebitda = ts["entry_ebitda"]

    # Single source of truth for cash: the engine's transaction_summary. This is
    # the same figure that flows into Total Uses, so Sources & Uses ties out.
    cash_to_bs = ts.get("cash_to_balance_sheet", 0.0)

    row = 3
    _section(ws, row, "Uses", span=2)
    row += 1
    uses_rows = [
        ("Entry EV", entry_ev, FMT_MONEY),
        ("Transaction Fees", fees, FMT_MONEY),
        ("Cash to Balance Sheet", cash_to_bs, FMT_MONEY),
        ("Total Uses", total_uses, FMT_MONEY),
    ]
    row = _write_pairs(ws, row, uses_rows)
    row += 1

    _section(ws, row, "Sources", span=2)
    row += 1
    if is_multi:
        for t in cap_summary.get("tranches") or []:
            _cell(ws, row, 1, t.get("name"), bold=True, align="left")
            _cell(ws, row, 2, t.get("opening_balance"), fmt=FMT_MONEY)
            row += 1
        row = _write_pairs(ws, row, [
            ("Total Opening Debt", cap_summary.get("total_opening_debt"), FMT_MONEY),
            ("Sponsor Equity", sponsor_equity, FMT_MONEY),
        ])
    else:
        row = _write_pairs(ws, row, [
            ("Debt Amount", debt_amount, FMT_MONEY),
            ("Sponsor Equity", sponsor_equity, FMT_MONEY),
        ])
    row += 1

    total_sources = debt_amount + sponsor_equity
    _section(ws, row, "Ratios", span=2)
    row += 1
    row = _write_pairs(ws, row, [
        ("Debt % of Uses", (debt_amount / total_uses) if total_uses else None, FMT_PCT),
        ("Sponsor Equity % of Uses", (sponsor_equity / total_uses) if total_uses else None, FMT_PCT),
        ("Entry Leverage (Debt / EBITDA)", (debt_amount / entry_ebitda) if entry_ebitda else None, FMT_MULT1),
    ])
    row += 1

    _section(ws, row, "Check", span=2)
    row += 1
    sources_less_uses = total_sources - total_uses
    tie = abs(sources_less_uses) < 0.01
    row = _write_pairs(ws, row, [
        ("Total Sources", total_sources, FMT_MONEY),
        ("Total Uses", total_uses, FMT_MONEY),
        ("Sources - Uses", sources_less_uses, FMT_MONEY),
        ("Sources = Uses", _yn(tie)),
    ])
    row += 1
    _note(ws, row, 1,
          "Sources and uses are displayed from exported values; the check row "
          "compares Total Sources and Total Uses without an Excel formula.")
    row += 1
    _note(ws, row, 1,
          "Existing debt repayment, rollover equity and seller notes are deferred scope.")
    ws.freeze_panes = "A3"


def _build_operating_forecast(ws, result, is_multi) -> None:
    _title(ws, "Operating Forecast", span=9)
    of = result["operating_forecast"]
    years = of["years"]
    ds = result["debt_schedule"]
    widths = {"A": 26}
    for i in range(len(years)):
        widths[get_column_letter(2 + i)] = 14
    _set_widths(ws, widths)
    _header_row(ws, 3, ["Metric"] + [f"Y{y}" for y in years])
    metric_rows = [
        ("Revenue", of.get("revenue")),
        ("EBITDA", of.get("ebitda")),
        ("Gross Cash Taxes", of.get("gross_cash_taxes") or of.get("cash_taxes")),
        ("Tax Shield", of.get("tax_shield")),
        ("Levered Cash Taxes", of.get("levered_cash_taxes") or of.get("cash_taxes")),
        ("CapEx", of.get("capex")),
        ("Change in NWC", of.get("change_in_nwc")),
    ]
    if is_multi:
        metric_rows.append(("Cash Flow Before Debt Service", [r.get("cash_flow_before_debt_service") for r in ds]))
    else:
        metric_rows.append(("Cash Flow Before Debt Service", [r.get("cash_flow_before_debt_service") for r in ds]))
    r = 4
    for label, values in metric_rows:
        _cell(ws, r, 1, label, bold=True, align="left")
        for c, v in enumerate(values or [], 2):
            _cell(ws, r, c, v, fmt=FMT_MONEY)
        r += 1
    r += 1
    _note(ws, r, 1, "Operating drivers are shown as values; debt-capacity outputs "
          "are computed by the Python engine, not by Excel.", end_col=2 + len(years) - 1)
    ws.freeze_panes = "B4"


def _build_debt_schedule(ws, result, is_multi) -> None:
    _title(ws, "Debt Schedule", span=12 if is_multi else 9)
    if is_multi:
        _set_widths(ws, {
            "A": 10, "B": 18, "C": 16, "D": 16, "E": 16, "F": 16,
            "G": 18, "H": 16, "I": 16, "J": 16, "K": 16, "L": 18,
        })
        _section(ws, 3, "Annual Totals", span=12)
        year_headers = [
            "Year", "Total Beg Debt", "Cash Flow Before Debt Service", "Cash Interest",
            "Gross Cash Taxes", "Tax Shield", "Levered Cash Taxes",
            "Mandatory Amortization", "Revolver Draw", "Cash After Interest & Mandatory Amort.",
            "Optional Repayment", "Ending Cash",
            "Total End Debt", "Net Debt", "Debt Service Failure",
        ]
        year_fmts = [FMT_YEAR, FMT_MONEY, FMT_MONEY, FMT_MONEY, FMT_MONEY,
                     FMT_MONEY, FMT_MONEY, FMT_MONEY, FMT_MONEY, FMT_MONEY,
                     FMT_MONEY, FMT_MONEY, FMT_MONEY, FMT_MONEY,
                     FMT_MONEY, None]
        _header_row(ws, 4, year_headers)
        r = 5
        for row in result["debt_schedule"]:
            net_debt = row["total_ending_debt"] - row["ending_cash_balance"]
            values = [
                row["year"], row["total_beginning_debt"],
                row.get("cash_flow_before_debt_service", row.get("cash_before_debt_service")),
                row["total_cash_interest"], row.get("gross_cash_taxes"), row.get("tax_shield"),
                row.get("levered_cash_taxes"), row["total_mandatory_amortization"],
                row["revolver_draw"],
                row.get("cash_after_interest_and_mandatory_amortization", row.get("cash_after_required_debt_service")),
                row["total_optional_repayment"],
                row["ending_cash_balance"], row["total_ending_debt"], net_debt,
                _yn(row["debt_service_failure"]),
            ]
            for c, (value, fmt) in enumerate(zip(values, year_fmts), 1):
                _cell(ws, r, c, value, fmt=fmt)
            r += 1
        r += 1
        _section(ws, r, "Tranche Detail", span=9)
        r += 1
        detail_headers = [
            "Year", "Tranche", "Beginning Balance", "Draw", "Cash Interest",
            "Mandatory Amort.", "Optional Repayment", "Ending Balance", "Available Capacity",
        ]
        detail_fmts = [FMT_YEAR, None, FMT_MONEY, FMT_MONEY, FMT_MONEY,
                       FMT_MONEY, FMT_MONEY, FMT_MONEY, FMT_MONEY]
        _header_row(ws, r, detail_headers)
        r += 1
        for row in result["debt_schedule"]:
            for tr in row["tranches"]:
                values = [
                    row["year"], tr.get("name") or tr["id"], tr["beginning_balance"],
                    tr["draw"], tr["cash_interest"], tr["mandatory_amortization"],
                    tr["optional_repayment"], tr["ending_balance"], tr.get("available_capacity"),
                ]
                for c, (value, fmt) in enumerate(zip(values, detail_fmts), 1):
                    _cell(ws, r, c, value, fmt=fmt)
                r += 1
    else:
        _set_widths(ws, {
            "A": 10, "B": 16, "C": 14, "D": 18, "E": 18,
            "F": 18, "G": 16, "H": 14, "I": 18,
        })
        _section(ws, 3, "Annual Debt Service", span=9)
        headers = [
            "Year", "Beginning Debt", "Cash Flow Before Debt Service", "Cash Interest",
            "Gross Cash Taxes", "Tax Shield", "Levered Cash Taxes",
            "Mandatory Amortization", "Cash After Interest & Mandatory Amort.",
            "Optional Repayment",
            "Ending Debt", "Debt Service Failure",
        ]
        fmts = [FMT_YEAR, FMT_MONEY, FMT_MONEY, FMT_MONEY, FMT_MONEY,
                FMT_MONEY, FMT_MONEY, FMT_MONEY, FMT_MONEY, FMT_MONEY,
                FMT_MONEY, None]
        _header_row(ws, 4, headers)
        for r, row in enumerate(result["debt_schedule"], 5):
            values = [
                row["year"], row["beginning_debt"],
                row.get("cash_flow_before_debt_service", row["cash_available_for_debt"] + row["cash_interest"]),
                row["cash_interest"], row.get("gross_cash_taxes"), row.get("tax_shield"),
                row.get("levered_cash_taxes"), row["mandatory_amortization"],
                row.get("cash_after_interest_and_mandatory_amortization", row["fcf_available_for_sweep"]),
                row["optional_repayment"],
                row["ending_debt"], _yn(row["debt_service_failure"]),
            ]
            for c, (value, fmt) in enumerate(zip(values, fmts), 1):
                _cell(ws, r, c, value, fmt=fmt)
    ws.freeze_panes = "A5"


def _build_covenant_check(ws, result, is_multi) -> None:
    _title(ws, "Covenant Check", span=12)
    _set_widths(ws, {
        "A": 10, "B": 16, "C": 16, "D": 14, "E": 14, "F": 18, "G": 16,
        "H": 14, "I": 16, "J": 16, "K": 16, "L": 18,
    })
    if not is_multi:
        _section(ws, 3, "Not Applicable", span=4)
        _note(ws, 4, 1,
              "Covenant checks apply to the multi-tranche capital structure. The "
              "single-tranche simplified mode does not model covenants.")
        ws.freeze_panes = "A3"
        return

    cov = result.get("covenant_summary") or {}
    ds = result["debt_schedule"]
    row = 3
    row = _write_pairs(ws, row, [
        ("Status", cov.get("status")),
        ("Breach Years", ", ".join(str(y) for y in (cov.get("breach_years") or [])) or "(none)"),
    ])
    row += 1
    headers = [
        "Year", "EBITDA", "Ending Debt", "Ending Cash", "Net Debt",
        "Net Debt / EBITDA", "Max Leverage", "Leverage Breach", "Cash Interest",
        "Interest Coverage", "Min Coverage", "Coverage Breach",
    ]
    _header_row(ws, row, headers)
    row += 1
    by_year = {r["year"]: r for r in ds}
    for chk in cov.get("checks") or []:
        dsr = by_year.get(chk.get("year"), {})
        nde = "Net Cash" if chk.get("is_net_cash") else chk.get("net_debt_ebitda")
        ic = chk.get("interest_coverage")
        _cell(ws, row, 1, chk.get("year"), fmt=FMT_YEAR)
        _cell(ws, row, 2, chk.get("ebitda"), fmt=FMT_MONEY)
        _cell(ws, row, 3, dsr.get("total_ending_debt"), fmt=FMT_MONEY)
        _cell(ws, row, 4, dsr.get("ending_cash_balance"), fmt=FMT_MONEY)
        _cell(ws, row, 5, chk.get("net_debt"), fmt=FMT_MONEY)
        _cell(ws, row, 6, nde, fmt=FMT_MULT2 if isinstance(nde, (int, float)) else None)
        _cell(ws, row, 7, chk.get("max_net_debt_ebitda"), fmt=FMT_MULT2)
        _cell(ws, row, 8, _yn(chk.get("leverage_breach")), align="left")
        _cell(ws, row, 9, dsr.get("total_cash_interest"), fmt=FMT_MONEY)
        _cell(ws, row, 10, "n/a" if ic is None else ic, fmt=FMT_MULT2 if isinstance(ic, (int, float)) else None)
        _cell(ws, row, 11, chk.get("min_interest_coverage"), fmt=FMT_MULT2)
        _cell(ws, row, 12, _yn(chk.get("interest_coverage_breach")), align="left")
        row += 1
    row += 1
    _note(ws, row, 1,
          "Net Debt / Forecast EBITDA uses forecast-year EBITDA. Covenant check is detection-only: it flags leverage / coverage breaches "
          "against thresholds. No cure, reset, or refinancing is modeled, and it is "
          "not a deal recommendation. Net cash years show 'Net Cash'; non-positive "
          "interest shows 'n/a' rather than infinity.", end_col=8)
    ws.freeze_panes = "A6"


def _build_returns_summary(ws, result, is_multi, inputs=None) -> None:
    _title(ws, "Returns Summary", span=2)
    _set_widths(ws, {"A": 30, "B": 20})
    ret = result["returns"]
    ex = result["exit"]
    row = 3
    _section(ws, row, "Headline Returns", span=2)
    row += 1
    headline = [
        ("IRR", ret["irr"], FMT_PCT),
        ("MOIC", ret["moic"], FMT_MULT2),
        ("Sponsor Equity", ret["sponsor_equity"], FMT_MONEY),
        ("Exit Equity Value", ret["exit_equity_value"], FMT_MONEY),
        ("Debt Paydown", ret["debt_paydown"], FMT_MONEY),
        ("Remaining Debt", ret["remaining_debt"], FMT_MONEY),
    ]
    if is_multi:
        headline.append(("Ending Cash Balance", ret.get("ending_cash_balance"), FMT_MONEY))
    row = _write_pairs(ws, row, headline)
    row += 1
    _section(ws, row, "Exit Equity Bridge", span=2)
    row += 1
    exit_rows = [
        ("Exit EBITDA", ex.get("exit_ebitda"), FMT_MONEY),
        ("Exit Multiple", ex.get("exit_multiple"), FMT_MULT1),
        ("Exit Enterprise Value", ex.get("exit_ev"), FMT_MONEY),
        ("Less: Remaining Debt", ex.get("remaining_debt"), FMT_MONEY),
        ("Plus: Ending Cash", ex.get("ending_cash_balance", 0.0), FMT_MONEY),
    ]
    exit_rows.append(("Exit Equity Value", ex.get("exit_equity_value"), FMT_MONEY))
    row = _write_pairs(ws, row, exit_rows)
    row += 1

    # V4.7 lightweight return context: a single compact badge line, only if
    # triggered. It is context, not a warning and not a recommendation.
    context = build_return_context(result, (inputs or {}).get("attribution"))
    if context.get("status") == "ok":
        _cell(ws, row, 1, "Return Context", bold=True, align="left")
        _note(ws, row, 2, context.get("badge_en"))
        row += 2

    _note(ws, row, 1,
          "IRR / MOIC and the value bridge come from the Python run_lbo engine. "
          "Excel performs no secondary calculation. Attribution detail is on the "
          "Returns Attribution sheet.")
    ws.freeze_panes = "A3"


def _build_attribution(ws, inputs) -> None:
    _title(ws, "Returns Attribution", span=5)
    _set_widths(ws, {"A": 30, "B": 16, "C": 18, "D": 12, "E": 60})
    attribution = inputs.get("attribution") or {}
    if attribution.get("status") != "ok":
        _section(ws, 3, "Attribution Unavailable", span=5)
        _note(ws, 4, 1,
              "Returns attribution is unavailable for this run. It requires a valid "
              "LBO result with positive sponsor equity and exit value.")
        ws.freeze_panes = "A3"
        return

    row = 3
    _section(ws, row, "Equity Value Bridge", span=5)
    row += 1
    _header_row(ws, row, ["Component", "Value", "MOIC Contribution", "Direction", "Rationale"])
    row += 1
    for comp in attribution.get("components") or []:
        _cell(ws, row, 1, comp.get("label_en") or comp.get("label_cn"), align="left")
        _cell(ws, row, 2, comp.get("value"), fmt=FMT_MONEY)
        _cell(ws, row, 3, comp.get("moic_contribution"), fmt="0.0000x")
        _cell(ws, row, 4, comp.get("direction"), align="left")
        _cell(ws, row, 5, comp.get("rationale_en") or comp.get("rationale_cn"), wrap=True)
        row += 1
    notes = attribution.get("notes") or {}
    _cell(ws, row, 1, "Note", bold=True, align="left")
    _note(ws, row, 2, notes.get("moic_contribution_note_en") or MOIC_CONTRIBUTION_NOTE_EN)
    row += 1
    _note(ws, row, 2, notes.get("moic_contribution_note_cn") or MOIC_CONTRIBUTION_NOTE_CN)
    row += 2

    tie = attribution.get("tie_out") or {}
    _section(ws, row, "Tie-out", span=2)
    row += 1
    row = _write_pairs(ws, row, [
        ("Attribution Method", attribution.get("method")),
        ("Target Equity Value Creation", tie.get("target_equity_value_creation"), FMT_MONEY),
        ("Component Sum Before Residual", tie.get("component_sum_before_residual"), FMT_MONEY),
        ("Residual", tie.get("residual"), FMT_MONEY),
        ("Mathematical Tie-out Pass", _yn(tie.get("mathematical_tie_out_pass"))),
        ("Directional Bridge Pass", _yn(tie.get("directional_bridge_pass"))),
    ])
    row += 1
    _section(ws, row, "Disclosure", span=2)
    row += 1
    _note(ws, row, 1, notes.get("disclosure_en") or
          "Simplified equity value bridge. It explains model return sources "
          "directionally, does not split component-level IRR, and is not an "
          "investment or acquisition recommendation.")
    ws.freeze_panes = "A3"


def _build_maturity_wall(ws, result, is_multi) -> None:
    _title(ws, "Maturity Wall", span=5)
    _set_widths(ws, {"A": 16, "B": 20, "C": 36, "D": 18, "E": 30})
    if not is_multi:
        _section(ws, 3, "Not Applicable", span=4)
        _note(ws, 4, 1,
              "The maturity wall applies to multi-tranche capital structures. The "
              "single-tranche simplified mode models a single debt balance.")
        ws.freeze_panes = "A3"
        return

    exit_year = result["transaction_summary"]["exit_year"]
    row = 3
    _header_row(ws, row, ["Maturity Year", "Outstanding at Exit", "Tranches", "Within Hold Period?", "Notes"])
    row += 1
    for bucket in result.get("maturity_wall") or []:
        within = bucket.get("year") is not None and bucket["year"] <= exit_year
        _cell(ws, row, 1, bucket.get("year"), fmt=FMT_YEAR)
        _cell(ws, row, 2, bucket.get("maturing_debt"), fmt=FMT_MONEY)
        _cell(ws, row, 3, ", ".join(bucket.get("tranches") or []), align="left")
        _cell(ws, row, 4, _yn(within), align="left")
        _cell(ws, row, 5, "Refinancing not modeled" if within else "", wrap=True)
        row += 1
    row += 1
    _note(ws, row, 1,
          "Maturity wall shows debt outstanding at exit by maturity year, not "
          "original face value. Display / audit only; refinancing and maturity "
          "default are not modeled.")
    ws.freeze_panes = "A4"


def _build_audit(ws, inputs, result, is_multi, cap_summary, *,
                 formula_native: bool = False) -> None:
    _title(ws, "Audit & Disclosures", span=6)
    _set_widths(ws, {"A": 28, "B": 22, "C": 16, "D": 16, "E": 16, "F": 18})
    row = 3

    _section(ws, row, "Source of Truth", span=6)
    row += 1
    for text in (result.get("audit", {}).get("disclosures") or AUDIT_DISCLOSURES):
        _note(ws, row, 1, text)
        row += 1
    if formula_native:
        for text in [
            "Python run_lbo() remains the calculation oracle for the exported "
            "initial case; the headline IRR / MOIC shown on the Cover are its output.",
            "The core transaction, operating, debt-schedule and returns sections "
            "(Assumptions, Sources & Uses, Operating Forecast, Debt Schedule, "
            "Returns Summary / Exit Bridge) include live Excel formulas so the "
            "exported case can be re-checked and edited inside Excel.",
            "V4.1.1: the multi-tranche debt waterfall is now formula-driven — "
            "tranche-level beginning balance, beginning-balance interest, "
            "mandatory amortization, revolver draw, priority-ordered cash sweep, "
            "ending balance, total / net debt and ending cash are live Excel "
            "formulas, and the Returns / Exit Bridge references the formula-driven "
            "total ending debt and ending cash. It is no longer Python-precomputed.",
            "Formulas use only beginning-balance interest, so there is no Excel "
            "circular reference and no external link.",
            "V4.1.2: Covenant Check and Maturity Wall are now formula-linked to the "
            "Debt Schedule and Operating Forecast — leverage / coverage ratios, "
            "pass/fail flags, tranche ending balances and share of total debt "
            "recalculate when debt or operating assumptions change. Covenant "
            "thresholds and tranche maturity years are editable inputs and pass/fail "
            "uses Excel IF().",
            "V4.1.2: the Returns Attribution sheet adds live formula tie-out checks "
            "(Sources = Uses, Sponsor Equity, Exit Equity bridge, MOIC, debt paydown, "
            "ending cash, and an attribution component-sum-vs-bridge check) that "
            "reference the workbook named-range chain, so a broken tie-out surfaces as "
            "a CHECK flag rather than a silent mismatch.",
            "Input cells use the light-blue fill; cross-sheet links use the link font; "
            "formula and check cells are computed — the Cover, Model Map and these "
            "notes mark the input / output / check regions.",
            "Scenario Summary and Sensitivity Analysis remain Python-precomputed "
            "oracle-backed presentation grids; they are not live Excel Data Tables.",
            "V4.1.3: the Scenario Summary adds a scenario assumption map, an output "
            "bridge and live formula tie-out checks (MOIC x sponsor = exit equity, "
            "component sum = equity value creation, base case = main exported case, "
            "debt paydown, initial cash funding) plus an IC review block; the "
            "Sensitivity Analysis adds an axis map and live formula checks (grid "
            "dimensions, best / base / worst IRR). The underlying scenario / "
            "sensitivity grids stay Python run_lbo() oracle output.",
            "V4.1.4 (final release candidate): a formatting / disclosure hardening "
            "pass only — no engine math, run_lbo() truth, debt-schedule, covenant, "
            "attribution or scenario / sensitivity grid logic was changed. It marks "
            "the four workbook regions consistently (light-blue input fill, green "
            "cross-sheet link font, peach headline-output fill, demarcated pass/fail "
            "check sections), and adds frozen panes and fit-to-width print setup so "
            "the workbook prints as a deliverable PE / IB model. The scenario and "
            "sensitivity grids remain Python-precomputed, oracle-backed grids — they "
            "are not live Excel Data Tables. No network, LLM or external runtime "
            "dependency is introduced; the export stays deterministic.",
        ]:
            _note(ws, row, 1, text)
            row += 1
    else:
        _note(ws, row, 1, "Python run_lbo is the IRR / MOIC source of truth. Excel "
              "displays values only — no circular references, no external links.")
    row += 2

    _section(ws, row, "Calculation Discipline", span=6)
    row += 1
    _note(ws, row, 1, "Interest is based on the beginning debt balance; there is no "
          "iterative circular calculation.")
    row += 2

    _section(ws, row, "Tax / Cash Bridge", span=6)
    row += 1
    for text in [
        "Simplified tax shield uses cash interest x tax rate. No full tax schedule, deferred tax or three-statement engine is modeled.",
        "Single-tranche mode does not model a cash balance bridge.",
        "Multi-tranche mode captures the ending cash balance.",
    ]:
        _note(ws, row, 1, text)
        row += 1
    if is_multi and cap_summary.get("single_vs_multi_cash_bridge_note_cn"):
        _note(ws, row, 1, cap_summary.get("single_vs_multi_cash_bridge_note_cn"))
        row += 1
    row += 1

    suitability = inputs.get("suitability") or {}
    _section(ws, row, "Suitability Disclosure", span=6)
    row += 1
    _note(ws, row, 1, (suitability.get("disclosure_en") if suitability else None) or
          "Suitability assesses whether LBO is an appropriate modeling framework. "
          "It is not an investment or acquisition recommendation.")
    row += 2

    _section(ws, row, "Attribution Disclosure", span=6)
    row += 1
    _note(ws, row, 1, "Simplified equity value bridge; it does not split component "
          "IRR and is not an investment or acquisition recommendation.")
    row += 2

    _section(ws, row, "Covenant Disclosure", span=6)
    row += 1
    _note(ws, row, 1, "Covenant detection only; no cure, reset, or refinancing is "
          "modeled.")
    row += 2

    _section(ws, row, "Maturity Wall Disclosure", span=6)
    row += 1
    _note(ws, row, 1, "Maturity wall shows debt outstanding at exit, not original "
          "face value.")
    row += 2

    _section(ws, row, "Deferred Scope", span=6)
    row += 1
    _note(ws, row, 1, "Not implemented in the current version (deferred future "
          "scope, not permanent prohibitions): PIK, dividend recap, refinancing, "
          "covenant cure, scenario store, add-on M&A, engine-driven three-statement "
          "LBO with balance-sheet balancing, and further professional model features "
          "beyond V4.5.")
    ws.freeze_panes = "A3"


def _scenario_comparison(inputs: dict[str, Any]) -> dict | None:
    """Return the scenario comparison block if usable scenario data is attached
    to the export payload, else None. Accepts either the full build_lbo_scenarios
    result under inputs['scenarios'] or a bare comparison dict."""
    data = inputs.get("scenarios")
    if not isinstance(data, dict):
        return None
    comparison = data.get("comparison")
    if isinstance(comparison, dict) and comparison.get("rows"):
        return data
    if data.get("rows"):  # bare comparison dict
        return {"comparison": data}
    return None


def _scenario_cell(value: Any, fmt: str) -> tuple[Any, str | None]:
    """Map a comparison cell value + canonical format to (excel_value, number_format)."""
    if isinstance(value, bool):
        return _yn(value), None
    if isinstance(value, str):
        return value, None
    if value is None:
        return "n/a", None
    if fmt == "percent":
        return value, FMT_PCT
    if fmt == "multiple":
        return value, FMT_MULT2
    if fmt == "amount":
        return value, FMT_MONEY
    if fmt == "boolean":
        return _yn(value), None
    return value, None


def _build_scenario_summary(ws, scenario_data: dict) -> None:
    _title(ws, "Scenario Summary", span=5)
    _set_widths(ws, {"A": 34, "B": 18, "C": 18, "D": 18, "E": 50})
    comparison = scenario_data.get("comparison") or {}
    rows = comparison.get("rows") or []
    statuses = comparison.get("scenario_statuses") or {}

    _header_row(ws, 3, [
        "Metric",
        SCENARIO_LABELS["base"],
        SCENARIO_LABELS["upside"],
        SCENARIO_LABELS["downside"],
        "Notes",
    ])

    r = 4
    for row in rows:
        fmt = row.get("format", "text")
        _cell(ws, r, 1, row.get("metric"), bold=True, align="left")
        for col, key in ((2, "base"), (3, "upside"), (4, "downside")):
            value, number_format = _scenario_cell(row.get(key), fmt)
            _cell(ws, r, col, value, fmt=number_format, align="right")
        r += 1

    # Per-scenario availability (single reason per unavailable scenario).
    r += 1
    _section(ws, r, "Scenario Availability", span=5)
    r += 1
    for key in ("base", "upside", "downside"):
        status = statuses.get(key) or {}
        state = status.get("status", "ok")
        _cell(ws, r, 1, SCENARIO_LABELS[key], bold=True, align="left")
        if state == "ok":
            _cell(ws, r, 2, "Available")
        else:
            _cell(ws, r, 2, "Unavailable")
            _note(ws, r, 3, status.get("reason") or "Unavailable.", end_col=5)
        r += 1

    # Required disclosures.
    r += 1
    _section(ws, r, "Disclosures", span=5)
    r += 1
    _note(ws, r, 1, DISCLOSURE_EXCEL_DETERMINISTIC, end_col=5)
    r += 1
    _note(ws, r, 1, DISCLOSURE_EXCEL_MULTI_FACTOR, end_col=5)
    r += 1
    _note(ws, r, 1,
          "Scenario labels are Base / Upside / Downside modeling cases. They are "
          "modeling assumptions, not market long / short views, and do not "
          "constitute a recommendation to transact.", end_col=5)
    ws.freeze_panes = "A4"


def _sensitivity_data(inputs: dict[str, Any]) -> dict | None:
    data = inputs.get("sensitivity")
    if isinstance(data, dict) and data.get("status") == "ok" and isinstance(data.get("grids"), dict):
        return data
    return None


def _sens_value(cell: dict[str, Any], metric: str) -> tuple[Any, str | None]:
    if not isinstance(cell, dict) or cell.get("status") != "ok":
        return "n/a", None
    value = cell.get(metric)
    if value is None:
        return "n/a", None
    return value, FMT_PCT if metric == "irr" else FMT_MULT2


def _build_sensitivity_grid(ws, start_row: int, title: str, grid: dict[str, Any]) -> int:
    _section(ws, start_row, title, span=8)
    row = start_row + 1
    cols = grid.get("cols") or []
    _header_row(ws, row, [grid.get("row_label") or "Row"] + [f"{col:.1f}x" for col in cols])
    row += 1
    for r_idx, row_value in enumerate(grid.get("rows") or []):
        _cell(ws, row, 1, f"{row_value:.1f}x", bold=True, align="left")
        cells = (grid.get("cells") or [])
        line = cells[r_idx] if r_idx < len(cells) else []
        for c_idx, _ in enumerate(cols):
            cell = line[c_idx] if c_idx < len(line) else {}
            value, fmt = _sens_value(cell, "irr")
            out = _cell(ws, row, 2 + c_idx, value, fmt=fmt, align="right")
            if isinstance(cell, dict) and cell.get("is_base"):
                out.fill = _FILL_INPUT
        row += 1
    row += 1
    _cell(ws, row, 1, "MOIC Detail", bold=True, align="left")
    row += 1
    _header_row(ws, row, [grid.get("row_label") or "Row"] + [f"{col:.1f}x" for col in cols])
    row += 1
    for r_idx, row_value in enumerate(grid.get("rows") or []):
        _cell(ws, row, 1, f"{row_value:.1f}x", bold=True, align="left")
        line = (grid.get("cells") or [])[r_idx] if r_idx < len(grid.get("cells") or []) else []
        for c_idx, _ in enumerate(cols):
            cell = line[c_idx] if c_idx < len(line) else {}
            value, fmt = _sens_value(cell, "moic")
            out = _cell(ws, row, 2 + c_idx, value, fmt=fmt, align="right")
            if isinstance(cell, dict) and cell.get("is_base"):
                out.fill = _FILL_INPUT
        row += 1
    return row + 2


def _build_sensitivity_summary(ws, sensitivity: dict) -> None:
    _title(ws, SENSITIVITY_SHEET, span=8)
    _set_widths(ws, {"A": 24, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 22})
    base = sensitivity.get("base") or {}
    row = 3
    row = _write_pairs(ws, row, [
        ("Method", sensitivity.get("method")),
        ("Base Entry Multiple", base.get("entry_multiple"), FMT_MULT1),
        ("Base Exit Multiple", base.get("exit_multiple"), FMT_MULT1),
        ("Base Debt / EBITDA", base.get("leverage"), FMT_MULT1),
    ])
    row += 1
    _note(ws, row, 1, "Deterministic sensitivity; not probability-weighted. IRR / MOIC values come from Python run_lbo.", end_col=8)
    row += 2
    grids = sensitivity.get("grids") or {}
    row = _build_sensitivity_grid(ws, row, "Entry Multiple x Exit Multiple", grids.get("entry_exit") or {})
    row = _build_sensitivity_grid(ws, row, "Debt / EBITDA x Exit Multiple", grids.get("leverage_exit") or {})
    ws.freeze_panes = "B9"


def generate_lbo_excel(inputs: dict[str, Any], result: dict[str, Any]) -> BytesIO:
    if result.get("status") != "ok" or not result.get("returns"):
        raise ValueError("Cannot export LBO workbook for non-ok result.")

    cap_summary = result.get("capital_structure_summary") or {}
    is_multi = cap_summary.get("mode") == "multi_tranche"

    scenario_data = _scenario_comparison(inputs)
    sensitivity = _sensitivity_data(inputs)

    wb = Workbook()
    ws_cover = wb.active
    ws_cover.title = SHEET_NAMES[0]
    sheets = {SHEET_NAMES[0]: ws_cover}
    for name in SHEET_NAMES[1:]:
        sheets[name] = wb.create_sheet(name)

    if scenario_data is not None:
        sheets[SCENARIO_SUMMARY_SHEET] = wb.create_sheet(
            SCENARIO_SUMMARY_SHEET, _SCENARIO_INSERT_INDEX
        )
    if sensitivity is not None:
        insert_at = _SCENARIO_INSERT_INDEX + (1 if scenario_data is not None else 0)
        sheets[SENSITIVITY_SHEET] = wb.create_sheet(SENSITIVITY_SHEET, insert_at)

    _build_cover(sheets["Cover"], inputs, result, is_multi)
    _build_model_map(sheets["Model Map"], scenario_data is not None, sensitivity is not None)
    if scenario_data is not None:
        _build_scenario_summary(sheets[SCENARIO_SUMMARY_SHEET], scenario_data)
    if sensitivity is not None:
        _build_sensitivity_summary(sheets[SENSITIVITY_SHEET], sensitivity)
    _build_assumptions(sheets["Assumptions"], inputs, result, is_multi, cap_summary)
    _build_sources_uses(sheets["Sources & Uses"], inputs, result, is_multi, cap_summary)
    _build_operating_forecast(sheets["Operating Forecast"], result, is_multi)
    _build_debt_schedule(sheets["Debt Schedule"], result, is_multi)
    _build_covenant_check(sheets["Covenant Check"], result, is_multi)
    _build_returns_summary(sheets["Returns Summary"], result, is_multi, inputs)
    _build_attribution(sheets["Returns Attribution"], inputs)
    _build_maturity_wall(sheets["Maturity Wall"], result, is_multi)
    _build_audit(sheets["Audit & Disclosures"], inputs, result, is_multi, cap_summary)

    # V4.7 RC hotfix: consistent amount-unit note on amount-bearing sheets.
    for name in ("Assumptions", "Sources & Uses", "Operating Forecast",
                 "Debt Schedule", "Returns Summary"):
        _append_units_note(sheets[name], result)
    if scenario_data is not None:
        _append_units_note(sheets[SCENARIO_SUMMARY_SHEET], result)
    if sensitivity is not None:
        _append_units_note(sheets[SENSITIVITY_SHEET], result)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# Explicit legacy alias. V4.1.0 introduces a formula-native workbook
# (modeling.lbo_formula_workbook.generate_lbo_formula_excel) as the user-facing
# export. This values-only builder is retained as the Python-oracle / gold-master
# reference path and keeps its 0-formula, values-only contract.
generate_lbo_values_only_excel = generate_lbo_excel
