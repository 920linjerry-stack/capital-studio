"""V4.1.1 Multi-Tranche Debt Waterfall & Cash Sweep Formulaization.

This module builds the *formula-native* LBO Excel export that the user receives
from ``/api/modeling/lbo/excel``. Unlike the legacy values-only workbook
(``modeling.lbo_excel_exporter.generate_lbo_values_only_excel``), the core model
sections here contain real Excel formulas (cells beginning with ``=``) wired
through stable Assumptions anchors / workbook named ranges, so a user can open
the file, edit a key assumption, and let Excel recalculate the core
transaction / operating / returns chain.

Discipline
----------
* ``modeling.lbo_calculator.run_lbo()`` remains the calculation oracle. This
  module never re-implements engine math; the initial exported case is the
  engine's output and the Excel formulas reproduce the same closed-form
  relationships (beginning-balance interest, so no Excel circular reference).
* Formula-native sheets (single-tranche): Assumptions, Sources & Uses,
  Operating Forecast, Debt Schedule, Returns Summary / Exit Bridge.
* Formula-native sheets (multi-tranche, V4.1.1): Assumptions anchors (incl.
  per-tranche opening / rate / amort / commitment inputs), Sources & Uses,
  Operating Forecast, the full tranche-level Debt Schedule waterfall
  (beginning balance, beginning-balance interest, mandatory amortization,
  revolver draw, priority-ordered cash sweep, ending balance, total / net debt,
  ending cash) and the Returns Summary / Exit Bridge. The bridge references the
  formula-driven total ending debt and ending cash named ranges rather than
  hard-coded engine outputs.
* Scenario Summary and Sensitivity Analysis remain Python-precomputed
  presentation layers (values only), reused unchanged from the legacy exporter.
* No network, no LLM, no new dependency, no V5 M&A surface is touched.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.properties import PageSetupProperties

from modeling.templates.goldman_dcf import (
    border_thin,
    font_formula,
    font_label,
    font_link,
)
from modeling.lbo_attribution import MOIC_CONTRIBUTION_NOTE_EN, MOIC_CONTRIBUTION_NOTE_CN
from modeling.lbo_capital_structure import validate_capital_structure
from modeling.lbo_scenarios import SCENARIO_LABELS
from modeling.lbo_excel_exporter import (
    FMT_MONEY,
    FMT_MULT1,
    FMT_MULT2,
    FMT_PCT,
    FMT_YEAR,
    SCENARIO_SUMMARY_SHEET,
    SENSITIVITY_SHEET,
    SHEET_NAMES,
    _FILL_INPUT,
    _SCENARIO_INSERT_INDEX,
    _append_units_note,
    _assumptions_extra_sections,
    _build_audit,
    _build_covenant_check,
    _build_cover,
    _build_debt_schedule,
    _build_maturity_wall,
    _build_model_map,
    _build_operating_forecast,
    _build_scenario_summary,
    _build_sensitivity_grid,
    _build_sensitivity_summary,
    _cell,
    _header_row,
    _note,
    _scenario_comparison,
    _section,
    _sensitivity_data,
    _set_widths,
    _title,
    _write_pairs,
    _yn,
)

FORMULA_WORKBOOK_VERSION = "V4.1.4 Final Formula-Native LBO Workbook"

_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")

# Pale-peach fill marking headline OUTPUT cells (e.g. MOIC / IRR), so the four
# workbook regions read at a glance: light-blue = editable input, green link
# font = cross-sheet reference, plain = formula, peach = headline output. The
# pass/fail check region is demarcated by its "...Checks" section titles.
_FILL_OUTPUT = PatternFill("solid", fgColor="FCE4D6")

# Wide, multi-column sheets print best in landscape; the narrow label/value
# sheets stay portrait. Both fit to one page wide so a banker can print the
# model without manual page setup.
_LANDSCAPE_SHEETS = frozenset({
    "Model Map", "Scenario Summary", "Sensitivity Analysis", "Operating Forecast",
    "Debt Schedule", "Covenant Check", "Returns Attribution", "Maturity Wall",
})


def _emphasize_output(cell) -> None:
    """Apply the headline-output fill to an already-written formula/value cell."""
    cell.fill = _FILL_OUTPUT


def _apply_print_setup(ws, landscape: bool) -> None:
    """Deterministic banker print setup: fit-to-one-page-wide, A4, light margins.
    No volatile content; affects presentation only, never the formula chain."""
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = False

# Workbook-level named ranges (stable anchors). Downstream formulas reference
# these names rather than fragile A1 addresses, so inserting a row elsewhere does
# not break the model chain.
N_ENTRY_EBITDA = "LBO_Entry_EBITDA"
N_ENTRY_MULT = "LBO_Entry_Multiple"
N_EXIT_MULT = "LBO_Exit_Multiple"
N_HOLDING = "LBO_Holding_Period"
N_FEES_PCT = "LBO_Fees_Pct"
N_CASH_TO_BS = "LBO_Cash_To_BS"
N_OPENING_DEBT = "LBO_Opening_Debt"
N_INTEREST_RATE = "LBO_Interest_Rate"
N_MAND_AMORT = "LBO_Mand_Amort_Pct"
N_CASH_SWEEP = "LBO_Cash_Sweep_Pct"
N_TAX_RATE = "LBO_Tax_Rate"
N_TAX_SHIELD_ON = "LBO_Tax_Shield_On"
N_TOTAL_USES = "LBO_Total_Uses"
N_TOTAL_DEBT = "LBO_Total_Debt"
N_SPONSOR_EQUITY = "LBO_Sponsor_Equity"
N_EXIT_EQUITY = "LBO_Exit_Equity"
N_MOIC = "LBO_MOIC"
# Multi-tranche (V4.1.1) anchors / outputs.
N_MIN_CASH = "LBO_Min_Cash"
N_CASH_SWEEP_ON = "LBO_Cash_Sweep_On"
N_TOTAL_ENDING_DEBT = "LBO_Total_Ending_Debt"
N_ENDING_CASH = "LBO_Ending_Cash"
# Covenant thresholds (V4.1.2; editable inputs on the Covenant Check sheet that
# the leverage / coverage pass-fail IF() formulas reference).
N_MAX_LEVERAGE = "LBO_Max_Leverage"
N_MIN_COVERAGE = "LBO_Min_Coverage"

# Tiny epsilon used in the Excel gating formulas to mirror the engine's EPS
# comparisons (written as a decimal literal so the lazy test evaluator's cell
# regex never misreads scientific notation as a cell address).
_EPS_LIT = "0.000000001"


def _tranche_name(prefix: str, idx: int) -> str:
    """Stable per-tranche named-range name (e.g. ``LBO_Tr0_Opening``)."""
    return f"LBO_Tr{idx}_{prefix}"


def _define(wb: Workbook, name: str, sheet: str, row: int, col: int) -> None:
    ref = f"{quote_sheetname(sheet)}!${get_column_letter(col)}${row}"
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name, attr_text=ref))


def _formula(ws, row: int, col: int, formula: str, fmt: str | None = None,
             cross: bool = False):
    """Write an Excel formula cell. ``cross=True`` uses the cross-sheet link font
    so users can visually distinguish references to other sheets."""
    c = ws.cell(row, col, formula)
    c.font = font_link() if cross else font_formula()
    c.border = border_thin()
    if fmt is not None:
        c.number_format = fmt
    c.alignment = _RIGHT
    return c


def _input_cell(ws, row: int, col: int, value: Any, fmt: str | None = None):
    """Write a user-editable input value with the light-blue input fill."""
    c = _cell(ws, row, col, value, fmt=fmt)
    c.fill = _FILL_INPUT
    return c


def _bool_to_flag(value: Any) -> int:
    return 1 if bool(value) else 0


# ── Assumptions (input anchors + named ranges) ───────────────────────────────


def _build_assumptions_formula(ws, wb, inputs, result, is_multi, cap_summary,
                               settings=None) -> None:
    _title(ws, "Assumptions", span=5)
    _set_widths(ws, {"A": 32, "B": 18, "C": 22, "D": 16, "E": 46})
    ts = result["transaction_summary"]
    of = result.get("operating_forecast") or {}
    debt = inputs.get("debt") or {}

    entry_ev = ts["entry_ev"]
    fees_pct = (ts["transaction_fees"] / entry_ev) if entry_ev else 0.0
    tax_rate = of.get("tax_rate", inputs.get("tax_rate", 0.25))
    tax_on = _bool_to_flag(of.get("tax_shield_enabled", inputs.get("tax_shield_enabled", True)))

    if is_multi:
        interest_rate = cap_summary.get("weighted_avg_cash_interest_rate_year_1") or 0.0
        mand_amort = 0.0
        cash_sweep = 1.0
    else:
        interest_rate = debt.get("interest_rate") or 0.0
        mand_amort = debt.get("mandatory_amortization_pct") or 0.0
        cash_sweep = debt.get("cash_sweep_pct", 1.0)

    _section(ws, 3, "Transaction Assumptions", span=5)
    # Fixed-row anchor block. Label in col 1, editable value in col 2.
    anchors = [
        (4, "Entry EBITDA", ts["entry_ebitda"], FMT_MONEY, N_ENTRY_EBITDA),
        (5, "Entry Multiple", ts["entry_multiple"], FMT_MULT1, N_ENTRY_MULT),
        (6, "Exit Multiple", ts["exit_multiple"], FMT_MULT1, N_EXIT_MULT),
        (7, "Holding Period (years)", ts["exit_year"], FMT_YEAR, N_HOLDING),
        (8, "Transaction Fees % of EV", fees_pct, FMT_PCT, N_FEES_PCT),
        (9, "Cash to Balance Sheet", ts.get("cash_to_balance_sheet", 0.0), FMT_MONEY, N_CASH_TO_BS),
    ]
    _section(ws, 11, "Debt Assumptions", span=5)
    debt_label = "Total Opening Debt" if is_multi else "Opening Debt"
    anchors += [
        (12, debt_label, ts["debt_amount"], FMT_MONEY, N_OPENING_DEBT),
        (13, "Interest Rate" + (" (Wtd Avg Y1)" if is_multi else ""), interest_rate, FMT_PCT, N_INTEREST_RATE),
        (14, "Mandatory Amortization %", mand_amort, FMT_PCT, N_MAND_AMORT),
        (15, "Cash Sweep %", cash_sweep, FMT_PCT, N_CASH_SWEEP),
    ]
    _section(ws, 17, "Tax Assumptions", span=5)
    anchors += [
        (18, "Tax Rate", tax_rate, FMT_PCT, N_TAX_RATE),
        (19, "Tax Shield Enabled (1 = on)", tax_on, FMT_YEAR, N_TAX_SHIELD_ON),
    ]
    for row, label, value, fmt, name in anchors:
        _cell(ws, row, 1, label, bold=True, align="left")
        _input_cell(ws, row, 2, value, fmt=fmt)
        _define(wb, name, ws.title, row, 2)

    row = 21
    if is_multi and settings is not None:
        # V4.1.1: per-tranche inputs are editable anchors (named ranges) that the
        # Debt Schedule waterfall formulas reference, so editing a tranche's
        # opening balance / rate / amortization / commitment recalculates the
        # whole multi-tranche chain in Excel.
        _section(ws, row, "Multi-Tranche Detail (editable inputs)", span=7)
        row += 1
        _header_row(ws, row, ["Tranche", "Opening Balance", "Interest Rate",
                              "Mand Amort %", "Commitment", "Maturity Yr", "Type"])
        row += 1
        for i, t in enumerate(settings["tranches"]):
            _cell(ws, row, 1, t["name"], align="left")
            _input_cell(ws, row, 2, t["opening_balance"], fmt=FMT_MONEY)
            _define(wb, _tranche_name("Opening", i), ws.title, row, 2)
            _input_cell(ws, row, 3, t["interest_rate"], fmt=FMT_PCT)
            _define(wb, _tranche_name("Rate", i), ws.title, row, 3)
            _input_cell(ws, row, 4, t["mandatory_amortization_pct"], fmt=FMT_PCT)
            _define(wb, _tranche_name("Amort", i), ws.title, row, 4)
            _input_cell(ws, row, 5, t["commitment"], fmt=FMT_MONEY)
            _define(wb, _tranche_name("Commit", i), ws.title, row, 5)
            _cell(ws, row, 6, t["maturity_year"], fmt=FMT_YEAR)
            _cell(ws, row, 7, t["type"], align="left")
            row += 1
        row += 1
        _cell(ws, row, 1, "Minimum Cash Balance", bold=True, align="left")
        _input_cell(ws, row, 2, settings.get("minimum_cash_balance", 0.0), fmt=FMT_MONEY)
        _define(wb, N_MIN_CASH, ws.title, row, 2)
        row += 1
        _cell(ws, row, 1, "Cash Sweep Enabled (1 = on)", bold=True, align="left")
        _input_cell(ws, row, 2, _bool_to_flag(settings.get("cash_sweep_enabled", True)), fmt=FMT_YEAR)
        _define(wb, N_CASH_SWEEP_ON, ws.title, row, 2)
        row += 2

    row = _assumptions_extra_sections(ws, inputs, row)
    ws.freeze_panes = "A3"


# ── Sources & Uses (fully formula-driven in both modes) ──────────────────────


def _build_sources_uses_formula(ws, wb, inputs, result, is_multi, cap_summary) -> None:
    _title(ws, "Sources & Uses", span=2)
    _set_widths(ws, {"A": 34, "B": 20})

    _section(ws, 3, "Uses", span=2)
    _cell(ws, 4, 1, "Entry EV", bold=True, align="left")
    _formula(ws, 4, 2, f"={N_ENTRY_EBITDA}*{N_ENTRY_MULT}", FMT_MONEY)
    _cell(ws, 5, 1, "Transaction Fees", bold=True, align="left")
    _formula(ws, 5, 2, f"=B4*{N_FEES_PCT}", FMT_MONEY)
    _cell(ws, 6, 1, "Cash to Balance Sheet", bold=True, align="left")
    _formula(ws, 6, 2, f"={N_CASH_TO_BS}", FMT_MONEY)
    _cell(ws, 7, 1, "Total Uses", bold=True, align="left")
    _formula(ws, 7, 2, "=B4+B5+B6", FMT_MONEY)
    _define(wb, N_TOTAL_USES, ws.title, 7, 2)

    _section(ws, 9, "Sources", span=2)
    row = 10
    if is_multi:
        for i, t in enumerate(cap_summary.get("tranches") or []):
            _cell(ws, row, 1, t.get("name"), bold=True, align="left")
            _formula(ws, row, 2, f"={_tranche_name('Opening', i)}", FMT_MONEY)
            row += 1
        _cell(ws, row, 1, "Total Opening Debt", bold=True, align="left")
        _formula(ws, row, 2, f"={N_OPENING_DEBT}", FMT_MONEY)
        debt_row = row
        row += 1
    else:
        _cell(ws, row, 1, "Debt Amount", bold=True, align="left")
        _formula(ws, row, 2, f"={N_OPENING_DEBT}", FMT_MONEY)
        debt_row = row
        row += 1
    _define(wb, N_TOTAL_DEBT, ws.title, debt_row, 2)

    _cell(ws, row, 1, "Sponsor Equity", bold=True, align="left")
    _formula(ws, row, 2, f"=B7-{N_TOTAL_DEBT}", FMT_MONEY)
    _define(wb, N_SPONSOR_EQUITY, ws.title, row, 2)
    sponsor_row = row
    row += 2

    _section(ws, row, "Ratios", span=2)
    row += 1
    _cell(ws, row, 1, "Debt % of Uses", bold=True, align="left")
    _formula(ws, row, 2, f"=IFERROR({N_TOTAL_DEBT}/{N_TOTAL_USES},0)", FMT_PCT)
    row += 1
    _cell(ws, row, 1, "Sponsor Equity % of Uses", bold=True, align="left")
    _formula(ws, row, 2, f"=IFERROR(B{sponsor_row}/{N_TOTAL_USES},0)", FMT_PCT)
    row += 1
    _cell(ws, row, 1, "Entry Leverage (Debt / EBITDA)", bold=True, align="left")
    _formula(ws, row, 2, f"=IFERROR({N_TOTAL_DEBT}/{N_ENTRY_EBITDA},0)", FMT_MULT1)
    row += 2

    _section(ws, row, "Check", span=2)
    row += 1
    _cell(ws, row, 1, "Total Sources", bold=True, align="left")
    _formula(ws, row, 2, f"={N_TOTAL_DEBT}+B{sponsor_row}", FMT_MONEY)
    total_sources_row = row
    row += 1
    _cell(ws, row, 1, "Total Uses", bold=True, align="left")
    _formula(ws, row, 2, f"={N_TOTAL_USES}", FMT_MONEY)
    total_uses_row = row
    row += 1
    _cell(ws, row, 1, "Sources - Uses", bold=True, align="left")
    _formula(ws, row, 2, f"=B{total_sources_row}-B{total_uses_row}", FMT_MONEY)
    diff_row = row
    row += 1
    _cell(ws, row, 1, "Sources = Uses", bold=True, align="left")
    _formula(ws, row, 2, f'=IF(ABS(B{diff_row})<0.01,"Yes","No")')
    ws.freeze_panes = "A3"


# ── Operating Forecast (single-tranche formula chain) ────────────────────────

# Row map for the single-tranche formula Operating Forecast / Debt Schedule.
_OF_ROW = {
    "Revenue": 4,
    "EBITDA": 5,
    "EBITDA Margin (%)": 6,
    "Gross Cash Taxes": 7,
    "Tax Shield": 8,
    "Levered Cash Taxes": 9,
    "CapEx": 10,
    "Change in NWC": 11,
    "Cash Flow Before Debt Service": 12,
}
_DS_ROW = {
    "Beginning Debt": 4,
    "Cash Interest": 5,
    "Cash Flow Before Debt Service": 6,
    "Cash Available for Debt": 7,
    "Mandatory Amortization": 8,
    "FCF Available for Sweep": 9,
    "Optional Repayment": 10,
    "Ending Debt": 11,
}


def _build_operating_forecast_formula(ws, result, ds_interest_row: int | None = None) -> None:
    # ``ds_interest_row`` is the Debt Schedule row carrying total cash interest
    # that the Tax Shield references. Single-tranche uses the per-year Cash
    # Interest row; multi-tranche passes its Total Cash Interest row so the same
    # operating-forecast layout serves both modes.
    if ds_interest_row is None:
        ds_interest_row = _DS_ROW["Cash Interest"]
    _title(ws, "Operating Forecast", span=9)
    of = result["operating_forecast"]
    years = of["years"]
    n = len(years)
    widths = {"A": 28}
    for i in range(n):
        widths[get_column_letter(2 + i)] = 14
    _set_widths(ws, widths)
    _header_row(ws, 3, ["Metric"] + [f"Y{y}" for y in years])

    revenue = of.get("revenue") or [None] * n
    ebitda = of.get("ebitda") or [None] * n
    gross_tax = of.get("gross_cash_taxes") or of.get("cash_taxes") or [None] * n
    capex = of.get("capex") or [None] * n
    nwc = of.get("change_in_nwc") or [None] * n

    # Input rows (manual forecast values; downstream formulas reference them).
    for label, values in (("Revenue", revenue), ("EBITDA", ebitda),
                          ("Gross Cash Taxes", gross_tax), ("CapEx", capex),
                          ("Change in NWC", nwc)):
        _cell(ws, _OF_ROW[label], 1, label, bold=True, align="left")
        for i, v in enumerate(values):
            _input_cell(ws, _OF_ROW[label], 2 + i, v, fmt=FMT_MONEY)

    # Formula rows (continuous per-year chain).
    for label in ("EBITDA Margin (%)", "Tax Shield", "Levered Cash Taxes",
                  "Cash Flow Before Debt Service"):
        _cell(ws, _OF_ROW[label], 1, label, bold=True, align="left")

    ds_sheet = quote_sheetname("Debt Schedule")
    for i in range(n):
        col = get_column_letter(2 + i)
        # EBITDA Margin = EBITDA / Revenue
        _formula(ws, _OF_ROW["EBITDA Margin (%)"], 2 + i,
                 f"=IFERROR({col}{_OF_ROW['EBITDA']}/{col}{_OF_ROW['Revenue']},0)", FMT_PCT)
        # Tax Shield = Cash Interest (Debt Schedule) * Tax Rate * Tax Shield On
        _formula(ws, _OF_ROW["Tax Shield"], 2 + i,
                 f"={ds_sheet}!{col}{ds_interest_row}*{N_TAX_RATE}*{N_TAX_SHIELD_ON}",
                 FMT_MONEY, cross=True)
        # Levered Cash Taxes = MAX(0, Gross Cash Taxes - Tax Shield)
        _formula(ws, _OF_ROW["Levered Cash Taxes"], 2 + i,
                 f"=MAX(0,{col}{_OF_ROW['Gross Cash Taxes']}-{col}{_OF_ROW['Tax Shield']})", FMT_MONEY)
        # CFBDS = EBITDA - Levered Cash Taxes - CapEx - Change in NWC
        _formula(ws, _OF_ROW["Cash Flow Before Debt Service"], 2 + i,
                 f"={col}{_OF_ROW['EBITDA']}-{col}{_OF_ROW['Levered Cash Taxes']}"
                 f"-{col}{_OF_ROW['CapEx']}-{col}{_OF_ROW['Change in NWC']}", FMT_MONEY)
    ws.freeze_panes = "B4"


def _build_debt_schedule_formula(ws, result) -> None:
    _title(ws, "Debt Schedule", span=9)
    of = result["operating_forecast"]
    years = of["years"]
    n = len(years)
    widths = {"A": 30}
    for i in range(n):
        widths[get_column_letter(2 + i)] = 14
    _set_widths(ws, widths)
    _section(ws, 2, "Single-Tranche Debt Waterfall", span=1 + n)
    _header_row(ws, 3, ["Item"] + [f"Y{y}" for y in years])

    for label in _DS_ROW:
        _cell(ws, _DS_ROW[label], 1, label, bold=True, align="left")

    of_sheet = quote_sheetname("Operating Forecast")
    end_row = _DS_ROW["Ending Debt"]
    for i in range(n):
        col = get_column_letter(2 + i)
        # Beginning Debt: Y1 = Opening Debt; later years carry prior Ending Debt.
        if i == 0:
            _formula(ws, _DS_ROW["Beginning Debt"], 2 + i, f"={N_OPENING_DEBT}", FMT_MONEY)
        else:
            prev = get_column_letter(1 + i)
            _formula(ws, _DS_ROW["Beginning Debt"], 2 + i, f"={prev}{end_row}", FMT_MONEY)
        # Cash Interest = Beginning Debt * Interest Rate
        _formula(ws, _DS_ROW["Cash Interest"], 2 + i,
                 f"={col}{_DS_ROW['Beginning Debt']}*{N_INTEREST_RATE}", FMT_MONEY)
        # CFBDS pulled from Operating Forecast
        _formula(ws, _DS_ROW["Cash Flow Before Debt Service"], 2 + i,
                 f"={of_sheet}!{col}{_OF_ROW['Cash Flow Before Debt Service']}", FMT_MONEY, cross=True)
        # Cash Available for Debt = CFBDS - Cash Interest
        _formula(ws, _DS_ROW["Cash Available for Debt"], 2 + i,
                 f"={col}{_DS_ROW['Cash Flow Before Debt Service']}-{col}{_DS_ROW['Cash Interest']}", FMT_MONEY)
        # Mandatory Amortization = MIN(Beginning Debt, Opening Debt * Mand Amort %)
        _formula(ws, _DS_ROW["Mandatory Amortization"], 2 + i,
                 f"=MIN({col}{_DS_ROW['Beginning Debt']},{N_OPENING_DEBT}*{N_MAND_AMORT})", FMT_MONEY)
        # FCF Available for Sweep = Cash Available - Mandatory
        _formula(ws, _DS_ROW["FCF Available for Sweep"], 2 + i,
                 f"={col}{_DS_ROW['Cash Available for Debt']}-{col}{_DS_ROW['Mandatory Amortization']}", FMT_MONEY)
        # Optional Repayment = MIN(MAX(0, FCF * Sweep%), Beginning - Mandatory)
        _formula(ws, _DS_ROW["Optional Repayment"], 2 + i,
                 f"=MIN(MAX(0,{col}{_DS_ROW['FCF Available for Sweep']}*{N_CASH_SWEEP}),"
                 f"{col}{_DS_ROW['Beginning Debt']}-{col}{_DS_ROW['Mandatory Amortization']})", FMT_MONEY)
        # Ending Debt = Beginning - Mandatory - Optional
        _formula(ws, _DS_ROW["Ending Debt"], 2 + i,
                 f"={col}{_DS_ROW['Beginning Debt']}-{col}{_DS_ROW['Mandatory Amortization']}"
                 f"-{col}{_DS_ROW['Optional Repayment']}", FMT_MONEY)
    ws.freeze_panes = "B4"


# ── Multi-tranche Debt Schedule (V4.1.1 formula waterfall) ───────────────────

# Row keys for the multi-tranche totals / cash block. Per-tranche rows are laid
# out first (6 rows each), then this fixed-order totals block follows.
_DS_MULTI_TOTAL_KEYS = [
    "Total Beginning Debt",
    "Total Cash Interest",
    "Cash Flow Before Debt Service",
    "Beginning Cash",
    "Cash Before Debt Service",
    "Total Mandatory Amortization",
    "Required Debt Service",
    "Revolver Draw",
    "Cash After Required Debt Service",
    "Cash Available for Sweep",
    "Total Optional Repayment",
    "Ending Cash",
    "Total Ending Debt",
    "Net Debt",
]
_DS_MULTI_TRANCHE_LINES = (
    "Beginning Balance", "Cash Interest", "Mandatory Amortization",
    "Draw", "Optional Repayment", "Ending Balance",
)


def _multi_ds_layout(tranches):
    """Return ``(per_tranche, totals)`` row maps for the multi-tranche Debt
    Schedule. Header is row 3; per-tranche blocks start at row 4."""
    per_tranche: dict[str, dict[str, int]] = {}
    r = 4
    for t in tranches:
        per_tranche[t["id"]] = {line: r + off for off, line in enumerate(_DS_MULTI_TRANCHE_LINES)}
        r += len(_DS_MULTI_TRANCHE_LINES)
    totals: dict[str, int] = {}
    for key in _DS_MULTI_TOTAL_KEYS:
        totals[key] = r
        r += 1
    return per_tranche, totals


def _build_debt_schedule_formula_multi(ws, wb, result, settings, per_tranche, totals) -> None:
    _title(ws, "Debt Schedule", span=10)
    of = result["operating_forecast"]
    years = of["years"]
    n = len(years)
    tranches = settings["tranches"]
    widths = {"A": 34}
    for i in range(n):
        widths[get_column_letter(2 + i)] = 14
    _set_widths(ws, widths)
    _section(ws, 2, "Multi-Tranche Debt Waterfall", span=1 + n)
    _header_row(ws, 3, ["Item"] + [f"Y{y}" for y in years])

    # Row labels (col A).
    for t in tranches:
        rows = per_tranche[t["id"]]
        for line in _DS_MULTI_TRANCHE_LINES:
            _cell(ws, rows[line], 1, f"{t['name']} — {line}", bold=(line == "Beginning Balance"), align="left")
    for key in _DS_MULTI_TOTAL_KEYS:
        _cell(ws, totals[key], 1, key, bold=True, align="left")

    of_sheet = quote_sheetname("Operating Forecast")
    revolver = next((t for t in tranches if t["is_revolver"]), None)
    rev_idx = tranches.index(revolver) if revolver is not None else None
    # Optional cash sweep follows sweep_priority ascending (engine ordering).
    sweepable = sorted(
        [t for t in tranches if t["optional_repay_allowed"]],
        key=lambda t: (t["sweep_priority"], t["id"]),
    )

    for j in range(n):
        col = get_column_letter(2 + j)
        prev = get_column_letter(1 + j)

        # ── Per-tranche beginning / interest / mandatory ──────────────────
        for i, t in enumerate(tranches):
            rows = per_tranche[t["id"]]
            opening = _tranche_name("Opening", i)
            rate = _tranche_name("Rate", i)
            amort = _tranche_name("Amort", i)
            if j == 0:
                _formula(ws, rows["Beginning Balance"], 2 + j, f"={opening}", FMT_MONEY)
            else:
                _formula(ws, rows["Beginning Balance"], 2 + j, f"={prev}{rows['Ending Balance']}", FMT_MONEY)
            _formula(ws, rows["Cash Interest"], 2 + j,
                     f"={col}{rows['Beginning Balance']}*{rate}", FMT_MONEY)
            if t["is_revolver"]:
                _formula(ws, rows["Mandatory Amortization"], 2 + j, "=0", FMT_MONEY)
            else:
                # Mandatory uses original opening balance, capped by beginning.
                _formula(ws, rows["Mandatory Amortization"], 2 + j,
                         f"=MIN({col}{rows['Beginning Balance']},{opening}*{amort})", FMT_MONEY)

        # ── Aggregates that the cash waterfall needs ──────────────────────
        beg_sum = "+".join(f"{col}{per_tranche[t['id']]['Beginning Balance']}" for t in tranches)
        _formula(ws, totals["Total Beginning Debt"], 2 + j, f"={beg_sum}", FMT_MONEY)
        int_sum = "+".join(f"{col}{per_tranche[t['id']]['Cash Interest']}" for t in tranches)
        _formula(ws, totals["Total Cash Interest"], 2 + j, f"={int_sum}", FMT_MONEY)
        _formula(ws, totals["Cash Flow Before Debt Service"], 2 + j,
                 f"={of_sheet}!{col}{_OF_ROW['Cash Flow Before Debt Service']}", FMT_MONEY, cross=True)
        if j == 0:
            _formula(ws, totals["Beginning Cash"], 2 + j, f"={N_CASH_TO_BS}", FMT_MONEY)
        else:
            _formula(ws, totals["Beginning Cash"], 2 + j, f"={prev}{totals['Ending Cash']}", FMT_MONEY)
        _formula(ws, totals["Cash Before Debt Service"], 2 + j,
                 f"={col}{totals['Cash Flow Before Debt Service']}+{col}{totals['Beginning Cash']}", FMT_MONEY)
        mand_sum = "+".join(f"{col}{per_tranche[t['id']]['Mandatory Amortization']}" for t in tranches)
        _formula(ws, totals["Total Mandatory Amortization"], 2 + j, f"={mand_sum}", FMT_MONEY)
        _formula(ws, totals["Required Debt Service"], 2 + j,
                 f"={col}{totals['Total Cash Interest']}+{col}{totals['Total Mandatory Amortization']}", FMT_MONEY)

        # ── Revolver draw (only on a cash shortfall) ──────────────────────
        if revolver is not None:
            commit = _tranche_name("Commit", rev_idx)
            rev_beg = f"{col}{per_tranche[revolver['id']]['Beginning Balance']}"
            shortfall = f"({col}{totals['Required Debt Service']}-{col}{totals['Cash Before Debt Service']})"
            _formula(ws, totals["Revolver Draw"], 2 + j,
                     f"=IF({shortfall}>{_EPS_LIT},MIN({shortfall},MAX(0,{commit}-{rev_beg})),0)", FMT_MONEY)
        else:
            _formula(ws, totals["Revolver Draw"], 2 + j, "=0", FMT_MONEY)
        # Per-tranche draw rows: revolver mirrors the draw total, others zero.
        for t in tranches:
            rows = per_tranche[t["id"]]
            if t["is_revolver"]:
                _formula(ws, rows["Draw"], 2 + j, f"={col}{totals['Revolver Draw']}", FMT_MONEY)
            else:
                _formula(ws, rows["Draw"], 2 + j, "=0", FMT_MONEY)

        # ── Cash after required debt service, then sweepable cash ─────────
        _formula(ws, totals["Cash After Required Debt Service"], 2 + j,
                 f"={col}{totals['Cash Before Debt Service']}+{col}{totals['Revolver Draw']}"
                 f"-{col}{totals['Total Cash Interest']}-{col}{totals['Total Mandatory Amortization']}", FMT_MONEY)
        car = f"{col}{totals['Cash After Required Debt Service']}"
        # Sweep gated exactly as the engine: enabled, no revolver draw this year,
        # and cash above the minimum balance. When gated off this is 0, which
        # forces every per-tranche optional repayment below to 0.
        _formula(ws, totals["Cash Available for Sweep"], 2 + j,
                 f"=IF(AND({N_CASH_SWEEP_ON}>0.5,{col}{totals['Revolver Draw']}<={_EPS_LIT},"
                 f"{car}>{N_MIN_CASH}+{_EPS_LIT}),MAX(0,{car}-{N_MIN_CASH}),0)", FMT_MONEY)

        # ── Optional repayment cascade (priority order) ───────────────────
        consumed: list[str] = []
        for t in sweepable:
            rows = per_tranche[t["id"]]
            bal_after = (f"({col}{rows['Beginning Balance']}+{col}{rows['Draw']}"
                         f"-{col}{rows['Mandatory Amortization']})")
            minus = "".join(f"-{col}{per_tranche[e]['Optional Repayment']}" for e in consumed)
            _formula(ws, rows["Optional Repayment"], 2 + j,
                     f"=MIN(MAX(0,{col}{totals['Cash Available for Sweep']}{minus}),MAX(0,{bal_after}))", FMT_MONEY)
            consumed.append(t["id"])
        for t in tranches:
            if not t["optional_repay_allowed"]:
                _formula(ws, per_tranche[t["id"]]["Optional Repayment"], 2 + j, "=0", FMT_MONEY)

        # ── Ending balances and ending cash ───────────────────────────────
        for t in tranches:
            rows = per_tranche[t["id"]]
            _formula(ws, rows["Ending Balance"], 2 + j,
                     f"=MAX(0,{col}{rows['Beginning Balance']}+{col}{rows['Draw']}"
                     f"-{col}{rows['Mandatory Amortization']}-{col}{rows['Optional Repayment']})", FMT_MONEY)
        opt_sum = "+".join(f"{col}{per_tranche[t['id']]['Optional Repayment']}" for t in tranches)
        _formula(ws, totals["Total Optional Repayment"], 2 + j, f"={opt_sum}", FMT_MONEY)
        _formula(ws, totals["Ending Cash"], 2 + j,
                 f"={col}{totals['Cash After Required Debt Service']}-{col}{totals['Total Optional Repayment']}", FMT_MONEY)
        end_sum = "+".join(f"{col}{per_tranche[t['id']]['Ending Balance']}" for t in tranches)
        _formula(ws, totals["Total Ending Debt"], 2 + j, f"={end_sum}", FMT_MONEY)
        _formula(ws, totals["Net Debt"], 2 + j,
                 f"={col}{totals['Total Ending Debt']}-{col}{totals['Ending Cash']}", FMT_MONEY)

    last_col_idx = 2 + n - 1
    _define(wb, N_TOTAL_ENDING_DEBT, ws.title, totals["Total Ending Debt"], last_col_idx)
    _define(wb, N_ENDING_CASH, ws.title, totals["Ending Cash"], last_col_idx)
    ws.freeze_panes = "B4"


# ── Returns Summary / Exit Bridge (formula-driven in both modes) ─────────────


def _build_returns_summary_formula(ws, wb, result, is_multi) -> None:
    _title(ws, "Returns Summary", span=2)
    _set_widths(ws, {"A": 32, "B": 20})
    ex = result["exit"]
    of = result["operating_forecast"]
    n = len(of["years"])
    last_col = get_column_letter(2 + n - 1)

    _section(ws, 3, "Exit Equity Bridge", span=2)
    of_sheet = quote_sheetname("Operating Forecast")
    _cell(ws, 4, 1, "Exit EBITDA", bold=True, align="left")
    # Exit EBITDA references the (formula-native) Operating Forecast final year
    # in both modes.
    _formula(ws, 4, 2, f"={of_sheet}!{last_col}{_OF_ROW['EBITDA']}", FMT_MONEY, cross=True)
    _cell(ws, 5, 1, "Exit Multiple", bold=True, align="left")
    _formula(ws, 5, 2, f"={N_EXIT_MULT}", FMT_MULT1)
    _cell(ws, 6, 1, "Exit Enterprise Value", bold=True, align="left")
    _formula(ws, 6, 2, "=B4*B5", FMT_MONEY)
    _cell(ws, 7, 1, "Less: Remaining Debt", bold=True, align="left")
    if is_multi:
        # V4.1.1: reference the formula-driven multi-tranche total ending debt,
        # not a hard-coded engine value.
        _formula(ws, 7, 2, f"={N_TOTAL_ENDING_DEBT}", FMT_MONEY, cross=True)
    else:
        ds_sheet = quote_sheetname("Debt Schedule")
        _formula(ws, 7, 2, f"={ds_sheet}!{last_col}{_DS_ROW['Ending Debt']}", FMT_MONEY, cross=True)
    _cell(ws, 8, 1, "Plus: Ending Cash", bold=True, align="left")
    if is_multi:
        _formula(ws, 8, 2, f"={N_ENDING_CASH}", FMT_MONEY, cross=True)
    else:
        _cell(ws, 8, 2, 0.0, fmt=FMT_MONEY)
    _cell(ws, 9, 1, "Exit Equity Value", bold=True, align="left")
    _formula(ws, 9, 2, "=B6-B7+B8", FMT_MONEY)
    _define(wb, N_EXIT_EQUITY, ws.title, 9, 2)

    _section(ws, 11, "Headline Returns", span=2)
    _cell(ws, 12, 1, "Sponsor Equity", bold=True, align="left")
    _formula(ws, 12, 2, f"={N_SPONSOR_EQUITY}", FMT_MONEY, cross=True)
    _cell(ws, 13, 1, "Exit Equity Value", bold=True, align="left")
    _formula(ws, 13, 2, f"={N_EXIT_EQUITY}", FMT_MONEY)
    _cell(ws, 14, 1, "MOIC", bold=True, align="left")
    _emphasize_output(_formula(ws, 14, 2, f"=IFERROR({N_EXIT_EQUITY}/{N_SPONSOR_EQUITY},0)", FMT_MULT2))
    _define(wb, N_MOIC, ws.title, 14, 2)
    _cell(ws, 15, 1, "IRR (annualized)", bold=True, align="left")
    # Engine cash flows are a single outflow at close and a single inflow at exit,
    # so the bisection IRR equals MOIC^(1/holding) - 1. The formula ties out.
    _emphasize_output(_formula(ws, 15, 2, f"=IFERROR(POWER({N_MOIC},1/{N_HOLDING})-1,0)", FMT_PCT))
    _cell(ws, 16, 1, "Debt Paydown", bold=True, align="left")
    _formula(ws, 16, 2, f"={N_TOTAL_DEBT}-B7", FMT_MONEY)
    _cell(ws, 17, 1, "Remaining Debt", bold=True, align="left")
    _formula(ws, 17, 2, "=B7", FMT_MONEY)

    row = 19
    note = ws.cell(row, 1,
                   "IRR / MOIC and the exit bridge are live Excel formulas that "
                   "reproduce the run_lbo() oracle relationships; IRR uses "
                   "MOIC^(1/holding)-1, valid because the model has a single "
                   "equity outflow at close and a single inflow at exit. "
                   "Attribution detail is on the Returns Attribution sheet.")
    note.font = font_label()
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.row_dimensions[row].height = 60
    ws.freeze_panes = "A3"


# ── Covenant Check (V4.1.2 formula-linked, multi-tranche) ────────────────────


def _build_covenant_check_formula(ws, wb, result, totals) -> None:
    """Multi-tranche Covenant Check whose leverage / coverage ratios and
    pass/fail flags are live Excel formulas linked to the formula-native Debt
    Schedule (total ending debt, ending cash, total cash interest) and the
    Operating Forecast EBITDA. Thresholds are editable inputs; ``run_lbo()``'s
    covenant engine remains the oracle for the exported case."""
    _title(ws, "Covenant Check", span=12)
    _set_widths(ws, {
        "A": 22, "B": 16, "C": 16, "D": 14, "E": 14, "F": 18, "G": 16,
        "H": 14, "I": 16, "J": 16, "K": 16, "L": 18,
    })
    cov = result.get("covenant_summary") or {}
    of = result["operating_forecast"]
    years = of["years"]
    max_leverage = cov.get("max_net_debt_ebitda")
    min_coverage = cov.get("min_interest_coverage")

    ds_sheet = quote_sheetname("Debt Schedule")
    of_sheet = quote_sheetname("Operating Forecast")
    ebitda_row = _OF_ROW["EBITDA"]
    ted_row = totals["Total Ending Debt"]
    cash_row = totals["Ending Cash"]
    int_row = totals["Total Cash Interest"]

    row = 3
    _section(ws, row, "Covenant Thresholds (editable inputs)", span=4)
    row += 1
    _cell(ws, row, 1, "Max Net Debt / EBITDA", bold=True, align="left")
    have_lev = max_leverage is not None
    if have_lev:
        _input_cell(ws, row, 2, max_leverage, fmt=FMT_MULT2)
        _define(wb, N_MAX_LEVERAGE, ws.title, row, 2)
    else:
        _cell(ws, row, 2, "n/a", align="left")
    row += 1
    _cell(ws, row, 1, "Min Interest Coverage", bold=True, align="left")
    have_cov = min_coverage is not None
    if have_cov:
        _input_cell(ws, row, 2, min_coverage, fmt=FMT_MULT2)
        _define(wb, N_MIN_COVERAGE, ws.title, row, 2)
    else:
        _cell(ws, row, 2, "n/a", align="left")
    row += 1
    _cell(ws, row, 1, "Engine Covenant Status", bold=True, align="left")
    _cell(ws, row, 2, cov.get("status") or "n/a", align="left")
    row += 2

    _section(ws, row, "Covenant Schedule (formula-linked to Debt Schedule)", span=12)
    row += 1
    _header_row(ws, row, [
        "Year", "EBITDA", "Total Ending Debt", "Ending Cash", "Net Debt",
        "Net Debt / EBITDA", "Max Leverage", "Leverage Breach", "Cash Interest",
        "Interest Coverage", "Min Coverage", "Coverage Breach",
    ])
    row += 1
    for j, yr in enumerate(years):
        dcol = get_column_letter(2 + j)
        r = row
        _cell(ws, r, 1, yr, fmt=FMT_YEAR)
        _formula(ws, r, 2, f"={of_sheet}!{dcol}{ebitda_row}", FMT_MONEY, cross=True)
        _formula(ws, r, 3, f"={ds_sheet}!{dcol}{ted_row}", FMT_MONEY, cross=True)
        _formula(ws, r, 4, f"={ds_sheet}!{dcol}{cash_row}", FMT_MONEY, cross=True)
        _formula(ws, r, 5, f"=C{r}-D{r}", FMT_MONEY)
        # Net Debt / EBITDA, with the engine's net-cash convention.
        _formula(ws, r, 6, f'=IF(E{r}<0,"Net Cash",IFERROR(E{r}/B{r},"n/a"))', FMT_MULT2)
        if have_lev:
            _formula(ws, r, 7, f"={N_MAX_LEVERAGE}", FMT_MULT2)
            # Leverage breach: Net Debt > Max Leverage x EBITDA (avoids a division
            # so a zero-EBITDA year cannot raise a spurious #DIV/0! breach).
            _formula(ws, r, 8,
                     f'=IF(AND(B{r}>{_EPS_LIT},E{r}>{N_MAX_LEVERAGE}*B{r}+{_EPS_LIT}),"Yes","No")')
        else:
            _cell(ws, r, 7, "n/a", align="left")
            _cell(ws, r, 8, "n/a", align="left")
        _formula(ws, r, 9, f"={ds_sheet}!{dcol}{int_row}", FMT_MONEY, cross=True)
        # Interest coverage = EBITDA / Cash Interest; non-positive interest is n/a.
        _formula(ws, r, 10, f'=IF(I{r}>{_EPS_LIT},IFERROR(B{r}/I{r},"n/a"),"n/a")', FMT_MULT2)
        if have_cov:
            _formula(ws, r, 11, f"={N_MIN_COVERAGE}", FMT_MULT2)
            # Coverage breach: EBITDA < Min Coverage x Cash Interest.
            _formula(ws, r, 12,
                     f'=IF(AND(I{r}>{_EPS_LIT},B{r}<{N_MIN_COVERAGE}*I{r}-{_EPS_LIT}),"Yes","No")')
        else:
            _cell(ws, r, 11, "n/a", align="left")
            _cell(ws, r, 12, "n/a", align="left")
        row += 1
    row += 1
    _note(ws, row, 1,
          "Leverage and coverage ratios are live Excel formulas linked to the Debt "
          "Schedule (total ending debt, ending cash, total cash interest) and the "
          "Operating Forecast EBITDA; the thresholds above are editable inputs and "
          "pass/fail uses Excel IF(). Net cash years show 'Net Cash'; non-positive "
          "interest shows 'n/a'. Detection only — no cure, reset or refinancing is "
          "modeled. Python run_lbo() remains the covenant oracle for the exported case.",
          end_col=12)
    ws.freeze_panes = "A3"


# ── Maturity Wall (V4.1.2 formula-linked, multi-tranche) ─────────────────────


def _build_maturity_wall_formula(ws, wb, result, settings, per_tranche) -> None:
    """Per-tranche maturity wall whose ending balance at exit and share of total
    debt are live formulas referencing the formula-native Debt Schedule final
    column, so editing tranche opening balances / amortization / cash sweep moves
    the wall. Within-hold flag compares each maturity year to the holding
    period."""
    _title(ws, "Maturity Wall", span=6)
    _set_widths(ws, {"A": 22, "B": 14, "C": 22, "D": 18, "E": 18, "F": 44})
    of = result["operating_forecast"]
    n = len(of["years"])
    last_col = get_column_letter(2 + n - 1)
    ds_sheet = quote_sheetname("Debt Schedule")
    tranches = settings["tranches"]

    row = 3
    _header_row(ws, row, [
        "Tranche", "Maturity Year", "Ending Balance at Exit",
        "Share of Total Debt", "Within Hold Period?", "Notes",
    ])
    row += 1
    for t in tranches:
        r = row
        end_row = per_tranche[t["id"]]["Ending Balance"]
        _cell(ws, r, 1, t["name"], bold=True, align="left")
        _input_cell(ws, r, 2, t["maturity_year"], fmt=FMT_YEAR)
        _formula(ws, r, 3, f"={ds_sheet}!{last_col}{end_row}", FMT_MONEY, cross=True)
        _formula(ws, r, 4, f"=IFERROR(C{r}/{N_TOTAL_ENDING_DEBT},0)", FMT_PCT)
        _formula(ws, r, 5, f'=IF(B{r}<={N_HOLDING},"Yes","No")')
        _formula(ws, r, 6,
                 f'=IF(B{r}<={N_HOLDING},"Matures within hold period — refinancing not modeled","")')
        row += 1
    _cell(ws, row, 1, "Total", bold=True, align="left")
    _formula(ws, row, 3, f"={N_TOTAL_ENDING_DEBT}", FMT_MONEY, cross=True)
    _formula(ws, row, 4, f"=IFERROR(C{row}/{N_TOTAL_ENDING_DEBT},0)", FMT_PCT)
    row += 2
    _note(ws, row, 1,
          "Each tranche's debt outstanding at exit is the final Debt Schedule column "
          "(outstanding, not original face value); ending balance and share of total "
          "debt are live formulas, so they move when tranche opening balances, "
          "amortization or the cash sweep change. The within-hold flag compares each "
          "maturity year (editable) to the holding period. Refinancing and maturity "
          "default are not modeled.", end_col=6)
    ws.freeze_panes = "A4"


# ── Returns Attribution + formula tie-out checks (V4.1.2) ─────────────────────


def _build_attribution_formula(ws, wb, inputs, result, is_multi) -> None:
    """Returns Attribution sheet augmented with a live formula tie-out checks
    block (Sources = Uses, Sponsor Equity, Exit Equity bridge, MOIC, debt
    paydown, ending cash) that references the workbook named-range chain rather
    than Python literals, plus — when attribution detail is available — a live
    component-sum-vs-bridge check so a stale or tampered component is flagged
    instead of being silently absorbed by the residual."""
    _title(ws, "Returns Attribution", span=5)
    _set_widths(ws, {"A": 38, "B": 18, "C": 20, "D": 14, "E": 60})
    rs = quote_sheetname("Returns Summary")
    tol = "0.01"

    row = 3
    _section(ws, row, "Formula Tie-out Checks", span=5)
    row += 1
    _header_row(ws, row, ["Check", "Left", "Right", "Diff", "Result"])
    row += 1

    def _eq_check(label, left, right):
        nonlocal row
        r = row
        _cell(ws, r, 1, label, align="left")
        _formula(ws, r, 2, left, FMT_MONEY)
        _formula(ws, r, 3, right, FMT_MONEY)
        _formula(ws, r, 4, f"=B{r}-C{r}", FMT_MONEY)
        _formula(ws, r, 5, f'=IF(ABS(D{r})<{tol},"OK","CHECK")')
        row += 1

    def _ge_check(label, left):
        nonlocal row
        r = row
        _cell(ws, r, 1, label, align="left")
        _formula(ws, r, 2, left, FMT_MONEY)
        _cell(ws, r, 3, 0.0, fmt=FMT_MONEY)
        _formula(ws, r, 4, f"=B{r}-C{r}", FMT_MONEY)
        _formula(ws, r, 5, f'=IF(B{r}>=-{tol},"OK","CHECK")')
        row += 1

    _eq_check("Sources = Uses", f"={N_TOTAL_DEBT}+{N_SPONSOR_EQUITY}", f"={N_TOTAL_USES}")
    _eq_check("Sponsor Equity = Uses - Debt", f"={N_SPONSOR_EQUITY}",
              f"={N_TOTAL_USES}-{N_TOTAL_DEBT}")
    _eq_check("Exit Equity = EV - Debt + Cash", f"={N_EXIT_EQUITY}",
              f"={rs}!B6-{rs}!B7+{rs}!B8")
    _eq_check("MOIC x Sponsor = Exit Equity", f"={N_MOIC}*{N_SPONSOR_EQUITY}",
              f"={N_EXIT_EQUITY}")
    _ge_check("Debt Paydown >= 0", f"={N_OPENING_DEBT}-{rs}!B7")
    if is_multi:
        _eq_check("Ending Debt: Returns = Schedule", f"={rs}!B7", f"={N_TOTAL_ENDING_DEBT}")
        _ge_check("Ending Cash >= 0", f"={N_ENDING_CASH}")
    row += 1

    attribution = inputs.get("attribution") or {}
    if attribution.get("status") != "ok":
        _section(ws, row, "Equity Value Bridge", span=5)
        row += 1
        _note(ws, row, 1,
              "Component-level attribution is unavailable for this run (it requires a "
              "valid result with positive sponsor equity and exit value). The formula "
              "tie-out checks above remain live and reference the workbook named-range "
              "chain.", end_col=5)
        ws.freeze_panes = "A3"
        return

    _section(ws, row, "Equity Value Bridge", span=5)
    row += 1
    _header_row(ws, row, ["Component", "Value", "MOIC Contribution", "Direction", "Rationale"])
    row += 1
    comp_value_rows: list[int] = []
    for comp in attribution.get("components") or []:
        _cell(ws, row, 1, comp.get("label_en") or comp.get("label_cn"), align="left")
        _cell(ws, row, 2, comp.get("value"), fmt=FMT_MONEY)
        comp_value_rows.append(row)
        _cell(ws, row, 3, comp.get("moic_contribution"), fmt="0.0000x")
        _cell(ws, row, 4, comp.get("direction"), align="left")
        _cell(ws, row, 5, comp.get("rationale_en") or comp.get("rationale_cn"), wrap=True)
        row += 1
    notes = attribution.get("notes") or {}
    _cell(ws, row, 1, "Note", bold=True, align="left")
    _note(ws, row, 2, notes.get("moic_contribution_note_en") or MOIC_CONTRIBUTION_NOTE_EN, end_col=5)
    row += 1
    _note(ws, row, 2, notes.get("moic_contribution_note_cn") or MOIC_CONTRIBUTION_NOTE_CN, end_col=5)
    row += 2

    _section(ws, row, "Attribution Tie-out (formula)", span=3)
    row += 1
    _cell(ws, row, 1, "Equity Value Creation (live bridge)", bold=True, align="left")
    live_row = row
    _formula(ws, row, 2, f"={N_EXIT_EQUITY}-{N_SPONSOR_EQUITY}", FMT_MONEY)
    row += 1
    _cell(ws, row, 1, "Attribution Component Sum (incl. residual)", bold=True, align="left")
    sum_row = row
    if comp_value_rows:
        _formula(ws, row, 2, "=" + "+".join(f"B{r}" for r in comp_value_rows), FMT_MONEY)
    else:
        _cell(ws, row, 2, 0.0, fmt=FMT_MONEY)
    row += 1
    _cell(ws, row, 1, "Component Sum vs Live Bridge", bold=True, align="left")
    _formula(ws, row, 2, f"=B{sum_row}-B{live_row}", FMT_MONEY)
    _formula(ws, row, 3, f'=IF(ABS(B{row})<{tol},"OK","CHECK")')
    row += 2

    tie = attribution.get("tie_out") or {}
    _section(ws, row, "Python Tie-out (reference)", span=2)
    row += 1
    row = _write_pairs(ws, row, [
        ("Attribution Method", attribution.get("method")),
        ("Target Equity Value Creation", tie.get("target_equity_value_creation"), FMT_MONEY),
        ("Component Sum Before Residual", tie.get("component_sum_before_residual"), FMT_MONEY),
        ("Residual", tie.get("residual"), FMT_MONEY),
        ("Mathematical Tie-out Pass", _yn(tie.get("mathematical_tie_out_pass"))),
        ("Directional Bridge Pass", _yn(tie.get("directional_bridge_pass"))),
    ])
    row += 1
    _note(ws, row, 1,
          "The live bridge equals exit equity minus sponsor equity from the formula "
          "chain; the component sum (including the residual safety valve) ties to it, "
          "so a tampered or stale component is flagged rather than silently absorbed. "
          "Python build_lbo_attribution remains the component-level oracle.", end_col=5)
    row += 2
    _section(ws, row, "Disclosure", span=5)
    row += 1
    _note(ws, row, 1, notes.get("disclosure_en") or
          "Simplified equity value bridge. It explains model return sources "
          "directionally, does not split component-level IRR, and is not an "
          "investment or acquisition recommendation.", end_col=5)
    ws.freeze_panes = "A3"


# ── Scenario Summary (V4.1.3 oracle-backed + formula-checked) ────────────────

# The seven attribution-contribution metric keys that, by construction, must sum
# to a scenario's equity value creation (exit equity - sponsor equity). Mirrors
# tests/test_lbo_scenario_cash_funding_tieout._CONTRIB_KEYS, but here the sum is a
# live Excel formula so a silent mismatch surfaces as CHECK.
_SCENARIO_CONTRIB_KEYS = [
    "ebitda_growth_contribution",
    "multiple_movement_contribution",
    "deleveraging_contribution",
    "fees_drag",
    "ending_cash_balance_contribution",
    "initial_cash_funding_contribution",
    "residual",
]
# Comparison table column layout: Base = col B, Upside = col C, Downside = col D
# (fixed by _build_scenario_summary's render loop).
_SCENARIO_COLS = (("base", 2), ("upside", 3), ("downside", 4))


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _scenario_assumption_map(ws, row: int, scenario_data: dict) -> int:
    """Values-only 'Scenario Assumption Map': the key assumptions and the
    deterministic deltas behind each Base / Upside / Downside case. Sourced from
    the attached full ``build_lbo_scenarios`` payload (per-scenario inputs +
    scenario_config). Degrades gracefully to a note when only a bare comparison
    dict was attached."""
    scenarios = scenario_data.get("scenarios") or {}
    config = scenario_data.get("scenario_config") or {}
    _section(ws, row, "Scenario Assumption Map", span=4)
    row += 1
    _header_row(ws, row, ["Assumption", SCENARIO_LABELS["base"],
                          SCENARIO_LABELS["upside"], SCENARIO_LABELS["downside"]])
    row += 1
    if not scenarios:
        _note(ws, row, 1,
              "Per-scenario assumption detail is unavailable for this export (only "
              "the comparison summary was attached). The output bridge and tie-out "
              "checks below remain available.", end_col=4)
        return row + 2

    order = ["base", "upside", "downside"]

    def inp(key: str) -> dict:
        return (scenarios.get(key) or {}).get("inputs") or {}

    def tx(key: str, field: str) -> Any:
        return (inp(key).get("transaction") or {}).get(field)

    def y1_ebitda(key: str) -> Any:
        vals = (inp(key).get("operating_forecast") or {}).get("ebitda") or []
        return vals[0] if vals else None

    def interest(key: str) -> Any:
        d = inp(key)
        cs = d.get("capital_structure") or {}
        if cs.get("mode") == "multi_tranche":
            tranches = cs.get("tranches") or []
            return tranches[0].get("interest_rate") if tranches else None
        return (d.get("debt") or {}).get("interest_rate")

    def cash_bs(key: str) -> Any:
        return (inp(key).get("debt") or {}).get("cash_to_balance_sheet", 0.0)

    def delta(key: str, dkey: str) -> float:
        if key == "base":
            return 0.0
        return float((config.get(key) or {}).get(dkey, 0.0) or 0.0)

    value_rows = [
        ("Entry Multiple (held constant)", lambda k: tx(k, "entry_multiple"), FMT_MULT1),
        ("Exit Multiple", lambda k: tx(k, "exit_multiple"), FMT_MULT1),
        ("Year 1 EBITDA", y1_ebitda, FMT_MONEY),
        ("Interest Rate (tranche 1 / debt)", interest, FMT_PCT),
        ("Cash to Balance Sheet", cash_bs, FMT_MONEY),
    ]
    for label, fn, fmt in value_rows:
        _cell(ws, row, 1, label, bold=True, align="left")
        for i, key in enumerate(order):
            _cell(ws, row, 2 + i, fn(key), fmt=fmt)
        row += 1

    delta_rows = [
        ("EBITDA Delta %", "ebitda_pct_delta", FMT_PCT),
        ("Exit Multiple Delta", "exit_multiple_delta", FMT_MULT1),
        ("Interest Rate Delta", "interest_rate_delta", FMT_PCT),
        ("CapEx Delta %", "capex_pct_delta", FMT_PCT),
        ("Change in NWC Delta %", "nwc_pct_delta", FMT_PCT),
    ]
    for label, dkey, fmt in delta_rows:
        _cell(ws, row, 1, label, align="left")
        for i, key in enumerate(order):
            _cell(ws, row, 2 + i, delta(key, dkey), fmt=fmt)
        row += 1
    _note(ws, row, 1,
          "Scenario deltas are deterministic modeling shocks applied to the Base "
          "inputs (same transaction: entry valuation, opening debt and sponsor "
          "equity held constant). Values are Python-precomputed; they are not "
          "forecasts, probabilities or recommendations.", end_col=4)
    return row + 2


def _scenario_output_bridge(ws, row: int, key_row: dict, ok_cols: list) -> tuple[int, dict]:
    """'Scenario Output Bridge': the headline outputs per available scenario,
    written as live formulas that reference the comparison table cells, plus a
    live Equity Value Creation = Exit Equity - Sponsor Equity row. Returns the
    next row and a ``{metric_key: bridge_row}`` map for the tie-out checks."""
    _section(ws, row, "Scenario Output Bridge", span=4)
    row += 1
    _header_row(ws, row, ["Output", SCENARIO_LABELS["base"],
                          SCENARIO_LABELS["upside"], SCENARIO_LABELS["downside"]])
    row += 1
    bridge_rows: dict[str, int] = {}
    metrics = [
        ("Sponsor Equity", "sponsor_equity", FMT_MONEY),
        ("Exit Equity Value", "exit_equity_value", FMT_MONEY),
        ("MOIC", "moic", FMT_MULT2),
        ("IRR", "irr", FMT_PCT),
        ("Remaining Debt", "remaining_debt", FMT_MONEY),
        ("Ending Cash Balance", "ending_cash_balance", FMT_MONEY),
        ("Debt Paydown", "debt_paydown", FMT_MONEY),
        ("Covenant Status", "covenant_status", None),
    ]
    for label, mkey, fmt in metrics:
        _cell(ws, row, 1, label, bold=True, align="left")
        src = key_row.get(mkey)
        if src is not None:
            for _name, col in ok_cols:
                letter = get_column_letter(col)
                _formula(ws, row, col, f"={letter}{src}", fmt)
        bridge_rows[mkey] = row
        row += 1
    _cell(ws, row, 1, "Equity Value Creation", bold=True, align="left")
    for _name, col in ok_cols:
        letter = get_column_letter(col)
        _formula(ws, row, col,
                 f"={letter}{bridge_rows['exit_equity_value']}"
                 f"-{letter}{bridge_rows['sponsor_equity']}", FMT_MONEY)
    bridge_rows["equity_value_creation"] = row
    row += 1
    _note(ws, row, 1,
          "Output bridge cells are live formulas referencing the scenario "
          "comparison values above; Equity Value Creation = Exit Equity - Sponsor "
          "Equity. run_lbo() remains the per-scenario oracle.", end_col=4)
    return row + 2, bridge_rows


def _scenario_tieout_checks(ws, row: int, key_row: dict, row_value: dict,
                            bridge_rows: dict, ok_cols: list, is_multi_formula: bool,
                            cash_to_bs: float) -> int:
    """Live formula tie-out checks: per-scenario MOIC x Sponsor = Exit Equity,
    Debt Paydown = Opening - Remaining, and (when full attribution is available)
    Component Sum = Equity Value Creation; plus Base-case-ties-to-main-exported-
    case checks and the initial-cash-funding tie-out. Each check is an Excel
    IF(ABS(diff)<tol) that resolves to OK or CHECK."""
    tol = "0.01"
    _section(ws, row, "Scenario Tie-out Checks", span=5)
    row += 1
    _header_row(ws, row, ["Check", "Left", "Right", "Diff", "Result"])
    row += 1

    def eq_check(label, left, right, fmt=FMT_MONEY):
        nonlocal row
        r = row
        _cell(ws, r, 1, label, align="left")
        _formula(ws, r, 2, left, fmt)
        _formula(ws, r, 3, right, fmt)
        _formula(ws, r, 4, f"=B{r}-C{r}", fmt)
        _formula(ws, r, 5, f'=IF(ABS(D{r})<{tol},"OK","CHECK")')
        row += 1

    # Per available scenario: MOIC x Sponsor Equity = Exit Equity Value.
    for name, col in ok_cols:
        letter = get_column_letter(col)
        eq_check(f"{SCENARIO_LABELS[name]}: MOIC x Sponsor = Exit Equity",
                 f"={letter}{bridge_rows['moic']}*{letter}{bridge_rows['sponsor_equity']}",
                 f"={letter}{bridge_rows['exit_equity_value']}")

    # Per available scenario: Debt Paydown = Opening Debt - Remaining Debt. Opening
    # debt is held constant across scenarios, so it references the main named range.
    for name, col in ok_cols:
        letter = get_column_letter(col)
        eq_check(f"{SCENARIO_LABELS[name]}: Debt Paydown = Opening - Remaining",
                 f"={N_OPENING_DEBT}-{letter}{bridge_rows['remaining_debt']}",
                 f"={letter}{bridge_rows['debt_paydown']}")

    # Per available scenario with full attribution: component sum = value creation.
    for name, col in ok_cols:
        if not all(
            _is_number((row_value.get(k) or {}).get(name)) and k in key_row
            for k in _SCENARIO_CONTRIB_KEYS
        ):
            continue
        letter = get_column_letter(col)
        comp_sum = "+".join(f"{letter}{key_row[k]}" for k in _SCENARIO_CONTRIB_KEYS)
        eq_check(f"{SCENARIO_LABELS[name]}: Component Sum = Value Creation",
                 f"={comp_sum}",
                 f"={letter}{bridge_rows['equity_value_creation']}")

    # Base case ties to the main exported workbook formula chain (same inputs).
    base_ok = any(name == "base" for name, _ in ok_cols)
    if base_ok:
        eq_check("Base MOIC = Workbook MOIC",
                 f"=B{bridge_rows['moic']}", f"={N_MOIC}", fmt=FMT_MULT2)
        eq_check("Base Sponsor Equity = Workbook",
                 f"=B{bridge_rows['sponsor_equity']}", f"={N_SPONSOR_EQUITY}")
        eq_check("Base Exit Equity = Workbook",
                 f"=B{bridge_rows['exit_equity_value']}", f"={N_EXIT_EQUITY}")
        if is_multi_formula:
            eq_check("Base Remaining Debt = Schedule",
                     f"=B{bridge_rows['remaining_debt']}", f"={N_TOTAL_ENDING_DEBT}")

    # Initial cash funding contribution = -Cash to Balance Sheet (when funded).
    icf_key = "initial_cash_funding_contribution"
    if (cash_to_bs and cash_to_bs > 0 and base_ok and icf_key in key_row
            and _is_number((row_value.get(icf_key) or {}).get("base"))):
        eq_check("Initial Cash Funding = -Cash to BS",
                 f"=B{key_row[icf_key]}", f"=-{N_CASH_TO_BS}")
    row += 1
    _note(ws, row, 1,
          "Checks are live Excel formulas referencing the scenario comparison "
          "values, the output bridge and the main workbook named-range chain; a "
          "broken tie-out surfaces as CHECK rather than a silent mismatch. The "
          "Base column ties to the main exported case because both run the same "
          "inputs through run_lbo().", end_col=5)
    return row + 2


def _scenario_ic_review(ws, row: int, bridge_rows: dict, ok_cols: list) -> int:
    """'IC Review Notes': live formula summaries describing scenario spread and
    downside protection for model audit / IC defense. Not a recommendation."""
    have = {name for name, _ in ok_cols}
    _section(ws, row, "IC Review Notes (model audit)", span=4)
    row += 1
    if "upside" in have and "downside" in have:
        _cell(ws, row, 1, "Scenario IRR Spread (Upside - Downside)", bold=True, align="left")
        _formula(ws, row, 2, f"=C{bridge_rows['irr']}-D{bridge_rows['irr']}", FMT_PCT)
        row += 1
    if "downside" in have:
        _cell(ws, row, 1, "Downside MOIC", bold=True, align="left")
        _formula(ws, row, 2, f"=D{bridge_rows['moic']}", FMT_MULT2)
        row += 1
        _cell(ws, row, 1, "Downside Remaining Debt", bold=True, align="left")
        _formula(ws, row, 2, f"=D{bridge_rows['remaining_debt']}", FMT_MONEY)
        row += 1
    if "base" in have:
        _cell(ws, row, 1, "Base Equity Value Creation", bold=True, align="left")
        _formula(ws, row, 2, f"=B{bridge_rows['equity_value_creation']}", FMT_MONEY)
        row += 1
    _note(ws, row, 1,
          "IC review summaries are live formulas for model audit / defense only. "
          "They describe scenario spread and downside protection; they are not "
          "investment or acquisition recommendations.", end_col=4)
    return row + 2


def _build_scenario_summary_formula(ws, scenario_data: dict, is_multi_formula: bool,
                                    cash_to_bs: float) -> None:
    """V4.1.3 Scenario Summary: the existing values-only Base / Upside / Downside
    comparison (Python oracle output), augmented with an assumption map, an
    output bridge, live formula tie-out checks and an IC review block. The
    comparison metric values themselves stay Python-precomputed; only the bridge,
    checks and IC review are formulas."""
    _build_scenario_summary(ws, scenario_data)

    comparison = scenario_data.get("comparison") or {}
    rows = comparison.get("rows") or []
    statuses = comparison.get("scenario_statuses") or {}
    # Header at row 3; comparison data rows run from row 4 in schema order.
    key_row = {r.get("key"): 4 + i for i, r in enumerate(rows)}
    row_value = {r.get("key"): r for r in rows}
    ok_cols = [
        (name, col) for name, col in _SCENARIO_COLS
        if (statuses.get(name) or {}).get("status") == "ok"
    ]

    r = ws.max_row + 2
    r = _scenario_assumption_map(ws, r, scenario_data)
    if not ok_cols:
        _note(ws, r, 1,
              "No scenario produced valid LBO returns, so the output bridge and "
              "formula tie-out checks are unavailable for this export.", end_col=5)
        ws.freeze_panes = "A4"
        return
    r, bridge_rows = _scenario_output_bridge(ws, r, key_row, ok_cols)
    r = _scenario_tieout_checks(ws, r, key_row, row_value, bridge_rows, ok_cols,
                                is_multi_formula, cash_to_bs)
    r = _scenario_ic_review(ws, r, bridge_rows, ok_cols)
    ws.freeze_panes = "A4"


# ── Sensitivity Analysis (V4.1.3 oracle-backed + formula-checked) ────────────


def _sensitivity_axis_map(ws, row: int, grids: dict, specs: list) -> int:
    """Values-only metadata / axis map declaring each grid's row / column
    variable, the swept values and the displayed metric."""
    _section(ws, row, "Sensitivity Metadata / Axis Map", span=8)
    row += 1
    _header_row(ws, row, ["Grid", "Axis", "Variable", "Values", "Metric"])
    row += 1
    for key, title in specs:
        grid = grids.get(key) or {}
        rvals = ", ".join(f"{v:.1f}x" for v in (grid.get("rows") or []))
        cvals = ", ".join(f"{v:.1f}x" for v in (grid.get("cols") or []))
        _cell(ws, row, 1, title, bold=True, align="left")
        _cell(ws, row, 2, "Row", align="left")
        _cell(ws, row, 3, grid.get("row_label"), align="left")
        _cell(ws, row, 4, rvals, align="left")
        _cell(ws, row, 5, "IRR / MOIC", align="left")
        row += 1
        _cell(ws, row, 2, "Column", align="left")
        _cell(ws, row, 3, grid.get("col_label"), align="left")
        _cell(ws, row, 4, cvals, align="left")
        row += 1
    _note(ws, row, 1,
          "Axis map declares each grid's row / column variable, the swept values "
          "and the displayed metric. The grid checks below verify the rendered "
          "grid matches these dimensions and that the base case sits within the "
          "IRR range.", end_col=8)
    return row + 2


def _sensitivity_grid_checks(ws, row: int, start_row: int, title: str,
                             grid: dict) -> int:
    """Live formula checks over a single (already-rendered) sensitivity grid:
    dimensions match the axis metadata, and best / base / worst IRR summaries with
    a base-within-range sanity check. References only the OK numeric IRR cells, so
    unavailable cells never poison a MIN / MAX."""
    rows_vals = grid.get("rows") or []
    cols_vals = grid.get("cols") or []
    cells = grid.get("cells") or []
    # The IRR data block: section at start_row, header at start_row+1, first IRR
    # data row at start_row+2 (matches _build_sensitivity_grid's layout).
    irr_first = start_row + 2
    irr_coords: list[str] = []
    base_coord: str | None = None
    for r_idx in range(len(rows_vals)):
        line = cells[r_idx] if r_idx < len(cells) else []
        for c_idx in range(len(cols_vals)):
            cell = line[c_idx] if c_idx < len(line) else {}
            if isinstance(cell, dict) and cell.get("status") == "ok" and _is_number(cell.get("irr")):
                coord = f"{get_column_letter(2 + c_idx)}{irr_first + r_idx}"
                irr_coords.append(coord)
                if cell.get("is_base"):
                    base_coord = coord

    _section(ws, row, f"{title} - Grid Checks (formula)", span=5)
    row += 1
    n_rows = len(rows_vals)
    n_cols = len(cols_vals)
    present = sum(len(line) for line in cells)
    _cell(ws, row, 1, "Row Count", align="left")
    _cell(ws, row, 2, n_rows, fmt=FMT_YEAR)
    rc = row
    row += 1
    _cell(ws, row, 1, "Column Count", align="left")
    _cell(ws, row, 2, n_cols, fmt=FMT_YEAR)
    cc = row
    row += 1
    _cell(ws, row, 1, "Cells Present", align="left")
    _cell(ws, row, 2, present, fmt=FMT_YEAR)
    pr = row
    row += 1
    _cell(ws, row, 1, "Rows x Cols", align="left")
    _formula(ws, row, 2, f"=B{rc}*B{cc}", FMT_YEAR)
    prod = row
    row += 1
    _cell(ws, row, 1, "Dimensions Match", align="left")
    _formula(ws, row, 2, f'=IF(ABS(B{prod}-B{pr})<0.5,"OK","CHECK")')
    row += 1

    if irr_coords:
        joined = ",".join(irr_coords)
        _cell(ws, row, 1, "Best IRR (max)", align="left")
        _formula(ws, row, 2, f"=MAX({joined})", FMT_PCT)
        maxr = row
        row += 1
        _cell(ws, row, 1, "Worst IRR (min)", align="left")
        _formula(ws, row, 2, f"=MIN({joined})", FMT_PCT)
        minr = row
        row += 1
        if base_coord:
            _cell(ws, row, 1, "Base Case IRR", align="left")
            _formula(ws, row, 2, f"={base_coord}", FMT_PCT)
            br = row
            row += 1
            _cell(ws, row, 1, "Base Within IRR Range", align="left")
            _formula(ws, row, 2,
                     f'=IF(AND(B{minr}<=B{br}+{_EPS_LIT},B{br}<=B{maxr}+{_EPS_LIT}),"OK","CHECK")')
            row += 1
    else:
        _note(ws, row, 1,
              "No numeric IRR cells available in this grid for formula summaries.",
              end_col=5)
        row += 1
    return row + 2


def _build_sensitivity_summary_formula(ws, sensitivity: dict) -> None:
    """V4.1.3 Sensitivity Analysis: the existing Python-precomputed IRR / MOIC
    grids (oracle output, unchanged math) plus a metadata / axis map and live
    formula checks (grid dimensions, best / base / worst IRR). The grids are not
    live Excel Data Tables."""
    _title(ws, SENSITIVITY_SHEET, span=8)
    _set_widths(ws, {"A": 24, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 22})
    base = sensitivity.get("base") or {}
    row = 3
    row = _write_pairs(ws, row, [
        ("Method", sensitivity.get("method")),
        ("Base Entry Multiple", base.get("entry_multiple"), FMT_MULT1),
        ("Base Exit Multiple", base.get("exit_multiple"), FMT_MULT1),
        ("Base Debt / EBITDA", base.get("leverage"), FMT_MULT1),
    ])
    row += 1
    _note(ws, row, 1,
          "Deterministic sensitivity; not probability-weighted. IRR / MOIC values "
          "are Python run_lbo oracle output. V4.1.3 adds an axis map and live "
          "formula checks (grid dimensions, best / base / worst IRR) over the "
          "precomputed grid; the grid itself is not a live Excel Data Table.",
          end_col=8)
    row += 2
    grids = sensitivity.get("grids") or {}
    specs = [
        ("entry_exit", "Entry Multiple x Exit Multiple"),
        ("leverage_exit", "Debt / EBITDA x Exit Multiple"),
    ]
    row = _sensitivity_axis_map(ws, row, grids, specs)
    for key, title in specs:
        grid = grids.get(key) or {}
        start_row = row
        row = _build_sensitivity_grid(ws, start_row, title, grid)
        row = _sensitivity_grid_checks(ws, row, start_row, title, grid)
    ws.freeze_panes = "B9"


# ── Orchestration ────────────────────────────────────────────────────────────


def generate_lbo_formula_excel(inputs: dict[str, Any], result: dict[str, Any]) -> BytesIO:
    """Build the formula-native LBO workbook (V4.1.1; multi-tranche waterfall).

    ``run_lbo()`` output (``result``) is the oracle for the exported case; the
    core sheets express the same relationships as Excel formulas. Raises
    ``ValueError`` for a non-ok result, matching the legacy exporter contract.
    """
    if result.get("status") != "ok" or not result.get("returns"):
        raise ValueError("Cannot export LBO workbook for non-ok result.")

    cap_summary = result.get("capital_structure_summary") or {}
    is_multi = cap_summary.get("mode") == "multi_tranche"

    # Re-validate the capital structure to recover normalized tranche fields
    # (commitment, mandatory amortization %, sweep priority, revolver flag) that
    # the formula chain needs but that the summary does not expose. run_lbo()
    # already validated successfully (result is ok), so this is a pure re-parse,
    # not a second source of truth. Fall back to values-only if it ever fails.
    settings = None
    if is_multi:
        settings, cs_flags = validate_capital_structure(inputs.get("capital_structure") or {})
        if settings is None:
            is_multi_formula = False
        else:
            is_multi_formula = True
    else:
        is_multi_formula = False

    scenario_data = _scenario_comparison(inputs)
    sensitivity = _sensitivity_data(inputs)

    wb = Workbook()
    ws_cover = wb.active
    ws_cover.title = SHEET_NAMES[0]
    sheets = {SHEET_NAMES[0]: ws_cover}
    for name in SHEET_NAMES[1:]:
        sheets[name] = wb.create_sheet(name)

    if scenario_data is not None:
        sheets[SCENARIO_SUMMARY_SHEET] = wb.create_sheet(
            SCENARIO_SUMMARY_SHEET, _SCENARIO_INSERT_INDEX
        )
    if sensitivity is not None:
        insert_at = _SCENARIO_INSERT_INDEX + (1 if scenario_data is not None else 0)
        sheets[SENSITIVITY_SHEET] = wb.create_sheet(SENSITIVITY_SHEET, insert_at)

    # Cover / Model Map / Audit.
    _build_cover(sheets["Cover"], inputs, result, is_multi,
                 formula_native=True, version=FORMULA_WORKBOOK_VERSION)
    _build_model_map(sheets["Model Map"], scenario_data is not None, sensitivity is not None,
                     formula_native=True)

    # V4.1.3: Scenario Summary / Sensitivity Analysis remain Python oracle-backed
    # precomputed grids, but are now augmented with assumption mapping, an output
    # bridge, live formula tie-out / IC review checks (scenario) and an axis map +
    # grid formula checks (sensitivity). The underlying grid math is unchanged.
    cash_to_bs = result.get("transaction_summary", {}).get("cash_to_balance_sheet", 0.0) or 0.0
    if scenario_data is not None:
        _build_scenario_summary_formula(sheets[SCENARIO_SUMMARY_SHEET], scenario_data,
                                        is_multi_formula, cash_to_bs)
    if sensitivity is not None:
        _build_sensitivity_summary_formula(sheets[SENSITIVITY_SHEET], sensitivity)

    # Formula-native core.
    _build_assumptions_formula(sheets["Assumptions"], wb, inputs, result, is_multi,
                               cap_summary, settings)
    _build_sources_uses_formula(sheets["Sources & Uses"], wb, inputs, result, is_multi, cap_summary)
    if is_multi_formula:
        # V4.1.1: multi-tranche operating forecast + tranche-level waterfall are
        # now formula-native. The Operating Forecast tax shield references the
        # Debt Schedule's Total Cash Interest row.
        per_tranche, totals = _multi_ds_layout(settings["tranches"])
        _build_operating_forecast_formula(sheets["Operating Forecast"], result,
                                          ds_interest_row=totals["Total Cash Interest"])
        _build_debt_schedule_formula_multi(sheets["Debt Schedule"], wb, result,
                                           settings, per_tranche, totals)
        # V4.1.2: Covenant Check and Maturity Wall are now formula-linked to the
        # tranche-level Debt Schedule / Operating Forecast above.
        _build_covenant_check_formula(sheets["Covenant Check"], wb, result, totals)
        _build_maturity_wall_formula(sheets["Maturity Wall"], wb, result, settings, per_tranche)
    elif is_multi:
        # Defensive fallback: capital structure failed re-validation; keep the
        # values-only multi-tranche presentation rather than emitting no schedule.
        _build_operating_forecast(sheets["Operating Forecast"], result, is_multi)
        _build_debt_schedule(sheets["Debt Schedule"], result, is_multi)
        _build_covenant_check(sheets["Covenant Check"], result, is_multi)
        _build_maturity_wall(sheets["Maturity Wall"], result, is_multi)
    else:
        _build_operating_forecast_formula(sheets["Operating Forecast"], result)
        _build_debt_schedule_formula(sheets["Debt Schedule"], result)
        # Single-tranche: covenants / maturity wall are not modeled by the engine;
        # keep the values-only "Not Applicable" disclosure sheets.
        _build_covenant_check(sheets["Covenant Check"], result, is_multi)
        _build_maturity_wall(sheets["Maturity Wall"], result, is_multi)
    _build_returns_summary_formula(sheets["Returns Summary"], wb, result, is_multi_formula)

    # V4.1.2: Returns Attribution carries live formula tie-out checks referencing
    # the workbook named-range chain (works in both modes, with or without
    # precomputed attribution detail).
    _build_attribution_formula(sheets["Returns Attribution"], wb, inputs, result, is_multi_formula)
    _build_audit(sheets["Audit & Disclosures"], inputs, result, is_multi, cap_summary,
                 formula_native=True)

    for name in ("Assumptions", "Sources & Uses", "Operating Forecast",
                 "Debt Schedule", "Returns Summary"):
        _append_units_note(sheets[name], result)
    if scenario_data is not None:
        _append_units_note(sheets[SCENARIO_SUMMARY_SHEET], result)
    if sensitivity is not None:
        _append_units_note(sheets[SENSITIVITY_SHEET], result)

    # V4.1.4: deterministic banker print setup on every sheet (fit-to-one-page-
    # wide; landscape for the wide multi-column sheets, portrait for the narrow
    # label/value sheets). Presentation only — no effect on the formula chain.
    for name, ws in sheets.items():
        _apply_print_setup(ws, landscape=name in _LANDSCAPE_SHEETS)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
