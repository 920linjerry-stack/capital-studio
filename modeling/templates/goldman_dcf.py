# modeling/templates/goldman_dcf.py
# 职责：高盛风格 Excel DCF 模板的布局和样式定义。
# 只负责"怎么排版"，不负责计算，计算结果由 excel_exporter.py 传入。

from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── 颜色常量（高盛风格配色）────────────────────────────────────────────────────
COLOR_HEADER_BG   = "1A3A5C"  # 深蓝色表头背景
COLOR_HEADER_FG   = "FFFFFF"  # 白色表头文字
COLOR_INPUT_FG    = "003399"  # 蓝色字体 = 用户可编辑输入
COLOR_FORMULA_FG  = "000000"  # 黑色字体 = 公式计算结果
COLOR_LINK_FG     = "006600"  # 绿色字体 = 跨表链接
COLOR_SECTION_BG  = "D9E1F2"  # 淡蓝色分区背景
COLOR_HIGHLIGHT   = "FFF2CC"  # 黄色高亮（敏感性矩阵当前格）
COLOR_NEGATIVE    = "C00000"  # 深红色（负数）
COLOR_BORDER      = "BFBFBF"  # 边框灰色

# ── 字体预设 ──────────────────────────────────────────────────────────────────
def font_title(bold=True):
    return Font(name="Arial", size=12, bold=bold, color=COLOR_FORMULA_FG)

def font_header():
    return Font(name="Arial", size=10, bold=True, color=COLOR_HEADER_FG)

def font_input():
    """蓝色字体 = 假设输入，用户可在 Excel 里直接改"""
    return Font(name="Arial", size=10, color=COLOR_INPUT_FG)

def font_formula():
    return Font(name="Arial", size=10, color=COLOR_FORMULA_FG)

def font_link():
    """绿色字体 = 跨 Sheet 引用"""
    return Font(name="Arial", size=10, color=COLOR_LINK_FG)

def font_label():
    return Font(name="Arial", size=10, bold=True, color=COLOR_FORMULA_FG)

def font_watermark():
    return Font(name="Arial", size=8, color="AAAAAA", italic=True)

# ── 填充预设 ──────────────────────────────────────────────────────────────────
def fill_header():
    return PatternFill("solid", fgColor=COLOR_HEADER_BG)

def fill_section():
    return PatternFill("solid", fgColor=COLOR_SECTION_BG)

def fill_highlight():
    return PatternFill("solid", fgColor=COLOR_HIGHLIGHT)

# ── 边框预设 ──────────────────────────────────────────────────────────────────
def border_thin():
    s = Side(style="thin", color=COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom_only():
    return Border(bottom=Side(style="medium", color="000000"))

# ── 数字格式 ──────────────────────────────────────────────────────────────────
FMT_COMMA    = '#,##0'           # 千分位整数
FMT_COMMA2   = '#,##0.00'        # 千分位两位小数
FMT_PCT1     = '0.0%'            # 百分比一位小数
FMT_PCT2     = '0.00%'           # 百分比两位小数
FMT_MULTIPLE = '0.0x'            # 倍数（如 15.0x）

# ── 列宽配置（各 sheet 共用）────────────────────────────────────────────────
COLUMN_WIDTHS = {
    "A": 32,
    "B": 16,
    "C": 16,
    "D": 16,
    "E": 16,
    "F": 16,
    "G": 16,
    "H": 16,
}

def apply_column_widths(ws, widths=None):
    """统一设置列宽，传入 dict 覆盖默认值"""
    w = {**COLUMN_WIDTHS, **(widths or {})}
    for col, width in w.items():
        ws.column_dimensions[col].width = width

def write_header_row(ws, row, labels, start_col=1):
    """写一行深蓝色表头"""
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=label)
        c.font       = font_header()
        c.fill       = fill_header()
        c.alignment  = Alignment(horizontal="center", vertical="center")
        c.border     = border_thin()

def write_section_title(ws, row, title, span_cols=7):
    """写分区标题（淡蓝底色，粗体）"""
    c = ws.cell(row=row, column=1, value=title)
    c.font      = font_label()
    c.fill      = fill_section()
    c.border    = border_thin()
    # 视觉上合并单元格（实际不合并，避免 openpyxl 麻烦）
    for col in range(2, span_cols + 1):
        cc = ws.cell(row=row, column=col)
        cc.fill   = fill_section()
        cc.border = border_thin()

def write_input_row(ws, row, label, value, fmt=None, comment=None):
    """写一行"蓝色输入"：标签 + 值（蓝色字体，用户可改）"""
    lc = ws.cell(row=row, column=1, value=label)
    lc.font   = font_formula()
    lc.border = border_thin()

    vc = ws.cell(row=row, column=2, value=value)
    vc.font   = font_input()
    vc.border = border_thin()
    if fmt:
        vc.number_format = fmt
    if comment:
        cc = ws.cell(row=row, column=3, value=comment)
        cc.font = font_watermark()

def write_formula_row(ws, row, label, value, fmt=None):
    """写一行"黑色计算结果"：标签 + 值"""
    lc = ws.cell(row=row, column=1, value=label)
    lc.font   = font_formula()
    lc.border = border_thin()

    vc = ws.cell(row=row, column=2, value=value)
    vc.font   = font_formula()
    vc.border = border_thin()
    if fmt:
        vc.number_format = fmt
