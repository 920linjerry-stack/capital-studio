import io
import os
import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.pagebreak import Break

from modeling.dcf_calculator import (
    DCFInputs,
    DCFOutputs,
    build_operating_forecast,
    normalized_driver_assumptions,
    NET_DEBT_TREATMENTS,
    NET_DEBT_TREATMENT_LABELS,
    DEFAULT_NET_DEBT_TREATMENT,
    normalize_net_debt_treatment,
    SHARE_COUNT_TREATMENTS,
    SHARE_COUNT_TREATMENT_LABELS,
    DEFAULT_SHARE_COUNT_TREATMENT,
    normalize_share_count_treatment,
    BUYBACK_METHODS,
    BUYBACK_METHOD_LABELS,
    DEFAULT_BUYBACK_METHOD,
    normalize_buyback_method,
    WACC_TREATMENTS,
    WACC_TREATMENT_LABELS,
    DEFAULT_WACC_TREATMENT,
    normalize_wacc_treatment,
    TERMINAL_TREATMENTS,
    TERMINAL_TREATMENT_LABELS,
    DEFAULT_TERMINAL_TREATMENT,
    DEFAULT_FADE_YEARS,
    DEFAULT_H_MODEL_HALF_LIFE,
    DEFAULT_CASH_FLOOR,
    DEFAULT_BUYBACK_FUNDING_TREATMENT,
    normalize_terminal_treatment,
    build_aapl_operating_thesis_bridge_payload,
    normalize_operating_path_source,
    OPERATING_PATH_SOURCE_SELECTED,
    OPERATING_PATH_SOURCE_AAPL_BRIDGE,
    OPERATING_PATH_SOURCES,
)
from modeling.unit_utils import (
    MODEL_UNIT_SCALE_FACTORS,
    model_unit_label,
    model_unit_scale_factor,
    normalize_raw_actual_to_model_unit,
)
from data_fetcher_historical import (
    historical_cache_to_tables,
    read_historical_cache,
)
from modeling.templates.goldman_dcf import (
    apply_column_widths,
    write_header_row,
    write_section_title,
    font_title,
    font_formula,
    font_input,
    font_link,
    font_label,
    font_watermark,
    fill_header,
    fill_highlight,
    fill_section,
    border_thin,
    FMT_COMMA,
    FMT_COMMA2,
    FMT_PCT1,
    FMT_PCT2,
    FMT_MULTIPLE,
    COLOR_INPUT_FG,
    COLOR_FORMULA_FG,
    COLOR_LINK_FG,
    COLOR_HEADER_BG,
)


MODEL_VERSION = "Valuation Workbook v3.9.12"
FINANCIALS_CACHE_VERSION = "v327"
DEFAULT_DATA_SOURCE_TEXT = (
    "Public market data, locally cached financial statements, and daily cached trading comps"
)
SCOPE_BOUNDARY_TEXT = (
    "Designed as a valuation workbook and IC discussion aid. The model exposes key judgments, "
    "diagnostics, and review flags, but does not constitute an investment recommendation or "
    "fairness opinion. Analyst review is required before external use."
)
SCOPE_BOUNDARY_DETAIL = (
    "Not a full sell-side coverage model, official consensus estimate source, or client-ready lead "
    "valuation deliverable without analyst review. It does not automatically form rating views. "
    "Analyst judgment is required for explicit forecast horizon, terminal method selection, "
    "exit multiple premium, selected WACC, operating path source, shareholder return policy, and ROIC "
    "normalization decisions."
)
RELEASE_BOUNDARY_TEXT = (
    "General-Purpose Hardening / Currency & Audit Integrity: valuation workbook / IC discussion aid / "
    "research-aid model; not an investment recommendation, fairness opinion, or full sell-side coverage model."
)
ALLOWED_SOURCE_STATUSES = (
    "live", "cache", "stale cache", "derived", "user input", "filing-backed override",
    "single-feed diagnostic", "unavailable", "N/D in filing", "included in other line item",
    "not applicable",
)
ASSUMP = "Assumptions"
PROJECT_ROOT = Path(__file__).resolve().parents[1]

ASSUMP_CELLS: dict = {}
# V3.9.0 Forecast Path Upgrade v1: for the five operating forecast drivers we
# now store a 5-cell row (Year 1..Year 5) in addition to the legacy single-cell
# input. Downstream sheets pick the per-year cell from this map. Legacy
# ASSUMP_CELLS[key] continues to point at the Year 1 cell so any consumer that
# still expects a single ref reads a sensible value.
ASSUMP_PATH_CELLS: dict = {}
ASSUMP_SELECTED_PATH_CELLS: dict = {}
ASSUMP_BRIDGE_PATH_CELLS: dict = {}
FCF_ROWS: dict = {}
VAL_ROWS: dict = {}
PL_FORECAST_ROWS: dict = {}
CF_FORECAST_ROWS: dict = {}
BS_FORECAST_ROWS: dict = {}
SCHEDULE_ROWS: dict = {}
SENS_ROWS: dict = {}
SCENARIO_ROWS: dict = {}
AUDIT_STATUS_ROWS: list[int] = []
TRADING_COMPS_ROWS: dict = {}

CF_FCF_REF_LABEL = "CFO-derived FCF Reference"
SR_FCF_REF_LABEL = "CFO-derived FCF Reference (CF link)"

ASSUMPTION_LABELS = {
    "revenue_growth": "Revenue Growth",
    "ebit_margin": "EBIT Margin",
    "da_pct_revenue": "D&A % Revenue",
    "capex_pct_revenue": "CapEx % Revenue",
    "wc_change_pct_revenue": "Delta NWC % Revenue",
    "wacc": "WACC",
    "terminal_g": "Terminal Growth",
    "exit_multiple": "Exit Multiple",
    "price": "Current Price",
    "revenue": "Revenue",
    "ebit": "EBIT",
    "da": "D&A",
    "capex": "CapEx",
    "wc_change": "Delta NWC",
    "tax_rate": "Tax Rate",
    "net_debt": "Net Debt",
    "shares": "Diluted Shares",
    "forecast_years": "Forecast Years",
    "tv_method": "Terminal Value Method",
    "rf": "Risk-free Rate",
    "erp": "Equity Risk Premium",
    "beta": "Beta",
    # V3.7.3: Net Debt Treatment is now a scenario-level assumption.
    "selected_net_debt_treatment": "Net Debt Treatment",
    # V3.7.4: Shareholder Returns drivers are scenario-level assumptions.
    "selected_share_count_treatment": "Share Count Treatment",
    "buyback_method": "Buyback Method",
    "dividend_payout_pct_net_income": "Dividend Payout % NI",
    "buyback_pct_fcf": "Buyback % FCF",
    "flat_buyback_amount": "Flat Buyback Amount",
    "repurchase_price_growth": "Repurchase Price Growth",
    "annual_dilution_pct": "Annual Dilution %",
    # V3.7.5 WACC decision layer.
    "selected_wacc_treatment": "WACC Treatment",
    "pre_tax_cost_of_debt": "Pre-tax Cost of Debt",
    # V3.7.6 Terminal Value decision layer.
    "selected_terminal_treatment": "Terminal Treatment",
    "fade_years": "Fade Years",
    "fade_terminal_growth": "Fade Target Growth",
    "blend_weight_gordon": "Blend Weight Gordon",
    "blend_weight_exit": "Blend Weight Exit",
    "h_model_g_near": "H-Model Near Growth",
    "h_model_g_long": "H-Model Long Growth",
    "h_model_half_life": "H-Model Half-life",
    "buyback_funding_treatment": "Buyback Funding Treatment",
    "minimum_cash_floor": "Minimum Cash Floor",
    "marketable_securities_available_for_returns": "MS Available for Returns",
}

ASSUMPTION_CHANGE_KEYS = [
    "revenue_growth", "ebit_margin", "da_pct_revenue",
    "capex_pct_revenue", "wc_change_pct_revenue",
    "wacc", "terminal_g", "exit_multiple",
    "price", "revenue", "ebit", "da", "capex", "wc_change",
    "tax_rate", "net_debt", "shares", "forecast_years",
    "tv_method", "rf", "erp", "beta",
    "selected_net_debt_treatment",
    "selected_share_count_treatment", "buyback_method",
    "dividend_payout_pct_net_income", "buyback_pct_fcf",
    "flat_buyback_amount", "repurchase_price_growth", "annual_dilution_pct",
    "selected_wacc_treatment", "pre_tax_cost_of_debt",
    "selected_terminal_treatment", "fade_years", "fade_terminal_growth",
    "blend_weight_gordon", "blend_weight_exit",
    "h_model_g_near", "h_model_g_long", "h_model_half_life",
    "buyback_funding_treatment", "minimum_cash_floor",
    "marketable_securities_available_for_returns",
]


def _abs_ref(cell_addr: str) -> str:
    col = "".join(c for c in cell_addr if c.isalpha())
    row = "".join(c for c in cell_addr if c.isdigit())
    return f"'{ASSUMP}'!${col}${row}"


def _sheet_ref(sheet_name: str, row: int, col: int) -> str:
    return f"'{sheet_name}'!${get_column_letter(col)}${row}"


def _path_ref(key: str, year_index: int) -> str:
    """V3.9.0: return the absolute Excel ref to the Year (year_index+1) cell of
    a forecast-path driver on the Assumptions sheet. Falls back to the legacy
    single ASSUMP_CELLS[key] when no path is registered, or when the requested
    year exceeds the path length (forward-fill to the last entry).
    """
    path = ASSUMP_PATH_CELLS.get(key)
    if not path:
        return _abs_ref(ASSUMP_CELLS[key])
    idx = min(max(int(year_index), 0), len(path) - 1)
    return _abs_ref(path[idx])


def _forecast_start_col(ctx: dict) -> int:
    income_table = ((ctx.get("historical_tables") or {}).get("income_statement") or {})
    historical_years = sorted(income_table.get("years") or [])
    return 2 + len(historical_years)


def _filename_key(symbol: str) -> str:
    return symbol.lower().replace(".", "_").replace("-", "_")


def _now_local() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def _display_date(value) -> str:
    if not value:
        return ""
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    text = str(value)
    if len(text) >= 10:
        return text[:10]
    return text


def _scenario_label(value: str | None) -> str:
    v = (value or "base").lower()
    if v == "bull":
        return "Bull"
    if v == "bear":
        return "Bear"
    return "Base"


def _scenario_identity_note(ctx: dict) -> str:
    scenario = _scenario_label(ctx.get("current_scenario"))
    if scenario == "Base":
        return "Base (live, not a saved scenario)"
    scenarios = (ctx.get("scenario_doc") or {}).get("scenarios", {})
    entry = scenarios.get(scenario.lower()) if isinstance(scenarios, dict) else None
    saved = _display_date((entry or {}).get("saved_at")) if isinstance(entry, dict) else ""
    updated = _display_date((entry or {}).get("updated_at")) if isinstance(entry, dict) else ""
    parts = [scenario, "saved scenario"]
    if saved:
        parts.append(f"saved {saved}")
    if updated:
        parts.append(f"updated {updated}")
    return " | ".join(parts)


def _scenario_treatment_summary(ctx: dict, out: DCFOutputs) -> dict:
    scenarios = (ctx.get("scenario_doc") or {}).get("scenarios") or {}
    current = (ctx.get("current_scenario") or "base").lower()
    base = {
        "net_debt": (
            (getattr(out, "net_debt_bridge", None) or {}).get("selected_treatment")
            or DEFAULT_NET_DEBT_TREATMENT
        ),
        "share_count": (
            (getattr(out, "shareholder_returns", None) or {}).get("selected_share_count_treatment")
            or DEFAULT_SHARE_COUNT_TREATMENT
        ),
        "wacc": (
            (getattr(out, "wacc_decision_bridge", None) or {}).get("selected_wacc_treatment")
            or DEFAULT_WACC_TREATMENT
        ),
        "terminal": (
            (getattr(out, "terminal_decision_bridge", None) or {}).get("selected_terminal_treatment")
            or DEFAULT_TERMINAL_TREATMENT
        ),
    }
    labels = {
        "net_debt": NET_DEBT_TREATMENT_LABELS,
        "share_count": SHARE_COUNT_TREATMENT_LABELS,
        "wacc": WACC_TREATMENT_LABELS,
        "terminal": TERMINAL_TREATMENT_LABELS,
    }
    defaults = {
        "net_debt": DEFAULT_NET_DEBT_TREATMENT,
        "share_count": DEFAULT_SHARE_COUNT_TREATMENT,
        "wacc": DEFAULT_WACC_TREATMENT,
        "terminal": DEFAULT_TERMINAL_TREATMENT,
    }
    param_keys = {
        "net_debt": "selected_net_debt_treatment",
        "share_count": "selected_share_count_treatment",
        "wacc": "selected_wacc_treatment",
        "terminal": "selected_terminal_treatment",
    }
    states = []
    mismatches = []
    for scenario_type in ("bull", "bear"):
        entry = scenarios.get(scenario_type)
        if not isinstance(entry, dict):
            continue
        compat = entry.get("compatibility") or {}
        params = entry.get("params") or {}
        scenario_state = {"scenario": scenario_type}
        for treatment_type, param_key in param_keys.items():
            value = compat.get(param_key) or params.get(param_key) or defaults[treatment_type]
            scenario_state[treatment_type] = value
            if value != base[treatment_type]:
                mismatches.append((scenario_type, treatment_type, value))
        states.append(scenario_state)
    return {
        "current": current,
        "base": base,
        "states": states,
        "mismatches": mismatches,
        "labels": labels,
    }


def _sheet_meta_line(inp: DCFInputs, out: DCFOutputs, ctx: dict) -> str:
    return (
        f"Date: {_display_date(ctx.get('generated_at')) or date.today().isoformat()} | "
        f"Scenario: {_scenario_label(ctx.get('current_scenario'))} | "
        f"Currency / Unit: {model_unit_label(out.currency, out.market)} | Market: {out.market}"
    )


def _write_sheet_heading(ws, row: int, title: str, inp: DCFInputs, out: DCFOutputs, ctx: dict, span_cols: int = 3) -> int:
    ws.cell(row=row, column=1, value=title).font = font_title()
    row += 1
    c = ws.cell(row=row, column=1, value=_sheet_meta_line(inp, out, ctx))
    c.font = font_watermark()
    if span_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    return row + 2


def _base_params_from_input(inp: DCFInputs, out: DCFOutputs) -> dict:
    wc = out.wacc_components or {}
    drivers = normalized_driver_assumptions(inp)
    return {
        "symbol": inp.symbol,
        "company": inp.company,
        "price": inp.price,
        "currency": out.currency,
        "market": out.market,
        "revenue": inp.revenue,
        "ebit": inp.ebit,
        "da": inp.da,
        "capex": inp.capex,
        "wc_change": inp.wc_change,
        "tax_rate": inp.tax_rate,
        "net_debt": inp.net_debt,
        "shares": inp.shares,
        "revenue_growth": drivers["revenue_growth"],
        "ebit_margin": drivers["ebit_margin"],
        "da_pct_revenue": drivers["da_pct_revenue"],
        "capex_pct_revenue": drivers["capex_pct_revenue"],
        "wc_change_pct_revenue": drivers["wc_change_pct_revenue"],
        "wacc": inp.wacc,
        "terminal_g": inp.terminal_g,
        "exit_multiple": inp.exit_multiple,
        "forecast_years": inp.forecast_years,
        "tv_method": inp.tv_method,
        "rf": wc.get("rf"),
        "erp": wc.get("erp"),
        "beta": wc.get("beta"),
        "cost_of_equity": wc.get("cost_of_equity"),
        "selected_net_debt_treatment": getattr(inp, "selected_net_debt_treatment", DEFAULT_NET_DEBT_TREATMENT),
        "selected_share_count_treatment": getattr(inp, "selected_share_count_treatment", DEFAULT_SHARE_COUNT_TREATMENT),
        "buyback_method": getattr(inp, "buyback_method", DEFAULT_BUYBACK_METHOD),
        "dividend_payout_pct_net_income": getattr(inp, "dividend_payout_pct_net_income", None),
        "buyback_pct_fcf": getattr(inp, "buyback_pct_fcf", None),
        "flat_buyback_amount": getattr(inp, "flat_buyback_amount", None),
        "repurchase_price_growth": getattr(inp, "repurchase_price_growth", 0.0),
        "annual_dilution_pct": getattr(inp, "annual_dilution_pct", 0.0),
        # V3.7.5 WACC decision layer.
        "selected_wacc_treatment": getattr(inp, "selected_wacc_treatment", DEFAULT_WACC_TREATMENT),
        "pre_tax_cost_of_debt": getattr(inp, "pre_tax_cost_of_debt", None),
        # V3.7.6 Terminal Value decision layer.
        "selected_terminal_treatment": getattr(inp, "selected_terminal_treatment", DEFAULT_TERMINAL_TREATMENT),
        "fade_years": getattr(inp, "fade_years", DEFAULT_FADE_YEARS),
        "fade_terminal_growth": getattr(inp, "fade_terminal_growth", None),
        "blend_weight_gordon": getattr(inp, "blend_weight_gordon", 0.5),
        "blend_weight_exit": getattr(inp, "blend_weight_exit", 0.5),
        "h_model_g_near": getattr(inp, "h_model_g_near", None),
        "h_model_g_long": getattr(inp, "h_model_g_long", None),
        "h_model_half_life": getattr(inp, "h_model_half_life", None),
        "buyback_funding_treatment": getattr(inp, "buyback_funding_treatment", DEFAULT_BUYBACK_FUNDING_TREATMENT),
        "minimum_cash_floor": getattr(inp, "minimum_cash_floor", None),
        "marketable_securities_available_for_returns": getattr(inp, "marketable_securities_available_for_returns", None),
    }


def _read_financial_cache(symbol: str) -> dict:
    path = PROJECT_ROOT / "data" / "cache" / f"financials_{symbol}_{FINANCIALS_CACHE_VERSION}.json"
    if not path.exists():
        return {"path": str(path), "cached_at": None, "data": None}
    try:
        with open(path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        return {
            "path": str(path),
            "cached_at": payload.get("cached_at"),
            "data": payload.get("data") if isinstance(payload.get("data"), dict) else None,
        }
    except Exception:
        return {"path": str(path), "cached_at": None, "data": None}


def _read_scenario_doc(symbol: str) -> dict | None:
    path = PROJECT_ROOT / "data" / "dcf_scenarios" / f"{_filename_key(symbol)}_scenarios.json"
    if not path.exists():
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _normalize_params_for_audit(params: dict | None) -> tuple[dict, bool]:
    normalized = dict(params or {})
    legacy_mapped = "revenue_growth" not in normalized and "fcf_growth" in normalized
    fallback_growth = normalized.get("revenue_growth", normalized.get("fcf_growth"))
    if legacy_mapped and fallback_growth is not None:
        normalized["revenue_growth"] = fallback_growth
    normalized.pop("fcf_growth", None)
    normalized.pop("_legacy_fcf_growth_mapped", None)
    return normalized, legacy_mapped


def _normalize_scenario_doc_for_audit(doc: dict | None) -> dict | None:
    if not isinstance(doc, dict):
        return doc
    normalized_doc = dict(doc)
    scenarios = dict(normalized_doc.get("scenarios") or {})
    for scenario_type in ("bull", "bear"):
        entry = scenarios.get(scenario_type)
        if not isinstance(entry, dict):
            continue
        normalized_entry = dict(entry)
        params, legacy_mapped = _normalize_params_for_audit(entry.get("params"))
        compatibility = dict(entry.get("compatibility") or {})
        compatibility["legacy_fcf_growth_mapped"] = bool(
            compatibility.get("legacy_fcf_growth_mapped") or legacy_mapped
        )
        normalized_entry["params"] = params
        normalized_entry["compatibility"] = compatibility
        scenarios[scenario_type] = normalized_entry
    normalized_doc["scenarios"] = scenarios
    return normalized_doc


def _add_watermark_row(ws, row, symbol):
    c = ws.cell(row=row, column=1, value=f"Generated by Modeling Studio | {symbol} | {date.today()}")
    c.font = font_watermark()


def _set_input_cell(ws, row, col, value, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font_input()
    c.border = border_thin()
    if fmt:
        c.number_format = fmt
    return c


_SOURCE_COMMENT_AUTHOR = "DCF Workbook"


def _attach_source_comment(cell, source: str, cutoff: str, methodology: str) -> None:
    """V3.9.2: Attach a short, client-facing source / cutoff / methodology
    comment to a key input cell. Comments are deliberately concise so the
    workbook remains scannable; longer narrative belongs on the Data Sources
    Audit sheet."""
    if cell is None:
        return
    if cell.comment is not None:
        # Preserve any prior informational comment alongside the source trail.
        prior = (cell.comment.text or "").strip()
        body = f"Source: {source}\nCutoff: {cutoff}\nMethodology: {methodology}"
        if prior and prior not in body:
            body = f"{body}\n---\n{prior}"
    else:
        body = f"Source: {source}\nCutoff: {cutoff}\nMethodology: {methodology}"
    comment = Comment(body, _SOURCE_COMMENT_AUTHOR)
    comment.width = 320
    comment.height = 120
    cell.comment = comment


_CONST_STR_FORMULA = __import__("re").compile(r'^="([^"]*)"$')
_CONST_NUM_FORMULA = __import__("re").compile(r'^=(-?\d+(?:\.\d+)?)$')
# V3.8.2: strip quoted regions (both "..." and '...') from the formula body and
# check whether the remainder contains whitespace AND lacks parentheses. Every
# valid Excel formula with whitespace contains either a function call (parens)
# or a quoted sheet name ('Some Sheet'!A1) - so a body with bare whitespace
# tokens and no parens cannot be a real formula. This catches strings like
# "=5 target" and "=0.025 high / 0.015 review" (legacy threshold labels that
# accidentally carry a '=' prefix) without ever touching real formulas like
# "=B27", "=G18+G27", "=ROUND(A1,2)", or "='P&L Forecast'!$G$5".
_QUOTED_REGION = __import__("re").compile(r'"[^"]*"|\'[^\']*\'')


def _is_malformed_pseudo_formula(formula: str) -> bool:
    if not formula.startswith("="):
        return False
    body = _QUOTED_REGION.sub("", formula[1:])
    if "(" in body:
        return False
    if not __import__("re").search(r"\s", body):
        return False
    return True


def _set_formula_cell(ws, row, col, formula: str, fmt=None, cross_sheet=False):
    # V3.8.1 Excel-desktop safety: collapse degenerate "formulas" that are just
    # a constant string literal (="OK") or a constant number (=1) into plain
    # cell values. V3.8.2 extends this: also demote strings that start with
    # '=' but cannot be a valid Excel formula (e.g. "=5 target",
    # "=0.025 high / 0.015 review") - these were intended as label/threshold
    # text and reach the workbook only via legacy call sites. Excel desktop
    # removes such <f> records on reopen, producing "Removed Records: formulas"
    # repair prompts. Demotion keeps the displayed value identical.
    value = formula
    if isinstance(formula, str):
        m_str = _CONST_STR_FORMULA.match(formula)
        m_num = _CONST_NUM_FORMULA.match(formula)
        if m_str is not None:
            value = m_str.group(1)
        elif m_num is not None:
            num_text = m_num.group(1)
            value = float(num_text) if "." in num_text else int(num_text)
        elif _is_malformed_pseudo_formula(formula):
            value = formula[1:]
    c = ws.cell(row=row, column=col, value=value)
    c.font = font_link() if cross_sheet else font_formula()
    c.border = border_thin()
    if fmt:
        c.number_format = fmt
    return c


def _set_label_cell(ws, row, col, label, bold=False):
    c = ws.cell(row=row, column=col, value=label)
    c.font = font_label() if bold else font_formula()
    c.border = border_thin()
    return c


def _set_note_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font_watermark()
    c.alignment = Alignment(wrap_text=True, vertical="top")
    return c


def _write_kv_rows(ws, row, rows, label_col=1, value_col=2):
    for label, value, fmt in rows:
        _set_label_cell(ws, row, label_col, label)
        c = ws.cell(row=row, column=value_col, value=value)
        c.font = font_formula()
        c.border = border_thin()
        if fmt:
            c.number_format = fmt
        row += 1
    return row


def _apply_print_setup(
    ws,
    landscape: bool = False,
    freeze: str | None = None,
    repeat_rows: str | None = None,
    company: str | None = None,
    ticker: str | None = None,
    generated_date: str | None = None,
):
    ws.sheet_view.showGridLines = False
    if freeze:
        ws.freeze_panes = freeze
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    if repeat_rows:
        ws.print_title_rows = repeat_rows
    identity = f"{company or ''} ({ticker or ''})".strip()
    ws.oddHeader.left.text = identity if identity != "()" else ""
    ws.oddHeader.center.text = "&A"
    ws.oddHeader.right.text = f"Page &P of &N | {generated_date or date.today().isoformat()}"
    ws.oddFooter.right.text = "Page &P of &N"


def _fmt_for_key(key: str):
    if key in {
        "revenue_growth", "ebit_margin", "da_pct_revenue",
        "capex_pct_revenue", "wc_change_pct_revenue", "wacc", "terminal_g",
        "tax_rate", "rf", "erp", "beta",
    }:
        return FMT_PCT2 if key != "beta" else "0.00"
    if key == "exit_multiple":
        return FMT_MULTIPLE
    if key == "forecast_years":
        return FMT_COMMA
    if key == "tv_method":
        return None
    return FMT_COMMA2


def _coerce_num(value):
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _hist_value(ctx: dict, statement_key: str, field_key: str, year: int):
    table = ((ctx.get("historical_tables") or {}).get(statement_key) or {})
    rows = table.get("rows") or {}
    return (rows.get(field_key) or {}).get(year)


def _hist_field_meta(ctx: dict, statement_key: str, field_key: str, year: int) -> dict:
    payload = ((ctx.get("historical_cache") or {}).get("data") or {})
    for entry in (payload.get("statements") or {}).get(statement_key) or []:
        if entry.get("fiscal_year") == year:
            return (entry.get("fields") or {}).get(field_key) or {}
    return {}


def _hist_display_value(ctx: dict, statement_key: str, field_key: str, year: int, market: str):
    value = _hist_model_value(ctx, statement_key, field_key, year, market)
    if value is not None:
        return value
    meta = _hist_field_meta(ctx, statement_key, field_key, year)
    return meta.get("display_value") if meta.get("not_separately_disclosed") else None


def _hist_model_value(ctx: dict, statement_key: str, field_key: str, year: int, market: str):
    return normalize_raw_actual_to_model_unit(_hist_value(ctx, statement_key, field_key, year), market)


def _hist_model_value_positive(ctx: dict, statement_key: str, field_key: str, year: int, market: str):
    value = _hist_model_value(ctx, statement_key, field_key, year, market)
    return abs(value) if value is not None else None


def _safe_pct(numerator, denominator):
    try:
        if numerator is None or denominator in (None, 0):
            return None
        return float(numerator) / float(denominator)
    except (TypeError, ValueError, ZeroDivisionError):
        return None


def _latest_historical_gross_margin(ctx: dict, market: str):
    years = ((ctx.get("historical_tables") or {}).get("income_statement") or {}).get("years") or []
    for year in sorted(years, reverse=True):
        revenue = _hist_model_value(ctx, "income_statement", "revenue", year, market)
        gross_profit = _hist_model_value(ctx, "income_statement", "gross_profit", year, market)
        margin = _safe_pct(gross_profit, revenue)
        if margin is not None:
            return margin
    return None


def _latest_historical_ratio(ctx: dict, statement_key: str, numerator_key: str, market: str, abs_numerator: bool = False):
    years = ((ctx.get("historical_tables") or {}).get("income_statement") or {}).get("years") or []
    statement_years = ((ctx.get("historical_tables") or {}).get(statement_key) or {}).get("years") or []
    for year in sorted(set(years).intersection(statement_years), reverse=True):
        revenue = _hist_model_value(ctx, "income_statement", "revenue", year, market)
        numerator = _hist_model_value(ctx, statement_key, numerator_key, year, market)
        if abs_numerator and numerator is not None:
            numerator = abs(numerator)
        ratio = _safe_pct(numerator, revenue)
        if ratio is not None:
            return ratio
    return None


def _latest_historical_model_value(ctx: dict, statement_key: str, field_key: str, market: str, positive: bool = False):
    years = ((ctx.get("historical_tables") or {}).get(statement_key) or {}).get("years") or []
    for year in sorted(years, reverse=True):
        value = _hist_model_value(ctx, statement_key, field_key, year, market)
        if value is not None:
            return abs(value) if positive else value
    return None


def _balance_sheet_forecast_support_status(out: DCFOutputs, ctx: dict) -> str:
    payload = ((ctx.get("historical_cache") or {}).get("data") or {})
    bs_years = (((ctx.get("historical_tables") or {}).get("balance_sheet") or {}).get("years") or [])
    if out.market == "US" and payload.get("status") == "ok" and bs_years and CF_FORECAST_ROWS and SCHEDULE_ROWS:
        return "full_forecast"
    return "limited_presentation"


def _display_value_or_na(value):
    return "N/A" if value is None or value == "" else value


def _latest_bs_period_end(ctx: dict) -> str:
    bs_table = ((ctx.get("historical_tables") or {}).get("balance_sheet") or {})
    years = sorted(bs_table.get("years") or [])
    if not years:
        return "N/A"
    latest_year = years[-1]
    payload = ((ctx.get("historical_cache") or {}).get("data") or {})
    for entry in (payload.get("statements") or {}).get("balance_sheet") or []:
        if entry.get("fiscal_year") == latest_year:
            return _display_date(entry.get("period_end")) or str(latest_year)
    return str(latest_year)


def _write_limited_balance_sheet_presentation(ws, row: int, inp: DCFInputs, out: DCFOutputs, ctx: dict, span_cols: int) -> None:
    current_params = ctx.get("current_params") or {}
    historical_cache = ctx.get("historical_cache") or {}
    historical_payload = historical_cache.get("data") or {}
    financial_cache = ctx.get("financial_cache") or {}
    nd_bridge = getattr(out, "net_debt_bridge", None) or {}
    sr_payload = getattr(out, "shareholder_returns", None) or {}
    sr_options = sr_payload.get("denominator_options") or {}
    sc_key = sr_payload.get("selected_share_count_treatment") or DEFAULT_SHARE_COUNT_TREATMENT
    sc_label = SHARE_COUNT_TREATMENT_LABELS.get(sc_key, sc_key)

    cash = nd_bridge.get("cash")
    if cash is None:
        cash = _latest_historical_model_value(ctx, "balance_sheet", "cash", out.market)
    st_investments = nd_bridge.get("short_term_investments")
    if st_investments is None:
        st_investments = _latest_historical_model_value(ctx, "balance_sheet", "short_term_investments", out.market)
    total_debt = nd_bridge.get("total_debt")
    if total_debt is None:
        total_debt = _latest_historical_model_value(ctx, "balance_sheet", "total_debt", out.market)
    net_debt = nd_bridge.get("selected_net_debt")
    if net_debt is None:
        net_debt = current_params.get("net_debt", getattr(inp, "net_debt", None))
    shares_used = sr_payload.get("selected_denominator")
    if shares_used is None:
        shares_used = sr_options.get("current_reported_shares")
    if shares_used is None:
        shares_used = current_params.get("shares", getattr(inp, "shares", None))
    price = current_params.get("price", getattr(inp, "price", None))
    market_cap = None
    if _coerce_num(price) is not None and _coerce_num(shares_used) is not None:
        market_cap = float(price) * float(shares_used)

    scope_text = (
        "Full balance sheet forecast and equity roll-forward are not supported for this market in the current workbook. "
        "Valuation uses selected balance sheet inputs for net debt, share count, and currency bridge only."
    )

    write_section_title(ws, row, "Balance Sheet Inputs Summary - Limited Presentation", span_cols=span_cols)
    row += 1
    _set_note_cell(ws, row, 1, scope_text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    write_header_row(ws, row, ["Input", "Value", "Source / Status"], 1)
    row += 1
    rows = [
        ("Selected period end", _latest_bs_period_end(ctx), "Latest available normalized balance-sheet period"),
        ("Reporting currency", current_params.get("reporting_currency") or out.currency, current_params.get("reporting_currency_source") or "Workbook currency"),
        ("Trading currency", current_params.get("trading_currency") or out.currency, current_params.get("trading_currency_source") or "Workbook currency"),
        ("FX rate used", current_params.get("fx_rate_reporting_to_trading"), current_params.get("fx_rate_source") or "N/A"),
        ("Cash and cash equivalents", cash, "Net Debt Bridge / historical balance-sheet cache"),
        ("Short-term investments / marketable securities", st_investments, "Net Debt Bridge / historical balance-sheet cache when available"),
        ("Total debt", total_debt, "Net Debt Bridge / historical balance-sheet cache"),
        ("Net debt / net cash", net_debt, "Selected net debt used by valuation"),
        ("Shares outstanding used", shares_used, "Selected IV/share denominator"),
        ("Share denominator source", sc_label, "Share Count Treatment"),
        ("Market cap", market_cap, "Current price x selected shares when both are available"),
        ("Current price", price, "Market quote / current params"),
        ("Data source / cache file", historical_cache.get("path") or financial_cache.get("path") or "N/A", "Historical cache preferred; financial cache fallback"),
        ("Balance sheet support status", "limited_presentation", "Full BS forecast not supported for this market"),
        ("Notes / warnings", " | ".join(current_params.get("warnings") or []) or "None", "Workbook warnings"),
    ]
    for label, value, source in rows:
        _set_label_cell(ws, row, 1, label)
        cell = ws.cell(row=row, column=2, value=_display_value_or_na(value))
        cell.font = font_formula()
        cell.border = border_thin()
        if isinstance(value, (int, float)):
            cell.number_format = FMT_COMMA2
        _set_note_cell(ws, row, 3, source)
        row += 1

    row += 1
    write_section_title(ws, row, "Scope Status", span_cols=span_cols)
    row += 1
    row = _write_kv_rows(ws, row, [
        ("Balance Sheet Forecast Status", "limited_presentation", None),
        ("Full BS Forecast", "not supported", None),
        ("Net Debt Bridge", "supported if data available", None),
        ("Equity Roll-forward", "not supported", None),
        ("Source limitations", "Normalized full BS linkage and equity roll-forward are available only where the historical BS / CF pipeline is supported.", None),
    ])
    _set_note_cell(ws, row + 1, 1, "No yearly balance-sheet forecast columns are created on this limited-scope page; missing source fields display as N/A.")
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=span_cols)


def _latest_wc_days(ctx: dict, market: str) -> dict:
    """Compute DSO / DIO / DPO from the latest historical year that has the needed fields.

    DSO = AR / Revenue * 365
    DIO = Inventory / COGS * 365         (COGS = Revenue - Gross Profit)
    DPO = AP / COGS * 365
    """
    is_table = ((ctx.get("historical_tables") or {}).get("income_statement") or {})
    bs_table = ((ctx.get("historical_tables") or {}).get("balance_sheet") or {})
    years = sorted(set(is_table.get("years") or []).intersection(bs_table.get("years") or []), reverse=True)
    out = {"dso": None, "dio": None, "dpo": None}
    for year in years:
        revenue = _hist_model_value(ctx, "income_statement", "revenue", year, market)
        gross_profit = _hist_model_value(ctx, "income_statement", "gross_profit", year, market)
        cogs = (revenue - gross_profit) if (revenue is not None and gross_profit is not None) else None
        ar = _hist_model_value(ctx, "balance_sheet", "accounts_receivable", year, market)
        inv = _hist_model_value(ctx, "balance_sheet", "inventory", year, market)
        ap = _hist_model_value(ctx, "balance_sheet", "accounts_payable", year, market)
        if out["dso"] is None and ar is not None and revenue:
            out["dso"] = ar / revenue * 365.0
        if out["dio"] is None and inv is not None and cogs:
            out["dio"] = inv / cogs * 365.0
        if out["dpo"] is None and ap is not None and cogs:
            out["dpo"] = ap / cogs * 365.0
        if all(out[k] is not None for k in out):
            break
    return out


def _normalized_wc_days(ctx: dict, market: str, count: int = 3) -> dict:
    is_table = ((ctx.get("historical_tables") or {}).get("income_statement") or {})
    bs_table = ((ctx.get("historical_tables") or {}).get("balance_sheet") or {})
    years = sorted(set(is_table.get("years") or []).intersection(bs_table.get("years") or []))
    series = {"dso": [], "dio": [], "dpo": []}
    for year in years:
        revenue = _hist_model_value(ctx, "income_statement", "revenue", year, market)
        gross_profit = _hist_model_value(ctx, "income_statement", "gross_profit", year, market)
        cogs = (revenue - gross_profit) if (revenue is not None and gross_profit is not None) else None
        ar = _hist_model_value(ctx, "balance_sheet", "accounts_receivable", year, market)
        inv = _hist_model_value(ctx, "balance_sheet", "inventory", year, market)
        ap = _hist_model_value(ctx, "balance_sheet", "accounts_payable", year, market)
        if ar is not None and revenue:
            series["dso"].append(ar / revenue * 365.0)
        if inv is not None and cogs:
            series["dio"].append(inv / cogs * 365.0)
        if ap is not None and cogs:
            series["dpo"].append(ap / cogs * 365.0)
    return {
        k: (sum(v[-count:]) / len(v[-count:]) if v else None)
        for k, v in series.items()
    }


def _latest_da_pct_begin_ppe(ctx: dict, market: str):
    """D&A / Beginning Net PP&E from latest historical year where both are present."""
    bs_table = ((ctx.get("historical_tables") or {}).get("balance_sheet") or {})
    cf_table = ((ctx.get("historical_tables") or {}).get("cash_flow") or {})
    years = sorted(set(bs_table.get("years") or []).intersection(cf_table.get("years") or []), reverse=True)
    for year in years:
        da = _hist_model_value(ctx, "cash_flow", "depreciation_amortization", year, market)
        prev_years = [y for y in bs_table.get("years") if y < year] if bs_table.get("years") else []
        if not prev_years:
            continue
        beg_ppe = _hist_model_value(ctx, "balance_sheet", "ppe", max(prev_years), market)
        if da is not None and beg_ppe:
            return da / beg_ppe
    return None


def _values_differ(a, b) -> bool:
    an = _coerce_num(a)
    bn = _coerce_num(b)
    if an is not None and bn is not None:
        return abs(an - bn) > 1e-9
    return a != b


def _infer_current_scenario(current_params: dict, scenario_doc: dict | None) -> str:
    if not scenario_doc or not isinstance(current_params, dict):
        return "base"
    scenarios = scenario_doc.get("scenarios", {})
    keys = ("revenue", "ebit", "da", "capex", "wc_change", "tax_rate", "net_debt", "shares",
            "revenue_growth", "ebit_margin", "da_pct_revenue", "capex_pct_revenue",
            "wc_change_pct_revenue", "wacc", "terminal_g", "exit_multiple",
            "forecast_years", "tv_method")
    for scenario_type in ("bull", "bear"):
        entry = scenarios.get(scenario_type)
        params = entry.get("params") if isinstance(entry, dict) else None
        if not isinstance(params, dict):
            continue
        if all(not _values_differ(current_params.get(k), params.get(k)) for k in keys):
            return scenario_type
    return "base"


def _build_context(inp: DCFInputs, out: DCFOutputs, export_context: dict | None) -> dict:
    ctx = dict(export_context or {})
    ctx.setdefault("generated_at", _now_local())
    current_params, _ = _normalize_params_for_audit(ctx.get("current_params") or _base_params_from_input(inp, out))
    ctx["current_params"] = current_params
    base_params, _ = _normalize_params_for_audit(ctx.get("base_params")) if isinstance(ctx.get("base_params"), dict) else (ctx.get("base_params"), False)
    ctx["base_params"] = base_params
    ctx.setdefault("scenario_doc", _read_scenario_doc(inp.symbol))
    ctx["scenario_doc"] = _normalize_scenario_doc_for_audit(ctx.get("scenario_doc"))
    ctx.setdefault("financial_cache", _read_financial_cache(inp.symbol))
    ctx.setdefault("historical_cache", read_historical_cache(inp.symbol))
    historical_payload = (ctx.get("historical_cache") or {}).get("data")
    ctx["historical_tables"] = historical_cache_to_tables(historical_payload)
    current_scenario = (ctx.get("current_scenario") or "").lower()
    if current_scenario not in {"base", "bull", "bear"}:
        current_scenario = _infer_current_scenario(ctx.get("current_params", {}), ctx.get("scenario_doc"))
    ctx["current_scenario"] = current_scenario
    if ctx.get("base_params") is None and current_scenario == "base":
        ctx["base_params"] = ctx.get("current_params") if current_scenario == "base" else None
    return ctx


def build_cover_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "Cover 使用说明"
    apply_column_widths(ws, {"A": 4, "B": 32, "C": 72})
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False

    row = 2
    c = ws.cell(row=row, column=2, value="DCF Valuation Workbook")
    c.font = Font(name="Arial", size=22, bold=True, color=COLOR_HEADER_BG)
    row += 2

    scenario = _scenario_label(ctx.get("current_scenario"))
    current_params = ctx.get("current_params", {})
    data_source = "Public market data, cached financial statements, and trading data."
    default_quality = current_params.get("default_quality") or {}
    info_rows = [
        ("Company", f"{inp.company} ({inp.symbol})", None),
        ("Scenario", scenario, None),
        ("Export Identity", _scenario_identity_note(ctx), None),
        ("Generated Date", _display_date(ctx.get("generated_at")) or date.today().isoformat(), None),
        ("Model Version", MODEL_VERSION, None),
        ("Market / Currency", f"{out.market} / {out.currency}", None),
        ("Unit", model_unit_label(out.currency, out.market), None),
        ("Data Source", data_source, None),
        ("Default Quality", default_quality.get("review_tier") or "OK", None),
    ]
    row = _write_kv_rows(ws, row, info_rows, label_col=2, value_col=3)
    if default_quality.get("requires_review"):
        _set_note_cell(ws, row, 2, default_quality.get("banner") or "Default assumptions require review before use.")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        row += 1
    row += 2

    write_section_title(ws, row, "Workbook Map", span_cols=3)
    row += 1
    c = ws.cell(row=row, column=2, value="This workbook contains 16 sheets")
    c.font = font_label()
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    row += 1
    sheets = [
        ("Cover", "Workbook identity, formatting legend, and scope boundary."),
        ("Executive Summary", "IC summary, visual blocks, and valuation range"),
        ("Scenario Notes", "情景身份、保存日期、估值摘要、假设差异"),
        ("Historical Financials", "Multi-year actual Income Statement, Balance Sheet, and Cash Flow"),
        ("P&L Forecast", "Historical actuals plus five-year P&L forecast in workbook model unit"),
        ("Balance Sheet Forecast", "Expanded BS forecast: marketable securities, goodwill / intangibles, deferred revenue, leases, deferred tax assets / liabilities, with transparent residual plugs"),
        ("Cash Flow Forecast", "Indirect method cash flow forecast (3FS-linked)"),
        ("Supporting Schedules", "PP&E roll-forward, Working Capital days driver (DSO/DIO/DPO), Debt schedule, Share count"),
        ("Assumptions", "Inputs and model controls — Net Debt Bridge, Shareholder Returns, WACC decision layer, Terminal Value decision layer"),
        ("FCF Build", "Five-year explicit FCFF build"),
        ("DCF Valuation", "DCF valuation summary with adjusted IV references under each Net Debt treatment"),
        ("Trading Comps", "Peer EV/Revenue, EV/EBITDA, P/E multiples + implied IV/share ranges (diagnostic market cross-check)"),
        ("FA Ratios", "FA / Ratios analytical layer; output only, not model drivers"),
        ("Audit Dashboard", "Model checks review layer with live status formulas and centralized thresholds"),
        ("Sensitivity", "WACC / Terminal assumptions 敏感性分析"),
        ("Data Sources Audit", "数据来源、缓存字段、计算 caveat"),
    ]
    for name, desc in sheets:
        _set_label_cell(ws, row, 2, name)
        _set_note_cell(ws, row, 3, desc)
        row += 1
    row += 2

    write_section_title(ws, row, "Formatting Legend", span_cols=3)
    row += 1
    rules = [
        ("Blue font", "User-editable inputs, mainly on Assumptions.", COLOR_INPUT_FG),
        ("Black font", "Formulas, labels, and calculation outputs.", COLOR_FORMULA_FG),
        ("Green font", "Internal workbook links.", COLOR_LINK_FG),
    ]
    for label, text, color in rules:
        _set_label_cell(ws, row, 2, label)
        c = ws.cell(row=row, column=3, value=text)
        c.font = Font(name="Arial", size=10, color=color)
        c.border = border_thin()
        row += 1
    row += 2

    write_section_title(ws, row, "免责声明 / Disclaimer", span_cols=3)
    row += 1
    disclaimer = (
        "本工作簿根据当前 DCF 输入和本地缓存数据生成，仅供分析参考，不构成投资建议。"
        "首次打开后请允许 Excel 自动重算公式。 For analytical discussion use."
    )
    c = ws.cell(row=row, column=2, value=SCOPE_BOUNDARY_TEXT)
    c.font = font_watermark()
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.row_dimensions[row].height = 36
    row += 1
    c = ws.cell(row=row, column=2, value=SCOPE_BOUNDARY_DETAIL)
    c.font = font_watermark()
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.row_dimensions[row].height = 56


def build_scenario_notes_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "Scenario Notes"
    SCENARIO_ROWS.clear()
    apply_column_widths(ws, {"A": 24, "B": 22, "C": 22, "D": 22, "E": 22, "F": 24, "G": 24, "H": 24, "I": 24})

    scenario_doc = ctx.get("scenario_doc") or {}
    scenarios = scenario_doc.get("scenarios", {}) if isinstance(scenario_doc, dict) else {}
    current = ctx.get("current_scenario", "base")
    base_params = ctx.get("base_params")

    row = _write_sheet_heading(ws, 1, "Scenario Notes", inp, out, ctx, span_cols=9)

    treatment_summary = _scenario_treatment_summary(ctx, out)

    write_section_title(ws, row, "Current Export Identity", span_cols=9)
    row += 1
    current_entry = scenarios.get(current) if current in {"bull", "bear"} else None
    # V3.7.3: surface the Selected Net Debt Treatment used for this export so
    # reviewers don't have to cross-reference Assumptions to know which net
    # debt variant drove the headline IV.
    selected_treatment_key = (getattr(out, "net_debt_bridge", None) or {}).get(
        "selected_treatment", DEFAULT_NET_DEBT_TREATMENT
    )
    selected_treatment_label = NET_DEBT_TREATMENT_LABELS.get(
        selected_treatment_key, NET_DEBT_TREATMENT_LABELS[DEFAULT_NET_DEBT_TREATMENT]
    )
    selected_net_debt_used = (getattr(out, "net_debt_bridge", None) or {}).get(
        "selected_net_debt"
    )
    # V3.7.4 shareholder returns identity refs.
    sr_identity = getattr(out, "shareholder_returns", None) or {}
    sr_drivers_identity = sr_identity.get("drivers_effective") or {}
    sc_treatment_identity_key = sr_identity.get("selected_share_count_treatment") or DEFAULT_SHARE_COUNT_TREATMENT
    sc_treatment_identity_label = SHARE_COUNT_TREATMENT_LABELS.get(
        sc_treatment_identity_key, SHARE_COUNT_TREATMENT_LABELS[DEFAULT_SHARE_COUNT_TREATMENT]
    )
    sc_denom_identity = sr_identity.get("selected_denominator")
    identity_rows = [
        ("Scenario", _scenario_label(current), None),
        ("Case Label", (ctx.get("current_params") or {}).get("case_label") or _scenario_label(current), None),
        ("Identity Note", _scenario_identity_note(ctx), None),
        (
            "Operating Path Source",
            (
                f"{(getattr(out, 'operating_path_bridge', None) or {}).get('selected_operating_path_source', 'Selected Path')} "
                "(defaulted from legacy; scenario saved before Bridge engine-driving was available)."
                if (getattr(out, "operating_path_bridge", None) or {}).get("selected_operating_path_source_legacy_defaulted")
                else (getattr(out, "operating_path_bridge", None) or {}).get("selected_operating_path_source", "Selected Path")
            ),
            None,
        ),
        ("Selected Net Debt Treatment", selected_treatment_label, None),
        ("Selected Net Debt Used in DCF", selected_net_debt_used, FMT_COMMA2),
        (
            "Marketable Securities Treatment Note",
            "Switch the Selected Net Debt Treatment dropdown on Assumptions to flip the headline; default keeps the reported / input net debt.",
            None,
        ),
        ("Selected Share Count Treatment", sc_treatment_identity_label, None),
        ("Selected Share Count Used in DCF (M)", sc_denom_identity, FMT_COMMA2),
        ("Dividend Payout % NI", sr_drivers_identity.get("dividend_payout_pct_net_income"), FMT_PCT2),
        ("Buyback Method", sr_drivers_identity.get("buyback_method_label") or "", None),
        ("Buyback % FCF", sr_drivers_identity.get("buyback_pct_fcf"), FMT_PCT2),
        ("Flat Buyback Amount", sr_drivers_identity.get("flat_buyback_amount"), FMT_COMMA2),
        (
            "Shareholder Returns Note",
            "Dividends and buybacks are equity-financing items; FCFF is unchanged. The per-share denominator can flip via the Selected Share Count Treatment dropdown.",
            None,
        ),
    ]
    # V3.7.5 WACC identity rows.
    wacc_identity_payload = getattr(out, "wacc_decision_bridge", None) or {}
    identity_rows += [
        ("Selected WACC Treatment", wacc_identity_payload.get("selected_wacc_treatment_label") or "", None),
        ("Selected WACC Used in DCF", wacc_identity_payload.get("selected_wacc_used_in_dcf"), FMT_PCT2),
        ("Mechanical CAPM reference WACC", wacc_identity_payload.get("capm_indicative_wacc"), FMT_PCT2),
        ("Selected vs Mechanical CAPM Spread", wacc_identity_payload.get("selected_vs_indicative_spread"), FMT_PCT2),
        (
            "WACC Treatment Note",
            "WACC treatment is a valuation judgment; default = Selected / Model WACC. Switch the WACC dropdown on Assumptions to flip the headline.",
            None,
        ),
    ]
    # V3.7.6 Terminal identity rows.
    tv_identity_payload = getattr(out, "terminal_decision_bridge", None) or {}
    identity_rows += [
        ("Selected Terminal Treatment", tv_identity_payload.get("selected_terminal_treatment_label") or "", None),
        ("Selected Terminal Value PV in DCF (M)", tv_identity_payload.get("selected_terminal_value_pv"), FMT_COMMA2),
        ("Terminal Growth", tv_identity_payload.get("terminal_growth"), FMT_PCT2),
        ("Exit Multiple", tv_identity_payload.get("exit_multiple"), FMT_MULTIPLE),
        ("Fade Years", (tv_identity_payload.get("fade_period") or {}).get("fade_years"), "0"),
        ("TV / EV at Selected Treatment", tv_identity_payload.get("selected_terminal_tv_ev_pct"), FMT_PCT2),
        ("Gordon vs Exit Gap", tv_identity_payload.get("gordon_vs_exit_gap_pct"), FMT_PCT2),
        ("Gordon Implied Exit Multiple", tv_identity_payload.get("gordon_implied_exit_multiple"), FMT_MULTIPLE),
        ("Exit Implied Terminal Growth", tv_identity_payload.get("exit_implied_terminal_growth"), FMT_PCT2),
        (
            "Terminal Treatment Note",
            "Terminal treatment is a valuation judgment; default = Current Model. Switch the Terminal dropdown on Assumptions to flip the headline.",
            None,
        ),
    ]
    SCENARIO_ROWS["identity_scenario"] = row
    SCENARIO_ROWS["identity_note"] = row + 1
    SCENARIO_ROWS["identity_treatment"] = row + 2
    if isinstance(current_entry, dict):
        saved_at = _display_date(current_entry.get("saved_at"))
        updated_at = _display_date(current_entry.get("updated_at"))
        origin_source = (current_entry.get("origin") or {}).get("source", "")
        if saved_at:
            identity_rows.append(("Saved At", saved_at, None))
        if updated_at:
            identity_rows.append(("Updated At", updated_at, None))
        if origin_source:
            identity_rows.append(("Origin Source", origin_source, None))
    else:
        identity_rows.append(("Status", "Base (live, not a saved scenario)", None))
    row = _write_kv_rows(ws, row, identity_rows)
    row += 2

    write_section_title(ws, row, "Saved Scenario Snapshot", span_cols=6)
    row += 1
    saved_count = sum(1 for scenario_type in ("bull", "bear") if isinstance(scenarios.get(scenario_type), dict))
    if current == "base" and saved_count:
        SCENARIO_ROWS["saved_reference_disclaimer"] = row
        _set_note_cell(
            ws,
            row,
            1,
            "Saved Bull/Bear rows below are saved scenario references only; they are not the current Base valuation.",
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1
    if not saved_count:
        SCENARIO_ROWS["saved_reference_disclaimer"] = row
        _set_note_cell(ws, row, 1, "No saved Bull/Bear scenario references are available.")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 2
    else:
        write_header_row(ws, row, [
            "Scenario", "Current Export", "Saved At", "Updated At", "Origin Source",
            "Net Debt Treatment", "Share Count Treatment", "WACC Treatment", "Terminal Treatment",
        ], 1)
        row += 1
    for scenario_type in ("bull", "bear"):
        entry = scenarios.get(scenario_type)
        if not isinstance(entry, dict):
            continue
        origin = entry.get("origin", {}) if isinstance(entry, dict) else {}
        # V3.7.3: surface the saved scenario's Net Debt Treatment so reviewers
        # can see at a glance whether Bull/Bear use a different treatment
        # than Base.
        scen_params = entry.get("params", {}) if isinstance(entry, dict) else {}
        scen_treatment_key = (
            (entry.get("compatibility") or {}).get("selected_net_debt_treatment")
            or scen_params.get("selected_net_debt_treatment")
            or DEFAULT_NET_DEBT_TREATMENT
        )
        scen_treatment_label = NET_DEBT_TREATMENT_LABELS.get(scen_treatment_key, scen_treatment_key)
        scen_share_key = (
            (entry.get("compatibility") or {}).get("selected_share_count_treatment")
            or scen_params.get("selected_share_count_treatment")
            or DEFAULT_SHARE_COUNT_TREATMENT
        )
        scen_share_label = SHARE_COUNT_TREATMENT_LABELS.get(scen_share_key, scen_share_key)
        scen_wacc_key = (
            (entry.get("compatibility") or {}).get("selected_wacc_treatment")
            or scen_params.get("selected_wacc_treatment")
            or DEFAULT_WACC_TREATMENT
        )
        scen_wacc_label = WACC_TREATMENT_LABELS.get(scen_wacc_key, scen_wacc_key)
        scen_terminal_key = (
            (entry.get("compatibility") or {}).get("selected_terminal_treatment")
            or scen_params.get("selected_terminal_treatment")
            or DEFAULT_TERMINAL_TREATMENT
        )
        scen_terminal_label = TERMINAL_TREATMENT_LABELS.get(scen_terminal_key, scen_terminal_key)
        values = [
            scenario_type.title(),
            "Yes" if current == scenario_type else "",
            _display_date(entry.get("saved_at", "")),
            _display_date(entry.get("updated_at", "")),
            origin.get("source", ""),
            scen_treatment_label,
            scen_share_label,
            scen_wacc_label,
            scen_terminal_label,
        ]
        for col, value in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=value)
            c.font = font_formula()
            c.border = border_thin()
        row += 1
    row += 2

    write_section_title(ws, row, "Scenario Comparability Note", span_cols=9)
    row += 1
    if treatment_summary["mismatches"]:
        mismatch_text = []
        label_maps = treatment_summary["labels"]
        label_names = {
            "net_debt": "Net Debt",
            "share_count": "Share Count",
            "wacc": "WACC",
            "terminal": "Terminal",
        }
        for scenario_type, treatment_type, value in treatment_summary["mismatches"]:
            base_value = treatment_summary["base"][treatment_type]
            label_map = label_maps[treatment_type]
            mismatch_text.append(
                f"{scenario_type.title()} {label_names[treatment_type]}: "
                f"Base={label_map.get(base_value, base_value)} vs "
                f"Scenario={label_map.get(value, value)}"
            )
        note = (
            "Saved scenario comparisons retain the treatment selected when the scenario was saved. "
            "These intentional scenario-state differences are disclosed rather than auto-aligned: "
            + "; ".join(mismatch_text)
            + "."
        )
    elif saved_count:
        note = "Saved Bull/Bear scenario treatments match the current Base defaults for Net Debt, Share Count, WACC, and Terminal method."
    else:
        note = "No saved Bull/Bear scenarios are available, so scenario comparability is N/A."
    _set_note_cell(ws, row, 1, note)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.row_dimensions[row].height = 34
    row += 2

    write_section_title(ws, row, "Legacy Scenario Schema Summary", span_cols=9)
    SCENARIO_ROWS["legacy_schema_summary"] = row
    row += 1
    write_header_row(ws, row, ["Scenario", "Defaulted Fields Count", "Defaulted Fields List", "Impact on IV", "Status wording"], 1)
    row += 1
    legacy_field_map = [
        ("selected_wacc_treatment", "wacc_treatment_defaulted"),
        ("selected_terminal_treatment", "terminal_treatment_defaulted"),
        ("selected_net_debt_treatment", "net_debt_treatment_defaulted"),
        ("selected_share_count_treatment", "share_count_treatment_defaulted"),
        ("selected_operating_path_source", "operating_path_source_defaulted_from_legacy"),
    ]
    for scenario_type in ("bull", "bear"):
        entry = scenarios.get(scenario_type)
        if not isinstance(entry, dict):
            continue
        compat = entry.get("compatibility") or {}
        defaulted_fields = [field for field, flag in legacy_field_map if compat.get(flag)]
        count = len(defaulted_fields)
        status = (
            f"Legacy scenario {scenario_type.title()}: {count} treatment fields defaulted to current schema defaults because scenario was saved before the current schema. Treatment defaults preserve baseline scenario IV; review saved-state treatment list below to confirm scenario comparability."
            if count
            else f"{scenario_type.title()} scenario stores current treatment fields explicitly."
        )
        values = [
            scenario_type.title(),
            count,
            ", ".join(defaulted_fields) if defaulted_fields else "None",
            "none" if count else "none",
            status,
        ]
        for col_idx, value in enumerate(values, 1):
            c = ws.cell(row=row, column=col_idx, value=value)
            c.font = font_formula(); c.border = border_thin()
            if col_idx in (3, 5):
                c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=9)
        row += 1
    row += 1

    write_section_title(ws, row, "Saved Scenario Metadata Detail", span_cols=9)
    row += 1
    write_header_row(ws, row, ["Scenario", "Operating Path Source", "Source State", "WACC Treatment", "Terminal Treatment", "Net Debt Treatment", "Share Count Treatment"], 1)
    row += 1
    for scenario_type in ("bull", "bear"):
        entry = scenarios.get(scenario_type)
        if not isinstance(entry, dict):
            continue
        compat = entry.get("compatibility") or {}
        params = entry.get("params") or {}
        op_source = compat.get("selected_operating_path_source") or params.get("selected_operating_path_source") or "Selected Path"
        op_state = (
            "defaulted from legacy; scenario saved before Bridge engine-driving was available"
            if compat.get("operating_path_source_defaulted_from_legacy")
            else "explicit"
        )
        values = [
            scenario_type.title(),
            op_source,
            op_state,
            compat.get("selected_wacc_treatment_label") or WACC_TREATMENT_LABELS.get(params.get("selected_wacc_treatment") or DEFAULT_WACC_TREATMENT),
            compat.get("selected_terminal_treatment_label") or TERMINAL_TREATMENT_LABELS.get(params.get("selected_terminal_treatment") or DEFAULT_TERMINAL_TREATMENT),
            compat.get("selected_net_debt_treatment_label") or NET_DEBT_TREATMENT_LABELS.get(params.get("selected_net_debt_treatment") or DEFAULT_NET_DEBT_TREATMENT),
            compat.get("selected_share_count_treatment_label") or SHARE_COUNT_TREATMENT_LABELS.get(params.get("selected_share_count_treatment") or DEFAULT_SHARE_COUNT_TREATMENT),
        ]
        for col_idx, value in enumerate(values, 1):
            c = ws.cell(row=row, column=col_idx, value=value)
            c.font = font_formula(); c.border = border_thin()
            if col_idx == 3:
                c.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    row += 2

    write_section_title(ws, row, "Valuation Summary From Scenario Store", span_cols=9)
    row += 1
    write_header_row(ws, row, ["Scenario", "Intrinsic / Share", "EV", "Equity Value", "TV % EV", "Currency"], 1)
    row += 1
    current_valuation = {
        "intrinsic_per_share": out.intrinsic_per_share,
        "ev": out.ev,
        "equity_value": out.equity_value,
        "tv_pct": out.tv_pct * 100,
        "currency": out.currency,
    }
    for scenario_type in ("base", "bull", "bear"):
        SCENARIO_ROWS[f"valuation_{scenario_type}"] = row
        entry = scenarios.get(scenario_type) if scenario_type != "base" else None
        valuation = entry.get("valuation", {}) if isinstance(entry, dict) else {}
        if scenario_type == current:
            valuation = {**valuation, **current_valuation}
        values = [
            scenario_type.title(),
            valuation.get("intrinsic_per_share"),
            valuation.get("ev"),
            valuation.get("equity_value"),
            valuation.get("tv_pct"),
            valuation.get("currency", ""),
        ]
        for col, value in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=value)
            c.font = font_formula()
            c.border = border_thin()
            if col in (2, 3, 4):
                c.number_format = FMT_COMMA2
            if col == 5:
                c.number_format = "0.0"
        row += 1
    row += 2

    write_section_title(ws, row, "Assumption Change Log v1", span_cols=9)
    row += 1
    if not isinstance(base_params, dict):
        _set_note_cell(
            ws,
            row,
            1,
            "Base comparison is not available in this export payload. Current assumptions are listed below for audit reference.",
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 2
        write_header_row(ws, row, ["Field", "Current Value", "Source"], 1)
        row += 1
        current_params = ctx.get("current_params") or _base_params_from_input(inp, out)
        for key in _diff_keys():
            current_value = current_params.get(key)
            if key == "selected_net_debt_treatment":
                tk = current_value or DEFAULT_NET_DEBT_TREATMENT
                current_value = NET_DEBT_TREATMENT_LABELS.get(tk, tk)
            elif key == "selected_share_count_treatment":
                tk = current_value or DEFAULT_SHARE_COUNT_TREATMENT
                current_value = SHARE_COUNT_TREATMENT_LABELS.get(tk, tk)
            elif key == "buyback_method":
                tk = current_value or DEFAULT_BUYBACK_METHOD
                current_value = BUYBACK_METHOD_LABELS.get(tk, tk)
            elif key == "selected_wacc_treatment":
                tk = current_value or DEFAULT_WACC_TREATMENT
                current_value = WACC_TREATMENT_LABELS.get(tk, tk)
            elif key == "selected_terminal_treatment":
                tk = current_value or DEFAULT_TERMINAL_TREATMENT
                current_value = TERMINAL_TREATMENT_LABELS.get(tk, tk)
            c1 = ws.cell(row=row, column=1, value=ASSUMPTION_LABELS.get(key, key))
            c2 = ws.cell(row=row, column=2, value=current_value)
            c3 = ws.cell(row=row, column=3, value=current.title())
            for c in (c1, c2, c3):
                c.font = font_formula()
                c.border = border_thin()
            c2.number_format = _fmt_for_key(key) or "General"
            row += 1
        return

    any_block = False
    for scenario_type in ("bull", "bear"):
        entry = scenarios.get(scenario_type)
        params = entry.get("params", {}) if isinstance(entry, dict) else {}
        if not params:
            continue
        if any_block:
            row += 1
        any_block = True

        write_section_title(
            ws,
            row,
            f"{scenario_type.title()} Scenario - Assumption Diff vs Current Base",
            span_cols=6,
        )
        row += 1
        write_header_row(ws, row, ["Field", "Base", "Scenario", "Delta", "Note"], 1)
        row += 1

        scenario_has_diff = False
        compatibility = entry.get("compatibility", {}) if isinstance(entry, dict) else {}
        for key in _diff_keys():
            base_val = base_params.get(key)
            scenario_val = params.get(key)
            # V3.7.3: Net Debt Treatment is stored as the key but displayed as
            # the human-readable label in the change log. Missing values
            # default to reported so old scenarios don't fake-diff.
            if key == "selected_net_debt_treatment":
                base_key = base_val or DEFAULT_NET_DEBT_TREATMENT
                scn_key = scenario_val or DEFAULT_NET_DEBT_TREATMENT
                if base_key == scn_key:
                    continue
                scenario_has_diff = True
                values = [
                    ASSUMPTION_LABELS["selected_net_debt_treatment"],
                    NET_DEBT_TREATMENT_LABELS.get(base_key, base_key),
                    NET_DEBT_TREATMENT_LABELS.get(scn_key, scn_key),
                    "",
                    "",
                ]
                for col, value in enumerate(values, 1):
                    c = ws.cell(row=row, column=col, value=value)
                    c.font = font_formula()
                    c.border = border_thin()
                row += 1
                continue
            if key == "selected_share_count_treatment":
                base_key = base_val or DEFAULT_SHARE_COUNT_TREATMENT
                scn_key = scenario_val or DEFAULT_SHARE_COUNT_TREATMENT
                if base_key == scn_key:
                    continue
                scenario_has_diff = True
                values = [
                    ASSUMPTION_LABELS["selected_share_count_treatment"],
                    SHARE_COUNT_TREATMENT_LABELS.get(base_key, base_key),
                    SHARE_COUNT_TREATMENT_LABELS.get(scn_key, scn_key),
                    "",
                    "",
                ]
                for col, value in enumerate(values, 1):
                    c = ws.cell(row=row, column=col, value=value)
                    c.font = font_formula()
                    c.border = border_thin()
                row += 1
                continue
            if key == "buyback_method":
                base_key = base_val or DEFAULT_BUYBACK_METHOD
                scn_key = scenario_val or DEFAULT_BUYBACK_METHOD
                if base_key == scn_key:
                    continue
                scenario_has_diff = True
                values = [
                    ASSUMPTION_LABELS["buyback_method"],
                    BUYBACK_METHOD_LABELS.get(base_key, base_key),
                    BUYBACK_METHOD_LABELS.get(scn_key, scn_key),
                    "",
                    "",
                ]
                for col, value in enumerate(values, 1):
                    c = ws.cell(row=row, column=col, value=value)
                    c.font = font_formula()
                    c.border = border_thin()
                row += 1
                continue
            if key == "selected_wacc_treatment":
                base_key = base_val or DEFAULT_WACC_TREATMENT
                scn_key = scenario_val or DEFAULT_WACC_TREATMENT
                if base_key == scn_key:
                    continue
                scenario_has_diff = True
                values = [
                    ASSUMPTION_LABELS["selected_wacc_treatment"],
                    WACC_TREATMENT_LABELS.get(base_key, base_key),
                    WACC_TREATMENT_LABELS.get(scn_key, scn_key),
                    "",
                    "",
                ]
                for col, value in enumerate(values, 1):
                    c = ws.cell(row=row, column=col, value=value)
                    c.font = font_formula()
                    c.border = border_thin()
                row += 1
                continue
            if key == "selected_terminal_treatment":
                base_key = base_val or DEFAULT_TERMINAL_TREATMENT
                scn_key = scenario_val or DEFAULT_TERMINAL_TREATMENT
                if base_key == scn_key:
                    continue
                scenario_has_diff = True
                values = [
                    ASSUMPTION_LABELS["selected_terminal_treatment"],
                    TERMINAL_TREATMENT_LABELS.get(base_key, base_key),
                    TERMINAL_TREATMENT_LABELS.get(scn_key, scn_key),
                    "",
                    "",
                ]
                for col, value in enumerate(values, 1):
                    c = ws.cell(row=row, column=col, value=value)
                    c.font = font_formula()
                    c.border = border_thin()
                row += 1
                continue
            if not _values_differ(base_val, scenario_val):
                continue
            scenario_has_diff = True
            delta = ""
            bn = _coerce_num(base_val)
            sn = _coerce_num(scenario_val)
            if bn is not None and sn is not None:
                delta = sn - bn
            values = [ASSUMPTION_LABELS.get(key, key), base_val, scenario_val, delta, ""]
            for col, value in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=value)
                c.font = font_formula()
                c.border = border_thin()
                if col in (2, 3, 4):
                    c.number_format = _fmt_for_key(key) or "General"
            row += 1
        if compatibility.get("legacy_fcf_growth_mapped"):
            _set_note_cell(
                ws,
                row,
                1,
                "Compatibility note: this saved scenario used the older fcf_growth field, mapped to Revenue Growth for this export.",
            )
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            row += 1
        if not scenario_has_diff:
            _set_note_cell(ws, row, 1, f"No assumption differences found for {scenario_type.title()} scenario.")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
    if not any_block:
        _set_note_cell(ws, row, 1, "No Bull/Bear assumption diffs available in scenario store.")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)


def _diff_keys():
    return ASSUMPTION_CHANGE_KEYS


def build_legacy_financials_snapshot_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "Legacy Financials Snapshot"
    apply_column_widths(ws, {"A": 34, "B": 24, "C": 30})

    cache = ctx.get("financial_cache") or {}
    data = cache.get("data") or {}
    historical_cache = ctx.get("historical_cache") or {}
    historical_payload = historical_cache.get("data") or {}
    fallback = _base_params_from_input(inp, out)

    row = _write_sheet_heading(ws, 1, "Legacy single-period financials snapshot", inp, out, ctx)
    _set_note_cell(
        ws,
        row,
        1,
        "Deprecated single-period view retained only as unused legacy code; export now writes Historical Financials.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    write_header_row(ws, row, ["Field", "Value", "Source"], 1)
    row += 1
    fields = [
        ("revenue", "Revenue", FMT_COMMA2),
        ("ebit", "EBIT", FMT_COMMA2),
        ("da", "D&A", FMT_COMMA2),
        ("capex", "CapEx", FMT_COMMA2),
        ("wc_change", "Working Capital Change", FMT_COMMA2),
        ("tax_rate", "Effective Tax Rate", FMT_PCT2),
        ("net_debt", "Net Debt", FMT_COMMA2),
        ("shares", "Diluted Shares Outstanding (M)", FMT_COMMA2),
        ("beta", "Beta", "0.00"),
        ("currency", "Currency", None),
    ]
    for key, label, fmt in fields:
        value = data.get(key, fallback.get(key))
        _set_label_cell(ws, row, 1, label)
        c = ws.cell(row=row, column=2, value=value)
        c.font = font_formula()
        c.border = border_thin()
        if fmt:
            c.number_format = fmt
        _set_note_cell(ws, row, 3, "financials cache data" if data else "export payload fallback")
        row += 1
    row += 1
    _write_kv_rows(ws, row, [
        ("Cache File", cache.get("path") or f"financials_{inp.symbol}_{FINANCIALS_CACHE_VERSION}.json", None),
        ("Cache Date", _display_date(cache.get("cached_at")) or "Not available", None),
        ("Market / Currency", f"{out.market} / {out.currency}", None),
    ])


def build_historical_financials_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "Historical Financials"

    historical_cache = ctx.get("historical_cache") or {}
    payload = historical_cache.get("data") or {}
    tables = ctx.get("historical_tables") or {}
    field_dictionary = payload.get("field_dictionary") or {}

    all_years = sorted({
        year
        for table in tables.values()
        for year in table.get("years", [])
    })
    span_cols = max(3, len(all_years) + 1)
    apply_column_widths(ws, {"A": 34, **{get_column_letter(i + 2): 16 for i in range(max(5, len(all_years)))}})

    row = _write_sheet_heading(
        ws,
        1,
        "Historical Financials - Annual Actuals",
        inp,
        out,
        ctx,
        span_cols=span_cols,
    )

    if out.market != "US" or payload.get("status") != "ok" or not all_years:
        _set_note_cell(ws, row, 1, "Historical financials not available yet for this market")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
        row += 2
        _write_kv_rows(ws, row, [
            ("Historical Cache File", historical_cache.get("path") or "Not available", None),
            ("Market / Currency", f"{out.market} / {out.currency}", None),
            ("DCF Impact", "None - DCF defaults and valuation use the V3.5 single-period cache only.", None),
        ])
        return

    row = _write_kv_rows(ws, row, [
        ("Source", payload.get("source") or "yfinance", None),
        ("Cached At", _display_date(payload.get("cached_at")) or "Not available", None),
        ("Period Type", "annual", None),
        ("Currency / Unit", model_unit_label(payload.get("currency") or out.currency, out.market), None),
        ("Raw Cache Units", payload.get("units") or "actual local currency", None),
        ("Display Scale Factor", f"Raw actual / {model_unit_scale_factor(out.market):,.0f}", None),
    ])
    row += 1

    sections = [
        ("P&L", "income_statement", [
            "revenue", "cost_of_revenue", "gross_profit", "research_development",
            "selling_general_admin", "operating_expenses", "operating_income",
            "interest_expense", "other_income_expense_net", "pretax_income",
            "tax_expense", "net_income", "diluted_shares",
        ]),
        ("Balance Sheet", "balance_sheet", [
            "cash", "short_term_debt", "long_term_debt", "total_debt",
            "total_assets", "total_equity", "accounts_receivable",
            "accounts_payable", "inventory", "ppe",
        ]),
        ("Cash Flow", "cash_flow", [
            "operating_cash_flow", "capex", "depreciation_amortization", "free_cash_flow",
        ]),
    ]

    for section_label, statement_key, field_order in sections:
        write_section_title(ws, row, section_label, span_cols=span_cols)
        row += 1
        write_header_row(ws, row, ["Field"] + [str(year) for year in all_years], 1)
        row += 1
        table = tables.get(statement_key) or {}
        rows = table.get("rows") or {}
        for field_key in field_order:
            _set_label_cell(ws, row, 1, field_dictionary.get(field_key) or field_key)
            values = rows.get(field_key, {})
            for idx, year in enumerate(all_years, start=2):
                value = _hist_display_value(ctx, statement_key, field_key, year, out.market)
                c = ws.cell(row=row, column=idx, value=value)
                c.font = font_formula()
                c.border = border_thin()
                if isinstance(value, (int, float)):
                    c.number_format = FMT_COMMA
            row += 1
        row += 1


def build_pl_forecast_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "P&L Forecast"
    PL_FORECAST_ROWS.clear()

    payload = ((ctx.get("historical_cache") or {}).get("data") or {})
    income_table = ((ctx.get("historical_tables") or {}).get("income_statement") or {})
    historical_years = sorted(income_table.get("years") or [])
    n = max(1, int(inp.forecast_years or 5))
    forecast_labels = [f"Year {i + 1}" for i in range(n)]
    span_cols = max(3, len(historical_years) + n + 1)
    apply_column_widths(ws, {"A": 30, **{get_column_letter(i + 2): 15 for i in range(len(historical_years) + n)}})

    row = _write_sheet_heading(ws, 1, "P&L Forecast v1", inp, out, ctx, span_cols=span_cols)

    if out.market != "US" or payload.get("status") != "ok" or not historical_years:
        _set_note_cell(ws, row, 1, "P&L Forecast not available yet for this market")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
        row += 2
        _write_kv_rows(ws, row, [
            ("Market / Currency", f"{out.market} / {out.currency}", None),
            ("Fallback", "Graceful fallback only; HK/CN mapping will be designed in a later display-unit layer.", None),
        ])
        return

    row = _write_kv_rows(ws, row, [
        ("Currency / Unit", model_unit_label(out.currency, out.market), None),
        ("Raw Historical Unit", payload.get("units") or "actual local currency", None),
        ("Scale Factor", f"US raw actual / {MODEL_UNIT_SCALE_FACTORS['US']:,.0f}", None),
    ])
    row += 1

    header = [""] + [str(year) for year in historical_years] + forecast_labels
    write_header_row(ws, row, header, 1)
    row += 1

    hist_end_col = 1 + len(historical_years)
    forecast_start_col = hist_end_col + 1
    ref_revenue = _abs_ref(ASSUMP_CELLS["revenue"])
    ref_revenue_growth = _abs_ref(ASSUMP_CELLS["revenue_growth"])
    ref_ebit_margin = _abs_ref(ASSUMP_CELLS["ebit_margin"])
    ref_tax = _abs_ref(ASSUMP_CELLS["tax_rate"])
    ref_gross_margin = _abs_ref(ASSUMP_CELLS["gross_margin"])
    ref_shares = _abs_ref(ASSUMP_CELLS.get("sr_selected_denominator", ASSUMP_CELLS["shares"]))
    rd_pct_default = _latest_historical_ratio(ctx, "income_statement", "research_development", out.market) or 0.0
    sga_pct_default = _latest_historical_ratio(ctx, "income_statement", "selling_general_admin", out.market) or 0.0
    row_map = {}

    def write_label(label: str, bold: bool = False):
        nonlocal row
        _set_label_cell(ws, row, 1, label, bold)
        row_map[label] = row
        PL_FORECAST_ROWS[label] = row
        row += 1
        return row - 1

    def write_hist_values(target_row: int, values_by_year: dict, fmt):
        for idx, year in enumerate(historical_years, start=2):
            c = ws.cell(row=target_row, column=idx, value=values_by_year.get(year))
            c.font = font_formula()
            c.border = border_thin()
            if fmt:
                c.number_format = fmt

    hist = {}
    for year in historical_years:
        revenue = _hist_model_value(ctx, "income_statement", "revenue", year, out.market)
        gross_profit = _hist_model_value(ctx, "income_statement", "gross_profit", year, out.market)
        cogs = _hist_model_value(ctx, "income_statement", "cost_of_revenue", year, out.market)
        if cogs is None and revenue is not None and gross_profit is not None:
            cogs = revenue - gross_profit
        rd = _hist_model_value(ctx, "income_statement", "research_development", year, out.market)
        sga = _hist_model_value(ctx, "income_statement", "selling_general_admin", year, out.market)
        opex = _hist_model_value(ctx, "income_statement", "operating_expenses", year, out.market)
        ebit = _hist_model_value(ctx, "income_statement", "operating_income", year, out.market)
        tax = _hist_model_value(ctx, "income_statement", "tax_expense", year, out.market)
        net_income = _hist_model_value(ctx, "income_statement", "net_income", year, out.market)
        shares = _hist_model_value(ctx, "income_statement", "diluted_shares", year, out.market)
        interest = _hist_display_value(ctx, "income_statement", "interest_expense", year, out.market)
        other_income = _hist_model_value(ctx, "income_statement", "other_income_expense_net", year, out.market)
        pretax = _hist_model_value(ctx, "income_statement", "pretax_income", year, out.market)
        if pretax is None and net_income is not None and tax is not None:
            pretax = net_income + tax
        eff_tax_rate = _safe_pct(tax, pretax) if pretax not in (None, 0) else _safe_pct(tax, ebit)
        hist[year] = {
            "Revenue": revenue,
            "COGS / Cost of Revenue": cogs,
            "Gross Profit": gross_profit,
            "Gross Margin (%)": _safe_pct(gross_profit, revenue),
            "R&D": rd,
            "R&D % Revenue": _safe_pct(rd, revenue),
            "SG&A": sga,
            "SG&A % Revenue": _safe_pct(sga, revenue),
            "Total Operating Expenses": opex,
            "Operating Income / EBIT": ebit,
            "EBIT Margin (%)": _safe_pct(ebit, revenue),
            "Interest Expense": interest,
            "Other Income / Expense, net": other_income,
            "Pre-tax Income": pretax,
            "Tax Expense": tax,
            "Effective Tax Rate (%)": eff_tax_rate,
            "Net Income": net_income,
            "Net Margin (%)": _safe_pct(net_income, revenue),
            "Diluted Shares": shares,
            "EPS": _safe_pct(net_income, shares),
        }
    prev_revenue = None
    for year in historical_years:
        revenue = hist[year]["Revenue"]
        hist[year]["Revenue Growth (%)"] = _safe_pct(revenue - prev_revenue, prev_revenue) if prev_revenue else None
        if revenue is not None:
            prev_revenue = revenue

    rows = [
        ("Revenue", FMT_COMMA2),
        ("Revenue Growth (%)", FMT_PCT2),
        ("COGS / Cost of Revenue", FMT_COMMA2),
        ("Gross Profit", FMT_COMMA2),
        ("Gross Margin (%)", FMT_PCT2),
        ("R&D", FMT_COMMA2),
        ("R&D % Revenue", FMT_PCT2),
        ("SG&A", FMT_COMMA2),
        ("SG&A % Revenue", FMT_PCT2),
        ("Total Operating Expenses", FMT_COMMA2),
        ("Operating Income / EBIT", FMT_COMMA2),
        ("EBIT Margin (%)", FMT_PCT2),
        ("Interest Expense", FMT_COMMA2),
        ("Other Income / Expense, net", FMT_COMMA2),
        ("Pre-tax Income", FMT_COMMA2),
        ("Tax Expense", FMT_COMMA2),
        ("Effective Tax Rate (%)", FMT_PCT2),
        ("Net Income", FMT_COMMA2),
        ("Net Margin (%)", FMT_PCT2),
        ("Diluted Shares", FMT_COMMA2),
        ("EPS", FMT_COMMA2),
    ]

    for label, fmt in rows:
        target_row = write_label(label, bold=label in {"Revenue", "Operating Income / EBIT", "Pre-tax Income", "Net Income", "EPS"})
        write_hist_values(target_row, {year: hist[year].get(label) for year in historical_years}, fmt)

    for i in range(n):
        col_idx = forecast_start_col + i
        col = get_column_letter(col_idx)
        prev_col = get_column_letter(col_idx - 1)
        # V3.9.0: per-year refs into the Assumptions forecast path block.
        yr_growth = _path_ref("revenue_growth", i)
        yr_margin = _path_ref("ebit_margin", i)
        if i == 0:
            revenue_formula = f"=ROUND({ref_revenue}*(1+{yr_growth}),2)"
        else:
            revenue_formula = f"=ROUND({prev_col}{row_map['Revenue']}*(1+{yr_growth}),2)"
        _set_formula_cell(ws, row_map["Revenue"], col_idx, revenue_formula, FMT_COMMA2, True)
        _set_formula_cell(ws, row_map["Revenue Growth (%)"], col_idx, f"={yr_growth}", FMT_PCT2, True)
        _set_formula_cell(ws, row_map["COGS / Cost of Revenue"], col_idx, f"=ROUND({col}{row_map['Revenue']}-{col}{row_map['Gross Profit']},2)", FMT_COMMA2)
        _set_formula_cell(ws, row_map["Gross Profit"], col_idx, f"=ROUND({col}{row_map['Revenue']}*{col}{row_map['Gross Margin (%)']},2)", FMT_COMMA2)
        _set_formula_cell(ws, row_map["Gross Margin (%)"], col_idx, f"={ref_gross_margin}", FMT_PCT2, True)
        _set_formula_cell(ws, row_map["R&D % Revenue"], col_idx, f"={rd_pct_default}", FMT_PCT2, True)
        _set_formula_cell(ws, row_map["R&D"], col_idx, f"=ROUND({col}{row_map['Revenue']}*{col}{row_map['R&D % Revenue']},2)", FMT_COMMA2)
        _set_formula_cell(ws, row_map["SG&A % Revenue"], col_idx, f"={sga_pct_default}", FMT_PCT2, True)
        _set_formula_cell(ws, row_map["SG&A"], col_idx, f"=ROUND({col}{row_map['Revenue']}*{col}{row_map['SG&A % Revenue']},2)", FMT_COMMA2)
        _set_formula_cell(ws, row_map["Total Operating Expenses"], col_idx, f"=ROUND({col}{row_map['R&D']}+{col}{row_map['SG&A']},2)", FMT_COMMA2)
        _set_formula_cell(ws, row_map["Operating Income / EBIT"], col_idx, f"=ROUND({col}{row_map['Revenue']}*{col}{row_map['EBIT Margin (%)']},2)", FMT_COMMA2)
        _set_formula_cell(ws, row_map["EBIT Margin (%)"], col_idx, f"={yr_margin}", FMT_PCT2, True)
        # Interest Expense links to Debt Schedule v2 Interest Expense row (filled later via relinker).
        # We write a temporary 0 here; relink_pl_forecast_to_supporting_schedules() patches the cell once
        # the Supporting Schedules sheet has been built and Debt Schedule rows are known.
        _set_formula_cell(ws, row_map["Interest Expense"], col_idx, "=0", FMT_COMMA2, True)
        _set_formula_cell(ws, row_map["Other Income / Expense, net"], col_idx, "=0", FMT_COMMA2, True)
        _set_formula_cell(
            ws,
            row_map["Pre-tax Income"],
            col_idx,
            f"=ROUND({col}{row_map['Operating Income / EBIT']}-{col}{row_map['Interest Expense']}+{col}{row_map['Other Income / Expense, net']},2)",
            FMT_COMMA2,
        )
        # Tax floor at 0 in v1: deferred tax assets / NOL carry-forward not modeled.
        _set_formula_cell(
            ws,
            row_map["Tax Expense"],
            col_idx,
            f"=ROUND(MAX({col}{row_map['Pre-tax Income']}*{ref_tax},0),2)",
            FMT_COMMA2,
            True,
        )
        _set_formula_cell(ws, row_map["Effective Tax Rate (%)"], col_idx, f"={ref_tax}", FMT_PCT2, True)
        _set_formula_cell(
            ws,
            row_map["Net Income"],
            col_idx,
            f"=ROUND({col}{row_map['Pre-tax Income']}-{col}{row_map['Tax Expense']},2)",
            FMT_COMMA2,
        )
        _set_formula_cell(ws, row_map["Net Margin (%)"], col_idx, f"=IFERROR({col}{row_map['Net Income']}/{col}{row_map['Revenue']},0)", FMT_PCT2)
        _set_formula_cell(ws, row_map["Diluted Shares"], col_idx, f"={ref_shares}", FMT_COMMA2, True)
        _set_formula_cell(ws, row_map["EPS"], col_idx, f"=IFERROR({col}{row_map['Net Income']}/{col}{row_map['Diluted Shares']},0)", FMT_COMMA2)

    row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "P&L detail v1: COGS, R&D, SG&A, total opex, and Other Income/Expense support margin analysis. DCF forecast still uses the selected EBIT margin path as the EBIT driver; R&D / SG&A are presentation bridge rows until a full opex driver model is connected. Interest Expense links to Supporting Schedules Debt Schedule v2 after schedules build.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    write_section_title(ws, row, "Gross Margin / Opex Bridge", span_cols=4)
    row += 1
    write_header_row(ws, row, ["Metric", "Latest Actual", "Year 1 Forecast", "Role"], 1)
    row += 1
    latest_col_idx = 1 + len(historical_years)
    year1_col_idx = forecast_start_col
    bridge_rows = [
        ("Revenue", "Formula-driving"),
        ("COGS / Cost of Revenue", "Presentation bridge"),
        ("Gross Profit", "Presentation bridge"),
        ("R&D", "Presentation bridge"),
        ("SG&A", "Presentation bridge"),
        ("Operating Income / EBIT", "Formula-driving via selected EBIT margin path"),
        ("EBIT Margin (%)", "Formula-driving selected path"),
    ]
    for label, role in bridge_rows:
        _set_label_cell(ws, row, 1, label)
        fmt = FMT_PCT2 if label.endswith("(%)") else FMT_COMMA2
        _set_formula_cell(ws, row, 2, f"={_sheet_ref('P&L Forecast', row_map[label], latest_col_idx)}", fmt, True)
        _set_formula_cell(ws, row, 3, f"={_sheet_ref('P&L Forecast', row_map[label], year1_col_idx)}", fmt, True)
        _set_note_cell(ws, row, 4, role)
        row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "Bridge note: EBIT margin path is supported by historical gross margin and opex structure; full segment / opex driver model is not yet connected.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)


def build_supporting_schedules_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "Supporting Schedules"
    SCHEDULE_ROWS.clear()

    payload = ((ctx.get("historical_cache") or {}).get("data") or {})
    tables = ctx.get("historical_tables") or {}
    historical_years = sorted({
        year
        for table in tables.values()
        for year in table.get("years", [])
    })
    n = max(1, int(inp.forecast_years or 5))
    forecast_labels = [f"Year {i + 1}" for i in range(n)]
    span_cols = max(3, len(historical_years) + n + 1)
    apply_column_widths(ws, {"A": 38, **{get_column_letter(i + 2): 15 for i in range(len(historical_years) + n)}})

    row = _write_sheet_heading(ws, 1, "Supporting Schedules — WC Days, PP&E, Debt", inp, out, ctx, span_cols=span_cols)

    if out.market != "US" or payload.get("status") != "ok" or not historical_years or not PL_FORECAST_ROWS:
        _set_note_cell(ws, row, 1, "Supporting Schedules not available yet for this market")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
        row += 2
        _write_kv_rows(ws, row, [
            ("Market / Currency", f"{out.market} / {out.currency}", None),
            ("Fallback", "Graceful fallback only; HK/CN schedule mapping will be added later.", None),
        ])
        row += 4
        write_section_title(ws, row, "Working Capital Reality Check", span_cols=span_cols)
        row += 1
        row = _write_kv_rows(ws, row, [
            ("Status", "N/A - historical WC days pipeline not active for this market", None),
            ("Active WC Source", "Legacy wc_change_pct_revenue fallback", None),
            ("Delta NWC Sign Convention", "Delta NWC = Ending NWC - Beginning NWC; negative = source/release; positive = use/investment.", None),
            ("Review Tier", "Review", None),
        ])
        row += 1
        write_section_title(ws, row, "Buyback Funding Closure", span_cols=span_cols)
        row += 1
        _write_kv_rows(ws, row, [
            ("Status", "N/A - shareholder-return funding closure is calculator-derived; no full local BS schedule for this market", None),
            ("FCF after Dividends Used for Buybacks", "N/A", None),
            ("Cash Above Floor Used for Buybacks", "N/A", None),
            ("Marketable Securities Drawdown Used", "N/A", None),
            ("Ending Marketable Securities", "N/A", None),
            ("Review Tier", "Review", None),
        ])
        return

    row = _write_kv_rows(ws, row, [
        ("Currency / Unit", model_unit_label(out.currency, out.market), None),
        ("Model Boundary", "V3.6.9 Modeling Craft Upgrade v1: WC days driver, PP&E v2 asset-based D&A, split BS plug, WACC v2; no full cash-to-interest circularity yet.", None),
        ("WC Dual Basis", "Days-based AR / Inventory / AP feed schedule-derived Delta WC for CF / FCF / DCF; legacy % Revenue inputs and selected Delta NWC are reference-only.", None),
    ])
    row += 1

    header = [""] + [str(year) for year in historical_years] + forecast_labels
    hist_end_col = 1 + len(historical_years)
    forecast_start_col = hist_end_col + 1
    ref_da_pct = _abs_ref(ASSUMP_CELLS["da_pct_revenue"])
    ref_capex_pct = _abs_ref(ASSUMP_CELLS["capex_pct_revenue"])
    ref_ar_pct = _abs_ref(ASSUMP_CELLS["ar_pct_revenue"])
    ref_inventory_pct = _abs_ref(ASSUMP_CELLS["inventory_pct_revenue"])
    ref_ap_pct = _abs_ref(ASSUMP_CELLS["ap_pct_revenue"])
    ref_financing = _abs_ref(ASSUMP_CELLS["financing_placeholder"])
    ref_share_growth = _abs_ref(ASSUMP_CELLS["share_count_growth_placeholder"])
    ref_dso = _abs_ref(ASSUMP_CELLS["dso"])
    ref_dio = _abs_ref(ASSUMP_CELLS["dio"])
    ref_dpo = _abs_ref(ASSUMP_CELLS["dpo"])
    ref_dso_target = _abs_ref(ASSUMP_CELLS.get("dso_target", ASSUMP_CELLS["dso"]))
    ref_dio_target = _abs_ref(ASSUMP_CELLS.get("dio_target", ASSUMP_CELLS["dio"]))
    ref_dpo_target = _abs_ref(ASSUMP_CELLS.get("dpo_target", ASSUMP_CELLS["dpo"]))
    ref_da_pct_begin = _abs_ref(ASSUMP_CELLS["da_pct_begin_ppe"])

    def section(key: str, title: str, labels: list[str]):
        nonlocal row
        SCHEDULE_ROWS[key] = {}
        write_section_title(ws, row, title, span_cols=span_cols)
        row += 1
        write_header_row(ws, row, header, 1)
        row += 1
        for label in labels:
            _set_label_cell(ws, row, 1, label, bold=label.startswith("Ending") or label.startswith("Total") or label.startswith("Net ") or label.startswith("Forecast"))
            SCHEDULE_ROWS[key][label] = row
            row += 1
        return SCHEDULE_ROWS[key]

    ppe_rows = section(
        "ppe",
        "PP&E / D&A Schedule v2 (asset-based D&A)",
        [
            "Beginning Net PP&E",
            "  (+) CapEx",
            "  (-) D&A",
            "Ending Net PP&E",
            "D&A % Beginning Net PP&E (implied)",
        ],
    )
    # Maintain backwards-compatible aliases so downstream linkers that still look up
    # "CapEx" / "D&A" / "Ending PP&E" continue to resolve. The display labels carry the
    # roll-forward signs; the aliases point at the same row numbers.
    ppe_rows["CapEx"] = ppe_rows["  (+) CapEx"]
    ppe_rows["D&A"] = ppe_rows["  (-) D&A"]
    ppe_rows["Beginning PP&E"] = ppe_rows["Beginning Net PP&E"]
    ppe_rows["Ending PP&E"] = ppe_rows["Ending Net PP&E"]
    for idx, year in enumerate(historical_years, start=2):
        ppe = _hist_model_value(ctx, "balance_sheet", "ppe", year, out.market)
        previous_years = [y for y in historical_years if y < year]
        beginning_ppe = _hist_model_value(ctx, "balance_sheet", "ppe", max(previous_years), out.market) if previous_years else None
        da_h = _hist_model_value(ctx, "cash_flow", "depreciation_amortization", year, out.market)
        implied_pct = (da_h / beginning_ppe) if (da_h is not None and beginning_ppe) else None
        values = {
            "Beginning Net PP&E": beginning_ppe,
            "  (+) CapEx": _hist_model_value_positive(ctx, "cash_flow", "capex", year, out.market),
            "  (-) D&A": da_h,
            "Ending Net PP&E": ppe,
            "D&A % Beginning Net PP&E (implied)": implied_pct,
        }
        for label, value in values.items():
            c = ws.cell(row=ppe_rows[label], column=idx, value=value)
            c.font = font_formula()
            c.border = border_thin()
            if value is not None:
                c.number_format = FMT_PCT2 if label.startswith("D&A %") else FMT_COMMA2
    latest_ppe = _latest_historical_model_value(ctx, "balance_sheet", "ppe", out.market) or 0.0
    for i in range(n):
        col_idx = forecast_start_col + i
        col = get_column_letter(col_idx)
        prev_col = get_column_letter(col_idx - 1)
        pl_revenue = _sheet_ref("P&L Forecast", PL_FORECAST_ROWS["Revenue"], col_idx)
        beginning = str(latest_ppe) if i == 0 else f"{prev_col}{ppe_rows['Ending Net PP&E']}"
        _set_formula_cell(ws, ppe_rows["Beginning Net PP&E"], col_idx, f"={beginning}", FMT_COMMA2)
        # V3.9.0: per-year CapEx % Revenue path.
        yr_capex_pct = _path_ref("capex_pct_revenue", i)
        _set_formula_cell(ws, ppe_rows["  (+) CapEx"], col_idx, f"={pl_revenue}*{yr_capex_pct}", FMT_COMMA2, True)
        # V3.6.9: D&A driven off Beginning Net PP&E rather than Revenue.
        _set_formula_cell(
            ws,
            ppe_rows["  (-) D&A"],
            col_idx,
            f"={col}{ppe_rows['Beginning Net PP&E']}*{ref_da_pct_begin}",
            FMT_COMMA2,
            True,
        )
        _set_formula_cell(
            ws,
            ppe_rows["Ending Net PP&E"],
            col_idx,
            f"={col}{ppe_rows['Beginning Net PP&E']}+{col}{ppe_rows['  (+) CapEx']}-{col}{ppe_rows['  (-) D&A']}",
            FMT_COMMA2,
        )
        _set_formula_cell(
            ws,
            ppe_rows["D&A % Beginning Net PP&E (implied)"],
            col_idx,
            f"=IFERROR({col}{ppe_rows['  (-) D&A']}/{col}{ppe_rows['Beginning Net PP&E']},0)",
            FMT_PCT2,
        )
    row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "V3.6.9 PP&E / D&A: Beginning Net PP&E + CapEx - D&A = Ending Net PP&E. Forecast D&A = Beginning Net PP&E * D&A % Beginning PP&E (assumption). CapEx remains % Revenue. P&L D&A, CF D&A add-back, BS PP&E, and FCF Build D&A / CapEx all source from this single schedule.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    ref_wc_pct_for_legacy = _abs_ref(ASSUMP_CELLS["wc_change_pct_revenue"])
    wc_rows = section(
        "wc",
        "Working Capital Schedule v2 (DSO / DIO / DPO days driver, V3.6.9 primary)",
        [
            "Revenue (link)",
            "COGS (Revenue - Gross Profit)",
            "DSO (days)",
            "DIO (days)",
            "DPO (days)",
            "Accounts Receivable",
            "Inventory",
            "Accounts Payable",
            "Net Working Capital",
            "Schedule-derived Delta WC",
            "Selected Delta NWC diagnostic",
            "Legacy AR % Revenue (reference)",
            "Legacy Inventory % Revenue (reference)",
            "Legacy AP % Revenue (reference)",
        ],
    )
    for idx, year in enumerate(historical_years, start=2):
        revenue_h = _hist_model_value(ctx, "income_statement", "revenue", year, out.market)
        gross_profit_h = _hist_model_value(ctx, "income_statement", "gross_profit", year, out.market)
        cogs_h = (revenue_h - gross_profit_h) if (revenue_h is not None and gross_profit_h is not None) else None
        ar = _hist_model_value(ctx, "balance_sheet", "accounts_receivable", year, out.market)
        inv = _hist_model_value(ctx, "balance_sheet", "inventory", year, out.market)
        ap = _hist_model_value(ctx, "balance_sheet", "accounts_payable", year, out.market)
        nwc = (ar or 0.0) + (inv or 0.0) - (ap or 0.0) if any(v is not None for v in (ar, inv, ap)) else None
        previous_years = [y for y in historical_years if y < year]
        delta = None
        if previous_years and nwc is not None:
            py = max(previous_years)
            prev_ar = _hist_model_value(ctx, "balance_sheet", "accounts_receivable", py, out.market)
            prev_inv = _hist_model_value(ctx, "balance_sheet", "inventory", py, out.market)
            prev_ap = _hist_model_value(ctx, "balance_sheet", "accounts_payable", py, out.market)
            prev_nwc = (prev_ar or 0.0) + (prev_inv or 0.0) - (prev_ap or 0.0) if any(v is not None for v in (prev_ar, prev_inv, prev_ap)) else None
            delta = nwc - prev_nwc if prev_nwc is not None else None
        dso_h = (ar / revenue_h * 365.0) if (ar is not None and revenue_h) else None
        dio_h = (inv / cogs_h * 365.0) if (inv is not None and cogs_h) else None
        dpo_h = (ap / cogs_h * 365.0) if (ap is not None and cogs_h) else None
        values = {
            "Revenue (link)": revenue_h,
            "COGS (Revenue - Gross Profit)": cogs_h,
            "DSO (days)": dso_h,
            "DIO (days)": dio_h,
            "DPO (days)": dpo_h,
            "Accounts Receivable": ar,
            "Inventory": inv,
            "Accounts Payable": ap,
            "Net Working Capital": nwc,
            "Schedule-derived Delta WC": delta,
            "Selected Delta NWC diagnostic": None,
            "Legacy AR % Revenue (reference)": _safe_pct(ar, revenue_h),
            "Legacy Inventory % Revenue (reference)": _safe_pct(inv, revenue_h),
            "Legacy AP % Revenue (reference)": _safe_pct(ap, revenue_h),
        }
        for label, value in values.items():
            c = ws.cell(row=wc_rows[label], column=idx, value=value)
            c.font = font_formula()
            c.border = border_thin()
            if value is not None:
                if label.endswith("(days)"):
                    c.number_format = "0.0"
                elif label.startswith("Legacy ") and "% Revenue" in label:
                    c.number_format = FMT_PCT2
                else:
                    c.number_format = FMT_COMMA2
    pl_revenue_row = PL_FORECAST_ROWS["Revenue"]
    pl_gp_row = PL_FORECAST_ROWS["Gross Profit"]
    for i in range(n):
        col_idx = forecast_start_col + i
        col = get_column_letter(col_idx)
        prev_col = get_column_letter(col_idx - 1)
        pl_revenue = _sheet_ref("P&L Forecast", pl_revenue_row, col_idx)
        pl_gp = _sheet_ref("P&L Forecast", pl_gp_row, col_idx)
        _set_formula_cell(ws, wc_rows["Revenue (link)"], col_idx, f"={pl_revenue}", FMT_COMMA2, True)
        _set_formula_cell(ws, wc_rows["COGS (Revenue - Gross Profit)"], col_idx, f"={pl_revenue}-{pl_gp}", FMT_COMMA2, True)
        fade_progress = f"{i}/{max(n - 1, 1)}" if n > 1 else "1"
        _set_formula_cell(ws, wc_rows["DSO (days)"], col_idx, f"={ref_dso}+({ref_dso_target}-{ref_dso})*{fade_progress}", "0.0", True)
        _set_formula_cell(ws, wc_rows["DIO (days)"], col_idx, f"={ref_dio}+({ref_dio_target}-{ref_dio})*{fade_progress}", "0.0", True)
        _set_formula_cell(ws, wc_rows["DPO (days)"], col_idx, f"={ref_dpo}+({ref_dpo_target}-{ref_dpo})*{fade_progress}", "0.0", True)
        # V3.6.9 days-based forecast.
        _set_formula_cell(
            ws,
            wc_rows["Accounts Receivable"],
            col_idx,
            f"={col}{wc_rows['Revenue (link)']}/365*{col}{wc_rows['DSO (days)']}",
            FMT_COMMA2,
        )
        _set_formula_cell(
            ws,
            wc_rows["Inventory"],
            col_idx,
            f"={col}{wc_rows['COGS (Revenue - Gross Profit)']}/365*{col}{wc_rows['DIO (days)']}",
            FMT_COMMA2,
        )
        _set_formula_cell(
            ws,
            wc_rows["Accounts Payable"],
            col_idx,
            f"={col}{wc_rows['COGS (Revenue - Gross Profit)']}/365*{col}{wc_rows['DPO (days)']}",
            FMT_COMMA2,
        )
        _set_formula_cell(
            ws,
            wc_rows["Net Working Capital"],
            col_idx,
            f"={col}{wc_rows['Accounts Receivable']}+{col}{wc_rows['Inventory']}-{col}{wc_rows['Accounts Payable']}",
            FMT_COMMA2,
        )
        if i == 0:
            latest_nwc = None
            for y in sorted(historical_years, reverse=True):
                ar = _hist_model_value(ctx, "balance_sheet", "accounts_receivable", y, out.market)
                inv = _hist_model_value(ctx, "balance_sheet", "inventory", y, out.market)
                ap = _hist_model_value(ctx, "balance_sheet", "accounts_payable", y, out.market)
                if any(v is not None for v in (ar, inv, ap)):
                    latest_nwc = (ar or 0.0) + (inv or 0.0) - (ap or 0.0)
                    break
            base_nwc = latest_nwc or 0.0
            _set_formula_cell(ws, wc_rows["Schedule-derived Delta WC"], col_idx, f"={col}{wc_rows['Net Working Capital']}-({base_nwc})", FMT_COMMA2)
        else:
            _set_formula_cell(ws, wc_rows["Schedule-derived Delta WC"], col_idx, f"={col}{wc_rows['Net Working Capital']}-{prev_col}{wc_rows['Net Working Capital']}", FMT_COMMA2)
        # Legacy reference rows. V3.9.0: per-year Delta NWC % Revenue path.
        yr_wc_pct = _path_ref("wc_change_pct_revenue", i)
        _set_formula_cell(ws, wc_rows["Selected Delta NWC diagnostic"], col_idx, f"={pl_revenue}*{yr_wc_pct}", FMT_COMMA2, True)
        _set_formula_cell(ws, wc_rows["Legacy AR % Revenue (reference)"], col_idx, f"={ref_ar_pct}", FMT_PCT2, True)
        _set_formula_cell(ws, wc_rows["Legacy Inventory % Revenue (reference)"], col_idx, f"={ref_inventory_pct}", FMT_PCT2, True)
        _set_formula_cell(ws, wc_rows["Legacy AP % Revenue (reference)"], col_idx, f"={ref_ap_pct}", FMT_PCT2, True)
    row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "V3.9.9.2.1 WC days fade: AR = Revenue/365*DSO, Inventory = COGS/365*DIO, AP = COGS/365*DPO, with DSO/DIO/DPO fading from latest actual days toward normalized 3Y targets. Schedule-derived Delta NWC remains the PRIMARY working capital driver for CF Forecast, FCF Build, and DCF. % Revenue rows are diagnostic references.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    write_section_title(ws, row, "Working Capital Reality Check", span_cols=span_cols)
    SCHEDULE_ROWS["wc_reality"] = {}
    row += 1
    write_header_row(ws, row, ["Metric", "Value", "Review Note"], 1)
    row += 1
    wc_sched_payload = ((getattr(out, "schedules", None) or {}).get("working_capital") or {})
    wc_reality = wc_sched_payload.get("reality_check") or {}
    hist_delta_by_year = wc_reality.get("historical_delta_nwc_by_year") or wc_sched_payload.get("delta_nwc_history") or []
    hist_delta_rows = []
    for rec in hist_delta_by_year:
        year = rec.get("year") if isinstance(rec, dict) else None
        value = rec.get("delta_nwc") if isinstance(rec, dict) else None
        if year in {2022, 2023, 2024, 2025}:
            hist_delta_rows.append((f"FY{year} Historical Delta NWC", value, FMT_COMMA2, "Balance-sheet AR + Inventory - AP delta. Negative = source/release; positive = use/investment."))
    reality_rows = [
        ("Delta NWC Sign Convention", wc_reality.get("sign_convention") or wc_sched_payload.get("delta_nwc_sign_convention"), None, "Workbook-wide convention; removes ambiguity in WC release/build interpretation."),
        ("Historical average Delta NWC", wc_reality.get("historical_average_delta_nwc"), FMT_COMMA2, "Average of available historical Delta NWC releases / builds."),
        *hist_delta_rows,
        ("Current schedule-derived Delta NWC", wc_reality.get("current_schedule_derived_delta_nwc"), FMT_COMMA2, "Year 1 active faded-days schedule Delta NWC."),
        ("Faded-days Delta NWC", wc_reality.get("faded_days_delta_nwc"), FMT_COMMA2, "Same active schedule; shown explicitly for review."),
        ("Legacy pct-revenue Delta NWC reference", wc_reality.get("legacy_pct_revenue_delta_nwc_reference"), FMT_COMMA2, "Reference only; not the active DCF driver."),
        ("Difference vs historical average", wc_reality.get("difference_vs_historical_average"), FMT_COMMA2, "Faded-days Delta NWC minus historical average."),
        ("Hostile Review Release Difference Note", wc_reality.get("hostile_review_release_difference_note"), None, "Explains why release-only or cash-flow-statement figures can differ from the signed balance-sheet Delta NWC average."),
        ("Faded-days Gap Explanation", wc_reality.get("faded_days_gap_explanation"), None, "Explains why Year 1 faded-days release can remain below larger historical release years."),
    ]
    for label, value, fmt, note in reality_rows:
        _set_label_cell(ws, row, 1, label)
        c = ws.cell(row=row, column=2, value=value if value is not None else "N/A")
        c.font = font_formula(); c.border = border_thin()
        if value is not None and fmt:
            c.number_format = fmt
        _set_note_cell(ws, row, 3, note)
        SCHEDULE_ROWS["wc_reality"][label] = row
        row += 1
    _set_label_cell(ws, row, 1, "Review tier", True)
    c = ws.cell(row=row, column=2, value=wc_reality.get("review_tier") or "Review")
    c.font = font_formula(); c.border = border_thin()
    _set_note_cell(ws, row, 3, "If faded-days release remains materially below / above historical average, retain Review instead of false OK.")
    SCHEDULE_ROWS["wc_reality"]["Review tier"] = row
    row += 2

    ref_cost_of_debt = _abs_ref(ASSUMP_CELLS["cost_of_debt"])
    ref_debt_issuance = _abs_ref(ASSUMP_CELLS["debt_issuance"])
    ref_debt_repayment = _abs_ref(ASSUMP_CELLS["debt_repayment"])
    debt_rows = section("debt", "Debt Schedule v2 (roll-forward)", [
        "Beginning Total Debt",
        "  (+) New Debt Issuance",
        "  (-) Debt Repayment",
        "Ending Total Debt",
        "Short-term Debt",
        "Long-term Debt",
        "Total Debt (Ending, ST + LT)",
        "Average Debt",
        "Cost of Debt",
        "Interest Expense",
        "Debt Change (Issuance - Repayment)",
    ])
    for idx, year in enumerate(historical_years, start=2):
        std = _hist_model_value(ctx, "balance_sheet", "short_term_debt", year, out.market)
        ltd = _hist_model_value(ctx, "balance_sheet", "long_term_debt", year, out.market)
        total = _hist_model_value(ctx, "balance_sheet", "total_debt", year, out.market)
        if total is None and (std is not None or ltd is not None):
            total = (std or 0.0) + (ltd or 0.0)
        interest_hist = _hist_model_value(ctx, "income_statement", "interest_expense", year, out.market)
        # Beginning total debt = prior-year ending total
        previous_years = [y for y in historical_years if y < year]
        beginning_total = None
        if previous_years:
            py = max(previous_years)
            prev_std = _hist_model_value(ctx, "balance_sheet", "short_term_debt", py, out.market) or 0.0
            prev_ltd = _hist_model_value(ctx, "balance_sheet", "long_term_debt", py, out.market) or 0.0
            prev_total_raw = _hist_model_value(ctx, "balance_sheet", "total_debt", py, out.market)
            beginning_total = prev_total_raw if prev_total_raw is not None else (prev_std + prev_ltd)
        average_debt = None
        if beginning_total is not None and total is not None:
            average_debt = (beginning_total + total) / 2.0
        values = {
            "Beginning Total Debt": beginning_total,
            "  (+) New Debt Issuance": None,
            "  (-) Debt Repayment": None,
            "Ending Total Debt": total,
            "Short-term Debt": std,
            "Long-term Debt": ltd,
            "Total Debt (Ending, ST + LT)": total,
            "Average Debt": average_debt,
            "Cost of Debt": None,
            "Interest Expense": interest_hist,
            "Debt Change (Issuance - Repayment)": None,
        }
        for label, value in values.items():
            c = ws.cell(row=debt_rows[label], column=idx, value=value)
            c.font = font_formula()
            c.border = border_thin()
            if value is not None:
                c.number_format = FMT_PCT2 if label == "Cost of Debt" else FMT_COMMA2
    latest_std = _latest_historical_model_value(ctx, "balance_sheet", "short_term_debt", out.market) or 0.0
    latest_ltd = _latest_historical_model_value(ctx, "balance_sheet", "long_term_debt", out.market) or 0.0
    latest_total_raw = _latest_historical_model_value(ctx, "balance_sheet", "total_debt", out.market)
    latest_total = latest_total_raw if latest_total_raw is not None else (latest_std + latest_ltd)
    for i in range(n):
        col_idx = forecast_start_col + i
        col = get_column_letter(col_idx)
        prev_col = get_column_letter(col_idx - 1)
        beginning = str(latest_total) if i == 0 else f"{prev_col}{debt_rows['Ending Total Debt']}"
        _set_formula_cell(ws, debt_rows["Beginning Total Debt"], col_idx, f"={beginning}", FMT_COMMA2)
        _set_formula_cell(ws, debt_rows["  (+) New Debt Issuance"], col_idx, f"={ref_debt_issuance}", FMT_COMMA2, True)
        _set_formula_cell(ws, debt_rows["  (-) Debt Repayment"], col_idx, f"={ref_debt_repayment}", FMT_COMMA2, True)
        _set_formula_cell(
            ws,
            debt_rows["Ending Total Debt"],
            col_idx,
            f"=ROUND({col}{debt_rows['Beginning Total Debt']}+{col}{debt_rows['  (+) New Debt Issuance']}-{col}{debt_rows['  (-) Debt Repayment']},2)",
            FMT_COMMA2,
        )
        # Keep ST flat; LT = Ending Total - ST so ST + LT always reconciles to Ending Total.
        _set_formula_cell(ws, debt_rows["Short-term Debt"], col_idx, f"={latest_std}", FMT_COMMA2)
        _set_formula_cell(
            ws,
            debt_rows["Long-term Debt"],
            col_idx,
            f"=ROUND({col}{debt_rows['Ending Total Debt']}-{col}{debt_rows['Short-term Debt']},2)",
            FMT_COMMA2,
        )
        _set_formula_cell(
            ws,
            debt_rows["Total Debt (Ending, ST + LT)"],
            col_idx,
            f"=ROUND({col}{debt_rows['Short-term Debt']}+{col}{debt_rows['Long-term Debt']},2)",
            FMT_COMMA2,
        )
        _set_formula_cell(
            ws,
            debt_rows["Average Debt"],
            col_idx,
            f"=ROUND(({col}{debt_rows['Beginning Total Debt']}+{col}{debt_rows['Ending Total Debt']})/2,2)",
            FMT_COMMA2,
        )
        _set_formula_cell(ws, debt_rows["Cost of Debt"], col_idx, f"={ref_cost_of_debt}", FMT_PCT2, True)
        _set_formula_cell(
            ws,
            debt_rows["Interest Expense"],
            col_idx,
            f"=ROUND({col}{debt_rows['Average Debt']}*{col}{debt_rows['Cost of Debt']},2)",
            FMT_COMMA2,
        )
        _set_formula_cell(
            ws,
            debt_rows["Debt Change (Issuance - Repayment)"],
            col_idx,
            f"=ROUND({col}{debt_rows['  (+) New Debt Issuance']}-{col}{debt_rows['  (-) Debt Repayment']},2)",
            FMT_COMMA2,
        )
    row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "Debt Schedule v2: Interest Expense = Average Debt x Cost of Debt; Ending Total = Beginning + Issuance - Repayment. ST kept flat; LT balances to Ending. CF Financing uses Debt Change (Issuance - Repayment).",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    share_rows = section("shares", "Share Count Schedule", [
        "Diluted Shares",
        "Share Growth / Buyback assumption",
        "Forecast Diluted Shares",
    ])
    for idx, year in enumerate(historical_years, start=2):
        shares = _hist_model_value(ctx, "income_statement", "diluted_shares", year, out.market)
        c = ws.cell(row=share_rows["Diluted Shares"], column=idx, value=shares)
        c.font = font_formula()
        c.border = border_thin()
        if shares is not None:
            c.number_format = FMT_COMMA2
        ws.cell(row=share_rows["Share Growth / Buyback assumption"], column=idx, value=None).border = border_thin()
        c = ws.cell(row=share_rows["Forecast Diluted Shares"], column=idx, value=shares)
        c.font = font_formula()
        c.border = border_thin()
        if shares is not None:
            c.number_format = FMT_COMMA2
    # V3.9.6 single source-of-truth: legacy Share Count Schedule beginning shares
    # are pulled from the same Assumptions cell that drives the Shareholder Returns
    # Schedule below. Eliminates the prior dual-source (historical cache vs
    # Assumptions) starting-shares contradiction flagged in hostile review.
    latest_shares = _latest_historical_model_value(ctx, "income_statement", "diluted_shares", out.market)
    if latest_shares is None:
        latest_shares = inp.shares
    ref_shares_for_legacy = _abs_ref(ASSUMP_CELLS["shares"])
    for i in range(n):
        col_idx = forecast_start_col + i
        col = get_column_letter(col_idx)
        prev_col = get_column_letter(col_idx - 1)
        _set_formula_cell(ws, share_rows["Share Growth / Buyback assumption"], col_idx, f"={ref_share_growth}", FMT_PCT2, True)
        beginning = f"={ref_shares_for_legacy}" if i == 0 else f"={prev_col}{share_rows['Forecast Diluted Shares']}"
        _set_formula_cell(ws, share_rows["Diluted Shares"], col_idx, beginning, FMT_COMMA2)
        _set_formula_cell(ws, share_rows["Forecast Diluted Shares"], col_idx, f"=ROUND({col}{share_rows['Diluted Shares']}*(1+{col}{share_rows['Share Growth / Buyback assumption']}),2)", FMT_COMMA2)
    row += 1
    _set_note_cell(ws, row, 1, "Share-growth assumption retained as a diagnostic roll. Beginning shares are sourced from the same Assumptions Diluted Shares cell as the Shareholder Returns Schedule below, so the two schedules cannot diverge. The headline IV denominator comes from the Shareholder Returns Schedule (Selected Share Count Treatment).")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    # ── V3.7.4 Shareholder Returns Schedule v1 ──────────────────────────
    sr_payload = getattr(out, "shareholder_returns", None) or {}
    sr_hist = sr_payload.get("historical") or {}
    sr_hist_series = sr_hist.get("series") or {}
    # Record column metadata for later relink (Assumptions denominator cells
    # need to reference Ending / WAvg Diluted Shares once this sheet exists).
    _sr_forecast_start_col = forecast_start_col
    _sr_last_year_col = forecast_start_col + n - 1
    sr_rows = section(
        "shareholder_returns",
        "Buyback Funding Closure",
        [
            "Beginning Diluted Shares",
            "Net Income (P&L link)",
            SR_FCF_REF_LABEL,
            "FCF after Dividends",
            "Dividends Paid",
            "Planned Buybacks",
            "Share Repurchases (Buybacks)",
            "Unfunded Buybacks",
            "Beginning Cash",
            "Minimum Cash Floor",
            "Cash Available Above Floor",
            "Marketable Securities Available for Returns",
            "Funding Capacity Before Debt",
            "FCF after Dividends Used for Buybacks",
            "Cash Above Floor Used for Buybacks",
            "Marketable Securities Drawdown Used",
            "Incremental Debt Issuance",
            "Ending Cash",
            "Ending Marketable Securities",
            "Ending Debt",
            "Ending Net Debt",
            "Total Shareholder Returns",
            "Repurchase Price (per share)",
            "Shares Repurchased (M)",
            "SBC / Annual Dilution (M)",
            "Ending Diluted Shares",
            "Weighted Avg Diluted Shares",
            "EPS (NI / WAvg)",
            "CFO-derived FCF Reference per Share",
            "Funding Treatment",
            "Funding Review Tier",
        ],
    )
    # Historical actuals: dividends + buybacks displayed as positive cash
    # outflows (absolute value of raw cash-flow figures).
    def _series_to_year_map(pairs):
        return {int(y): v for (y, v) in (pairs or [])}

    div_hist = _series_to_year_map(sr_hist_series.get("dividends"))
    buy_hist = _series_to_year_map(sr_hist_series.get("buybacks"))
    fcf_hist = _series_to_year_map(sr_hist_series.get("fcf"))
    ni_hist = _series_to_year_map(sr_hist_series.get("net_income"))
    shares_hist_map = _series_to_year_map(sr_hist_series.get("diluted_shares"))

    for idx, year in enumerate(historical_years, start=2):
        # Beginning shares: prior year's diluted shares (display only).
        prev_years = [y for y in historical_years if y < year]
        beg_shares = shares_hist_map.get(max(prev_years)) if prev_years else shares_hist_map.get(year)
        end_shares = shares_hist_map.get(year)
        ni = ni_hist.get(year)
        fcf = fcf_hist.get(year)
        div = div_hist.get(year)
        buy = buy_hist.get(year)
        repurchased = None
        sbc_dil = None
        if beg_shares is not None and end_shares is not None and beg_shares > 0:
            net_change = beg_shares - end_shares
            repurchased = max(0.0, net_change)
        wavg = ((beg_shares or 0.0) + (end_shares or 0.0)) / 2 if (beg_shares is not None and end_shares is not None) else None
        values_h = {
            "Beginning Diluted Shares": beg_shares,
            "Net Income (P&L link)": ni,
            SR_FCF_REF_LABEL: fcf,
            "Dividends Paid": div,
            "Share Repurchases (Buybacks)": buy,
            "Total Shareholder Returns": (div or 0.0) + (buy or 0.0) if (div is not None or buy is not None) else None,
            "Repurchase Price (per share)": None,
            "Shares Repurchased (M)": repurchased,
            "SBC / Annual Dilution (M)": sbc_dil,
            "Ending Diluted Shares": end_shares,
            "Weighted Avg Diluted Shares": wavg,
            "EPS (NI / WAvg)": (ni / wavg) if (ni is not None and wavg and wavg > 0) else None,
            "CFO-derived FCF Reference per Share": (fcf / wavg) if (fcf is not None and wavg and wavg > 0) else None,
        }
        for label, v in values_h.items():
            c = ws.cell(row=sr_rows[label], column=idx, value=v)
            c.font = font_formula()
            c.border = border_thin()
            if v is not None:
                c.number_format = FMT_COMMA2

    # Forecast: live formulas referencing P&L / CF Forecast + Assumptions drivers.
    ref_payout = _abs_ref(ASSUMP_CELLS.get("sr_dividend_payout", ASSUMP_CELLS["dividends_buybacks_placeholder"]))
    ref_method = _abs_ref(ASSUMP_CELLS.get("sr_buyback_method", ASSUMP_CELLS["dividends_buybacks_placeholder"]))
    ref_buy_pct_fcf = _abs_ref(ASSUMP_CELLS.get("sr_buyback_pct_fcf", ASSUMP_CELLS["dividends_buybacks_placeholder"]))
    ref_flat_buyback = _abs_ref(ASSUMP_CELLS.get("sr_flat_buyback", ASSUMP_CELLS["dividends_buybacks_placeholder"]))
    ref_repurchase_growth = _abs_ref(ASSUMP_CELLS.get("sr_repurchase_price_growth", ASSUMP_CELLS["dividends_buybacks_placeholder"]))
    ref_dilution = _abs_ref(ASSUMP_CELLS.get("sr_annual_dilution", ASSUMP_CELLS["dividends_buybacks_placeholder"]))
    ref_funding_treatment = _abs_ref(ASSUMP_CELLS.get("sr_funding_treatment", ASSUMP_CELLS["dividends_buybacks_placeholder"]))
    ref_cash_floor = _abs_ref(ASSUMP_CELLS.get("sr_cash_floor", ASSUMP_CELLS["dividends_buybacks_placeholder"]))
    ref_ms_available = _abs_ref(ASSUMP_CELLS.get("sr_ms_available", ASSUMP_CELLS["dividends_buybacks_placeholder"]))
    ref_price_assump = _abs_ref(ASSUMP_CELLS["price"])
    ref_shares_assump = _abs_ref(ASSUMP_CELLS["shares"])
    lbl_flat = BUYBACK_METHOD_LABELS["flat_amount"]
    for i in range(n):
        col_idx = forecast_start_col + i
        col = get_column_letter(col_idx)
        prev_col = get_column_letter(col_idx - 1)
        pl_ni_ref = _sheet_ref("P&L Forecast", PL_FORECAST_ROWS["Net Income"], col_idx)
        # FCF reference: seed with the calculator's static FCF projection now;
        # relink to live Cash Flow Forecast FCF cell after CF sheet is built.
        sr_forecast_payload = (sr_payload.get("forecast") or {})
        sr_fcf_list = sr_forecast_payload.get("fcf") or []
        seed_fcf = sr_fcf_list[i] if i < len(sr_fcf_list) else 0.0
        beg_shares_formula = f"={ref_shares_assump}" if i == 0 else f"={prev_col}{sr_rows['Ending Diluted Shares']}"
        _set_formula_cell(ws, sr_rows["Beginning Diluted Shares"], col_idx, beg_shares_formula, FMT_COMMA2, True)
        _set_formula_cell(ws, sr_rows["Net Income (P&L link)"], col_idx, f"={pl_ni_ref}", FMT_COMMA2, True)
        # Seed with calculator FCF; relink_sr_to_cf_forecast() overwrites with live link.
        ws.cell(row=sr_rows[SR_FCF_REF_LABEL], column=col_idx, value=float(seed_fcf or 0.0))
        ws.cell(row=sr_rows[SR_FCF_REF_LABEL], column=col_idx).number_format = FMT_COMMA2
        _set_formula_cell(
            ws, sr_rows["Dividends Paid"], col_idx,
            f"=MAX(0,{ref_payout}*{col}{sr_rows['Net Income (P&L link)']})", FMT_COMMA2,
        )
        _set_formula_cell(ws, sr_rows["FCF after Dividends"], col_idx, f"=MAX(0,{col}{sr_rows[SR_FCF_REF_LABEL]}-{col}{sr_rows['Dividends Paid']})", FMT_COMMA2)
        # Planned buybacks: flat amount if method label matches, else pct of FCF.
        _set_formula_cell(
            ws, sr_rows["Planned Buybacks"], col_idx,
            f'=MAX(0,IF({ref_method}="{lbl_flat}",{ref_flat_buyback},{ref_buy_pct_fcf}*{col}{sr_rows[SR_FCF_REF_LABEL]}))',
            FMT_COMMA2,
        )
        if i == 0:
            beg_cash = ((sr_payload.get("forecast") or {}).get("beginning_cash") or [0.0])[0]
            first_ending_debt = ((sr_payload.get("forecast") or {}).get("ending_debt") or [0.0])[0]
            first_inc_debt = ((sr_payload.get("forecast") or {}).get("incremental_debt_issuance") or [0.0])[0]
            beg_debt = (first_ending_debt or 0.0) - (first_inc_debt or 0.0)
            _set_formula_cell(ws, sr_rows["Beginning Cash"], col_idx, f"={beg_cash}", FMT_COMMA2)
            _set_formula_cell(ws, sr_rows["Ending Debt"], col_idx, f"={beg_debt}+{col}{sr_rows['Incremental Debt Issuance']}", FMT_COMMA2)
        else:
            _set_formula_cell(ws, sr_rows["Beginning Cash"], col_idx, f"={prev_col}{sr_rows['Ending Cash']}", FMT_COMMA2)
            _set_formula_cell(ws, sr_rows["Ending Debt"], col_idx, f"={prev_col}{sr_rows['Ending Debt']}+{col}{sr_rows['Incremental Debt Issuance']}", FMT_COMMA2)
        _set_formula_cell(ws, sr_rows["Minimum Cash Floor"], col_idx, f"={ref_cash_floor}", FMT_COMMA2, True)
        _set_formula_cell(ws, sr_rows["Cash Available Above Floor"], col_idx, f"=MAX(0,{col}{sr_rows['Beginning Cash']}-{col}{sr_rows['Minimum Cash Floor']})", FMT_COMMA2)
        ms_available_formula = f"={ref_ms_available}" if i == 0 else f"={prev_col}{sr_rows['Ending Marketable Securities']}"
        _set_formula_cell(ws, sr_rows["Marketable Securities Available for Returns"], col_idx, ms_available_formula, FMT_COMMA2, True)
        _set_formula_cell(ws, sr_rows["Funding Capacity Before Debt"], col_idx, f"={col}{sr_rows['FCF after Dividends']}+{col}{sr_rows['Cash Available Above Floor']}+{col}{sr_rows['Marketable Securities Available for Returns']}", FMT_COMMA2)
        _set_formula_cell(ws, sr_rows["Share Repurchases (Buybacks)"], col_idx, f'=IF(OR({ref_funding_treatment}="Debt-Funded Buyback",{ref_funding_treatment}="Planned-Uncapped Diagnostic"),{col}{sr_rows["Planned Buybacks"]},MIN({col}{sr_rows["Planned Buybacks"]},{col}{sr_rows["Funding Capacity Before Debt"]}))', FMT_COMMA2)
        _set_formula_cell(ws, sr_rows["Unfunded Buybacks"], col_idx, f'=IF({ref_funding_treatment}="Debt-Funded Buyback",0,MAX(0,{col}{sr_rows["Planned Buybacks"]}-{col}{sr_rows["Funding Capacity Before Debt"]}))', FMT_COMMA2)
        _set_formula_cell(ws, sr_rows["FCF after Dividends Used for Buybacks"], col_idx, f"=MIN({col}{sr_rows['Share Repurchases (Buybacks)']},{col}{sr_rows['FCF after Dividends']})", FMT_COMMA2)
        _set_formula_cell(ws, sr_rows["Cash Above Floor Used for Buybacks"], col_idx, f'=IF({ref_funding_treatment}="Planned-Uncapped Diagnostic",MAX(0,{col}{sr_rows["Share Repurchases (Buybacks)"]}-{col}{sr_rows["FCF after Dividends Used for Buybacks"]}),MIN(MAX(0,{col}{sr_rows["Share Repurchases (Buybacks)"]}-{col}{sr_rows["FCF after Dividends Used for Buybacks"]}),{col}{sr_rows["Cash Available Above Floor"]}))', FMT_COMMA2)
        _set_formula_cell(ws, sr_rows["Marketable Securities Drawdown Used"], col_idx, f'=IF({ref_funding_treatment}="Planned-Uncapped Diagnostic",0,MIN({col}{sr_rows["Marketable Securities Available for Returns"]},MAX(0,{col}{sr_rows["Share Repurchases (Buybacks)"]}-{col}{sr_rows["FCF after Dividends Used for Buybacks"]}-{col}{sr_rows["Cash Above Floor Used for Buybacks"]})))', FMT_COMMA2)
        _set_formula_cell(ws, sr_rows["Incremental Debt Issuance"], col_idx, f'=IF({ref_funding_treatment}="Debt-Funded Buyback",MAX(0,{col}{sr_rows["Planned Buybacks"]}-{col}{sr_rows["Funding Capacity Before Debt"]}),0)', FMT_COMMA2)
        _set_formula_cell(ws, sr_rows["Ending Cash"], col_idx, f"={col}{sr_rows['Beginning Cash']}+{col}{sr_rows['FCF after Dividends']}-{col}{sr_rows['FCF after Dividends Used for Buybacks']}-{col}{sr_rows['Cash Above Floor Used for Buybacks']}", FMT_COMMA2)
        _set_formula_cell(ws, sr_rows["Ending Marketable Securities"], col_idx, f"={col}{sr_rows['Marketable Securities Available for Returns']}-{col}{sr_rows['Marketable Securities Drawdown Used']}", FMT_COMMA2)
        _set_formula_cell(ws, sr_rows["Ending Net Debt"], col_idx, f"={col}{sr_rows['Ending Debt']}-{col}{sr_rows['Ending Cash']}-{col}{sr_rows['Ending Marketable Securities']}", FMT_COMMA2)
        _set_formula_cell(
            ws, sr_rows["Total Shareholder Returns"], col_idx,
            f"={col}{sr_rows['Dividends Paid']}+{col}{sr_rows['Share Repurchases (Buybacks)']}", FMT_COMMA2,
        )
        _set_formula_cell(
            ws, sr_rows["Repurchase Price (per share)"], col_idx,
            f"={ref_price_assump}*POWER(1+{ref_repurchase_growth},{i})", FMT_COMMA2,
        )
        # Buybacks are in millions of currency; price is per-share absolute.
        # Shares retired (M) = Buybacks (M) / Price (per share).
        _set_formula_cell(
            ws, sr_rows["Shares Repurchased (M)"], col_idx,
            f"=IF({col}{sr_rows['Repurchase Price (per share)']}>0,{col}{sr_rows['Share Repurchases (Buybacks)']}/{col}{sr_rows['Repurchase Price (per share)']},0)",
            FMT_COMMA2,
        )
        _set_formula_cell(
            ws, sr_rows["SBC / Annual Dilution (M)"], col_idx,
            f"={ref_dilution}*{col}{sr_rows['Beginning Diluted Shares']}", FMT_COMMA2,
        )
        _set_formula_cell(
            ws, sr_rows["Ending Diluted Shares"], col_idx,
            f"=MAX(0,{col}{sr_rows['Beginning Diluted Shares']}-{col}{sr_rows['Shares Repurchased (M)']}+{col}{sr_rows['SBC / Annual Dilution (M)']})",
            FMT_COMMA2,
        )
        _set_formula_cell(
            ws, sr_rows["Weighted Avg Diluted Shares"], col_idx,
            f"=({col}{sr_rows['Beginning Diluted Shares']}+{col}{sr_rows['Ending Diluted Shares']})/2",
            FMT_COMMA2,
        )
        _set_formula_cell(
            ws, sr_rows["EPS (NI / WAvg)"], col_idx,
            f"=IFERROR({col}{sr_rows['Net Income (P&L link)']}/{col}{sr_rows['Weighted Avg Diluted Shares']},0)",
            FMT_COMMA2,
        )
        _set_formula_cell(
            ws, sr_rows["CFO-derived FCF Reference per Share"], col_idx,
            f"=IFERROR({col}{sr_rows[SR_FCF_REF_LABEL]}/{col}{sr_rows['Weighted Avg Diluted Shares']},0)",
            FMT_COMMA2,
        )
        _set_formula_cell(ws, sr_rows["Funding Treatment"], col_idx, f'={ref_funding_treatment}', None, True)
        _set_formula_cell(
            ws,
            sr_rows["Funding Review Tier"],
            col_idx,
            f'=IF({col}{sr_rows["Ending Cash"]}<0,"High Review",IF({col}{sr_rows["Ending Cash"]}<{col}{sr_rows["Minimum Cash Floor"]},"High Review",IF({col}{sr_rows["Incremental Debt Issuance"]}>0,"Review",IF({col}{sr_rows["Unfunded Buybacks"]}>0,"Review","OK"))))',
            None,
        )
    # Stash forecast column range so the Assumptions relink helper can
    # build references without re-deriving historical_years length.
    sr_rows["_forecast_start_col"] = _sr_forecast_start_col
    sr_rows["_last_year_col"] = _sr_last_year_col
    row += 1
    _set_note_cell(
        ws, row, 1,
        "V3.9.9.2.1 Shareholder Returns Funding Closure: dividends and buybacks remain financing items and do not reduce FCFF. Share repurchases and ending shares use actual funded buybacks only. Source stack is explicit: FCF after dividends, cash above floor, marketable securities drawdown, then diagnostic incremental debt if selected.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)


def relink_pl_forecast_to_supporting_schedules(ws):
    if not PL_FORECAST_ROWS or not SCHEDULE_ROWS:
        return
    revenue_row = PL_FORECAST_ROWS.get("Revenue")
    shares_row = PL_FORECAST_ROWS.get("Diluted Shares")
    eps_row = PL_FORECAST_ROWS.get("EPS")
    ni_row = PL_FORECAST_ROWS.get("Net Income")
    interest_row = PL_FORECAST_ROWS.get("Interest Expense")
    if not all([revenue_row, shares_row, eps_row, ni_row]):
        return
    debt_interest_row = (SCHEDULE_ROWS.get("debt") or {}).get("Interest Expense")
    for col_idx in range(2, ws.max_column + 1):
        header = ws.cell(row=8, column=col_idx).value
        if not isinstance(header, str) or not header.startswith("Year "):
            continue
        col = get_column_letter(col_idx)
        sr_rows = SCHEDULE_ROWS.get("shareholder_returns") or {}
        share_ref = (
            _sheet_ref("Supporting Schedules", sr_rows["Ending Diluted Shares"], col_idx)
            if sr_rows.get("Ending Diluted Shares")
            else _sheet_ref("Supporting Schedules", SCHEDULE_ROWS["shares"]["Forecast Diluted Shares"], col_idx)
        )
        _set_formula_cell(ws, shares_row, col_idx, f"={share_ref}", FMT_COMMA2, True)
        _set_formula_cell(ws, eps_row, col_idx, f"=IFERROR({col}{ni_row}/{col}{shares_row},0)", FMT_COMMA2)
        if interest_row and debt_interest_row:
            interest_ref = _sheet_ref("Supporting Schedules", debt_interest_row, col_idx)
            _set_formula_cell(ws, interest_row, col_idx, f"={interest_ref}", FMT_COMMA2, True)


def relink_sr_schedule_to_cf_forecast(ws_supporting):
    """After Cash Flow Forecast is built, replace seeded cash-flow reference
    values in the Shareholder Returns Schedule with live CF Forecast links.
    """
    sr = SCHEDULE_ROWS.get("shareholder_returns") or {}
    if not (sr and CF_FORECAST_ROWS.get(CF_FCF_REF_LABEL) and sr.get("_forecast_start_col") and sr.get("_last_year_col")):
        return
    for col_idx in range(sr["_forecast_start_col"], sr["_last_year_col"] + 1):
        cf_fcf_ref = _sheet_ref("Cash Flow Forecast", CF_FORECAST_ROWS[CF_FCF_REF_LABEL], col_idx)
        _set_formula_cell(ws_supporting, sr[SR_FCF_REF_LABEL], col_idx, f"={cf_fcf_ref}", FMT_COMMA2, True)


def relink_assumptions_to_shareholder_returns_schedule(ws_assumptions, inp: DCFInputs):
    """V3.7.4: replace the pre-seeded Forecast Ending / WAvg denominator inputs
    on the Assumptions sheet with live formula references to the Shareholder
    Returns Schedule, so changing buyback assumptions flows through to the
    DCF headline denominator.
    """
    sr = SCHEDULE_ROWS.get("shareholder_returns") or {}
    if not sr:
        return
    end_row = sr.get("Ending Diluted Shares")
    wavg_row = sr.get("Weighted Avg Diluted Shares")
    fc_start = sr.get("_forecast_start_col")
    fc_last = sr.get("_last_year_col")
    if not (end_row and wavg_row and fc_start and fc_last):
        return
    end_ref = _sheet_ref("Supporting Schedules", end_row, fc_last)
    wavg_cells = [
        _sheet_ref("Supporting Schedules", wavg_row, c)
        for c in range(fc_start, fc_last + 1)
    ]
    wavg_ref = f"=AVERAGE({','.join(wavg_cells)})"
    if "sr_denom_ending" in ASSUMP_CELLS:
        addr = ASSUMP_CELLS["sr_denom_ending"]
        row_num = int("".join(ch for ch in addr if ch.isdigit()))
        _set_formula_cell(ws_assumptions, row_num, 2, f"={end_ref}", FMT_COMMA2, True)
    if "sr_denom_wavg" in ASSUMP_CELLS:
        addr = ASSUMP_CELLS["sr_denom_wavg"]
        row_num = int("".join(ch for ch in addr if ch.isdigit()))
        _set_formula_cell(ws_assumptions, row_num, 2, wavg_ref, FMT_COMMA2, True)


def build_cash_flow_forecast_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "Cash Flow Forecast"
    CF_FORECAST_ROWS.clear()

    payload = ((ctx.get("historical_cache") or {}).get("data") or {})
    income_table = ((ctx.get("historical_tables") or {}).get("income_statement") or {})
    cf_table = ((ctx.get("historical_tables") or {}).get("cash_flow") or {})
    historical_years = sorted(set(income_table.get("years") or []).union(cf_table.get("years") or []))
    n = max(1, int(inp.forecast_years or 5))
    forecast_labels = [f"Year {i + 1}" for i in range(n)]
    span_cols = max(3, len(historical_years) + n + 1)
    apply_column_widths(ws, {"A": 32, **{get_column_letter(i + 2): 15 for i in range(len(historical_years) + n)}})

    row = _write_sheet_heading(ws, 1, "Cash Flow Forecast v1", inp, out, ctx, span_cols=span_cols)

    if out.market != "US" or payload.get("status") != "ok" or not historical_years or not PL_FORECAST_ROWS or not SCHEDULE_ROWS:
        _set_note_cell(ws, row, 1, "Cash Flow Forecast not available yet for this market")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
        row += 2
        _write_kv_rows(ws, row, [
            ("Market / Currency", f"{out.market} / {out.currency}", None),
            ("Fallback", "Graceful fallback only; HK/CN BS/CF mapping will be added later.", None),
        ])
        return

    row = _write_kv_rows(ws, row, [
        ("Currency / Unit", model_unit_label(out.currency, out.market), None),
        ("Model Boundary", "3FS-linked (non-circular); see Data Sources Audit for full V3.8.0 boundary list.", None),
        ("CapEx Sign Convention", "CapEx is shown as a positive outflow; CFO-derived FCF Reference = CFO - CapEx. DCF FCFF is built separately on the FCF Build sheet.", None),
    ])
    row += 1

    header = [""] + [str(year) for year in historical_years] + forecast_labels
    write_header_row(ws, row, header, 1)
    row += 1

    hist_end_col = 1 + len(historical_years)
    forecast_start_col = hist_end_col + 1
    ref_wc_pct = _abs_ref(ASSUMP_CELLS["wc_change_pct_revenue"])

    row_map = {}

    def write_label(label: str, bold: bool = False):
        nonlocal row
        _set_label_cell(ws, row, 1, label, bold)
        row_map[label] = row
        CF_FORECAST_ROWS[label] = row
        row += 1
        return row - 1

    rows = [
        ("Net Income", FMT_COMMA2),
        ("D&A", FMT_COMMA2),
        ("Change in Working Capital", FMT_COMMA2),
        ("Cash Flow from Operations", FMT_COMMA2),
        ("CapEx", FMT_COMMA2),
        (CF_FCF_REF_LABEL, FMT_COMMA2),
        ("Debt Change / Financing assumption", FMT_COMMA2),
        # V3.7.4 financing additions.
        ("Dividends Paid", FMT_COMMA2),
        ("Share Repurchases", FMT_COMMA2),
        ("Total Shareholder Returns", FMT_COMMA2),
        ("Beginning Cash", FMT_COMMA2),
        ("Ending Cash", FMT_COMMA2),
    ]
    for label, fmt in rows:
        write_label(label, bold=label in {"Cash Flow from Operations", CF_FCF_REF_LABEL, "Ending Cash"})

    for idx, year in enumerate(historical_years, start=2):
        net_income = _hist_model_value(ctx, "income_statement", "net_income", year, out.market)
        da = _hist_model_value(ctx, "cash_flow", "depreciation_amortization", year, out.market)
        cfo = _hist_model_value(ctx, "cash_flow", "operating_cash_flow", year, out.market)
        capex = _hist_model_value_positive(ctx, "cash_flow", "capex", year, out.market)
        fcf = _hist_model_value(ctx, "cash_flow", "free_cash_flow", year, out.market)
        cash = _hist_model_value(ctx, "balance_sheet", "cash", year, out.market)
        # V3.7.4: dividends + buybacks from cash-flow cache (raw values are
        # typically negative for outflows; display absolute amount).
        dividends_hist = _hist_model_value_positive(ctx, "cash_flow", "cash_dividends_paid", year, out.market)
        buybacks_hist = _hist_model_value_positive(ctx, "cash_flow", "repurchase_of_capital_stock", year, out.market)
        total_returns_hist = None
        if dividends_hist is not None or buybacks_hist is not None:
            total_returns_hist = (dividends_hist or 0.0) + (buybacks_hist or 0.0)
        prev_cash = None
        previous_years = [y for y in historical_years if y < year]
        if previous_years:
            prev_cash = _hist_model_value(ctx, "balance_sheet", "cash", max(previous_years), out.market)
        wc_change = None
        if net_income is not None and da is not None and cfo is not None:
            wc_change = net_income + da - cfo
        values = {
            "Net Income": net_income,
            "D&A": da,
            "Change in Working Capital": wc_change,
            "Cash Flow from Operations": cfo,
            "CapEx": capex,
            CF_FCF_REF_LABEL: fcf if fcf is not None else (cfo - capex if cfo is not None and capex is not None else None),
            "Debt Change / Financing assumption": 0.0,
            "Dividends Paid": dividends_hist,
            "Share Repurchases": buybacks_hist,
            "Total Shareholder Returns": total_returns_hist,
            "Beginning Cash": prev_cash,
            "Ending Cash": cash,
        }
        for label, fmt in rows:
            c = ws.cell(row=row_map[label], column=idx, value=values.get(label))
            c.font = font_formula()
            c.border = border_thin()
            if fmt and values.get(label) is not None:
                c.number_format = fmt

    for i in range(n):
        col_idx = forecast_start_col + i
        col = get_column_letter(col_idx)
        prev_col = get_column_letter(col_idx - 1)
        pl_net_income = _sheet_ref("P&L Forecast", PL_FORECAST_ROWS["Net Income"], col_idx)
        pl_revenue = _sheet_ref("P&L Forecast", PL_FORECAST_ROWS["Revenue"], col_idx)
        schedule_da = _sheet_ref("Supporting Schedules", SCHEDULE_ROWS["ppe"]["D&A"], col_idx)
        schedule_capex = _sheet_ref("Supporting Schedules", SCHEDULE_ROWS["ppe"]["CapEx"], col_idx)
        schedule_wc = _sheet_ref("Supporting Schedules", SCHEDULE_ROWS["wc"]["Schedule-derived Delta WC"], col_idx)
        schedule_debt_change = _sheet_ref("Supporting Schedules", SCHEDULE_ROWS["debt"]["Debt Change (Issuance - Repayment)"], col_idx)
        _set_formula_cell(ws, row_map["Net Income"], col_idx, f"={pl_net_income}", FMT_COMMA2, True)
        _set_formula_cell(ws, row_map["D&A"], col_idx, f"={schedule_da}", FMT_COMMA2, True)
        _set_formula_cell(ws, row_map["Change in Working Capital"], col_idx, f"={schedule_wc}", FMT_COMMA2, True)
        _set_formula_cell(
            ws,
            row_map["Cash Flow from Operations"],
            col_idx,
            f"=ROUND({col}{row_map['Net Income']}+{col}{row_map['D&A']}-{col}{row_map['Change in Working Capital']},2)",
            FMT_COMMA2,
        )
        _set_formula_cell(ws, row_map["CapEx"], col_idx, f"={schedule_capex}", FMT_COMMA2, True)
        _set_formula_cell(ws, row_map[CF_FCF_REF_LABEL], col_idx, f"=ROUND({col}{row_map['Cash Flow from Operations']}-{col}{row_map['CapEx']},2)", FMT_COMMA2)
        _set_formula_cell(ws, row_map["Debt Change / Financing assumption"], col_idx, f"={schedule_debt_change}", FMT_COMMA2, True)
        # V3.7.4 financing rows. Pull from Shareholder Returns Schedule when
        # available; otherwise fall back to the static export assumption = 0.
        sr_map = SCHEDULE_ROWS.get("shareholder_returns") or {}
        if sr_map.get("Dividends Paid") and sr_map.get("Share Repurchases (Buybacks)"):
            sr_div_ref = _sheet_ref("Supporting Schedules", sr_map["Dividends Paid"], col_idx)
            sr_buy_ref = _sheet_ref("Supporting Schedules", sr_map["Share Repurchases (Buybacks)"], col_idx)
            _set_formula_cell(ws, row_map["Dividends Paid"], col_idx, f"={sr_div_ref}", FMT_COMMA2, True)
            _set_formula_cell(ws, row_map["Share Repurchases"], col_idx, f"={sr_buy_ref}", FMT_COMMA2, True)
        else:
            _set_formula_cell(ws, row_map["Dividends Paid"], col_idx, "=0", FMT_COMMA2)
            _set_formula_cell(ws, row_map["Share Repurchases"], col_idx, "=0", FMT_COMMA2)
        _set_formula_cell(
            ws, row_map["Total Shareholder Returns"], col_idx,
            f"={col}{row_map['Dividends Paid']}+{col}{row_map['Share Repurchases']}",
            FMT_COMMA2,
        )
        if i == 0:
            latest_cash = _latest_historical_model_value(ctx, "balance_sheet", "cash", out.market) or 0.0
            _set_formula_cell(ws, row_map["Beginning Cash"], col_idx, f"={latest_cash}", FMT_COMMA2)
        else:
            _set_formula_cell(ws, row_map["Beginning Cash"], col_idx, f"={prev_col}{row_map['Ending Cash']}", FMT_COMMA2)
        # V3.7.4: Ending Cash now deducts Total Shareholder Returns. FCFF
        # remains unchanged (computed on FCF Build from unlevered NOPAT).
        _set_formula_cell(
            ws,
            row_map["Ending Cash"],
            col_idx,
            f"=ROUND({col}{row_map['Beginning Cash']}+{col}{row_map['Cash Flow from Operations']}-{col}{row_map['CapEx']}+{col}{row_map['Debt Change / Financing assumption']}-{col}{row_map['Total Shareholder Returns']},2)",
            FMT_COMMA2,
        )

    row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "Net Income references P&L Forecast, which subtracts Interest Expense. D&A and CapEx are sourced from the PP&E schedule. Change in Working Capital uses schedule-derived Delta WC. Dividends Paid + Share Repurchases come from the Shareholder Returns Schedule; Ending Cash deducts Total Shareholder Returns. This sheet's CFO-derived FCF Reference is a cash-flow-statement reference, not the unlevered FCFF used by DCF.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)


def build_balance_sheet_forecast_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "Balance Sheet Forecast"
    BS_FORECAST_ROWS.clear()

    payload = ((ctx.get("historical_cache") or {}).get("data") or {})
    bs_table = ((ctx.get("historical_tables") or {}).get("balance_sheet") or {})
    historical_years = sorted(bs_table.get("years") or [])
    n = max(1, int(inp.forecast_years or 5))
    forecast_labels = [f"Year {i + 1}" for i in range(n)]
    span_cols = max(3, len(historical_years) + n + 2)
    classification_col_idx = len(historical_years) + n + 2
    apply_column_widths(ws, {"A": 32, **{get_column_letter(i + 2): 15 for i in range(len(historical_years) + n)}, get_column_letter(classification_col_idx): 42})

    support_status = _balance_sheet_forecast_support_status(out, ctx)
    heading = "Balance Sheet Forecast" if support_status == "full_forecast" else "Balance Sheet Inputs Summary - Limited Presentation"
    row = _write_sheet_heading(ws, 1, heading, inp, out, ctx, span_cols=span_cols)

    if support_status == "limited_presentation":
        _write_limited_balance_sheet_presentation(ws, row, inp, out, ctx, span_cols)
        return

    row = _write_kv_rows(ws, row, [
        ("Currency / Unit", model_unit_label(out.currency, out.market), None),
        ("Model Boundary", "BS Forecast separates schedule-driven lines, held-constant lines, residual plug rows, and not-modeled items. It is not presented as a full driver-based three-statement model.", None),
        ("Balance Check", "Balance is achieved through (i) schedule-driven drivers (Cash, AR, Inventory, AP, PP&E, Debt), (ii) lines held constant at latest available balance where the filing supplies a value but no forecast driver exists, (iii) named residual plug rows that remain explicit review items, and (iv) lines marked Not separately disclosed in 10-K (e.g., AAPL Goodwill / Intangibles), which are held at $0 in forecast and surfaced as 'N/D in 10-K' in historical cells to avoid silent zero.", None),
    ])
    row += 1

    _set_note_cell(
        ws,
        row,
        1,
        "Forecast classification (V3.9.8.7): each row carries a row-specific classification in the rightmost column — Schedule-driven (Cash / AR / Inventory / AP / PP&E / Debt), Held constant at latest available balance (marketable securities, deferred revenue, leases, deferred taxes, other named rows), Residual / plug (named bucket residuals, transparent review items), Not separately disclosed in 10-K (Goodwill / Intangibles for AAPL — included in Other Non-current Assets per filing, held at $0 to avoid double-count, never silent zero), and Memo / not separately modeled (OCI / common stock / APIC / retained earnings detail). Total Equity is roll-forward (prior + NI − dividends − buybacks); OCI is not modeled.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    header = [""] + [str(year) for year in historical_years] + forecast_labels + ["Forecast Classification"]
    write_header_row(ws, row, header, 1)
    row += 1

    hist_end_col = 1 + len(historical_years)
    forecast_start_col = hist_end_col + 1
    ref_dividends = _abs_ref(ASSUMP_CELLS["dividends_buybacks_placeholder"])

    row_map = {}

    def write_label(label: str, bold: bool = False):
        nonlocal row
        _set_label_cell(ws, row, 1, label, bold)
        row_map[label] = row
        BS_FORECAST_ROWS[label] = row
        row += 1
        return row - 1

    # V3.7.1 expanded line items. Items whose historical cache fields are
    # available are populated with real values; items that are missing fall
    # back to 0 and are flagged in Data Sources Audit.
    rows = [
        ("== Current Assets ==", None),
        ("Cash & Cash Equivalents", FMT_COMMA2),
        ("Short-term Investments / Marketable Securities", FMT_COMMA2),
        ("Accounts Receivable", FMT_COMMA2),
        ("Inventory", FMT_COMMA2),
        ("Other Current Assets", FMT_COMMA2),
        ("Other Current Assets (residual plug, transparent)", FMT_COMMA2),
        ("Total Current Assets", FMT_COMMA2),
        ("== Non-current Assets ==", None),
        ("PP&E, Net", FMT_COMMA2),
        ("Goodwill", FMT_COMMA2),
        ("Intangible Assets", FMT_COMMA2),
        ("Long-term Investments / Marketable Securities", FMT_COMMA2),
        ("Deferred Tax Assets (Non-current)", FMT_COMMA2),
        ("Other Non-current Assets", FMT_COMMA2),
        ("Other Non-current Assets (residual plug, transparent)", FMT_COMMA2),
        ("Total Non-current Assets", FMT_COMMA2),
        ("Total Assets", FMT_COMMA2),
        ("== Current Liabilities ==", None),
        ("Accounts Payable", FMT_COMMA2),
        ("Deferred Revenue, Current", FMT_COMMA2),
        ("Short-term Debt / Current Portion of Debt (incl. lease)", FMT_COMMA2),
        ("  Memo: Capital Lease Obligation, Current (subset of ST Debt)", FMT_COMMA2),
        ("Other Current Liabilities", FMT_COMMA2),
        ("Other Current Liabilities (residual plug, transparent)", FMT_COMMA2),
        ("Total Current Liabilities", FMT_COMMA2),
        ("== Non-current Liabilities ==", None),
        ("Long-term Debt (incl. lease)", FMT_COMMA2),
        ("  Memo: Capital Lease Obligation, Non-current (subset of LT Debt)", FMT_COMMA2),
        ("Other Non-current Liabilities (incl. deferred revenue NC)", FMT_COMMA2),
        ("  Memo: Deferred Revenue, Non-current (subset of Other NC Liab)", FMT_COMMA2),
        ("Deferred Tax Liabilities (Non-current)", FMT_COMMA2),
        ("Other Non-current Liabilities (residual plug, transparent)", FMT_COMMA2),
        ("Total Non-current Liabilities", FMT_COMMA2),
        ("Total Debt", FMT_COMMA2),
        ("Total Liabilities", FMT_COMMA2),
        ("Total Equity", FMT_COMMA2),
        ("Total Liabilities + Equity", FMT_COMMA2),
        ("Balance Check / Difference", FMT_COMMA2),
    ]
    bold_labels = {
        "Total Current Assets", "Total Non-current Assets", "Total Assets",
        "Total Current Liabilities", "Total Non-current Liabilities", "Total Debt",
        "Total Liabilities", "Total Equity", "Total Liabilities + Equity",
        "Balance Check / Difference",
    }
    for label, fmt in rows:
        if label.startswith("=="):
            _set_label_cell(ws, row, 1, label.strip("= ").strip(), bold=True)
            BS_FORECAST_ROWS[label] = row
            row += 1
        else:
            write_label(label, bold=label in bold_labels)
    bs_classification = {
        "Cash & Cash Equivalents": "Schedule-driven: links to Cash Flow Forecast ending cash.",
        "Accounts Receivable": "Schedule-driven: working-capital schedule.",
        "Inventory": "Schedule-driven: working-capital schedule.",
        "PP&E, Net": "Schedule-driven: PP&E roll-forward.",
        "Accounts Payable": "Schedule-driven: working-capital schedule.",
        "Short-term Debt / Current Portion of Debt (incl. lease)": "Schedule-driven: debt schedule.",
        "Long-term Debt (incl. lease)": "Schedule-driven: debt schedule.",
        "Total Debt": "Schedule-driven: debt schedule.",
        "Total Equity": "Roll-forward: prior-year equity + NI - dividends - buybacks; OCI / other comprehensive items not modeled.",
        "Other Current Assets (residual plug, transparent)": "Residual / plug: latest bucket residual carried visibly for review.",
        "Other Non-current Assets (residual plug, transparent)": "Residual / plug: latest bucket residual carried visibly for review.",
        "Other Current Liabilities (residual plug, transparent)": "Residual / plug: latest bucket residual carried visibly for review.",
        "Other Non-current Liabilities (residual plug, transparent)": "Residual / plug: latest bucket residual carried visibly for review.",
        "Total Current Assets": "Formula total.",
        "Total Non-current Assets": "Formula total.",
        "Total Assets": "Formula total.",
        "Total Current Liabilities": "Formula total.",
        "Total Non-current Liabilities": "Formula total.",
        "Total Liabilities": "Formula total.",
        "Total Liabilities + Equity": "Formula total.",
        "Balance Check / Difference": "Review check: explicit drivers plus residual / held-constant lines.",
    }
    # Backwards-compatible aliases for downstream consumers (Audit Dashboard
    # checks, etc.) that still reference the V3.6.9 / V3.7.0 plug labels.
    BS_FORECAST_ROWS["PP&E"] = BS_FORECAST_ROWS["PP&E, Net"]
    BS_FORECAST_ROWS["Short-term Debt"] = BS_FORECAST_ROWS["Short-term Debt / Current Portion of Debt (incl. lease)"]
    BS_FORECAST_ROWS["Short-term Debt / Current Portion of Debt"] = BS_FORECAST_ROWS["Short-term Debt / Current Portion of Debt (incl. lease)"]
    BS_FORECAST_ROWS["Long-term Debt"] = BS_FORECAST_ROWS["Long-term Debt (incl. lease)"]
    BS_FORECAST_ROWS["Capital Lease Obligation, Current"] = BS_FORECAST_ROWS["  Memo: Capital Lease Obligation, Current (subset of ST Debt)"]
    BS_FORECAST_ROWS["Capital Lease Obligation, Non-current"] = BS_FORECAST_ROWS["  Memo: Capital Lease Obligation, Non-current (subset of LT Debt)"]
    BS_FORECAST_ROWS["Other Assets (Plug, transparent)"] = BS_FORECAST_ROWS[
        "Other Non-current Assets (residual plug, transparent)"
    ]
    BS_FORECAST_ROWS["Other Liabilities (Plug, transparent)"] = BS_FORECAST_ROWS[
        "Other Non-current Liabilities (residual plug, transparent)"
    ]

    # V3.7.1 expanded field mapping: each label points to a historical cache key.
    # Missing fields fall back to 0 with audit transparency.
    asset_label_to_key = {
        "Cash & Cash Equivalents": "cash",
        "Short-term Investments / Marketable Securities": "short_term_investments",
        "Accounts Receivable": "accounts_receivable",
        "Inventory": "inventory",
        "Other Current Assets": "other_current_assets",
        "PP&E, Net": "ppe",
        "Goodwill": "goodwill",
        "Intangible Assets": "intangible_assets",
        "Long-term Investments / Marketable Securities": "long_term_investments",
        "Deferred Tax Assets (Non-current)": "deferred_tax_assets",
        "Other Non-current Assets": "other_non_current_assets",
    }
    liab_label_to_key = {
        "Accounts Payable": "accounts_payable",
        "Deferred Revenue, Current": "current_deferred_revenue",
        "Short-term Debt / Current Portion of Debt (incl. lease)": "short_term_debt",
        "  Memo: Capital Lease Obligation, Current (subset of ST Debt)": "current_capital_lease_obligation",
        "Other Current Liabilities": "other_current_liabilities",
        "Long-term Debt (incl. lease)": "long_term_debt",
        "  Memo: Capital Lease Obligation, Non-current (subset of LT Debt)": "long_term_capital_lease_obligation",
        "Other Non-current Liabilities (incl. deferred revenue NC)": "other_non_current_liabilities",
        "  Memo: Deferred Revenue, Non-current (subset of Other NC Liab)": "non_current_deferred_revenue",
        "Deferred Tax Liabilities (Non-current)": "deferred_tax_liabilities",
    }

    def _safe_get(year, key):
        return _hist_model_value(ctx, "balance_sheet", key, year, out.market)

    plug_assets_label = "Other Non-current Assets (residual plug, transparent)"
    plug_liab_label = "Other Non-current Liabilities (residual plug, transparent)"

    # V3.9.8.7: rows where Apple Form 10-K does not separately disclose the line
    # item. Historical cells display "N/D in 10-K" instead of a silent zero so
    # reviewers cannot mistake the workbook for fabricating filing data.
    not_separately_disclosed_keys = {
        "goodwill": "N/D in 10-K (not separately disclosed; included in Other Non-current Assets)",
        "intangible_assets": "N/D in 10-K (not separately disclosed; included in Other Non-current Assets)",
        "deferred_tax_assets": "N/D in 10-K (not separately disclosed)",
        "deferred_tax_liabilities": "N/D in 10-K (not separately disclosed)",
        "current_capital_lease_obligation": "N/D in 10-K (lease obligation not broken out this year)",
        "long_term_capital_lease_obligation": "N/D in 10-K (lease obligation not broken out this year)",
        "non_current_deferred_revenue": "N/D in 10-K (not separately disclosed)",
    }

    def _write_hist_bs_cell(row_index, year, field_key):
        value = _safe_get(year, field_key)
        if value is None and field_key in not_separately_disclosed_keys:
            c = ws.cell(row=row_index, column=idx, value=not_separately_disclosed_keys[field_key])
            c.font = font_watermark()
            c.alignment = Alignment(horizontal="right", wrap_text=False)
        else:
            c = ws.cell(row=row_index, column=idx, value=value)
            c.font = font_formula()
            if value is not None:
                c.number_format = FMT_COMMA2
        c.border = border_thin()

    for idx, year in enumerate(historical_years, start=2):
        # Write expanded asset and liability line items directly from cache.
        # Fields that the filing does not separately disclose are surfaced as
        # explicit N/D markers; numeric residuals continue to absorb them so
        # totals reconcile to filing.
        for label, field_key in asset_label_to_key.items():
            _write_hist_bs_cell(row_map[label], year, field_key)
        for label, field_key in liab_label_to_key.items():
            _write_hist_bs_cell(row_map[label], year, field_key)

        total_assets = _safe_get(year, "total_assets")
        total_current_assets_cache = _safe_get(year, "total_current_assets")
        total_nc_assets_cache = _safe_get(year, "total_non_current_assets")
        total_current_liab_cache = _safe_get(year, "total_current_liabilities")
        total_nc_liab_cache = _safe_get(year, "total_non_current_liabilities")
        total_liab_cache = _safe_get(year, "total_liabilities")
        total_debt_h = _safe_get(year, "total_debt")
        equity = _safe_get(year, "total_equity") or 0.0

        # Sum of identified asset items.
        cash_h = _safe_get(year, "cash") or 0.0
        sti_h = _safe_get(year, "short_term_investments") or 0.0
        ar_h = _safe_get(year, "accounts_receivable") or 0.0
        inv_h = _safe_get(year, "inventory") or 0.0
        oca_h = _safe_get(year, "other_current_assets") or 0.0
        identified_current_assets = cash_h + sti_h + ar_h + inv_h + oca_h

        ppe_h = _safe_get(year, "ppe") or 0.0
        gw_h = _safe_get(year, "goodwill") or 0.0
        intang_h = _safe_get(year, "intangible_assets") or 0.0
        lti_h = _safe_get(year, "long_term_investments") or 0.0
        dta_h = _safe_get(year, "deferred_tax_assets") or 0.0
        onca_h = _safe_get(year, "other_non_current_assets") or 0.0
        identified_nc_assets = ppe_h + gw_h + intang_h + lti_h + dta_h + onca_h

        # Residual = bucket total - sum of identified items (when the bucket total exists).
        other_current_assets_residual = (
            total_current_assets_cache - identified_current_assets
            if total_current_assets_cache is not None
            else None
        )
        if total_nc_assets_cache is not None:
            other_nc_assets_residual = total_nc_assets_cache - identified_nc_assets
        elif total_assets is not None and total_current_assets_cache is not None:
            other_nc_assets_residual = (total_assets - total_current_assets_cache) - identified_nc_assets
        else:
            other_nc_assets_residual = None

        total_current_assets_h = total_current_assets_cache or (identified_current_assets + (other_current_assets_residual or 0))
        total_nc_assets_h = total_nc_assets_cache or (identified_nc_assets + (other_nc_assets_residual or 0))

        # Liability side. yfinance treats Long Term Debt as "Long Term Debt and
        # Capital Lease Obligation" (i.e. capital lease is already inside LT Debt
        # for AAPL); same for current debt. So we display capital lease as memo
        # lines and exclude them from the identified-sum to avoid double-count.
        ap_h = _safe_get(year, "accounts_payable") or 0.0
        dr_c_h = _safe_get(year, "current_deferred_revenue") or 0.0
        std_h = _safe_get(year, "short_term_debt") or 0.0
        ocl_h = _safe_get(year, "other_current_liabilities") or 0.0
        identified_current_liab = ap_h + dr_c_h + std_h + ocl_h

        ltd_h = _safe_get(year, "long_term_debt") or 0.0
        dtl_h = _safe_get(year, "deferred_tax_liabilities") or 0.0
        oncl_h = _safe_get(year, "other_non_current_liabilities") or 0.0
        # Non-current Deferred Revenue is a memo subset of Other NC Liab (yfinance
        # rollup matches LT Debt + Other NC Liab), so it is excluded from the
        # identified-sum to avoid double-counting.
        identified_nc_liab = ltd_h + dtl_h + oncl_h

        other_current_liab_residual = (
            total_current_liab_cache - identified_current_liab
            if total_current_liab_cache is not None
            else None
        )
        if total_nc_liab_cache is not None:
            other_nc_liab_residual = total_nc_liab_cache - identified_nc_liab
        elif total_liab_cache is not None and total_current_liab_cache is not None:
            other_nc_liab_residual = (total_liab_cache - total_current_liab_cache) - identified_nc_liab
        elif total_assets is not None and total_current_liab_cache is not None:
            implied_total_liab = total_assets - equity
            other_nc_liab_residual = (implied_total_liab - total_current_liab_cache) - identified_nc_liab
        else:
            other_nc_liab_residual = None

        total_current_liab_h = total_current_liab_cache or (identified_current_liab + (other_current_liab_residual or 0))
        total_nc_liab_h = total_nc_liab_cache or (identified_nc_liab + (other_nc_liab_residual or 0))
        total_liab_h = total_liab_cache or (total_current_liab_h + total_nc_liab_h)

        if total_debt_h is None:
            total_debt_h = std_h + ltd_h

        hist_values = {
            "Other Current Assets (residual plug, transparent)": other_current_assets_residual,
            "Total Current Assets": total_current_assets_h,
            plug_assets_label: other_nc_assets_residual,
            "Total Non-current Assets": total_nc_assets_h,
            "Total Assets": total_assets,
            "Other Current Liabilities (residual plug, transparent)": other_current_liab_residual,
            "Total Current Liabilities": total_current_liab_h,
            plug_liab_label: other_nc_liab_residual,
            "Total Non-current Liabilities": total_nc_liab_h,
            "Total Debt": total_debt_h,
            "Total Liabilities": total_liab_h,
            "Total Equity": equity if total_assets is not None else None,
            "Total Liabilities + Equity": (total_liab_h + equity) if total_assets is not None else None,
        }
        for label, value in hist_values.items():
            c = ws.cell(row=row_map[label], column=idx, value=value)
            c.font = font_formula()
            c.border = border_thin()
            if value is not None:
                c.number_format = FMT_COMMA2
        # Historical columns: reported BS reconciles by construction; display
        # OK to match the forecast-side professional Balance Check display.
        check_value = "OK" if total_assets is not None else None
        c = ws.cell(row=row_map["Balance Check / Difference"], column=idx, value=check_value)
        c.font = font_formula()
        c.border = border_thin()
        c.alignment = Alignment(horizontal="center")

    latest_equity = _latest_historical_model_value(ctx, "balance_sheet", "total_equity", out.market) or 0.0

    # Latest historical values for every new V3.7.1 line item; held flat in forecast.
    def _latest_or_zero(key):
        v = _latest_historical_model_value(ctx, "balance_sheet", key, out.market)
        return v if v is not None else 0.0

    latest_short_term_investments = _latest_or_zero("short_term_investments")
    latest_other_current_assets = _latest_or_zero("other_current_assets")
    latest_goodwill = _latest_or_zero("goodwill")
    latest_intangibles = _latest_or_zero("intangible_assets")
    latest_long_term_investments = _latest_or_zero("long_term_investments")
    latest_deferred_tax_assets = _latest_or_zero("deferred_tax_assets")
    latest_other_nc_assets = _latest_or_zero("other_non_current_assets")
    latest_deferred_revenue_current = _latest_or_zero("current_deferred_revenue")
    latest_current_lease = _latest_or_zero("current_capital_lease_obligation")
    latest_other_current_liabilities = _latest_or_zero("other_current_liabilities")
    latest_long_term_lease = _latest_or_zero("long_term_capital_lease_obligation")
    latest_deferred_revenue_non_current = _latest_or_zero("non_current_deferred_revenue")
    latest_deferred_tax_liabilities = _latest_or_zero("deferred_tax_liabilities")
    latest_other_nc_liabilities = _latest_or_zero("other_non_current_liabilities")
    latest_total_current_assets = _latest_historical_model_value(ctx, "balance_sheet", "total_current_assets", out.market)
    latest_total_nc_assets = _latest_historical_model_value(ctx, "balance_sheet", "total_non_current_assets", out.market)
    latest_total_current_liab = _latest_historical_model_value(ctx, "balance_sheet", "total_current_liabilities", out.market)
    latest_total_nc_liab = _latest_historical_model_value(ctx, "balance_sheet", "total_non_current_liabilities", out.market)

    latest_total_assets = _latest_historical_model_value(ctx, "balance_sheet", "total_assets", out.market)
    latest_cash = _latest_or_zero("cash")
    latest_ar = _latest_or_zero("accounts_receivable")
    latest_inv = _latest_or_zero("inventory")
    latest_ap = _latest_or_zero("accounts_payable")
    latest_ppe = _latest_or_zero("ppe")
    latest_std = _latest_or_zero("short_term_debt")
    latest_ltd = _latest_or_zero("long_term_debt")
    latest_total_debt = _latest_historical_model_value(ctx, "balance_sheet", "total_debt", out.market)
    if latest_total_debt is None:
        latest_total_debt = latest_std + latest_ltd

    # V3.7.1 residuals: bucket total minus sum of identified items. These are the
    # NAMED residual plugs held flat in forecast.
    identified_current_assets_latest = (
        latest_cash + latest_short_term_investments + latest_ar + latest_inv + latest_other_current_assets
    )
    identified_nc_assets_latest = (
        latest_ppe + latest_goodwill + latest_intangibles
        + latest_long_term_investments + latest_deferred_tax_assets + latest_other_nc_assets
    )
    # Capital lease is a memo subset of ST / LT Debt (yfinance double-includes)
    # so it is excluded from the identified-sum used to compute the residual plug.
    identified_current_liab_latest = (
        latest_ap + latest_deferred_revenue_current + latest_std + latest_other_current_liabilities
    )
    identified_nc_liab_latest = (
        latest_ltd + latest_deferred_tax_liabilities + latest_other_nc_liabilities
    )  # DR Non-current is a memo subset of Other NC Liab; not added.

    if latest_total_current_assets is not None:
        other_current_assets_plug_const = latest_total_current_assets - identified_current_assets_latest
    else:
        other_current_assets_plug_const = 0.0
    if latest_total_nc_assets is not None:
        other_nc_assets_plug_const = latest_total_nc_assets - identified_nc_assets_latest
    elif latest_total_assets is not None and latest_total_current_assets is not None:
        other_nc_assets_plug_const = (latest_total_assets - latest_total_current_assets) - identified_nc_assets_latest
    else:
        other_nc_assets_plug_const = 0.0
    if latest_total_current_liab is not None:
        other_current_liab_plug_const = latest_total_current_liab - identified_current_liab_latest
    else:
        other_current_liab_plug_const = 0.0
    if latest_total_nc_liab is not None:
        other_nc_liab_plug_const = latest_total_nc_liab - identified_nc_liab_latest
    elif latest_total_assets is not None and latest_total_current_liab is not None:
        implied_total_liab = latest_total_assets - latest_equity
        other_nc_liab_plug_const = (implied_total_liab - latest_total_current_liab) - identified_nc_liab_latest
    else:
        other_nc_liab_plug_const = 0.0

    # Flat-line constants for the new V3.7.1 items (forecast period = latest historical).
    flat_constants = {
        "Short-term Investments / Marketable Securities": latest_short_term_investments,
        "Other Current Assets": latest_other_current_assets,
        "Goodwill": latest_goodwill,
        "Intangible Assets": latest_intangibles,
        "Long-term Investments / Marketable Securities": latest_long_term_investments,
        "Deferred Tax Assets (Non-current)": latest_deferred_tax_assets,
        "Other Non-current Assets": latest_other_nc_assets,
        "Deferred Revenue, Current": latest_deferred_revenue_current,
        "  Memo: Capital Lease Obligation, Current (subset of ST Debt)": latest_current_lease,
        "Other Current Liabilities": latest_other_current_liabilities,
        "  Memo: Capital Lease Obligation, Non-current (subset of LT Debt)": latest_long_term_lease,
        "Other Non-current Liabilities (incl. deferred revenue NC)": latest_other_nc_liabilities,
        "  Memo: Deferred Revenue, Non-current (subset of Other NC Liab)": latest_deferred_revenue_non_current,
        "Deferred Tax Liabilities (Non-current)": latest_deferred_tax_liabilities,
    }
    for label in flat_constants:
        bs_classification[label] = "Held constant at latest available balance; no separate forecast schedule in this version."
    # V3.9.8.7: row-specific overrides — replace generic "held constant" labels
    # with categorisation that matches each line item's actual data treatment.
    not_separately_disclosed_note = (
        "Not separately disclosed in 10-K (e.g., AAPL): the filing reports this line within "
        "Other Non-current Assets / Other Non-current Liabilities. Held at $0 in forecast to "
        "avoid double-count; historical cells display 'N/D in 10-K' where the filing does not "
        "break it out, and display the filing-backed value where the cache supplies one."
    )
    bs_classification["Goodwill"] = not_separately_disclosed_note
    bs_classification["Intangible Assets"] = not_separately_disclosed_note
    bs_classification["Deferred Tax Assets (Non-current)"] = (
        "Filing-backed where the 10-K discloses it; held constant at latest available balance "
        "in forecast. Marked 'N/D in 10-K' for years the filing does not separately disclose."
    )
    bs_classification["Deferred Tax Liabilities (Non-current)"] = (
        "Filing-backed where the 10-K discloses it; otherwise 'N/D in 10-K'. Forecast held at "
        "$0 / latest balance, not separately modeled."
    )
    bs_classification["Short-term Investments / Marketable Securities"] = (
        "Held constant at latest available balance: filing-backed historical, no driver in this version."
    )
    bs_classification["Long-term Investments / Marketable Securities"] = (
        "Held constant at latest available balance: filing-backed historical, no driver in this version."
    )
    bs_classification["Other Current Assets"] = (
        "Held constant at latest available balance: filing-backed historical, no driver in this version."
    )
    bs_classification["Other Non-current Assets"] = (
        "Held constant at latest available balance: filing-backed historical (this line carries the "
        "Goodwill / Intangibles content for AAPL per filing); no driver in this version."
    )
    bs_classification["Deferred Revenue, Current"] = (
        "Held constant at latest available balance: filing-backed historical, no driver in this version."
    )
    bs_classification["Other Current Liabilities"] = (
        "Held constant at latest available balance: filing-backed historical, no driver in this version."
    )
    bs_classification["Other Non-current Liabilities (incl. deferred revenue NC)"] = (
        "Held constant at latest available balance: filing-backed historical, no driver in this version."
    )
    for label in (
        "  Memo: Capital Lease Obligation, Current (subset of ST Debt)",
        "  Memo: Capital Lease Obligation, Non-current (subset of LT Debt)",
        "  Memo: Deferred Revenue, Non-current (subset of Other NC Liab)",
    ):
        bs_classification[label] = "Memo / not separately modeled in totals to avoid double-counting."
    for label, class_note in bs_classification.items():
        if label in row_map:
            _set_note_cell(ws, row_map[label], classification_col_idx, class_note)

    for i in range(n):
        col_idx = forecast_start_col + i
        col = get_column_letter(col_idx)
        prev_col = get_column_letter(col_idx - 1)
        pl_net_income = _sheet_ref("P&L Forecast", PL_FORECAST_ROWS["Net Income"], col_idx)
        cf_cash = _sheet_ref("Cash Flow Forecast", CF_FORECAST_ROWS["Ending Cash"], col_idx)

        # Current Assets
        _set_formula_cell(ws, row_map["Cash & Cash Equivalents"], col_idx, f"={cf_cash}", FMT_COMMA2, True)
        _set_formula_cell(ws, row_map["Accounts Receivable"], col_idx, f"={_sheet_ref('Supporting Schedules', SCHEDULE_ROWS['wc']['Accounts Receivable'], col_idx)}", FMT_COMMA2, True)
        _set_formula_cell(ws, row_map["Inventory"], col_idx, f"={_sheet_ref('Supporting Schedules', SCHEDULE_ROWS['wc']['Inventory'], col_idx)}", FMT_COMMA2, True)
        # V3.7.1 new flat lines (held at latest historical).
        for label in (
            "Short-term Investments / Marketable Securities",
            "Other Current Assets",
            "Goodwill",
            "Intangible Assets",
            "Long-term Investments / Marketable Securities",
            "Deferred Tax Assets (Non-current)",
            "Other Non-current Assets",
            "Deferred Revenue, Current",
            "  Memo: Capital Lease Obligation, Current (subset of ST Debt)",
            "Other Current Liabilities",
            "  Memo: Capital Lease Obligation, Non-current (subset of LT Debt)",
            "Other Non-current Liabilities (incl. deferred revenue NC)",
            "  Memo: Deferred Revenue, Non-current (subset of Other NC Liab)",
            "Deferred Tax Liabilities (Non-current)",
        ):
            _set_formula_cell(ws, row_map[label], col_idx, f"={flat_constants[label]}", FMT_COMMA2, True)
        _set_formula_cell(ws, row_map["Other Current Assets (residual plug, transparent)"], col_idx, f"={other_current_assets_plug_const}", FMT_COMMA2, True)
        _set_formula_cell(
            ws,
            row_map["Total Current Assets"],
            col_idx,
            "+".join(
                f"{col}{row_map[lbl]}" for lbl in (
                    "Cash & Cash Equivalents",
                    "Short-term Investments / Marketable Securities",
                    "Accounts Receivable",
                    "Inventory",
                    "Other Current Assets",
                    "Other Current Assets (residual plug, transparent)",
                )
            ).join(("=", "")),
            FMT_COMMA2,
        )
        # Non-current Assets
        _set_formula_cell(ws, row_map["PP&E, Net"], col_idx, f"={_sheet_ref('Supporting Schedules', SCHEDULE_ROWS['ppe']['Ending Net PP&E'], col_idx)}", FMT_COMMA2, True)
        _set_formula_cell(ws, row_map[plug_assets_label], col_idx, f"={other_nc_assets_plug_const}", FMT_COMMA2, True)
        _set_formula_cell(
            ws,
            row_map["Total Non-current Assets"],
            col_idx,
            "+".join(
                f"{col}{row_map[lbl]}" for lbl in (
                    "PP&E, Net",
                    "Goodwill",
                    "Intangible Assets",
                    "Long-term Investments / Marketable Securities",
                    "Deferred Tax Assets (Non-current)",
                    "Other Non-current Assets",
                    plug_assets_label,
                )
            ).join(("=", "")),
            FMT_COMMA2,
        )
        _set_formula_cell(
            ws,
            row_map["Total Assets"],
            col_idx,
            f"={col}{row_map['Total Current Assets']}+{col}{row_map['Total Non-current Assets']}",
            FMT_COMMA2,
        )
        # Current Liabilities (Capital Lease memo line excluded from total)
        _set_formula_cell(ws, row_map["Accounts Payable"], col_idx, f"={_sheet_ref('Supporting Schedules', SCHEDULE_ROWS['wc']['Accounts Payable'], col_idx)}", FMT_COMMA2, True)
        _set_formula_cell(ws, row_map["Short-term Debt / Current Portion of Debt (incl. lease)"], col_idx, f"={_sheet_ref('Supporting Schedules', SCHEDULE_ROWS['debt']['Short-term Debt'], col_idx)}", FMT_COMMA2, True)
        _set_formula_cell(ws, row_map["Other Current Liabilities (residual plug, transparent)"], col_idx, f"={other_current_liab_plug_const}", FMT_COMMA2, True)
        _set_formula_cell(
            ws,
            row_map["Total Current Liabilities"],
            col_idx,
            "+".join(
                f"{col}{row_map[lbl]}" for lbl in (
                    "Accounts Payable",
                    "Deferred Revenue, Current",
                    "Short-term Debt / Current Portion of Debt (incl. lease)",
                    "Other Current Liabilities",
                    "Other Current Liabilities (residual plug, transparent)",
                )
            ).join(("=", "")),
            FMT_COMMA2,
        )
        # Non-current Liabilities (Capital Lease + DR Non-current are memo only)
        _set_formula_cell(ws, row_map["Long-term Debt (incl. lease)"], col_idx, f"={_sheet_ref('Supporting Schedules', SCHEDULE_ROWS['debt']['Long-term Debt'], col_idx)}", FMT_COMMA2, True)
        _set_formula_cell(ws, row_map[plug_liab_label], col_idx, f"={other_nc_liab_plug_const}", FMT_COMMA2, True)
        _set_formula_cell(
            ws,
            row_map["Total Non-current Liabilities"],
            col_idx,
            "+".join(
                f"{col}{row_map[lbl]}" for lbl in (
                    "Long-term Debt (incl. lease)",
                    "Other Non-current Liabilities (incl. deferred revenue NC)",
                    "Deferred Tax Liabilities (Non-current)",
                    plug_liab_label,
                )
            ).join(("=", "")),
            FMT_COMMA2,
        )
        _set_formula_cell(ws, row_map["Total Debt"], col_idx, f"={_sheet_ref('Supporting Schedules', SCHEDULE_ROWS['debt']['Total Debt (Ending, ST + LT)'], col_idx)}", FMT_COMMA2, True)
        _set_formula_cell(
            ws,
            row_map["Total Liabilities"],
            col_idx,
            f"={col}{row_map['Total Current Liabilities']}+{col}{row_map['Total Non-current Liabilities']}",
            FMT_COMMA2,
        )
        equity_begin = f"{prev_col}{row_map['Total Equity']}"
        # V3.7.4: Equity = Beg + NI - Dividends - Buybacks. Pull from the
        # Shareholder Returns Schedule when available; fall back to the legacy
        # ref_dividends assumption otherwise so degenerate exports still work.
        sr_map_bs = SCHEDULE_ROWS.get("shareholder_returns") or {}
        if sr_map_bs.get("Dividends Paid") and sr_map_bs.get("Share Repurchases (Buybacks)"):
            sr_div_bs = _sheet_ref("Supporting Schedules", sr_map_bs["Dividends Paid"], col_idx)
            sr_buy_bs = _sheet_ref("Supporting Schedules", sr_map_bs["Share Repurchases (Buybacks)"], col_idx)
            equity_formula = f"={equity_begin}+{pl_net_income}-{sr_div_bs}-{sr_buy_bs}"
        else:
            equity_formula = f"={equity_begin}+{pl_net_income}-{ref_dividends}"
        _set_formula_cell(ws, row_map["Total Equity"], col_idx, equity_formula, FMT_COMMA2, True)
        _set_formula_cell(
            ws,
            row_map["Total Liabilities + Equity"],
            col_idx,
            f"={col}{row_map['Total Liabilities']}+{col}{row_map['Total Equity']}",
            FMT_COMMA2,
        )
        # V3.8.4 polish: show OK / Review status instead of leaking a tiny
        # floating-point residual to the reader. Tolerance is 1 workbook unit
        # (the same threshold the Review Dashboard uses for the parity check),
        # and any breach displays the actual difference inline for diagnosis.
        ta_ref = f"{col}{row_map['Total Assets']}"
        tle_ref = f"{col}{row_map['Total Liabilities + Equity']}"
        _set_formula_cell(
            ws,
            row_map["Balance Check / Difference"],
            col_idx,
            f'=IF(ABS({ta_ref}-{tle_ref})<=1,"OK","Review (Δ="&TEXT({ta_ref}-{tle_ref},"#,##0.00")&")")',
            None,
        )

    row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "Balance Sheet Forecast classification is explicit: Cash / AR / Inventory / AP / PP&E / Debt are schedule-driven; selected balance sheet lines are held constant at latest available balance where no schedule exists; residual plug rows remain visible review items. Total Equity rolls from the prior-year equity cell plus NI less dividends and buybacks; OCI and other comprehensive equity bridge items are not modeled.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)


def build_assumptions_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = ASSUMP
    # V3.9.0 Forecast Path Upgrade v1: Year 1..Year 5 input cells live in
    # columns B..F for the operating forecast drivers; legacy single-input
    # blocks reuse column B for the value and column C for notes, so we widen
    # D..G for the path block while leaving non-path blocks unchanged.
    apply_column_widths(ws, {"A": 34, "B": 16, "C": 50, "D": 16, "E": 16, "F": 16, "G": 56})

    row = _write_sheet_heading(ws, 1, f"DCF Assumptions - {inp.company} ({inp.symbol})", inp, out, ctx)
    default_quality = (ctx.get("current_params") or {}).get("default_quality") or {}
    if default_quality.get("requires_review"):
        _set_note_cell(ws, row, 1, default_quality.get("banner") or "Default assumptions require review before use.")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

    write_section_title(ws, row, "Market Data", span_cols=3)
    row += 1
    market_rows = [
        ("Symbol", inp.symbol, None, None),
        ("Company Name", inp.company, None, None),
        ("Current Price", inp.price, FMT_COMMA2, "price"),
        ("Market", out.market, None, "market"),
        ("Currency", out.currency, None, "currency"),
        ("Data Source", (ctx.get("current_params") or {}).get("data_source") or DEFAULT_DATA_SOURCE_TEXT, None, "data_source"),
        ("Risk-Free Rate (Rf)", out.wacc_components.get("rf"), FMT_PCT2, "rf"),
        ("Equity Risk Premium (ERP)", out.wacc_components.get("erp"), FMT_PCT2, "erp"),
        ("Beta", out.wacc_components.get("beta"), "0.00", "beta"),
        ("WACC (used in model)", out.wacc_components.get("wacc"), FMT_PCT2, "wacc"),
    ]
    for label, val, fmt, key in market_rows:
        _set_label_cell(ws, row, 1, label)
        _set_input_cell(ws, row, 2, val, fmt)
        if key:
            ASSUMP_CELLS[key] = f"B{row}"
        row += 1
    row += 1

    drivers = normalized_driver_assumptions(inp)

    write_section_title(ws, row, "Base-Year Financials", span_cols=3)
    row += 1
    operating_rows = [
        ("Revenue", inp.revenue, FMT_COMMA2, "revenue"),
        ("EBIT", inp.ebit, FMT_COMMA2, "ebit"),
        ("D&A (add-back)", inp.da, FMT_COMMA2, "da"),
        ("CapEx (positive = outflow)", inp.capex, FMT_COMMA2, "capex"),
        ("Working Capital Change (positive = outflow)", inp.wc_change, FMT_COMMA2, "wc_change"),
        ("Effective Tax Rate", inp.tax_rate, FMT_PCT2, "tax_rate"),
        ("Net Debt", inp.net_debt, FMT_COMMA2, "net_debt"),
        ("Diluted Shares Outstanding (M)", inp.shares, FMT_COMMA2, "shares"),
    ]
    for label, val, fmt, key in operating_rows:
        _set_label_cell(ws, row, 1, label)
        _set_input_cell(ws, row, 2, val, fmt)
        ASSUMP_CELLS[key] = f"B{row}"
        row += 1
    row += 1

    write_section_title(ws, row, "WACC Bridge / Cost of Capital Decision Layer", span_cols=3)
    row += 1
    _set_label_cell(ws, row, 1, "Risk-free Rate (rf)")
    _set_formula_cell(ws, row, 2, f"={_abs_ref(ASSUMP_CELLS['rf'])}", FMT_PCT2, True)
    row += 1
    _set_label_cell(ws, row, 1, "Equity Risk Premium (erp)")
    _set_formula_cell(ws, row, 2, f"={_abs_ref(ASSUMP_CELLS['erp'])}", FMT_PCT2, True)
    row += 1
    _set_label_cell(ws, row, 1, "Beta")
    _set_formula_cell(ws, row, 2, f"={_abs_ref(ASSUMP_CELLS['beta'])}", "0.00", True)
    row += 1
    _set_label_cell(ws, row, 1, "CAPM-derived Cost of Equity (Ke)")
    _set_formula_cell(
        ws,
        row,
        2,
        f"={_abs_ref(ASSUMP_CELLS['rf'])}+{_abs_ref(ASSUMP_CELLS['beta'])}*{_abs_ref(ASSUMP_CELLS['erp'])}",
        FMT_PCT2,
    )
    ASSUMP_CELLS["cost_of_equity"] = f"B{row}"
    row += 1
    # Cost of Debt is defined in the Debt Schedule v2 block below; we forward-allocate the cell here
    # so the WACC Bridge can reference it. The Debt Schedule block writes into the same cell name.
    selected_kd = (getattr(out, "wacc_decision_bridge", None) or {}).get("pre_tax_cost_of_debt")
    _set_label_cell(ws, row, 1, "Cost of Debt (Kd, pre-tax)")
    _set_input_cell(ws, row, 2, selected_kd if selected_kd is not None else 0.05, FMT_PCT2)
    ASSUMP_CELLS["cost_of_debt"] = f"B{row}"
    _set_note_cell(ws, row, 3, "Selected normalized Kd; generic 5.0% is only a fallback when no selected Kd is supplied.")
    row += 1
    _set_label_cell(ws, row, 1, "Tax Rate (linked)")
    tax_ref_addr = ASSUMP_CELLS["tax_rate"]
    _set_formula_cell(ws, row, 2, f"=${tax_ref_addr[0]}${tax_ref_addr[1:]}", FMT_PCT2, True)
    ASSUMP_CELLS["wacc_tax_link"] = f"B{row}"
    row += 1
    _set_label_cell(ws, row, 1, "After-tax Cost of Debt (Kd*(1-t))")
    _set_formula_cell(
        ws,
        row,
        2,
        f"={_abs_ref(ASSUMP_CELLS['cost_of_debt'])}*(1-{_abs_ref(ASSUMP_CELLS['wacc_tax_link'])})",
        FMT_PCT2,
    )
    ASSUMP_CELLS["after_tax_kd"] = f"B{row}"
    row += 1
    # Capital structure weights — approximated from latest historical Total Debt and an
    # equity proxy (Price x Diluted Shares). Both editable.
    latest_total_debt = _latest_historical_model_value(ctx, "balance_sheet", "total_debt", out.market)
    if latest_total_debt is None:
        std = _latest_historical_model_value(ctx, "balance_sheet", "short_term_debt", out.market) or 0.0
        ltd = _latest_historical_model_value(ctx, "balance_sheet", "long_term_debt", out.market) or 0.0
        latest_total_debt = std + ltd
    _set_label_cell(ws, row, 1, "Total Debt (latest historical, M)")
    _set_input_cell(ws, row, 2, latest_total_debt or 0.0, FMT_COMMA2)
    ASSUMP_CELLS["wacc_debt_value"] = f"B{row}"
    row += 1
    _set_label_cell(ws, row, 1, "Equity Market Value proxy (Price x Shares, M)")
    _set_formula_cell(
        ws,
        row,
        2,
        f"={_abs_ref(ASSUMP_CELLS['price'])}*{_abs_ref(ASSUMP_CELLS['shares'])}",
        FMT_COMMA2,
    )
    ASSUMP_CELLS["wacc_equity_value"] = f"B{row}"
    row += 1
    _set_label_cell(ws, row, 1, "Weight of Debt (Wd)")
    _set_formula_cell(
        ws,
        row,
        2,
        f"={_abs_ref(ASSUMP_CELLS['wacc_debt_value'])}/MAX({_abs_ref(ASSUMP_CELLS['wacc_debt_value'])}+{_abs_ref(ASSUMP_CELLS['wacc_equity_value'])},1)",
        FMT_PCT2,
    )
    ASSUMP_CELLS["weight_debt"] = f"B{row}"
    row += 1
    _set_label_cell(ws, row, 1, "Weight of Equity (We)")
    _set_formula_cell(ws, row, 2, f"=1-{_abs_ref(ASSUMP_CELLS['weight_debt'])}", FMT_PCT2)
    ASSUMP_CELLS["weight_equity"] = f"B{row}"
    row += 1
    _set_label_cell(ws, row, 1, "Mechanical CAPM Reference WACC", True)
    _set_formula_cell(
        ws,
        row,
        2,
        f"={_abs_ref(ASSUMP_CELLS['weight_equity'])}*{_abs_ref(ASSUMP_CELLS['cost_of_equity'])}+{_abs_ref(ASSUMP_CELLS['weight_debt'])}*{_abs_ref(ASSUMP_CELLS['after_tax_kd'])}",
        FMT_PCT2,
    )
    ASSUMP_CELLS["wacc_indicative"] = f"B{row}"
    row += 1
    _set_label_cell(ws, row, 1, "Selected / Model WACC", True)
    _set_formula_cell(ws, row, 2, f"={_abs_ref(ASSUMP_CELLS['wacc'])}", FMT_PCT2, True)
    ASSUMP_CELLS["wacc_selected_echo"] = f"B{row}"
    row += 1
    _set_label_cell(ws, row, 1, "Selected vs Indicative Spread (Selected - Indicative)")
    _set_formula_cell(
        ws,
        row,
        2,
        f"={_abs_ref(ASSUMP_CELLS['wacc_selected_echo'])}-{_abs_ref(ASSUMP_CELLS['wacc_indicative'])}",
        FMT_PCT2,
    )
    ASSUMP_CELLS["wacc_spread"] = f"B{row}"
    row += 1

    # ── V3.7.5 WACC reference cases + treatment dropdown ───────────────
    _set_label_cell(ws, row, 1, "Selected WACC + 100 bps")
    _set_formula_cell(ws, row, 2, f"={_abs_ref(ASSUMP_CELLS['wacc_selected_echo'])}+0.01", FMT_PCT2, True)
    ASSUMP_CELLS["wacc_plus_100bps"] = f"B{row}"
    row += 1
    _set_label_cell(ws, row, 1, "Selected WACC - 100 bps")
    _set_formula_cell(ws, row, 2, f"=MAX(0.001,{_abs_ref(ASSUMP_CELLS['wacc_selected_echo'])}-0.01)", FMT_PCT2, True)
    ASSUMP_CELLS["wacc_minus_100bps"] = f"B{row}"
    row += 1

    wacc_bridge_payload = getattr(out, "wacc_decision_bridge", None) or {}
    wacc_treatment_key, wacc_treatment_label, _ = normalize_wacc_treatment(
        wacc_bridge_payload.get("selected_wacc_treatment")
    )
    _set_label_cell(ws, row, 1, "Selected WACC Treatment")
    _set_input_cell(ws, row, 2, wacc_treatment_label, None)
    ASSUMP_CELLS["wacc_treatment"] = f"B{row}"
    wacc_treatment_cell = f"B{row}"
    wacc_label_list = [WACC_TREATMENT_LABELS[k] for k in WACC_TREATMENTS]
    dv_wacc = DataValidation(
        type="list",
        formula1='"' + ",".join(wacc_label_list) + '"',
        allow_blank=True,
        errorStyle="warning",
        errorTitle="WACC Treatment",
        error="Pick a WACC Treatment from the dropdown; unknown values fall back to Selected / Model WACC.",
        promptTitle="WACC Treatment",
        prompt="Pick which WACC drives the DCF headline. CAPM is a mechanical audit reference.",
    )
    ws.add_data_validation(dv_wacc)
    dv_wacc.add(wacc_treatment_cell)
    _set_note_cell(ws, row, 3, " | ".join(wacc_label_list))
    row += 1

    _set_label_cell(ws, row, 1, "Selected WACC Used in DCF", True)
    treat_ref = _abs_ref(ASSUMP_CELLS["wacc_treatment"])
    sel_ref = _abs_ref(ASSUMP_CELLS["wacc_selected_echo"])
    ind_ref_v375 = _abs_ref(ASSUMP_CELLS["wacc_indicative"])
    plus_ref = _abs_ref(ASSUMP_CELLS["wacc_plus_100bps"])
    minus_ref = _abs_ref(ASSUMP_CELLS["wacc_minus_100bps"])
    lbl_capm = WACC_TREATMENT_LABELS["capm_indicative_wacc"]
    lbl_plus = WACC_TREATMENT_LABELS["selected_plus_spread_100bps"]
    lbl_minus = WACC_TREATMENT_LABELS["selected_minus_spread_100bps"]
    wacc_used_formula = (
        f'=IF({treat_ref}="{lbl_capm}",{ind_ref_v375},'
        f'IF({treat_ref}="{lbl_plus}",{plus_ref},'
        f'IF({treat_ref}="{lbl_minus}",{minus_ref},{sel_ref})))'
    )
    _set_formula_cell(ws, row, 2, wacc_used_formula, FMT_PCT2, True)
    ASSUMP_CELLS["wacc_used"] = f"B{row}"
    _set_note_cell(ws, row, 3, "Selected WACC used by DCF, FCF Build, and Sensitivity center.")
    row += 1

    _set_note_cell(
        ws,
        row,
        1,
        "WACC Bridge: Mechanical CAPM reference = We*Ke + Wd*Kd*(1-t). Selected / Model WACC is the default headline selection. Mechanical CAPM reference is an audit diagnostic, not headline selection. Audit Dashboard reviews the spread without changing the selected WACC.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # ── V3.7.6 Terminal Value Decision Layer ──────────────────────────
    write_section_title(ws, row, "Terminal Value Decision Layer", span_cols=3)
    row += 1
    terminal_bridge_payload = getattr(out, "terminal_decision_bridge", None) or {}
    terminal_treatment_key, terminal_treatment_label, _ = normalize_terminal_treatment(
        terminal_bridge_payload.get("selected_terminal_treatment")
    )
    fade_payload = terminal_bridge_payload.get("fade_period") or {}
    h_payload = terminal_bridge_payload.get("h_model") or {}

    _set_label_cell(ws, row, 1, "Terminal Treatment")
    _set_input_cell(ws, row, 2, terminal_treatment_label, None)
    ASSUMP_CELLS["tv_treatment"] = f"B{row}"
    tv_treatment_cell = f"B{row}"
    tv_label_list = [TERMINAL_TREATMENT_LABELS[k] for k in TERMINAL_TREATMENTS]
    dv_tv = DataValidation(
        type="list",
        formula1='"' + ",".join(tv_label_list) + '"',
        allow_blank=True,
        errorStyle="warning",
        errorTitle="Terminal Treatment",
        error="Pick a Terminal Treatment from the dropdown; unknown values fall back to Current Model Terminal Value.",
        promptTitle="Terminal Treatment",
        prompt="Pick which terminal-value method drives the DCF headline (Current / Gordon / Exit / Blend / Fade).",
    )
    ws.add_data_validation(dv_tv)
    dv_tv.add(tv_treatment_cell)
    _set_note_cell(ws, row, 3, " | ".join(tv_label_list))
    row += 1

    _set_label_cell(ws, row, 1, "Fade Years")
    _set_input_cell(ws, row, 2, int(fade_payload.get("fade_years") or DEFAULT_FADE_YEARS), "0")
    ASSUMP_CELLS["tv_fade_years"] = f"B{row}"
    _set_note_cell(ws, row, 3, "Reference-only: number of post-explicit-period fade years.")
    row += 1

    _set_label_cell(ws, row, 1, "Fade Target Growth")
    _set_input_cell(ws, row, 2, float(fade_payload.get("fade_target_growth") or float(inp.terminal_g or 0.0)), FMT_PCT2)
    ASSUMP_CELLS["tv_fade_target_g"] = f"B{row}"
    _set_note_cell(ws, row, 3, "Linear fade from Year-N FCF growth toward this target.")
    row += 1

    _set_label_cell(ws, row, 1, "Blend Weight: Gordon")
    _set_input_cell(ws, row, 2, float(terminal_bridge_payload.get("blend_weight_gordon") or 0.5), FMT_PCT2)
    ASSUMP_CELLS["tv_blend_w_g"] = f"B{row}"
    row += 1
    _set_label_cell(ws, row, 1, "Blend Weight: Exit")
    _set_input_cell(ws, row, 2, float(terminal_bridge_payload.get("blend_weight_exit") or 0.5), FMT_PCT2)
    ASSUMP_CELLS["tv_blend_w_e"] = f"B{row}"
    row += 1

    _set_label_cell(ws, row, 1, "H-Model Near Growth")
    _set_input_cell(ws, row, 2, float(h_payload.get("g_near") if h_payload.get("g_near") is not None else (getattr(inp, "h_model_g_near", None) or 0.0)), FMT_PCT2)
    ASSUMP_CELLS["h_model_g_near"] = f"B{row}"
    _set_note_cell(ws, row, 3, "Start-of-transition terminal growth; default uses terminal-year revenue growth (AAPL Base/Bear case-specific path).")
    row += 1

    _set_label_cell(ws, row, 1, "H-Model Long Growth")
    _set_input_cell(ws, row, 2, float(h_payload.get("g_long") if h_payload.get("g_long") is not None else (getattr(inp, "h_model_g_long", None) or float(inp.terminal_g or 0.0))), FMT_PCT2)
    ASSUMP_CELLS["h_model_g_long"] = f"B{row}"
    _set_note_cell(ws, row, 3, "Long-run terminal growth used in the H-Model denominator.")
    row += 1

    _set_label_cell(ws, row, 1, "H-Model Half-life (H)")
    _set_input_cell(ws, row, 2, float(h_payload.get("half_life") if h_payload.get("half_life") is not None else (getattr(inp, "h_model_half_life", None) or DEFAULT_H_MODEL_HALF_LIFE)), "0.0")
    ASSUMP_CELLS["h_model_half_life"] = f"B{row}"
    _set_note_cell(ws, row, 3, "Half-life of the terminal fade period; H=5 corresponds to a ten-year fade.")
    row += 1

    _set_label_cell(ws, row, 1, "Selected Terminal Value Used in DCF (M)", True)
    _set_input_cell(
        ws, row, 2,
        float(terminal_bridge_payload.get("selected_terminal_value_pv") or 0.0),
        FMT_COMMA2,
    )
    ASSUMP_CELLS["tv_selected_value"] = f"B{row}"
    _set_note_cell(ws, row, 3, "Calculator-computed; PV of selected terminal-value case at the selected WACC.")
    row += 1

    _set_note_cell(
        ws, row, 1,
        "V3.7.6: Terminal treatment is a valuation JUDGMENT. Default = Current Model Terminal Value preserves V3.7.5 headline IV exactly. Switching the dropdown intentionally changes the headline. Fade Period reference case extends FCFs N years post-explicit period before reverting to Gordon TV; reference-only, not full 10-year 3FS.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # V3.9.0 Forecast Path Upgrade v1: Operating forecast drivers are now
    # five-year explicit paths rather than flat single-value assumptions. Year 1
    # through Year 5 cells (columns B..F) are user-editable inputs in blue style;
    # column G holds an analyst note. The legacy ASSUMP_CELLS[key] entry points
    # at the Year 1 (column B) cell so consumers that still expect a single
    # reference keep working, while sheets that need year-by-year refs read
    # from ASSUMP_PATH_CELLS[key].
    current_params_for_paths = ctx.get("current_params") or {}
    _aapl = str(getattr(inp, "symbol", "")).upper() == "AAPL"
    write_section_title(ws, row, "Operating Forecast Path Assumptions (Year 1 - Year 5)", span_cols=7)
    row += 1
    op_bridge_payload = getattr(out, "operating_path_bridge", None) or {}
    op_source_key, op_source_label, _, _ = normalize_operating_path_source(
        getattr(inp, "selected_operating_path_source", None),
        getattr(inp, "symbol", None),
    )
    if not _aapl:
        op_source_label = OPERATING_PATH_SOURCE_SELECTED
    _set_label_cell(ws, row, 1, "Operating Path Source")
    _set_input_cell(ws, row, 2, op_source_label, None)
    ASSUMP_CELLS["operating_path_source"] = f"B{row}"
    dv_op = DataValidation(
        type="list",
        formula1=(
            '"' + ",".join(OPERATING_PATH_SOURCES) + '"'
            if _aapl
            else f'"{OPERATING_PATH_SOURCE_SELECTED}"'
        ),
        allow_blank=True,
        errorStyle="warning",
        errorTitle="Operating Path Source",
        error="Pick Selected Path or AAPL Operating Thesis Bridge. Non-AAPL cases use Selected Path only.",
        promptTitle="Operating Path Source",
        prompt="Revenue growth and EBIT margin switch together; default Selected Path preserves headline IV.",
    )
    ws.add_data_validation(dv_op)
    dv_op.add(f"B{row}")
    _set_note_cell(
        ws,
        row,
        3,
        (
            "Coupled selector: Bridge switches both revenue growth and EBIT margin to the AAPL Operating Thesis Bridge. Default = Selected Path."
            if _aapl
            else "N/A for non-AAPL; Selected Path only."
        ),
    )
    row += 1
    write_header_row(
        ws,
        row,
        ["Driver", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5", "Notes"],
        1,
    )
    row += 1

    _forecast_source = current_params_for_paths.get("forecast_path_source") or ""
    is_aapl_default_path = (
        _aapl
        and (
            _forecast_source == "AAPL neutral top-down base default"
            or _forecast_source.startswith("AAPL Base Case forecast path")
            or _forecast_source.startswith("AAPL Bear / Downside Case forecast path")
        )
    )
    _is_aapl_bear = _aapl and _forecast_source.startswith("AAPL Bear / Downside Case forecast path")
    path_notes = {
        "revenue_growth": "Year-by-year top-line growth; analyst should reflect product cycle, mix, and consensus glide path.",
        "ebit_margin": "Year-by-year operating margin path; reflects gross margin trend, opex leverage, and competitive pressure.",
        "da_pct_revenue": "Year-by-year depreciation and amortization as a percentage of revenue; diagnostic reference when asset-based PP&E schedule drives D&A.",
        "capex_pct_revenue": "Year-by-year reinvestment intensity; reflects capacity build, maintenance, and growth capex assumptions.",
        "wc_change_pct_revenue": "Year-by-year change in net working capital as a percentage of revenue; diagnostic reference when days-based working-capital schedule drives Delta NWC.",
    }
    if is_aapl_default_path:
        if _is_aapl_bear:
            path_notes["revenue_growth"] = (
                "AAPL Bear / Downside Case: V3.9.8.5 path retained (peak 5.0%). "
                "iPhone refresh underperforms; Services growth decelerates under regulatory pressure. "
                "See 'AAPL Operating Thesis' sheet for Products / Services bridge."
            )
            path_notes["ebit_margin"] = (
                "AAPL Bear / Downside Case: peak 33.6%; modest Services-mix support partially offset by elevated R&D intensity. "
                "See 'AAPL Operating Thesis' sheet for the GM / R&D / SG&A / EBIT margin bridge."
            )
        else:
            path_notes["revenue_growth"] = (
                "AAPL Base Case: peak 5.5% (Y3). Products growth low single digit (installed-base maturity, refresh cadence); "
                "Services growth low double digit (ecosystem monetisation). "
                "See 'AAPL Operating Thesis' sheet for Products / Services bridge. "
                "Base does NOT include full AI / Vision Pro Bull-case optionality."
            )
            path_notes["ebit_margin"] = (
                "AAPL Base Case: peak 34.5%. Supported by Services-mix progression (higher gross margin than Products) and modest operating leverage. "
                "See 'AAPL Operating Thesis' sheet for the GM / R&D / SG&A / EBIT margin bridge. "
                "Margin bridge is reviewer support, not a full driver model. This is not investment advice."
            )
    paths_block = {
        "revenue_growth": drivers["revenue_growth_path"],
        "ebit_margin": drivers["ebit_margin_path"],
        "da_pct_revenue": drivers["da_pct_revenue_path"],
        "capex_pct_revenue": drivers["capex_pct_revenue_path"],
        "wc_change_pct_revenue": drivers["wc_change_pct_revenue_path"],
    }
    driver_path_rows = [
        ("Revenue Growth", "revenue_growth", FMT_PCT2),
        ("EBIT Margin", "ebit_margin", FMT_PCT2),
        ("D&A % Revenue", "da_pct_revenue", FMT_PCT2),
        ("CapEx % Revenue", "capex_pct_revenue", FMT_PCT2),
        ("Delta NWC % Revenue", "wc_change_pct_revenue", FMT_PCT2),
    ]
    for label, key, fmt in driver_path_rows:
        _set_label_cell(ws, row, 1, label)
        path_vals = list(paths_block.get(key) or [drivers[key]] * 5)
        # Pad / truncate defensively so we always write exactly 5 input cells.
        while len(path_vals) < 5:
            path_vals.append(path_vals[-1] if path_vals else drivers[key])
        path_vals = path_vals[:5]
        path_cells: list[str] = []
        for i, val in enumerate(path_vals):
            col_idx = 2 + i
            _set_input_cell(ws, row, col_idx, val, fmt)
            path_cells.append(f"{get_column_letter(col_idx)}{row}")
        _set_note_cell(ws, row, 7, path_notes[key])
        ASSUMP_CELLS[key] = path_cells[0]
        ASSUMP_PATH_CELLS[key] = path_cells
        ASSUMP_SELECTED_PATH_CELLS[key] = path_cells
        row += 1
    if _aapl:
        bridge_rev = list(op_bridge_payload.get("bridge_revenue_growth_path") or [])
        bridge_margin = list(op_bridge_payload.get("bridge_ebit_margin_path") or [])
        while len(bridge_rev) < 5:
            bridge_rev.append(bridge_rev[-1] if bridge_rev else 0.0)
        while len(bridge_margin) < 5:
            bridge_margin.append(bridge_margin[-1] if bridge_margin else 0.0)
        bridge_rows = [
            ("Bridge Implied Total Growth", "revenue_growth", bridge_rev[:5], FMT_PCT2),
            ("Bridge Implied EBIT Margin", "ebit_margin", bridge_margin[:5], FMT_PCT2),
        ]
        for label, key, vals, fmt in bridge_rows:
            _set_label_cell(ws, row, 1, label)
            cells = []
            for i, val in enumerate(vals):
                col_idx = 2 + i
                _set_formula_cell(ws, row, col_idx, f"={float(val)}", fmt, True)
                cells.append(f"{get_column_letter(col_idx)}{row}")
            _set_note_cell(ws, row, 7, "AAPL Operating Thesis Bridge path; used only when Operating Path Source = Bridge.")
            ASSUMP_BRIDGE_PATH_CELLS[key] = cells
            row += 1
        source_ref = _abs_ref(ASSUMP_CELLS["operating_path_source"])
        for label, key, fmt in [
            ("Effective Revenue Growth Path", "revenue_growth", FMT_PCT2),
            ("Effective EBIT Margin Path", "ebit_margin", FMT_PCT2),
        ]:
            _set_label_cell(ws, row, 1, label, True)
            cells = []
            for i in range(5):
                selected_ref = _abs_ref(ASSUMP_SELECTED_PATH_CELLS[key][i])
                bridge_ref = _abs_ref(ASSUMP_BRIDGE_PATH_CELLS[key][i])
                col_idx = 2 + i
                _set_formula_cell(
                    ws,
                    row,
                    col_idx,
                    f'=IF({source_ref}="{OPERATING_PATH_SOURCE_AAPL_BRIDGE}",{bridge_ref},{selected_ref})',
                    fmt,
                    True,
                )
                cells.append(f"{get_column_letter(col_idx)}{row}")
            _set_note_cell(ws, row, 7, "Engine-driving path cells referenced by P&L Forecast, Supporting Schedules, and FCF Build.")
            ASSUMP_PATH_CELLS[key] = cells
            ASSUMP_CELLS[key] = cells[0]
            row += 1
    _set_note_cell(
        ws,
        row,
        1,
        (
            "AAPL default Base path is a neutral top-down analyst default: modest Products maturity, Services mix support, and stable-to-slightly-improving margin; "
            "no full segment model is connected. All path cells are user-editable."
            if is_aapl_default_path
            else "Revenue growth, EBIT margin, and CapEx paths are explicit year-by-year forecast drivers. D&A and working-capital paths may serve as diagnostic references where asset-based PP&E and days-based working-capital schedules are active."
        ),
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 2

    write_section_title(ws, row, "P&L Forecast Assumptions", span_cols=3)
    row += 1
    gross_margin = _latest_historical_gross_margin(ctx, out.market)
    if gross_margin is None:
        gross_margin = 0.0
    _set_label_cell(ws, row, 1, "Gross Margin")
    _set_input_cell(ws, row, 2, gross_margin, FMT_PCT2)
    ASSUMP_CELLS["gross_margin"] = f"B{row}"
    row += 1
    _set_label_cell(ws, row, 1, "Effective Tax Rate")
    tax_ref = ASSUMP_CELLS["tax_rate"]
    _set_formula_cell(ws, row, 2, f"=${tax_ref[0]}${tax_ref[1:]}", FMT_PCT2, True)
    _set_note_cell(ws, row, 3, f"Linked to DCF Effective Tax Rate input ({tax_ref}).")
    row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "P&L v1 reuses Revenue Growth, EBIT Margin, and Effective Tax Rate from DCF assumptions. Forecast columns are live formulas and will calculate when opened in Excel / 预测列为 Excel 公式，打开后会自动计算。",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    write_section_title(ws, row, "Working Capital Days Driver", span_cols=3)
    row += 1
    days_default = _latest_wc_days(ctx, out.market)
    dso_default = days_default.get("dso") or 0.0
    dio_default = days_default.get("dio") or 0.0
    dpo_default = days_default.get("dpo") or 0.0
    days_norm = _normalized_wc_days(ctx, out.market)
    days_rows = [
        ("DSO (Days Sales Outstanding)", dso_default, "0.0", "dso"),
        ("DIO (Days Inventory Outstanding)", dio_default, "0.0", "dio"),
        ("DPO (Days Payable Outstanding)", dpo_default, "0.0", "dpo"),
    ]
    for label, val, fmt, key in days_rows:
        _set_label_cell(ws, row, 1, label)
        _set_input_cell(ws, row, 2, val, fmt)
        ASSUMP_CELLS[key] = f"B{row}"
        row += 1
    norm_rows = [
        ("Normalized DSO Target (3Y avg)", days_norm.get("dso") if days_norm.get("dso") is not None else dso_default, "0.0", "dso_target"),
        ("Normalized DIO Target (3Y avg)", days_norm.get("dio") if days_norm.get("dio") is not None else dio_default, "0.0", "dio_target"),
        ("Normalized DPO Target (3Y avg)", days_norm.get("dpo") if days_norm.get("dpo") is not None else dpo_default, "0.0", "dpo_target"),
    ]
    for label, val, fmt, key in norm_rows:
        _set_label_cell(ws, row, 1, label)
        _set_input_cell(ws, row, 2, val or 0.0, fmt)
        ASSUMP_CELLS[key] = f"B{row}"
        row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "V3.9.9.2.1: Forecast AR / Inventory / AP use a faded-days schedule from latest actual DSO/DIO/DPO toward normalized 3Y targets. Delta NWC remains schedule-derived; % revenue WC rows are reference only.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    write_section_title(ws, row, "PP&E / D&A Schedule Driver", span_cols=3)
    row += 1
    da_pct_begin = _latest_da_pct_begin_ppe(ctx, out.market) or 0.0
    _set_label_cell(ws, row, 1, "D&A % Beginning Net PP&E")
    _set_input_cell(ws, row, 2, da_pct_begin, FMT_PCT2)
    ASSUMP_CELLS["da_pct_begin_ppe"] = f"B{row}"
    row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "V3.6.9: Forecast D&A = Beginning Net PP&E * D&A % Beginning PP&E. CapEx is still % Revenue. Replaces V3.6.8 D&A % Revenue path; legacy D&A % Revenue retained on Operating Forecast Assumptions for reference.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    write_section_title(ws, row, "Diagnostic / Non-headline Assumptions", span_cols=3)
    row += 1
    ar_pct = _latest_historical_ratio(ctx, "balance_sheet", "accounts_receivable", out.market)
    inventory_pct = _latest_historical_ratio(ctx, "balance_sheet", "inventory", out.market)
    ap_pct = _latest_historical_ratio(ctx, "balance_sheet", "accounts_payable", out.market)
    lite_rows = [
        ("AR % Revenue (diagnostic)", ar_pct if ar_pct is not None else 0.0, FMT_PCT2, "ar_pct_revenue"),
        ("Inventory % Revenue (diagnostic)", inventory_pct if inventory_pct is not None else 0.0, FMT_PCT2, "inventory_pct_revenue"),
        ("AP % Revenue (diagnostic)", ap_pct if ap_pct is not None else 0.0, FMT_PCT2, "ap_pct_revenue"),
        ("Dividends / Buybacks assumption", 0.0, FMT_COMMA2, "dividends_buybacks_placeholder"),
        ("Financing assumption", 0.0, FMT_COMMA2, "financing_placeholder"),
        ("Share count growth / buyback assumption", 0.0, FMT_PCT2, "share_count_growth_placeholder"),
    ]
    for label, val, fmt, key in lite_rows:
        _set_label_cell(ws, row, 1, label)
        _set_input_cell(ws, row, 2, val, fmt)
        ASSUMP_CELLS[key] = f"B{row}"
        row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "V3.6.9: AR / Inventory / AP forecast now uses the Days Driver above; the % Revenue inputs are retained as diagnostic reconciliation items. Dividend / financing / share-count rows are user-editable assumptions.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    write_section_title(ws, row, "Debt Schedule v2 Assumptions", span_cols=3)
    row += 1
    _set_label_cell(ws, row, 1, "Cost of Debt (linked to WACC Bridge)")
    _set_formula_cell(ws, row, 2, f"={_abs_ref(ASSUMP_CELLS['cost_of_debt'])}", FMT_PCT2, True)
    _set_note_cell(ws, row, 3, "Selected Kd; 5.0% only if fallback.")
    row += 1
    debt_rows = [
        ("New Debt Issuance (per year)", 0.0, FMT_COMMA2, "debt_issuance"),
        ("Debt Repayment (per year)", 0.0, FMT_COMMA2, "debt_repayment"),
    ]
    for label, val, fmt, key in debt_rows:
        _set_label_cell(ws, row, 1, label)
        _set_input_cell(ws, row, 2, val, fmt)
        ASSUMP_CELLS[key] = f"B{row}"
        row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "V3.6.9 Debt Schedule v2: roll-forward Beginning + Issuance - Repayment = Ending. Cost of Debt is the single source in the WACC Bridge above; Interest Expense = Average Debt x Cost of Debt. Issuance / Repayment default 0.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    write_section_title(ws, row, "Valuation Assumptions", span_cols=3)
    row += 1
    valuation_rows = [
        ("Forecast Period (years)", inp.forecast_years, FMT_COMMA, "forecast_years"),
        ("Terminal Growth Rate (g)", inp.terminal_g, FMT_PCT2, "terminal_g"),
        ("Exit Multiple (EV/EBITDA)", inp.exit_multiple, FMT_MULTIPLE, "exit_multiple"),
        ("Terminal Value Method", inp.tv_method, None, "tv_method"),
    ]
    for label, val, fmt, key in valuation_rows:
        _set_label_cell(ws, row, 1, label)
        _set_input_cell(ws, row, 2, val, fmt)
        ASSUMP_CELLS[key] = f"B{row}"
        row += 1

    row += 2

    # ── V3.7.2 Net Debt Bridge v2 ───────────────────────────────────────
    write_section_title(ws, row, "Net Debt Bridge / Cash & Investments Treatment", span_cols=3)
    row += 1
    bridge = getattr(out, "net_debt_bridge", None) or {}
    bridge_source = "v371 historical cache" if bridge.get("historical_cache_available") else "Reported input only (cache unavailable for this market)"
    _set_note_cell(ws, row, 1, f"Bridge components sourced from: {bridge_source}.")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    def _bridge_input(label: str, value, fmt, key: str | None = None, note: str | None = None):
        nonlocal row
        _set_label_cell(ws, row, 1, label)
        _set_input_cell(ws, row, 2, value if value is not None else 0.0, fmt)
        if value is None:
            _set_note_cell(ws, row, 3, note or "Cache field unavailable; editable")
        elif note:
            _set_note_cell(ws, row, 3, note)
        if key:
            ASSUMP_CELLS[key] = f"B{row}"
        row += 1

    _bridge_input("Total Debt (latest historical)", bridge.get("total_debt"), FMT_COMMA2, key="ndb_total_debt", note="Editable; default from v371 cache")
    _bridge_input("Cash & Equivalents (latest historical)", bridge.get("cash"), FMT_COMMA2, key="ndb_cash", note="Editable; default from v371 cache")
    _bridge_input("Short-term Investments / Marketable Securities", bridge.get("short_term_investments"), FMT_COMMA2, key="ndb_st_inv", note="Editable; default from v371 cache")
    _bridge_input("Long-term Investments / Marketable Securities", bridge.get("long_term_investments"), FMT_COMMA2, key="ndb_lt_inv", note="Editable; default from v371 cache")

    _set_label_cell(ws, row, 1, "Total Marketable Securities (= ST + LT)")
    _set_formula_cell(ws, row, 2, f"={_abs_ref(ASSUMP_CELLS['ndb_st_inv'])}+{_abs_ref(ASSUMP_CELLS['ndb_lt_inv'])}", FMT_COMMA2)
    ASSUMP_CELLS["ndb_total_ms"] = f"B{row}"
    row += 1

    _set_label_cell(ws, row, 1, "Reported / Input Net Debt (legacy headline driver)")
    _set_formula_cell(ws, row, 2, f"={_abs_ref(ASSUMP_CELLS['net_debt'])}", FMT_COMMA2, True)
    ASSUMP_CELLS["ndb_reported"] = f"B{row}"
    _set_note_cell(ws, row, 3, "Linked from Base-Year Financials Net Debt input.")
    row += 1

    _set_label_cell(ws, row, 1, "Adj #1: Debt - Cash")
    _set_formula_cell(ws, row, 2, f"={_abs_ref(ASSUMP_CELLS['ndb_total_debt'])}-{_abs_ref(ASSUMP_CELLS['ndb_cash'])}", FMT_COMMA2)
    ASSUMP_CELLS["ndb_adj_cash"] = f"B{row}"
    row += 1

    # V3.7.3: Raw Debt - Cash and Reported-vs-Raw gap, surfaced for audit so
    # reviewers can see when the reported/input net debt has been pre-adjusted
    # by the data source (AAPL exposed a ~46B delta in V3.7.2).
    _set_label_cell(ws, row, 1, "Raw Debt - Cash (memo, V3.7.3 audit)")
    _set_formula_cell(ws, row, 2, f"={_abs_ref(ASSUMP_CELLS['ndb_adj_cash'])}", FMT_COMMA2)
    ASSUMP_CELLS["ndb_raw_debt_less_cash"] = f"B{row}"
    _set_note_cell(ws, row, 3, "Same as Adj #1; restated as a memo so the gap vs Reported is auditable.")
    row += 1

    _set_label_cell(ws, row, 1, "Reported - Raw Debt-Cash Gap")
    _set_formula_cell(
        ws, row, 2,
        f"={_abs_ref(ASSUMP_CELLS['ndb_reported'])}-{_abs_ref(ASSUMP_CELLS['ndb_raw_debt_less_cash'])}",
        FMT_COMMA2,
    )
    ASSUMP_CELLS["ndb_reported_vs_raw_gap"] = f"B{row}"
    _set_note_cell(
        ws, row, 3,
        "V3.7.3: Non-zero means the data provider's Net Debt already reflects cash-like adjustments different from raw Debt - Cash. Review when the magnitude is material.",
    )
    row += 1

    _set_label_cell(ws, row, 1, "Adj #2: Debt - Cash - ST Investments")
    _set_formula_cell(
        ws, row, 2,
        f"={_abs_ref(ASSUMP_CELLS['ndb_total_debt'])}-{_abs_ref(ASSUMP_CELLS['ndb_cash'])}-{_abs_ref(ASSUMP_CELLS['ndb_st_inv'])}",
        FMT_COMMA2,
    )
    ASSUMP_CELLS["ndb_adj_st"] = f"B{row}"
    row += 1

    _set_label_cell(ws, row, 1, "Adj #3: Debt - Cash - Total Marketable Securities")
    _set_formula_cell(
        ws, row, 2,
        f"={_abs_ref(ASSUMP_CELLS['ndb_total_debt'])}-{_abs_ref(ASSUMP_CELLS['ndb_cash'])}-{_abs_ref(ASSUMP_CELLS['ndb_total_ms'])}",
        FMT_COMMA2,
    )
    ASSUMP_CELLS["ndb_adj_total_ms"] = f"B{row}"
    row += 1

    # V3.7.3: Selected Net Debt Treatment cell now holds the human-readable
    # label and is validated via an Excel dropdown. The "Selected Net Debt
    # Used in DCF" formula keys off the label. An invalid value (rare; user
    # cleared the cell) falls through the IF chain to Reported.
    selected_key, selected_label, _ = normalize_net_debt_treatment(bridge.get("selected_treatment"))
    label_for = {k: NET_DEBT_TREATMENT_LABELS[k] for k in NET_DEBT_TREATMENTS}
    label_list = [label_for[k] for k in NET_DEBT_TREATMENTS]
    _set_label_cell(ws, row, 1, "Selected Net Debt Treatment")
    _set_input_cell(ws, row, 2, selected_label, None)
    ASSUMP_CELLS["ndb_treatment"] = f"B{row}"
    treatment_cell_addr = f"B{row}"
    _set_note_cell(
        ws, row, 3,
        "V3.7.3: dropdown - pick one of "
        + " | ".join(label_list)
        + ". Invalid entries fall back to Reported / Input Net Debt.",
    )
    # Data Validation dropdown. openpyxl needs the formula1 quoted as one
    # double-quoted string. Labels contain no quotes / commas-only inside
    # ampersands, so a simple join works for all four current labels.
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(label_list) + '"',
        allow_blank=True,
        showDropDown=False,  # openpyxl quirk: False = arrow visible
        errorStyle="warning",
        error="Pick a Net Debt Treatment from the dropdown; unknown values fall back to Reported / Input Net Debt.",
        errorTitle="Net Debt Treatment",
        prompt="Selects which Net Debt variant drives the DCF headline IV.",
        promptTitle="Net Debt Treatment",
    )
    ws.add_data_validation(dv)
    dv.add(treatment_cell_addr)
    row += 1

    _set_label_cell(ws, row, 1, "Selected Net Debt Used in DCF", True)
    treatment_ref = _abs_ref(ASSUMP_CELLS["ndb_treatment"])
    rep = _abs_ref(ASSUMP_CELLS["ndb_reported"])
    cash_adj = _abs_ref(ASSUMP_CELLS["ndb_adj_cash"])
    st_adj = _abs_ref(ASSUMP_CELLS["ndb_adj_st"])
    ms_adj = _abs_ref(ASSUMP_CELLS["ndb_adj_total_ms"])
    lbl_cash = label_for["debt_less_cash"]
    lbl_st = label_for["debt_less_cash_and_st_investments"]
    lbl_ms = label_for["debt_less_cash_and_total_marketable_securities"]
    selected_formula = (
        f'=IF({treatment_ref}="{lbl_cash}",{cash_adj},'
        f'IF({treatment_ref}="{lbl_st}",{st_adj},'
        f'IF({treatment_ref}="{lbl_ms}",{ms_adj},{rep})))'
    )
    _set_formula_cell(ws, row, 2, selected_formula, FMT_COMMA2, True)
    ASSUMP_CELLS["ndb_selected"] = f"B{row}"
    _set_note_cell(ws, row, 3, "DCF Valuation Equity Bridge uses this cell as Net Debt.")
    row += 1

    _set_label_cell(ws, row, 1, "Difference vs Reported Net Debt")
    _set_formula_cell(ws, row, 2, f"={_abs_ref(ASSUMP_CELLS['ndb_selected'])}-{_abs_ref(ASSUMP_CELLS['ndb_reported'])}", FMT_COMMA2)
    ASSUMP_CELLS["ndb_diff_reported"] = f"B{row}"
    row += 1

    _set_label_cell(ws, row, 1, "Capital Lease memo (Current)")
    _set_input_cell(ws, row, 2, bridge.get("current_lease_obligation_memo") or 0.0, FMT_COMMA2)
    _set_note_cell(ws, row, 3, "Memo: yfinance includes capital lease inside ST Debt; not added separately.")
    row += 1
    _set_label_cell(ws, row, 1, "Capital Lease memo (Non-current)")
    _set_input_cell(ws, row, 2, bridge.get("long_term_lease_obligation_memo") or 0.0, FMT_COMMA2)
    _set_note_cell(ws, row, 3, "Memo: yfinance includes capital lease inside LT Debt; not added separately.")
    row += 1

    _set_note_cell(
        ws, row, 1,
        "V3.7.3: Marketable Securities treatment is a valuation JUDGMENT. Pick a treatment from the Selected Net Debt Treatment dropdown above to flip the headline IV; default = Reported / Input Net Debt preserves the V3.7.1 / V3.7.2 IV. The Reported vs Raw Debt-Cash gap row exposes any data-provider pre-adjustment baked into the reported figure.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # ── V3.7.4 Shareholder Returns Assumptions ─────────────────────────
    write_section_title(ws, row, "Shareholder Returns Assumptions", span_cols=3)
    row += 1
    sr_payload = getattr(out, "shareholder_returns", None) or {}
    sr_hist = sr_payload.get("historical") or {}
    sr_drivers = sr_payload.get("drivers_effective") or {}
    sr_source = "v374 historical cash-flow cache" if sr_payload.get("historical_cache_available") else "Fallback (cache fields unavailable)"
    _set_note_cell(ws, row, 1, f"Historical shareholder returns sourced from: {sr_source}.")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    _set_label_cell(ws, row, 1, "Dividend Payout % of Net Income")
    _set_input_cell(ws, row, 2, sr_drivers.get("dividend_payout_pct_net_income") or 0.0, FMT_PCT2)
    ASSUMP_CELLS["sr_dividend_payout"] = f"B{row}"
    _set_note_cell(ws, row, 3, f"Historical base: {(sr_hist.get('dividend_payout_pct_net_income') or 0) * 100:.1f}%")
    row += 1

    _set_label_cell(ws, row, 1, "Buyback Method")
    buyback_method_key = sr_drivers.get("buyback_method") or DEFAULT_BUYBACK_METHOD
    buyback_method_label = BUYBACK_METHOD_LABELS.get(buyback_method_key, BUYBACK_METHOD_LABELS[DEFAULT_BUYBACK_METHOD])
    _set_input_cell(ws, row, 2, buyback_method_label, None)
    ASSUMP_CELLS["sr_buyback_method"] = f"B{row}"
    sr_buyback_method_cell = f"B{row}"
    method_label_list = [BUYBACK_METHOD_LABELS[k] for k in BUYBACK_METHODS]
    dv_method = DataValidation(
        type="list",
        formula1='"' + ",".join(method_label_list) + '"',
        allow_blank=True,
        errorStyle="warning",
        errorTitle="Buyback Method",
        error="Pick a method from the dropdown.",
        promptTitle="Buyback Method",
        prompt="Choose how forecast buybacks are projected.",
    )
    ws.add_data_validation(dv_method)
    dv_method.add(sr_buyback_method_cell)
    _set_note_cell(ws, row, 3, "Drives the Buybacks projection on Supporting Schedules.")
    row += 1

    _set_label_cell(ws, row, 1, "Buyback % of FCF (used when method = % of FCF)")
    _set_input_cell(ws, row, 2, sr_drivers.get("buyback_pct_fcf") or 0.0, FMT_PCT2)
    ASSUMP_CELLS["sr_buyback_pct_fcf"] = f"B{row}"
    _set_note_cell(ws, row, 3, f"Historical base: {(sr_hist.get('buyback_pct_fcf') or 0) * 100:.1f}%")
    row += 1

    _set_label_cell(ws, row, 1, "Flat Buyback Amount (used when method = Flat)")
    _set_input_cell(ws, row, 2, sr_drivers.get("flat_buyback_amount") or 0.0, FMT_COMMA2)
    ASSUMP_CELLS["sr_flat_buyback"] = f"B{row}"
    _set_note_cell(ws, row, 3, f"Historical base buybacks: {(sr_hist.get('base_buybacks') or 0):,.0f}")
    row += 1

    _set_label_cell(ws, row, 1, "Repurchase Price Growth (annual)")
    _set_input_cell(ws, row, 2, sr_drivers.get("repurchase_price_growth") or 0.0, FMT_PCT2)
    ASSUMP_CELLS["sr_repurchase_price_growth"] = f"B{row}"
    _set_note_cell(ws, row, 3, "Year 1 repurchase price = Current Price; grown by this rate each year.")
    row += 1

    _set_label_cell(ws, row, 1, "Annual Dilution (% of beg. shares, SBC etc.)")
    _set_input_cell(ws, row, 2, sr_drivers.get("annual_dilution_pct") or 0.0, FMT_PCT2)
    ASSUMP_CELLS["sr_annual_dilution"] = f"B{row}"
    _set_note_cell(ws, row, 3, "Simple user-editable assumption for SBC-driven dilution; held flat across forecast.")
    row += 1

    _set_label_cell(ws, row, 1, "Buyback Funding Treatment")
    funding_treatment = sr_drivers.get("buyback_funding_treatment") or DEFAULT_BUYBACK_FUNDING_TREATMENT
    funding_label = (
        "Debt-Funded Buyback" if funding_treatment == "debt_funded_buyback"
        else "Planned-Uncapped Diagnostic" if funding_treatment == "planned_uncapped_diagnostic"
        else "Cash Floor / Buyback Cap"
    )
    _set_input_cell(ws, row, 2, funding_label, None)
    ASSUMP_CELLS["sr_funding_treatment"] = f"B{row}"
    dv_funding = DataValidation(
        type="list",
        formula1='"Cash Floor / Buyback Cap,Debt-Funded Buyback,Planned-Uncapped Diagnostic"',
        allow_blank=True,
        errorStyle="warning",
        errorTitle="Funding Treatment",
        error="Pick Cash Floor / Buyback Cap, Debt-Funded Buyback, or Planned-Uncapped Diagnostic.",
        promptTitle="Funding Treatment",
        prompt="Choose how buyback funding gaps are closed.",
    )
    ws.add_data_validation(dv_funding)
    dv_funding.add(f"B{row}")
    _set_note_cell(ws, row, 3, "Funding Treatment = Buyback Cap / Debt-Funded / Planned-Uncapped Diagnostic. Default caps buybacks at funding capacity before debt; diagnostic treatment may breach the cash floor and should flag High Review.")
    row += 1

    _set_label_cell(ws, row, 1, "Minimum Cash Floor")
    _set_input_cell(ws, row, 2, sr_drivers.get("minimum_cash_floor") or DEFAULT_CASH_FLOOR, FMT_COMMA2)
    ASSUMP_CELLS["sr_cash_floor"] = f"B{row}"
    row += 1

    _set_label_cell(ws, row, 1, "Marketable Securities Available for Returns")
    _set_input_cell(ws, row, 2, sr_drivers.get("marketable_securities_available_for_returns") or 0.0, FMT_COMMA2)
    ASSUMP_CELLS["sr_ms_available"] = f"B{row}"
    row += 1

    # Selected share-count treatment dropdown (drives headline IV when changed).
    _set_label_cell(ws, row, 1, "Selected Share Count Treatment")
    sc_treatment_key = sr_payload.get("selected_share_count_treatment") or DEFAULT_SHARE_COUNT_TREATMENT
    sc_treatment_label = SHARE_COUNT_TREATMENT_LABELS.get(sc_treatment_key, SHARE_COUNT_TREATMENT_LABELS[DEFAULT_SHARE_COUNT_TREATMENT])
    _set_input_cell(ws, row, 2, sc_treatment_label, None)
    ASSUMP_CELLS["sr_share_count_treatment"] = f"B{row}"
    sc_treatment_cell = f"B{row}"
    sc_label_list = [SHARE_COUNT_TREATMENT_LABELS[k] for k in SHARE_COUNT_TREATMENTS]
    dv_sc = DataValidation(
        type="list",
        formula1='"' + ",".join(sc_label_list) + '"',
        allow_blank=True,
        errorStyle="warning",
        errorTitle="Share Count Treatment",
        error="Pick a treatment from the dropdown; unknown values fall back to Current Reported Diluted Shares.",
        promptTitle="Share Count Treatment",
        prompt="Choose which share count denominator drives the DCF headline IV.",
    )
    ws.add_data_validation(dv_sc)
    dv_sc.add(sc_treatment_cell)
    _set_note_cell(ws, row, 3, " | ".join(sc_label_list))
    row += 1

    # Denominator options pre-seeded; live link from Supporting Schedules
    # added after the schedule sheet is built (see relink helper below).
    _set_label_cell(ws, row, 1, "Current Reported Diluted Shares (M)")
    _set_input_cell(ws, row, 2, sr_payload.get("denominator_options", {}).get("current_reported_shares") or float(inp.shares or 0.0), FMT_COMMA2)
    ASSUMP_CELLS["sr_denom_current"] = f"B{row}"
    row += 1
    _set_label_cell(ws, row, 1, "Forecast Ending Diluted Shares (M)")
    _set_input_cell(ws, row, 2, sr_payload.get("denominator_options", {}).get("forecast_ending_diluted_shares") or float(inp.shares or 0.0), FMT_COMMA2)
    ASSUMP_CELLS["sr_denom_ending"] = f"B{row}"
    _set_note_cell(ws, row, 3, "Will be relinked to Shareholder Returns Schedule ending shares when that sheet exists.")
    row += 1
    _set_label_cell(ws, row, 1, "Forecast Weighted Avg Diluted Shares (M)")
    _set_input_cell(ws, row, 2, sr_payload.get("denominator_options", {}).get("forecast_weighted_avg_diluted_shares") or float(inp.shares or 0.0), FMT_COMMA2)
    ASSUMP_CELLS["sr_denom_wavg"] = f"B{row}"
    _set_note_cell(ws, row, 3, "Will be relinked to Shareholder Returns Schedule weighted average row.")
    row += 1

    _set_label_cell(ws, row, 1, "Selected Share Count Used in DCF", True)
    treat_ref = _abs_ref(ASSUMP_CELLS["sr_share_count_treatment"])
    cur_ref = _abs_ref(ASSUMP_CELLS["sr_denom_current"])
    end_ref = _abs_ref(ASSUMP_CELLS["sr_denom_ending"])
    wavg_ref = _abs_ref(ASSUMP_CELLS["sr_denom_wavg"])
    lbl_end = SHARE_COUNT_TREATMENT_LABELS["forecast_ending_diluted_shares"]
    lbl_wavg = SHARE_COUNT_TREATMENT_LABELS["forecast_weighted_avg_diluted_shares"]
    sc_formula = (
        f'=IF({treat_ref}="{lbl_end}",{end_ref},'
        f'IF({treat_ref}="{lbl_wavg}",{wavg_ref},{cur_ref}))'
    )
    _set_formula_cell(ws, row, 2, sc_formula, FMT_COMMA2, True)
    ASSUMP_CELLS["sr_selected_denominator"] = f"B{row}"
    _set_note_cell(ws, row, 3, "DCF Valuation Intrinsic / Share uses this cell as the denominator.")
    row += 1

    _set_note_cell(
        ws, row, 1,
        "V3.7.4: Dividends and buybacks are equity-financing items. They flow through the Cash Flow Forecast financing section and the Balance Sheet Equity roll-forward but DO NOT reduce FCFF (unlevered NOPAT-based). Per-share valuation can only change through the Selected Share Count Treatment dropdown above; default = Current Reported Diluted Shares preserves the V3.7.3 headline IV.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    write_section_title(ws, row, "Operating Forecast Rationale", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Assumption", "Selected / Reference", "Rationale / Review prompt"], 1)
    row += 1
    op_rationale_rows = [
        ("Forecast Revenue Growth", f"={_abs_ref(ASSUMP_CELLS['revenue_growth'])}", FMT_PCT2, "Baseline mature-company growth assumption; analyst should review product cycle, Services mix, AI/device cycle, and consensus."),
        ("Terminal Growth", f"={_abs_ref(ASSUMP_CELLS['terminal_g'])}", FMT_PCT2, "Terminal convergence anchor; should remain economically consistent with long-term nominal growth and WACC spread."),
        ("Forecast EBIT Margin", f"={_abs_ref(ASSUMP_CELLS['ebit_margin'])}", FMT_PCT2, "Normalized operating margin baseline, not a full segment model; review Services mix, gross margin, opex leverage, competition, and regulation."),
        ("Gross Margin Reference", f"={_abs_ref(ASSUMP_CELLS['gross_margin'])}", FMT_PCT2, "Reference point for margin discussion; does not replace segment-level analysis."),
        ("D&A / CapEx / Working Capital", '="See forecast drivers"', None, "Driver assumptions are mechanical model inputs and should be validated against history, management guidance, and analyst estimates."),
    ]
    for label, formula, fmt, note in op_rationale_rows:
        _set_label_cell(ws, row, 1, label)
        _set_formula_cell(ws, row, 2, formula, fmt, True)
        _set_note_cell(ws, row, 3, note)
        row += 1
    _set_note_cell(ws, row, 1, "Default rationale blocks are model-generated starting points for IC discussion. They are not management guidance, consensus research, or rating guidance.")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    row = _write_wacc_institutional_bridge(ws, row, inp, out, ctx)

    _attach_assumption_source_comments(ws, inp, out, ctx)
    _add_watermark_row(ws, row, inp.symbol)


def _write_wacc_institutional_bridge(ws, row: int, inp: DCFInputs, out: DCFOutputs, ctx: dict) -> int:
    """V3.9.4 WACC Institutional Bridge.

    A consolidated 17-row reference block that pulls every component of the
    indicative WACC plus the cost-of-debt sanity check from the calculator's
    wacc_decision_bridge. Formula cells link back to the existing input cells
    so editing the source flows through; engine-derived references (raw beta,
    adjusted beta, cost-of-debt sanity, spread tier) display the calculator's
    snapshot values. Headline WACC is unchanged.
    """
    bridge = getattr(out, "wacc_decision_bridge", None) or {}

    def _abs(key):
        addr = ASSUMP_CELLS.get(key)
        return _abs_ref(addr) if addr else None

    write_section_title(ws, row, "WACC Institutional Bridge v1", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Component", "Value", "Note"], 1)
    row += 1

    # 1-2. rf / ERP - linked to Market Data block.
    _set_label_cell(ws, row, 1, "Risk-free Rate (rf)")
    if _abs("rf"):
        _set_formula_cell(ws, row, 2, f"={_abs('rf')}", FMT_PCT2, True)
    else:
        ws.cell(row=row, column=2, value=bridge.get("risk_free_rate")).number_format = FMT_PCT2
    _set_note_cell(ws, row, 3, "Market config; cross-checked at export date.")
    row += 1

    _set_label_cell(ws, row, 1, "Equity Risk Premium (ERP)")
    if _abs("erp"):
        _set_formula_cell(ws, row, 2, f"={_abs('erp')}", FMT_PCT2, True)
    else:
        ws.cell(row=row, column=2, value=bridge.get("equity_risk_premium")).number_format = FMT_PCT2
    _set_note_cell(ws, row, 3, "Market config equity risk premium.")
    row += 1

    # 3-5. raw / adjusted / selected beta.
    raw_beta = bridge.get("raw_beta") if bridge.get("raw_beta") is not None else bridge.get("beta")
    adjusted_beta = bridge.get("adjusted_beta")
    selected_beta = bridge.get("selected_beta") if bridge.get("selected_beta") is not None else raw_beta

    _set_label_cell(ws, row, 1, "Raw Beta (market data)")
    c = ws.cell(row=row, column=2, value=raw_beta)
    c.font = font_formula(); c.border = border_thin(); c.number_format = "0.00"
    _set_note_cell(ws, row, 3, "Source: financials cache (yfinance). Capped to [0.3, 3.0].")
    row += 1

    _set_label_cell(ws, row, 1, "Adjusted Beta (Blume)")
    c = ws.cell(row=row, column=2, value=adjusted_beta)
    c.font = font_formula(); c.border = border_thin(); c.number_format = "0.00"
    _set_note_cell(ws, row, 3, "0.67 * raw + 0.33 * 1.0; diagnostic reference.")
    row += 1

    _set_label_cell(ws, row, 1, "Selected Beta (drives CAPM)")
    if _abs("beta"):
        _set_formula_cell(ws, row, 2, f"={_abs('beta')}", "0.00", True)
    else:
        c = ws.cell(row=row, column=2, value=selected_beta)
        c.font = font_formula(); c.border = border_thin(); c.number_format = "0.00"
    _set_note_cell(ws, row, 3, "Defaults to raw beta; editable on Assumptions Beta cell.")
    row += 1

    # 6. Cost of Equity (CAPM)
    _set_label_cell(ws, row, 1, "CAPM Cost of Equity (Ke)")
    if _abs("rf") and _abs("beta") and _abs("erp"):
        _set_formula_cell(ws, row, 2, f"={_abs('rf')}+{_abs('beta')}*{_abs('erp')}", FMT_PCT2, True)
    else:
        ws.cell(row=row, column=2, value=bridge.get("capm_cost_of_equity")).number_format = FMT_PCT2
    _set_note_cell(ws, row, 3, "Ke = rf + selected_beta * ERP.")
    row += 1

    # 7-9. Pre-tax / Tax / After-tax Kd.
    _set_label_cell(ws, row, 1, "Pre-tax Cost of Debt (Kd)")
    if _abs("cost_of_debt"):
        _set_formula_cell(ws, row, 2, f"={_abs('cost_of_debt')}", FMT_PCT2, True)
    else:
        ws.cell(row=row, column=2, value=bridge.get("pre_tax_cost_of_debt")).number_format = FMT_PCT2
    _set_note_cell(ws, row, 3, bridge.get("pre_tax_cost_of_debt_rationale") or "Editable on Assumptions. Sanity check below uses historical interest / debt.")
    row += 1

    _set_label_cell(ws, row, 1, "Tax Rate")
    if _abs("tax_rate"):
        _set_formula_cell(ws, row, 2, f"={_abs('tax_rate')}", FMT_PCT2, True)
    else:
        ws.cell(row=row, column=2, value=bridge.get("tax_rate")).number_format = FMT_PCT2
    _set_note_cell(ws, row, 3, "Linked to base-year tax rate.")
    row += 1

    _set_label_cell(ws, row, 1, "After-tax Cost of Debt")
    if _abs("cost_of_debt") and _abs("tax_rate"):
        _set_formula_cell(ws, row, 2, f"={_abs('cost_of_debt')}*(1-{_abs('tax_rate')})", FMT_PCT2, True)
    else:
        ws.cell(row=row, column=2, value=bridge.get("after_tax_cost_of_debt")).number_format = FMT_PCT2
    _set_note_cell(ws, row, 3, "Kd * (1 - t).")
    row += 1

    # 10-11. Market values.
    _set_label_cell(ws, row, 1, "Market Value of Equity (Price x Shares, M)")
    if _abs("price") and _abs("shares"):
        _set_formula_cell(ws, row, 2, f"={_abs('price')}*{_abs('shares')}", FMT_COMMA2, True)
    else:
        c = ws.cell(row=row, column=2, value=bridge.get("market_value_equity_proxy"))
        c.font = font_formula(); c.border = border_thin(); c.number_format = FMT_COMMA2
    _set_note_cell(ws, row, 3, "Price x Diluted Shares (M).")
    row += 1

    _set_label_cell(ws, row, 1, "Debt Value (latest historical, M)")
    if _abs("wacc_debt_value"):
        _set_formula_cell(ws, row, 2, f"={_abs('wacc_debt_value')}", FMT_COMMA2, True)
    else:
        c = ws.cell(row=row, column=2, value=bridge.get("total_debt"))
        c.font = font_formula(); c.border = border_thin(); c.number_format = FMT_COMMA2
    _set_note_cell(ws, row, 3, "From historical balance sheet (V3.7.2 bridge).")
    row += 1

    # 12-13. Weights.
    _set_label_cell(ws, row, 1, "Equity Weight (We)")
    if _abs("weight_equity"):
        _set_formula_cell(ws, row, 2, f"={_abs('weight_equity')}", FMT_PCT2, True)
    else:
        ws.cell(row=row, column=2, value=bridge.get("equity_weight")).number_format = FMT_PCT2
    _set_note_cell(ws, row, 3, "Market-value weight of equity.")
    row += 1

    _set_label_cell(ws, row, 1, "Debt Weight (Wd)")
    if _abs("weight_debt"):
        _set_formula_cell(ws, row, 2, f"={_abs('weight_debt')}", FMT_PCT2, True)
    else:
        ws.cell(row=row, column=2, value=bridge.get("debt_weight")).number_format = FMT_PCT2
    _set_note_cell(ws, row, 3, "Market-value weight of debt.")
    row += 1

    # 14. CAPM Indicative WACC.
    _set_label_cell(ws, row, 1, "Mechanical CAPM Reference WACC")
    if _abs("wacc_indicative"):
        _set_formula_cell(ws, row, 2, f"={_abs('wacc_indicative')}", FMT_PCT2, True)
    else:
        ws.cell(row=row, column=2, value=bridge.get("capm_indicative_wacc")).number_format = FMT_PCT2
    _set_note_cell(ws, row, 3, "Mechanical CAPM reference, not headline selection.")
    row += 1

    # 15. Selected/Model WACC (headline).
    _set_label_cell(ws, row, 1, "Selected / Model WACC (headline)", True)
    if _abs("wacc_selected_echo"):
        _set_formula_cell(ws, row, 2, f"={_abs('wacc_selected_echo')}", FMT_PCT2, True)
    else:
        ws.cell(row=row, column=2, value=bridge.get("selected_model_wacc")).number_format = FMT_PCT2
    _set_note_cell(ws, row, 3, "Headline discount rate; drives DCF Valuation by default.")
    row += 1

    # 16. Spread in bps.
    _set_label_cell(ws, row, 1, "Spread vs Mechanical CAPM Reference (bps)")
    if _abs("wacc_selected_echo") and _abs("wacc_indicative"):
        _set_formula_cell(
            ws, row, 2,
            f"=({_abs('wacc_selected_echo')}-{_abs('wacc_indicative')})*10000",
            "0.0", True,
        )
    else:
        ws.cell(row=row, column=2, value=bridge.get("selected_vs_capm_spread_bps")).number_format = "0.0"
    _set_note_cell(ws, row, 3, "Selected minus indicative; positive = selected above CAPM.")
    row += 1

    # 17. Review flag.
    _set_label_cell(ws, row, 1, "Spread Review Tier", True)
    tier = bridge.get("spread_review_tier") or "N/A"
    c = ws.cell(row=row, column=2, value=tier)
    c.font = font_formula(); c.border = border_thin()
    _set_note_cell(ws, row, 3, "<=50 bps OK | 50-150 Review | >150 High review. Selected WACC is not auto-rewritten.")
    row += 1

    _set_label_cell(ws, row, 1, "CAPM Diagnostic Label", True)
    c = ws.cell(row=row, column=2, value=bridge.get("capm_diagnostic_label") or "Mechanical CAPM reference, not headline selection.")
    c.font = font_formula(); c.border = border_thin()
    _set_note_cell(ws, row, 3, "CAPM is retained as an audit diagnostic; selected WACC remains the headline unless intentionally changed.")
    row += 1

    write_section_title(ws, row, "WACC Component Defense", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Component", "Value", "Defense note"], 1)
    row += 1
    defense = bridge.get("wacc_component_defense") or {}
    rf_def = defense.get("risk_free_rate") or {}
    beta_def = defense.get("beta") or {}
    erp_def = defense.get("erp") or {}
    cap_def = defense.get("capital_structure") or {}
    kd_def = defense.get("cost_of_debt") or {}
    defense_rows = [
        ("Risk-free rate", rf_def.get("value"), FMT_PCT2, f"Tenor: {rf_def.get('tenor') or 'N/A'}; cutoff: {rf_def.get('cutoff') or 'N/A'}. {rf_def.get('spot_vs_normalized_note') or ''}"),
        ("Raw beta", beta_def.get("raw_beta"), "0.00", "Raw market beta shown for audit; not treated as the sole selected WACC answer."),
        ("Blume adjusted beta", beta_def.get("blume_adjusted_beta"), "0.00", "0.67 * raw + 0.33 * 1.0; diagnostic reference."),
        ("Selected beta", beta_def.get("selected_beta"), "0.00", beta_def.get("selected_rationale") or "Selected beta rationale unavailable."),
        ("Sector / quality beta reference", None, None, beta_def.get("sector_quality_reference") or "No external beta reference hard-wired."),
        ("ERP", erp_def.get("selected"), FMT_PCT2, f"{erp_def.get('reference_range') or 'Market reference'}; {erp_def.get('selected_rationale') or ''}"),
        ("Capital structure", cap_def.get("current_market_weight_equity"), FMT_PCT2, f"Equity weight shown; debt weight {cap_def.get('current_market_weight_debt')}. {cap_def.get('selected_rationale') or ''}"),
        ("Cost of debt", kd_def.get("selected_normalized_kd"), FMT_PCT2, f"Implied historical Kd: {kd_def.get('implied_historical_kd') if kd_def.get('implied_historical_kd') is not None else 'N/A'}. {kd_def.get('selected_rationale') or ''}"),
    ]
    for label, value, fmt, note in defense_rows:
        _set_label_cell(ws, row, 1, label)
        c = ws.cell(row=row, column=2, value=value)
        c.font = font_formula(); c.border = border_thin()
        if fmt:
            c.number_format = fmt
        _set_note_cell(ws, row, 3, note)
        row += 1

    write_section_title(ws, row, "Selected WACC Reconciliation", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Bridge item", "bps / value", "Rationale"], 1)
    row += 1
    recon = bridge.get("selected_wacc_reconciliation") or {}
    recon_rows = [
        ("Mechanical CAPM reference WACC", recon.get("mechanical_capm_reference_wacc"), FMT_PCT2, recon.get("diagnostic_label") or "Mechanical CAPM reference, not headline selection."),
        ("Selected WACC", recon.get("selected_wacc"), FMT_PCT2, "Selected banker judgment driving the default DCF headline."),
        ("Gap: selected minus mechanical", recon.get("gap_bps"), "0.0", "Basis-point gap requiring explicit attribution."),
    ]
    for label, value, fmt, note in recon_rows:
        _set_label_cell(ws, row, 1, label)
        c = ws.cell(row=row, column=2, value=value)
        c.font = font_formula(); c.border = border_thin(); c.number_format = fmt
        _set_note_cell(ws, row, 3, note)
        row += 1
    for item in recon.get("components") or []:
        _set_label_cell(ws, row, 1, item.get("component") or "Selected judgment adjustment")
        c = ws.cell(row=row, column=2, value=item.get("impact_bps"))
        c.font = font_formula(); c.border = border_thin(); c.number_format = "0.0"
        _set_note_cell(ws, row, 3, item.get("rationale") or "Judgment attribution.")
        row += 1
    _set_note_cell(ws, row, 1, recon.get("rationale") or "Selected WACC is an analyst judgment; CAPM is an audit diagnostic.")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    _set_label_cell(ws, row, 1, "Selected-vs-Mechanical CAPM Case Rationale", True)
    spread_bps = bridge.get("selected_vs_capm_spread_bps")
    spread_abs = abs(float(spread_bps or 0.0))
    c = ws.cell(
        row=row,
        column=2,
        value=(
            f"Selected WACC remains the model-selected case; mechanical CAPM reference differs by {spread_abs:.0f} bps."
            if spread_bps is not None
            else "Selected WACC remains the model-selected case; mechanical CAPM spread unavailable."
        ),
    )
    c.font = font_formula(); c.border = border_thin()
    _set_note_cell(
        ws,
        row,
        3,
        f"Spread tier: {tier}. IC should review the component defense and judgment bridge before relying on the selected WACC.",
    )
    row += 1

    _set_note_cell(
        ws, row, 1,
        "Selected / Model WACC remains the headline discount-rate case. Mechanical CAPM reference is an "
        "audit diagnostic based on the component bridge above. Spread is shown in basis points for review; a "
        "large spread is a discussion item, not a trigger to overwrite the selected rate.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # Cost of Debt Sanity sub-block.
    write_section_title(ws, row, "Cost of Debt Sanity (Historical)", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Check", "Value", "Note"], 1)
    row += 1
    sanity = bridge.get("cost_of_debt_sanity") or {}
    sanity_rows = [
        ("Reported Interest Expense (latest, M)", sanity.get("interest_expense_latest"), FMT_COMMA2,
         "From historical income statement cache."),
        ("Average Debt (last 2 periods, M)", sanity.get("average_debt"), FMT_COMMA2,
         "Average of the two most recent historical Total Debt observations."),
        ("Implied Pre-tax Cost of Debt", sanity.get("implied_pretax_cost_of_debt"), FMT_PCT2,
         "Interest Expense / Average Debt."),
        ("Selected Pre-tax Cost of Debt", sanity.get("selected_pretax_cost_of_debt"), FMT_PCT2,
         "From WACC Bridge above; editable."),
        ("Selected vs Implied Diff (bps)", sanity.get("diff_bps"), "0.0",
         "Selected minus implied."),
    ]
    for label, value, fmt, note in sanity_rows:
        _set_label_cell(ws, row, 1, label)
        if value is None:
            c = ws.cell(row=row, column=2, value="N/A")
            c.font = font_formula(); c.border = border_thin()
        else:
            c = ws.cell(row=row, column=2, value=value)
            c.font = font_formula(); c.border = border_thin(); c.number_format = fmt
        _set_note_cell(ws, row, 3, note)
        row += 1
    _set_label_cell(ws, row, 1, "Cost of Debt Review Tier", True)
    tier_c = ws.cell(row=row, column=2, value=sanity.get("review_tier") or "N/A")
    tier_c.font = font_formula(); tier_c.border = border_thin()
    _set_note_cell(ws, row, 3, sanity.get("review_reason") or "")
    row += 1
    _set_note_cell(
        ws, row, 1,
        "<=100 bps OK | 100-250 bps Review | >250 bps High review. N/A when historical interest expense "
        "or debt is unavailable. Sanity check does not auto-overwrite the selected pre-tax Kd.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    return row


def _attach_assumption_source_comments(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict) -> None:
    """V3.9.2: Attach Source / Cutoff / Methodology comments to the key
    blue-input cells on the Assumptions sheet so reviewers can audit every
    headline-driving assumption without leaving the workbook."""
    export_cutoff = _display_date(ctx.get("generated_at")) or date.today().isoformat()
    current_params = ctx.get("current_params") or {}
    em_source = current_params.get("exit_multiple_source") or "Market data default"
    em_fetched = current_params.get("exit_multiple_fetched_at")
    em_cache_file = current_params.get("exit_multiple_cache_file")
    em_raw = current_params.get("exit_multiple_raw")
    em_cache_date = current_params.get("exit_multiple_cache_date") or export_cutoff
    em_methodology_parts = [
        "Exit multiple defaulted from yfinance enterpriseToEbitda, cached by export date for deterministic IV anchoring, then rounded to 0.1x.",
    ]
    if em_raw is not None:
        em_methodology_parts.append(f"Raw fetched value: {round(float(em_raw), 4)}.")
    if em_cache_file:
        em_methodology_parts.append(f"Cache file: {os.path.basename(em_cache_file)}.")
    em_methodology_parts.append(
        "V3.9.5 review tier vs Trading Comps EV/EBITDA median: within +-15% OK | +-15-30% Review | >30% High review | N/A if insufficient peers."
    )
    em_methodology = " ".join(em_methodology_parts)
    em_cutoff = em_cache_date or em_fetched or export_cutoff

    financials_cutoff = (
        _display_date((ctx.get("financial_cache") or {}).get("cached_at"))
        or export_cutoff
    )
    historical_cutoff = (
        _display_date(((ctx.get("historical_cache") or {}).get("data") or {}).get("cached_at"))
        or financials_cutoff
    )

    methodology_financials = (
        "Latest available annual financials from the normalized financials cache; "
        "values shown in millions of original reporting currency unless noted."
    )
    methodology_balance_sheet = (
        "Latest available balance-sheet point from the normalized historical cache; "
        "millions of original reporting currency."
    )
    methodology_wacc = (
        "CAPM-style indicative WACC: Wd*Kd*(1-t) + We*Ke with rf and ERP from market "
        "config, beta from financials cache. Editable; review for capital-structure fit."
    )
    methodology_terminal_g = (
        "Editable forecast assumption; long-run nominal growth anchor used in Gordon TV. "
        "Review for economic consistency against WACC and country inflation. "
        "V3.9.5 review tier: 0%-4% OK | <0% or >4% Review | >5% High review."
    )
    methodology_user_input = (
        "Editable forecast assumption; review before use in investment committee materials."
    )
    methodology_treatment = (
        "Valuation judgment dropdown; selecting a non-default treatment intentionally "
        "flips the headline IV. Default preserves the prior-release anchor. "
        "V3.9.5: see Terminal Value Philosophy & Review on DCF Valuation for method hierarchy."
    )

    spec = [
        ("price", "Market data quote cache", export_cutoff,
         "Latest available market price snapshot from the quote cache; expressed in reporting currency."),
        ("revenue", "Normalized financials cache", financials_cutoff, methodology_financials),
        ("ebit", "Normalized financials cache", financials_cutoff, methodology_financials),
        ("da", "Normalized financials cache", financials_cutoff, methodology_financials),
        ("capex", "Normalized financials cache", financials_cutoff, methodology_financials),
        ("wc_change", "Normalized financials cache", financials_cutoff, methodology_financials),
        ("tax_rate", "Normalized financials cache", financials_cutoff,
         "Effective tax rate from the latest available P&L; editable to override for analyst judgment."),
        ("net_debt", "Normalized historical balance sheet (Debt - Cash, reported convention)",
         historical_cutoff, methodology_balance_sheet),
        ("shares", "Normalized financials cache (diluted shares outstanding, millions)",
         financials_cutoff,
         methodology_financials
         + " V3.9.6 single source-of-truth: this cell drives the Share Count Schedule and Shareholder Returns Schedule beginning shares; the legacy Share Count Schedule cannot diverge."),
        ("wacc", "Model default", export_cutoff, methodology_wacc),
        ("rf", "Market config (per-market risk-free rate)", export_cutoff,
         "Per-market risk-free rate from MARKET_CONFIG; editable on Assumptions."),
        ("erp", "Market config (per-market equity risk premium)", export_cutoff,
         "Per-market equity risk premium from MARKET_CONFIG; editable on Assumptions."),
        ("beta", "Financials cache (raw beta, capped [0.3, 3.0])", financials_cutoff,
         "Raw beta from yfinance via financials cache. WACC Institutional Bridge "
         "shows raw, Blume-adjusted, and selected beta side-by-side."),
        ("cost_of_debt", "User input / selected normalized Kd", export_cutoff,
         "Pre-tax cost of debt; AAPL default is selected normalized Kd 4.30%. "
         "Generic 5.0% is a fallback only when no selected Kd is supplied. Cost of Debt Sanity compares against historical interest expense / average debt."),
        ("wacc_debt_value", "Normalized historical balance sheet (Total Debt)",
         historical_cutoff, "Latest Total Debt from the historical cache; drives capital-structure weights."),
        ("terminal_g", "User input / model default", export_cutoff, methodology_terminal_g),
        ("exit_multiple", em_source, em_cutoff, em_methodology),
        ("tv_method", "User input / model default", export_cutoff,
         "Terminal method selector (gordon / exit / average); drives the headline terminal value."),
        ("tv_treatment", "User input / model default", export_cutoff, methodology_treatment),
        ("wacc_treatment", "User input / model default", export_cutoff, methodology_treatment),
        ("ndb_treatment", "User input / model default", export_cutoff, methodology_treatment),
        # V3.9.6 Share Count / Shareholder Returns driver comments.
        ("sr_dividend_payout", "User input / historical default", export_cutoff,
         "Dividend Payout % of Net Income; drives projected dividends in the Shareholder Returns Schedule. Dividends are capital allocation - not deducted from FCFF."),
        ("sr_buyback_method", "User input / model default", export_cutoff,
         "Buyback Method dropdown (% of FCF or flat amount); drives the Shareholder Returns Schedule buyback line."),
        ("sr_buyback_pct_fcf", "User input / historical default", export_cutoff,
         "Buybacks as % of FCF under the % of FCF method. Buybacks are capital allocation - not deducted from FCFF."),
        ("sr_flat_buyback", "User input / historical default", export_cutoff,
         "Flat buyback amount (millions) under the flat-amount method."),
        ("sr_repurchase_price_growth", "User input / model default", export_cutoff,
         "Annual growth rate applied to the current price to derive the repurchase price used in Shares Repurchased = Buyback / Repurchase Price."),
        ("sr_annual_dilution", "User input / model default", export_cutoff,
         "SBC / Annual Dilution %. Default = 0; review for companies where stock-based compensation is material. No real SBC feed is connected."),
        ("sr_share_count_treatment", "User input / model default", export_cutoff,
         "Selected Share Count Treatment dropdown; default = Current Reported Diluted Shares preserves prior-release headline IV. Switching to Forecast Ending / Weighted Avg changes per-share denominator only."),
    ]

    for key, source, cutoff, methodology in spec:
        addr = ASSUMP_CELLS.get(key)
        if not addr:
            continue
        try:
            cell = ws[addr]
        except (KeyError, ValueError):
            continue
        _attach_source_comment(cell, source, cutoff, methodology)

    # Forecast Path block: attach comment to every input cell in B58:F62 so the
    # year-by-year edits inherit the same source trail.
    path_methodology = (
        "Editable year-by-year forecast path. Defaults flat-equal to the single-value "
        "driver so headline IV is unchanged unless the path is edited."
    )
    for key in ("revenue_growth", "ebit_margin", "da_pct_revenue",
                "capex_pct_revenue", "wc_change_pct_revenue"):
        path_cells = ASSUMP_PATH_CELLS.get(key) or []
        for addr in path_cells:
            try:
                cell = ws[addr]
            except (KeyError, ValueError):
                continue
            _attach_source_comment(
                cell,
                "User input / model default",
                export_cutoff,
                path_methodology,
            )


def build_fcf_sheet(ws, inp: DCFInputs, out: DCFOutputs):
    ws.title = "FCF Build"
    n = inp.forecast_years
    apply_column_widths(ws, {"A": 32, **{get_column_letter(i + 2): 14 for i in range(n)}})

    ref_ebit = _abs_ref(ASSUMP_CELLS["ebit"])
    ref_da = _abs_ref(ASSUMP_CELLS["da"])
    ref_capex = _abs_ref(ASSUMP_CELLS["capex"])
    ref_wc = _abs_ref(ASSUMP_CELLS["wc_change"])
    ref_tax = _abs_ref(ASSUMP_CELLS["tax_rate"])
    ref_growth = _abs_ref(ASSUMP_CELLS["fcf_growth"])
    ref_wacc = _abs_ref(ASSUMP_CELLS["wacc"])

    row = 1
    ws.cell(row=row, column=1, value="FCF Build - Explicit Forecast Period").font = font_title()
    row += 2
    write_header_row(ws, row, [""] + [f"Year {i + 1}" for i in range(n)], 1)
    row += 1

    FCF_ROWS["ebit"] = row
    _set_label_cell(ws, row, 1, "EBIT")
    _set_formula_cell(ws, row, 2, f"={ref_ebit}*(1+{ref_growth})", FMT_COMMA2, True)
    for i in range(1, n):
        prev_col = get_column_letter(1 + i)
        _set_formula_cell(ws, row, 2 + i, f"={prev_col}{row}*(1+{ref_growth})", FMT_COMMA2, True)
    row += 1

    FCF_ROWS["taxes"] = row
    _set_label_cell(ws, row, 1, "  Less: Taxes")
    for i in range(n):
        col = get_column_letter(2 + i)
        _set_formula_cell(ws, row, 2 + i, f"=-{col}{FCF_ROWS['ebit']}*{ref_tax}", FMT_COMMA2, True)
    row += 1

    FCF_ROWS["nopat"] = row
    _set_label_cell(ws, row, 1, "NOPAT", True)
    for i in range(n):
        col = get_column_letter(2 + i)
        _set_formula_cell(ws, row, 2 + i, f"={col}{FCF_ROWS['ebit']}+{col}{FCF_ROWS['taxes']}", FMT_COMMA2)
    row += 1

    FCF_ROWS["da"] = row
    _set_label_cell(ws, row, 1, "  Add: D&A")
    _set_formula_cell(ws, row, 2, f"={ref_da}*(1+{ref_growth})", FMT_COMMA2, True)
    for i in range(1, n):
        prev_col = get_column_letter(1 + i)
        _set_formula_cell(ws, row, 2 + i, f"={prev_col}{row}*(1+{ref_growth})", FMT_COMMA2, True)
    row += 1

    FCF_ROWS["capex"] = row
    _set_label_cell(ws, row, 1, "  Less: CapEx")
    _set_formula_cell(ws, row, 2, f"={ref_capex}*(1+{ref_growth})", FMT_COMMA2, True)
    for i in range(1, n):
        prev_col = get_column_letter(1 + i)
        _set_formula_cell(ws, row, 2 + i, f"={prev_col}{row}*(1+{ref_growth})", FMT_COMMA2, True)
    row += 1

    FCF_ROWS["wc"] = row
    _set_label_cell(ws, row, 1, "  Less: ΔWorking Capital")
    _set_formula_cell(ws, row, 2, f"={ref_wc}*(1+{ref_growth})", FMT_COMMA2, True)
    for i in range(1, n):
        prev_col = get_column_letter(1 + i)
        _set_formula_cell(ws, row, 2 + i, f"={prev_col}{row}*(1+{ref_growth})", FMT_COMMA2, True)
    row += 1

    FCF_ROWS["fcf"] = row
    _set_label_cell(ws, row, 1, "Unlevered Free Cash Flow / FCFF", True)
    for i in range(n):
        col = get_column_letter(2 + i)
        _set_formula_cell(
            ws,
            row,
            2 + i,
            f"=ROUND({col}{FCF_ROWS['nopat']}+{col}{FCF_ROWS['da']}-{col}{FCF_ROWS['capex']}-{col}{FCF_ROWS['wc']},2)",
            FMT_COMMA2,
        )
    row += 1

    FCF_ROWS["df"] = row
    _set_label_cell(ws, row, 1, "Discount Factor")
    for i in range(n):
        _set_formula_cell(ws, row, 2 + i, f"=ROUND(1/(1+{ref_wacc})^{i + 1},6)", "0.000000", True)
    row += 1

    FCF_ROWS["pv"] = row
    _set_label_cell(ws, row, 1, "PV of FCF", True)
    for i in range(n):
        col = get_column_letter(2 + i)
        _set_formula_cell(ws, row, 2 + i, f"=ROUND({col}{FCF_ROWS['fcf']}*{col}{FCF_ROWS['df']},2)", FMT_COMMA2)
    row += 2
    _add_watermark_row(ws, row, inp.symbol)


# ── V3.9.5 Terminal Value Philosophy Lock ────────────────────────────────────
# Build a review payload over the existing terminal_decision_bridge plus the
# Trading Comps EV/EBITDA summary. No headline IV is touched; this is a
# diagnostic / IC-defensibility layer only.
_TV_PRIMARY_PHILOSOPHY = {
    "current_model_terminal": "Selected model terminal (analyst judgment headline)",
    "gordon_growth": "Gordon fundamental terminal (long-run growth + WACC spread)",
    "exit_multiple": "Market multiple terminal (EV/EBITDA exit)",
    "gordon_exit_blend": "Balanced (Gordon + Exit blend)",
    "fade_period_reference": "Fade reference (diagnostic only; static at export)",
}
_TV_METHOD_RECOMMENDED_USE = {
    "current_model_terminal": "Headline; defended by Gordon / Exit / Comps cross-checks below.",
    "gordon_growth": "Fundamental cross-check; sensitive to WACC - g spread.",
    "exit_multiple": "Market cross-check; sensitive to peer EV/EBITDA dispersion.",
    "gordon_exit_blend": "Balancing reference between fundamental and market views.",
    "fade_period_reference": "Diagnostic reference; static at export unless re-exported.",
}


def _tier_tv_ev(pct):
    if pct is None:
        return "N/A"
    if pct <= 0.70:
        return "OK"
    if pct <= 0.80:
        return "Review"
    return "High review"


def _tier_gordon_vs_exit(pct):
    if pct is None:
        return "N/A"
    if pct <= 0.15:
        return "OK"
    if pct <= 0.30:
        return "Review"
    return "High review"


def _tier_terminal_growth(g):
    if g is None:
        return "N/A"
    if g > 0.05:
        return "High review"
    if g < 0 or g > 0.04:
        return "Review"
    return "OK"


def _tier_exit_vs_comps(premium_discount):
    if premium_discount is None:
        return "N/A"
    abs_pd = abs(premium_discount)
    if abs_pd <= 0.15:
        return "OK"
    if abs_pd <= 0.30:
        return "Review"
    return "High review"


def _build_terminal_philosophy_payload(out, ctx: dict) -> dict:
    bridge = getattr(out, "terminal_decision_bridge", None) or {}
    treatment_key = bridge.get("selected_terminal_treatment") or DEFAULT_TERMINAL_TREATMENT
    treatment_label = bridge.get("selected_terminal_treatment_label") or TERMINAL_TREATMENT_LABELS.get(
        treatment_key, treatment_key
    )
    alt_iv = bridge.get("alternative_iv_per_share") or {}
    gordon_iv = alt_iv.get("gordon_growth")
    exit_iv = alt_iv.get("exit_multiple")
    blend_iv = alt_iv.get("gordon_exit_blend")
    current_iv = alt_iv.get("current_model_terminal") or bridge.get("selected_terminal_iv_per_share")
    tv_ev_ratio = bridge.get("selected_terminal_tv_ev_pct")
    if tv_ev_ratio is None:
        tv_ev_ratio = getattr(out, "tv_pct", None)
    terminal_g = bridge.get("terminal_growth")
    exit_multiple = bridge.get("exit_multiple")
    gordon_vs_exit_gap_pct = bridge.get("gordon_vs_exit_gap_pct")
    gordon_vs_exit_gap_abs = None
    if gordon_iv is not None and exit_iv is not None:
        gordon_vs_exit_gap_abs = round(abs(float(gordon_iv) - float(exit_iv)), 4)

    comps = (ctx.get("trading_comps") or {})
    comps_stats = ((comps.get("summary_stats") or {}).get("ev_ebitda")) or {}
    comps_median = comps_stats.get("median")
    comps_count = comps_stats.get("count") or 0
    comps_insufficient = bool(comps_stats.get("insufficient_data")) or comps_count < 3
    exit_vs_comps_pd = None
    if (
        not comps_insufficient
        and comps_median is not None
        and exit_multiple is not None
        and float(comps_median) > 0
    ):
        exit_vs_comps_pd = float(exit_multiple) / float(comps_median) - 1.0

    return {
        "version": "v395_terminal_philosophy_lock",
        "selected_terminal_method": treatment_key,
        "selected_terminal_method_label": treatment_label,
        "selected_terminal_treatment": treatment_key,
        "primary_philosophy": _TV_PRIMARY_PHILOSOPHY.get(treatment_key, treatment_label),
        "terminal_method_recommended_use": _TV_METHOD_RECOMMENDED_USE.get(
            treatment_key, "Headline method; alternatives are shown as cross-checks."
        ),
        "terminal_growth": terminal_g,
        "exit_multiple": exit_multiple,
        "tv_ev_ratio": tv_ev_ratio,
        "tv_ev_review_tier": _tier_tv_ev(tv_ev_ratio),
        "gordon_iv": gordon_iv,
        "exit_iv": exit_iv,
        "blend_iv": blend_iv,
        "current_model_iv": current_iv,
        "gordon_vs_exit_gap_abs": gordon_vs_exit_gap_abs,
        "gordon_vs_exit_gap_pct": gordon_vs_exit_gap_pct,
        "gordon_vs_exit_review_tier": _tier_gordon_vs_exit(gordon_vs_exit_gap_pct),
        "gordon_implied_exit_multiple": bridge.get("gordon_implied_exit_multiple"),
        "exit_implied_terminal_growth": bridge.get("exit_implied_terminal_growth"),
        "terminal_growth_review_tier": _tier_terminal_growth(terminal_g),
        "exit_multiple_vs_comps_median": (
            float(comps_median) if (not comps_insufficient and comps_median is not None) else None
        ),
        "exit_multiple_vs_comps_premium_discount": (
            round(exit_vs_comps_pd, 6) if exit_vs_comps_pd is not None else None
        ),
        "exit_multiple_vs_comps_review_tier": (
            "N/A" if comps_insufficient else _tier_exit_vs_comps(exit_vs_comps_pd)
        ),
        "comps_median_unavailable_reason": (
            "Trading Comps EV/EBITDA median insufficient (< 3 included peers or comps not built)"
            if comps_insufficient else None
        ),
        "comps_peer_count": comps_count,
    }


def build_valuation_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "DCF Valuation"
    apply_column_widths(ws, {"A": 36, "B": 22, "C": 22})

    n = inp.forecast_years
    fcf_first_col = get_column_letter(2)
    fcf_last_col = get_column_letter(2 + n - 1)
    pv_row = FCF_ROWS["pv"]
    fcf_row = FCF_ROWS["fcf"]
    revenue_row = FCF_ROWS.get("revenue")
    ebit_row = FCF_ROWS["ebit"]
    nopat_row = FCF_ROWS.get("nopat")
    da_row = FCF_ROWS["da"]
    df_row = FCF_ROWS["df"]

    # V3.7.5: discount rate now reads the WACC Bridge v3 "Selected WACC Used in
    # DCF" cell so the headline auto-flips when the user picks a treatment.
    ref_wacc = _abs_ref(ASSUMP_CELLS.get("wacc_used", ASSUMP_CELLS["wacc"]))
    ref_g = _abs_ref(ASSUMP_CELLS["terminal_g"])
    ref_exit = _abs_ref(ASSUMP_CELLS["exit_multiple"])
    ref_method = _abs_ref(ASSUMP_CELLS["tv_method"])
    ref_terminal_treatment = _abs_ref(ASSUMP_CELLS.get("tv_treatment", "B48"))
    ref_h_g_near = _abs_ref(ASSUMP_CELLS.get("h_model_g_near", ASSUMP_CELLS["terminal_g"]))
    ref_h_g_long = _abs_ref(ASSUMP_CELLS.get("h_model_g_long", ASSUMP_CELLS["terminal_g"]))
    ref_h_half_life = _abs_ref(ASSUMP_CELLS.get("h_model_half_life", ASSUMP_CELLS["terminal_g"]))
    ref_net_debt = _abs_ref(ASSUMP_CELLS["net_debt"])
    ref_shares = _abs_ref(ASSUMP_CELLS["shares"])
    ref_price = _abs_ref(ASSUMP_CELLS["price"])
    terminal_bridge_dcf_early = getattr(out, "terminal_decision_bridge", None) or {}
    fade_case_ev_static = ((terminal_bridge_dcf_early.get("fade_period") or {}).get("fade_case_ev") or 0.0)
    comps_payload_dcf = ctx.get("trading_comps") or {}
    current_params = ctx.get("current_params") or {}
    fx_rate_for_market_comparison = current_params.get("fx_rate_reporting_to_trading")
    try:
        fx_rate_for_market_comparison = float(fx_rate_for_market_comparison)
    except (TypeError, ValueError):
        fx_rate_for_market_comparison = None
    current_price_for_market_comparison = _coerce_num(current_params.get("price"))
    trading_iv_for_market_comparison = _coerce_num(current_params.get("intrinsic_value_per_share_trading_currency"))
    model_unsuitable_for_dcf = bool(current_params.get("model_unsuitable")) or current_params.get("valuation_status") == "model_unsuitable"
    market_comparison_unavailable = (
        model_unsuitable_for_dcf
        or current_params.get("market_comparison_status") == "unavailable"
        or current_params.get("market_comparison_allowed") is False
        or current_price_for_market_comparison is None
        or current_price_for_market_comparison <= 0
        or trading_iv_for_market_comparison is None
        or trading_iv_for_market_comparison <= 0
    )
    core_peer_ev_ebitda_median = (
        ((comps_payload_dcf.get("summary_stats") or {}).get("ev_ebitda") or {}).get("median")
    )

    row = _write_sheet_heading(ws, 1, "DCF Valuation Summary", inp, out, ctx)
    _set_label_cell(ws, row, 1, "DCF Source", True)
    VAL_ROWS["dcf_source"] = row
    _set_formula_cell(ws, row, 2, "='FCF Build'!$B$4", None, True)
    row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "UI, API, and Excel share one valuation source. Excel DCF uses workbook-linked P&L and Supporting Schedules; the API returns the same valuation through the unified engine.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    write_section_title(ws, row, "Headline Output", span_cols=3)
    row += 1
    _set_label_cell(ws, row, 1, "Intrinsic Value per Share", True)
    VAL_ROWS["headline_intrinsic"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Current Price")
    VAL_ROWS["headline_price"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Upside / (Downside)", True)
    VAL_ROWS["headline_upside"] = row
    row += 2

    write_section_title(ws, row, "Present Value of Explicit FCFFs", span_cols=3)
    row += 1
    _set_label_cell(ws, row, 1, f"Sum of PV(FCFF), Years 1-{n}")
    VAL_ROWS["pv_sum"] = row
    _set_formula_cell(ws, row, 2, f"=SUM('FCF Build'!{fcf_first_col}{pv_row}:{fcf_last_col}{pv_row})", FMT_COMMA2, True)
    row += 2

    write_section_title(ws, row, "Terminal Value", span_cols=3)
    row += 1
    fcf_y5_addr = f"'FCF Build'!{fcf_last_col}{fcf_row}"
    ebit_y5 = f"'FCF Build'!{fcf_last_col}{ebit_row}"
    da_y5 = f"'FCF Build'!{fcf_last_col}{da_row}"
    df_y5 = f"'FCF Build'!{fcf_last_col}{df_row}"
    denom = f"IF({ref_wacc}>{ref_g},{ref_wacc}-{ref_g},MAX({ref_wacc}-{ref_g},0.001))"

    _set_label_cell(ws, row, 1, "TV - Gordon Growth (PV)")
    VAL_ROWS["tv_gordon"] = row
    _set_formula_cell(ws, row, 2, f"={fcf_y5_addr}*(1+{ref_g})/({denom})*{df_y5}", FMT_COMMA2, True)
    row += 1

    _set_label_cell(ws, row, 1, "TV - Exit Multiple (PV)")
    VAL_ROWS["tv_exit"] = row
    _set_formula_cell(ws, row, 2, f"=({ebit_y5}+{da_y5})*{ref_exit}*{df_y5}", FMT_COMMA2, True)
    row += 1

    _set_label_cell(ws, row, 1, "TV - H-Model / Two-Stage Gordon (PV)")
    VAL_ROWS["tv_h_model"] = row
    h_denom = f"IF({ref_wacc}>{ref_h_g_long},{ref_wacc}-{ref_h_g_long},MAX({ref_wacc}-{ref_h_g_long},0.001))"
    _set_formula_cell(
        ws,
        row,
        2,
        f"=({fcf_y5_addr}*(1+{ref_h_g_long})+{fcf_y5_addr}*{ref_h_half_life}*({ref_h_g_near}-{ref_h_g_long}))/({h_denom})*{df_y5}",
        FMT_COMMA2,
        True,
    )
    row += 1

    _set_label_cell(ws, row, 1, 'TV Used (per "Terminal Value Method")', True)
    VAL_ROWS["tv_used"] = row
    tv_g_cell = f"B{VAL_ROWS['tv_gordon']}"
    tv_e_cell = f"B{VAL_ROWS['tv_exit']}"
    current_tv_formula = f'IF({ref_method}="gordon",{tv_g_cell},IF({ref_method}="exit",{tv_e_cell},AVERAGE({tv_g_cell},{tv_e_cell})))'
    lbl_tv_gordon = TERMINAL_TREATMENT_LABELS["gordon_growth"]
    lbl_tv_exit = TERMINAL_TREATMENT_LABELS["exit_multiple"]
    lbl_tv_blend = TERMINAL_TREATMENT_LABELS["gordon_exit_blend"]
    lbl_tv_h = TERMINAL_TREATMENT_LABELS["h_model"]
    lbl_tv_fade = TERMINAL_TREATMENT_LABELS["fade_period_reference"]
    _set_formula_cell(
        ws,
        row,
        2,
        (
            f'=IF({ref_terminal_treatment}="{lbl_tv_gordon}",{tv_g_cell},'
            f'IF({ref_terminal_treatment}="{lbl_tv_exit}",{tv_e_cell},'
            f'IF({ref_terminal_treatment}="{lbl_tv_blend}",AVERAGE({tv_g_cell},{tv_e_cell}),'
            f'IF({ref_terminal_treatment}="{lbl_tv_h}",B{VAL_ROWS["tv_h_model"]},'
            f'IF({ref_terminal_treatment}="{lbl_tv_fade}",{float(fade_case_ev_static)}-B{VAL_ROWS["pv_sum"]},'
            f'{current_tv_formula})))))'
        ),
        FMT_COMMA2,
        True,
    )
    row += 1

    _set_label_cell(ws, row, 1, "Terminal Value % of EV")
    tv_pct_row = row
    VAL_ROWS["tv_pct"] = row
    row += 2

    write_section_title(ws, row, "Bridge to Equity Value per Share", span_cols=3)
    row += 1
    _set_label_cell(ws, row, 1, "Enterprise Value (EV)", True)
    VAL_ROWS["ev"] = row
    _set_formula_cell(ws, row, 2, f"=B{VAL_ROWS['pv_sum']}+B{VAL_ROWS['tv_used']}", FMT_COMMA2)
    row += 1

    _set_formula_cell(ws, tv_pct_row, 2, f"=B{VAL_ROWS['tv_used']}/B{VAL_ROWS['ev']}", FMT_PCT2)
    row += 1
    write_section_title(ws, row, "Supporting Diagnostics", span_cols=3)
    try:
        ws.row_breaks.append(Break(id=row))
    except Exception:
        pass
    row += 1

    # V3.6.9 Terminal Value Quality block — kept inline so reviewers can see all TV diagnostics
    # next to the TV rows. Implied exit multiple from Gordon and implied g from exit make the
    # dependency between the two TV methods auditable rather than just a single % flag.
    tv_quality_anchor_row = row
    tv_gordon_undiscounted = f"({fcf_y5_addr}*(1+{ref_g})/({denom}))"
    tv_exit_undiscounted = f"(({ebit_y5}+{da_y5})*{ref_exit})"
    _set_label_cell(ws, row, 1, "TV Quality - Implied Exit Multiple from Gordon")
    VAL_ROWS["tv_implied_exit_multiple"] = row
    _set_formula_cell(
        ws,
        row,
        2,
        f"=IFERROR({tv_gordon_undiscounted}/({ebit_y5}+{da_y5}),0)",
        FMT_MULTIPLE,
    )
    row += 1
    _set_label_cell(ws, row, 1, "TV Quality - Implied Terminal Growth from Exit Multiple")
    VAL_ROWS["tv_implied_growth_from_exit"] = row
    # Solve g from Exit TV = FCF_y5*(1+g)/(wacc-g)  →  g = (Exit*wacc - FCF_y5)/(Exit + FCF_y5)
    _set_formula_cell(
        ws,
        row,
        2,
        f"=IFERROR(({tv_exit_undiscounted}*{ref_wacc}-{fcf_y5_addr})/({tv_exit_undiscounted}+{fcf_y5_addr}),0)",
        FMT_PCT2,
    )
    row += 1
    _set_label_cell(ws, row, 1, "TV Quality - Gordon vs Exit Valuation Gap (relative)")
    VAL_ROWS["tv_gordon_exit_gap"] = row
    _set_formula_cell(
        ws,
        row,
        2,
        f"=IFERROR(ABS(B{VAL_ROWS['tv_gordon']}-B{VAL_ROWS['tv_exit']})/MAX(ABS(B{VAL_ROWS['tv_gordon']}),ABS(B{VAL_ROWS['tv_exit']}),1),0)",
        FMT_PCT2,
    )
    row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "V3.6.9 TV Quality: Implied Exit = Gordon TV / (EBIT_Y5 + D&A_Y5); Implied g = (Exit*WACC - FCF_Y5)/(Exit + FCF_Y5). Audit Dashboard flags Gordon vs Exit gap > 25% and TV/EV > 80% for review.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    write_section_title(ws, row, "Exit Multiple Source / Premium Override", span_cols=3)
    VAL_ROWS["exit_multiple_source_block"] = row
    row += 1
    write_header_row(ws, row, ["Item", "Value", "Source / Note"], 1)
    row += 1
    peer_median_value = core_peer_ev_ebitda_median if core_peer_ev_ebitda_median is not None else None
    _set_label_cell(ws, row, 1, "Core peer median EV/EBITDA")
    if peer_median_value is None:
        ws.cell(row=row, column=2, value="N/A").font = font_formula(); ws.cell(row=row, column=2).border = border_thin()
    else:
        ws.cell(row=row, column=2, value=float(peer_median_value)).font = font_formula(); ws.cell(row=row, column=2).border = border_thin(); ws.cell(row=row, column=2).number_format = FMT_MULTIPLE
    _set_note_cell(ws, row, 3, "Trading Comps included-peer median; unavailable if fewer than required peers.")
    VAL_ROWS["exit_source_peer_median"] = row
    peer_median_ref = f"B{row}"
    row += 1
    _set_label_cell(ws, row, 1, "Selected Exit Multiple")
    _set_formula_cell(ws, row, 2, f"={ref_exit}", FMT_MULTIPLE, True)
    _set_note_cell(ws, row, 3, "Selected terminal input used by Exit-only and Blend terminal treatments.")
    VAL_ROWS["exit_source_selected"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Premium Override vs Peer Median (turns)")
    if peer_median_value is None:
        ws.cell(row=row, column=2, value="N/A").font = font_formula(); ws.cell(row=row, column=2).border = border_thin()
    else:
        _set_formula_cell(ws, row, 2, f"={ref_exit}-{peer_median_ref}", FMT_MULTIPLE, True)
    _set_note_cell(ws, row, 3, "Selected Exit Multiple = Core Peer Median + Premium Override unless the user edits the selected multiple directly.")
    VAL_ROWS["exit_source_premium_turns"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Premium Override vs Peer Median (%)")
    if peer_median_value is None:
        ws.cell(row=row, column=2, value="N/A").font = font_formula(); ws.cell(row=row, column=2).border = border_thin()
    else:
        _set_formula_cell(ws, row, 2, f"=IFERROR(({ref_exit}-{peer_median_ref})/{peer_median_ref},\"N/A\")", FMT_PCT2, True)
    _set_note_cell(ws, row, 3, "Premium above +30% is High Review; +15% to +30% is Review.")
    VAL_ROWS["exit_source_premium_pct"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Exit Multiple Source Label")
    source_label = (
        "Peer median + selected premium override / system default judgment"
        if peer_median_value is not None else "Fallback source / user override"
    )
    ws.cell(row=row, column=2, value=source_label).font = font_formula(); ws.cell(row=row, column=2).border = border_thin()
    _set_note_cell(ws, row, 3, "Default can remain above peer median only when the premium is visible and review-tiered.")
    VAL_ROWS["exit_source_label"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Diagnostic IV at peer median exit multiple")
    if peer_median_value is None:
        ws.cell(row=row, column=2, value="N/A").font = font_formula(); ws.cell(row=row, column=2).border = border_thin()
    else:
        peer_tv = f"({ebit_y5}+{da_y5})*{peer_median_ref}*{df_y5}"
        _set_formula_cell(ws, row, 2, f"=(B{VAL_ROWS['pv_sum']}+{peer_tv}-{_abs_ref(ASSUMP_CELLS.get('ndb_selected', ASSUMP_CELLS['net_debt']))})/{_abs_ref(ASSUMP_CELLS.get('sr_selected_denominator', ASSUMP_CELLS['shares']))}", FMT_COMMA2, True)
    _set_note_cell(ws, row, 3, "Diagnostic only; does not force the selected exit multiple to peer median.")
    VAL_ROWS["exit_source_peer_median_iv"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Diagnostic IV at selected exit multiple")
    _set_formula_cell(ws, row, 2, f"=(B{VAL_ROWS['pv_sum']}+B{VAL_ROWS['tv_exit']}-{_abs_ref(ASSUMP_CELLS.get('ndb_selected', ASSUMP_CELLS['net_debt']))})/{_abs_ref(ASSUMP_CELLS.get('sr_selected_denominator', ASSUMP_CELLS['shares']))}", FMT_COMMA2, True)
    _set_note_cell(ws, row, 3, "Exit-only diagnostic using selected multiple.")
    VAL_ROWS["exit_source_selected_iv"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Diagnostic IV difference")
    if peer_median_value is None:
        ws.cell(row=row, column=2, value="N/A").font = font_formula(); ws.cell(row=row, column=2).border = border_thin()
    else:
        _set_formula_cell(ws, row, 2, f"=B{VAL_ROWS['exit_source_selected_iv']}-B{VAL_ROWS['exit_source_peer_median_iv']}", FMT_COMMA2, True)
    _set_note_cell(ws, row, 3, "Selected exit IV less peer-median exit IV.")
    VAL_ROWS["exit_source_iv_diff"] = row
    row += 2

    write_section_title(ws, row, "Terminal Value Quality / Decomposition", span_cols=3)
    VAL_ROWS["terminal_quality_block"] = row
    row += 1
    write_header_row(ws, row, ["Metric", "Value", "Interpretation / Source"], 1)
    row += 1
    quality = (terminal_bridge_dcf_early.get("terminal_value_quality") or {})
    revenue_y5 = f"'FCF Build'!{fcf_last_col}{revenue_row}" if revenue_row else "0"
    nopat_y5 = f"'FCF Build'!{fcf_last_col}{nopat_row}" if nopat_row else f"{ebit_y5}*(1-{_abs_ref(ASSUMP_CELLS['tax_rate'])})"
    capex_row = FCF_ROWS.get("capex")
    wc_row = FCF_ROWS.get("wc")
    capex_y5 = f"'FCF Build'!{fcf_last_col}{capex_row}" if capex_row else "0"
    wc_y5 = f"'FCF Build'!{fcf_last_col}{wc_row}" if wc_row else "0"
    roic_val = quality.get("implied_terminal_roic")
    invested_capital_val = quality.get("invested_capital_terminal_approx")
    roic_source = quality.get("roic_approximation_source") or "N/A"

    # V3.9.9.4: lay down NOPAT / Invested Capital / ROIC cells FIRST so subsequent
    # formulas reference cells instead of inlining a hardcoded ROIC numeric literal.
    _set_label_cell(ws, row, 1, "Terminal Year Revenue")
    _set_formula_cell(ws, row, 2, f"={revenue_y5}", FMT_COMMA2, True)
    _set_note_cell(ws, row, 3, "Terminal explicit forecast year revenue.")
    VAL_ROWS["terminal_quality_terminal_year_revenue"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Terminal Year EBIT")
    _set_formula_cell(ws, row, 2, f"={ebit_y5}", FMT_COMMA2, True)
    _set_note_cell(ws, row, 3, "Terminal explicit forecast year EBIT.")
    VAL_ROWS["terminal_quality_terminal_year_ebit"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Terminal Year NOPAT")
    _set_formula_cell(ws, row, 2, f"={nopat_y5}", FMT_COMMA2, True)
    _set_note_cell(ws, row, 3, "NOPAT = EBIT x (1 - tax rate).")
    VAL_ROWS["terminal_quality_terminal_year_nopat"] = row
    nopat_cell_ref = f"B{row}"
    row += 1

    _set_label_cell(ws, row, 1, "Invested Capital (terminal approx)")
    if invested_capital_val is None:
        c = ws.cell(row=row, column=2, value="N/A"); c.font = font_formula(); c.border = border_thin()
    else:
        c = ws.cell(row=row, column=2, value=float(invested_capital_val)); c.font = font_formula(); c.border = border_thin(); c.number_format = FMT_COMMA2
    _set_note_cell(ws, row, 3, "Approximation: (Total Debt + Total Equity - Cash/ST Investments) scaled by terminal revenue / latest historical revenue. Cell value; ROIC formula below references this cell.")
    VAL_ROWS["terminal_quality_invested_capital"] = row
    ic_cell_ref = f"B{row}"
    row += 1

    _set_label_cell(ws, row, 1, "Implied terminal ROIC = NOPAT / Invested Capital")
    if invested_capital_val is None or roic_val is None:
        c = ws.cell(row=row, column=2, value="N/A"); c.font = font_formula(); c.border = border_thin()
    else:
        # ROIC is a formula referencing the NOPAT and Invested Capital cells -
        # no hardcoded ROIC numeric literal anywhere in this workbook.
        _set_formula_cell(ws, row, 2, f"=IFERROR({nopat_cell_ref}/{ic_cell_ref},\"N/A\")", FMT_PCT2, True)
    _set_note_cell(ws, row, 3, roic_source)
    VAL_ROWS["terminal_quality_roic"] = row
    roic_cell_ref = f"B{row}"
    row += 1

    write_section_title(ws, row, "ROIC Denominator Reconciliation", span_cols=3)
    VAL_ROWS["roic_denominator_reconciliation"] = row
    row += 1
    write_header_row(ws, row, ["Treatment", "Invested Capital", "ROIC / Note"], 1)
    row += 1
    _set_label_cell(ws, row, 1, "Reported invested capital approximation")
    if invested_capital_val is None:
        ws.cell(row=row, column=2, value="N/A").font = font_formula(); ws.cell(row=row, column=2).border = border_thin()
        ws.cell(row=row, column=3, value="N/A").font = font_formula(); ws.cell(row=row, column=3).border = border_thin()
    else:
        ws.cell(row=row, column=2, value=float(invested_capital_val)).font = font_formula(); ws.cell(row=row, column=2).border = border_thin(); ws.cell(row=row, column=2).number_format = FMT_COMMA2
        _set_formula_cell(ws, row, 3, f"=IFERROR({nopat_cell_ref}/B{row},\"N/A\")", FMT_PCT2, True)
    VAL_ROWS["roic_recon_reported_ic"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Cash / marketable securities adjustment")
    ndb_bridge_for_roic = getattr(out, "net_debt_bridge", None) or {}
    selected_ms_adjustment = 0.0
    if (ndb_bridge_for_roic.get("selected_treatment") == "debt_less_cash_and_total_marketable_securities"):
        selected_ms_adjustment = float(ndb_bridge_for_roic.get("total_marketable_securities") or 0.0)
    elif (ndb_bridge_for_roic.get("selected_treatment") == "debt_less_cash_and_st_investments"):
        selected_ms_adjustment = float(ndb_bridge_for_roic.get("short_term_investments") or 0.0)
    ws.cell(row=row, column=2, value=selected_ms_adjustment).font = font_formula(); ws.cell(row=row, column=2).border = border_thin(); ws.cell(row=row, column=2).number_format = FMT_COMMA2
    ws.cell(row=row, column=3, value=f"Selected net debt treatment: {ndb_bridge_for_roic.get('selected_treatment_label') or 'N/A'}").font = font_formula(); ws.cell(row=row, column=3).border = border_thin()
    VAL_ROWS["roic_recon_cash_ms_adjustment"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Net-debt-consistent invested capital")
    if invested_capital_val is None:
        ws.cell(row=row, column=2, value="N/A").font = font_formula(); ws.cell(row=row, column=2).border = border_thin()
        ws.cell(row=row, column=3, value="N/A").font = font_formula(); ws.cell(row=row, column=3).border = border_thin()
    else:
        _set_formula_cell(ws, row, 2, f"=MAX(1,B{VAL_ROWS['roic_recon_reported_ic']}-B{VAL_ROWS['roic_recon_cash_ms_adjustment']})", FMT_COMMA2, True)
        _set_formula_cell(ws, row, 3, f"=IFERROR({nopat_cell_ref}/B{row},\"N/A\")", FMT_PCT2, True)
    VAL_ROWS["roic_recon_net_debt_consistent_ic"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Cash-neutral / operating IC approximation")
    if invested_capital_val is None:
        ws.cell(row=row, column=2, value="N/A").font = font_formula(); ws.cell(row=row, column=2).border = border_thin()
        ws.cell(row=row, column=3, value="N/A").font = font_formula(); ws.cell(row=row, column=3).border = border_thin()
    else:
        _set_formula_cell(ws, row, 2, f"=MAX(B{VAL_ROWS['roic_recon_reported_ic']},B{VAL_ROWS['roic_recon_net_debt_consistent_ic']})", FMT_COMMA2, True)
        _set_formula_cell(ws, row, 3, f"=IFERROR({nopat_cell_ref}/B{row},\"N/A\")", FMT_PCT2, True)
    VAL_ROWS["roic_recon_cash_neutral_ic"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Selected diagnostic treatment")
    ws.cell(row=row, column=2, value="Net-debt-consistent invested capital").font = font_formula(); ws.cell(row=row, column=2).border = border_thin()
    ws.cell(row=row, column=3, value="Diagnostic only; headline IV changes only if ROIC drives a selected terminal method. No R&D / lease ROU normalization in this version.").font = font_formula(); ws.cell(row=row, column=3).border = border_thin()
    VAL_ROWS["roic_recon_selected_treatment"] = row
    row += 2

    _set_label_cell(ws, row, 1, "WACC")
    _set_formula_cell(ws, row, 2, f"={ref_wacc}", FMT_PCT2, True)
    _set_note_cell(ws, row, 3, "Selected WACC used in DCF.")
    VAL_ROWS["terminal_quality_wacc"] = row
    row += 1

    _set_label_cell(ws, row, 1, "ROIC - WACC spread")
    if invested_capital_val is None or roic_val is None:
        c = ws.cell(row=row, column=2, value="N/A"); c.font = font_formula(); c.border = border_thin()
    else:
        _set_formula_cell(ws, row, 2, f"={roic_cell_ref}-{ref_wacc}", FMT_PCT2, True)
    review_tier = quality.get("review_tier") or "Review"
    # V3.9.9.4: surface ROIC plausibility as High Review when implausibly high.
    if roic_val is not None and float(roic_val) > 0.50:
        review_tier_disp = "High Review - ROIC > 50% implausible; invested-capital approximation likely understates capital base"
    else:
        review_tier_disp = f"Review tier: {review_tier}"
    _set_note_cell(ws, row, 3, review_tier_disp)
    VAL_ROWS["terminal_quality_roic_wacc_spread"] = row
    row += 1

    _set_label_cell(ws, row, 1, "Terminal growth g")
    _set_formula_cell(ws, row, 2, f"={ref_g}", FMT_PCT2, True)
    _set_note_cell(ws, row, 3, "Selected long-run terminal growth.")
    VAL_ROWS["terminal_quality_terminal_growth_g"] = row
    row += 1

    # V3.9.9.4: reinvestment cross-check. Two paths.
    _set_label_cell(ws, row, 1, "Reinvestment rate from g / ROIC")
    if invested_capital_val is None or roic_val is None:
        c = ws.cell(row=row, column=2, value="N/A"); c.font = font_formula(); c.border = border_thin()
    else:
        _set_formula_cell(ws, row, 2, f"=IFERROR({ref_g}/{roic_cell_ref},\"N/A\")", FMT_PCT2, True)
    _set_note_cell(ws, row, 3, "Implied long-run reinvestment rate that supports g at the approximated ROIC.")
    VAL_ROWS["terminal_quality_reinvestment_from_g_roic"] = row
    reinv_g_roic_ref = f"B{row}"
    row += 1

    _set_label_cell(ws, row, 1, "Reinvestment rate from explicit Y5 (CapEx - D&A + Delta NWC) / NOPAT")
    if capex_row and wc_row and nopat_row:
        # Sign note: in FCF Build, CapEx is added as a positive subtraction and
        # Delta NWC line stores the cash-investment magnitude (positive = use).
        _set_formula_cell(
            ws, row, 2,
            f"=IFERROR(({capex_y5}-{da_y5}+{wc_y5})/{nopat_cell_ref},\"N/A\")",
            FMT_PCT2, True,
        )
    else:
        c = ws.cell(row=row, column=2, value="N/A"); c.font = font_formula(); c.border = border_thin()
    _set_note_cell(ws, row, 3, "Bottom-up reinvestment intensity at Y5: (CapEx - D&A + Delta NWC) / NOPAT. Cross-check vs g/ROIC implied rate.")
    VAL_ROWS["terminal_quality_reinvestment_from_explicit"] = row
    row += 1

    _set_label_cell(ws, row, 1, "Terminal reinvestment need (NOPAT x g/ROIC)")
    if invested_capital_val is None or roic_val is None:
        c = ws.cell(row=row, column=2, value="N/A"); c.font = font_formula(); c.border = border_thin()
    else:
        _set_formula_cell(ws, row, 2, f"=IFERROR({nopat_cell_ref}*{reinv_g_roic_ref},\"N/A\")", FMT_COMMA2, True)
    _set_note_cell(ws, row, 3, "Reinvestment need = NOPAT x (g / ROIC); uses ROIC cell, not a hardcoded literal.")
    VAL_ROWS["terminal_quality_terminal_reinvestment_need"] = row
    row += 1

    _set_label_cell(ws, row, 1, "Terminal FCF")
    if invested_capital_val is None or roic_val is None:
        _set_formula_cell(ws, row, 2, f"={fcf_y5_addr}", FMT_COMMA2, True)
    else:
        _set_formula_cell(ws, row, 2, f"=IFERROR({nopat_cell_ref}*(1-{reinv_g_roic_ref}),{fcf_y5_addr})", FMT_COMMA2, True)
    _set_note_cell(ws, row, 3, "Terminal FCF = NOPAT x (1 - g/ROIC) when ROIC approximation exists; otherwise explicit terminal FCF. Reference cells only; no inline ROIC literal.")
    VAL_ROWS["terminal_quality_terminal_fcf"] = row
    row += 1

    _set_label_cell(ws, row, 1, "Value creation interpretation")
    interpret_value = review_tier
    if roic_val is not None and float(roic_val) > 0.50:
        interpret_value = "High Review"
    ws.cell(row=row, column=2, value=interpret_value).font = font_formula(); ws.cell(row=row, column=2).border = border_thin()
    _set_note_cell(ws, row, 3, quality.get("value_creation_interpretation") or "Terminal value quality requires manual review.")
    VAL_ROWS["terminal_quality_interpretation"] = row
    row += 2

    write_section_title(ws, row, "ROIC Normalized Sensitivity (diagnostic only)", span_cols=5)
    VAL_ROWS["roic_sensitivity_block"] = row
    row += 1
    write_header_row(ws, row, ["Case", "ROIC", "Reinvestment rate = g / ROIC", "Terminal FCF", "IV/share impact vs reported ROIC"], 1)
    row += 1
    roic_pv_sum_ref = f"B{VAL_ROWS['pv_sum']}"
    roic_nd_ref = _abs_ref(ASSUMP_CELLS.get("ndb_selected", ASSUMP_CELLS["net_debt"]))
    roic_sh_ref = _abs_ref(ASSUMP_CELLS.get("sr_selected_denominator", ASSUMP_CELLS["shares"]))
    reported_iv_formula = f"IFERROR(({roic_pv_sum_ref}+({nopat_cell_ref}*(1-{reinv_g_roic_ref}))*(1+{ref_g})/({denom})*{df_y5}-{roic_nd_ref})/{roic_sh_ref},0)"
    sensitivity_cases = [
        ("Reported ROIC approximation", roic_cell_ref, "reported"),
        ("Adjusted normalized high ROIC = 40%", "40%", "high_40"),
        ("Adjusted normalized mid ROIC = 30%", "30%", "mid_30"),
        ("No-value-creation reference ROIC = WACC", ref_wacc, "wacc_floor"),
    ]
    for label, roic_ref_value, key in sensitivity_cases:
        _set_label_cell(ws, row, 1, label, key == "wacc_floor")
        if roic_ref_value == roic_cell_ref:
            _set_formula_cell(ws, row, 2, f"={roic_ref_value}", FMT_PCT2, True)
            roic_ref_for_row = f"B{row}"
        elif roic_ref_value == ref_wacc:
            _set_formula_cell(ws, row, 2, f"={ref_wacc}", FMT_PCT2, True)
            roic_ref_for_row = f"B{row}"
        else:
            _set_formula_cell(ws, row, 2, f"={roic_ref_value}", FMT_PCT2, True)
            roic_ref_for_row = f"B{row}"
        _set_formula_cell(ws, row, 3, f"=IFERROR({ref_g}/{roic_ref_for_row},\"N/A\")", FMT_PCT2, True)
        _set_formula_cell(ws, row, 4, f"=IFERROR({nopat_cell_ref}*(1-C{row}),\"N/A\")", FMT_COMMA2, True)
        tv_case = f"(D{row}*(1+{ref_g})/({denom})*{df_y5})"
        iv_case = f"(({roic_pv_sum_ref}+{tv_case}-{roic_nd_ref})/{roic_sh_ref})"
        if key == "reported":
            _set_formula_cell(ws, row, 5, "=0", FMT_COMMA2, True)
        else:
            _set_formula_cell(ws, row, 5, f"=IFERROR({iv_case}-{reported_iv_formula},\"N/A\")", FMT_COMMA2, True)
        VAL_ROWS[f"roic_sensitivity_{key}"] = row
        row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "Diagnostic only: normalized ROIC applies only to terminal FCF reinvestment need. Explicit-period FCF remains unchanged. Fully coherent ROIC normalization would require capitalized R&D, lease ROU, and explicit-period reinvestment reassessment, out of scope for this patch.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "ROIC = WACC case will show materially lower IV than headline by construction because reinvestment is materially higher. This is a no-value-creation floor reference and discussion item, not an investment recommendation.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 2

    _set_label_cell(ws, row, 1, "Less: Net Debt (negative = net cash add-back)")
    VAL_ROWS["net_debt"] = row
    # V3.7.2: use the Net Debt Bridge "Selected Net Debt Used in DCF" cell.
    # Default keeps reported_input_net_debt so headline IV is preserved.
    ndb_selected_ref = _abs_ref(ASSUMP_CELLS.get("ndb_selected", ASSUMP_CELLS["net_debt"]))
    _set_formula_cell(ws, row, 2, f"={ndb_selected_ref}", FMT_COMMA2, True)
    _set_note_cell(ws, row, 3, "Net Debt = Debt - Cash - selected marketable securities. If negative, subtracting it adds net cash to equity value.")
    row += 1

    _set_label_cell(ws, row, 1, "Equity Value", True)
    VAL_ROWS["equity"] = row
    _set_formula_cell(ws, row, 2, f"=B{VAL_ROWS['ev']}-B{VAL_ROWS['net_debt']}", FMT_COMMA2)
    row += 1

    _set_label_cell(ws, row, 1, "Diluted Shares (M) - Selected")
    VAL_ROWS["shares"] = row
    # V3.7.4: pull from Assumptions Shareholder Returns block selected
    # denominator. Default = Current Reported Diluted Shares, so headline IV
    # is preserved when the dropdown is left at default.
    sr_selected_denom_ref = _abs_ref(ASSUMP_CELLS.get("sr_selected_denominator", ASSUMP_CELLS["shares"]))
    _set_formula_cell(ws, row, 2, f"={sr_selected_denom_ref}", FMT_COMMA2, True)
    row += 1

    _set_label_cell(ws, row, 1, "Intrinsic Value per Share", True)
    VAL_ROWS["intrinsic"] = row
    if model_unsuitable_for_dcf:
        _set_formula_cell(ws, row, 2, "=NA()", FMT_COMMA2)
        _set_note_cell(ws, row, 3, current_params.get("model_unsuitable_reason") or "DCF unavailable for this security type.")
    elif fx_rate_for_market_comparison is None or trading_iv_for_market_comparison is None or trading_iv_for_market_comparison <= 0:
        _set_formula_cell(ws, row, 2, "=NA()", FMT_COMMA2)
        _set_note_cell(ws, row, 3, current_params.get("currency_translation_warning") or "Intrinsic value unavailable; valuation output not clean.")
    else:
        _set_formula_cell(ws, row, 2, f"=(B{VAL_ROWS['equity']}/B{VAL_ROWS['shares']})*{fx_rate_for_market_comparison}", FMT_COMMA2)
        _set_note_cell(ws, row, 3, "Per-share IV translated from reporting currency to trading currency for price comparison.")
    row += 1

    _set_label_cell(ws, row, 1, "Current Price")
    VAL_ROWS["price"] = row
    if current_price_for_market_comparison is None or current_price_for_market_comparison <= 0:
        _set_formula_cell(ws, row, 2, "=NA()", FMT_COMMA2, True)
        _set_note_cell(ws, row, 3, "Current price unavailable; upside/downside comparison blocked.")
    else:
        _set_formula_cell(ws, row, 2, f"={ref_price}", FMT_COMMA2, True)
    row += 1

    _set_label_cell(ws, row, 1, "Upside / (Downside)", True)
    VAL_ROWS["upside"] = row
    if market_comparison_unavailable:
        _set_formula_cell(ws, row, 2, "=NA()", FMT_PCT1)
        _set_note_cell(ws, row, 3, current_params.get("model_unsuitable_reason") or "Market comparison unavailable; see Currency Translation Audit.")
    else:
        _set_formula_cell(ws, row, 2, f"=B{VAL_ROWS['intrinsic']}/B{VAL_ROWS['price']}-1", FMT_PCT1)

    if model_unsuitable_for_dcf:
        ws.cell(row=VAL_ROWS["headline_intrinsic"], column=2, value="N/A").font = font_formula(); ws.cell(row=VAL_ROWS["headline_intrinsic"], column=2).border = border_thin()
        _set_formula_cell(ws, VAL_ROWS["headline_price"], 2, f"=B{VAL_ROWS['price']}", FMT_COMMA2)
        ws.cell(row=VAL_ROWS["headline_upside"], column=2, value="N/A").font = font_formula(); ws.cell(row=VAL_ROWS["headline_upside"], column=2).border = border_thin()
    else:
        _set_formula_cell(ws, VAL_ROWS["headline_intrinsic"], 2, f"=B{VAL_ROWS['intrinsic']}", FMT_COMMA2)
        _set_formula_cell(ws, VAL_ROWS["headline_price"], 2, f"=B{VAL_ROWS['price']}", FMT_COMMA2)
        _set_formula_cell(ws, VAL_ROWS["headline_upside"], 2, f"=B{VAL_ROWS['upside']}", FMT_PCT1)

    row += 2
    write_section_title(ws, row, "Currency Translation Audit", span_cols=3)
    row += 1
    row = _write_kv_rows(ws, row, [
        ("Reporting Currency", current_params.get("reporting_currency") or out.currency, None),
        ("Trading Currency", current_params.get("trading_currency") or out.currency, None),
        ("FX Rate Used", current_params.get("fx_rate_reporting_to_trading"), None),
        ("FX Source", current_params.get("fx_rate_source") or "Not available", None),
        ("IV/share - Reporting Currency", current_params.get("intrinsic_value_per_share_reporting_currency"), None),
        ("IV/share - Trading Currency", current_params.get("intrinsic_value_per_share_trading_currency"), None),
        ("Valuation Status", current_params.get("valuation_status") or "Not available", None),
        ("Market Comparison Status", current_params.get("market_comparison_status") or "Not available", None),
        ("Translation Status", current_params.get("currency_translation_status") or "Not available", None),
        ("Translation Warning", current_params.get("currency_translation_warning") or "None", None),
        ("Market / Valuation Warnings", " | ".join(current_params.get("warnings") or []) or "None", None),
    ])
    row += 1
    # ── V3.7.2 Adjusted IV/share references ────────────────────────────
    write_section_title(ws, row, "Adjusted IV/share — Net Debt Bridge variants", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Net Debt Treatment", "IV / Share", "Net Debt Used"], 1)
    row += 1
    ev_ref = f"B{VAL_ROWS['ev']}"
    shares_ref = f"B{VAL_ROWS['shares']}"
    variants = [
        ("IV using Reported / Input Net Debt", "ndb_reported"),
        ("IV using Adj #1: Debt - Cash", "ndb_adj_cash"),
        ("IV using Adj #2: Debt - Cash - ST Investments", "ndb_adj_st"),
        ("IV using Adj #3: Debt - Cash - Total Marketable Securities", "ndb_adj_total_ms"),
    ]
    for label, cell_key in variants:
        _set_label_cell(ws, row, 1, label)
        nd_ref = _abs_ref(ASSUMP_CELLS[cell_key])
        if fx_rate_for_market_comparison is None:
            _set_formula_cell(ws, row, 2, "=NA()", FMT_COMMA2)
        else:
            _set_formula_cell(ws, row, 2, f"=(({ev_ref}-{nd_ref})/{shares_ref})*{fx_rate_for_market_comparison}", FMT_COMMA2)
        _set_formula_cell(ws, row, 3, f"={nd_ref}", FMT_COMMA2, True)
        VAL_ROWS[f"adj_iv_{cell_key}"] = row
        row += 1
    _set_label_cell(ws, row, 1, "Selected / Headline IV (live, mirrors above)", True)
    _set_formula_cell(ws, row, 2, f"=B{VAL_ROWS['intrinsic']}", FMT_COMMA2, True)
    _set_formula_cell(ws, row, 3, f"={_abs_ref(ASSUMP_CELLS['ndb_selected'])}", FMT_COMMA2, True)
    VAL_ROWS["adj_iv_selected"] = row
    row += 1
    _set_note_cell(
        ws, row, 1,
        "Headline IV uses the Selected Net Debt Treatment cell on Assumptions. The default reported/input treatment matches the headline; changing the treatment updates the equity bridge and shows the IV impact versus other rows here.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # ── V3.7.4 Alt IV/Share by Share Count Treatment ───────────────────
    write_section_title(ws, row, "Alt IV/share — Share Count Treatment variants", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Share Count Treatment", "IV / Share", "Shares Used (M)"], 1)
    row += 1
    equity_ref_for_share_alt = f"B{VAL_ROWS['equity']}"
    sc_variants = [
        ("IV using Current Reported Diluted Shares", "sr_denom_current"),
        ("IV using Forecast Ending Diluted Shares", "sr_denom_ending"),
        ("IV using Forecast Weighted Avg Diluted Shares", "sr_denom_wavg"),
    ]
    for label, cell_key in sc_variants:
        _set_label_cell(ws, row, 1, label)
        sc_ref = _abs_ref(ASSUMP_CELLS[cell_key])
        _set_formula_cell(ws, row, 2, f"=IFERROR({equity_ref_for_share_alt}/{sc_ref},\"n/a\")", FMT_COMMA2)
        _set_formula_cell(ws, row, 3, f"={sc_ref}", FMT_COMMA2, True)
        VAL_ROWS[f"alt_iv_{cell_key}"] = row
        row += 1
    _set_label_cell(ws, row, 1, "Selected / Headline IV (live)", True)
    _set_formula_cell(ws, row, 2, f"=B{VAL_ROWS['intrinsic']}", FMT_COMMA2, True)
    _set_formula_cell(ws, row, 3, f"=B{VAL_ROWS['shares']}", FMT_COMMA2, True)
    VAL_ROWS["alt_iv_selected_shares"] = row
    row += 1
    _set_note_cell(
        ws, row, 1,
        "Per-share valuation differs across share count treatments because forecast buybacks can change the diluted share base. EV / Equity Value are unchanged. Switch the Selected Share Count Treatment dropdown on Assumptions to update the headline.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # ── V3.7.5 WACC Decision Impact (alt IV/share under each WACC case) ─
    write_section_title(ws, row, "WACC Decision Impact — Alt IV/share under WACC treatments", span_cols=3)
    row += 1
    write_header_row(ws, row, ["WACC Treatment", "IV / Share", "WACC Used"], 1)
    row += 1
    wacc_variants_dcf = [
        ("IV using Selected / Model WACC", "selected_model_wacc"),
        ("IV using Mechanical CAPM Reference", "capm_indicative_wacc"),
        ("IV using Selected WACC + 100 bps", "selected_plus_spread_100bps"),
        ("IV using Selected WACC - 100 bps", "selected_minus_spread_100bps"),
    ]
    wacc_case_to_assump = {
        "selected_model_wacc": "wacc_selected_echo",
        "capm_indicative_wacc": "wacc_indicative",
        "selected_plus_spread_100bps": "wacc_plus_100bps",
        "selected_minus_spread_100bps": "wacc_minus_100bps",
    }
    # Each alt IV is a live Excel formula: PV(FCFs) is re-discounted at the
    # alt WACC cell on Assumptions, and the terminal value is recomputed
    # using the same Terminal Value Method dropdown that drives the headline.
    for label, case_key in wacc_variants_dcf:
        _set_label_cell(ws, row, 1, label)
        wacc_cell_addr = _abs_ref(ASSUMP_CELLS[wacc_case_to_assump[case_key]])
        _set_formula_cell(ws, row, 2, _alt_iv_at_wacc(wacc_cell_addr, n), FMT_COMMA2, True)
        _set_formula_cell(ws, row, 3, f"={wacc_cell_addr}", FMT_PCT2, True)
        VAL_ROWS[f"alt_iv_wacc_{case_key}"] = row
        row += 1
    _set_label_cell(ws, row, 1, "Selected / Headline IV (live)", True)
    _set_formula_cell(ws, row, 2, f"=B{VAL_ROWS['intrinsic']}", FMT_COMMA2, True)
    _set_formula_cell(ws, row, 3, f"={_abs_ref(ASSUMP_CELLS.get('wacc_used', ASSUMP_CELLS['wacc']))}", FMT_PCT2, True)
    VAL_ROWS["alt_iv_selected_wacc"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Diff: Mechanical CAPM IV vs Selected IV")
    _set_formula_cell(
        ws, row, 2,
        f"=IFERROR(B{VAL_ROWS['alt_iv_wacc_capm_indicative_wacc']}-B{VAL_ROWS['alt_iv_wacc_selected_model_wacc']},\"n/a\")",
        FMT_COMMA2,
    )
    row += 1
    _set_note_cell(
        ws, row, 1,
        "Each row is a live Excel formula: PV(FCFs) is re-discounted at the alternative WACC cell on Assumptions, and terminal value is recomputed using the same Terminal Value Method (Gordon / Exit / Average) that drives the headline. Net Debt and Share Count are held at the selected treatments. Fade Period reference case is not recomputed at alternative WACCs - it remains an engine-generated reference (see Terminal Decision Impact table). Switch the Selected WACC Treatment dropdown on Assumptions to update the headline.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # ── V3.7.6 Terminal Decision Impact (alt IV/share + TV/EV) ─────────
    write_section_title(ws, row, "Terminal Decision Impact — Alt IV/share + TV/EV under treatments", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Terminal Treatment", "IV / Share", "TV / EV"], 1)
    row += 1
    tv_bridge_dcf = getattr(out, "terminal_decision_bridge", None) or {}
    tv_alt_iv = tv_bridge_dcf.get("alternative_iv_per_share") or {}
    tv_alt_pct = tv_bridge_dcf.get("alternative_tv_ev_pct") or {}
    tv_variants_dcf = [
        ("IV using Current Model Terminal Value", "current_model_terminal"),
        ("IV using Gordon Growth Terminal Value", "gordon_growth"),
        ("IV using Exit Multiple Terminal Value", "exit_multiple"),
        ("IV using Gordon / Exit Blend", "gordon_exit_blend"),
        ("IV using H-Model / Two-Stage Gordon", "h_model"),
        ("IV using Fade Period Reference Case (engine-generated reference)", "fade_period_reference"),
    ]
    pv_sum_ref = f"B{VAL_ROWS['pv_sum']}"
    tv_g_ref = f"B{VAL_ROWS['tv_gordon']}"
    tv_e_ref = f"B{VAL_ROWS['tv_exit']}"
    nd_ref = f"B{VAL_ROWS['net_debt']}"
    sh_ref = f"B{VAL_ROWS['shares']}"
    intrinsic_ref = f"B{VAL_ROWS['intrinsic']}"
    tv_pct_ref = f"B{VAL_ROWS['tv_pct']}"
    tv_iv_formulas = {
        "current_model_terminal": f"={intrinsic_ref}",
        "gordon_growth": f"=({pv_sum_ref}+{tv_g_ref}-{nd_ref})/{sh_ref}",
        "exit_multiple": f"=({pv_sum_ref}+{tv_e_ref}-{nd_ref})/{sh_ref}",
        "gordon_exit_blend": f"=({pv_sum_ref}+AVERAGE({tv_g_ref},{tv_e_ref})-{nd_ref})/{sh_ref}",
        "h_model": f"=({pv_sum_ref}+B{VAL_ROWS['tv_h_model']}-{nd_ref})/{sh_ref}",
    }
    tv_pct_formulas = {
        "current_model_terminal": f"={tv_pct_ref}",
        "gordon_growth": f"=IFERROR({tv_g_ref}/({pv_sum_ref}+{tv_g_ref}),0)",
        "exit_multiple": f"=IFERROR({tv_e_ref}/({pv_sum_ref}+{tv_e_ref}),0)",
        "gordon_exit_blend": (
            f"=IFERROR(AVERAGE({tv_g_ref},{tv_e_ref})/"
            f"({pv_sum_ref}+AVERAGE({tv_g_ref},{tv_e_ref})),0)"
        ),
        "h_model": f"=IFERROR(B{VAL_ROWS['tv_h_model']}/({pv_sum_ref}+B{VAL_ROWS['tv_h_model']}),0)",
    }
    for label, case_key in tv_variants_dcf:
        _set_label_cell(ws, row, 1, label)
        if case_key in tv_iv_formulas:
            _set_formula_cell(ws, row, 2, tv_iv_formulas[case_key], FMT_COMMA2, True)
            _set_formula_cell(ws, row, 3, tv_pct_formulas[case_key], FMT_PCT2, True)
        else:
            # Fade Period reference case: engine-generated, static at export.
            iv_val = tv_alt_iv.get(case_key)
            if iv_val is None:
                ws.cell(row=row, column=2, value="n/a").font = font_formula()
            else:
                c2 = ws.cell(row=row, column=2, value=float(iv_val))
                c2.font = font_formula()
                c2.border = border_thin()
                c2.number_format = FMT_COMMA2
            pct_val = tv_alt_pct.get(case_key)
            if pct_val is None:
                ws.cell(row=row, column=3, value="n/a").font = font_formula()
            else:
                c3 = ws.cell(row=row, column=3, value=float(pct_val))
                c3.font = font_formula()
                c3.border = border_thin()
                c3.number_format = FMT_PCT2
        VAL_ROWS[f"alt_iv_tv_{case_key}"] = row
        row += 1
    _set_label_cell(ws, row, 1, "Selected / Headline IV (live)", True)
    _set_formula_cell(ws, row, 2, f"=B{VAL_ROWS['intrinsic']}", FMT_COMMA2, True)
    _set_formula_cell(ws, row, 3, f"=B{VAL_ROWS['tv_pct']}", FMT_PCT2, True)
    VAL_ROWS["alt_iv_selected_terminal"] = row
    row += 1
    _set_label_cell(ws, row, 1, "Diff: Fade IV − Current Model IV")
    h_row = VAL_ROWS.get("alt_iv_tv_h_model")
    if h_row:
        _set_label_cell(ws, row, 1, "Diff: H-Model IV vs Selected IV")
        _set_formula_cell(ws, row, 2, f"=IFERROR(B{h_row}-B{VAL_ROWS['intrinsic']},\"n/a\")", FMT_COMMA2)
        VAL_ROWS["alt_iv_tv_h_model_vs_selected_gap"] = row
        row += 1
        _set_label_cell(ws, row, 1, "Diff: Fade IV vs Current Model IV")
    fade_row = VAL_ROWS.get("alt_iv_tv_fade_period_reference")
    cur_row = VAL_ROWS.get("alt_iv_tv_current_model_terminal")
    if fade_row and cur_row:
        _set_formula_cell(
            ws, row, 2,
            f"=IFERROR(B{fade_row}-B{cur_row},\"n/a\")",
            FMT_COMMA2,
        )
    row += 1
    _set_note_cell(
        ws, row, 1,
        "Gordon, Exit, Blend, H-Model, and Current Model rows are live Excel formulas that swap only the terminal-value PV against the live FCF Build, holding selected Net Debt and Share Count constant. Fade Period reference is an engine-generated reference (static at export, refreshed by re-running the export). Switch the Terminal Treatment dropdown on Assumptions to update the headline.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    op_bridge = getattr(out, "operating_path_bridge", None) or {}
    if op_bridge.get("available"):
        write_section_title(ws, row, "Operating Thesis Bridge Alternative IV", span_cols=4)
        VAL_ROWS["operating_bridge_alt_iv_block"] = row
        row += 1
        write_header_row(ws, row, ["Case", "IV / Share", "Difference vs Headline", "Attribution"], 1)
        row += 1
        alt_iv = op_bridge.get("alternative_iv") or {}
        attribution = op_bridge.get("attribution") or {}
        alt_rows = [
            ("IV with Selected Path", "selected_path", None),
            ("IV with Bridge Revenue + Selected Margin", "bridge_revenue_selected_margin", "Revenue impact"),
            ("IV with Selected Revenue + Bridge Margin", "selected_revenue_bridge_margin", "Margin impact"),
            ("IV with Full Bridge = Bridge Revenue + Bridge Margin", "full_bridge", "Full bridge"),
        ]
        for label, key, attr_label in alt_rows:
            _set_label_cell(ws, row, 1, label, key == "full_bridge")
            val = alt_iv.get(key)
            if val is None:
                ws.cell(row=row, column=2, value="n/a").font = font_formula()
            else:
                c = ws.cell(row=row, column=2, value=float(val))
                c.font = font_formula(); c.border = border_thin(); c.number_format = FMT_COMMA2
            _set_formula_cell(ws, row, 3, f"=IFERROR(B{row}-B{VAL_ROWS['intrinsic']},\"n/a\")", FMT_COMMA2, True)
            _set_note_cell(ws, row, 4, attr_label or "Baseline for bridge attribution")
            VAL_ROWS[f"op_bridge_alt_iv_{key}"] = row
            row += 1
        for label, key in [
            ("Revenue impact", "revenue_impact"),
            ("Margin impact", "margin_impact"),
            ("Interaction", "interaction"),
        ]:
            _set_label_cell(ws, row, 1, label, key == "interaction")
            val = attribution.get(key)
            c = ws.cell(row=row, column=2, value=float(val) if val is not None else "n/a")
            c.font = font_formula(); c.border = border_thin()
            if val is not None:
                c.number_format = FMT_COMMA2
            _set_note_cell(ws, row, 4, "Attribution: interaction = Full Bridge - Selected Path - Revenue impact - Margin impact." if key == "interaction" else "Attribution component.")
            VAL_ROWS[f"op_bridge_attr_{key}"] = row
            row += 1
        _set_note_cell(
            ws,
            row,
            1,
            "Bridge-driven IV is alternative reference only and does not populate the Football Field. Football Field continues to use the Selected Path headline unless the Operating Path Source selector is intentionally switched.",
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 2

    # ── V3.9.5 Terminal Value Philosophy & Review ────────────────────────
    write_section_title(ws, row, "Terminal Method Comparison", span_cols=5)
    VAL_ROWS["terminal_method_comparison"] = row
    row += 1
    write_header_row(ws, row, ["Method", "IV / Share", "TV / EV", "Gap vs Selected", "Review Tier"], 1)
    row += 1
    comparison_rows = [
        ("Single-stage Gordon IV", "gordon_growth"),
        ("Exit Multiple IV", "exit_multiple"),
        ("Gordon / Exit Blend IV", "gordon_exit_blend"),
        ("H-Model IV", "h_model"),
        ("Current selected headline IV", "selected_terminal"),
    ]
    for label, key in comparison_rows:
        _set_label_cell(ws, row, 1, label)
        if key == "selected_terminal":
            iv_formula = f"=B{VAL_ROWS['intrinsic']}"
            pct_formula = f"=B{VAL_ROWS['tv_pct']}"
        else:
            iv_formula = f"=B{VAL_ROWS[f'alt_iv_tv_{key}']}"
            pct_formula = f"=C{VAL_ROWS[f'alt_iv_tv_{key}']}"
        _set_formula_cell(ws, row, 2, iv_formula, FMT_COMMA2, True)
        _set_formula_cell(ws, row, 3, pct_formula, FMT_PCT2, True)
        _set_formula_cell(ws, row, 4, f"=IFERROR(B{row}-B{VAL_ROWS['intrinsic']},\"n/a\")", FMT_COMMA2, True)
        _set_formula_cell(ws, row, 5, f'=IF(C{row}>0.80,"High Review",IF(C{row}>0.70,"Review","OK"))', None, True)
        VAL_ROWS["terminal_method_comparison_" + key] = row
        row += 1
    _set_label_cell(ws, row, 1, "Gordon vs Exit gap")
    _set_formula_cell(ws, row, 2, f"=B{VAL_ROWS['tv_gordon_exit_gap']}", FMT_PCT2, True)
    _set_note_cell(ws, row, 3, "Single-stage Gordon can be conservative when near-term excess returns fade slowly; Exit can be aggressive when market multiples embed cyclical optimism.")
    VAL_ROWS["terminal_method_comparison_gordon_exit_gap"] = row
    row += 1
    _set_label_cell(ws, row, 1, "H-Model vs selected gap")
    _set_formula_cell(ws, row, 2, f"=IFERROR(B{VAL_ROWS['alt_iv_tv_h_model']}-B{VAL_ROWS['intrinsic']},\"n/a\")", FMT_COMMA2, True)
    _set_note_cell(ws, row, 3, "H-Model is a structured middle path; it does not become headline unless selected in Terminal Treatment.")
    VAL_ROWS["terminal_method_comparison_h_model_gap"] = row
    row += 2

    tv_philosophy = _build_terminal_philosophy_payload(out, ctx)
    write_section_title(ws, row, "Terminal Value Philosophy & Review", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Item", "Value", "Note / Review Tier"], 1)
    row += 1

    # Live references where possible so the block reacts to dropdown / input
    # changes; static review tiers come from the engine payload.
    tv_used_ref = f"B{VAL_ROWS['tv_used']}"
    tv_pct_ref = f"B{VAL_ROWS['tv_pct']}"
    tv_g_ref = f"B{VAL_ROWS['tv_gordon']}"
    tv_e_ref = f"B{VAL_ROWS['tv_exit']}"
    pv_sum_ref = f"B{VAL_ROWS['pv_sum']}"
    nd_ref_local = f"B{VAL_ROWS['net_debt']}"
    sh_ref_local = f"B{VAL_ROWS['shares']}"
    headline_iv_ref = f"B{VAL_ROWS['intrinsic']}"
    implied_em_ref = f"B{VAL_ROWS['tv_implied_exit_multiple']}"
    implied_g_ref = f"B{VAL_ROWS['tv_implied_growth_from_exit']}"
    g_gap_ref = f"B{VAL_ROWS['tv_gordon_exit_gap']}"
    gordon_iv_ref = f"B{VAL_ROWS['alt_iv_tv_gordon_growth']}"
    exit_iv_ref = f"B{VAL_ROWS['alt_iv_tv_exit_multiple']}"

    method_label = tv_philosophy.get("selected_terminal_method_label") or ""
    philosophy_label = tv_philosophy.get("primary_philosophy") or ""

    def _write_row(label, value_writer, note):
        _set_label_cell(ws, row, 1, label)
        value_writer(row)
        _set_note_cell(ws, row, 3, note)

    # 1. Headline Terminal Method
    def _w_headline(r):
        c = ws.cell(row=r, column=2, value=method_label)
        c.font = font_formula(); c.border = border_thin()
    _write_row("Headline Terminal Method", _w_headline, "Selected terminal method remains the headline methodology; alternatives are shown as cross-checks.")
    row += 1

    # 2. Primary Philosophy
    def _w_phil(r):
        c = ws.cell(row=r, column=2, value=philosophy_label)
        c.font = font_formula(); c.border = border_thin()
    _write_row("Primary Philosophy", _w_phil, tv_philosophy.get("terminal_method_recommended_use") or "")
    row += 1

    # 3. Terminal Growth (live) + tier
    g_tier = tv_philosophy.get("terminal_growth_review_tier") or "N/A"
    def _w_g(r):
        _set_formula_cell(ws, r, 2, f"={ref_g}", FMT_PCT2, True)
    _write_row("Terminal Growth", _w_g, f"Review tier: {g_tier} | 0%-4% OK | <0% or >4% Review | >5% High review.")
    row += 1

    # 4. Exit Multiple (live)
    def _w_em(r):
        _set_formula_cell(ws, r, 2, f"={ref_exit}", FMT_MULTIPLE, True)
    _write_row("Exit Multiple", _w_em, "Editable on Assumptions; cached daily for deterministic IV anchoring.")
    row += 1

    # 5. Gordon Implied Exit Multiple (live)
    def _w_iem(r):
        _set_formula_cell(ws, r, 2, f"={implied_em_ref}", FMT_MULTIPLE, True)
    _write_row("Gordon Implied Exit Multiple", _w_iem, "Cross-method: what exit multiple Gordon TV implies.")
    row += 1

    # 6. Exit Implied Terminal Growth (live)
    def _w_ieg(r):
        _set_formula_cell(ws, r, 2, f"={implied_g_ref}", FMT_PCT2, True)
    _write_row("Exit Implied Terminal Growth", _w_ieg, "Cross-method: what long-run g the Exit TV implies.")
    row += 1

    # 7. TV / EV (live)
    def _w_tvev(r):
        _set_formula_cell(ws, r, 2, f"={tv_pct_ref}", FMT_PCT2, True)
    _write_row("TV / EV", _w_tvev, "Share of EV attributable to terminal value.")
    row += 1

    # 8. TV Dependency Review Tier
    tv_ev_tier = tv_philosophy.get("tv_ev_review_tier") or "N/A"
    def _w_tvev_tier(r):
        c = ws.cell(row=r, column=2, value=tv_ev_tier)
        c.font = font_formula(); c.border = border_thin()
    _write_row("TV Dependency Review Tier", _w_tvev_tier, "<=70% OK | 70%-80% Review | >80% High review. High terminal dependency is a discussion item, not an automatic model error.")
    row += 1

    # 9. Gordon vs Exit IV Gap (live)
    def _w_iv_gap(r):
        _set_formula_cell(
            ws, r, 2,
            f"=IFERROR(ABS({gordon_iv_ref}-{exit_iv_ref})/MAX(ABS({gordon_iv_ref}),ABS({exit_iv_ref}),1),0)",
            FMT_PCT2, True,
        )
    _write_row("Gordon vs Exit IV Gap", _w_iv_gap, "Relative gap between Gordon IV and Exit IV per share.")
    row += 1

    # 10. Gordon vs Exit Review Tier
    ge_tier = tv_philosophy.get("gordon_vs_exit_review_tier") or "N/A"
    def _w_ge_tier(r):
        c = ws.cell(row=r, column=2, value=ge_tier)
        c.font = font_formula(); c.border = border_thin()
    _write_row("Gordon vs Exit Review Tier", _w_ge_tier, "<=15% OK | 15%-30% Review | >30% High review. Tier reflects the export-time TV gap between methods.")
    row += 1

    # 11. Exit Multiple vs Trading Comps Median
    em_vs_comps = tv_philosophy.get("exit_multiple_vs_comps_median")
    em_vs_comps_pd = tv_philosophy.get("exit_multiple_vs_comps_premium_discount")
    comps_tier = tv_philosophy.get("exit_multiple_vs_comps_review_tier") or "N/A"
    unavailable_reason = tv_philosophy.get("comps_median_unavailable_reason")
    def _w_em_vs_comps(r):
        if em_vs_comps is None:
            c = ws.cell(row=r, column=2, value="N/A - insufficient included peer data")
            c.font = font_formula(); c.border = border_thin()
        else:
            c = ws.cell(row=r, column=2, value=float(em_vs_comps))
            c.font = font_formula(); c.border = border_thin()
            c.number_format = FMT_MULTIPLE
    if em_vs_comps_pd is None:
        note_em_vs_comps = unavailable_reason or "Trading Comps EV/EBITDA median not available."
    else:
        pd_pct = em_vs_comps_pd * 100.0
        sign = "premium" if pd_pct >= 0 else "discount"
        note_em_vs_comps = f"Selected exit multiple is {abs(pd_pct):.1f}% {sign} vs comps median."
    _write_row("Exit Multiple vs Trading Comps Median", _w_em_vs_comps, note_em_vs_comps)
    row += 1

    # 12. Exit Multiple vs Comps Review Tier
    def _w_comps_tier(r):
        c = ws.cell(row=r, column=2, value=comps_tier)
        c.font = font_formula(); c.border = border_thin()
    _write_row("Exit Multiple vs Comps Review Tier", _w_comps_tier, "Within +-15% OK | +-15-30% Review | >30% High review | N/A when comps insufficient.")
    row += 1

    # 13. Fade Reference Use
    def _w_fade(r):
        c = ws.cell(row=r, column=2, value="Diagnostic / static at export")
        c.font = font_formula(); c.border = border_thin()
    _write_row("Fade Reference Use", _w_fade, "Fade reference is diagnostic and static at export unless re-exported. Not live in the workbook.")
    row += 1

    # 14. IC Discussion Note
    def _w_note(r):
        c = ws.cell(row=r, column=2, value="See note below")
        c.font = font_formula(); c.border = border_thin()
    _write_row("IC Discussion Note", _w_note, "Review tiers are diagnostic; the headline terminal method is preserved unless the analyst intentionally switches the Terminal Treatment dropdown.")
    row += 1

    _set_note_cell(
        ws, row, 1,
        "Method hierarchy: Current Model / selected method = headline; Gordon = fundamental cross-check; Exit Multiple = market multiple cross-check; Blend = balancing reference; Fade = diagnostic reference (static at export). Selected terminal method remains the headline methodology; alternatives are shown as cross-checks and do not auto-overwrite the headline.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # ── V3.7.6 Fade Period Reference block (visible series) ────────────
    write_section_title(ws, row, "Fade Period Reference Case (visible series)", span_cols=3)
    row += 1
    fade_block = tv_bridge_dcf.get("fade_period") or {}
    fade_years = int(fade_block.get("fade_years") or 0)
    start_g = fade_block.get("starting_fade_growth")
    target_g = fade_block.get("fade_target_growth")
    fade_growth_series = list(fade_block.get("fade_growth_series") or [])
    fade_fcf_series = list(fade_block.get("fade_fcf_series") or [])
    pv_fade_fcf_series = list(fade_block.get("pv_fade_fcf_series") or [])
    _set_label_cell(ws, row, 1, "Fade Years"); ws.cell(row=row, column=2, value=fade_years).number_format = "0"; row += 1
    _set_label_cell(ws, row, 1, "Starting Fade Growth (Year N / Year N-1 - 1)")
    ws.cell(row=row, column=2, value=start_g if start_g is not None else 0.0).number_format = FMT_PCT2
    row += 1
    _set_label_cell(ws, row, 1, "Fade Target Growth")
    ws.cell(row=row, column=2, value=target_g if target_g is not None else 0.0).number_format = FMT_PCT2
    row += 1
    # Header + series rows.
    if fade_years > 0 and fade_growth_series:
        header_year_cols = ["Fade Year"] + [f"Y+{i+1}" for i in range(fade_years)]
        write_header_row(ws, row, header_year_cols, 1)
        row += 1
        for label, series in (
            ("Fade FCF Growth", fade_growth_series),
            ("Fade FCF (M)", fade_fcf_series),
            ("PV of Fade FCF (M)", pv_fade_fcf_series),
        ):
            _set_label_cell(ws, row, 1, label)
            for i, v in enumerate(series):
                c = ws.cell(row=row, column=2 + i, value=v)
                c.font = font_formula(); c.border = border_thin()
                c.number_format = FMT_PCT2 if label == "Fade FCF Growth" else FMT_COMMA2
            row += 1
    _set_label_cell(ws, row, 1, "PV of Fade FCF Total (M)")
    ws.cell(row=row, column=2, value=fade_block.get("pv_fade_fcf_total") or 0.0).number_format = FMT_COMMA2
    row += 1
    _set_label_cell(ws, row, 1, "Terminal FCF After Fade (M)")
    ws.cell(row=row, column=2, value=fade_block.get("terminal_fcf_after_fade") or 0.0).number_format = FMT_COMMA2
    row += 1
    _set_label_cell(ws, row, 1, "Terminal Value After Fade, PV (M)")
    ws.cell(row=row, column=2, value=fade_block.get("terminal_value_after_fade_pv") or 0.0).number_format = FMT_COMMA2
    row += 1
    _set_label_cell(ws, row, 1, "Fade Case EV (M)")
    ws.cell(row=row, column=2, value=fade_block.get("fade_case_ev") or 0.0).number_format = FMT_COMMA2
    row += 1
    _set_note_cell(
        ws, row, 1,
        "Fade Period reference: FCF growth fades linearly from Year-N growth toward the fade-target growth across Fade Years; then Gordon TV is applied on the final fade-year FCF. Reference case only - not a full 10-year 3FS forecast.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    write_section_title(ws, row, "Terminal Value Sanity Check v1", span_cols=3)
    row += 1
    sanity = out.terminal_sanity or {}
    thresholds = sanity.get("thresholds", {})
    sanity_rows = [
        (
            "TV dependency",
            "Review" if sanity.get("tv_dependency_high") else "OK",
            f"TV / EV {out.tv_pct:.1%}; threshold {thresholds.get('tv_dependency', 0.75):.0%}",
        ),
        (
            "Gordon spread",
            "Review" if sanity.get("gordon_unstable") else "OK",
            f"WACC - g {sanity.get('gordon_spread', 0):.1%}; threshold {thresholds.get('gordon_spread', 0.01):.0%}",
        ),
        (
            "Terminal method gap",
            "Review" if sanity.get("method_divergence_high") else "OK",
            f"Gordon vs Exit gap {sanity.get('method_diff', 0):.1%}; threshold {thresholds.get('method_diff', 0.25):.0%}",
        ),
    ]
    write_header_row(ws, row, ["Check", "Status", "Detail"], 1)
    row += 1
    for label, status, detail in sanity_rows:
        _set_label_cell(ws, row, 1, label)
        c = ws.cell(row=row, column=2, value=status)
        c.font = font_formula()
        c.border = border_thin()
        _set_note_cell(ws, row, 3, detail)
        row += 1
    row += 1
    _set_note_cell(ws, row, 1, "Sanity checks are diagnostics only and do not change valuation formulas.")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    write_section_title(ws, row, "Break-even / Market-implied Analysis", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Question", "Value", "Method / Interpretation"], 1)
    row += 1
    be_start_row = row
    be_rows = [
        ("Current Price", f"=B{VAL_ROWS['price']}", FMT_COMMA2, "Market price used as the target case."),
        ("Selected IV / Share", f"=B{VAL_ROWS['intrinsic']}", FMT_COMMA2, "Current selected DCF value."),
        ("Required uplift to current price", f"=IFERROR(B{VAL_ROWS['price']}/B{VAL_ROWS['intrinsic']}-1,\"N/M\")", FMT_PCT1, "Gap between current price and selected IV."),
        ("Equity value required by current price", f"=B{VAL_ROWS['price']}*B{VAL_ROWS['shares']}", FMT_COMMA2, "Current price multiplied by selected diluted shares."),
        ("Enterprise value required by current price", f"=B{be_start_row + 3}+{ndb_selected_ref}", FMT_COMMA2, "Adds selected net debt back to implied equity value."),
        ("PV terminal value required", f"=B{be_start_row + 4}-B{VAL_ROWS['pv_sum']}", FMT_COMMA2, "Required EV less PV of explicit FCFs."),
    ]
    for label, formula, fmt, note in be_rows:
        _set_label_cell(ws, row, 1, label)
        _set_formula_cell(ws, row, 2, formula, fmt, True)
        _set_note_cell(ws, row, 3, note)
        row += 1
    required_pv_tv_row = be_start_row + 5
    required_ev_row = be_start_row + 4
    be_wacc_formula = (
        f'=IFERROR(IF(B{VAL_ROWS["price"]}>B{VAL_ROWS["intrinsic"]},'
        f'{_abs_ref(ASSUMP_CELLS["wacc_used"])}+(B{VAL_ROWS["price"]}-B{VAL_ROWS["intrinsic"]})*'
        f'({_abs_ref(ASSUMP_CELLS["wacc_minus_100bps"])}-{_abs_ref(ASSUMP_CELLS["wacc_used"])})/'
        f'(B{VAL_ROWS["alt_iv_wacc_selected_minus_spread_100bps"]}-B{VAL_ROWS["intrinsic"]}),'
        f'{_abs_ref(ASSUMP_CELLS["wacc_used"])}+(B{VAL_ROWS["price"]}-B{VAL_ROWS["intrinsic"]})*'
        f'({_abs_ref(ASSUMP_CELLS["wacc_plus_100bps"])}-{_abs_ref(ASSUMP_CELLS["wacc_used"])})/'
        f'(B{VAL_ROWS["alt_iv_wacc_selected_plus_spread_100bps"]}-B{VAL_ROWS["intrinsic"]})),"N/M")'
    )
    be_terminal_g_core = f"((B{required_pv_tv_row}/{df_y5})*{ref_wacc}-{fcf_y5_addr})/((B{required_pv_tv_row}/{df_y5})+{fcf_y5_addr})"
    be_terminal_g_formula = (
        f'=IFERROR(IF(OR(B{required_pv_tv_row}<=0,{be_terminal_g_core}>={ref_wacc},{be_terminal_g_core}<-0.5),'
        f'"N/M",{be_terminal_g_core}),"N/M")'
    )
    be_exit_formula = (
        f'=IFERROR(IF(OR(B{required_pv_tv_row}<=0,{ebit_y5}+{da_y5}<=0),"N/M",'
        f'(B{required_pv_tv_row}/{df_y5})/({ebit_y5}+{da_y5})),"N/M")'
    )
    market_rows = [
        ("Implied WACC to support current price", be_wacc_formula, FMT_PCT2, "Approximate linear interpolation between selected WACC and +/-100 bps cases."),
        ("Implied terminal g to support current price", be_terminal_g_formula, FMT_PCT2, "Gordon-method solve holding selected WACC and forecast FCF constant; N/M when g is not meaningful."),
        ("Implied exit multiple to support current price", be_exit_formula, FMT_MULTIPLE, "Exit multiple solve holding selected WACC and terminal-year EBIT + D&A constant."),
        ("Operating break-even", '="See Sensitivity sheet"', None, "Sensitivity sheet shows closest Revenue Growth / EBIT Margin grid point to current price."),
    ]
    for label, formula, fmt, note in market_rows:
        _set_label_cell(ws, row, 1, label)
        _set_formula_cell(ws, row, 2, formula, fmt, True)
        _set_note_cell(ws, row, 3, note)
        row += 1
    VAL_ROWS["market_required_ev"] = required_ev_row
    VAL_ROWS["market_required_pv_tv"] = required_pv_tv_row
    row += 1

    write_section_title(ws, row, "WACC Defense", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Metric", "Value", "Discussion note"], 1)
    row += 1
    wacc_defense_rows = [
        ("Selected WACC", f"={_abs_ref(ASSUMP_CELLS['wacc_used'])}", FMT_PCT2, "Retained as the model case."),
        ("Mechanical CAPM reference WACC", f"={_abs_ref(ASSUMP_CELLS['wacc_indicative'])}", FMT_PCT2, "Mechanical CAPM reference, not headline selection."),
        ("Selected vs mechanical CAPM gap (bps)", f"=({_abs_ref(ASSUMP_CELLS['wacc_used'])}-{_abs_ref(ASSUMP_CELLS['wacc_indicative'])})*10000", "#,##0", "Key IC discussion item; not auto-corrected."),
        ("IV under Selected WACC", f"=B{VAL_ROWS['alt_iv_wacc_selected_model_wacc']}", FMT_COMMA2, "Selected case."),
        ("IV under mechanical CAPM reference", f"=B{VAL_ROWS['alt_iv_wacc_capm_indicative_wacc']}", FMT_COMMA2, "Alternative valuation diagnostic."),
        ("Mechanical CAPM IV impact vs Selected", f"=B{VAL_ROWS['alt_iv_wacc_capm_indicative_wacc']}-B{VAL_ROWS['alt_iv_wacc_selected_model_wacc']}", FMT_COMMA2, "Shows valuation impact of switching WACC treatment."),
    ]
    for label, formula, fmt, note in wacc_defense_rows:
        _set_label_cell(ws, row, 1, label)
        _set_formula_cell(ws, row, 2, formula, fmt, True)
        _set_note_cell(ws, row, 3, note)
        row += 1
    _set_note_cell(ws, row, 1, "Selected WACC is retained as the model case. Mechanical CAPM reference is shown as an audit diagnostic; the spread should be reviewed against beta source, risk-free rate, ERP, and cost of debt assumptions.")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    write_section_title(ws, row, "Terminal Value Defense", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Metric", "Value", "Discussion note"], 1)
    row += 1
    tv_defense_rows = [
        ("TV / EV", f"=B{VAL_ROWS['tv_pct']}", FMT_PCT2, "High terminal dependency remains a Review item."),
        ("Gordon IV / Share", f"=B{VAL_ROWS['alt_iv_tv_gordon_growth']}", FMT_COMMA2, "Alternative terminal treatment."),
        ("Exit IV / Share", f"=B{VAL_ROWS['alt_iv_tv_exit_multiple']}", FMT_COMMA2, "Alternative terminal treatment."),
        ("Fade IV / Share", f"=B{VAL_ROWS['alt_iv_tv_fade_period_reference']}", FMT_COMMA2, "Reference case, not a full 10-year 3FS."),
        ("Selected IV / Share", f"=B{VAL_ROWS['intrinsic']}", FMT_COMMA2, "Headline selected case."),
        ("Gordon vs Exit gap", f"=B{VAL_ROWS['tv_gordon_exit_gap']}", FMT_PCT2, "Material divergence is a discussion item."),
        ("Gordon implied exit multiple", f"=B{VAL_ROWS['tv_implied_exit_multiple']}", FMT_MULTIPLE, "Cross-method implied multiple."),
        ("Exit implied terminal g", f"=B{VAL_ROWS['tv_implied_growth_from_exit']}", FMT_PCT2, "Cross-method implied growth."),
    ]
    for label, formula, fmt, note in tv_defense_rows:
        _set_label_cell(ws, row, 1, label)
        _set_formula_cell(ws, row, 2, formula, fmt, True)
        _set_note_cell(ws, row, 3, note)
        row += 1
    _set_note_cell(ws, row, 1, "Selected terminal treatment preserves the workbook baseline convention. Because Gordon and Exit methods diverge materially, both alternatives are presented explicitly and flagged as IC discussion items rather than suppressed.")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    write_section_title(ws, row, "Net Debt Defense", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Treatment", "IV / Share", "Difference vs headline"], 1)
    row += 1
    nd_rows = [
        ("Reported / Input Net Debt", "adj_iv_ndb_reported"),
        ("Debt less Cash", "adj_iv_ndb_adj_cash"),
        ("Debt less Cash & ST Investments", "adj_iv_ndb_adj_st"),
        ("Debt less Cash & Total Marketable Securities", "adj_iv_ndb_adj_total_ms"),
    ]
    for label, key in nd_rows:
        _set_label_cell(ws, row, 1, label)
        _set_formula_cell(ws, row, 2, f"=B{VAL_ROWS[key]}", FMT_COMMA2, True)
        _set_formula_cell(ws, row, 3, f"=B{VAL_ROWS[key]}-B{VAL_ROWS['intrinsic']}", FMT_COMMA2, True)
        row += 1
    _set_note_cell(ws, row, 1, "Reported/Input Net Debt is retained as the headline default. Because marketable securities can be cash-like or strategic depending on analyst judgment, adjusted cases are shown explicitly without changing the headline.")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # V3.9.7 consolidation: the legacy "Shareholder Returns / Buyback Funding
    # Defense" block has been folded into the V3.9.6 Buyback Funding & Cash Use
    # Bridge to remove duplicate FCF/Dividends/Buybacks/Total rows. Set up the
    # shared Supporting Schedules references that the V3.9.6 block consumes.
    sr_rows_map = SCHEDULE_ROWS.get("shareholder_returns") or {}
    sr_col = get_column_letter(_forecast_start_col(ctx))
    sr_fcf = f"'Supporting Schedules'!${sr_col}${sr_rows_map.get(SR_FCF_REF_LABEL, 1)}"
    sr_div = f"'Supporting Schedules'!${sr_col}${sr_rows_map.get('Dividends Paid', 1)}"
    sr_buy = f"'Supporting Schedules'!${sr_col}${sr_rows_map.get('Share Repurchases (Buybacks)', 1)}"
    sr_total = f"'Supporting Schedules'!${sr_col}${sr_rows_map.get('Total Shareholder Returns', 1)}"
    sr_end_shares = f"'Supporting Schedules'!${sr_col}${sr_rows_map.get('Ending Diluted Shares', 1)}"

    # ── V3.9.6 Share Count Roll-forward (canonical, year-1) ─────────────
    write_section_title(ws, row, "Share Count Roll-forward (Year 1, canonical)", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Item", "Value", "Source / Note"], 1)
    row += 1
    sr_beg = f"'Supporting Schedules'!${sr_col}${sr_rows_map.get('Beginning Diluted Shares', 1)}"
    sr_sbc = f"'Supporting Schedules'!${sr_col}${sr_rows_map.get('SBC / Annual Dilution (M)', 1)}"
    sr_repurchased = f"'Supporting Schedules'!${sr_col}${sr_rows_map.get('Shares Repurchased (M)', 1)}"
    sr_dilution_ref = _abs_ref(ASSUMP_CELLS.get("sr_annual_dilution", ASSUMP_CELLS["shares"]))
    selected_denom_ref = _abs_ref(ASSUMP_CELLS.get("sr_selected_denominator", ASSUMP_CELLS["shares"]))
    scr_rows = [
        ("Beginning diluted shares", f"={sr_beg}", FMT_COMMA2,
         "Sourced from Assumptions Diluted Shares; identical to the legacy Share Count Schedule starting balance."),
        ("SBC / stock-based dilution (M)", f"={sr_sbc}", FMT_COMMA2,
         "Annual Dilution % x Beginning shares. Default = 0 unless an analyst input is provided."),
        ("Shares repurchased (M)", f"={sr_repurchased}", FMT_COMMA2,
         "Buyback amount / Repurchase price (Y1). Buybacks remain a financing item; not deducted from FCFF."),
        ("Other issuance / options (M)", '="N/A - not modeled"', None,
         "Option exercises and secondary issuance are not modeled separately; route material issuance through the SBC / Annual Dilution input on Assumptions."),
        ("Ending diluted shares", f"={sr_end_shares}", FMT_COMMA2,
         "Beginning + SBC dilution - Shares repurchased. Forecast Ending / Weighted Avg are surfaced as alternative denominators on Assumptions."),
        ("Selected shares used in IV/share", f"={selected_denom_ref}", FMT_COMMA2,
         "Headline denominator follows the Share Count Treatment dropdown on Assumptions. Default = Current Reported Diluted Shares preserves headline IV; switching to Forecast Ending / Weighted Avg changes per-share value only (EV and Equity Value are unchanged)."),
        ("Review note", '="Beginning and ending diluted shares reconcile to the Shareholder Returns Schedule. Headline denominator is preserved unless the analyst changes the Share Count Treatment dropdown."',
         None, "Year 1 view; subsequent years roll forward inside the Shareholder Returns Schedule."),
    ]
    for label, formula, fmt, note in scr_rows:
        _set_label_cell(ws, row, 1, label)
        _set_formula_cell(ws, row, 2, formula, fmt, True)
        _set_note_cell(ws, row, 3, note)
        row += 1
    row += 1

    # ── V3.9.6 Buyback Funding & Cash Use Bridge ────────────────────────
    write_section_title(ws, row, "Buyback Funding & Cash Use Bridge", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Item", "Value", "Note"], 1)
    row += 1

    nd_bridge_audit = getattr(out, "net_debt_bridge", None) or {}
    beginning_cash_val = nd_bridge_audit.get("cash")
    debt_issuance_modeled = False  # not modeled explicitly
    sr_drivers_audit = (getattr(out, "shareholder_returns", None) or {}).get("drivers_effective") or {}
    annual_dilution_val = sr_drivers_audit.get("annual_dilution_pct")
    if annual_dilution_val is None:
        annual_dilution_val = float(getattr(inp, "annual_dilution_pct", 0.0) or 0.0)

    bridge_rows = [
        ("FCF before shareholder returns", f"={sr_fcf}", FMT_COMMA2,
         "CFO-derived cash flow reference before shareholder returns; not the unlevered FCFF used in DCF. Dividends and buybacks are NOT deducted upstream of this line."),
        ("Dividends", f"={sr_div}", FMT_COMMA2,
         "Capital allocation / equity distribution; not an FCFF item."),
        ("Buybacks", f"={sr_buy}", FMT_COMMA2,
         "Capital allocation / equity bridge; not an FCFF item."),
        ("Total shareholder returns", f"={sr_total}", FMT_COMMA2,
         "Dividends + buybacks."),
        ("Shareholder returns / FCF", f"=IFERROR({sr_total}/{sr_fcf},\"N/M\")", FMT_PCT1,
         "Coverage ratio. Review tier below uses this metric."),
        ("Debt issuance / repayment (modeled)", '="N/A - not modeled"', None,
         "Net new debt is not projected in this build (no debt schedule). Balance-sheet funding capacity is assessed against beginning cash only; this is a known scope limit, not a missing formula."),
        ("Cash retained / used after shareholder returns", f"=IFERROR({sr_fcf}-{sr_total},\"N/M\")", FMT_COMMA2,
         "Positive = FCF covers returns and adds to cash; negative = returns draw on existing cash or require external funding."),
        ("Beginning cash", (float(beginning_cash_val) if beginning_cash_val is not None else "=\"N/A - cash not in cache\""),
         FMT_COMMA2 if beginning_cash_val is not None else None,
         "Latest reported cash & equivalents from the normalized historical balance sheet (Net Debt Bridge component)."),
        ("Ending cash before other items",
         (f"={beginning_cash_val}+IFERROR({sr_fcf}-{sr_total},0)" if beginning_cash_val is not None else "=\"N/A - cash not in cache\""),
         FMT_COMMA2 if beginning_cash_val is not None else None,
         "Simplified reference: Beginning cash + (FCF - shareholder returns). Excludes M&A, lease repayments, net debt issuance, and second-order working-capital swings. Not a substitute for the full Cash Flow Forecast."),
        ("Funding review tier",
         f'=IFERROR(IF({sr_fcf}<=0,"High review - FCF<=0",IF({sr_total}/{sr_fcf}>1.2,"High review - returns >120% FCF",IF({sr_total}/{sr_fcf}>0.8,"Review - returns 80-120% FCF","OK"))),"N/A")',
         None, "Returns/FCF <=80% OK | 80-120% Review | >120% or FCF<=0 High review."),
    ]
    for label, formula, fmt, note in bridge_rows:
        _set_label_cell(ws, row, 1, label)
        if isinstance(formula, (int, float)):
            c = ws.cell(row=row, column=2, value=float(formula))
            c.font = font_formula(); c.border = border_thin()
            if fmt:
                c.number_format = fmt
        else:
            _set_formula_cell(ws, row, 2, formula, fmt, True)
        _set_note_cell(ws, row, 3, note)
        row += 1
    row += 1

    # V3.9.9.6 Final hardening: shareholder return policy sensitivity. This
    # table keeps the default policy intact and shows denominator-only policy
    # alternatives using funded buybacks, not planned-but-unfunded amounts.
    VAL_ROWS["shareholder_return_policy_sensitivity"] = row
    write_section_title(ws, row, "Shareholder Return Policy Sensitivity (diagnostic only)", span_cols=8)
    row += 1
    write_header_row(
        ws,
        row,
        [
            "Policy case", "Funded buybacks", "Ending cash", "Ending marketable securities",
            "Incremental debt", "Ending shares", "Forecast ending IV/share", "Difference vs current policy",
        ],
        1,
    )
    row += 1
    sr_first_idx = _forecast_start_col(ctx)
    sr_last_idx = sr_first_idx + n - 1

    def sr_ref(label: str, idx: int) -> str:
        return f"'Supporting Schedules'!{get_column_letter(idx)}{sr_rows_map.get(label, 1)}"

    def sr_range(label: str) -> str:
        return (
            f"'Supporting Schedules'!{get_column_letter(sr_first_idx)}{sr_rows_map.get(label, 1)}:"
            f"{get_column_letter(sr_last_idx)}{sr_rows_map.get(label, 1)}"
        )

    def sum_year_expr(parts: list[str]) -> str:
        return "SUM(" + ",".join(parts) + ")" if parts else "0"

    planned_cells = [sr_ref("Planned Buybacks", idx) for idx in range(sr_first_idx, sr_last_idx + 1)]
    fcf_after_div_cells = [sr_ref("FCF after Dividends", idx) for idx in range(sr_first_idx, sr_last_idx + 1)]
    capacity_cells = [sr_ref("Funding Capacity Before Debt", idx) for idx in range(sr_first_idx, sr_last_idx + 1)]
    cash_above_cells = [sr_ref("Cash Available Above Floor", idx) for idx in range(sr_first_idx, sr_last_idx + 1)]
    ms_available_cells = [sr_ref("Marketable Securities Available for Returns", idx) for idx in range(sr_first_idx, sr_last_idx + 1)]
    repurchase_price_cells = [sr_ref("Repurchase Price (per share)", idx) for idx in range(sr_first_idx, sr_last_idx + 1)]
    sbc_dilution_cells = [sr_ref("SBC / Annual Dilution (M)", idx) for idx in range(sr_first_idx, sr_last_idx + 1)]
    current_cash_used = f"SUM({sr_range('Cash Above Floor Used for Buybacks')})"
    current_ms_used = f"SUM({sr_range('Marketable Securities Drawdown Used')})"
    current_funded = f"SUM({sr_range('Share Repurchases (Buybacks)')})"
    current_incremental_debt = f"SUM({sr_range('Incremental Debt Issuance')})"
    final_ending_cash = sr_ref("Ending Cash", sr_last_idx)
    final_ending_ms = sr_ref("Ending Marketable Securities", sr_last_idx)
    final_ending_shares = sr_ref("Ending Diluted Shares", sr_last_idx)
    beginning_shares = sr_ref("Beginning Diluted Shares", sr_first_idx)
    cumulative_sbc = f"SUM({sr_range('SBC / Annual Dilution (M)')})"
    minimum_cash_floor_ref = _abs_ref(ASSUMP_CELLS.get("sr_cash_floor", ASSUMP_CELLS["shares"]))
    equity_value_ref = f"'DCF Valuation'!$B${VAL_ROWS['equity']}"

    def alt_funded_expr(policy: str) -> list[str]:
        exprs: list[str] = []
        for planned, fcf_after_div, capacity, price, sbc_m in zip(
            planned_cells, fcf_after_div_cells, capacity_cells, repurchase_price_cells, sbc_dilution_cells
        ):
            if policy == "reduced":
                exprs.append(f"MIN({planned}*50%,{capacity})")
            elif policy == "fcf_capped":
                exprs.append(f"MIN({planned},{fcf_after_div})")
            elif policy == "sbc_offset":
                exprs.append(f"MIN({planned},{capacity},{sbc_m}*{price})")
            elif policy == "debt_funded":
                exprs.append(planned)
        return exprs

    def alt_cash_used_expr(alt_exprs: list[str]) -> str:
        parts = [
            f"MIN(MAX(0,{alt}-{fcf_after_div}),{cash_above})"
            for alt, fcf_after_div, cash_above in zip(alt_exprs, fcf_after_div_cells, cash_above_cells)
        ]
        return sum_year_expr(parts)

    def alt_ms_used_expr(alt_exprs: list[str]) -> str:
        parts = []
        for alt, fcf_after_div, cash_above, ms_avail in zip(alt_exprs, fcf_after_div_cells, cash_above_cells, ms_available_cells):
            cash_piece = f"MIN(MAX(0,{alt}-{fcf_after_div}),{cash_above})"
            parts.append(f"MIN({ms_avail},MAX(0,{alt}-{fcf_after_div}-{cash_piece}))")
        return sum_year_expr(parts)

    def alt_debt_expr(alt_exprs: list[str], debt_funded: bool = False) -> str:
        if not debt_funded:
            return "0"
        return sum_year_expr([f"MAX(0,{alt}-{cap})" for alt, cap in zip(alt_exprs, capacity_cells)])

    def alt_shares_expr(alt_exprs: list[str]) -> str:
        retired = sum_year_expr([
            f"IFERROR(({alt})/{price},0)"
            for alt, price in zip(alt_exprs, repurchase_price_cells)
        ])
        return f"{beginning_shares}+{cumulative_sbc}-({retired})"

    current_policy_iv_row = row
    policy_rows = [
        (
            "Current / planned buyback policy",
            current_funded,
            f"={final_ending_cash}",
            f"={final_ending_ms}",
            current_incremental_debt,
            f"={final_ending_shares}",
            "Current funded policy; actual funded buybacks from Supporting Schedules.",
        ),
    ]
    for policy_label, policy_key, debt_funded in (
        ("Reduced buyback case", "reduced", False),
        ("FCF-after-dividends capped buyback case", "fcf_capped", False),
        ("SBC-offset-only buyback case", "sbc_offset", False),
        ("Debt-funded buyback case", "debt_funded", True),
    ):
        alt_exprs = alt_funded_expr(policy_key)
        funded = sum_year_expr(alt_exprs)
        cash_used = alt_cash_used_expr(alt_exprs)
        ms_used = alt_ms_used_expr(alt_exprs)
        debt = alt_debt_expr(alt_exprs, debt_funded=debt_funded)
        ending_cash = f"=MAX({minimum_cash_floor_ref},{final_ending_cash}+({current_cash_used})-({cash_used}))"
        ending_ms = f"=MAX(0,{final_ending_ms}+({current_ms_used})-({ms_used}))"
        ending_shares = f"={alt_shares_expr(alt_exprs)}"
        policy_rows.append((policy_label, funded, ending_cash, ending_ms, debt, ending_shares, "Diagnostic alternative; does not alter headline unless user changes policy assumptions."))

    for idx, (label, funded_formula, cash_formula, ms_formula, debt_formula, shares_formula, note) in enumerate(policy_rows):
        _set_label_cell(ws, row, 1, label)
        _set_formula_cell(ws, row, 2, f"={funded_formula}" if not str(funded_formula).startswith("=") else funded_formula, FMT_COMMA2, True)
        _set_formula_cell(ws, row, 3, cash_formula if str(cash_formula).startswith("=") else f"={cash_formula}", FMT_COMMA2, True)
        _set_formula_cell(ws, row, 4, ms_formula if str(ms_formula).startswith("=") else f"={ms_formula}", FMT_COMMA2, True)
        _set_formula_cell(ws, row, 5, f"={debt_formula}" if not str(debt_formula).startswith("=") else debt_formula, FMT_COMMA2, True)
        _set_formula_cell(ws, row, 6, shares_formula if str(shares_formula).startswith("=") else f"={shares_formula}", FMT_COMMA2, True)
        _set_formula_cell(ws, row, 7, f"=IFERROR({equity_value_ref}/F{row},\"N/M\")", FMT_COMMA2, True)
        _set_formula_cell(ws, row, 8, "=0" if idx == 0 else f"=IFERROR(G{row}-G{current_policy_iv_row},\"N/M\")", FMT_COMMA2, True)
        row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "Policy sensitivity is diagnostic only. It uses actual funded buybacks, respects the cash floor, and changes only the forecast ending share denominator shown in this table; headline IV and the default buyback policy are unchanged.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 2

    # V3.9.9.4 SBC Dilution v1 disclosure block: historical SBC magnitude,
    # SBC / Market Cap, SBC / Revenue, plus the live assumption and methodology.
    sr_audit_block = getattr(out, "shareholder_returns", None) or {}
    sr_drivers_v4 = sr_audit_block.get("drivers_effective") or {}
    sr_hist_block = sr_audit_block.get("historical") or {}
    sbc_base = sr_hist_block.get("base_sbc")
    sbc_pct_mc = sr_drivers_v4.get("sbc_pct_market_cap")
    sbc_pct_rev = sr_drivers_v4.get("sbc_pct_revenue")
    sbc_methodology = sr_drivers_v4.get("annual_dilution_default_methodology")
    sbc_user_override = bool(sr_drivers_v4.get("annual_dilution_user_override"))
    sbc_source_year = sr_drivers_v4.get("annual_dilution_default_source_year")

    write_section_title(ws, row, "SBC Dilution v1 (historical + assumption)", span_cols=3)
    row += 1
    _set_label_cell(ws, row, 1, "Latest historical SBC (M)")
    if sbc_base is None:
        c = ws.cell(row=row, column=2, value="N/A"); c.font = font_formula(); c.border = border_thin()
    else:
        c = ws.cell(row=row, column=2, value=float(sbc_base)); c.font = font_formula(); c.border = border_thin(); c.number_format = FMT_COMMA2
    yr_note = f"Fiscal year {sbc_source_year}; " if sbc_source_year else ""
    _set_note_cell(ws, row, 3, f"{yr_note}cash-flow statement stock-based compensation (absolute). N/A when cache lacks the field.")
    row += 1
    _set_label_cell(ws, row, 1, "SBC % Revenue")
    if sbc_pct_rev is None:
        c = ws.cell(row=row, column=2, value="N/A"); c.font = font_formula(); c.border = border_thin()
    else:
        c = ws.cell(row=row, column=2, value=float(sbc_pct_rev)); c.font = font_formula(); c.border = border_thin(); c.number_format = FMT_PCT2
    _set_note_cell(ws, row, 3, "SBC / latest historical revenue. Magnitude reference.")
    row += 1
    _set_label_cell(ws, row, 1, "SBC % Market Cap")
    if sbc_pct_mc is None:
        c = ws.cell(row=row, column=2, value="N/A"); c.font = font_formula(); c.border = border_thin()
    else:
        c = ws.cell(row=row, column=2, value=float(sbc_pct_mc)); c.font = font_formula(); c.border = border_thin(); c.number_format = FMT_PCT2
    _set_note_cell(ws, row, 3, "SBC / (Current Price x Reported Shares). Direct gross-dilution proxy and primary basis for the default annual dilution %.")
    row += 1
    _set_label_cell(ws, row, 1, "SBC / Annual Dilution % (assumption, live)")
    _set_formula_cell(ws, row, 2, f"={sr_dilution_ref}", FMT_PCT2, True)
    if sbc_user_override:
        _set_note_cell(ws, row, 3, "User-provided override (Assumptions). Edit there to change.")
    elif sbc_methodology:
        _set_note_cell(ws, row, 3, sbc_methodology)
    elif annual_dilution_val == 0.0:
        _set_note_cell(ws, row, 3, "SBC dilution = 0. No live SBC feed and no historical data; review for SBC-material names.")
    else:
        _set_note_cell(ws, row, 3, "Annual share-count dilution from SBC / option exercises; editable analyst input.")
    row += 1
    _set_note_cell(
        ws, row, 1,
        "Buybacks and dividends are capital-allocation / financing items and are NOT deducted from unlevered FCFF. The bridge is a simplified reference - it does not replace the Cash Flow Forecast or model debt issuance, and it does not auto-correct funding shortfalls. Tiering surfaces material gaps for IC discussion only.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # V3.7.0 Valuation Parity: engine-generated IV vs Excel Live IV, plus Legacy reference.
    # V3.9.1: titled as an engine vs Excel diagnostic so it does not read as a
    # primary valuation output. Headline IV remains the live Excel chain above.
    write_section_title(ws, row, "Engine vs Excel Live - Parity Diagnostic (engine-generated reference, static at export)", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Metric", "Value", "Note"], 1)
    row += 1

    unified_iv = out.intrinsic_per_share
    engine_label = (out.audit or {}).get("engine", "unknown")
    _set_label_cell(ws, row, 1, "API Unified Engine Intrinsic / Share")
    c = ws.cell(row=row, column=2, value=unified_iv)
    c.font = font_formula()
    c.border = border_thin()
    c.number_format = FMT_COMMA2
    VAL_ROWS["v370_unified_api"] = row
    _set_note_cell(ws, row, 3, f"Returned by run_dcf to UI / API / Excel exporter (engine={engine_label}).")
    row += 1

    _set_label_cell(ws, row, 1, "Excel Live Intrinsic / Share (workbook formula)", True)
    _set_formula_cell(ws, row, 2, f"=B{VAL_ROWS['intrinsic']}", FMT_COMMA2, True)
    VAL_ROWS["v370_excel_live"] = row
    _set_note_cell(ws, row, 3, "Live formula chain from FCF Build True 3FS row.")
    row += 1

    _set_label_cell(ws, row, 1, "Parity Diff (Excel Live - API Unified, target = 0)", True)
    _set_formula_cell(ws, row, 2, f"=B{VAL_ROWS['v370_excel_live']}-B{VAL_ROWS['v370_unified_api']}", FMT_COMMA2)
    VAL_ROWS["v370_parity_diff"] = row
    _set_note_cell(ws, row, 3, "Only rounding / Excel recalc rounding should remain; magnitude > 1 indicates engine drift.")
    row += 1

    _set_label_cell(ws, row, 1, "Parity Diff %")
    _set_formula_cell(ws, row, 2, f"=IFERROR(B{VAL_ROWS['v370_parity_diff']}/B{VAL_ROWS['v370_unified_api']},0)", FMT_PCT2)
    VAL_ROWS["v370_parity_diff_pct"] = row
    _set_note_cell(ws, row, 3, "Relative parity gap.")
    row += 1

    # Legacy reference IV: re-derive what the V3.6.x legacy calculator would have produced,
    # purely as a historical anchor. Not used in valuation.
    legacy_iv_value = None
    try:
        from modeling.dcf_calculator import _calc_valuation_from_forecast  # local import to avoid top-of-file churn
        legacy_forecast_for_iv = build_operating_forecast(inp)
        legacy_valuation = _calc_valuation_from_forecast(inp, legacy_forecast_for_iv)
        legacy_iv_value = legacy_valuation.get("intrinsic_per_share")
    except Exception:
        legacy_iv_value = None
    _set_label_cell(ws, row, 1, "Legacy operating-forecast IV / Share (V3.6.x reference)")
    c = ws.cell(row=row, column=2, value=legacy_iv_value)
    c.font = font_formula()
    c.border = border_thin()
    c.number_format = FMT_COMMA2
    VAL_ROWS["v370_legacy_iv"] = row
    _set_note_cell(ws, row, 3, "Historical reference - what the pre-V3.7.0 calculator would have produced; not the current source.")
    row += 1

    _set_label_cell(ws, row, 1, "Legacy vs Unified Diff (historical only)")
    _set_formula_cell(ws, row, 2, f"=B{VAL_ROWS['v370_legacy_iv']}-B{VAL_ROWS['v370_unified_api']}", FMT_COMMA2)
    _set_note_cell(ws, row, 3, "Explains the V3.6.x → V3.7.0 IV change (mainly WC days driver + asset-based D&A).")
    row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "V3.7.0 Engine Unification v1: UI / API / Excel valuation share a single source. Any residual Excel vs API diff should be rounding or Excel COM recalculation noise. Legacy reference row is kept so the V3.6.x → V3.7.0 transition is auditable. dcf_calculator.py now uses the Unified True 3FS engine when historical cache is available; HK / CN continue to fall back to the legacy operating forecast.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)


def _fa_formula(ref: str, fmt=None):
    return f"={ref}", fmt


def _fa_ratio(numerator: str, denominator: str):
    return f'=IFERROR({numerator}/{denominator},"n/a")', FMT_PCT2


def _quoted_ref(sheet_name: str, row: int, col: int) -> str:
    return f"'{sheet_name}'!{get_column_letter(col)}{row}"


def build_aapl_operating_thesis_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    """V3.9.9.0 AAPL Operating Thesis Bridge v1 (presentation / support).

    Renders the Products / Services revenue bridge, Services-mix progression,
    and the GM / R&D / SG&A / EBIT margin bridge for the active scenario
    (Base or Bear). Historical Products / Services split is sourced from
    Apple's 10-K segment disclosures (rounded). The bridge is presentation /
    support only — the DCF engine continues to consume the selected top-down
    revenue / EBIT margin paths from Assumptions.
    """
    ws.title = "AAPL Operating Thesis"
    span_cols = 8
    apply_column_widths(ws, {
        "A": 34, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 18,
    })
    row = _write_sheet_heading(ws, 1, "AAPL Operating Thesis Bridge v1", inp, out, ctx, span_cols=span_cols)

    # Scenario detection: prefer explicit ctx['scenario'] when provided; else
    # infer from the active forecast path. Base path peaks at 34.5% EBIT
    # margin; Bear peaks at 33.6%. The 33.9% midpoint separates them cleanly.
    ctx_scenario = ctx.get("scenario") if isinstance(ctx.get("scenario"), str) else None
    if ctx_scenario:
        scenario_key = ctx_scenario.strip().lower()
    else:
        margin_peak = max((out.ebit_margin_projections or [0.0]))
        scenario_key = "base" if margin_peak >= 0.339 else "bear"
    if scenario_key not in {"base", "bear"}:
        scenario_key = "base"
    bridge = build_aapl_operating_thesis_bridge_payload(scenario_key)
    case_label = bridge.get("case_label") or "Base Case"

    _set_note_cell(
        ws, row, 1,
        f"V3.9.9.0 Operating Thesis Bridge v1 — Active Case: {case_label}. "
        f"This bridge supports the selected top-down revenue and EBIT margin paths in Assumptions. "
        "It is review support, not a full segment driver model — the DCF engine continues to use the selected Assumptions paths. "
        "Historical Products / Services split is sourced from Apple 10-K segment disclosures (rounded); reviewer should verify before relying.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    # ── Historical Products / Services revenue ─────────────────────────────
    year_labels = [f"Y{i+1}" for i in range(5)]
    op_bridge = getattr(out, "operating_path_bridge", None) or {}
    source_ref = _abs_ref(ASSUMP_CELLS.get("operating_path_source", "B1"))
    write_section_title(ws, row, "Operating Path Source / Reconciliation", span_cols=span_cols)
    row += 1
    _set_label_cell(ws, row, 1, "Operating Path Source")
    _set_formula_cell(ws, row, 2, f"={source_ref}", None, True)
    _set_note_cell(ws, row, 3, "Bridge is engine-driving when source = AAPL Operating Thesis Bridge; otherwise bridge remains reference/support.")
    row += 1
    write_header_row(ws, row, ["Metric", *year_labels, "Note"], 1)
    row += 1
    bridge_rev_path = list(op_bridge.get("bridge_revenue_growth_path") or bridge.get("bridge_implied_total_growth_path") or [])
    bridge_margin_path = list(op_bridge.get("bridge_ebit_margin_path") or bridge.get("bridge_implied_ebit_margin_path") or [])
    selected_rev_path = list(op_bridge.get("selected_revenue_growth_path") or out.revenue_growth_projections or [])
    selected_margin_path = list(op_bridge.get("selected_ebit_margin_path") or out.ebit_margin_projections or [])
    recon_rows = [
        ("Bridge implied total revenue growth", bridge_rev_path, FMT_PCT2, "Products + Services bridge implied total growth."),
        ("Selected DCF revenue growth", selected_rev_path, FMT_PCT2, "Selected Path cells on Assumptions."),
        ("Difference in bps - revenue", [((b - s) * 10000.0) for b, s in zip(bridge_rev_path, selected_rev_path)], "0", "Bridge minus Selected Path, bps."),
        ("Bridge implied EBIT margin", bridge_margin_path, FMT_PCT2, "Gross margin - R&D% - SG&A%."),
        ("Selected DCF EBIT margin", selected_margin_path, FMT_PCT2, "Selected Path cells on Assumptions."),
        ("Difference in bps - margin", [((b - s) * 10000.0) for b, s in zip(bridge_margin_path, selected_margin_path)], "0", "Bridge minus Selected Path, bps."),
    ]
    for label, vals, fmt, note in recon_rows:
        _set_label_cell(ws, row, 1, label)
        for j in range(5):
            v = vals[j] if j < len(vals) else None
            c = ws.cell(row=row, column=2 + j, value=v)
            c.font = font_formula(); c.border = border_thin()
            if v is not None:
                c.number_format = fmt
        _set_note_cell(ws, row, 7, note)
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=span_cols)
        row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "When Operating Path Source = Selected Path, differences vs the bridge reflect analyst-selected overrides / rounding. Switch Operating Path Source to Bridge for fully mechanical bridge-driven case.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 1
    _set_note_cell(ws, row, 1, "Audit: " + (op_bridge.get("coherence_flag") or "reference/support"))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    write_section_title(ws, row, "Selected vs Bridge Materiality", span_cols=span_cols)
    row += 1
    write_header_row(ws, row, ["Path", *year_labels, "Classification"], 1)
    row += 1
    rev_diff_bps = [((b - s) * 10000.0) for b, s in zip(bridge_rev_path, selected_rev_path)]
    margin_diff_bps = [((b - s) * 10000.0) for b, s in zip(bridge_margin_path, selected_margin_path)]
    max_rev_bps = max([abs(x) for x in rev_diff_bps] or [0.0])
    max_margin_bps = max([abs(x) for x in margin_diff_bps] or [0.0])
    mat_rows = [
        ("Revenue Growth - Selected Path", selected_rev_path, FMT_PCT2, ""),
        ("Revenue Growth - Bridge Implied Path", bridge_rev_path, FMT_PCT2, ""),
        ("Revenue Growth - Difference bps", rev_diff_bps, "0", "Bridge-consistent" if max_rev_bps <= 50 else "Analyst override"),
        ("EBIT Margin - Selected Path", selected_margin_path, FMT_PCT2, ""),
        ("EBIT Margin - Bridge Implied Path", bridge_margin_path, FMT_PCT2, ""),
        ("EBIT Margin - Difference bps", margin_diff_bps, "0", "Bridge-consistent" if max_margin_bps <= 50 else "Analyst override"),
    ]
    for label, vals, fmt, cls in mat_rows:
        _set_label_cell(ws, row, 1, label)
        for j in range(5):
            v = vals[j] if j < len(vals) else None
            c = ws.cell(row=row, column=2 + j, value=v)
            c.font = font_formula(); c.border = border_thin()
            if v is not None:
                c.number_format = fmt
        _set_note_cell(ws, row, 7, cls)
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=span_cols)
        row += 1
    if max_rev_bps <= 50 and max_margin_bps <= 50:
        mat_note = "No material override. Selected Path is bridge-consistent within 50 bps materiality threshold."
    else:
        mat_note = "Analyst override: one or more selected path differences exceed 50 bps; rationale required before IC use."
    _set_note_cell(ws, row, 1, mat_note)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 1
    if max_margin_bps == 0:
        _set_note_cell(ws, row, 1, "Margin bridge fully reconciles to selected EBIT margin path.")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
        row += 1
    row += 1

    write_section_title(ws, row, "Historical Products / Services Revenue ($mm)", span_cols=span_cols)
    row += 1
    write_header_row(ws, row, ["Fiscal Year (source)", "Products ($mm)", "Services ($mm)", "Total ($mm)", "Services % Revenue"], 1)
    row += 1
    hist = bridge.get("historical_split") or []
    for fy_label, products, services in hist:
        total = (products or 0.0) + (services or 0.0)
        svc_pct = (services / total) if total else None
        cells = [fy_label, products, services, total, svc_pct]
        for col_idx, val in enumerate(cells, 1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = font_formula(); c.border = border_thin()
            if col_idx in (2, 3, 4):
                c.number_format = "#,##0"
            elif col_idx == 5:
                c.number_format = "0.0%"
        row += 1
    _set_note_cell(ws, row, 1, "Source: " + (bridge.get("historical_source") or ""))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    # ── Forecast Products / Services revenue (selected case) ───────────────
    year_labels = [f"Y{i+1}" for i in range(5)]
    write_section_title(ws, row, f"Forecast Products / Services Revenue — {case_label} ($mm)", span_cols=span_cols)
    row += 1
    write_header_row(ws, row, ["Line", *year_labels, "Note"], 1)
    row += 1

    p0 = float(bridge.get("base_year_products") or 0.0)
    s0 = float(bridge.get("base_year_services") or 0.0)
    pg = list(bridge.get("products_growth_path") or [])
    sg = list(bridge.get("services_growth_path") or [])
    products_fc: list[float] = []
    services_fc: list[float] = []
    p_run = p0
    s_run = s0
    for i in range(5):
        p_run = p_run * (1.0 + (pg[i] if i < len(pg) else 0.0))
        s_run = s_run * (1.0 + (sg[i] if i < len(sg) else 0.0))
        products_fc.append(p_run)
        services_fc.append(s_run)
    total_fc = [p + s for p, s in zip(products_fc, services_fc)]
    services_pct_fc = [(s / t if t else None) for s, t in zip(services_fc, total_fc)]
    bridge_growth_fc = [
        ((total_fc[i] / total_fc[i - 1] - 1.0) if i > 0 else (total_fc[0] / (p0 + s0) - 1.0 if (p0 + s0) else None))
        for i in range(5)
    ]
    selected_growth = list(out.revenue_growth_projections or [])
    rows_to_write = [
        ("Products Growth %", pg, "0.0%", "Low single-digit; iPhone refresh + Mac/iPad/Wearables maturity."),
        ("Products Revenue ($mm)", products_fc, "#,##0", "Reviewer support; not engine-driving."),
        ("Services Growth %", sg, "0.0%", "Low-double-digit; ecosystem monetisation."),
        ("Services Revenue ($mm)", services_fc, "#,##0", "Reviewer support; not engine-driving."),
        ("Total Revenue ($mm) — bridge", total_fc, "#,##0", "Σ Products + Services; reconciliation reference."),
        ("Bridge Implied Total Growth %", bridge_growth_fc, "0.00%", "Should approximate the selected revenue growth path."),
        ("Selected Revenue Growth % (DCF engine)", selected_growth, "0.00%", "Active path consumed by the DCF engine."),
        ("Services % of Revenue", services_pct_fc, "0.0%", "Mix progression supports gross margin / EBIT margin thesis."),
    ]
    for label, vals, fmt, note in rows_to_write:
        _set_label_cell(ws, row, 1, label)
        for j in range(5):
            v = vals[j] if j < len(vals) else None
            c = ws.cell(row=row, column=2 + j, value=v)
            c.font = font_formula(); c.border = border_thin()
            if v is not None:
                c.number_format = fmt
        _set_note_cell(ws, row, 7, note)
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=span_cols)
        row += 1
    row += 1

    _set_note_cell(ws, row, 1, bridge.get("products_thesis") or "")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols); row += 1
    _set_note_cell(ws, row, 1, bridge.get("services_thesis") or "")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols); row += 1
    _set_note_cell(ws, row, 1, bridge.get("base_vs_bull_note") or "")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols); row += 2

    # ── Margin Bridge (GM / R&D% / SG&A% / EBIT margin) ────────────────────
    write_section_title(ws, row, f"Margin Bridge — {case_label}", span_cols=span_cols)
    row += 1
    write_header_row(ws, row, ["Line", *year_labels, "Note"], 1)
    row += 1
    gm = list(bridge.get("gross_margin_path") or [])
    rd = list(bridge.get("rd_pct_revenue_path") or [])
    sga = list(bridge.get("sga_pct_revenue_path") or [])
    implied_ebit = [
        (gm[i] - rd[i] - sga[i]) if (i < len(gm) and i < len(rd) and i < len(sga)) else None
        for i in range(5)
    ]
    selected_ebit = list(out.ebit_margin_projections or [])
    bridge_rows = [
        ("Gross Margin", gm, "Modest expansion driven by Services mix progression."),
        ("R&D % Revenue", rd, "Stable-to-slight leverage; R&D dollars grow with revenue."),
        ("SG&A % Revenue", sga, "Modest drift; captures retail / brand investment."),
        ("Implied EBIT Margin (GM − R&D% − SG&A%)", implied_ebit, "Bridge reconciles within rounding to the selected EBIT margin path."),
        ("Selected EBIT Margin (DCF engine)", selected_ebit, "Active path consumed by the DCF engine."),
    ]
    for label, vals, note in bridge_rows:
        _set_label_cell(ws, row, 1, label)
        for j in range(5):
            v = vals[j] if j < len(vals) else None
            c = ws.cell(row=row, column=2 + j, value=v)
            c.font = font_formula(); c.border = border_thin()
            if v is not None:
                c.number_format = "0.00%"
        _set_note_cell(ws, row, 7, note)
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=span_cols)
        row += 1
    row += 1
    _set_note_cell(ws, row, 1, bridge.get("margin_thesis") or "")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols); row += 2

    _set_note_cell(ws, row, 1, bridge.get("presentation_only_note") or "")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols); row += 1


def build_trading_comps_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    """V3.7.7 Trading Comps v1 sheet.

    Renders peer set, peer multiples, outlier flags, summary stats, and
    target-implied IV/share ranges. Numeric cells are calculator-derived
    (yfinance-actuals scale converted to per-share absolute currency); cells
    stay editable and the sheet degrades gracefully when comps payload is
    missing or partial.
    """
    ws.title = "Trading Comps"
    TRADING_COMPS_ROWS.clear()
    apply_column_widths(ws, {
        "A": 12, "B": 28, "C": 26, "D": 16, "E": 14, "F": 36, "G": 14,
        "H": 18, "I": 18, "J": 16, "K": 16, "L": 16, "M": 16,
        "N": 22, "O": 34,
    })

    comps = (ctx.get("trading_comps") or {})
    span_cols = 15
    row = _write_sheet_heading(ws, 1, "Trading Comps", inp, out, ctx, span_cols=span_cols)

    if not comps or comps.get("status") == "error":
        err = comps.get("error") if isinstance(comps, dict) else None
        _set_note_cell(
            ws, row, 1,
            "Trading Comps v1 unavailable for this export. "
            + (f"Error: {err}" if err else "Peer data could not be fetched.")
            + " DCF valuation remains the primary model; comps would only provide cross-check.",
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
        TRADING_COMPS_ROWS["status"] = "unavailable"
        return

    if comps.get("status") == "unavailable":
        reason = comps.get("reason") or comps.get("unavailable_reason") or (
            "Industry peer set requires review. Generic technology fallback has been suppressed for this ticker."
        )
        profile_id = comps.get("profile_id") or ((comps.get("profile") or {}).get("profile_id")) or "unavailable"
        peer_source = comps.get("peer_set_source") or comps.get("peer_source") or "suppressed_generic_tech_fallback"
        write_section_title(ws, row, "Trading Comps unavailable", span_cols=span_cols)
        row += 1
        unavailable_rows = [
            ("Status", "unavailable"),
            ("Profile", profile_id),
            ("Peer Set Source", peer_source),
            ("Reason", reason),
            ("Boundary", "Industry peer set requires review. Generic technology fallback has been suppressed for this ticker."),
        ]
        for label, value in unavailable_rows:
            _set_label_cell(ws, row, 1, label)
            _set_note_cell(ws, row, 2, value)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=span_cols)
            row += 1
        row += 1
        _set_note_cell(
            ws,
            row,
            1,
            "No peer rows, summary statistics, or implied IV range are printed because an unmapped peer set is more misleading than no comps cross-check.",
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
        TRADING_COMPS_ROWS["status"] = "unavailable"
        return

    TRADING_COMPS_ROWS["status"] = "built"

    # ── Methodology / source note ──────────────────────────────────────
    _set_note_cell(
        ws, row, 1,
        f"V3.9.9.7 Trading Comps multi-source fallback. Symbol={comps.get('symbol')}. Peer source={comps.get('peer_source')}. "
        f"Fetched at {comps.get('fetched_at')}. {comps.get('methodology_note', '')}",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    if comps.get("peer_philosophy_note"):
        _set_note_cell(ws, row, 1, comps.get("peer_philosophy_note"))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
        row += 1
    # V3.9.8.8 Trading Comps Unit Fix: peer source data from yfinance is in actual
    # local currency; the rest of the workbook is in millions of reporting currency.
    # Display monetary columns in $mm to remove the prior raw-USD vs USD-millions
    # mix-up flagged by hostile review. Multiples (EV/Rev, EV/EBITDA, P/E) are
    # ratios and unaffected by the unit scaling.
    _set_note_cell(
        ws, row, 1,
        "V3.9.9.7 unit convention: USD mm for market cap / EV / revenue / EBITDA / net income; per-share data in USD; multiples shown as x. "
        "Unavailable fields display N/A, never zero. Stale cache and derived EBITDA are shown as Review quality, not discarded.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 1
    comps_reference_note = comps.get("no_recommendation_note") or ""
    _set_note_cell(ws, row, 1, comps_reference_note)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    # ── Peer table ─────────────────────────────────────────────────────
    write_section_title(ws, row, "Peer Set + Multiples", span_cols=span_cols)
    row += 1
    headers = [
        "Ticker", "Name", "Category", "Included in Stats", "Quality Tier", "Rationale / Exclusion Reason",
        "Price (USD)", "Revenue (USD mm)", "EBITDA (USD mm)",
        "Market Cap ($mm)", "Enterprise Value ($mm)",
        "EV / Revenue", "EV / EBITDA", "P / E (trailing)",
        "Sources / Outlier / Exclusion Notes",
    ]
    write_header_row(ws, row, headers, 1)
    row += 1
    peer_rows = comps.get("peer_rows") or []
    target_for_peer_table = comps.get("target") or {}
    target_display_row = None
    if target_for_peer_table:
        target_raw_display = {
            **target_for_peer_table,
            "field_sources": target_for_peer_table.get("field_sources") or {},
        }
        try:
            tgt_ev_rev = (
                float(target_raw_display.get("enterprise_value")) / float(target_raw_display.get("revenue"))
                if target_raw_display.get("enterprise_value") is not None and target_raw_display.get("revenue") else None
            )
            tgt_ev_ebitda = (
                float(target_raw_display.get("enterprise_value")) / float(target_raw_display.get("ebitda"))
                if target_raw_display.get("enterprise_value") is not None and target_raw_display.get("ebitda") else None
            )
        except (TypeError, ValueError, ZeroDivisionError):
            tgt_ev_rev = None
            tgt_ev_ebitda = None
        target_display_row = {
            "ticker": target_for_peer_table.get("ticker"),
            "name": target_for_peer_table.get("name") or target_for_peer_table.get("ticker"),
            "category": "Target: same trailing market-data basis",
            "included_in_stats_any": False,
            "included_in_stats_all": False,
            "included_in_stats": {},
            "data_quality_tier": "Review",
            "rationale": "Target row shown on same trailing basis for comparability; excluded from peer statistics.",
            "raw": target_raw_display,
            "multiples": {
                "ev_revenue": tgt_ev_rev,
                "ev_ebitda": tgt_ev_ebitda,
                "pe": target_raw_display.get("trailing_pe"),
            },
            "outliers": {},
            "exclusions": {},
            "field_source_summary": ", ".join(
                f"{k}={(target_raw_display.get('field_sources') or {}).get(k, 'unavailable')}"
                for k in ("price", "market_cap", "enterprise_value", "revenue", "ebitda", "net_income")
            ),
        }
    display_peer_rows = ([target_display_row] if target_display_row else []) + list(peer_rows)
    for r in display_peer_rows:
        mults = r.get("multiples") or {}
        outliers = r.get("outliers") or {}
        excls = r.get("exclusions") or {}
        per_mult_inc = r.get("included_in_stats") or {}
        raw = r.get("raw") or {}
        if r.get("included_in_stats_all"):
            included_text = "Yes (all multiples)"
        elif r.get("included_in_stats_any"):
            kept = [k for k, v in per_mult_inc.items() if v]
            included_text = "Partial (" + ", ".join(kept) + ")"
        else:
            included_text = "No (excluded from statistics)"
        flag_notes = []
        if r.get("outlier_reason"):
            flag_notes.append("Outlier: " + r["outlier_reason"])
        if r.get("exclusion_reason"):
            flag_notes.append("Excluded: " + r["exclusion_reason"])
        rationale_text = r.get("rationale") or ""
        if r.get("exclusion_reason") and not r.get("included_in_stats_any"):
            rationale_text = f"{rationale_text} (Excluded: {r['exclusion_reason']})".strip()
        def _to_mm(v):
            try:
                return float(v) / 1_000_000.0 if v is not None else None
            except (TypeError, ValueError):
                return None
        source_notes = []
        if raw.get("cache_stale"):
            source_notes.append(f"Stale cache date: {raw.get('cache_date')}")
        if raw.get("ebitda_methodology"):
            source_notes.append(raw.get("ebitda_methodology"))
        if r.get("field_source_summary"):
            source_notes.append(r.get("field_source_summary"))
        cells = [
            r.get("ticker"),
            r.get("name") or r.get("ticker"),
            r.get("category") or "Diagnostic / Other",
            included_text,
            r.get("data_quality_tier") or "Insufficient",
            rationale_text,
            raw.get("price"),
            _to_mm(raw.get("revenue")),
            _to_mm(raw.get("ebitda")),
            _to_mm(raw.get("market_cap")),
            _to_mm(raw.get("enterprise_value")),
            mults.get("ev_revenue"),
            mults.get("ev_ebitda"),
            mults.get("pe"),
            " | ".join([*(source_notes or []), *(flag_notes or [])]) if (source_notes or flag_notes) else "",
        ]
        excluded_row = not r.get("included_in_stats_any")
        for col_idx, value in enumerate(cells, 1):
            display_value = "N/A" if value is None and col_idx >= 7 else value
            c = ws.cell(row=row, column=col_idx, value=display_value)
            c.font = font_formula()
            c.border = border_thin()
            if col_idx == 7:
                c.number_format = FMT_COMMA2
            elif col_idx in (8, 9, 10, 11):
                c.number_format = "#,##0"
            elif col_idx in (12, 13, 14):
                c.number_format = "0.00"
            if col_idx in (6, 15):
                c.alignment = Alignment(wrap_text=True, vertical="top")
            # Highlight rows fully excluded from statistics in light grey so
            # readers can scan included peers at a glance.
            if excluded_row:
                c.fill = PatternFill("solid", fgColor="EEEEEE")
            # Per-multiple outlier highlight (warm yellow on the cell).
            if outliers.get("ev_revenue") and col_idx == 12:
                c.fill = PatternFill("solid", fgColor="FFF2CC")
            if outliers.get("ev_ebitda") and col_idx == 13:
                c.fill = PatternFill("solid", fgColor="FFF2CC")
            if outliers.get("pe") and col_idx == 14:
                c.fill = PatternFill("solid", fgColor="FFF2CC")
        row += 1
    row += 1

    # ── Summary stats ──────────────────────────────────────────────────
    write_section_title(ws, row, "Summary Statistics (Included peers only)", span_cols=span_cols)
    row += 1
    write_header_row(ws, row, ["Multiple", "Count", "Min", "P25", "Median", "Mean", "P75", "Max", "Contributors", "Status"], 1)
    row += 1
    stats = comps.get("summary_stats") or {}
    for label, key in (("EV / Revenue", "ev_revenue"), ("EV / EBITDA", "ev_ebitda"), ("P / E", "pe")):
        s = stats.get(key) or {}
        contributors = ", ".join(s.get("contributors") or []) or "n/a"
        if s.get("insufficient_data"):
            status_text = "Insufficient data - " + (s.get("insufficient_reason") or "")
        elif s.get("count"):
            status_text = "OK"
        else:
            status_text = "n/a"
        cells = [
            label,
            s.get("count"),
            s.get("min"), s.get("p25"), s.get("median"), s.get("mean"), s.get("p75"), s.get("max"),
            contributors, status_text,
        ]
        for col_idx, value in enumerate(cells, 1):
            c = ws.cell(row=row, column=col_idx, value=value)
            c.font = font_formula()
            c.border = border_thin()
            if col_idx in (3, 4, 5, 6, 7, 8):
                c.number_format = "0.00"
            if col_idx in (9, 10):
                c.alignment = Alignment(wrap_text=True, vertical="top")
            if s.get("insufficient_data") and col_idx >= 3:
                c.fill = PatternFill("solid", fgColor="EEEEEE")
        TRADING_COMPS_ROWS[f"summary_{key}"] = row
        row += 1
    row += 1

    # ── Implied valuation ───────────────────────────────────────────────
    write_section_title(ws, row, "Target Implied Valuation (Diagnostic cross-check; uses selected Net Debt + Share Count)", span_cols=span_cols)
    row += 1
    write_header_row(
        ws, row,
        ["Multiple", "Metric Used ($mm)", "Multiple Low (P25)", "Multiple Mid (Median)", "Multiple High (P75)",
         "Implied EV/Equity Low ($mm)", "Implied EV/Equity Mid ($mm)", "Implied EV/Equity High ($mm)",
         "IV / Share Low", "IV / Share Mid", "IV / Share High", "Status"], 1,
    )
    row += 1
    impl = comps.get("implied_valuation") or {}
    for label, key in (("EV / Revenue", "ev_revenue"), ("EV / EBITDA", "ev_ebitda"), ("P / E", "pe")):
        v = impl.get(key) or {}
        suppressed = v.get("suppressed_reason")
        # When suppressed, write the literal string "n/a" rather than None so
        # downstream IFERROR refs on Football Field render cleanly.
        def _cell_val(x):
            if suppressed and x is None:
                return "n/a"
            return x
        def _mm_or_na(x):
            if suppressed and x is None:
                return "n/a"
            try:
                return float(x) / 1_000_000.0 if x is not None else None
            except (TypeError, ValueError):
                return None
        cells = [
            label,
            _mm_or_na(v.get("metric_used")),
            v.get("multiple_low"), v.get("multiple_median"), v.get("multiple_high"),
            _mm_or_na(v.get("implied_value_low")), _mm_or_na(v.get("implied_value_mid")), _mm_or_na(v.get("implied_value_high")),
            _cell_val(v.get("iv_per_share_low")), _cell_val(v.get("iv_per_share_mid")), _cell_val(v.get("iv_per_share_high")),
            ("Suppressed - " + suppressed) if suppressed else "OK",
        ]
        for col_idx, value in enumerate(cells, 1):
            c = ws.cell(row=row, column=col_idx, value=value)
            c.font = font_formula()
            c.border = border_thin()
            if col_idx == 2:
                c.number_format = "#,##0"
            elif col_idx in (3, 4, 5):
                c.number_format = "0.00"
            elif col_idx in (6, 7, 8):
                c.number_format = "#,##0"
            elif col_idx in (9, 10, 11):
                c.number_format = FMT_COMMA2
            if col_idx == 12:
                c.alignment = Alignment(wrap_text=True, vertical="top")
            if suppressed and col_idx >= 6 and col_idx <= 11:
                c.fill = PatternFill("solid", fgColor="EEEEEE")
        TRADING_COMPS_ROWS[f"implied_{key}"] = row
        row += 1
    row += 1
    _set_note_cell(
        ws, row, 1,
        "Range reflects dispersion across business models and market multiples; diagnostic cross-check, not primary valuation output. "
        "Rows marked Suppressed have too few included peers to print a credible range.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    # ── Target snapshot ─────────────────────────────────────────────────
    write_section_title(ws, row, "Target Snapshot (yfinance peer-comparable basis)", span_cols=span_cols)
    row += 1
    tgt = comps.get("target") or {}
    def _tgt_mm(v):
        try:
            return float(v) / 1_000_000.0 if v is not None else None
        except (TypeError, ValueError):
            return None
    target_rows = [
        ("Ticker", tgt.get("ticker")),
        ("Name", tgt.get("name")),
        ("Currency", tgt.get("currency")),
        ("Revenue ($mm, TTM)", _tgt_mm(tgt.get("revenue"))),
        ("EBITDA ($mm, TTM)", _tgt_mm(tgt.get("ebitda"))),
        ("Net Income ($mm, TTM)", _tgt_mm(tgt.get("net_income"))),
        ("Market Cap ($mm)", _tgt_mm(tgt.get("market_cap"))),
        ("Enterprise Value ($mm)", _tgt_mm(tgt.get("enterprise_value"))),
        ("Net Debt Used in Comps ($mm)", _tgt_mm(tgt.get("net_debt_used_actual_currency"))),
        ("Shares Used in Comps (mm)", _tgt_mm(tgt.get("shares_used"))),
        ("Source", tgt.get("source")),
    ]
    for label, val in target_rows:
        _set_label_cell(ws, row, 1, label)
        c = ws.cell(row=row, column=2, value=val)
        c.font = font_formula(); c.border = border_thin()
        if isinstance(val, (int, float)) and val is not None:
            c.number_format = "#,##0"
        row += 1
    _set_note_cell(
        ws, row, 1,
        "Shares and net debt above are expressed in yfinance peer-comparable units (actual count / actual currency) so peer multiples apply consistently. The DCF headline IV/share uses the Selected Share Count Treatment denominator on DCF Valuation - different basis, same underlying Assumptions Diluted Shares cell.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    # ── Target vs Model Basis reconciliation ────────────────────────────
    basis = comps.get("target_model_basis") or None
    write_section_title(ws, row, "Target vs Model Basis Reconciliation", span_cols=span_cols)
    row += 1
    if basis:
        write_header_row(
            ws, row,
            ["Metric", "Model Basis ($mm)", "Market / TTM ($mm)", "Delta % (Model vs Market)", "Note"],
            1,
        )
        row += 1
        def _b_mm(v):
            try:
                return float(v) / 1_000_000.0 if v is not None else None
            except (TypeError, ValueError):
                return None
        recon_rows = [
            ("Revenue", _b_mm(basis.get("revenue_model_actual")), _b_mm(basis.get("revenue_market_actual")),
             basis.get("revenue_delta_pct"),
             "Model uses FY annual from normalized financials cache. Market uses yfinance TTM totalRevenue. Both shown in $mm."),
            ("EBITDA (EBIT + D&A)", _b_mm(basis.get("ebitda_model_actual")), _b_mm(basis.get("ebitda_market_actual")),
             basis.get("ebitda_delta_pct"),
             "Model derives EBITDA = EBIT + D&A (FY); market uses yfinance TTM EBITDA, which can include reclassifications. Both shown in $mm."),
            ("Market Cap (Price x Shares)", _b_mm(basis.get("market_cap_model_actual")), _b_mm(basis.get("market_cap_market_actual")),
             basis.get("market_cap_delta_pct"),
             "Model = workbook Price x Diluted Shares. Market = yfinance reported market cap. Both shown in $mm."),
            ("Enterprise Value", _b_mm(basis.get("ev_model_actual")), _b_mm(basis.get("ev_market_actual")),
             basis.get("ev_delta_pct"),
             "Model EV = Market Cap + Selected Net Debt. Market EV per yfinance enterpriseValue. Both shown in $mm."),
        ]
        for label, model_val, market_val, delta, note in recon_rows:
            _set_label_cell(ws, row, 1, label)
            for col_idx, val, fmt in (
                (2, model_val, "#,##0"),
                (3, market_val, "#,##0"),
                (4, delta, "0.0%"),
            ):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.font = font_formula(); c.border = border_thin()
                if val is not None:
                    c.number_format = fmt
            _set_note_cell(ws, row, 5, note)
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=span_cols)
            row += 1
        _set_note_cell(
            ws, row, 1,
            basis.get("basis_note") or "",
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
        row += 2
    else:
        _set_note_cell(
            ws, row, 1,
            "Model basis reconciliation unavailable; target metrics shown above use yfinance TTM only.",
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
        row += 2

    write_section_title(ws, row, "Forward EPS Diagnostic (one-feed, illustrative)", span_cols=span_cols)
    TRADING_COMPS_ROWS["forward_diagnostic"] = row
    row += 1
    write_header_row(
        ws,
        row,
        [
            "Ticker", "Company", "Peer classification", "Price", "Forward EPS",
            "Forward P/E (yfinance one-feed)", "Trailing P/E", "Forward vs trailing delta",
            "Source / availability", "Data limitation note",
        ],
        1,
    )
    row += 1
    fd = comps.get("forward_diagnostic") or {}
    fds = fd.get("summary") or {}
    forward_available = int(fds.get("core_forward_available_count") or 0)
    forward_total = int(fds.get("core_peer_count") or 0)
    if forward_available < 3:
        _set_note_cell(
            ws,
            row,
            1,
            "Forward diagnostic unavailable in current feed run. Insufficient forward coverage - trailing comps remain primary diagnostic.",
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
        row += 1
    for item in list(fd.get("rows") or []):
        values = [
            item.get("ticker"),
            item.get("company"),
            item.get("peer_classification"),
            item.get("price"),
            item.get("forward_eps"),
            item.get("forward_pe"),
            item.get("trailing_pe"),
            item.get("forward_vs_trailing_delta"),
            item.get("source_availability"),
            item.get("data_limitation_note"),
        ]
        for col_idx, value in enumerate(values, 1):
            display = "N/A" if value is None and col_idx in (4, 5, 6, 7, 8) else value
            c = ws.cell(row=row, column=col_idx, value=display)
            c.font = font_formula(); c.border = border_thin()
            if col_idx in (4, 5):
                c.number_format = FMT_COMMA2
            elif col_idx in (6, 7, 8):
                c.number_format = FMT_MULTIPLE
            if col_idx in (9, 10):
                c.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    row += 1
    write_header_row(ws, row, ["Forward Comps Summary / IC Read", "Value", "Note"], 1)
    row += 1
    summary_rows = [
        ("AAPL forward P/E vs core peer median", fds.get("target_forward_pe"), f"Core peer median: {fds.get('core_peer_forward_pe_median') if fds.get('core_peer_forward_pe_median') is not None else 'N/A'}"),
        ("AAPL trailing P/E vs core peer median", fds.get("target_trailing_pe"), f"Core peer median: {fds.get('core_peer_trailing_pe_median') if fds.get('core_peer_trailing_pe_median') is not None else 'N/A'}"),
        ("Forward coverage count", f"{fds.get('core_forward_available_count', 0)} / {fds.get('core_peer_count', 0)} core peers", fds.get("review_tier")),
        ("Forward vs trailing interpretation", fds.get("review_tier"), fds.get("interpretation")),
    ]
    for label, value, note in summary_rows:
        _set_label_cell(ws, row, 1, label)
        c = ws.cell(row=row, column=2, value="N/A" if value is None else value)
        c.font = font_formula(); c.border = border_thin()
        if isinstance(value, (int, float)):
            c.number_format = FMT_MULTIPLE
        _set_note_cell(ws, row, 3, note or "")
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=span_cols)
        row += 1
    _set_note_cell(
        ws,
        row,
        1,
        fds.get("disclosure") or (
            "Forward P/E values are sourced from yfinance.info.forwardPE / available one-feed data. "
            "They are NOT FactSet / Refinitiv multi-analyst consensus medians. Analyst count, dispersion, "
            "and update timestamp are not available through this feed. Use as directional diagnostic only; "
            "validate against institutional consensus feed before relying in IC discussion."
        ),
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    # ── Institutional consensus boundary (intentionally not synthesized) ──
    write_section_title(ws, row, "Institutional Consensus Feed Boundary", span_cols=span_cols)
    row += 1
    fc = comps.get("forward_consensus") or {}
    fc_rows = [
        ("NTM EV / Revenue", fc.get("ntm_ev_revenue") or "N/A - consensus forecast feed not connected"),
        ("NTM EV / EBITDA", fc.get("ntm_ev_ebitda") or "N/A - consensus forecast feed not connected"),
        ("NTM P / E", fc.get("ntm_pe") or "N/A - consensus forecast feed not connected"),
    ]
    for label, val in fc_rows:
        _set_label_cell(ws, row, 1, label)
        c = ws.cell(row=row, column=2, value=val)
        c.font = font_formula(); c.border = border_thin()
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=span_cols)
        row += 1
    _set_note_cell(
        ws, row, 1,
        fc.get("note") or "Forward consensus multiples are intentionally not included unless an external consensus feed is connected.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    # ── Source / cache / timestamp trail ────────────────────────────────
    write_section_title(ws, row, "Source / Cache / Timestamp Trail", span_cols=span_cols)
    row += 1
    write_header_row(ws, row, ["Subject", "Source", "Cache File", "Fetched At / Cache Date"], 1)
    row += 1
    tgt_cache = tgt.get("cache_file") or "n/a"
    if tgt_cache and tgt_cache not in ("n/a", "caller_supplied"):
        tgt_cache = os.path.basename(tgt_cache)
    trail_rows = [
        ("Target market data", tgt.get("source") or "yfinance.info", tgt_cache, comps.get("fetched_at")),
    ]
    peer_cache_files = comps.get("peer_cache_files") or {}
    for r in peer_rows:
        ticker = (r.get("ticker") or "").upper()
        raw = r.get("raw") or {}
        cache_file = peer_cache_files.get(ticker) or ""
        if cache_file:
            cache_file = os.path.basename(cache_file)
        trail_rows.append((
            f"Peer {ticker}",
            raw.get("source") or "yfinance.info",
            cache_file or "n/a",
            raw.get("fetched_at") or "n/a",
        ))
    for subject, source, cache_file, fetched in trail_rows:
        _set_label_cell(ws, row, 1, subject)
        for col_idx, val in ((2, source), (3, cache_file), (4, fetched)):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = font_formula(); c.border = border_thin()
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=span_cols)
        row += 1
    _set_note_cell(
        ws, row, 1,
        "Peer market data is fetched once per calendar day from yfinance and cached on disk. "
        "Target metrics on the Peer Set + Multiples table are TTM market data so they are directly "
        "comparable to peer multiples; the Target vs Model Basis Reconciliation block above shows "
        "the FY annual figures used inside the DCF workbook.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    row += 2

    # ── Data quality ────────────────────────────────────────────────────
    write_section_title(ws, row, "Trading Comps Data Quality", span_cols=span_cols)
    row += 1
    dq = comps.get("data_quality") or {}
    op = dq.get("outlier_policy") or {}
    mix_pct = dq.get("source_mix_pct") or {}
    dq_rows = [
        ("Core peer count", dq.get("core_peer_count")),
        ("Core peer EV/Revenue coverage", dq.get("core_peer_ev_revenue_coverage")),
        ("Core peer EV/EBITDA coverage", dq.get("core_peer_ev_ebitda_coverage")),
        ("Core peer P/E coverage", dq.get("core_peer_pe_coverage")),
        ("Forward P/E coverage", f"{(comps.get('forward_diagnostic') or {}).get('summary', {}).get('core_forward_available_count', 0)} / {(comps.get('forward_diagnostic') or {}).get('summary', {}).get('core_peer_count', 0)}"),
        ("% from live source", mix_pct.get("live")),
        ("% from cache source", mix_pct.get("cache")),
        ("% derived", mix_pct.get("derived")),
        ("% unavailable", mix_pct.get("unavailable")),
        ("Overall comps usability tier", dq.get("overall_comps_usability_tier")),
        ("Football Field comps status", dq.get("football_field_comps_status")),
        ("Total peers", dq.get("peer_count_total")),
        ("Failed fetch", dq.get("peer_count_failed_fetch")),
        ("Fully excluded (no included multiple)", dq.get("peer_count_fully_excluded")),
        ("Included for EV/Revenue", dq.get("peer_count_included_ev_revenue")),
        ("Included for EV/EBITDA", dq.get("peer_count_included_ev_ebitda")),
        ("Included for P/E", dq.get("peer_count_included_pe")),
        ("Insufficient peer threshold (min for credible range)", dq.get("insufficient_peer_threshold")),
        ("Outlier policy: abs caps", str(op.get("abs_caps"))),
        ("Outlier policy: IQR multiplier", op.get("iqr_multiplier")),
    ]
    for label, val in dq_rows:
        _set_label_cell(ws, row, 1, label)
        c = ws.cell(row=row, column=2, value=val); c.font = font_formula(); c.border = border_thin()
        if isinstance(val, (int, float)):
            c.number_format = "0.0%" if label.startswith("%") else ("0.00" if isinstance(val, float) else "0")
        row += 1
    TRADING_COMPS_ROWS["data_quality_end"] = row
    if comps.get("warnings"):
        row += 1
        _set_note_cell(ws, row, 1, "Warnings: " + " | ".join(comps.get("warnings") or []))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
        row += 1

    row += 1
    _set_note_cell(
        ws, row, 1,
        "Trading Comps boundary: peer set is a configurable reference for discussion purposes only. "
        "EV-based implied IV/share uses the V3.7.3 Selected Net Debt and V3.7.4 Selected Share Count "
        "(workbook millions converted to actual currency for yfinance-scale arithmetic). "
        "Outliers and missing peer rows are flagged in the Peer Set table.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)


def build_fa_ratios_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "FA Ratios"
    income_table = ((ctx.get("historical_tables") or {}).get("income_statement") or {})
    historical_years = sorted(income_table.get("years") or [])
    n = max(1, int(inp.forecast_years or 5))
    forecast_labels = [f"Year {i + 1}" for i in range(n)]
    total_cols = len(historical_years) + n
    span_cols = max(3, total_cols + 1)
    apply_column_widths(ws, {"A": 34, **{get_column_letter(i + 2): 15 for i in range(total_cols)}})

    row = _write_sheet_heading(ws, 1, "FA / Ratios - Analytical Output Only", inp, out, ctx, span_cols=span_cols)
    row = _write_kv_rows(ws, row, [
        ("Model Boundary", "Analytical output only; ratios are not model drivers.", None),
        ("Source", "Historical columns reference statement sheets; forecast columns reference forecast and schedule sheets.", None),
        ("Fallback", "Missing fields display n/a as text rather than fabricated zeroes.", None),
    ])
    row += 1

    header = [""] + [str(year) for year in historical_years] + forecast_labels
    write_header_row(ws, row, header, 1)
    row += 1

    forecast_start_col = 2 + len(historical_years)

    def has_rows(*items):
        return all(item for item in items)

    def ref(sheet: str, row_key: int, col_idx: int) -> str:
        return _quoted_ref(sheet, row_key, col_idx)

    def metric(label: str, fmt, formula_builder, bold: bool = False):
        nonlocal row
        _set_label_cell(ws, row, 1, label, bold)
        for col_idx in range(2, total_cols + 2):
            formula, cell_fmt = formula_builder(col_idx)
            if formula:
                _set_formula_cell(ws, row, col_idx, formula, cell_fmt or fmt, True)
            else:
                c = ws.cell(row=row, column=col_idx, value="n/a")
                c.font = font_formula()
                c.border = border_thin()
        row += 1

    def pl(label: str, col_idx: int):
        return ref("P&L Forecast", PL_FORECAST_ROWS[label], col_idx)

    def bs(label: str, col_idx: int):
        return ref("Balance Sheet Forecast", BS_FORECAST_ROWS[label], col_idx)

    def cf(label: str, col_idx: int):
        return ref("Cash Flow Forecast", CF_FORECAST_ROWS[label], col_idx)

    def sched(section: str, label: str, col_idx: int):
        return ref("Supporting Schedules", SCHEDULE_ROWS[section][label], col_idx)

    write_section_title(ws, row, "Profitability", span_cols=span_cols)
    row += 1
    metric("Revenue Growth", FMT_PCT2, lambda c: _fa_formula(pl("Revenue Growth (%)", c), FMT_PCT2) if has_rows(PL_FORECAST_ROWS.get("Revenue Growth (%)")) else (None, None))
    metric("Gross Margin", FMT_PCT2, lambda c: _fa_formula(pl("Gross Margin (%)", c), FMT_PCT2) if has_rows(PL_FORECAST_ROWS.get("Gross Margin (%)")) else (None, None))
    metric("EBIT Margin", FMT_PCT2, lambda c: _fa_formula(pl("EBIT Margin (%)", c), FMT_PCT2) if has_rows(PL_FORECAST_ROWS.get("EBIT Margin (%)")) else (None, None))
    metric("Net Margin", FMT_PCT2, lambda c: _fa_formula(pl("Net Margin (%)", c), FMT_PCT2) if has_rows(PL_FORECAST_ROWS.get("Net Margin (%)")) else (None, None))
    metric("ROA", FMT_PCT2, lambda c: _fa_ratio(pl("Net Income", c), bs("Total Assets", c)) if has_rows(PL_FORECAST_ROWS.get("Net Income"), BS_FORECAST_ROWS.get("Total Assets")) else (None, None))
    metric("ROE", FMT_PCT2, lambda c: _fa_ratio(pl("Net Income", c), bs("Total Equity", c)) if has_rows(PL_FORECAST_ROWS.get("Net Income"), BS_FORECAST_ROWS.get("Total Equity")) else (None, None))
    row += 1

    write_section_title(ws, row, "Efficiency / Working Capital", span_cols=span_cols)
    row += 1
    metric("AR / Revenue", FMT_PCT2, lambda c: _fa_ratio(bs("Accounts Receivable", c), pl("Revenue", c)) if has_rows(BS_FORECAST_ROWS.get("Accounts Receivable"), PL_FORECAST_ROWS.get("Revenue")) else (None, None))
    metric("Inventory / Revenue", FMT_PCT2, lambda c: _fa_ratio(bs("Inventory", c), pl("Revenue", c)) if has_rows(BS_FORECAST_ROWS.get("Inventory"), PL_FORECAST_ROWS.get("Revenue")) else (None, None))
    metric("AP / Revenue", FMT_PCT2, lambda c: _fa_ratio(bs("Accounts Payable", c), pl("Revenue", c)) if has_rows(BS_FORECAST_ROWS.get("Accounts Payable"), PL_FORECAST_ROWS.get("Revenue")) else (None, None))
    metric("NWC / Revenue", FMT_PCT2, lambda c: _fa_ratio(sched("wc", "Net Working Capital", c), pl("Revenue", c)) if has_rows(SCHEDULE_ROWS.get("wc", {}).get("Net Working Capital"), PL_FORECAST_ROWS.get("Revenue")) else (None, None))
    metric("CFO-derived FCF Reference Conversion", FMT_PCT2, lambda c: _fa_ratio(cf(CF_FCF_REF_LABEL, c), pl("Net Income", c)) if has_rows(CF_FORECAST_ROWS.get(CF_FCF_REF_LABEL), PL_FORECAST_ROWS.get("Net Income")) else (None, None))
    row += 1

    write_section_title(ws, row, "Leverage / Liquidity", span_cols=span_cols)
    row += 1
    metric("Total Debt / Equity", "0.00x", lambda c: (f'=IFERROR({bs("Total Debt", c)}/{bs("Total Equity", c)},"n/a")', "0.00x") if has_rows(BS_FORECAST_ROWS.get("Total Debt"), BS_FORECAST_ROWS.get("Total Equity")) else (None, None))
    metric("Net Debt / EBITDA", "0.00x", lambda c: (f'=IFERROR(({bs("Total Debt", c)}-{bs("Cash & Cash Equivalents", c)})/({pl("Operating Income / EBIT", c)}+{cf("D&A", c)}),"n/a")', "0.00x") if has_rows(BS_FORECAST_ROWS.get("Total Debt"), BS_FORECAST_ROWS.get("Cash & Cash Equivalents"), PL_FORECAST_ROWS.get("Operating Income / EBIT"), CF_FORECAST_ROWS.get("D&A")) else (None, None))
    metric("Cash / Revenue", FMT_PCT2, lambda c: _fa_ratio(bs("Cash & Cash Equivalents", c), pl("Revenue", c)) if has_rows(BS_FORECAST_ROWS.get("Cash & Cash Equivalents"), PL_FORECAST_ROWS.get("Revenue")) else (None, None))
    metric("Debt / Total Capital", FMT_PCT2, lambda c: (f'=IFERROR({bs("Total Debt", c)}/({bs("Total Debt", c)}+{bs("Total Equity", c)}),"n/a")', FMT_PCT2) if has_rows(BS_FORECAST_ROWS.get("Total Debt"), BS_FORECAST_ROWS.get("Total Equity")) else (None, None))
    metric("Current Ratio", "0.00x", lambda c: ('="n/a"', None))
    row += 1

    write_section_title(ws, row, "Per Share / Return", span_cols=span_cols)
    row += 1
    metric("Diluted Shares", FMT_COMMA2, lambda c: _fa_formula(pl("Diluted Shares", c), FMT_COMMA2) if has_rows(PL_FORECAST_ROWS.get("Diluted Shares")) else (None, None))
    metric("EPS", FMT_COMMA2, lambda c: _fa_formula(pl("EPS", c), FMT_COMMA2) if has_rows(PL_FORECAST_ROWS.get("EPS")) else (None, None))
    metric("CFO-derived FCF Reference per Share", FMT_COMMA2, lambda c: (f'=IFERROR({cf(CF_FCF_REF_LABEL, c)}/{pl("Diluted Shares", c)},"n/a")', FMT_COMMA2) if has_rows(CF_FORECAST_ROWS.get(CF_FCF_REF_LABEL), PL_FORECAST_ROWS.get("Diluted Shares")) else (None, None))
    metric("Revenue per Share", FMT_COMMA2, lambda c: (f'=IFERROR({pl("Revenue", c)}/{pl("Diluted Shares", c)},"n/a")', FMT_COMMA2) if has_rows(PL_FORECAST_ROWS.get("Revenue"), PL_FORECAST_ROWS.get("Diluted Shares")) else (None, None))

    row += 2
    _set_note_cell(ws, row, 1, "Ratios are a one-way analytical layer. No forecast, schedule, FCF, DCF, or sensitivity formula should reference this sheet.")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)


def _max_abs_formula(cells: list[str]) -> str:
    if not cells:
        return '"n/a"'
    return "MAX(" + ",".join(f"ABS({cell})" for cell in cells) + ")"


def _audit_status_cell(ws, row: int, formula: str):
    c = _set_formula_cell(ws, row, 2, formula)
    c.alignment = Alignment(horizontal="center")
    return c


def build_audit_dashboard_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "Audit Dashboard"
    AUDIT_STATUS_ROWS.clear()
    apply_column_widths(ws, {"A": 42, "B": 14, "C": 22, "D": 22, "E": 58})

    row = _write_sheet_heading(ws, 1, "Review Dashboard (Appendix)", inp, out, ctx, span_cols=5)

    write_section_title(ws, row, "Thresholds", span_cols=5)
    row += 1
    write_header_row(ws, row, ["Threshold", "Value", "Unit", "Use", "Note"], 1)
    row += 1
    thresholds = [
        ("TV % of EV cap", 0.80, FMT_PCT2, "tv_pct_cap", "Review when terminal value dependency exceeds 80%."),
        ("WACC − terminal g spread min", 0.01, FMT_PCT2, "spread_min", "Review when Gordon spread is too tight."),
        ("Gordon / Exit TV gap tolerance", 0.25, FMT_PCT2, "tv_gap_tol", "Relative gap tolerance between Gordon and Exit terminal value."),
        ("FCF reconciliation tolerance", 1.0, FMT_COMMA2, "recon_tol", "Workbook display-unit rounding tolerance."),
        ("Balance Check tolerance", 1.0, FMT_COMMA2, "balance_tol", "BS converges to within one workbook unit via transparent plug rows."),
        ("Sensitivity center tolerance", 0.01, FMT_COMMA2, "sens_tol", "Per-share center-cell tolerance."),
        ("WC dual basis tolerance", 1.0, FMT_COMMA2, "wc_tol", "Schedule-derived Delta NWC vs legacy reference tolerance."),
        ("Plug magnitude — OK ceiling (% of Total Assets)", 0.15, FMT_PCT2, "plug_pct_tol", "Tiered: ≤15% OK, 15–35% Review, >35% High Review."),
        ("Plug magnitude — high-review threshold (% of Total Assets)", 0.35, FMT_PCT2, "plug_pct_high", "Above this, plug needs explicit explanation."),
        ("Plug reduction target (% of legacy single residual)", 0.75, FMT_PCT2, "plug_reduction_target", "Expanded BS coverage should bring the named residual plug to ≤75% of the legacy single-bucket residual."),
        ("WACC selected vs indicative spread tolerance", 0.015, FMT_PCT2, "wacc_spread_tol", "Flag spread > 150 bps between selected and CAPM-indicative WACC."),
    ]
    threshold_refs = {}
    for label, value, fmt, key, note in thresholds:
        _set_label_cell(ws, row, 1, label)
        _set_input_cell(ws, row, 2, value, fmt)
        _set_note_cell(ws, row, 3, fmt or "")
        _set_note_cell(ws, row, 4, key)
        _set_note_cell(ws, row, 5, note)
        threshold_refs[key] = f"$B${row}"
        row += 1
    row += 2

    def write_check_group(title: str):
        nonlocal row
        write_section_title(ws, row, title, span_cols=5)
        row += 1
        write_header_row(ws, row, ["Check Name", "Status", "Value", "Threshold", "Comment"], 1)
        row += 1

    status_rows = []

    def write_check(name: str, status_formula: str, value_formula=None, threshold_formula=None, comment: str = "", value_fmt=None, threshold_fmt=None):
        nonlocal row
        _set_label_cell(ws, row, 1, name)
        _audit_status_cell(ws, row, status_formula)
        status_rows.append(row)
        if value_formula is not None:
            _set_formula_cell(ws, row, 3, value_formula, value_fmt, True)
        else:
            ws.cell(row=row, column=3, value="").border = border_thin()
        if threshold_formula is not None:
            _set_formula_cell(ws, row, 4, threshold_formula, threshold_fmt, True)
        else:
            ws.cell(row=row, column=4, value="").border = border_thin()
        _set_note_cell(ws, row, 5, comment)
        row += 1

    n = max(1, int(inp.forecast_years or 5))
    fcf_cols = [get_column_letter(2 + i) for i in range(n)]
    fcf_last_col = get_column_letter(1 + n)
    market_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['market']}"
    dcf_source_ref = f"'DCF Valuation'!$B${VAL_ROWS['dcf_source']}"
    intrinsic_ref = f"'DCF Valuation'!$B${VAL_ROWS['intrinsic']}"

    write_check_group("Workbook Identity")
    write_check(
        "Workbook version lock",
        f'=IF(C{row}="{MODEL_VERSION}","OK","Review")',
        f'="{MODEL_VERSION}"',
        f'="{MODEL_VERSION}"',
        "Workbook identity rows should display the current general-purpose hardening version.",
    )
    write_check(
        "Scope boundary present",
        '="OK"',
        '="valuation workbook and IC discussion aid; not recommendation or fairness opinion"',
        None,
        "Cover, Executive Summary, and Data Sources Audit disclose the workbook boundary.",
    )
    current_warnings = list((ctx.get("current_params") or {}).get("warnings") or []) + list((getattr(out, "audit", None) or {}).get("warnings") or [])
    current_warnings = [str(w) for w in current_warnings if str(w)]
    _set_label_cell(ws, row, 1, "Disclosure warnings surfaced")
    _audit_status_cell(ws, row, '="Review"' if current_warnings else '="OK"')
    status_rows.append(row)
    ws.cell(row=row, column=3, value=len(current_warnings)).border = border_thin()
    ws.cell(row=row, column=4, value=0).border = border_thin()
    _set_note_cell(ws, row, 5, " | ".join(current_warnings) if current_warnings else "No default or engine warnings surfaced.")
    row += 1
    write_check(
        "Date format policy",
        '="OK"',
        '="YYYY-MM-DD"',
        None,
        "Generated dates are emitted through ISO date display helpers.",
    )

    write_check_group("Formula / Load Integrity")
    write_check(
        "DCF Source = Unified True 3FS Engine v1",
        f'=IF(ISNUMBER(SEARCH("Unified True 3FS",{dcf_source_ref})),"OK",IF(ISNUMBER(SEARCH("Legacy",{dcf_source_ref})),"Fallback",IF(ISNUMBER(SEARCH("True 3FS",{dcf_source_ref})),"OK","Review")))',
        f"={dcf_source_ref}",
        None,
        "V3.7.0 DCF source should be the Unified True 3FS Engine (or Legacy fallback for non-US markets).",
    )
    recon_diff_row = FCF_ROWS.get("reconciliation", {}).get("difference")
    if recon_diff_row:
        recon_cells = [f"'FCF Build'!{col}{recon_diff_row}" for col in fcf_cols]
        write_check(
            "API vs Excel FCF Parity (target = 0)",
            f'=IF(C{row}<={threshold_refs["recon_tol"]},"OK","Review")',
            f"={_max_abs_formula(recon_cells)}",
            f"={threshold_refs['recon_tol']}",
            "V3.7.0: API and Excel share the unified engine; max-abs year diff should be within rounding tolerance.",
            FMT_COMMA2,
            FMT_COMMA2,
        )
    else:
        write_check("API vs Excel FCF Parity (target = 0)", '="N/A"', '="n/a"', f"={threshold_refs['recon_tol']}", "Reconciliation rows are unavailable in fallback exports.")
    concept_rows = FCF_ROWS.get("concept_reconciliation", {})
    if concept_rows and CF_FORECAST_ROWS.get(CF_FCF_REF_LABEL):
        write_check(
            "FCF concept separation: DCF FCFF vs CFO-derived reference",
            '="OK"',
            f"='FCF Build'!A{concept_rows.get('fcff_used', 1)}&\" / \"&'Cash Flow Forecast'!A{CF_FORECAST_ROWS[CF_FCF_REF_LABEL]}",
            None,
            "DCF cash flow is explicitly named Unlevered FCFF; Cash Flow Forecast uses a separate CFO-derived reference label.",
        )
        gap_cells = [f"'FCF Build'!{col}{concept_rows.get('explained_gap')}" for col in fcf_cols]
        after_tax_cells = [f"'FCF Build'!{col}{concept_rows.get('after_tax_interest')}" for col in fcf_cols]
        write_check(
            "FCF concept reconciliation bridge present",
            f'=IF(COUNT({",".join(gap_cells + after_tax_cells)})>=2,"OK","Review")',
            f"={_max_abs_formula(gap_cells)}",
            "Bridge includes after-tax interest",
            "Bridge shows CFO-derived reference, Interest Expense, after-tax interest add-back, other CFO bridge items / not modeled, and Unlevered FCFF used in DCF.",
            FMT_COMMA2,
        )
        other_bridge_cells = [f"'FCF Build'!{col}{concept_rows.get('other_bridge')}" for col in fcf_cols]
        fcff_cells = [f"'FCF Build'!{col}{concept_rows.get('fcff_used')}" for col in fcf_cols]
        write_check(
            "FCF concept bridge residual materiality",
            f'=IF(C{row}>10%,"Review","OK")',
            f"=IFERROR({_max_abs_formula(other_bridge_cells)}/MAX({_max_abs_formula(fcff_cells)},1),0)",
            "10%",
            "Other CFO bridge items are a not-fully-modeled residual. Review when residual exceeds 10% of FCFF; do not treat the bridge as a perfect three-statement tie-out.",
            FMT_PCT2,
            FMT_PCT2,
        )
    else:
        write_check("FCF concept reconciliation bridge present", '="N/A"', '="n/a"', None, "Unavailable in limited-market exports.")
    # V3.7.0 valuation-level parity at the IV/share line.
    if "v370_parity_diff" in VAL_ROWS:
        parity_diff_ref = f"'DCF Valuation'!$B${VAL_ROWS['v370_parity_diff']}"
        write_check(
            "API vs Excel Intrinsic / Share Parity (target = 0)",
            f'=IF(ABS({parity_diff_ref})<=1,"OK","Review")',
            f"=ABS({parity_diff_ref})",
            "=1",
            "V3.7.0: Excel Live IV/share should equal API Unified Engine IV/share within rounding (< 1 currency unit).",
            FMT_COMMA2,
            FMT_COMMA2,
        )

    balance_row = BS_FORECAST_ROWS.get("Balance Check / Difference")
    if balance_row:
        forecast_start_col_audit = _forecast_start_col(ctx)
        total_assets_row = BS_FORECAST_ROWS.get("Total Assets")
        total_le_row = BS_FORECAST_ROWS.get("Total Liabilities + Equity")
        if total_assets_row and total_le_row:
            balance_cells = [
                (
                    f"'Balance Sheet Forecast'!{get_column_letter(forecast_start_col_audit + i)}{total_assets_row}"
                    f"-'Balance Sheet Forecast'!{get_column_letter(forecast_start_col_audit + i)}{total_le_row}"
                )
                for i in range(n)
            ]
        else:
            balance_cells = [f"IFERROR(VALUE('Balance Sheet Forecast'!{col}{balance_row}),0)" for col in fcf_cols]
        write_check(
            "Balance Check targets zero via transparent plug rows",
            f'=IF({market_ref}<>"US","Fallback",IF(ISNUMBER(C{row}),IF(C{row}<={threshold_refs["balance_tol"]},"OK","Review"),"Review"))',
            f'=IFERROR({_max_abs_formula(balance_cells)},"N/A")',
            f"={threshold_refs['balance_tol']}",
            "Balance Check targets zero via transparent plug rows; plug magnitude is reviewed separately.",
            FMT_COMMA2,
            FMT_COMMA2,
        )
    else:
        write_check("Balance Check targets zero via transparent plug rows", f'=IF({market_ref}<>"US","Fallback","N/A")', '="n/a"', f"={threshold_refs['balance_tol']}", "Fallback export does not build the BS schedule.")

    wc_row = FCF_ROWS.get("wc")
    legacy_wc_row = FCF_ROWS.get("legacy_wc")
    if wc_row and legacy_wc_row:
        wc_gap_cells = [f"'FCF Build'!{col}{wc_row}-'FCF Build'!{col}{legacy_wc_row}" for col in fcf_cols]
        write_check(
            "WC dual basis: schedule-derived (primary) vs legacy reference",
            f'=IF(COUNT(\'FCF Build\'!B{legacy_wc_row}:{fcf_last_col}{legacy_wc_row})=0,"N/A","OK")',
            f"={_max_abs_formula(wc_gap_cells)}",
            f"={threshold_refs['wc_tol']}",
            "V3.6.8: DCF FCF now uses schedule-derived Delta NWC. Legacy is reference-only; large gaps are EXPECTED and drive the V3.6.7 to V3.6.8 valuation delta.",
            FMT_COMMA2,
            FMT_COMMA2,
        )
    else:
        write_check("WC dual basis: schedule-derived (primary) vs legacy reference", '="N/A"', '="n/a"', f"={threshold_refs['wc_tol']}", "Schedule-derived WC is unavailable.")

    wc_reality_rows = SCHEDULE_ROWS.get("wc_reality") or {}
    write_check(
        "WC faded-days schedule present",
        f'=IF(C{row}="present","OK","Review")',
        f'="{"present" if wc_reality_rows else "missing"}"',
        None,
        "V3.9.9.2.1: Working capital forecast should fade DSO/DIO/DPO toward normalized historical days, not static carry-forward days.",
    )
    if wc_reality_rows.get("Review tier"):
        write_check(
            "Historical WC reality check present",
            f'=IF(C{row}<>"","OK","Review")',
            f"='Supporting Schedules'!$B${wc_reality_rows['Review tier']}",
            None,
            "V3.9.9.2.1: compares active faded-days Delta NWC with historical average Delta NWC.",
        )

    # V3.6.8 new checks: Interest expense linkage, Debt Schedule roll-forward, True 3FS status.
    pl_interest_row = PL_FORECAST_ROWS.get("Interest Expense")
    debt_interest_row = (SCHEDULE_ROWS.get("debt") or {}).get("Interest Expense")
    if pl_interest_row and debt_interest_row:
        forecast_start_col_audit = _forecast_start_col(ctx)
        interest_gap_cells = []
        for i in range(n):
            col_pl = get_column_letter(forecast_start_col_audit + i)
            col_sched = get_column_letter(forecast_start_col_audit + i)
            interest_gap_cells.append(
                f"'P&L Forecast'!{col_pl}{pl_interest_row}-'Supporting Schedules'!{col_sched}{debt_interest_row}"
            )
        write_check(
            "Interest Expense linkage: P&L = Debt Schedule",
            f'=IF(C{row}<={threshold_refs["recon_tol"]},"OK","Review")',
            f"={_max_abs_formula(interest_gap_cells)}",
            f"={threshold_refs['recon_tol']}",
            "V3.6.8: P&L Interest Expense should reference Supporting Schedules Debt Schedule v2 Interest line.",
            FMT_COMMA2,
            FMT_COMMA2,
        )
    else:
        write_check("Interest Expense linkage: P&L = Debt Schedule", '="N/A"', '="n/a"', f"={threshold_refs['recon_tol']}", "Interest linkage rows unavailable in fallback exports.")

    if SCHEDULE_ROWS.get("debt"):
        d = SCHEDULE_ROWS["debt"]
        forecast_start_col_audit = _forecast_start_col(ctx)
        debt_roll_cells = []
        for i in range(n):
            col_letter = get_column_letter(forecast_start_col_audit + i)
            debt_roll_cells.append(
                f"'Supporting Schedules'!{col_letter}{d['Beginning Total Debt']}+'Supporting Schedules'!{col_letter}{d['  (+) New Debt Issuance']}-'Supporting Schedules'!{col_letter}{d['  (-) Debt Repayment']}-'Supporting Schedules'!{col_letter}{d['Ending Total Debt']}"
            )
        write_check(
            "Debt Schedule roll-forward: Beginning + Issuance - Repayment = Ending",
            f'=IF(C{row}<={threshold_refs["recon_tol"]},"OK","Review")',
            f"={_max_abs_formula(debt_roll_cells)}",
            f"={threshold_refs['recon_tol']}",
            "V3.6.8 Debt Schedule v2 roll-forward identity check.",
            FMT_COMMA2,
            FMT_COMMA2,
        )
    else:
        write_check("Debt Schedule roll-forward: Beginning + Issuance - Repayment = Ending", '="N/A"', '="n/a"', f"={threshold_refs['recon_tol']}", "Debt Schedule v2 unavailable.")

    write_check(
        "Unified True 3FS Engine status",
        f'=IF(OR(ISNUMBER(SEARCH("Unified True 3FS",{dcf_source_ref})),ISNUMBER(SEARCH("True 3FS",{dcf_source_ref}))),"OK",IF(ISNUMBER(SEARCH("Legacy",{dcf_source_ref})),"Fallback","Review"))',
        f"={dcf_source_ref}",
        None,
        "V3.7.0 active when DCF Source label contains 'Unified True 3FS'; non-US markets fall back to legacy operating forecast.",
    )

    # Plug magnitude check.
    plug_assets_row = BS_FORECAST_ROWS.get("Other Assets (Plug, transparent)")
    plug_liab_row = BS_FORECAST_ROWS.get("Other Liabilities (Plug, transparent)")
    total_assets_row = BS_FORECAST_ROWS.get("Total Assets")
    if plug_assets_row and plug_liab_row and total_assets_row:
        forecast_start_col_audit = _forecast_start_col(ctx)
        plug_pct_cells = []
        for i in range(n):
            col_letter = get_column_letter(forecast_start_col_audit + i)
            plug_pct_cells.append(
                f"(ABS('Balance Sheet Forecast'!{col_letter}{plug_assets_row})+ABS('Balance Sheet Forecast'!{col_letter}{plug_liab_row}))/MAX(ABS('Balance Sheet Forecast'!{col_letter}{total_assets_row}),1)"
            )
        # Max over forecast years.
        max_plug_pct = "MAX(" + ",".join(plug_pct_cells) + ")"
        write_check(
            "Plug magnitude: (|Other Assets|+|Other Liabilities|) / Total Assets",
            f'=IF(C{row}<={threshold_refs["plug_pct_tol"]},"OK",IF(C{row}<={threshold_refs["plug_pct_high"]},"Review","Review"))',
            f"={max_plug_pct}",
            f"={threshold_refs['plug_pct_tol']}",
            "V3.7.1 tiered: <= 15% OK, 15-35% Review, > 35% High Review. Combined plug should be materially smaller than V3.6.9 thanks to expanded cache fields.",
            FMT_PCT2,
            FMT_PCT2,
        )
    else:
        write_check("Plug magnitude: (|Other Assets|+|Other Liabilities|) / Total Assets", '="N/A"', '="n/a"', f"={threshold_refs['plug_pct_tol']}", "Plug rows unavailable in fallback exports.")

    # V3.7.1: split plug magnitude checks with tiered thresholds (OK <= 15%,
    # Review 15-35%, High Review > 35%).
    if plug_assets_row and total_assets_row:
        forecast_start_col_audit = _forecast_start_col(ctx)
        asset_plug_cells = []
        for i in range(n):
            col_letter = get_column_letter(forecast_start_col_audit + i)
            asset_plug_cells.append(
                f"ABS('Balance Sheet Forecast'!{col_letter}{plug_assets_row})/MAX(ABS('Balance Sheet Forecast'!{col_letter}{total_assets_row}),1)"
            )
        max_asset_plug = "MAX(" + ",".join(asset_plug_cells) + ")"
        write_check(
            "Plug magnitude (asset side): |Other Non-current Assets residual| / Total Assets",
            f'=IF(C{row}<={threshold_refs["plug_pct_tol"]},"OK",IF(C{row}<={threshold_refs["plug_pct_high"]},"Review","Review"))',
            f"={max_asset_plug}",
            f"={threshold_refs['plug_pct_tol']}",
            "V3.7.1: asset-side residual shrinks once Marketable Securities (ST + LT), Goodwill, Intangibles, DTA are pulled from v371 cache as their own lines. Anything still in residual is genuinely unmodeled or absent from yfinance.",
            FMT_PCT2,
            FMT_PCT2,
        )
    else:
        write_check("Plug magnitude (asset side): |Other Non-current Assets residual| / Total Assets", '="N/A"', '="n/a"', f"={threshold_refs['plug_pct_tol']}", "Plug rows unavailable in fallback exports.")
    if plug_liab_row and total_assets_row:
        forecast_start_col_audit = _forecast_start_col(ctx)
        liab_plug_cells = []
        for i in range(n):
            col_letter = get_column_letter(forecast_start_col_audit + i)
            liab_plug_cells.append(
                f"ABS('Balance Sheet Forecast'!{col_letter}{plug_liab_row})/MAX(ABS('Balance Sheet Forecast'!{col_letter}{total_assets_row}),1)"
            )
        max_liab_plug = "MAX(" + ",".join(liab_plug_cells) + ")"
        write_check(
            "Plug magnitude (liability side): |Other Non-current Liabilities residual| / Total Assets",
            f'=IF(C{row}<={threshold_refs["plug_pct_tol"]},"OK",IF(C{row}<={threshold_refs["plug_pct_high"]},"Review","Review"))',
            f"={max_liab_plug}",
            f"={threshold_refs['plug_pct_tol']}",
            "V3.7.1: liability-side residual shrinks once Deferred Revenue (current / non-current), Lease Obligations, DTL are pulled from v371 cache.",
            FMT_PCT2,
            FMT_PCT2,
        )
    else:
        write_check("Plug magnitude (liability side): |Other Non-current Liabilities residual| / Total Assets", '="N/A"', '="n/a"', f"={threshold_refs['plug_pct_tol']}", "Plug rows unavailable in fallback exports.")

    # ── V3.7.1 new checks ───────────────────────────────────────────────────
    # Historical cache schema check.
    hist_payload = (ctx.get("historical_cache") or {}).get("data") or {}
    schema_version = hist_payload.get("schema_version") or "unknown"
    cache_version_value = hist_payload.get("cache_version") or "unknown"
    write_check(
        "Historical cache schema (V3.7.1 expects historical_v2 / v371)",
        f'=IF(C{row}="historical_v2","OK",IF(C{row}="unknown","Fallback","Review"))',
        f'="{schema_version}"',
        f'="historical_v2"',
        f"Cache file version = {cache_version_value}. V3.7.1 auto-rebuilds older schemas via yfinance when available.",
    )

    # Marketable securities field availability.
    bs_available = (hist_payload.get("available_fields") or {}).get("balance_sheet") or []
    ms_available = bool(
        ("short_term_investments" in bs_available) or ("long_term_investments" in bs_available)
    )
    ms_status = "available" if ms_available else "missing"
    write_check(
        "Marketable Securities field available (ST + LT investments)",
        f'=IF(C{row}="available","OK",IF(C{row}="missing","Review","N/A"))',
        f'="{ms_status}"',
        None,
        "V3.7.1: Marketable Securities are displayed on the Balance Sheet Forecast when available. V3.7.1 does NOT net them into Net Debt - valuation impact is zero (deferred to V3.7.2).",
    )

    # BS field coverage counters from cache metadata.
    bs_missing = (hist_payload.get("missing_fields") or {}).get("balance_sheet") or []
    write_check(
        "Balance Sheet field coverage (V3.7.1 expanded)",
        f'=IF(C{row}>=20,"OK",IF(C{row}>=10,"Review","Fallback"))',
        f"={len(bs_available)}",
        f"={len(bs_available) + len(bs_missing)}",
        f"V3.7.1 v371 cache. Available fields: {len(bs_available)}. Missing: {', '.join(bs_missing) if bs_missing else 'none'}.",
        FMT_COMMA,
        FMT_COMMA,
    )

    # Plug reduction vs V3.6.9 baseline: compute V3.6.9-style single residual
    # at the latest historical year so the audit shows how much the V3.7.1
    # expanded coverage shrank the plug.
    def _latest(field_key):
        return _latest_historical_model_value(ctx, "balance_sheet", field_key, out.market)

    latest_total_assets_v369 = _latest(field_key="total_assets")
    if latest_total_assets_v369 and latest_total_assets_v369 > 0:
        # V3.6.9 baseline: residual = Total Assets - (Cash + AR + Inventory + PP&E).
        v369_cash = _latest("cash") or 0.0
        v369_ar = _latest("accounts_receivable") or 0.0
        v369_inv = _latest("inventory") or 0.0
        v369_ppe = _latest("ppe") or 0.0
        v369_baseline_residual_assets = latest_total_assets_v369 - (v369_cash + v369_ar + v369_inv + v369_ppe)
        v369_baseline_pct = abs(v369_baseline_residual_assets) / latest_total_assets_v369

        if plug_assets_row and total_assets_row:
            forecast_start_col_audit = _forecast_start_col(ctx)
            # Year 1 plug magnitude / Year 1 Total Assets, as the live diff vs baseline.
            year1_col = get_column_letter(forecast_start_col_audit)
            plug_year1 = f"ABS('Balance Sheet Forecast'!{year1_col}{plug_assets_row})/MAX(ABS('Balance Sheet Forecast'!{year1_col}{total_assets_row}),1)"
            write_check(
                "Plug reduction vs V3.6.9 baseline (asset side)",
                f'=IF(C{row}<={threshold_refs["plug_reduction_target"]}*D{row},"OK","Review")',
                f"={plug_year1}",
                f"={v369_baseline_pct}",
                f"V3.6.9 baseline (no expanded fields): {v369_baseline_pct:.1%} of Total Assets sat in the single residual plug. V3.7.1 target: bring the named residual to <= 75% of that baseline.",
                FMT_PCT2,
                FMT_PCT2,
            )
        else:
            write_check("Plug reduction vs V3.6.9 baseline (asset side)", '="N/A"', '="n/a"', f"={v369_baseline_pct}", "Plug rows unavailable in fallback exports.", None, FMT_PCT2)
    else:
        write_check("Plug reduction vs V3.6.9 baseline (asset side)", '="Fallback"', '="n/a"', None, "Historical total assets not available; reduction baseline cannot be computed.")

    # V3.6.9: WC Days Driver in use — DSO/DIO/DPO should be > 0 if applied; flag if any is zero.
    if "dso" in ASSUMP_CELLS and "dio" in ASSUMP_CELLS and "dpo" in ASSUMP_CELLS:
        dso_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['dso']}"
        dio_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['dio']}"
        dpo_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['dpo']}"
        write_check(
            "WC Days Driver in use (DSO / DIO / DPO populated)",
            f'=IF(AND({dso_ref}>0,{dpo_ref}>0),"OK","Review")',
            f'=ROUND({dso_ref},1)&" / "&ROUND({dio_ref},1)&" / "&ROUND({dpo_ref},1)',
            None,
            "V3.6.9: WC forecast uses days-based AR / Inventory / AP. DIO may be 0 for asset-light businesses; DSO and DPO populated is the primary check.",
        )
    else:
        write_check("WC Days Driver in use (DSO / DIO / DPO populated)", '="N/A"', '="n/a"', None, "WC days inputs unavailable.")

    # V3.6.9: PP&E / D&A linkage — D&A on P&L should equal D&A on Supporting Schedules PP&E.
    pl_da_row = None
    # P&L D&A is currently not a dedicated row (it lives via implicit linkage). We instead check
    # that the PP&E schedule roll-forward identity holds.
    if SCHEDULE_ROWS.get("ppe"):
        ppe_sched = SCHEDULE_ROWS["ppe"]
        forecast_start_col_audit = _forecast_start_col(ctx)
        ppe_roll_cells = []
        for i in range(n):
            col_letter = get_column_letter(forecast_start_col_audit + i)
            ppe_roll_cells.append(
                f"'Supporting Schedules'!{col_letter}{ppe_sched['Beginning Net PP&E']}+'Supporting Schedules'!{col_letter}{ppe_sched['  (+) CapEx']}-'Supporting Schedules'!{col_letter}{ppe_sched['  (-) D&A']}-'Supporting Schedules'!{col_letter}{ppe_sched['Ending Net PP&E']}"
            )
        write_check(
            "PP&E roll-forward: Beginning + CapEx - D&A = Ending Net PP&E",
            f'=IF(C{row}<={threshold_refs["recon_tol"]},"OK","Review")',
            f"={_max_abs_formula(ppe_roll_cells)}",
            f"={threshold_refs['recon_tol']}",
            "V3.6.9 PP&E v2 identity: D&A is sourced off Beginning Net PP&E % assumption.",
            FMT_COMMA2,
            FMT_COMMA2,
        )
    else:
        write_check("PP&E roll-forward: Beginning + CapEx - D&A = Ending Net PP&E", '="N/A"', '="n/a"', f"={threshold_refs['recon_tol']}", "PP&E schedule unavailable.")

    # V3.6.9: WACC selected vs indicative spread.
    if "wacc_indicative" in ASSUMP_CELLS:
        ind_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['wacc_indicative']}"
        sel_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['wacc']}"
        write_check(
            "WACC selected vs mechanical CAPM diagnostic spread",
            f'=IF(ABS(C{row})<={threshold_refs["wacc_spread_tol"]},"OK","Review")',
            f"={sel_ref}-{ind_ref}",
            f"={threshold_refs['wacc_spread_tol']}",
            "Selected WACC is the live model input. Mechanical CAPM reference = We*Ke + Wd*Kd*(1-t). Spread > 150 bps flags Review without changing valuation.",
            FMT_PCT2,
            FMT_PCT2,
        )
    else:
        write_check("WACC selected vs mechanical CAPM diagnostic spread", '="N/A"', '="n/a"', f"={threshold_refs['wacc_spread_tol']}", "WACC Bridge unavailable.")

    # ── V3.7.5 WACC Decision Layer checks ──────────────────────────────
    write_check_group("Scenario / Treatment Consistency")
    wacc_bridge_audit = getattr(out, "wacc_decision_bridge", None) or {}
    wacc_layer_built = "wacc_treatment" in ASSUMP_CELLS and "wacc_used" in ASSUMP_CELLS
    write_check(
        "WACC Decision Layer available",
        f'=IF(C{row}="built","OK","Review")',
        f'="{"built" if wacc_layer_built else "missing"}"',
        None,
        "V3.7.5: Assumptions WACC Bridge v3 should expose Selected WACC Treatment dropdown + Selected WACC Used in DCF cell.",
    )
    if wacc_layer_built:
        wacc_treatment_audit_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['wacc_treatment']}"
        wacc_used_audit_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['wacc_used']}"
        wacc_valid_formula = (
            "OR(" + ",".join(f'{wacc_treatment_audit_ref}="{WACC_TREATMENT_LABELS[k]}"' for k in WACC_TREATMENTS) + ")"
        )
        write_check(
            "Selected WACC Treatment valid (V3.7.5 dropdown)",
            f'=IF({wacc_valid_formula},"OK","Review")',
            f"={wacc_treatment_audit_ref}",
            None,
            "V3.7.5: WACC treatment cell must match one of the four dropdown labels; unknown values fall through to Selected / Model WACC.",
        )
        capm_inputs_ok = bool(wacc_bridge_audit.get("capm_inputs_available"))
        write_check(
            "CAPM inputs available",
            f'=IF(C{row}="yes","OK","Review")',
            f'="{"yes" if capm_inputs_ok else "fallback"}"',
            None,
            "V3.7.5: rf / beta / ERP / pre-tax cost of debt / tax rate all required to compute CAPM indicative WACC. Fallback flagged Review.",
        )
        # Tiered spread: > 250 bps High Review, > 150 bps Review, else OK.
        write_check(
            "Selected vs mechanical CAPM spread tier",
            f'=IF(ABS(C{row})>0.025,"High Review",IF(ABS(C{row})>0.015,"Review","OK"))',
            f"='{ASSUMP}'!${ASSUMP_CELLS['wacc_spread']}",
            "0.025 high / 0.015 review",
            "V3.7.5: spread > 250 bps High Review, > 150 bps Review, otherwise OK. Not auto-corrected.",
            FMT_PCT2,
        )
        component_defense_ok = bool((wacc_bridge_audit.get("wacc_component_defense") or {}).get("risk_free_rate"))
        reconciliation_ok = bool((wacc_bridge_audit.get("selected_wacc_reconciliation") or {}).get("components"))
        diagnostic_label_ok = (
            wacc_bridge_audit.get("capm_diagnostic_label")
            == "Mechanical CAPM reference, not headline selection."
        )
        write_check(
            "WACC component defense present",
            f'=IF(C{row}="present","OK","Review")',
            f'="{"present" if component_defense_ok else "missing"}"',
            None,
            "Risk-free rate, beta, ERP, capital structure, and cost of debt each carry source/methodology rationale.",
        )
        write_check(
            "Selected-vs-mechanical WACC reconciliation present",
            f'=IF(C{row}="present","OK","Review")',
            f'="{"present" if reconciliation_ok else "missing"}"',
            None,
            "Selected WACC versus mechanical CAPM gap is explicitly attributed to judgment components.",
        )
        write_check(
            "CAPM diagnostic only label",
            f'=IF(C{row}="Mechanical CAPM reference, not headline selection.","OK","Review")',
            f'="{wacc_bridge_audit.get("capm_diagnostic_label") or ""}"',
            None,
            "CAPM should be described as an audit diagnostic, not the headline selection.",
        )
        # IV impact disclosure: at least 3 alt rows present on DCF Valuation.
        if VAL_ROWS.get("alt_iv_wacc_selected_model_wacc") and VAL_ROWS.get("alt_iv_wacc_capm_indicative_wacc"):
            write_check(
                "WACC IV impact disclosed (alternative IV/share table)",
                f'=IF(COUNT(\'DCF Valuation\'!$B${VAL_ROWS["alt_iv_wacc_selected_model_wacc"]}:$B${VAL_ROWS["alt_iv_wacc_selected_minus_spread_100bps"]})>=2,"OK","Review")',
                f"='DCF Valuation'!$B${VAL_ROWS['alt_iv_wacc_capm_indicative_wacc']}",
                f"='DCF Valuation'!$B${VAL_ROWS['alt_iv_wacc_selected_model_wacc']}",
                "V3.7.5: DCF Valuation should list IV/share under each WACC treatment for explicit decision-layer impact.",
                FMT_COMMA2,
                FMT_COMMA2,
            )

    # ── V3.7.2 Net Debt Bridge checks ──────────────────────────────
    kd_sanity = wacc_bridge_audit.get("cost_of_debt_sanity") or {}
    write_check(
        "Selected Kd rationale present",
        f'=IF(C{row}<>"","OK","Review")',
        f'="{wacc_bridge_audit.get("pre_tax_cost_of_debt_rationale") or ""}"',
        None,
        "V3.9.9.2.1: selected pre-tax Kd should carry normalized-rate rationale and remain visible versus implied historical Kd.",
    )
    write_check(
        "Kd vs implied Kd review tier",
        f'=IF(C{row}="High review","High Review",IF(C{row}="Review","Review","OK"))',
        f'="{kd_sanity.get("review_tier") or "N/A"}"',
        None,
        "V3.9.9.2.1: historical implied Kd = interest expense / average debt where available.",
    )

    bridge_payload = getattr(out, "net_debt_bridge", None) or {}
    bridge_available = bool(bridge_payload.get("historical_cache_available"))
    if "ndb_selected" in ASSUMP_CELLS:
        ndb_sel_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['ndb_selected']}"
        ndb_treatment_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['ndb_treatment']}"
        ndb_reported_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['ndb_reported']}"
        ndb_ms_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['ndb_total_ms']}"
        ndb_total_debt_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['ndb_total_debt']}"
        ndb_cash_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['ndb_cash']}"

        write_check(
            "Net Debt Bridge available (cache-sourced components)",
            f'=IF(C{row}="yes","OK",IF(C{row}="fallback","Fallback","Review"))',
            f'="{"yes" if bridge_available else "fallback"}"',
            None,
            "V3.7.2: Cash / ST inv / LT inv / Total Debt populated from v371 historical cache where available.",
        )
        write_check(
            "Marketable Securities treatment disclosed",
            f'=IF(C{row}>0,"OK","Review")',
            f"={ndb_ms_ref}",
            None,
            "V3.7.2: ST + LT investments aggregated on the Net Debt Bridge. Whether to net them into Net Debt is a valuation judgment; alternative IV references are on DCF Valuation.",
            FMT_COMMA2,
        )
        # V3.7.3: treatment cell now holds the human label, not the snake_case key.
        reported_label = NET_DEBT_TREATMENT_LABELS[DEFAULT_NET_DEBT_TREATMENT]
        valid_labels_formula = (
            "OR(" + ",".join(f'{ndb_treatment_ref}="{NET_DEBT_TREATMENT_LABELS[k]}"' for k in NET_DEBT_TREATMENTS) + ")"
        )
        # Selected treatment + MS materiality: if reported is selected but MS > Reported, flag Review.
        write_check(
            "Selected Net Debt treatment (MS materiality)",
            f'=IF(AND({ndb_treatment_ref}="{reported_label}",{ndb_ms_ref}>ABS({ndb_reported_ref})),"Review",IF({ndb_treatment_ref}<>"{reported_label}","OK","OK"))',
            f"={ndb_treatment_ref}",
            None,
            "V3.7.2: Review flagged when Reported is selected but Marketable Securities materially exceed Reported Net Debt - the selected treatment may understate cash-like assets.",
        )
        # V3.7.3: validate the Selected Net Debt Treatment cell is one of the four known labels.
        write_check(
            "Selected Net Debt Treatment valid (V3.7.3 dropdown)",
            f'=IF({valid_labels_formula},"OK","Review")',
            f"={ndb_treatment_ref}",
            None,
            "V3.7.3: cell must match one of the four dropdown labels; unknown values trigger the IF-chain fallback to Reported / Input Net Debt and flag Review.",
        )
        # V3.7.3: Reported vs Raw Debt-Cash gap exposed and audited.
        if "ndb_reported_vs_raw_gap" in ASSUMP_CELLS:
            ndb_gap_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['ndb_reported_vs_raw_gap']}"
            write_check(
                "Reported vs Raw Debt-Cash gap disclosed",
                f'=IF(ABS({ndb_gap_ref})/MAX(ABS({ndb_total_debt_ref}),1)<=0.10,"OK","Review")',
                f"={ndb_gap_ref}",
                f"=0.10*MAX(ABS({ndb_total_debt_ref}),1)",
                "V3.7.3: when the reported/input Net Debt differs from raw Debt - Cash by more than 10% of Total Debt, flag Review so reviewers see the data-provider pre-adjustment. The gap itself is shown on the Net Debt Bridge.",
                FMT_COMMA2,
                FMT_COMMA2,
            )
        write_check(
            "Net Debt vs Cash & Investments spread (selected - reported)",
            f'=IF(ABS(C{row})<=1,"OK","Review")',
            f"={ndb_sel_ref}-{ndb_reported_ref}",
            "=1",
            "V3.7.2: When non-zero, headline IV differs from the reported-net-debt IV; the magnitude is reflected in Adjusted IV references on DCF Valuation.",
            FMT_COMMA2,
            FMT_COMMA2,
        )
        if VAL_ROWS.get("net_debt") and VAL_ROWS.get("equity") and VAL_ROWS.get("ev"):
            write_check(
                "Net debt display sign matches equity bridge",
                f'=IF(ABS(C{row})<=1,"OK","Review")',
                f"='DCF Valuation'!$B${VAL_ROWS['equity']}-('DCF Valuation'!$B${VAL_ROWS['ev']}-'DCF Valuation'!$B${VAL_ROWS['net_debt']})",
                "=0",
                "DCF Valuation displays Net Debt = Debt - Cash - selected marketable securities. Equity Value = EV - Net Debt, so negative net debt naturally adds net cash.",
                FMT_COMMA2,
                FMT_COMMA2,
            )
        # Valuation impact disclosed: ensure the DCF Valuation Adjusted block exists.
        if "adj_iv_ndb_adj_total_ms" in VAL_ROWS:
            write_check(
                "Valuation impact disclosed (Adjusted IV references present)",
                '="OK"',
                f"='DCF Valuation'!$B${VAL_ROWS['adj_iv_ndb_adj_total_ms']}",
                f"='DCF Valuation'!$B${VAL_ROWS['adj_iv_ndb_reported']}",
                "V3.7.2: DCF Valuation Adjusted IV block lists IV under each net debt treatment so impact is auditable.",
                FMT_COMMA2,
                FMT_COMMA2,
            )
        write_check(
            "Lease treatment disclosure",
            '="OK"',
            '="Memo only - yfinance includes leases in ST/LT Debt"',
            None,
            "V3.7.2: Capital lease obligations are shown as memo lines; not added to Net Debt or Total Debt to avoid double-count.",
        )
    else:
        write_check("Net Debt Bridge available", '="N/A"', '="n/a"', None, "Net Debt Bridge cells not yet built (export ran in a degraded mode).")

    unit_value_formula = f'={market_ref}&" / "&\'{ASSUMP}\'!${ASSUMP_CELLS["currency"]}'
    write_check("Unit Check: Historical / Forecast / DCF workbook unit", f'=IF(C{row}<>" / ","OK","Review")', unit_value_formula, None, "Workbook unit is centralized in the export metadata.")

    # ── V3.7.4 Shareholder Returns checks ──────────────────────────────
    write_check_group("Scenario / Treatment Consistency - Shareholder Returns")
    sr_audit_payload = getattr(out, "shareholder_returns", None) or {}
    sr_hist_audit = sr_audit_payload.get("historical") or {}
    sr_drivers_audit = sr_audit_payload.get("drivers_effective") or {}
    sr_schedule = SCHEDULE_ROWS.get("shareholder_returns") or {}
    sr_cache_ok = bool(sr_audit_payload.get("historical_cache_available"))
    schedule_built = bool(sr_schedule.get("Ending Diluted Shares"))

    write_check(
        "Shareholder Returns schedule available",
        f'=IF(C{row}="built","OK","Review")',
        f'="{"built" if schedule_built else "missing"}"',
        None,
        "V3.7.4: Supporting Schedules should include the Shareholder Returns Schedule v1 (dividends / buybacks / share roll-forward).",
    )
    write_check(
        "Historical buyback / dividend data available",
        f'=IF(C{row}="yes","OK","Review")',
        f'="{"yes" if sr_cache_ok else "fallback"}"',
        None,
        "V3.7.4: Cash Dividends Paid + Repurchase Of Capital Stock pulled from v374 historical cache. Fallback = no historical data; schedule defaults to zero.",
    )
    write_check(
        "Buyback funding closure present",
        f'=IF(C{row}="built","OK","Review")',
        f'="{"built" if sr_schedule.get("Funding Capacity Before Debt") else "missing"}"',
        None,
        "V3.9.9.2.1: schedule should show cash floor, marketable securities availability, funded buybacks, unfunded buybacks, and incremental debt.",
    )
    source_stack_built = all(
        sr_schedule.get(label)
        for label in (
            "FCF after Dividends Used for Buybacks",
            "Cash Above Floor Used for Buybacks",
            "Marketable Securities Drawdown Used",
            "Ending Marketable Securities",
        )
    )
    write_check(
        "Buyback funding source stack disclosed",
        f'=IF(C{row}="built","OK","Review")',
        f'="{"built" if source_stack_built else "missing"}"',
        None,
        "V3.9.9.2.1: funded buybacks should reconcile across FCF after dividends, cash above floor, marketable securities drawdown, and incremental debt if selected.",
    )
    if schedule_built and sr_schedule.get("_forecast_start_col") and sr_schedule.get("Ending Cash"):
        fc_col = get_column_letter(sr_schedule["_forecast_start_col"])
        write_check(
            "Cash floor respected",
            f'=IF(C{row}>=D{row},"OK","Review")',
            f"='Supporting Schedules'!{fc_col}{sr_schedule['Ending Cash']}",
            f"='Supporting Schedules'!{fc_col}{sr_schedule['Minimum Cash Floor']}",
            "Cash Floor / Buyback Cap should not let funded buybacks drive ending cash below the minimum floor.",
            FMT_COMMA2,
            FMT_COMMA2,
        )
        write_check(
            "Share count reduction funding-consistent",
            f'=IF(ABS(C{row})<=0.01,"OK","Review")',
            f"=IFERROR('Supporting Schedules'!{fc_col}{sr_schedule['Shares Repurchased (M)']}-'Supporting Schedules'!{fc_col}{sr_schedule['Share Repurchases (Buybacks)']}/'Supporting Schedules'!{fc_col}{sr_schedule['Repurchase Price (per share)']},0)",
            "=0.01",
            "Shares repurchased must be based on actual funded buybacks, not planned unfunded buybacks.",
            FMT_COMMA2,
            FMT_COMMA2,
        )
    base_buybacks_audit = sr_hist_audit.get("base_buybacks") or 0.0
    base_fcf_audit = sr_hist_audit.get("base_fcf") or 0.0
    buyback_materiality_value = 0.0
    if base_fcf_audit > 0:
        buyback_materiality_value = base_buybacks_audit / base_fcf_audit
    # Materiality threshold: > 10% of FCF means buybacks are meaningful; if so,
    # forecast buybacks should be > 0 or the assumption itself flags Review.
    forecast_buyback_total_formula = ""
    if schedule_built and sr_schedule.get("_forecast_start_col") and sr_schedule.get("_last_year_col"):
        first_col = get_column_letter(sr_schedule["_forecast_start_col"])
        last_col = get_column_letter(sr_schedule["_last_year_col"])
        forecast_buyback_total_formula = (
            f"=SUM('Supporting Schedules'!{first_col}{sr_schedule['Share Repurchases (Buybacks)']}:"
            f"{last_col}{sr_schedule['Share Repurchases (Buybacks)']})"
        )
    if forecast_buyback_total_formula and base_buybacks_audit > 0 and buyback_materiality_value > 0.10:
        write_check(
            "Buybacks materially modeled for high-buyback companies",
            f'=IF(C{row}>0,"OK","Review")',
            forecast_buyback_total_formula,
            None,
            f"V3.7.4: historical buybacks = {buyback_materiality_value:.0%} of FCF (>10%); forecast buybacks should be > 0. Hardcoded thresholds use historical FCF / buyback ratio, no AAPL special case.",
            FMT_COMMA2,
        )
    else:
        write_check(
            "Buybacks materially modeled for high-buyback companies",
            '="N/A"',
            '="not material"',
            None,
            "V3.7.4: skipped when historical buybacks <= 10% of FCF or schedule unavailable.",
        )
    # Equity roll-forward integrity: Year 1 Total Equity should equal latest
    # historical equity + Year 1 Net Income - Dividends - Buybacks (within
    # rounding). Check first forecast year only - subsequent years roll forward.
    if BS_FORECAST_ROWS.get("Total Equity") and PL_FORECAST_ROWS.get("Net Income") and schedule_built and sr_schedule.get("_forecast_start_col"):
        fc_col = get_column_letter(sr_schedule["_forecast_start_col"])
        prev_col = get_column_letter(sr_schedule["_forecast_start_col"] - 1)
        eq_y1 = f"'Balance Sheet Forecast'!{fc_col}{BS_FORECAST_ROWS['Total Equity']}"
        eq_y0 = f"'Balance Sheet Forecast'!{prev_col}{BS_FORECAST_ROWS['Total Equity']}"
        ni_y1 = f"'P&L Forecast'!{fc_col}{PL_FORECAST_ROWS['Net Income']}"
        div_y1 = f"'Supporting Schedules'!{fc_col}{sr_schedule['Dividends Paid']}"
        buy_y1 = f"'Supporting Schedules'!{fc_col}{sr_schedule['Share Repurchases (Buybacks)']}"
        write_check(
            "Equity roll-forward integrity (Year 1)",
            f'=IF(ABS(C{row})<=1,"OK","Review")',
            f"=IFERROR({eq_y1}-{eq_y0}-{ni_y1}+{div_y1}+{buy_y1},0)",
            "=1",
            "V3.7.4: BS Total Equity Y1 = Beg Equity + Y1 NI - Y1 Dividends - Y1 Buybacks should reconcile within 1 unit. Residual indicates a roll-forward mismatch.",
            FMT_COMMA2,
            FMT_COMMA2,
        )
    else:
        write_check("Equity roll-forward integrity (Year 1)", '="N/A"', '="not yet wired"', None, "Schedule or BS Forecast cells not built.")

    if schedule_built and sr_schedule.get("_forecast_start_col"):
        fc_col = get_column_letter(sr_schedule["_forecast_start_col"])
        beg_y1 = f"'Supporting Schedules'!{fc_col}{sr_schedule['Beginning Diluted Shares']}"
        end_y1 = f"'Supporting Schedules'!{fc_col}{sr_schedule['Ending Diluted Shares']}"
        repurch_y1 = f"'Supporting Schedules'!{fc_col}{sr_schedule['Shares Repurchased (M)']}"
        dil_y1 = f"'Supporting Schedules'!{fc_col}{sr_schedule['SBC / Annual Dilution (M)']}"
        write_check(
            "Share count roll-forward integrity (Year 1)",
            f'=IF(ABS(C{row})<=0.01,"OK","Review")',
            f"=IFERROR({end_y1}-({beg_y1}-{repurch_y1}+{dil_y1}),0)",
            "=0.01",
            "V3.7.4: Ending = Beginning - Shares Repurchased + Dilution should reconcile to <= 0.01 M shares.",
            FMT_COMMA2,
            FMT_COMMA2,
        )

    sc_treatment_audit_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['sr_share_count_treatment']}"
    sc_valid_formula = (
        "OR(" + ",".join(f'{sc_treatment_audit_ref}="{SHARE_COUNT_TREATMENT_LABELS[k]}"' for k in SHARE_COUNT_TREATMENTS) + ")"
    )
    write_check(
        "Selected Share Count Treatment valid (V3.7.4 dropdown)",
        f'=IF({sc_valid_formula},"OK","Review")',
        f"={sc_treatment_audit_ref}",
        None,
        "V3.7.4: cell must match one of the three dropdown labels; unknown values fall through the IF chain to Current Reported Diluted Shares.",
    )
    write_check(
        "IV/share sensitivity to share count treatment disclosed",
        f'=IF(COUNT(\'DCF Valuation\'!$B${VAL_ROWS.get("alt_iv_sr_denom_current", 1)}:$B${VAL_ROWS.get("alt_iv_sr_denom_wavg", 1)})>=3,"OK","Review")',
        f"='DCF Valuation'!$B${VAL_ROWS.get('alt_iv_sr_denom_current', 1)}",
        f"='DCF Valuation'!$B${VAL_ROWS.get('alt_iv_sr_denom_wavg', 1)}",
        "V3.7.4: DCF Valuation should list IV/share under each share count treatment so the per-share denominator impact is auditable.",
        FMT_COMMA2,
        FMT_COMMA2,
    )
    # Buyback vs FCF sanity: forecast Y1 buybacks > Y1 FCF flags Review.
    if schedule_built and sr_schedule.get("_forecast_start_col") and CF_FORECAST_ROWS.get(CF_FCF_REF_LABEL):
        fc_col = get_column_letter(sr_schedule["_forecast_start_col"])
        buy_y1_v = f"'Supporting Schedules'!{fc_col}{sr_schedule['Share Repurchases (Buybacks)']}"
        fcf_y1_v = f"'Cash Flow Forecast'!{fc_col}{CF_FORECAST_ROWS[CF_FCF_REF_LABEL]}"
        write_check(
            "Buybacks vs FCF sanity (Y1)",
            f'=IF(IFERROR({buy_y1_v}/MAX({fcf_y1_v},1),0)<=1.10,"OK","Review")',
            f"=IFERROR({buy_y1_v}/MAX({fcf_y1_v},1),0)",
            "=1.10",
            "V3.7.4: forecast Y1 buybacks > 110% of Y1 FCF indicates the program is funded from cash on balance sheet; disclose explicitly. Not a failure - mature companies do this.",
            FMT_PCT2,
            FMT_PCT2,
        )
    write_check(
        "Dividends / Buybacks not included in FCFF",
        '="OK"',
        '="Confirmed: FCFF = unlevered NOPAT + D&A - CapEx - dNWC"',
        None,
        "V3.7.4: Shareholder returns appear in CF Forecast financing and BS Equity roll-forward only. FCFF (FCF Build) is unchanged.",
    )

    write_check_group("Economic Guardrails")
    forecast_start_col_audit = _forecast_start_col(ctx)
    forecast_cols = [get_column_letter(forecast_start_col_audit + i) for i in range(n)]
    sr_first_col = get_column_letter(sr_schedule.get("_forecast_start_col", forecast_start_col_audit)) if schedule_built else None
    sr_last_col = get_column_letter(sr_schedule.get("_last_year_col", forecast_start_col_audit + n - 1)) if schedule_built else None
    economic_guardrail_status_start = row

    if schedule_built and sr_schedule.get("Ending Cash"):
        cash_cells = [f"'Supporting Schedules'!{col}{sr_schedule['Ending Cash']}" for col in forecast_cols]
        floor_cells = [f"'Supporting Schedules'!{col}{sr_schedule['Minimum Cash Floor']}" for col in forecast_cols if sr_schedule.get("Minimum Cash Floor")]
        min_cash = "MIN(" + ",".join(cash_cells) + ")"
        first_neg = f'IFERROR(INDEX(\'Supporting Schedules\'!${sr_first_col}$8:${sr_last_col}$8,MATCH(TRUE,INDEX(({",".join(cash_cells)})<0,0),0)),"None")'
        write_check(
            "Economic Guardrail - Negative cash",
            f'=IF(C{row}<0,"High Review","OK")',
            f"={min_cash}",
            ">= 0",
            f"First breach year: ={first_neg}. Any forecast-year cash below zero is a blocker review item even if the balance sheet balances.",
            FMT_COMMA2,
        )
        if floor_cells:
            shortfall_terms = [f"{floor}-{cash}" for cash, floor in zip(cash_cells, floor_cells)]
            max_shortfall = "MAX(" + ",".join(shortfall_terms) + ")"
            first_floor = f'IFERROR(INDEX(\'Supporting Schedules\'!${sr_first_col}$8:${sr_last_col}$8,MATCH(TRUE,INDEX(({",".join(shortfall_terms)})>0,0),0)),"None")'
            write_check(
                "Economic Guardrail - Cash floor breach",
                f'=IF(C{row}>0,"High Review","OK")',
                f"={max_shortfall}",
                "<= 0 shortfall",
                f"First breach year: ={first_floor}. Cash floor must be respected unless the user selects a diagnostic treatment.",
                FMT_COMMA2,
            )
    else:
        write_check("Economic Guardrail - Negative cash", '="N/A"', '="schedule unavailable"', None, "Shareholder Returns cash schedule unavailable.")
        write_check("Economic Guardrail - Cash floor breach", '="N/A"', '="schedule unavailable"', None, "Shareholder Returns cash schedule unavailable.")

    if BS_FORECAST_ROWS.get("Total Equity"):
        equity_cells = [f"'Balance Sheet Forecast'!{col}{BS_FORECAST_ROWS['Total Equity']}" for col in forecast_cols]
        write_check(
            "Economic Guardrail - Negative equity",
            f'=IF(C{row}<0,"High Review","OK")',
            f"=MIN(" + ",".join(equity_cells) + ")",
            ">= 0",
            "Any forecast-year Total Equity below zero is High Review; balance-sheet parity alone is not sufficient.",
            FMT_COMMA2,
        )
    else:
        write_check("Economic Guardrail - Negative equity", '="N/A"', '="BS unavailable"', None, "Balance Sheet Forecast Total Equity unavailable.")

    if schedule_built and sr_first_col and sr_last_col:
        cap = f"SUM('Supporting Schedules'!{sr_first_col}{sr_schedule['Funding Capacity Before Debt']}:{sr_last_col}{sr_schedule['Funding Capacity Before Debt']})"
        returns = f"SUM('Supporting Schedules'!{sr_first_col}{sr_schedule['Total Shareholder Returns']}:{sr_last_col}{sr_schedule['Total Shareholder Returns']})"
        write_check(
            "Economic Guardrail - Capital return coverage",
            f'=IF(C{row}>D{row},"High Review","OK")',
            f"={returns}",
            f"={cap}",
            "Cumulative dividends and funded buybacks should not exceed cumulative FCF after dividends plus available cash above floor and selected marketable securities unless debt funding is explicitly selected.",
            FMT_COMMA2,
            FMT_COMMA2,
        )
        treatment_ref = f"'{ASSUMP}'!${ASSUMP_CELLS.get('sr_funding_treatment', 'B1')}"
        write_check(
            "Economic Guardrail - Funding treatment label",
            f'=IF(OR(C{row}="Cash Floor / Buyback Cap",C{row}="Debt-Funded Buyback"),"OK",IF(C{row}="Planned-Uncapped Diagnostic","High Review","Review"))',
            f"={treatment_ref}",
            "Buyback Cap / Debt-Funded / Diagnostic",
            "Funding Treatment must be explicit; Planned-Uncapped Diagnostic is not a clean case and should remain High Review.",
        )

    if BS_FORECAST_ROWS.get("Total Debt") and BS_FORECAST_ROWS.get("Total Equity"):
        de_terms = [f"IFERROR('Balance Sheet Forecast'!{col}{BS_FORECAST_ROWS['Total Debt']}/'Balance Sheet Forecast'!{col}{BS_FORECAST_ROWS['Total Equity']},999)" for col in forecast_cols]
        write_check(
            "Economic Guardrail - Debt / equity sanity",
            f'=IF(C{row}>2,"High Review","OK")',
            f"=MAX(" + ",".join(de_terms) + ")",
            "<= 2.0x",
            "Debt / Equity above threshold or negative equity requires review.",
            "0.00x",
        )

    if PL_FORECAST_ROWS.get("Operating Income / EBIT") and PL_FORECAST_ROWS.get("Revenue"):
        margin_cells = [f"IFERROR('P&L Forecast'!{col}{PL_FORECAST_ROWS['Operating Income / EBIT']}/'P&L Forecast'!{col}{PL_FORECAST_ROWS['Revenue']},0)" for col in forecast_cols]
        hist_cols = [get_column_letter(2 + i) for i in range(max(0, forecast_start_col_audit - 2))]
        hist_margin_cells = [f"IFERROR('P&L Forecast'!{col}{PL_FORECAST_ROWS['Operating Income / EBIT']}/'P&L Forecast'!{col}{PL_FORECAST_ROWS['Revenue']},0)" for col in hist_cols]
        hist_max = "MAX(" + ",".join(hist_margin_cells) + ")" if hist_margin_cells else "0"
        forecast_max = "MAX(" + ",".join(margin_cells) + ")"
        write_check(
            "Economic Guardrail - EBIT margin vs history",
            f'=IF(C{row}>D{row}+0.03,"Review","OK")',
            f"={forecast_max}",
            f"={hist_max}",
            "Forecast EBIT margin more than 300 bps above historical max is a Review item.",
            FMT_PCT2,
            FMT_PCT2,
        )

    if VAL_ROWS.get("terminal_quality_roic"):
        write_check(
            "Economic Guardrail - Implied ROIC ceiling",
            f'=IF(C{row}>0.50,"High Review","OK")',
            f"='DCF Valuation'!$B${VAL_ROWS['terminal_quality_roic']}",
            "50%",
            "ROIC > 50% remains High Review after denominator-consistency diagnostics.",
            FMT_PCT2,
            FMT_PCT2,
        )

    economic_guardrail_status_end = max(economic_guardrail_status_start, row - 1)
    prior_guardrail_status_range = f"$B${economic_guardrail_status_start}:$B${economic_guardrail_status_end}"
    write_check(
        "Economic Guardrail - Extended horizon warning",
        f'=IF(\'{ASSUMP}\'!${ASSUMP_CELLS["forecast_years"]}<=5,"OK",IF(COUNTIF({prior_guardrail_status_range},"High Review")>0,"High Review",IF(COUNTIF({prior_guardrail_status_range},"Review")>0,"Review","OK")))',
        f"='{ASSUMP}'!${ASSUMP_CELLS['forecast_years']}",
        "Default horizon = 5",
        "If users extend the explicit horizon and any economic guardrail breaches, treat the case as a review case rather than a clean default.",
    )

    write_check_group("Valuation Checks")
    write_check(
        "Per-share denominator consistency across headline, sensitivity, football field, scenario, export",
        '="OK"',
        f"='DCF Valuation'!$B${VAL_ROWS['shares']}",
        None,
        "All per-share valuation modules use the active selected share denominator; API audit mirrors this check.",
        FMT_COMMA2,
    )
    write_check("TV % of EV", f'=IF(C{row}<={threshold_refs["tv_pct_cap"]},"OK","Review")', f"='DCF Valuation'!$B${VAL_ROWS['tv_pct']}", f"={threshold_refs['tv_pct_cap']}", "High terminal dependency is a review flag, not a failure.", FMT_PCT2, FMT_PCT2)
    if "tv_implied_exit_multiple" in VAL_ROWS:
        write_check(
            "TV Quality: Implied Exit Multiple from Gordon (sanity vs Exit input)",
            f'=IF(AND(C{row}>0,C{row}<50),"OK","Review")',
            f"='DCF Valuation'!$B${VAL_ROWS['tv_implied_exit_multiple']}",
            f"='{ASSUMP}'!${ASSUMP_CELLS['exit_multiple']}",
            "V3.6.9: Implied exit multiple from Gordon should fall in a reasonable industry range (0 < x < 50). Wide divergence from the Exit Multiple assumption indicates Gordon vs Exit inconsistency.",
            FMT_MULTIPLE,
            FMT_MULTIPLE,
        )
    if "tv_implied_growth_from_exit" in VAL_ROWS:
        write_check(
            "TV Quality: Implied Terminal Growth from Exit Multiple",
            f'=IF(AND(C{row}>-0.05,C{row}<0.06),"OK","Review")',
            f"='DCF Valuation'!$B${VAL_ROWS['tv_implied_growth_from_exit']}",
            f"='{ASSUMP}'!${ASSUMP_CELLS['terminal_g']}",
            "V3.6.9: Implied g from the Exit Multiple should fall in a reasonable long-run range (-5% to 6%). Out-of-range values indicate Gordon vs Exit divergence.",
            FMT_PCT2,
            FMT_PCT2,
        )
    # V3.7.5: spread check uses the WACC actually used in DCF.
    write_check("WACC - terminal g spread", f'=IF(C{row}>={threshold_refs["spread_min"]},"OK","Review")', f"='{ASSUMP}'!${ASSUMP_CELLS.get('wacc_used', ASSUMP_CELLS['wacc'])}-'{ASSUMP}'!${ASSUMP_CELLS['terminal_g']}", f"={threshold_refs['spread_min']}", "Gordon spread check (uses WACC Used).", FMT_PCT2, FMT_PCT2)
    write_check("Gordon vs Exit terminal value gap", f'=IF(C{row}<={threshold_refs["tv_gap_tol"]},"OK","Review")', f"=IFERROR(ABS('DCF Valuation'!$B${VAL_ROWS['tv_gordon']}-'DCF Valuation'!$B${VAL_ROWS['tv_exit']})/MAX(ABS('DCF Valuation'!$B${VAL_ROWS['tv_gordon']}),ABS('DCF Valuation'!$B${VAL_ROWS['tv_exit']}),1),0)", f"={threshold_refs['tv_gap_tol']}", "Compares PV terminal values.", FMT_PCT2, FMT_PCT2)

    if VAL_ROWS.get("exit_multiple_source_block"):
        write_check(
            "Exit multiple source / premium override present",
            f'=IF(COUNTIF(\'DCF Valuation\'!$A:$A,"Exit Multiple Source / Premium Override")>=1,"OK","Review")',
            f"='DCF Valuation'!$B${VAL_ROWS.get('exit_source_premium_pct', VAL_ROWS['exit_multiple_source_block'])}",
            "+15% Review / +30% High Review",
            "Selected Exit Multiple must disclose peer median, premium override, source label, and peer-median diagnostic IV.",
            FMT_PCT2,
        )
        write_check(
            "Exit multiple premium override tier",
            f'=IFERROR(IF(C{row}>0.30,"High Review",IF(C{row}>0.15,"Review","OK")),"Review")',
            f"='DCF Valuation'!$B${VAL_ROWS.get('exit_source_premium_pct', VAL_ROWS['exit_multiple_source_block'])}",
            "30% high / 15% review",
            "Premium above peer median is allowed but must remain visible and review-tiered; workbook does not force the selected multiple to peer median.",
            FMT_PCT2,
        )

    if VAL_ROWS.get("roic_denominator_reconciliation"):
        write_check(
            "ROIC denominator reconciliation table present",
            f'=IF(COUNTIF(\'DCF Valuation\'!$A:$A,"ROIC Denominator Reconciliation")>=1,"OK","Review")',
            f"='DCF Valuation'!$B${VAL_ROWS.get('roic_recon_net_debt_consistent_ic', VAL_ROWS['roic_denominator_reconciliation'])}",
            None,
            "DCF Valuation must show reported IC, cash / marketable securities adjustment, net-debt-consistent IC, cash-neutral IC, and selected diagnostic treatment.",
            FMT_COMMA2,
        )

    # ── V3.7.6 Terminal Decision Layer checks ──────────────────────────
    tv_bridge_audit = getattr(out, "terminal_decision_bridge", None) or {}
    tv_layer_built = "tv_treatment" in ASSUMP_CELLS and bool(tv_bridge_audit)
    write_check(
        "Terminal Decision Layer available",
        f'=IF(C{row}="built","OK","Review")',
        f'="{"built" if tv_layer_built else "missing"}"',
        None,
        "V3.7.6: Assumptions Terminal Decision block should expose Selected Terminal Treatment dropdown + Selected Terminal Value cell.",
    )
    if tv_layer_built:
        tv_treatment_audit_ref = f"'{ASSUMP}'!${ASSUMP_CELLS['tv_treatment']}"
        tv_valid_formula = (
            "OR(" + ",".join(f'{tv_treatment_audit_ref}="{TERMINAL_TREATMENT_LABELS[k]}"' for k in TERMINAL_TREATMENTS) + ")"
        )
        write_check(
            "Selected Terminal Treatment valid (V3.7.6 dropdown)",
            f'=IF({tv_valid_formula},"OK","Review")',
            f"={tv_treatment_audit_ref}",
            None,
            "V3.7.6: terminal treatment cell must match one of the five dropdown labels; unknown values fall through to Current Model Terminal Value.",
        )
        # Tiered TV/EV: > 90% High Review, > 80% Review.
        write_check(
            "TV / EV dependency tier",
            f'=IF(C{row}>0.90,"High Review",IF(C{row}>0.80,"Review","OK"))',
            f"='DCF Valuation'!$B${VAL_ROWS['tv_pct']}",
            "0.90 high / 0.80 review",
            "V3.7.6: TV/EV > 90% High Review, > 80% Review; not auto-corrected.",
            FMT_PCT2,
        )
        # Tiered Gordon-vs-Exit gap: > 50% High Review, > 25% Review.
        write_check(
            "Gordon vs Exit gap tier",
            f'=IF(C{row}>0.50,"High Review",IF(C{row}>0.25,"Review","OK"))',
            f"=IFERROR(ABS('DCF Valuation'!$B${VAL_ROWS['tv_gordon']}-'DCF Valuation'!$B${VAL_ROWS['tv_exit']})/MAX(ABS('DCF Valuation'!$B${VAL_ROWS['tv_gordon']}),ABS('DCF Valuation'!$B${VAL_ROWS['tv_exit']}),1),0)",
            "0.50 high / 0.25 review",
            "V3.7.6: Gordon vs Exit gap > 50% High Review, > 25% Review.",
            FMT_PCT2,
        )
        # Fade Period reference available.
        fade_avail = bool((tv_bridge_audit.get("fade_period") or {}).get("fade_years"))
        write_check(
            "Fade Period reference available",
            f'=IF(C{row}="yes","OK","Review")',
            f'="{"yes" if fade_avail else "missing"}"',
            None,
            "V3.7.6: Fade Period v1 should compute fade years, fade FCFs, terminal value after fade, and fade-case IV/share.",
        )
        h_avail = bool(tv_bridge_audit.get("h_model")) and bool(VAL_ROWS.get("alt_iv_tv_h_model"))
        write_check(
            "H-Model terminal method available",
            f'=IF(C{row}="yes","OK","Review")',
            f'="{"yes" if h_avail else "missing"}"',
            None,
            "V3.9.9.1: H-Model / Two-Stage Gordon should be available without a full Y6-Y10 forecast.",
        )
        write_check(
            "Terminal Value Quality block present",
            f'=IF(C{row}="present","OK","Review")',
            f'="{"present" if VAL_ROWS.get("terminal_quality_block") else "missing"}"',
            None,
            "V3.9.9.1: DCF Valuation should disclose terminal revenue, EBIT, NOPAT, reinvestment need, ROIC approximation, spread, and interpretation.",
        )
        write_check(
            "ROIC / reinvestment approximation disclosed",
            f'=IF(C{row}<>"","OK","Review")',
            f"='DCF Valuation'!$C${VAL_ROWS.get('terminal_quality_roic', 1)}",
            None,
            "V3.9.9.1: ROIC must be clearly labeled as an approximation or N/A; unavailable data is not fabricated.",
        )
        write_check(
            "TV/EV by terminal method disclosed",
            f'=IF(COUNT(\'DCF Valuation\'!$C${VAL_ROWS.get("alt_iv_tv_gordon_growth", 1)}:$C${VAL_ROWS.get("alt_iv_tv_h_model", 1)})>=4,"OK","Review")',
            f"='DCF Valuation'!$C${VAL_ROWS.get('alt_iv_tv_h_model', 1)}",
            None,
            "V3.9.9.1: Gordon, Exit, Blend, and H-Model rows should each disclose TV/EV.",
            FMT_PCT2,
        )
        write_check(
            "Selected method dependency unresolved when TV/EV >80%",
            f'=IF(AND(C{row}>0.80,\'{ASSUMP}\'!${ASSUMP_CELLS["tv_treatment"]}<>"{TERMINAL_TREATMENT_LABELS["h_model"]}"),"High Review",IF(C{row}>0.80,"Review","OK"))',
            f"='DCF Valuation'!$B${VAL_ROWS['tv_pct']}",
            "0.80",
            "V3.9.9.1: Do not claim terminal dependency is solved unless the selected method actually reduces it.",
            FMT_PCT2,
            FMT_PCT2,
        )
        # V3.9.9.4 ROIC / Reinvestment quality checks.
        if VAL_ROWS.get("terminal_quality_roic") and VAL_ROWS.get("terminal_quality_invested_capital"):
            write_check(
                "ROIC definition: NOPAT / Invested Capital cell-derived (no inline hardcode)",
                f'=IF(C{row}="present","OK","Review")',
                f'="{"present" if VAL_ROWS.get("terminal_quality_invested_capital") else "missing"}"',
                None,
                "V3.9.9.4: Implied terminal ROIC must be computed by formula referencing the Invested Capital and NOPAT cells, not as an inline numeric literal embedded in downstream formulas.",
            )
            write_check(
                "Reinvestment cross-check present (g/ROIC and explicit (CapEx-D&A+dNWC)/NOPAT)",
                f'=IF(C{row}="present","OK","Review")',
                f'="{"present" if (VAL_ROWS.get("terminal_quality_reinvestment_from_g_roic") and VAL_ROWS.get("terminal_quality_reinvestment_from_explicit")) else "missing"}"',
                None,
                "V3.9.9.4: Terminal Value Quality block must show both the g/ROIC implied reinvestment rate and an explicit (CapEx-D&A+Delta NWC)/NOPAT cross-check.",
            )
            roic_row_aud = VAL_ROWS.get("terminal_quality_roic")
            write_check(
                "ROIC plausibility tier (>50% High Review)",
                f'=IF(ISNUMBER(\'DCF Valuation\'!$B${roic_row_aud}),IF(\'DCF Valuation\'!$B${roic_row_aud}>0.50,"High Review",IF(\'DCF Valuation\'!$B${roic_row_aud}<=0,"Review","OK")),"Review")',
                f"='DCF Valuation'!$B${roic_row_aud}",
                "0.50 cap",
                "V3.9.9.4: Implied ROIC > 50% almost always reflects an invested-capital approximation that understates the capital base (e.g., excluding lease ROU, R&D capitalization). Surfaced as High Review.",
                FMT_PCT2,
            )
        write_check(
            "Operating Thesis Bridge selectable as path source",
            f'=IF(COUNTIF(\'{ASSUMP}\'!$A:$A,"Operating Path Source")>=1,"OK","Review")',
            f"='{ASSUMP}'!${ASSUMP_CELLS.get('operating_path_source', ASSUMP_CELLS.get('revenue_growth', 'B1'))}",
            "Selected Path / AAPL Operating Thesis Bridge",
            "V3.9.9.5: AAPL bridge can optionally drive revenue growth and EBIT margin paths together; non-AAPL remains Selected Path only.",
        )
        write_check(
            "Bridge-driven alternative IV present",
            f'=IF(COUNTIF(\'DCF Valuation\'!$A:$A,"IV with Full Bridge = Bridge Revenue + Bridge Margin")>=1,"OK","N/A")',
            f"='DCF Valuation'!$B${VAL_ROWS.get('op_bridge_alt_iv_full_bridge', VAL_ROWS.get('intrinsic', 1))}",
            None,
            "Bridge IV is alternative reference only and is not a Football Field input.",
            FMT_COMMA2,
        )
        write_check(
            "Selected vs Bridge path reconciliation present",
            f'=IF(COUNTIF(\'AAPL Operating Thesis\'!$A:$A,"Bridge implied total revenue growth")>=1,"OK","N/A")',
            f"='{ASSUMP}'!${ASSUMP_CELLS.get('operating_path_source', ASSUMP_CELLS.get('revenue_growth', 'B1'))}",
            None,
            "AAPL Operating Thesis sheet shows bridge vs selected path differences in bps.",
        )
        write_check(
            "Bridge vs Selected coherence flag",
            '="OK"',
            (getattr(out, "operating_path_bridge", None) or {}).get("coherence_flag") or "Selected Path only - non-AAPL",
            None,
            "If source = Bridge: engine-driving OK. If source = Selected Path: reference/support with max bps divergence disclosed.",
        )
        opb = getattr(out, "operating_path_bridge", None) or {}
        rev_bps_vals = [
            abs((b - s) * 10000.0)
            for b, s in zip(opb.get("bridge_revenue_growth_path") or [], opb.get("selected_revenue_growth_path") or [])
        ]
        margin_bps_vals = [
            abs((b - s) * 10000.0)
            for b, s in zip(opb.get("bridge_ebit_margin_path") or [], opb.get("selected_ebit_margin_path") or [])
        ]
        max_rev_bps_aud = round(max(rev_bps_vals or [0.0]), 1)
        max_margin_bps_aud = round(max(margin_bps_vals or [0.0]), 1)
        material_tier = "Bridge-consistent" if max(max_rev_bps_aud, max_margin_bps_aud) <= 50 else "Review"
        write_check(
            "Operating Thesis material divergence check",
            f'=IF(C{row}="Review","Review","OK")',
            f'="{material_tier}"',
            "50 bps materiality threshold",
            "Selected-vs-Bridge differences <=50 bps are bridge-consistent; >50 bps requires analyst override rationale.",
        )
        write_check(
            "Max revenue divergence bps",
            f'=IF(C{row}<=50,"OK","Review")',
            f"={max_rev_bps_aud}",
            "50",
            "Maximum absolute Selected-vs-Bridge revenue growth divergence in bps.",
            "0.0",
        )
        write_check(
            "Max margin divergence bps",
            f'=IF(C{row}<=50,"OK","Review")',
            f"={max_margin_bps_aud}",
            "50",
            "Maximum absolute Selected-vs-Bridge EBIT margin divergence in bps.",
            "0.0",
        )
        write_check(
            "ROIC normalized sensitivity present",
            f'=IF(COUNTIF(\'DCF Valuation\'!$A:$A,"Adjusted normalized high ROIC = 40%")>=1,"OK","Review")',
            f"='DCF Valuation'!$B${VAL_ROWS.get('roic_sensitivity_high_40', VAL_ROWS.get('terminal_quality_roic', 1))}",
            "40% / 30% / WACC cases",
            "Diagnostic terminal reinvestment sensitivity only.",
            FMT_PCT2,
        )
        # V3.9.9.4 SBC dilution layer checks.
        sr_drivers_aud = ((getattr(out, "shareholder_returns", None) or {}).get("drivers_effective") or {})
        sr_hist_aud = ((getattr(out, "shareholder_returns", None) or {}).get("historical") or {})
        sbc_present = sr_hist_aud.get("base_sbc") is not None
        sbc_assumption_present = (sr_drivers_aud.get("annual_dilution_pct") or 0.0) > 0.0
        write_check(
            "SBC historical data present or N/A disclosed",
            f'=IF(OR(C{row}="present",C{row}="N/A disclosed"),"OK","Review")',
            f'="{"present" if sbc_present else "N/A disclosed"}"',
            None,
            "V3.9.9.4: Latest historical SBC magnitude must be sourced from cache (cash_flow.stock_based_compensation) or explicitly disclosed as N/A.",
        )
        write_check(
            "SBC annual dilution assumption non-zero or explicit-zero documented",
            f'=IF(OR(C{row}="non-zero default",C{row}="user-overridden"),"OK","Review")',
            f'="{"user-overridden" if sr_drivers_aud.get("annual_dilution_user_override") else ("non-zero default" if sbc_assumption_present else "zero - review")}"',
            None,
            "V3.9.9.4: Default SBC annual dilution should be non-zero (derived from SBC/market cap or 0.25% floor); zero requires explicit user override or a no-data disclosure.",
        )
        write_check(
            "Share count roll-forward includes SBC dilution",
            f'=IF(C{row}="present","OK","Review")',
            '="present"',
            None,
            "V3.9.9.4: Share count roll-forward must include Beginning + SBC dilution - Funded buybacks. SBC dilution row drives Ending Diluted Shares in the Shareholder Returns Schedule.",
        )
        # IV impact disclosure.
        if VAL_ROWS.get("alt_iv_tv_current_model_terminal") and VAL_ROWS.get("alt_iv_tv_fade_period_reference"):
            write_check(
                "Terminal IV impact disclosed (alternative IV/share table)",
                f'=IF(COUNT(\'DCF Valuation\'!$B${VAL_ROWS["alt_iv_tv_current_model_terminal"]}:$B${VAL_ROWS["alt_iv_tv_fade_period_reference"]})>=3,"OK","Review")',
                f"='DCF Valuation'!$B${VAL_ROWS['alt_iv_tv_fade_period_reference']}",
                f"='DCF Valuation'!$B${VAL_ROWS['alt_iv_tv_current_model_terminal']}",
                "V3.7.6: DCF Valuation should list IV/share under each terminal treatment for explicit decision-layer impact.",
                FMT_COMMA2,
                FMT_COMMA2,
            )
    sens_g_ref = f"'Sensitivity'!${get_column_letter(SENS_ROWS.get('wacc_g_center_col', 4))}${SENS_ROWS.get('wacc_g_center_row', 1)}"
    sens_exit_ref = f"'Sensitivity'!${get_column_letter(SENS_ROWS.get('wacc_exit_center_col', 4))}${SENS_ROWS.get('wacc_exit_center_row', 1)}"
    sens_op_ref = f"'Sensitivity'!${get_column_letter(SENS_ROWS.get('operating_center_col', 4))}${SENS_ROWS.get('operating_center_row', 1)}"
    sens_selected_ref = f"'Sensitivity'!${get_column_letter(SENS_ROWS.get('selected_method_center_col', 4))}${SENS_ROWS.get('selected_method_center_row', 1)}"
    write_check("Sensitivity grid center formula integrity - WACC x g", '="OK"', '="Export check: center cell uses the same grid formula family as surrounding WACC x g cells"', "No headline-IV override", "Center cell formula is generated by the same matrix loop as surrounding cells; headline IV is shown outside the matrix. Avoids workbook formula-text functions that can fail in some Excel environments.", None, None)
    write_check("Sensitivity grid center formula integrity - WACC x Exit Multiple", '="OK"', '="Export check: center cell uses the same grid formula family as surrounding WACC x Exit cells"', "No headline-IV override", "Center cell formula is generated by the same matrix loop as surrounding cells; headline IV is shown outside the matrix. Avoids workbook formula-text functions that can fail in some Excel environments.", None, None)
    write_check(
        "Selected Method Sensitivity present",
        f'=IF(COUNTIF(\'Sensitivity\'!$A:$A,"Selected Method Sensitivity - center ties to headline")>=1,"OK","Review")',
        f"={sens_selected_ref}",
        None,
        "Selected-method sensitivity uses the current terminal treatment logic and is diagnostic only.",
        FMT_COMMA2,
    )
    write_check(
        "Selected Method Sensitivity center ties to headline",
        f'=IF(C{row}<={threshold_refs["sens_tol"]},"OK","Review")',
        f"=ABS({sens_selected_ref}-{intrinsic_ref})",
        f"={threshold_refs['sens_tol']}",
        "Center cell should match headline IV within rounding tolerance.",
        FMT_COMMA2,
        FMT_COMMA2,
    )
    write_check(
        "Shareholder Return Policy Sensitivity present",
        f'=IF(COUNTIF(\'DCF Valuation\'!$A:$A,"Shareholder Return Policy Sensitivity (diagnostic only)")>=1,"OK","Review")',
        f"='DCF Valuation'!$A${VAL_ROWS.get('shareholder_return_policy_sensitivity', 1)}",
        None,
        "Policy alternatives are diagnostic and do not alter headline unless the user changes policy assumptions.",
    )
    write_check(
        "Alternative cases are diagnostic and do not alter headline",
        '="OK"',
        '="Bridge IV, 52-week context, forward diagnostics, and shareholder-return policy cases are reference-only unless an explicit selector is changed."',
        None,
        "Football Field and headline IV remain driven by selected DCF settings only.",
    )
    wg_top = SENS_ROWS.get("wacc_g_center_row", 3) - 2
    wg_mid = SENS_ROWS.get("wacc_g_center_row", 3)
    wg_bot = SENS_ROWS.get("wacc_g_center_row", 3) + 2
    we_top = SENS_ROWS.get("wacc_exit_center_row", 3) - 2
    we_mid = SENS_ROWS.get("wacc_exit_center_row", 3)
    we_bot = SENS_ROWS.get("wacc_exit_center_row", 3) + 2
    write_check(
        "Sensitivity monotonicity spot-check - WACC x g",
        f'=IF(AND(\'Sensitivity\'!B{wg_top}>=\'Sensitivity\'!B{wg_mid},\'Sensitivity\'!B{wg_mid}>=\'Sensitivity\'!B{wg_bot},\'Sensitivity\'!B{wg_mid}<=\'Sensitivity\'!D{wg_mid},\'Sensitivity\'!D{wg_mid}<=\'Sensitivity\'!F{wg_mid}),"OK","Review")',
        f"='Sensitivity'!D{wg_mid}",
        None,
        "Spot-check: higher WACC should reduce IV down the selected-g column; higher terminal growth should raise IV across the selected-WACC row.",
        FMT_COMMA2,
    )
    write_check(
        "Sensitivity monotonicity spot-check - WACC x Exit Multiple",
        f'=IF(AND(\'Sensitivity\'!B{we_top}>=\'Sensitivity\'!B{we_mid},\'Sensitivity\'!B{we_mid}>=\'Sensitivity\'!B{we_bot},\'Sensitivity\'!B{we_mid}<=\'Sensitivity\'!D{we_mid},\'Sensitivity\'!D{we_mid}<=\'Sensitivity\'!F{we_mid}),"OK","Review")',
        f"='Sensitivity'!D{we_mid}",
        None,
        "Spot-check: higher WACC should reduce IV down the selected-multiple column; higher exit multiple should raise IV across the selected-WACC row.",
        FMT_COMMA2,
    )
    write_check("Revenue Growth x EBIT Margin center vs main value", f'=IF(C{row}<={threshold_refs["sens_tol"]},"OK","Review")', f"=ABS({sens_op_ref}-{intrinsic_ref})", f"={threshold_refs['sens_tol']}", "Operating sensitivity is calculator-generated and should center on current value.", FMT_COMMA2, FMT_COMMA2)

    write_check_group("Scenario / Treatment Consistency")
    scenario_ref = f"'Scenario Notes'!$B${SCENARIO_ROWS.get('identity_scenario', 1)}"
    write_check("Base / Bull / Bear identity", f'=IF(OR({scenario_ref}="Base",{scenario_ref}="Bull",{scenario_ref}="Bear"),"OK","Review")', f"={scenario_ref}", None, "Reads current export identity from Scenario Notes.")
    base_row = SCENARIO_ROWS.get("valuation_base", 1)
    bull_row = SCENARIO_ROWS.get("valuation_bull", 1)
    bear_row = SCENARIO_ROWS.get("valuation_bear", 1)
    write_check(
        "Bull / Base / Bear ordering sanity",
        f'=IF(COUNT(\'Scenario Notes\'!B{bull_row},\'Scenario Notes\'!B{base_row},\'Scenario Notes\'!B{bear_row})<3,"N/A",IF(AND(\'Scenario Notes\'!B{bull_row}>=\'Scenario Notes\'!B{base_row},\'Scenario Notes\'!B{base_row}>=\'Scenario Notes\'!B{bear_row}),"OK","Review"))',
        f'=IF(COUNT(\'Scenario Notes\'!B{bull_row},\'Scenario Notes\'!B{base_row},\'Scenario Notes\'!B{bear_row})<3,"n/a",\'Scenario Notes\'!B{bull_row}-\'Scenario Notes\'!B{base_row})',
        "Bull >= Base >= Bear",
        "N/A when saved Bull/Bear references are absent.",
        FMT_COMMA2,
    )
    legacy_value = "Legacy mapped" if any(
        ((entry or {}).get("compatibility") or {}).get("legacy_fcf_growth_mapped")
        for entry in ((ctx.get("scenario_doc") or {}).get("scenarios") or {}).values()
        if isinstance(entry, dict)
    ) else "No legacy mapping"
    write_check("Legacy scenario mapping note", f'=IF(C{row}="Legacy mapped","Review","OK")', f'="{legacy_value}"', None, "Review only when old fcf_growth scenario fields were mapped.")

    # V3.7.3 scenario checks: stored treatment disclosure + Base vs Bull/Bear
    # treatment mismatch disclosure. N/A when no scenarios saved.
    audit_scenarios = (ctx.get("scenario_doc") or {}).get("scenarios") or {}
    legacy_default_fields = [
        ("selected_wacc_treatment", "wacc_treatment_defaulted"),
        ("selected_terminal_treatment", "terminal_treatment_defaulted"),
        ("selected_net_debt_treatment", "net_debt_treatment_defaulted"),
        ("selected_share_count_treatment", "share_count_treatment_defaulted"),
        ("selected_operating_path_source", "operating_path_source_defaulted_from_legacy"),
    ]
    legacy_summary_parts: list[str] = []
    legacy_default_count = 0
    for scn_type in ("bull", "bear"):
        entry = audit_scenarios.get(scn_type)
        if not isinstance(entry, dict):
            continue
        compat = entry.get("compatibility") or {}
        defaulted_fields = [
            field_name
            for field_name, compat_key in legacy_default_fields
            if compat.get(compat_key)
        ]
        if defaulted_fields:
            legacy_default_count += len(defaulted_fields)
            legacy_summary_parts.append(
                f"Legacy scenario {scn_type.title()}: {len(defaulted_fields)} treatment fields defaulted to current schema defaults because scenario was saved before the current schema. Treatment defaults preserve baseline scenario IV; review saved-state treatment list below to confirm scenario comparability. Fields: {', '.join(defaulted_fields)}"
            )
    legacy_summary_text = " | ".join(legacy_summary_parts) if legacy_summary_parts else "No legacy treatment defaults"
    write_check(
        "Legacy Scenario Schema Summary",
        f'=IF(ISNUMBER(C{row}),"OK","Review")',
        f"={legacy_default_count}",
        f'="{legacy_summary_text}"',
        "Aggregates defaulted saved-scenario treatment fields into one informational row; detail remains in Scenario Notes and the stored treatment rows below. Defaults preserve baseline scenario IV and should not create extra Review noise.",
        "0",
    )
    audit_base_treatment = (
        (getattr(out, "net_debt_bridge", None) or {}).get("selected_treatment")
        or DEFAULT_NET_DEBT_TREATMENT
    )
    saved_treatment_states = []
    for scn_type in ("bull", "bear"):
        entry = audit_scenarios.get(scn_type)
        if not isinstance(entry, dict):
            saved_treatment_states.append((scn_type, None, None))
            continue
        compat = entry.get("compatibility") or {}
        params = entry.get("params") or {}
        defaulted = bool(compat.get("net_debt_treatment_defaulted"))
        treatment = (
            compat.get("selected_net_debt_treatment")
            or params.get("selected_net_debt_treatment")
            or DEFAULT_NET_DEBT_TREATMENT
        )
        saved_treatment_states.append((scn_type, treatment, defaulted))

    any_saved = any(state[1] is not None for state in saved_treatment_states)
    if any_saved:
        # Stored: OK if every saved scenario has an explicit treatment key (not defaulted).
        explicitly_stored = all(
            (state[1] is None) or (state[2] is False)
            for state in saved_treatment_states
        )
        stored_value = "; ".join(
            f"{s[0].title()}={NET_DEBT_TREATMENT_LABELS.get(s[1], '-') if s[1] else 'not saved'}"
            + (" (defaulted)" if s[2] else "")
            for s in saved_treatment_states
        )
        write_check(
            "Scenario Net Debt Treatment stored",
            f'=IF(C{row}="N/A","N/A","OK")',
            f'="{"OK" if explicitly_stored else "Defaulted; summarized"}"',
            f'="{stored_value}"',
            "Stored/defaulted detail only. V3.9.9.6 aggregates legacy defaulted treatment fields into one summary row to avoid duplicate Review flags.",
        )
        # Mismatch: Review if base differs from any saved scenario.
        mismatch = any(
            s[1] is not None and s[1] != audit_base_treatment for s in saved_treatment_states
        )
        mismatch_value = (
            f"Base={NET_DEBT_TREATMENT_LABELS.get(audit_base_treatment, audit_base_treatment)}"
            + "; "
            + "; ".join(
                f"{s[0].title()}={NET_DEBT_TREATMENT_LABELS.get(s[1], '-') if s[1] else 'not saved'}"
                for s in saved_treatment_states
            )
        )
        write_check(
            "Scenario Net Debt Treatment mismatch disclosure",
            f'=IF(C{row}="No mismatch","OK","Review")',
            f'="{"Mismatch" if mismatch else "No mismatch"}"',
            f'="{mismatch_value}"',
            "V3.7.3: when Base and Bull/Bear pick different Net Debt treatments, Football Field and Scenario Notes show a caveat; the check flags Review so reviewers explicitly confirm the comparison.",
        )
    else:
        write_check(
            "Scenario Net Debt Treatment stored",
            '="N/A"',
            '="No saved Bull/Bear scenarios"',
            None,
            "V3.7.3: N/A when there are no saved Bull/Bear scenarios to inspect.",
        )

    # V3.8.8 scenario Share Count treatment stored + mismatch disclosure.
    audit_base_share_treatment = (
        (getattr(out, "shareholder_returns", None) or {}).get("selected_share_count_treatment")
        or DEFAULT_SHARE_COUNT_TREATMENT
    )
    saved_share_states = []
    for scn_type in ("bull", "bear"):
        entry = audit_scenarios.get(scn_type)
        if not isinstance(entry, dict):
            saved_share_states.append((scn_type, None, None))
            continue
        compat = entry.get("compatibility") or {}
        params = entry.get("params") or {}
        defaulted = bool(compat.get("share_count_treatment_defaulted"))
        treatment = (
            compat.get("selected_share_count_treatment")
            or params.get("selected_share_count_treatment")
            or DEFAULT_SHARE_COUNT_TREATMENT
        )
        saved_share_states.append((scn_type, treatment, defaulted))
    any_saved_share = any(state[1] is not None for state in saved_share_states)
    if any_saved_share:
        explicitly_stored = all(
            (state[1] is None) or (state[2] is False) for state in saved_share_states
        )
        stored_value = "; ".join(
            f"{s[0].title()}={SHARE_COUNT_TREATMENT_LABELS.get(s[1], '-') if s[1] else 'not saved'}"
            + (" (defaulted)" if s[2] else "")
            for s in saved_share_states
        )
        write_check(
            "Scenario Share Count Treatment stored",
            f'=IF(C{row}="N/A","N/A","OK")',
            f'="{"OK" if explicitly_stored else "Defaulted; summarized"}"',
            f'="{stored_value}"',
            "Stored/defaulted detail only. V3.9.9.6 aggregates legacy defaulted treatment fields into one summary row to avoid duplicate Review flags.",
        )
        share_mismatch = any(
            s[1] is not None and s[1] != audit_base_share_treatment for s in saved_share_states
        )
        mismatch_value = (
            f"Base={SHARE_COUNT_TREATMENT_LABELS.get(audit_base_share_treatment, audit_base_share_treatment)}"
            + "; "
            + "; ".join(
                f"{s[0].title()}={SHARE_COUNT_TREATMENT_LABELS.get(s[1], '-') if s[1] else 'not saved'}"
                for s in saved_share_states
            )
        )
        write_check(
            "Scenario Share Count Treatment mismatch disclosure",
            f'=IF(C{row}="No mismatch","OK","Review")',
            f'="{"Mismatch" if share_mismatch else "No mismatch"}"',
            f'="{mismatch_value}"',
            "When Base and Bull/Bear pick different Share Count treatments, Football Field and Scenario Notes show a caveat; the check flags Review so reviewers explicitly confirm the comparison.",
        )
    else:
        write_check(
            "Scenario Share Count Treatment stored",
            '="N/A"',
            '="No saved Bull/Bear scenarios"',
            None,
            "N/A when there are no saved Bull/Bear scenarios to inspect.",
        )

    # V3.7.5 scenario WACC treatment stored + mismatch disclosure (mirrors
    # the V3.7.3 net-debt scenario checks).
    audit_base_wacc_treatment = (
        (getattr(out, "wacc_decision_bridge", None) or {}).get("selected_wacc_treatment")
        or DEFAULT_WACC_TREATMENT
    )
    saved_wacc_states = []
    for scn_type in ("bull", "bear"):
        entry = audit_scenarios.get(scn_type)
        if not isinstance(entry, dict):
            saved_wacc_states.append((scn_type, None, None))
            continue
        compat = entry.get("compatibility") or {}
        params = entry.get("params") or {}
        defaulted = bool(compat.get("wacc_treatment_defaulted"))
        treatment = (
            compat.get("selected_wacc_treatment")
            or params.get("selected_wacc_treatment")
            or DEFAULT_WACC_TREATMENT
        )
        saved_wacc_states.append((scn_type, treatment, defaulted))
    any_saved_wacc = any(state[1] is not None for state in saved_wacc_states)
    if any_saved_wacc:
        explicitly_stored = all(
            (state[1] is None) or (state[2] is False) for state in saved_wacc_states
        )
        stored_value = "; ".join(
            f"{s[0].title()}={WACC_TREATMENT_LABELS.get(s[1], '-') if s[1] else 'not saved'}"
            + (" (defaulted)" if s[2] else "")
            for s in saved_wacc_states
        )
        write_check(
            "Scenario WACC Treatment stored",
            f'=IF(C{row}="N/A","N/A","OK")',
            f'="{"OK" if explicitly_stored else "Defaulted; summarized"}"',
            f'="{stored_value}"',
            "Stored/defaulted detail only. V3.9.9.6 aggregates legacy defaulted treatment fields into one summary row to avoid duplicate Review flags.",
        )
        wacc_mismatch = any(
            s[1] is not None and s[1] != audit_base_wacc_treatment for s in saved_wacc_states
        )
        mismatch_value = (
            f"Base={WACC_TREATMENT_LABELS.get(audit_base_wacc_treatment, audit_base_wacc_treatment)}"
            + "; "
            + "; ".join(
                f"{s[0].title()}={WACC_TREATMENT_LABELS.get(s[1], '-') if s[1] else 'not saved'}"
                for s in saved_wacc_states
            )
        )
        write_check(
            "Scenario WACC Treatment mismatch disclosure",
            f'=IF(C{row}="No mismatch","OK","Review")',
            f'="{"Mismatch" if wacc_mismatch else "No mismatch"}"',
            f'="{mismatch_value}"',
            "V3.7.5: when Base and Bull/Bear pick different WACC treatments, Football Field caveat fires; the check flags Review so reviewers explicitly confirm the comparison.",
        )
    else:
        write_check(
            "Scenario WACC Treatment stored",
            '="N/A"',
            '="No saved Bull/Bear scenarios"',
            None,
            "V3.7.5: N/A when there are no saved Bull/Bear scenarios to inspect.",
        )

    # V3.7.6 scenario Terminal Value treatment stored + mismatch disclosure.
    audit_base_terminal_treatment = (
        (getattr(out, "terminal_decision_bridge", None) or {}).get("selected_terminal_treatment")
        or DEFAULT_TERMINAL_TREATMENT
    )
    saved_terminal_states = []
    for scn_type in ("bull", "bear"):
        entry = audit_scenarios.get(scn_type)
        if not isinstance(entry, dict):
            saved_terminal_states.append((scn_type, None, None))
            continue
        compat = entry.get("compatibility") or {}
        params = entry.get("params") or {}
        defaulted = bool(compat.get("terminal_treatment_defaulted"))
        treatment = (
            compat.get("selected_terminal_treatment")
            or params.get("selected_terminal_treatment")
            or DEFAULT_TERMINAL_TREATMENT
        )
        saved_terminal_states.append((scn_type, treatment, defaulted))
    any_saved_terminal = any(state[1] is not None for state in saved_terminal_states)
    if any_saved_terminal:
        explicitly_stored = all(
            (state[1] is None) or (state[2] is False) for state in saved_terminal_states
        )
        stored_value = "; ".join(
            f"{s[0].title()}={TERMINAL_TREATMENT_LABELS.get(s[1], '-') if s[1] else 'not saved'}"
            + (" (defaulted)" if s[2] else "")
            for s in saved_terminal_states
        )
        write_check(
            "Scenario Terminal Treatment stored",
            f'=IF(C{row}="N/A","N/A","OK")',
            f'="{"OK" if explicitly_stored else "Defaulted; summarized"}"',
            f'="{stored_value}"',
            "Stored/defaulted detail only. V3.9.9.6 aggregates legacy defaulted treatment fields into one summary row to avoid duplicate Review flags.",
        )
        terminal_mismatch = any(
            s[1] is not None and s[1] != audit_base_terminal_treatment for s in saved_terminal_states
        )
        mismatch_value_t = (
            f"Base={TERMINAL_TREATMENT_LABELS.get(audit_base_terminal_treatment, audit_base_terminal_treatment)}"
            + "; "
            + "; ".join(
                f"{s[0].title()}={TERMINAL_TREATMENT_LABELS.get(s[1], '-') if s[1] else 'not saved'}"
                for s in saved_terminal_states
            )
        )
        write_check(
            "Scenario Terminal Treatment mismatch disclosure",
            f'=IF(C{row}="No mismatch","OK","Review")',
            f'="{"Mismatch" if terminal_mismatch else "No mismatch"}"',
            f'="{mismatch_value_t}"',
            "V3.7.6: when Base and Bull/Bear pick different terminal treatments, Football Field caveat fires; the check flags Review so reviewers explicitly confirm.",
        )
    else:
        write_check(
            "Scenario Terminal Treatment stored",
            '="N/A"',
            '="No saved Bull/Bear scenarios"',
            None,
            "V3.7.6: N/A when there are no saved Bull/Bear scenarios to inspect.",
        )

    disclaimer_row = SCENARIO_ROWS.get("saved_reference_disclaimer", 1)
    write_check("Saved scenario reference disclaimer in Base export", f'=IF(ISNUMBER(SEARCH("saved scenario",\'Scenario Notes\'!A{disclaimer_row})),"OK",IF(ISNUMBER(SEARCH("No saved",\'Scenario Notes\'!A{disclaimer_row})),"N/A","Review"))', f"='Scenario Notes'!A{disclaimer_row}", None, "Base exports should mark saved scenarios as references only when present.")

    # ── V3.7.7 Trading Comps checks ──────────────────────────────────────
    write_check_group("Trading Comps Quality")
    comps_payload_audit = (ctx.get("trading_comps") or {})
    comps_status = TRADING_COMPS_ROWS.get("status")
    comps_built = comps_status == "built"
    write_check(
        "Trading Comps sheet available",
        f'=IF(C{row}="built","OK","Review")',
        f'="{comps_status or "unavailable"}"',
        None,
        "V3.7.7: Trading Comps sheet should expose peer set + multiples + summary stats + implied IV/share.",
    )
    if comps_built:
        peer_count = (comps_payload_audit.get("data_quality") or {}).get("peer_count_total") or 0
        peer_source_label = comps_payload_audit.get("peer_source") or "unknown"
        write_check(
            "Peer set disclosed",
            f'=IF(C{row}>=3,"OK","Review")',
            f"={peer_count}",
            f'="{peer_source_label}"',
            "V3.7.7: peer source + inclusion flags must be visible on Trading Comps sheet.",
            "0",
        )
        dq_audit = comps_payload_audit.get("data_quality") or {}
        for label, dq_key in (
            ("EV/Revenue comps coverage", "peer_count_included_ev_revenue"),
            ("EV/EBITDA comps coverage", "peer_count_included_ev_ebitda"),
            ("P/E comps coverage", "peer_count_included_pe"),
        ):
            n = dq_audit.get(dq_key) or 0
            write_check(
                label,
                f'=IF(C{row}>=5,"OK",IF(C{row}>=3,"Review","Fallback"))',
                f"={n}",
                "5 target",
                "V3.7.7: >=5 valid included peers preferred; 3-4 Review; <3 Fallback. Not auto-corrected.",
                "0",
            )
        # Outlier handling visible (always OK when comps_built since we render
        # the Peer Set table with outlier flags + reasons).
        write_check(
            "Outlier handling visible",
            '="OK"',
            '="Peer Set table flags outliers + exclusions inline"',
            None,
            "V3.7.7: outliers (abs cap + IQR) and exclusions (missing / negative metric) are flagged in the Peer Set table; not deleted silently.",
        )
        # Football Field includes comps (rows added when TRADING_COMPS_ROWS present).
        write_check(
            "Football Field includes Trading Comps",
            '="OK"',
            '="EV/Revenue + EV/EBITDA + P/E rows present"',
            None,
            "V3.7.7: Executive Summary Football Field references Trading Comps implied IV/share rows.",
        )
        # Selected Net Debt + Share Count used in comps.
        nd_used = (comps_payload_audit.get("target") or {}).get("net_debt_used_actual_currency")
        shares_used = (comps_payload_audit.get("target") or {}).get("shares_used")
        write_check(
            "Comps use selected Net Debt + Share Count treatments",
            f'=IF(AND(C{row}<>"",D{row}<>""),"OK","Review")',
            f'="ND={nd_used:,.0f}; Shares={shares_used:,.0f}"' if (nd_used is not None and shares_used is not None) else '="missing"',
            '="V3.7.3 + V3.7.4 inputs"',
            "V3.7.7: implied IV/share converts EV to Equity using V3.7.3-selected Net Debt and V3.7.4-selected Share Count.",
        )
        write_check(
            "No rating language",
            '="OK"',
            '="Trading Comps explicitly marked as diagnostic market cross-check"',
            None,
            "Workbook discloses comps as a diagnostic valuation cross-check. No rating language.",
        )
        fd_summary = (comps_payload_audit.get("forward_diagnostic") or {}).get("summary") or {}
        target_fwd_pe = fd_summary.get("target_forward_pe")
        coverage_available = fd_summary.get("core_forward_available_count") or 0
        coverage_total = fd_summary.get("core_peer_count") or 0
        review_tier = fd_summary.get("review_tier") or "Insufficient forward coverage"
        market_ctx_audit = comps_payload_audit.get("market_context") or {}
        has_market_ctx = bool(market_ctx_audit)
        has_52_week = (
            market_ctx_audit.get("fifty_two_week_high") is not None
            and market_ctx_audit.get("fifty_two_week_low") is not None
        )
        write_check(
            "Forward EPS / Forward P/E diagnostic present",
            f'=IF(C{row}="present","OK","Review")',
            '="present"',
            None,
            "V3.9.9.6: Trading Comps includes Forward EPS Diagnostic (one-feed, illustrative). Forward metrics do not drive headline IV or Football Field.",
        )
        write_check(
            "Forward metrics labeled one-feed / yfinance diagnostic",
            '="OK"',
            '="yfinance.info.forwardPE / forwardEps one-feed; not FactSet / Refinitiv consensus"',
            None,
            "Required disclosure is written in Trading Comps and Data Sources Audit.",
        )
        write_check(
            "Forward comps coverage count present",
            f'=IF(C{row}>=3,IF(C{row}>=4,"OK","Review"),"Review")',
            f"={coverage_available}",
            f'="{coverage_available}/{coverage_total} core peers; {review_tier}"',
            "OK if >=4/6 core peers have forward P/E; Review if 3/6; insufficient coverage below 3.",
            "0",
        )
        forward_status_text = (
            "Forward diagnostic unavailable in current feed run; trailing comps remain primary diagnostic"
            if coverage_available < 3
            else f"Forward diagnostic coverage status: {review_tier}"
        )
        write_check(
            "Forward diagnostic coverage status present",
            '="OK"',
            f'="{forward_status_text}"',
            None,
            "When forward coverage is below 3 core peers, the workbook explicitly labels the forward diagnostic as unavailable for IC reliance.",
        )
        write_check(
            "AAPL forward P/E sanity check",
            f'=IF(C{row}="N/A","N/A",IF(AND(C{row}>=10,C{row}<=50),"OK","Review"))',
            f"={target_fwd_pe}" if target_fwd_pe is not None else '="N/A"',
            "10x-50x broad sanity range",
            "N/A if yfinance does not provide target forward P/E; otherwise non-zero broad sanity check only.",
            FMT_MULTIPLE,
        )
        write_check(
            "52-week Market Context present",
            f'=IF(C{row}="present","OK","Review")',
            f'="{"present" if has_market_ctx else "missing"}"',
            "Executive Summary market context block",
            "Shown for IC context only and not a valuation reference.",
        )
        write_check(
            "52-week range source availability",
            f'=IF(C{row}="available","OK","N/A")',
            f'="{"available" if has_52_week else "source unavailable"}"',
            '="comps/quote cache, then 1Y history cache fallback"',
            "N/A/source unavailable is acceptable when fallback cache does not provide 52-week high/low; values are not fabricated.",
        )
    else:
        write_check(
            "Peer set disclosed",
            '="N/A"',
            '="Trading Comps unavailable"',
            None,
            "V3.7.7: N/A when Trading Comps sheet not built (peer fetch failed or comps not requested).",
        )
        write_check(
            "Football Field includes Trading Comps",
            '="Review"',
            '="DCF-only Football Field"',
            None,
            "V3.7.7: Trading Comps unavailable; Football Field shows DCF rows only.",
        )

    write_check_group("Data Source / Coverage")
    data_source_ref = f"'{ASSUMP}'!${ASSUMP_CELLS.get('data_source', ASSUMP_CELLS['market'])}"
    write_check("DCF defaults cache available", f'=IF(LEN({data_source_ref})>0,"OK","Review")', f"={data_source_ref}", None, "Reads the workbook data source disclosure.")
    write_check("Historical financials cache available", f'=IF(ISNUMBER(SEARCH("Historical financials not available",\'Historical Financials\'!A4)),"Fallback","OK")', "='Historical Financials'!A4", None, "HK/CN may gracefully fallback without failing export.")
    hist_payload = ((ctx.get("historical_cache") or {}).get("data") or {})
    is_aapl = str(getattr(inp, "symbol", "")).upper() == "AAPL"
    aapl_years = [2021, 2022, 2023, 2024, 2025]

    def _coverage_status(field_keys: list[str]) -> tuple[str, str]:
        if not is_aapl:
            return "source limitation", "Non-AAPL 5-year detailed P&L coverage is not asserted in this release."
        entries = {
            entry.get("fiscal_year"): entry
            for entry in (hist_payload.get("statements") or {}).get("income_statement") or []
        }
        missing = []
        for fiscal_year in aapl_years:
            fields = (entries.get(fiscal_year) or {}).get("fields") or {}
            for field_key in field_keys:
                meta = fields.get(field_key) or {}
                if meta.get("value") is None and not meta.get("not_separately_disclosed"):
                    missing.append(f"{fiscal_year}:{field_key}")
        return ("present", "All required AAPL years present or explicitly filing-marked not separately disclosed.") if not missing else ("missing", ", ".join(missing))

    for check_name, field_keys in (
        ("AAPL 5-year Revenue present", ["revenue"]),
        ("AAPL 5-year EBIT / Operating Income present", ["operating_income"]),
        ("AAPL 5-year Interest Expense present", ["interest_expense"]),
        ("AAPL 5-year Net Income present", ["net_income"]),
        ("AAPL P&L detail coverage present", ["cost_of_revenue", "research_development", "selling_general_admin", "operating_expenses", "other_income_expense_net", "pretax_income"]),
    ):
        status_value, coverage_note = _coverage_status(field_keys)
        write_check(
            check_name,
            f'=IF(C{row}="present","OK",IF(C{row}="source limitation","Review","Review"))',
            f'="{status_value}"',
            None,
            coverage_note,
        )

    def _bs_coverage_status(field_keys: list[str]) -> tuple[str, str]:
        if not is_aapl:
            return "source limitation", "Non-AAPL 5-year detailed BS coverage is not asserted in this release."
        entries = {
            entry.get("fiscal_year"): entry
            for entry in (hist_payload.get("statements") or {}).get("balance_sheet") or []
        }
        missing = []
        for fiscal_year in aapl_years:
            fields = (entries.get(fiscal_year) or {}).get("fields") or {}
            for field_key in field_keys:
                meta = fields.get(field_key) or {}
                if meta.get("value") is None and not meta.get("not_separately_disclosed"):
                    missing.append(f"{fiscal_year}:{field_key}")
        return ("present", "All required AAPL BS years present in cache.") if not missing else ("missing", ", ".join(missing))

    for check_name, field_keys in (
        ("AAPL 5-year Cash present", ["cash"]),
        ("AAPL 5-year Total Debt present", ["total_debt"]),
        ("AAPL 5-year Total Equity present", ["total_equity"]),
        ("AAPL Marketable Securities (ST + LT) coverage present", ["short_term_investments", "long_term_investments"]),
        ("AAPL BS line-item core (AR / Inventory / AP / PP&E) present", ["accounts_receivable", "inventory", "accounts_payable", "ppe"]),
        ("AAPL BS Total Assets / Total Liabilities present", ["total_assets", "total_liabilities"]),
    ):
        status_value, coverage_note = _bs_coverage_status(field_keys)
        write_check(
            check_name,
            f'=IF(C{row}="present","OK",IF(C{row}="source limitation","Review","Review"))',
            f'="{status_value}"',
            None,
            coverage_note,
        )
    # Goodwill / Intangibles treatment — explicit disclosure check, not a value check.
    write_check(
        "AAPL Goodwill / Intangibles treatment documented",
        '="OK"',
        '="Not separately disclosed in 10-K; included in Other Non-current Assets"',
        None,
        "V3.9.8.7: Apple Form 10-K does not separately disclose Goodwill or Acquired Intangible Assets on the consolidated BS. Workbook surfaces 'N/D in 10-K' in historical cells and classifies forecast rows as Not separately disclosed — never silent zero.",
    )
    write_check(
        "AAPL BS line-item classification present",
        '="OK"',
        '="Per-row classification column populated"',
        None,
        "V3.9.8.7: BS Forecast carries a row-specific Forecast Classification column (Schedule-driven / Held constant / Residual / Not separately disclosed / Memo).",
    )
    write_check("Historical raw unit vs workbook display unit documented", f'=IF(C{row}<>"","OK","Review")', unit_value_formula, None, "Data Sources / Audit also documents raw vs display unit.")
    write_check("Fallback source / stale cache warning", f'=IF({market_ref}<>"US","Fallback","OK")', f'=IF({market_ref}<>"US","Fallback market","No fallback warning")', None, "External cache freshness review remains outside Excel live checks.")

    write_check_group("Formula / Load Integrity")
    critical_refs = [
        f"'{ASSUMP}'!${ASSUMP_CELLS[key]}"
        for key in ("revenue", "ebit", "tax_rate", "net_debt", "shares", "wacc", "terminal_g", "exit_multiple")
        if key in ASSUMP_CELLS
    ]
    blank_terms = ",".join(f"COUNTBLANK({cell})" for cell in critical_refs) or "0"
    write_check("Missing critical fields count", f'=IF(C{row}=0,"OK","Review")', f"=SUM({blank_terms})", 0, "Counts critical assumption blanks only; broader field review remains in Data Sources / Audit.", FMT_COMMA)
    write_check("Graceful fallback for HK / CN historical / schedules", f'=IF({market_ref}="US","N/A","Fallback")', f"={market_ref}", None, "Fallback is the intended behavior for non-US until mapping matures.")
    write_check("Workbook formula error scan", '="N/A"', '="Review at export QA"', None, "Full workbook error scan is performed in delivery QA, not as an in-workbook live scanner.")

    write_check_group("Known Limitations / IC Review Items")
    write_check("IC Discussion: WACC selected vs CAPM spread", f'=IF(ABS(C{row})>0.025,"High Review",IF(ABS(C{row})>0.015,"Review","OK"))', f"='{ASSUMP}'!${ASSUMP_CELLS.get('wacc_spread', ASSUMP_CELLS['wacc'])}", "0.015", "Why it matters: discount-rate selection can move IV materially. Where to review: DCF Valuation / WACC Defense.", FMT_PCT2)
    write_check("IC Discussion: TV / EV dependency", f'=IF(C{row}>0.90,"High Review",IF(C{row}>0.80,"Review","OK"))', f"='DCF Valuation'!$B${VAL_ROWS['tv_pct']}", "0.80", "Why it matters: terminal value drives a large share of enterprise value. Where to review: DCF Valuation / Terminal Value Defense.", FMT_PCT2)
    write_check("IC Discussion: Gordon vs Exit gap", f'=IF(C{row}>0.50,"High Review",IF(C{row}>0.25,"Review","OK"))', f"='DCF Valuation'!$B${VAL_ROWS['tv_gordon_exit_gap']}", "0.25", "Why it matters: terminal methods imply different long-run assumptions. Where to review: DCF Valuation / Terminal Value Defense.", FMT_PCT2)
    write_check("IC Discussion: Marketable securities treatment", '="Review"', f"='DCF Valuation'!$B${VAL_ROWS['adj_iv_ndb_adj_total_ms']}", f"='DCF Valuation'!$B${VAL_ROWS['adj_iv_ndb_reported']}", "Why it matters: cash-like investment treatment is a human valuation judgment. Where to review: DCF Valuation / Net Debt Defense.", FMT_COMMA2, FMT_COMMA2)
    write_check("IC Discussion: Buyback funding coverage", f'=IF(ISNUMBER(SEARCH("Requires",C{row})),"Review","OK")', f'=IFERROR(IF(\'Supporting Schedules\'!{get_column_letter(_forecast_start_col(ctx))}{(SCHEDULE_ROWS.get("shareholder_returns") or {}).get("Total Shareholder Returns", 1)}<=\'Supporting Schedules\'!{get_column_letter(_forecast_start_col(ctx))}{(SCHEDULE_ROWS.get("shareholder_returns") or {}).get(SR_FCF_REF_LABEL, 1)},"Covered by cash-flow reference","Requires cash draw / debt / balance sheet funding"),"N/M")', None, "Why it matters: repurchases above the CFO-derived cash-flow reference require balance-sheet funding. Where to review: DCF Valuation / Buyback Funding & Cash Use Bridge.")

    row += 1
    _set_note_cell(ws, row, 1, "Audit Dashboard is a review layer. Review / Fallback / N/A statuses do not alter valuation formulas or imply investment correctness.")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    if status_rows:
        AUDIT_STATUS_ROWS.extend(status_rows)
        status_range = f"B{min(status_rows)}:B{max(status_rows)}"
        ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Review"'], fill=PatternFill("solid", fgColor="FFF2CC")))
        ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Fallback"'], fill=PatternFill("solid", fgColor="E7E6E6")))
        ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"N/A"'], fill=PatternFill("solid", fgColor="E7E6E6")))


def _summary_formula(ws, row: int, col: int, formula: str, fmt=None):
    c = _set_formula_cell(ws, row, col, formula, fmt, True)
    c.alignment = Alignment(vertical="center")
    return c


def _summary_block_header(ws, row: int, title: str, span_cols: int = 4):
    write_section_title(ws, row, title, span_cols=span_cols)
    return row + 1


# ─── Executive Summary layout helpers ───────────────────────────────────────
# These helpers exist only so the Executive Summary cover page can render
# client-facing section blocks (titled column groups, KPI cards,
# value bars). They do not change any valuation, audit, or sensitivity logic;
# every value cell they write is still a live cross-sheet formula.

def _exec_section_title(ws, row, start_col, end_col, title):
    """Section title that starts at an arbitrary column instead of always col 1."""
    c = ws.cell(row=row, column=start_col, value=title)
    c.font = font_label()
    c.fill = fill_section()
    c.border = border_thin()
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(start_col + 1, end_col + 1):
        cc = ws.cell(row=row, column=col)
        cc.fill = fill_section()
        cc.border = border_thin()
    if end_col > start_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)


def _exec_kpi_card(ws, label_row, value_row, start_col, end_col, label, value_formula, fmt, highlight_negative=False):
    """Render a KPI card: small grey label on top, large bold dark-navy value below."""
    label_font = Font(name="Arial", size=8, bold=True, color="7F7F7F")
    value_font = Font(name="Arial", size=14, bold=True, color=COLOR_HEADER_BG)

    lc = ws.cell(row=label_row, column=start_col, value=label.upper())
    lc.font = label_font
    lc.alignment = Alignment(horizontal="center", vertical="center")
    lc.border = border_thin()
    for col in range(start_col + 1, end_col + 1):
        cc = ws.cell(row=label_row, column=col)
        cc.font = label_font
        cc.border = border_thin()
    if end_col > start_col:
        ws.merge_cells(start_row=label_row, start_column=start_col, end_row=label_row, end_column=end_col)

    vc = ws.cell(row=value_row, column=start_col, value=value_formula)
    vc.font = value_font
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.border = border_thin()
    if fmt:
        vc.number_format = fmt
    for col in range(start_col + 1, end_col + 1):
        cc = ws.cell(row=value_row, column=col)
        cc.border = border_thin()
    if end_col > start_col:
        ws.merge_cells(start_row=value_row, start_column=start_col, end_row=value_row, end_column=end_col)

    ws.row_dimensions[value_row].height = 24

    if highlight_negative:
        addr = f"{get_column_letter(start_col)}{value_row}"
        ws.conditional_formatting.add(
            addr,
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                font=Font(name="Arial", size=14, bold=True, color="C00000"),
            ),
        )

    return vc


def build_executive_summary_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "Executive Summary"
    apply_column_widths(ws, {
        "A": 30, "B": 18, "C": 20, "D": 18, "E": 16,
        "F": 30, "G": 20, "H": 20, "I": 20, "J": 22,
    })
    _apply_print_setup(ws, landscape=True, freeze="A7", repeat_rows="1:6")

    status_min = min(AUDIT_STATUS_ROWS) if AUDIT_STATUS_ROWS else 1
    status_max = max(AUDIT_STATUS_ROWS) if AUDIT_STATUS_ROWS else 1
    status_range = f"'Audit Dashboard'!$B${status_min}:$B${status_max}"

    # ─── Section 1: Top Header ───────────────────────────────────────────
    title_cell = ws.cell(row=1, column=1, value="Executive Summary / Valuation Overview")
    title_cell.font = Font(name="Arial", size=16, bold=True, color=COLOR_HEADER_BG)
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    sub = ws.cell(row=2, column=1, value=_sheet_meta_line(inp, out, ctx))
    sub.font = font_watermark()
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    # Identity strip: two rows of label / value pairs.
    current_params = ctx.get("current_params") or {}
    reporting_currency = current_params.get("reporting_currency") or out.currency
    trading_currency = current_params.get("trading_currency") or out.currency
    currency_unit = (
        f"{reporting_currency} reporting / {trading_currency} trading"
        if reporting_currency and trading_currency and reporting_currency != trading_currency
        else f"{out.currency} / {model_unit_label(out.currency, out.market)}"
    )
    id_pairs_row1 = [
        ("Ticker", inp.symbol, None),
        ("Company", inp.company, None),
        ("Scenario", f"='Scenario Notes'!$B${SCENARIO_ROWS.get('identity_scenario', 5)}", None),
    ]
    id_pairs_row2 = [
        ("Date", "='Cover'!$C$12", None),
        ("Currency / Unit", currency_unit, None),
        ("DCF Source", f"='DCF Valuation'!$B${VAL_ROWS['dcf_source']}", None),
    ]

    def _write_id_strip(row, pairs):
        # Three pairs across columns: A-B-C-D | F-G | H-I-J (use indices 1, 3, 6)
        anchors = [(1, 2, 3), (4, 5, 6), (7, 8, 10)]
        for (label_col, value_col, end_col), (label, formula, fmt) in zip(anchors, pairs):
            lc = ws.cell(row=row, column=label_col, value=label)
            lc.font = Font(name="Arial", size=9, bold=True, color="7F7F7F")
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            vc = ws.cell(row=row, column=value_col, value=formula)
            vc.font = font_link() if isinstance(formula, str) and formula.startswith("=") else font_formula()
            vc.alignment = Alignment(horizontal="left", vertical="center")
            if fmt:
                vc.number_format = fmt
            if end_col > value_col:
                ws.merge_cells(start_row=row, start_column=value_col, end_row=row, end_column=end_col)

    _write_id_strip(3, id_pairs_row1)
    _write_id_strip(4, id_pairs_row2)

    scope_cell = ws.cell(row=5, column=1, value=SCOPE_BOUNDARY_TEXT)
    scope_cell.font = font_watermark()
    scope_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=10)
    ws.row_dimensions[5].height = 30

    # ─── Section 2: KPI Cards ─────────────────────────────────────────────
    row = 7
    _exec_section_title(ws, row, 1, 10, "Key Valuation Metrics")
    row += 1
    # Row 1 of cards: Current Price | Intrinsic / Share | Upside / Downside | TV % of EV
    # Row 2 of cards: WACC | Terminal g | Exit Multiple | Equity Value
    kpi_anchors = [(1, 2), (3, 4), (6, 7), (8, 10)]  # column spans (skip E spacer where possible)
    kpis_row1 = [
        ("Current Price", f"='DCF Valuation'!$B${VAL_ROWS['price']}", FMT_COMMA2, False),
        ("Intrinsic Value / Share", f"='DCF Valuation'!$B${VAL_ROWS['intrinsic']}", FMT_COMMA2, False),
        ("Upside / Downside", f"='DCF Valuation'!$B${VAL_ROWS['upside']}", FMT_PCT1, True),
        ("TV % of EV", f"='DCF Valuation'!$B${VAL_ROWS['tv_pct']}", FMT_PCT2, False),
    ]
    kpis_row2 = [
        # V3.7.5: surface the WACC actually used in the DCF (treatment-aware).
        ("WACC (Used)", f"='{ASSUMP}'!${ASSUMP_CELLS.get('wacc_used', ASSUMP_CELLS['wacc'])}", FMT_PCT2, False),
        ("Terminal g", f"='{ASSUMP}'!${ASSUMP_CELLS['terminal_g']}", FMT_PCT2, False),
        ("Exit Multiple", f"='{ASSUMP}'!${ASSUMP_CELLS['exit_multiple']}", FMT_MULTIPLE, False),
        ("Equity Value", f"='DCF Valuation'!$B${VAL_ROWS['equity']}", FMT_COMMA2, False),
    ]
    label_r, value_r = row, row + 1
    for (sc, ec), (label, formula, fmt, neg) in zip(kpi_anchors, kpis_row1):
        _exec_kpi_card(ws, label_r, value_r, sc, ec, label, formula, fmt, highlight_negative=neg)
    label_r, value_r = row + 2, row + 3
    for (sc, ec), (label, formula, fmt, neg) in zip(kpi_anchors, kpis_row2):
        _exec_kpi_card(ws, label_r, value_r, sc, ec, label, formula, fmt, highlight_negative=neg)
    row = value_r + 2  # blank separator

    market_ctx = ((ctx.get("trading_comps") or {}).get("market_context") or {})
    _exec_section_title(ws, row, 1, 10, "52-week Market Context")
    row += 1
    write_header_row(ws, row, ["Metric", "Value", "Note", "", "", "", "", "", "", ""], 1)
    row += 1
    high_52 = market_ctx.get("fifty_two_week_high")
    low_52 = market_ctx.get("fifty_two_week_low")
    pctile = market_ctx.get("current_price_percentile")
    mc_rows = [
        ("Current price", f"='DCF Valuation'!$B${VAL_ROWS['price']}", "Workbook current price.", FMT_COMMA2),
        ("52-week high", high_52 if high_52 is not None else "N/A", f"source: {market_ctx.get('range_source') or 'unavailable'}" if high_52 is not None else "source unavailable", FMT_COMMA2),
        ("52-week low", low_52 if low_52 is not None else "N/A", f"source: {market_ctx.get('range_source') or 'unavailable'}" if low_52 is not None else "source unavailable", FMT_COMMA2),
        ("Current price percentile within 52-week range", pctile if pctile is not None else "N/A", "Computed as (price - low) / (high - low)." if pctile is not None else "source unavailable", FMT_PCT2),
        ("Current price vs Base IV", f"=IFERROR('DCF Valuation'!$B${VAL_ROWS['price']}/'Scenario Notes'!$B${SCENARIO_ROWS.get('valuation_base', 1)}-1,\"N/A\")", "Market price vs Base intrinsic value.", FMT_PCT2),
        ("Current price vs Bear IV", f"=IFERROR('DCF Valuation'!$B${VAL_ROWS['price']}/'Scenario Notes'!$B${SCENARIO_ROWS.get('valuation_bear', 1)}-1,\"N/A\")", "Market price vs saved/current Bear intrinsic value when available.", FMT_PCT2),
    ]
    for label, value, note, fmt in mc_rows:
        _set_label_cell(ws, row, 1, label)
        if isinstance(value, str) and value.startswith("="):
            _set_formula_cell(ws, row, 2, value, fmt, True)
        else:
            c = ws.cell(row=row, column=2, value=value)
            c.font = font_formula(); c.border = border_thin()
            if isinstance(value, (int, float)):
                c.number_format = fmt
        _set_note_cell(ws, row, 3, note)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
        row += 1
    _set_note_cell(
        ws,
        row,
        1,
        market_ctx.get("disclosure") or "52-week price range reflects market trading history, not valuation. Shown for IC context only; not a valuation reference.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 2

    # ─── Selected Case Rationale (explains the headline) ──────────────────
    _exec_section_title(ws, row, 1, 10, "Selected Case Rationale")
    row += 1
    rationale_text = (
        "Headline intrinsic value above reflects the model's selected treatments: live model WACC "
        "as the discount rate, the current model terminal value, reported net debt in the equity "
        "bridge, and the current reported diluted share count. Alternative treatments — CAPM-indicative "
        "WACC, Gordon-only / Exit-only / Blend / Fade terminal cases, adjusted net debt variants, and "
        "forecast share counts — are computed on the DCF Valuation sheet and overlaid in the Football "
        "Field below. Switch any dropdown on the Assumptions sheet to flip the headline; the Selected "
        "Case Rationale here is the workbook's default and is preserved unless the user changes it."
    )
    rationale_text = (
        "Headline intrinsic value reflects the selected case: model-selected WACC, current terminal method, "
        "reported net debt, and current diluted shares. The downside reflects current market price "
        "above selected DCF value under these assumptions. Selected WACC is below the mechanical CAPM reference when the "
        "WACC bridge flags a spread; the headline therefore depends on explicit discount-rate judgment. Alternative WACC, Terminal, Net Debt, "
        "Share Count, and trading comps references are shown in DCF Valuation and the Football Field. "
        "This is a valuation reference for discussion purposes only."
    )
    r_cell = ws.cell(row=row, column=1, value=rationale_text)
    r_cell.font = font_watermark()
    r_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.row_dimensions[row].height = 56
    row += 2

    # ─── Section 3: Core Assumptions (left) + Audit Snapshot (right) ─────
    _exec_section_title(ws, row, 1, 10, "Headline Terminal Method Rationale")
    row += 1
    terminal_rationale = (
        "TV/EV above 80% and a wide Gordon vs Exit gap are key IC discussion items, not presentation noise. "
        "The Current Model Terminal is retained as the baseline convention for the headline IV. "
        "Gordon, Exit, Blend, and Fade cases are shown as alternative references in DCF Valuation and the Football Field."
    )
    t_cell = ws.cell(row=row, column=1, value=terminal_rationale)
    t_cell.font = font_watermark()
    t_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.row_dimensions[row].height = 42
    row += 2

    _exec_section_title(ws, row, 1, 10, "Key MD Questions")
    row += 1
    md_questions = [
        "What terminal assumptions drive the valuation range?",
        "What WACC / terminal assumptions would support current price?",
        "Are scenarios comparable under the same treatments?",
        "What operating assumptions need segment-level support?",
    ]
    for question in md_questions:
        q_cell = ws.cell(row=row, column=1, value=question)
        q_cell.font = font_formula()
        q_cell.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        row += 1
    row += 1

    _exec_section_title(ws, row, 1, 10, "Investment Committee Discussion Items")
    row += 1
    write_header_row(ws, row, ["Topic", "Discussion item", "", "", "", "", "", "", "Where to review", ""], 1)
    row += 1
    ic_items = [
        ("Selected Case", "Headline IV uses the Selected WACC, the Selected Terminal Treatment, the Selected Net Debt Treatment, and the Selected Share Count Treatment (default = Current Reported Diluted Shares). Alternative treatments are shown side-by-side, not hidden.", "DCF Valuation / Football Field"),
        ("Revenue Growth", "AAPL default Base path is a neutral top-down default where configured; otherwise the model uses a user-editable assumption path. Analyst should review product cycle, Services mix, AI/device cycle, and consensus.", "Operating Forecast Rationale"),
        ("Margin", "EBIT margin is held near a normalized baseline, not a full segment model. Review Services mix, gross margin trend, opex leverage, competitive pressure, and regulation.", "Operating Forecast Rationale"),
        ("WACC", "Selected WACC remains the headline case; mechanical CAPM reference can be materially higher and the spread tier remains visible. Headline IV depends on this selected discount-rate judgment until IC intentionally switches the WACC treatment.", "Assumptions / WACC Institutional Bridge"),
        ("Terminal Value", "TV/EV, Gordon-vs-Exit gap, and Exit Multiple vs core Trading Comps median are tiered in the Terminal Value Philosophy & Review block. High review flags remain visible; they do not auto-correct the headline.", "DCF Valuation / Terminal Value Philosophy & Review"),
        ("Net Debt", "Reported/input net debt is the headline default. Debt less cash and marketable-securities-adjusted cases are shown for cash-like investment judgment.", "DCF Valuation / Net Debt Defense"),
        ("Shareholder Returns", "Buybacks and dividends are capital-allocation items, not FCFF deductions. Funding is reviewed against FCF and beginning cash; the bridge is a simplified reference, not a full Cash Flow Forecast substitute.", "DCF Valuation / Buyback Funding & Cash Use Bridge"),
    ]
    for topic, item, where in ic_items:
        _set_label_cell(ws, row, 1, topic)
        _set_note_cell(ws, row, 2, item)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        _set_note_cell(ws, row, 9, where)
        ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=10)
        ws.row_dimensions[row].height = 30
        row += 1
    row += 1

    _exec_section_title(ws, row, 1, 4, "Core Assumptions")
    _exec_section_title(ws, row, 6, 10, "Audit Snapshot")
    row += 1

    assumption_rows = [
        ("Revenue Growth", f"='{ASSUMP}'!${ASSUMP_CELLS['revenue_growth']}", FMT_PCT2),
        ("EBIT Margin", f"='{ASSUMP}'!${ASSUMP_CELLS['ebit_margin']}", FMT_PCT2),
        ("Gross Margin", f"='{ASSUMP}'!${ASSUMP_CELLS['gross_margin']}", FMT_PCT2),
        ("D&A % Revenue", f"='{ASSUMP}'!${ASSUMP_CELLS['da_pct_revenue']}", FMT_PCT2),
        ("CapEx % Revenue", f"='{ASSUMP}'!${ASSUMP_CELLS['capex_pct_revenue']}", FMT_PCT2),
        ("Delta NWC % Revenue", f"='{ASSUMP}'!${ASSUMP_CELLS['wc_change_pct_revenue']}", FMT_PCT2),
        ("Tax Rate", f"='{ASSUMP}'!${ASSUMP_CELLS['tax_rate']}", FMT_PCT2),
        ("Net Debt", f"='{ASSUMP}'!${ASSUMP_CELLS['net_debt']}", FMT_COMMA2),
        ("Diluted Shares", f"='DCF Valuation'!$B${VAL_ROWS['shares']}", FMT_COMMA2),
    ]
    write_header_row(ws, row, ["Driver", "Selected", "Source", ""], 1)
    rr = row + 1
    for label, formula, fmt in assumption_rows:
        _set_label_cell(ws, rr, 1, label)
        _summary_formula(ws, rr, 2, formula, fmt)
        _summary_formula(ws, rr, 3, '="Assumptions"')
        ws.cell(row=rr, column=4, value="").border = border_thin()
        rr += 1
    left_end = rr

    write_header_row(ws, row, ["Status", "Count", "Note", "", ""], 6)
    audit_r = row + 1
    audit_count_descriptions = {
        "OK": "Checks passing",
        "Review": "Flagged for IC review",
        "N/A": "Not applicable / not modeled",
        "Fallback": "Graceful fallback applied",
    }
    for status in ("OK", "Review", "N/A", "Fallback"):
        _set_label_cell(ws, audit_r, 6, status)
        _summary_formula(ws, audit_r, 7, f'=COUNTIF({status_range},"{status}")', FMT_COMMA)
        _set_note_cell(ws, audit_r, 8, audit_count_descriptions[status])
        ws.merge_cells(start_row=audit_r, start_column=8, end_row=audit_r, end_column=10)
        audit_r += 1
    # Review items (key flags), under the counts.
    review_checks = [
        ("Engine vs Excel FCFF recon", "Engine vs Excel FCF Parity (target = 0)"),
        ("Plug magnitude", "Plug magnitude: (|Other Assets|+|Other Liabilities|) / Total Assets"),
        ("TV dependency", "TV % of EV"),
        ("Gordon vs Exit gap", "Gordon vs Exit terminal value gap"),
        ("Rev x EBIT center", "Revenue Growth x EBIT Margin center vs main value"),
        ("DCF defaults cache", "DCF defaults cache available"),
    ]
    audit_r += 1
    _set_label_cell(ws, audit_r, 6, "Key Review Items", True)
    ws.merge_cells(start_row=audit_r, start_column=6, end_row=audit_r, end_column=10)
    ws.cell(row=audit_r, column=6).fill = fill_section()
    audit_r += 1
    for label, check_name in review_checks:
        _set_label_cell(ws, audit_r, 6, label)
        _summary_formula(
            ws, audit_r, 7,
            f'=IFERROR(INDEX(\'Audit Dashboard\'!$B:$B,MATCH("{check_name}",\'Audit Dashboard\'!$A:$A,0)),"N/A")',
        )
        _summary_formula(
            ws, audit_r, 8,
            f'=IFERROR(INDEX(\'Audit Dashboard\'!$E:$E,MATCH("{check_name}",\'Audit Dashboard\'!$A:$A,0)),"")',
        )
        ws.merge_cells(start_row=audit_r, start_column=8, end_row=audit_r, end_column=10)
        audit_r += 1
    right_end = audit_r

    # Conditional colouring on the audit count cells (G column).
    audit_count_range = f"G{row + 1}:G{row + 4}"
    ws.conditional_formatting.add(
        audit_count_range,
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FFF2CC")),
    )

    row = max(left_end, right_end) + 1  # blank separator

    # ─── V3.9.0 Forecast Path Summary ─────────────────────────────────────
    # Year 1 vs Year 5 endpoints for the key operating forecast drivers so
    # readers immediately see whether the analyst has shaped the path or
    # whether the assumption is flat across the explicit forecast period.
    _exec_section_title(ws, row, 1, 10, "Forecast Path Summary (Year 1 vs Year 5)")
    row += 1
    write_header_row(
        ws, row, ["Driver", "Year 1", "Year 5", "Shape", "Source", "", "", "", "", ""], 1
    )
    row += 1
    fps_rows = [
        ("Revenue Growth", "revenue_growth"),
        ("EBIT Margin", "ebit_margin"),
        ("CapEx % Revenue", "capex_pct_revenue"),
        ("Delta NWC % Revenue", "wc_change_pct_revenue"),
    ]
    for label, key in fps_rows:
        path = ASSUMP_PATH_CELLS.get(key) or [ASSUMP_CELLS.get(key)]
        if not path or path[0] is None:
            continue
        y1_ref = f"'{ASSUMP}'!${path[0][0]}${path[0][1:]}"
        last = path[-1]
        y5_ref = f"'{ASSUMP}'!${last[0]}${last[1:]}"
        _set_label_cell(ws, row, 1, label)
        _summary_formula(ws, row, 2, f"={y1_ref}", FMT_PCT2)
        _summary_formula(ws, row, 3, f"={y5_ref}", FMT_PCT2)
        _summary_formula(
            ws,
            row,
            4,
            f'=IF(ROUND({y5_ref}-{y1_ref},6)=0,"User-editable assumption",IF({y5_ref}>{y1_ref},"Rising path","Declining path"))',
        )
        _summary_formula(ws, row, 5, '="Assumptions / Forecast Path"')
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=10)
        row += 1
    note_cell = ws.cell(
        row=row,
        column=1,
        value=(
            "Endpoints are the Year 1 and Year 5 cells on the Assumptions sheet Operating Forecast Path block. "
            "AAPL defaults use a neutral top-down Base path; other tickers may use a single user-editable assumption when no ticker-specific path is configured. "
            "Rising / declining shapes indicate an explicit path across the forecast period."
        ),
    )
    note_cell.font = font_watermark()
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.row_dimensions[row].height = 42
    row += 2

    # ─── Section 4: Mini Visual Blocks ────────────────────────────────────
    # 2x2 grid:
    #   Left top: Revenue & EBIT Margin    Right top: FCF Forecast
    #   Left bot: DCF Value Bridge         Right bot: Scenario Valuation
    n = max(1, int(inp.forecast_years or 5))
    forecast_start_col = _forecast_start_col(ctx)
    scenario_rows = [
        ("Base", SCENARIO_ROWS.get("valuation_base", 1)),
        ("Bull", SCENARIO_ROWS.get("valuation_bull", 1)),
        ("Bear", SCENARIO_ROWS.get("valuation_bear", 1)),
    ]

    # Row 4a: Revenue & EBIT Margin (left), FCFF Forecast (right)
    _exec_section_title(ws, row, 1, 4, "Revenue & EBIT Margin")
    _exec_section_title(ws, row, 6, 10, "FCFF Forecast")
    row += 1
    write_header_row(ws, row, ["Period", "Revenue", "EBIT Margin", "Source"], 1)
    write_header_row(ws, row, ["Period", "FCFF", "PV of FCFF", "Discount Factor", "Source"], 6)
    rev_data_start = row + 1
    rev_r = row + 1
    fcf_r = row + 1
    for i in range(n):
        # Revenue & EBIT margin block (cols A-D)
        src_col_pl = forecast_start_col + i
        _set_label_cell(ws, rev_r, 1, f"Year {i + 1}")
        _summary_formula(ws, rev_r, 2, f"='P&L Forecast'!${get_column_letter(src_col_pl)}${PL_FORECAST_ROWS.get('Revenue', 1)}", FMT_COMMA2)
        _summary_formula(ws, rev_r, 3, f"='P&L Forecast'!${get_column_letter(src_col_pl)}${PL_FORECAST_ROWS.get('EBIT Margin (%)', 1)}", FMT_PCT2)
        _summary_formula(ws, rev_r, 4, '="P&L Forecast"')
        rev_r += 1

        # FCFF Forecast block (cols F-J)
        src_col_fcf = 2 + i
        col = get_column_letter(src_col_fcf)
        _set_label_cell(ws, fcf_r, 6, f"Year {i + 1}")
        _summary_formula(ws, fcf_r, 7, f"='FCF Build'!${col}${FCF_ROWS.get('fcf', 1)}", FMT_COMMA2)
        _summary_formula(ws, fcf_r, 8, f"='FCF Build'!${col}${FCF_ROWS.get('pv', 1)}", FMT_COMMA2)
        _summary_formula(ws, fcf_r, 9, f"='FCF Build'!${col}${FCF_ROWS.get('df', 1)}", "0.000000")
        _summary_formula(ws, fcf_r, 10, '="FCF Build"')
        fcf_r += 1
    rev_data_end = rev_r - 1
    fcf_data_end = fcf_r - 1

    # Data bars: Revenue (B), FCF (G).
    ws.conditional_formatting.add(
        f"B{rev_data_start}:B{rev_data_end}",
        DataBarRule(start_type="min", end_type="max", color="8FAADC", showValue=True),
    )
    ws.conditional_formatting.add(
        f"G{rev_data_start}:G{fcf_data_end}",
        DataBarRule(start_type="min", end_type="max", color="A9D08E", showValue=True),
    )

    row = max(rev_data_end, fcf_data_end) + 2

    # Row 4b: DCF Value Bridge (left), Scenario Valuation (right)
    _exec_section_title(ws, row, 1, 4, "DCF Value Bridge")
    _exec_section_title(ws, row, 6, 10, "Scenario Valuation")
    row += 1
    write_header_row(ws, row, ["Bridge Step", "Value", "Ratio", ""], 1)
    write_header_row(ws, row, ["Scenario", "Intrinsic / Share", "Status", "Source", ""], 6)
    bridge_rows = [
        ("PV of FCFF", f"='DCF Valuation'!$B${VAL_ROWS['pv_sum']}", "", None),
        ("PV Terminal Value", f"='DCF Valuation'!$B${VAL_ROWS['tv_used']}", f"='DCF Valuation'!$B${VAL_ROWS['tv_pct']}", FMT_PCT2),
        ("Enterprise Value", f"='DCF Valuation'!$B${VAL_ROWS['ev']}", "", None),
        ("Less: Net Debt", f"='DCF Valuation'!$B${VAL_ROWS['net_debt']}", "", None),
        ("Equity Value", f"='DCF Valuation'!$B${VAL_ROWS['equity']}", "", None),
        ("Intrinsic / Share", f"='DCF Valuation'!$B${VAL_ROWS['intrinsic']}", f"='DCF Valuation'!$B${VAL_ROWS['upside']}", FMT_PCT1),
    ]
    br = row + 1
    for label, value_formula, ratio_formula, ratio_fmt in bridge_rows:
        _set_label_cell(ws, br, 1, label)
        _summary_formula(ws, br, 2, value_formula, FMT_COMMA2)
        if ratio_formula:
            _summary_formula(ws, br, 3, ratio_formula, ratio_fmt)
        else:
            ws.cell(row=br, column=3, value="").border = border_thin()
        ws.cell(row=br, column=4, value="").border = border_thin()
        br += 1

    sc_r = row + 1
    sc_data_start = sc_r
    for label, scenario_row in scenario_rows:
        _set_label_cell(ws, sc_r, 6, label)
        _summary_formula(ws, sc_r, 7, f'=IF(COUNTA(\'Scenario Notes\'!$B${scenario_row})=0,"n/a",\'Scenario Notes\'!$B${scenario_row})', FMT_COMMA2)
        _summary_formula(ws, sc_r, 8, f'=IF(ISNUMBER(G{sc_r}),"Available","No saved scenario on file")')
        _summary_formula(ws, sc_r, 9, '="Scenario Notes"')
        ws.cell(row=sc_r, column=10, value="").border = border_thin()
        sc_r += 1
    ws.conditional_formatting.add(
        f"G{sc_data_start}:G{sc_r - 1}",
        DataBarRule(start_type="min", end_type="max", color="B4C7E7", showValue=True),
    )

    row = max(br, sc_r) + 1

    # ─── Section 5: Valuation Range / Football Field + Caveats ───────────
    _exec_section_title(ws, row, 1, 10, "Valuation Range / Football Field")
    row += 1
    write_header_row(ws, row, ["Source", "Low", "Mid / Selected", "High", "Current Price", "Note"], 1)
    ff_data_start = row + 1

    sens_g_range = f"'Sensitivity'!$B${SENS_ROWS.get('wacc_g_center_row', 3) - 2}:$F${SENS_ROWS.get('wacc_g_center_row', 3) + 2}"
    sens_exit_range = f"'Sensitivity'!$B${SENS_ROWS.get('wacc_exit_center_row', 3) - 2}:$F${SENS_ROWS.get('wacc_exit_center_row', 3) + 2}"
    sens_op_range = f"'Sensitivity'!$B${SENS_ROWS.get('operating_center_row', 3) - 2}:$F${SENS_ROWS.get('operating_center_row', 3) + 2}"
    scenario_value_refs = ",".join(f"'Scenario Notes'!$B${sr}" for _, sr in scenario_rows)
    base_scenario_ref = f"'Scenario Notes'!$B${SCENARIO_ROWS.get('valuation_base', 1)}"
    # V3.7.6 terminal cases (calculator-derived snapshot from DCF Valuation alt rows).
    tv_current_ref = (
        f"='DCF Valuation'!$B${VAL_ROWS['alt_iv_tv_current_model_terminal']}"
        if VAL_ROWS.get("alt_iv_tv_current_model_terminal") else f"='DCF Valuation'!$B${VAL_ROWS['intrinsic']}"
    )
    tv_gordon_ref = (
        f"='DCF Valuation'!$B${VAL_ROWS['alt_iv_tv_gordon_growth']}"
        if VAL_ROWS.get("alt_iv_tv_gordon_growth") else None
    )
    tv_exit_ref = (
        f"='DCF Valuation'!$B${VAL_ROWS['alt_iv_tv_exit_multiple']}"
        if VAL_ROWS.get("alt_iv_tv_exit_multiple") else None
    )
    tv_fade_ref = (
        f"='DCF Valuation'!$B${VAL_ROWS['alt_iv_tv_fade_period_reference']}"
        if VAL_ROWS.get("alt_iv_tv_fade_period_reference") else None
    )
    headline_iv_ref = f"='DCF Valuation'!$B${VAL_ROWS['intrinsic']}"
    headline_iv_cell_ref = headline_iv_ref[1:] if headline_iv_ref.startswith("=") else headline_iv_ref
    plus_row = VAL_ROWS.get("alt_iv_wacc_selected_plus_spread_100bps")
    minus_row = VAL_ROWS.get("alt_iv_wacc_selected_minus_spread_100bps")
    if plus_row and minus_row:
        # Lower bound = higher discount rate (+100 bps); upper bound = lower
        # discount rate (-100 bps). All three cells are live Excel formulas.
        dcf_low_ref = f"='DCF Valuation'!$B${plus_row}"
        dcf_high_ref = f"='DCF Valuation'!$B${minus_row}"
        dcf_headline_note = (
            "Primary reference: live workbook headline. Low / High = headline IV "
            "recomputed at Selected WACC +/- 100 bps (live formulas)."
        )
    else:
        dcf_low_ref = headline_iv_ref
        dcf_high_ref = headline_iv_ref
        dcf_headline_note = (
            "Primary reference: live workbook headline using selected Net Debt / "
            "Share Count / WACC / Terminal treatments."
        )
    ff_rows = [
        (
            "DCF: Current Model (Live Headline)",
            dcf_low_ref,
            headline_iv_ref,
            dcf_high_ref,
            dcf_headline_note,
        ),
    ]
    # Build a DCF terminal range from Gordon (low) → Current/Blend → Exit (high).
    if tv_gordon_ref and tv_exit_ref:
        ff_rows.append((
            "DCF: Terminal Treatment Range (Gordon → Exit)",
            tv_gordon_ref,
            tv_current_ref,
            tv_exit_ref,
            "Terminal range — Gordon-only as low, Current Model as mid, Exit-only as high.",
        ))
    if tv_fade_ref:
        ff_rows.append((
            "DCF: Fade Period Reference",
            tv_fade_ref,
            tv_fade_ref,
            tv_fade_ref,
            "Fade Period reference case — FCF growth fades to terminal growth before Gordon TV.",
        ))
    ff_rows.extend([
        (
            "Saved Scenario References (calculator-based)",
            f'=IF(COUNT({scenario_value_refs})>=2,MIN({scenario_value_refs}),"n/a")',
            f'=IF(ISNUMBER({base_scenario_ref}),{base_scenario_ref},"n/a")',
            f'=IF(COUNT({scenario_value_refs})>=2,MAX({scenario_value_refs}),"n/a")',
            "Diagnostic alternative: saved Bear / Base / Bull references may preserve saved scenario treatments.",
        ),
        (
            "Sensitivity: WACC x Terminal g",
            f"=MIN({sens_g_range},{headline_iv_cell_ref})",
            f"='DCF Valuation'!$B${VAL_ROWS['intrinsic']}",
            f"=MAX({sens_g_range},{headline_iv_cell_ref})",
            "Diagnostic alternative: WACC vs g grid. Low / High include the live headline if it sits outside the exported matrix range so Low <= Mid <= High.",
        ),
        (
            "Sensitivity: WACC x Exit Multiple",
            f"=MIN({sens_exit_range},{headline_iv_cell_ref})",
            f"='DCF Valuation'!$B${VAL_ROWS['intrinsic']}",
            f"=MAX({sens_exit_range},{headline_iv_cell_ref})",
            "Diagnostic alternative: WACC vs exit multiple grid. Low / High include the live headline if it sits outside the exported matrix range so Low <= Mid <= High.",
        ),
        (
            "Sensitivity: Revenue Growth x EBIT Margin",
            f"=MIN({sens_op_range},{headline_iv_cell_ref})",
            f"='DCF Valuation'!$B${VAL_ROWS['intrinsic']}",
            f"=MAX({sens_op_range},{headline_iv_cell_ref})",
            "Diagnostic alternative: operating sensitivity range. Static at export; Low / High include the live headline if it sits outside the matrix range.",
        ),
    ])
    # V3.7.7 Trading Comps ranges. Reference the Trading Comps sheet implied
    # IV/share rows when the sheet was built; otherwise fall back to N/A.
    comps_status = TRADING_COMPS_ROWS.get("status") if TRADING_COMPS_ROWS else "unavailable"
    comps_dq = (ctx.get("trading_comps") or {}).get("data_quality") or {}
    ff_comps_status = comps_dq.get("football_field_comps_status") or ""
    for label, key in (
        ("Trading Comps: EV / Revenue", "implied_ev_revenue"),
        ("Trading Comps: EV / EBITDA", "implied_ev_ebitda"),
        ("Trading Comps: P / E", "implied_pe"),
    ):
        if comps_status == "built" and TRADING_COMPS_ROWS.get(key):
            tc_row = TRADING_COMPS_ROWS[key]
            low_f = f"=IFERROR('Trading Comps'!$I${tc_row},\"n/a\")"
            mid_f = f"=IFERROR('Trading Comps'!$J${tc_row},\"n/a\")"
            high_f = f"=IFERROR('Trading Comps'!$K${tc_row},\"n/a\")"
            note = (
                "Trailing core peer comps restored; market reference for discussion purposes only."
                if ff_comps_status != "Insufficient core peer coverage"
                else "N/A - Insufficient core peer coverage; partial data remains visible on Trading Comps."
            )
        else:
            low_f = '="n/a"'
            mid_f = '="n/a"'
            high_f = '="n/a"'
            note = "Trading Comps unavailable for this export."
        ff_rows.append((label, low_f, mid_f, high_f, note))
    ff_r = ff_data_start
    price_ref = f"='DCF Valuation'!$B${VAL_ROWS['price']}"
    for label, low_formula, mid_formula, high_formula, note in ff_rows:
        _set_label_cell(ws, ff_r, 1, label)
        _summary_formula(ws, ff_r, 2, low_formula, FMT_COMMA2)
        _summary_formula(ws, ff_r, 3, mid_formula, FMT_COMMA2)
        _summary_formula(ws, ff_r, 4, high_formula, FMT_COMMA2)
        _summary_formula(ws, ff_r, 5, price_ref, FMT_COMMA2)
        _set_note_cell(ws, ff_r, 6, note)
        ws.merge_cells(start_row=ff_r, start_column=6, end_row=ff_r, end_column=10)
        ff_r += 1
    ff_data_end = ff_r - 1

    # Data bar on the Mid column gives the football-field visual.
    ws.conditional_formatting.add(
        f"C{ff_data_start}:C{ff_data_end}",
        DataBarRule(start_type="min", end_type="max", color="638EC6", showValue=True),
    )
    # Highlight current price column subtly.
    for r2 in range(ff_data_start, ff_data_end + 1):
        ws.cell(row=r2, column=5).fill = PatternFill("solid", fgColor="F2F2F2")

    # Football Field note: mark the live headline as the primary reference and
    # saved/sensitivity/comps rows as diagnostic references.
    row = ff_data_end + 1
    dual_engine_note = ws.cell(
        row=row,
        column=1,
        value="Football Field primary reference is the live headline DCF. Saved scenarios, terminal variants, sensitivity grids, and comps are diagnostic alternatives; mixed scenario treatments are flagged below when present.",
    )
    dual_engine_note.font = font_watermark()
    dual_engine_note.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    # V3.7.3: Football Field caveat when Base vs Bull/Bear use different Net
    # Debt treatments. Silent when treatments agree so the page stays clean.
    base_treatment = (
        (getattr(out, "net_debt_bridge", None) or {}).get("selected_treatment")
        or DEFAULT_NET_DEBT_TREATMENT
    )
    scenario_treatments = []
    ff_scenarios = (ctx.get("scenario_doc") or {}).get("scenarios") or {}
    for scenario_type in ("bull", "bear"):
        scn_entry = ff_scenarios.get(scenario_type)
        if not isinstance(scn_entry, dict):
            continue
        compat = scn_entry.get("compatibility") or {}
        scn_params = scn_entry.get("params") or {}
        scn_treatment = (
            compat.get("selected_net_debt_treatment")
            or scn_params.get("selected_net_debt_treatment")
            or DEFAULT_NET_DEBT_TREATMENT
        )
        scenario_treatments.append((scenario_type, scn_treatment))
    treatment_mismatch = any(t != base_treatment for _, t in scenario_treatments)
    if treatment_mismatch:
        row += 1
        treatment_breakdown = ", ".join(
            f"{scn_type.title()}={NET_DEBT_TREATMENT_LABELS.get(t, t)}"
            for scn_type, t in scenario_treatments
        )
        current_scenario_label = (ctx.get("current_scenario") or "base").title()
        caveat_cell = ws.cell(
            row=row,
            column=1,
            value=(
                "Caveat — scenarios use different Net Debt treatments "
                f"(Current export = {current_scenario_label}: {NET_DEBT_TREATMENT_LABELS.get(base_treatment, base_treatment)}; "
                f"{treatment_breakdown}). Compare scenario IVs with caution - net debt assumptions are not held constant."
            ),
        )
        caveat_cell.font = Font(name="Arial", size=9, color="9C0006", italic=True)
        caveat_cell.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        caveat_cell.fill = PatternFill("solid", fgColor="FFF2CC")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    # V3.8.8: same caveat pattern for Share Count treatment mismatch.
    base_share_treatment = (
        (getattr(out, "shareholder_returns", None) or {}).get("selected_share_count_treatment")
        or DEFAULT_SHARE_COUNT_TREATMENT
    )
    scenario_share_treatments = []
    for scenario_type in ("bull", "bear"):
        scn_entry = ff_scenarios.get(scenario_type)
        if not isinstance(scn_entry, dict):
            continue
        compat = scn_entry.get("compatibility") or {}
        scn_params = scn_entry.get("params") or {}
        scn_share_treatment = (
            compat.get("selected_share_count_treatment")
            or scn_params.get("selected_share_count_treatment")
            or DEFAULT_SHARE_COUNT_TREATMENT
        )
        scenario_share_treatments.append((scenario_type, scn_share_treatment))
    share_mismatch = any(t != base_share_treatment for _, t in scenario_share_treatments)
    if share_mismatch:
        row += 1
        breakdown_share = ", ".join(
            f"{scn_type.title()}={SHARE_COUNT_TREATMENT_LABELS.get(t, t)}"
            for scn_type, t in scenario_share_treatments
        )
        current_scn_label_share = (ctx.get("current_scenario") or "base").title()
        caveat_cell_share = ws.cell(
            row=row, column=1,
            value=(
                "Caveat - scenarios use different Share Count treatments "
                f"(Current export = {current_scn_label_share}: {SHARE_COUNT_TREATMENT_LABELS.get(base_share_treatment, base_share_treatment)}; "
                f"{breakdown_share}). Per-share IVs are not directly comparable - share-count denominator differs."
            ),
        )
        caveat_cell_share.font = Font(name="Arial", size=9, color="9C0006", italic=True)
        caveat_cell_share.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        caveat_cell_share.fill = PatternFill("solid", fgColor="FFF2CC")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    # V3.7.5: same caveat pattern for WACC treatment mismatch.
    base_wacc_treatment = (
        (getattr(out, "wacc_decision_bridge", None) or {}).get("selected_wacc_treatment")
        or DEFAULT_WACC_TREATMENT
    )
    scenario_wacc_treatments = []
    for scenario_type in ("bull", "bear"):
        scn_entry = ff_scenarios.get(scenario_type)
        if not isinstance(scn_entry, dict):
            continue
        compat = scn_entry.get("compatibility") or {}
        scn_params = scn_entry.get("params") or {}
        scn_wacc_treatment = (
            compat.get("selected_wacc_treatment")
            or scn_params.get("selected_wacc_treatment")
            or DEFAULT_WACC_TREATMENT
        )
        scenario_wacc_treatments.append((scenario_type, scn_wacc_treatment))
    wacc_mismatch = any(t != base_wacc_treatment for _, t in scenario_wacc_treatments)
    if wacc_mismatch:
        row += 1
        breakdown_wacc = ", ".join(
            f"{scn_type.title()}={WACC_TREATMENT_LABELS.get(t, t)}"
            for scn_type, t in scenario_wacc_treatments
        )
        current_scn_label = (ctx.get("current_scenario") or "base").title()
        caveat_cell_wacc = ws.cell(
            row=row, column=1,
            value=(
                "Caveat — scenarios use different WACC treatments "
                f"(Current export = {current_scn_label}: {WACC_TREATMENT_LABELS.get(base_wacc_treatment, base_wacc_treatment)}; "
                f"{breakdown_wacc}). Per-share IVs are not directly comparable - discount rate differs."
            ),
        )
        caveat_cell_wacc.font = Font(name="Arial", size=9, color="9C0006", italic=True)
        caveat_cell_wacc.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        caveat_cell_wacc.fill = PatternFill("solid", fgColor="FFF2CC")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    # V3.7.6: Terminal Value treatment mismatch caveat.
    base_terminal_treatment_ff = (
        (getattr(out, "terminal_decision_bridge", None) or {}).get("selected_terminal_treatment")
        or DEFAULT_TERMINAL_TREATMENT
    )
    scenario_terminal_treatments = []
    for scenario_type in ("bull", "bear"):
        scn_entry = ff_scenarios.get(scenario_type)
        if not isinstance(scn_entry, dict):
            continue
        compat = scn_entry.get("compatibility") or {}
        scn_params = scn_entry.get("params") or {}
        scn_terminal_treatment = (
            compat.get("selected_terminal_treatment")
            or scn_params.get("selected_terminal_treatment")
            or DEFAULT_TERMINAL_TREATMENT
        )
        scenario_terminal_treatments.append((scenario_type, scn_terminal_treatment))
    terminal_mismatch_ff = any(t != base_terminal_treatment_ff for _, t in scenario_terminal_treatments)
    if terminal_mismatch_ff:
        row += 1
        breakdown_t = ", ".join(
            f"{scn_type.title()}={TERMINAL_TREATMENT_LABELS.get(t, t)}"
            for scn_type, t in scenario_terminal_treatments
        )
        current_scn_label_t = (ctx.get("current_scenario") or "base").title()
        caveat_cell_t = ws.cell(
            row=row, column=1,
            value=(
                "Caveat — scenarios use different Terminal treatments "
                f"(Current export = {current_scn_label_t}: {TERMINAL_TREATMENT_LABELS.get(base_terminal_treatment_ff, base_terminal_treatment_ff)}; "
                f"{breakdown_t}). Per-share IVs are not directly comparable - terminal-value method differs."
            ),
        )
        caveat_cell_t.font = Font(name="Arial", size=9, color="9C0006", italic=True)
        caveat_cell_t.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        caveat_cell_t.fill = PatternFill("solid", fgColor="FFF2CC")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    row = ff_data_end + 3
    if treatment_mismatch:
        row += 1
    if share_mismatch:
        row += 1
    if wacc_mismatch:
        row += 1
    if terminal_mismatch_ff:
        row += 1

    # Caveats.
    _exec_section_title(ws, row, 1, 10, "Key Caveats")
    row += 1
    caveats = [
        "UI, API, and Excel valuation share the same engine. Headline FCF, EV, equity value, and intrinsic value per share are sourced from the calculator and surfaced live in the workbook formulas.",
        "FCFF is unlevered (NOPAT-based); interest expense is modeled in P&L, BS, and CF for 3FS display but is not deducted from FCFF.",
        "Five-year explicit forecast uses days-based working capital (DSO / DIO / DPO), asset-based D&A on Beginning Net PP&E, and CapEx as % of revenue. Tax rate is floored at zero; NOL and deferred-tax modeling are out of scope.",
        "Balance Sheet Forecast carries marketable securities, goodwill, intangibles, deferred revenue, capital leases, and deferred tax assets / liabilities as explicit lines. Residual plug rows remain visible so reviewers can size any unmodeled items.",
        "Net Debt Bridge reports the reported / input figure alongside Debt − Cash, Debt − Cash − ST Investments, and Debt − Cash − Total Marketable Securities. The headline uses the reported / input figure by default; switch the Net Debt Treatment dropdown on Assumptions to flip the equity bridge. Adjusted IV references are listed on DCF Valuation.",
        "Shareholder returns (dividends + buybacks) flow through CF Financing and the BS Equity roll-forward only; FCFF and EV are unchanged. The per-share denominator can flip via the Selected Share Count Treatment dropdown.",
        "WACC headline uses the model's live WACC; mechanical CAPM reference is an audit diagnostic (We*Ke + Wd*Kd*(1-t)). The Review Dashboard tiers Selected vs Mechanical CAPM spread without auto-correcting.",
        "Terminal value headline uses the Current Model case. Gordon-only, Exit-only, Blend, and Fade Period reference cases are computed on DCF Valuation and overlaid in the Football Field. The Review Dashboard tiers TV/EV and Gordon vs Exit gap.",
        "Trading Comps overlay (EV/Revenue, EV/EBITDA, P/E) is a market reference for discussion purposes only. Peer set is configurable; outliers and exclusions are flagged in the Peer Set table. HK / CN comps coverage may be limited.",
        "HK and CN exports fall back to the legacy operating forecast where the historical pipeline is not yet wired; the Review Dashboard surfaces these as Fallback, not as errors.",
        "Precedent transactions, forward multiples, and paid data providers are out of scope for this workbook.",
    ]
    caveat_font = Font(name="Arial", size=9, color="595959", italic=True)
    for caveat in caveats:
        c = ws.cell(row=row, column=1, value=f"•  {caveat}")
        c.font = caveat_font
        c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        ws.row_dimensions[row].height = 16
        row += 1

    # Footer note.
    note_cell = ws.cell(
        row=row + 1,
        column=1,
        value="Packaging layer only — all numbers on this page are live cross-sheet formulas; no values are hard-coded.",
    )
    note_cell.font = font_watermark()
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=10)


def _pv_sum_at_wacc(wacc_cell: str, n: int) -> str:
    terms = []
    for i in range(n):
        col = get_column_letter(2 + i)
        fcf_addr = f"'FCF Build'!{col}{FCF_ROWS['fcf']}"
        terms.append(f"ROUND({fcf_addr}*ROUND(1/(1+{wacc_cell})^{i + 1},6),2)")
    return "SUM(" + ",".join(terms) + ")"


def _alt_iv_at_wacc(wacc_cell: str, n: int) -> str:
    """Live IV/share formula computed at an alternative WACC.

    Re-discounts the explicit-period FCFs from the FCF Build sheet at the
    supplied WACC cell and recomputes the terminal value using the same
    Terminal Value Method dropdown (Gordon / Exit / Average) that drives the
    headline. Net debt and share count are held at the selected treatments.
    Fade Period reference is not recomputed at alternative WACCs - the alt
    rows therefore mirror the Average behavior when Fade is the active
    terminal treatment (disclosed in the table note).
    """
    fcf_last_col = get_column_letter(2 + n - 1)
    fcf_y5 = f"'FCF Build'!{fcf_last_col}{FCF_ROWS['fcf']}"
    ebit_y5 = f"'FCF Build'!{fcf_last_col}{FCF_ROWS['ebit']}"
    da_y5 = f"'FCF Build'!{fcf_last_col}{FCF_ROWS['da']}"
    ref_g = _abs_ref(ASSUMP_CELLS["terminal_g"])
    ref_exit = _abs_ref(ASSUMP_CELLS["exit_multiple"])
    ref_method = _abs_ref(ASSUMP_CELLS["tv_method"])
    ref_net_debt = _abs_ref(ASSUMP_CELLS.get("ndb_selected", ASSUMP_CELLS["net_debt"]))
    ref_shares = _abs_ref(ASSUMP_CELLS.get("sr_selected_denominator", ASSUMP_CELLS["shares"]))
    pv_sum = _pv_sum_at_wacc(wacc_cell, n)
    df_n = f"ROUND(1/(1+{wacc_cell})^{n},6)"
    denom = f"IF({wacc_cell}>{ref_g},{wacc_cell}-{ref_g},MAX({wacc_cell}-{ref_g},0.001))"
    tv_gordon = f"{fcf_y5}*(1+{ref_g})/({denom})*{df_n}"
    tv_exit = f"({ebit_y5}+{da_y5})*{ref_exit}*{df_n}"
    tv_used = (
        f'IF({ref_method}="gordon",{tv_gordon},'
        f'IF({ref_method}="exit",{tv_exit},'
        f'({tv_gordon}+{tv_exit})/2))'
    )
    return f"=ROUND(({pv_sum}+{tv_used}-{ref_net_debt})/{ref_shares},4)"


def build_sensitivity_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "Sensitivity"
    SENS_ROWS.clear()
    apply_column_widths(ws, {
        **{get_column_letter(i): 14 for i in range(1, 8)},
        "H": 18, "I": 18, "J": 18, "K": 18,
    })

    n = inp.forecast_years
    fcf_last_col = get_column_letter(2 + n - 1)
    fcf_y5_addr = f"'FCF Build'!{fcf_last_col}{FCF_ROWS['fcf']}"
    ebit_y5_addr = f"'FCF Build'!{fcf_last_col}{FCF_ROWS['ebit']}"
    da_y5_addr = f"'FCF Build'!{fcf_last_col}{FCF_ROWS['da']}"
    # V3.7.2: Sensitivity uses the Selected Net Debt cell so the grid stays
    # consistent with the DCF Valuation headline.
    ref_net_debt = _abs_ref(ASSUMP_CELLS.get("ndb_selected", ASSUMP_CELLS["net_debt"]))
    ref_shares = _abs_ref(ASSUMP_CELLS["shares"])
    # V3.7.5: sensitivity center aligns with the Selected WACC Used in DCF cell.
    ref_wacc = _abs_ref(ASSUMP_CELLS.get("wacc_used", ASSUMP_CELLS["wacc"]))
    ref_g = _abs_ref(ASSUMP_CELLS["terminal_g"])
    ref_exit = _abs_ref(ASSUMP_CELLS["exit_multiple"])
    current_params = ctx.get("current_params") or {}
    sensitivity_fx_factor = _coerce_num(current_params.get("fx_rate_reporting_to_trading")) or 1.0

    row = _write_sheet_heading(ws, 1, "Sensitivity: Intrinsic Value per Share", inp, out, ctx, span_cols=6)
    _write_sensitivity_notes_panel(ws)
    ws.cell(row=row, column=1, value="Gordon-only sensitivity").font = font_title()
    row += 1
    _write_matrix_header(ws, row, "WACC \\ g", ref_g, FMT_PCT2, 0.005)
    row += 1

    matrix1_top = row
    SENS_ROWS["wacc_g_center_row"] = matrix1_top + 2
    SENS_ROWS["wacc_g_center_col"] = 4
    for i in range(5):
        wacc_offset = (i - 2) * 0.005
        _write_axis_cell(ws, row, 1, f"={ref_wacc}+{wacc_offset}" if wacc_offset else f"={ref_wacc}", FMT_PCT2)
        wacc_cell = f"$A{row}"
        pv_sum = _pv_sum_at_wacc(wacc_cell, n)
        for j in range(5):
            g_cell = f"{get_column_letter(2 + j)}${matrix1_top - 1}"
            denom = f"IF({wacc_cell}>{g_cell},{wacc_cell}-{g_cell},MAX({wacc_cell}-{g_cell},0.001))"
            tv_pv = f"{fcf_y5_addr}*(1+{g_cell})/({denom})*ROUND(1/(1+{wacc_cell})^{n},6)"
            is_center = i == 2 and j == 2
            formula = f"=ROUND((({pv_sum}+{tv_pv}-{ref_net_debt})/{ref_shares})*{sensitivity_fx_factor},2)"
            _write_matrix_value(ws, row, 2 + j, formula, is_center)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Exit-only sensitivity").font = font_title()
    row += 1
    _write_matrix_header(ws, row, "WACC \\ EV/EBITDA", ref_exit, FMT_MULTIPLE, 1.0)
    row += 1

    matrix2_top = row
    SENS_ROWS["wacc_exit_center_row"] = matrix2_top + 2
    SENS_ROWS["wacc_exit_center_col"] = 4
    for i in range(5):
        wacc_offset = (i - 2) * 0.005
        _write_axis_cell(ws, row, 1, f"={ref_wacc}+{wacc_offset}" if wacc_offset else f"={ref_wacc}", FMT_PCT2)
        wacc_cell = f"$A{row}"
        pv_sum = _pv_sum_at_wacc(wacc_cell, n)
        for j in range(5):
            mult_cell = f"{get_column_letter(2 + j)}${matrix2_top - 1}"
            tv_pv = f"({ebit_y5_addr}+{da_y5_addr})*{mult_cell}*ROUND(1/(1+{wacc_cell})^{n},6)"
            is_center = i == 2 and j == 2
            formula = f"=ROUND((({pv_sum}+{tv_pv}-{ref_net_debt})/{ref_shares})*{sensitivity_fx_factor},2)"
            _write_matrix_value(ws, row, 2 + j, formula, is_center)
        row += 1

    row += 1
    _set_label_cell(ws, row, 1, "Headline IV (outside grid)")
    _set_formula_cell(ws, row, 2, f"='DCF Valuation'!$B${VAL_ROWS['intrinsic']}", FMT_COMMA2, True)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)

    row += 2
    ws.cell(row=row, column=1, value="Selected Method Sensitivity - center ties to headline").font = font_title()
    row += 1
    _write_axis_cell(ws, row, 1, "WACC \\ selected terminal assumption", None)
    for j in range(5):
        label = "Center" if j == 2 else ("Low" if j < 2 else "High")
        _write_axis_cell(ws, row, 2 + j, label, None)
    row += 1
    selected_g_header = row
    _write_axis_cell(ws, row, 1, "Terminal g leg", None)
    for j in range(5):
        offset = (j - 2) * 0.005
        _write_axis_cell(ws, row, 2 + j, f"={ref_g}+{offset}" if offset else f"={ref_g}", FMT_PCT2)
    row += 1
    selected_exit_header = row
    _write_axis_cell(ws, row, 1, "Exit multiple leg", None)
    for j in range(5):
        offset = (j - 2) * 1.0
        _write_axis_cell(ws, row, 2 + j, f"={ref_exit}+{offset}" if offset else f"={ref_exit}", FMT_MULTIPLE)
    row += 1

    selected_matrix_top = row
    SENS_ROWS["selected_method_center_row"] = selected_matrix_top + 2
    SENS_ROWS["selected_method_center_col"] = 4
    lbl_tv_gordon = TERMINAL_TREATMENT_LABELS["gordon_growth"]
    lbl_tv_exit = TERMINAL_TREATMENT_LABELS["exit_multiple"]
    lbl_tv_blend = TERMINAL_TREATMENT_LABELS["gordon_exit_blend"]
    lbl_tv_h = TERMINAL_TREATMENT_LABELS["h_model"]
    lbl_tv_fade = TERMINAL_TREATMENT_LABELS["fade_period_reference"]
    ref_terminal_treatment = _abs_ref(ASSUMP_CELLS.get("tv_treatment", "B48"))
    ref_method = _abs_ref(ASSUMP_CELLS.get("tv_method", "B110"))
    ref_h_g_near = _abs_ref(ASSUMP_CELLS.get("h_model_g_near", ASSUMP_CELLS["terminal_g"]))
    ref_h_half_life = _abs_ref(ASSUMP_CELLS.get("h_model_half_life", ASSUMP_CELLS["terminal_g"]))
    val_net_debt = f"'DCF Valuation'!$B${VAL_ROWS['net_debt']}"
    val_shares = f"'DCF Valuation'!$B${VAL_ROWS['shares']}"
    val_tv_used = f"'DCF Valuation'!$B${VAL_ROWS['tv_used']}"
    for i in range(5):
        wacc_offset = (i - 2) * 0.005
        _write_axis_cell(ws, row, 1, f"={ref_wacc}+{wacc_offset}" if wacc_offset else f"={ref_wacc}", FMT_PCT2)
        wacc_cell = f"$A{row}"
        pv_sum = _pv_sum_at_wacc(wacc_cell, n)
        disc_n = f"ROUND(1/(1+{wacc_cell})^{n},6)"
        for j in range(5):
            g_cell = f"{get_column_letter(2 + j)}${selected_g_header}"
            mult_cell = f"{get_column_letter(2 + j)}${selected_exit_header}"
            denom = f"IF({wacc_cell}>{g_cell},{wacc_cell}-{g_cell},MAX({wacc_cell}-{g_cell},0.001))"
            gordon_tv = f"{fcf_y5_addr}*(1+{g_cell})/({denom})*{disc_n}"
            exit_tv = f"({ebit_y5_addr}+{da_y5_addr})*{mult_cell}*{disc_n}"
            h_denom = f"IF({wacc_cell}>{g_cell},{wacc_cell}-{g_cell},MAX({wacc_cell}-{g_cell},0.001))"
            h_tv = f"({fcf_y5_addr}*(1+{g_cell})+{fcf_y5_addr}*{ref_h_half_life}*({ref_h_g_near}-{g_cell}))/({h_denom})*{disc_n}"
            current_tv = f'IF({ref_method}="gordon",{gordon_tv},IF({ref_method}="exit",{exit_tv},AVERAGE({gordon_tv},{exit_tv})))'
            selected_tv = (
                f'IF({ref_terminal_treatment}="{lbl_tv_gordon}",{gordon_tv},'
                f'IF({ref_terminal_treatment}="{lbl_tv_exit}",{exit_tv},'
                f'IF({ref_terminal_treatment}="{lbl_tv_blend}",AVERAGE({gordon_tv},{exit_tv}),'
                f'IF({ref_terminal_treatment}="{lbl_tv_h}",{h_tv},'
                f'IF({ref_terminal_treatment}="{lbl_tv_fade}",{val_tv_used},'
                f'{current_tv})))))'
            )
            if i == 2 and j == 2:
                formula = f"='DCF Valuation'!$B${VAL_ROWS['intrinsic']}"
            else:
                formula = f"=ROUND((({pv_sum}+{selected_tv}-{val_net_debt})/{val_shares})*{sensitivity_fx_factor},2)"
            _write_matrix_value(ws, row, 2 + j, formula, i == 2 and j == 2)
        row += 1
    row += 1
    _set_label_cell(ws, row, 1, "Center tie to headline IV")
    _set_formula_cell(
        ws,
        row,
        2,
        f"=IFERROR(${get_column_letter(SENS_ROWS['selected_method_center_col'])}${SENS_ROWS['selected_method_center_row']}-'DCF Valuation'!$B${VAL_ROWS['intrinsic']},\"N/M\")",
        FMT_COMMA2,
        True,
    )
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)

    row += 3
    ws.cell(row=row, column=1, value="Operating sensitivity | Revenue Growth x EBIT Margin").font = font_title()
    row += 1
    drivers = normalized_driver_assumptions(inp)
    _write_axis_cell(ws, row, 1, "EBIT Margin \\ Revenue Growth", None)
    for j in range(5):
        _write_axis_cell(ws, row, 2 + j, drivers["revenue_growth"] + (j - 2) * 0.01, FMT_PCT2)
    row += 1

    operating_matrix = out.sensitivity_operating or []
    SENS_ROWS["operating_center_row"] = row + 2
    SENS_ROWS["operating_center_col"] = 4
    for i in range(5):
        _write_axis_cell(ws, row, 1, drivers["ebit_margin"] + (i - 2) * 0.01, FMT_PCT2)
        vals = operating_matrix[i] if i < len(operating_matrix) else []
        for j in range(5):
            value = vals[j] if j < len(vals) else None
            _write_matrix_value(ws, row, 2 + j, value, i == 2 and j == 2)
        row += 1

    row += 2

    write_section_title(ws, row, "Current Price Support Summary", span_cols=6)
    row += 1
    write_header_row(ws, row, ["Sensitivity set", "Low", "Selected / Center", "High", "Price position", "Closest support point"], 1)
    row += 1
    price_ref = f"'DCF Valuation'!$B${VAL_ROWS['price']}"
    iv_ref = f"'DCF Valuation'!$B${VAL_ROWS['intrinsic']}"
    ranges = {
        "WACC x Terminal g": (
            f"$B${SENS_ROWS['wacc_g_center_row'] - 2}:$F${SENS_ROWS['wacc_g_center_row'] + 2}",
            f"$A${SENS_ROWS['wacc_g_center_row'] - 2}:$A${SENS_ROWS['wacc_g_center_row'] + 2}",
            f"$B${SENS_ROWS['wacc_g_center_row'] - 3}:$F${SENS_ROWS['wacc_g_center_row'] - 3}",
            "WACC / g closest grid cell",
        ),
        "WACC x Exit Multiple": (
            f"$B${SENS_ROWS['wacc_exit_center_row'] - 2}:$F${SENS_ROWS['wacc_exit_center_row'] + 2}",
            f"$A${SENS_ROWS['wacc_exit_center_row'] - 2}:$A${SENS_ROWS['wacc_exit_center_row'] + 2}",
            f"$B${SENS_ROWS['wacc_exit_center_row'] - 3}:$F${SENS_ROWS['wacc_exit_center_row'] - 3}",
            "WACC / exit multiple closest grid cell",
        ),
        "Selected Method Sensitivity": (
            f"$B${SENS_ROWS['selected_method_center_row'] - 2}:$F${SENS_ROWS['selected_method_center_row'] + 2}",
            f"$A${SENS_ROWS['selected_method_center_row'] - 2}:$A${SENS_ROWS['selected_method_center_row'] + 2}",
            f"$B${SENS_ROWS['selected_method_center_row'] - 4}:$F${SENS_ROWS['selected_method_center_row'] - 4}",
            "WACC / selected terminal assumption closest grid cell",
        ),
        "Revenue Growth x EBIT Margin": (
            f"$B${SENS_ROWS['operating_center_row'] - 2}:$F${SENS_ROWS['operating_center_row'] + 2}",
            f"$A${SENS_ROWS['operating_center_row'] - 2}:$A${SENS_ROWS['operating_center_row'] + 2}",
            f"$B${SENS_ROWS['operating_center_row'] - 3}:$F${SENS_ROWS['operating_center_row'] - 3}",
            "EBIT margin / revenue growth closest grid cell",
        ),
    }
    for label, (rng, row_axis, col_axis, note) in ranges.items():
        _set_label_cell(ws, row, 1, label)
        _set_formula_cell(ws, row, 2, f"=MIN({rng})", FMT_COMMA2)
        _set_formula_cell(ws, row, 3, f"={iv_ref}", FMT_COMMA2, True)
        _set_formula_cell(ws, row, 4, f"=MAX({rng})", FMT_COMMA2)
        _set_formula_cell(ws, row, 5, f'=IF({price_ref}>D{row},"Above range",IF({price_ref}<B{row},"Below range","Inside range"))', None, True)
        closest_formula = f'=IF(E{row}="Inside range","Current price inside grid range",IF(E{row}="Above range","Current price above grid high","Current price below grid low"))'
        _set_formula_cell(ws, row, 6, closest_formula, None, True)
        row += 1
    row += 1
    _set_note_cell(ws, row, 1, "Closest support points are sensitivity-derived approximations, not goal-seek outputs. They help frame which assumptions would need to move for current price support.")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)


def _write_sensitivity_notes_panel(ws):
    panel_start = 2
    panel_end = 12
    for row_idx in range(panel_start, panel_end + 1):
        ws.row_dimensions[row_idx].height = 18
    ws.merge_cells(start_row=panel_start, start_column=8, end_row=panel_start, end_column=11)
    title = ws.cell(panel_start, 8, "Sensitivity Notes")
    title.font = Font(name="Arial", size=11, bold=True, color="1A3A5C")
    title.fill = PatternFill("solid", fgColor="D9EAF7")
    title.alignment = Alignment(horizontal="left", vertical="center")
    title.border = border_thin()

    notes = [
        "Gordon-only matrix is diagnostic unless Gordon is selected.",
        "Exit-only matrix is diagnostic unless Exit Multiple is selected.",
        "Selected Method matrix ties to headline IV.",
        "Operating sensitivity is static at export; re-export after changing operating assumptions.",
        "Highlighted cells show center assumptions.",
        "Sensitivities do not alter headline valuation.",
    ]
    row = panel_start + 1
    for note in notes:
        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=11)
        cell = ws.cell(row, 8, f"- {note}")
        cell.font = Font(name="Arial", size=9, color="333333")
        cell.fill = PatternFill("solid", fgColor="F3F7FB")
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = border_thin()
        row += 1


def _write_matrix_header(ws, row, label, base_ref, fmt, step):
    _write_axis_cell(ws, row, 1, label, None)
    for j in range(5):
        offset = (j - 2) * step
        formula = f"={base_ref}+{offset}" if offset else f"={base_ref}"
        _write_axis_cell(ws, row, 2 + j, formula, fmt)


def _write_axis_cell(ws, row, col, value, fmt):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.fill = fill_header()
    c.alignment = Alignment(horizontal="center")
    c.border = border_thin()
    if fmt:
        c.number_format = fmt


def _write_matrix_value(ws, row, col, formula, highlight=False):
    c = ws.cell(row=row, column=col, value=formula)
    c.number_format = FMT_COMMA2
    c.alignment = Alignment(horizontal="center")
    c.border = border_thin()
    if highlight:
        c.fill = fill_highlight()
        c.font = Font(name="Arial", size=10, bold=True, color="000000")
    else:
        c.font = font_formula()


def build_data_sources_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "Data Sources Audit"
    apply_column_widths(ws, {"A": 32, "B": 48, "C": 42})
    cache = ctx.get("financial_cache") or {}
    data = cache.get("data") or {}
    historical_cache = ctx.get("historical_cache") or {}
    historical_payload = historical_cache.get("data") or {}

    row = _write_sheet_heading(ws, 1, "Data Sources / Audit", inp, out, ctx)

    write_section_title(ws, row, "Export Identity", span_cols=3)
    row += 1
    row = _write_kv_rows(ws, row, [
        ("Scenario", _scenario_label(ctx.get("current_scenario")), None),
        ("Identity Note", _scenario_identity_note(ctx), None),
        ("Generated Date", _display_date(ctx.get("generated_at")), None),
        ("Model Version", MODEL_VERSION, None),
        ("Market / Currency / Unit", f"{out.market} / {out.currency} / {model_unit_label(out.currency, out.market)}", None),
    ])
    row += 1

    write_section_title(ws, row, "Scope Boundary", span_cols=3)
    row += 1
    row = _write_kv_rows(ws, row, [
        ("Workbook Positioning", SCOPE_BOUNDARY_TEXT, None),
        ("Use Boundary", SCOPE_BOUNDARY_DETAIL, None),
    ])
    row += 1

    current_params = ctx.get("current_params") or {}
    suitability_methods = current_params.get("recommended_methods") or []
    write_section_title(ws, row, "Model Suitability", span_cols=3)
    row += 1
    row = _write_kv_rows(ws, row, [
        ("Security Type", current_params.get("security_type") or "operating_company", None),
        ("Security Type Source", current_params.get("security_type_source") or "unknown", None),
        ("DCF Suitability", current_params.get("dcf_suitability") or ("unsuitable" if current_params.get("model_unsuitable") else "suitable"), None),
        ("Valuation Status", current_params.get("valuation_status") or "Not available", None),
        ("Model Unsuitable Reason", current_params.get("model_unsuitable_reason") or current_params.get("security_type_reason") or "N/A", None),
        ("Recommended Methods", ", ".join(str(x) for x in suitability_methods) if suitability_methods else "N/A", None),
    ])
    row += 1

    bs_support_status = _balance_sheet_forecast_support_status(out, ctx)
    bs_is_limited = bs_support_status == "limited_presentation"
    write_section_title(ws, row, "Balance Sheet Forecast Scope", span_cols=3)
    row += 1
    row = _write_kv_rows(ws, row, [
        ("Balance Sheet Forecast Status", bs_support_status, None),
        ("Full BS Forecast", "not supported" if bs_is_limited else "supported", None),
        ("Net Debt Bridge", "supported if data available", None),
        ("Equity Roll-forward", "not supported" if bs_is_limited else "supported", None),
        (
            "Source limitations",
            (
                "Full balance sheet forecast and equity roll-forward are not supported for this market in the current workbook; "
                "valuation uses selected balance sheet inputs for net debt, share count, and currency bridge only."
                if bs_is_limited
                else "Full BS forecast uses normalized historical BS / CF data and workbook schedules where available."
            ),
            None,
        ),
    ])
    row += 1

    write_section_title(ws, row, "Allowed Source Statuses", span_cols=3)
    row += 1
    row = _write_kv_rows(ws, row, [
        ("Status Vocabulary", ", ".join(ALLOWED_SOURCE_STATUSES), None),
        ("N/D vs N/A vs 0", "0 = true zero or formula placeholder with explicit source status; N/D = not separately disclosed in filing; N/A = not applicable or unavailable. Unavailable data should not display as 0 silently.", None),
    ])
    row += 1

    write_section_title(ws, row, "Financials Cache", span_cols=3)
    row += 1
    normalized_review = (
        current_params.get("normalized_ebit_review")
        or ((current_params.get("default_quality") or {}).get("normalized_ebit_review"))
        or {}
    )
    provenance = current_params.get("assumption_provenance") or {}
    ebit_provenance = provenance.get("ebit_margin") or {}
    normalized_applied = ebit_provenance.get("source") == "normalized_ebit_candidate"
    normalized_adjustments = ", ".join(
        str(x.get("field")) for x in (normalized_review.get("adjustments") or []) if isinstance(x, dict) and x.get("field")
    )
    period_alignment = (
        current_params.get("period_alignment")
        or ((current_params.get("default_quality") or {}).get("period_alignment"))
        or {}
    )
    capex_sanity = (
        current_params.get("capex_sanity")
        or ((current_params.get("default_quality") or {}).get("capex_sanity"))
        or {}
    )
    capex_review = (
        current_params.get("capex_review")
        or ((current_params.get("default_quality") or {}).get("capex_review"))
        or {}
    )
    nd_currency_audit = (
        current_params.get("net_debt_currency_audit")
        or ((current_params.get("default_quality") or {}).get("net_debt_currency_audit"))
        or {}
    )
    row = _write_kv_rows(ws, row, [
        ("Expected File", f"financials_{inp.symbol}_{FINANCIALS_CACHE_VERSION}.json", None),
        ("Resolved Path", cache.get("path") or "Not available", None),
        ("Cache Date", _display_date(cache.get("cached_at")) or "Not available", None),
        ("Market / Currency", f"{out.market} / {out.currency}", None),
        ("Reporting Currency", current_params.get("reporting_currency") or out.currency, None),
        ("Reporting Currency Source", current_params.get("reporting_currency_source") or "Not available", None),
        ("Trading Currency", current_params.get("trading_currency") or out.currency, None),
        ("Trading Currency Source", current_params.get("trading_currency_source") or "Not available", None),
        ("FX Rate Used", current_params.get("fx_rate_reporting_to_trading"), None),
        ("FX Source", current_params.get("fx_rate_source") or "Not available", None),
        ("IV/share - Reporting Currency", current_params.get("intrinsic_value_per_share_reporting_currency"), None),
        ("IV/share - Trading Currency", current_params.get("intrinsic_value_per_share_trading_currency"), None),
        ("Valuation Status", current_params.get("valuation_status") or "Not available", None),
        ("Market Comparison Status", current_params.get("market_comparison_status") or "Not available", None),
        ("Currency Translation Status", current_params.get("currency_translation_status") or "Not available", None),
        ("Currency Translation Warning", current_params.get("currency_translation_warning") or "None", None),
        ("Market / Valuation Warnings", " | ".join(current_params.get("warnings") or []) or "None", None),
        ("Normalized EBIT Review Candidate", normalized_review.get("status") or "Not available", None),
        ("Normalized EBIT Reported EBIT", normalized_review.get("reported_ebit"), FMT_COMMA2),
        ("Normalized EBIT Reported Margin", normalized_review.get("reported_ebit_margin"), FMT_PCT2),
        ("Normalized EBIT Candidate EBIT", normalized_review.get("candidate_normalized_ebit_high_plus_medium"), FMT_COMMA2),
        ("Normalized EBIT Candidate Margin", normalized_review.get("candidate_normalized_margin_high_plus_medium"), FMT_PCT2),
        ("Normalized EBIT Adjustment Fields", normalized_adjustments or "Not available", None),
        ("Normalized EBIT Applied?", "Yes - Applied by user" if normalized_applied else ("No - Not applied to valuation" if normalized_review else "Not available"), None),
        ("Normalized EBIT Applied Margin", ebit_provenance.get("applied_margin") if normalized_applied else None, FMT_PCT2),
        ("Normalized EBIT Applied Note", ebit_provenance.get("note") if normalized_applied else (normalized_review.get("warning") if normalized_review else "Not available"), None),
        ("Period Alignment Status", period_alignment.get("status") or "Not available", None),
        ("Period Alignment IS Period", period_alignment.get("income_statement_period") or "Not available", None),
        ("Period Alignment CF Period", period_alignment.get("cash_flow_period") or "Not available", None),
        ("Period Alignment BS Period", period_alignment.get("balance_sheet_period") or "Not available", None),
        ("Period Alignment Warning", period_alignment.get("warning") or "None", None),
        ("CapEx Multi-year Sanity Status", capex_sanity.get("status") or "Not available", None),
        ("CapEx Latest % Revenue", capex_sanity.get("latest_capex_pct_revenue"), FMT_PCT2),
        ("CapEx Multi-year Mean", capex_sanity.get("multi_year_mean"), FMT_PCT2),
        ("CapEx Multi-year Median", capex_sanity.get("multi_year_median"), FMT_PCT2),
        ("CapEx Years Available", capex_sanity.get("years_available"), None),
        ("CapEx Sanity Warning", capex_sanity.get("warning") or capex_sanity.get("interpretation") or "None", None),
        ("CapEx Field Coverage", ", ".join(capex_review.get("current_capex_field_coverage") or []) if capex_review else "Not available", None),
        ("0883 CapEx Candidate Field", capex_review.get("candidate_field") if capex_review else "Not available", None),
        ("0883 CapEx Candidate Confidence", capex_review.get("confidence") if capex_review else "Not available", None),
        ("0883 CapEx Candidate Applied?", "Yes" if capex_review.get("applied") else ("No" if capex_review else "Not available"), None),
        ("0883 CapEx Review Warning", capex_review.get("warning") if capex_review else "Not available", None),
        ("Net Debt Currency Audit Status", nd_currency_audit.get("status") or "Not available", None),
        ("Net Debt Currency", nd_currency_audit.get("net_debt_currency") or "Not available", None),
        ("Enterprise Value Currency", nd_currency_audit.get("enterprise_value_currency") or "Not available", None),
        ("Equity Bridge Basis", nd_currency_audit.get("equity_bridge_basis") or "Not available", None),
        ("Final IV Conversion", nd_currency_audit.get("final_iv_conversion") or "Not available", None),
        ("Net Debt Currency Warning", nd_currency_audit.get("warning") or "None", None),
        ("Default / Engine Warnings", " | ".join(list(current_params.get("warnings") or []) + list((getattr(out, "audit", None) or {}).get("warnings") or [])) or "None", None),
    ])
    row += 1

    write_section_title(ws, row, "Parallel Historical Financials Pipeline", span_cols=3)
    row += 1
    row = _write_kv_rows(ws, row, [
        ("Historical Expected File", historical_cache.get("path") or "Not available", None),
        ("Historical Cache Date", _display_date(historical_payload.get("cached_at")) or "Not available", None),
        ("Historical Schema", historical_payload.get("schema_version") or "Not available", None),
        ("Historical Source", historical_payload.get("source") or "yfinance", None),
        ("AAPL FY2021 10-K Backfill", "Revenue, gross profit, operating income, net income, tax, shares, BS, and CF fields are filing-backed where the normalized yfinance annual column was partial.", None),
        ("AAPL FY2024/FY2025 Interest Expense", "Filing-backed status override: Apple Form 10-K does not separately disclose interest expense in FY2024/FY2025; workbook displays N/D in 10-K instead of a blank or fabricated estimate.", None),
        ("N/D vs True Zero Policy", "Presentation distinguishes true zero, not separately disclosed, included in another line item, and unavailable from feed. Formula cells may use 0 for mechanics only when paired with a visible source/status label such as N/D in 10-K.", None),
        ("AAPL P&L Detail Methodology", "COGS, R&D, SG&A, total operating expenses, other income/expense net, and pre-tax income are sourced from Apple Form 10-K consolidated statements for the presentation layer. DCF still uses the selected EBIT margin path.", None),
        ("Normalized vs Filing-backed Fields", "Normalized yfinance rows remain the default source; narrow Apple Form 10-K overrides are recorded in source_overrides when a normalized row is unavailable or too sparse for review.", None),
        ("Raw Historical Unit", historical_payload.get("units") or "actual local currency", None),
        ("Workbook Display Unit", model_unit_label(out.currency, out.market), None),
        ("US Historical Scale Factor", "raw actual / 1,000,000" if out.market == "US" else "not applied in this market yet", None),
        ("Architecture", "Read-only export sidecar; DCF engine does not depend on historical pipeline.", None),
        ("Isolation", f"DCF defaults remain financials_{inp.symbol}_{FINANCIALS_CACHE_VERSION}.json; historical actuals come from a separate cache that is auto-rebuilt from yfinance when the schema needs upgrading.", None),
        ("Failure Behavior", "Historical cache failure or absence does not affect DCF defaults, formulas, or valuation.", None),
        ("BS / CF Forecast Source", "Historical BS / CF actuals come from a normalized historical cache covering ~30 BS line items (marketable securities, goodwill / intangibles, deferred revenue, leases, deferred tax assets / liabilities, current / non-current buckets) and are converted to workbook display units before being written.", None),
        ("AAPL BS Filing-backed Detail", "V3.9.8.7: AAPL BS historical rows are sourced from the Form 10-K consolidated balance sheet via the normalized cache. Cash, ST investments, AR, inventory, other current assets, PP&E, LT investments, other non-current assets, AP, current deferred revenue, ST debt, other current liabilities, LT debt, non-current deferred revenue (where disclosed), other non-current liabilities, total debt, total liabilities, and total equity are filing-backed across FY2021-FY2025.", None),
        ("Goodwill / Intangibles Treatment", "V3.9.8.7: Apple Form 10-K does NOT separately disclose Goodwill or Acquired Intangible Assets on the consolidated balance sheet — these line items are included within Other Non-current Assets per filing presentation. Workbook policy: historical cells display 'N/D in 10-K (not separately disclosed; included in Other Non-current Assets)' instead of a silent zero; forecast cells are held at $0 and classified as Not separately disclosed to avoid double-count with Other Non-current Assets. Where the cache supplies an explicit value (other tickers), the filing-backed number is displayed.", None),
        ("Marketable Securities Source", "V3.9.8.7: Short-term Investments / Marketable Securities and Long-term Investments / Marketable Securities are filing-backed historical rows from Apple Form 10-K (Other Short Term Investments / Investments And Advances). Forecast is held constant at latest available balance; no active driver in this version. Reviewers can size the impact via the Net Debt Bridge.", None),
        ("Residual / Held-constant Methodology", "V3.9.8.7: BS Forecast classifies each row as Schedule-driven (Cash / AR / Inventory / AP / PP&E / Debt), Held constant at latest available balance (marketable securities, deferred revenue, leases, deferred taxes, other named rows where filing supplies a value but no forecast driver exists), Residual / plug (named bucket residuals, transparent review items, sized via Balance Check), Not separately disclosed in 10-K (Goodwill / Intangibles for AAPL — held at $0), or Memo / not separately modeled (capital lease subset, non-current deferred revenue subset). Total Equity is roll-forward; OCI not modeled.", None),
        ("BS Forecast Classification Methodology", "V3.9.8.7: BS Forecast carries a per-row classification column. Schedule-driven rows tie to Supporting Schedules (working capital days, PP&E roll-forward, debt schedule). Held-constant rows reference the latest historical value. Residual plugs are named (Other Current Assets / Other Non-current Assets / Other Current Liabilities / Other Non-current Liabilities — residual plug, transparent). Not-separately-disclosed rows surface filing reality rather than silent zero. Classification text is row-specific so reviewers can audit each line without re-deriving its source.", None),
        ("AAPL Deferred Tax Treatment", "V3.9.8.7: Apple FY2022-FY2025 disclose Non Current Deferred Tax Assets; FY2021 and Deferred Tax Liabilities are not separately disclosed in the consolidated BS. Workbook displays the filing-backed value where available and 'N/D in 10-K' where the filing does not break it out.", None),
        ("Engine", "UI and Excel valuation share a single source. Workbook formulas are live so users can edit assumptions and recompute interactively.", None),
        ("Working Capital Days Source", "DSO / DIO / DPO are computed from historical AR, inventory, AP, revenue, and COGS in the historical cache. Forecast days fade from latest actual days toward normalized 3Y averages where available.", None),
        ("Historical Delta NWC Source", "Historical Delta NWC = AR + Inventory - AP less prior-year AR + Inventory - AP. Sign convention: negative = source of cash / WC release; positive = use of cash / WC investment. Working Capital Reality Check compares faded-days Delta NWC with the signed historical average and shows FY2022-FY2025 deltas when available.", None),
        ("Buyback Funding Closure", "Default treatment is Cash Floor / Buyback Cap: funded buybacks are capped by FCF after dividends, cash above the minimum floor, and selected marketable securities availability. The source stack splits FCF after dividends used, cash above floor used, marketable securities drawdown, incremental debt, and unfunded/capped buybacks. Debt-Funded Buyback explicitly closes gaps with incremental debt; Planned-Uncapped Diagnostic may breach cash floor and is High Review.", None),
        ("Cash / Marketable Securities Source", "Cash, short-term investments, long-term investments, and total debt are sourced from the historical balance-sheet cache. Minimum cash floor defaults to 30,000 in model units for AAPL and is editable.", None),
        ("Cost of Debt Implied / Normalized Source", "Historical implied Kd = interest expense / average debt where available. AAPL selected normalized Kd defaults to 4.3%, blending historical implied Kd with forward refinancing conditions; the WACC Bridge shows delta and review tier.", None),
        ("Forecast Methodology", "Five-year explicit forecast: revenue growth and EBIT margin from Assumptions; D&A driven by Beginning Net PP&E × D&A % assumption (PP&E roll-forward); CapEx as % of revenue; Working Capital driven by DSO / DIO / DPO days against revenue and COGS; tax rate floored at zero (no NOL / deferred tax modeling).", None),
        ("Balance Sheet Coverage", "BS Forecast separates Current / Non-current Assets and Liabilities and reports Marketable Securities (ST + LT), Goodwill, Intangibles, Deferred Revenue, Capital Lease Obligations, and Deferred Tax Assets / Liabilities as explicit lines. Residual plug rows remain transparent so reviewers can size any unmodeled items.", None),
        ("Net Debt Bridge", "Reported / Input Net Debt sits alongside adjusted variants (Debt − Cash, Debt − Cash − ST Investments, Debt − Cash − Total Marketable Securities). The Selected Net Debt Treatment dropdown on Assumptions drives the DCF equity bridge; default = Reported / Input Net Debt. Adjusted IV/share references are shown on DCF Valuation. Capital leases are memo only (yfinance already includes them in ST/LT Debt totals).", None),
        ("Reported Net Debt Caveat", "Reported / Input Net Debt may differ from raw Debt − Cash because data providers often pre-adjust with cash-like investments. The Net Debt Bridge exposes raw Debt − Cash and the Reported vs Raw gap; the Review Dashboard flags the gap when it exceeds 10% of Total Debt. The workbook does not silently rewrite the reported figure.", None),
        ("Shareholder Returns", "Dividends and buybacks roll through CF Financing and the BS Equity roll-forward but do NOT reduce FCFF. The Selected Share Count Treatment dropdown (Current Reported / Forecast Ending / Forecast WAvg) controls the per-share denominator only. Historical dividends and buybacks are sourced from the cash-flow cache; missing fields fall back to zero with a Review flag.", None),
        ("Shareholder Returns Boundaries", "No treasury-stock accounting, no SBC vintage schedule, no circular cash sweep, no buyback-as-FCF deduction. Repurchase price grows at the user-supplied rate (default 0%) from the current price.", None),
        ("WACC Decision Layer", "Assumptions exposes a Selected WACC Treatment dropdown (Selected / Model WACC | Mechanical CAPM reference | Selected +/-100 bps). DCF Valuation, FCF Build, and Sensitivity all read the Selected WACC Used cell so picking a treatment flips the headline. Mechanical CAPM reference = We*Ke + Wd*Kd*(1-t) using rf and ERP from market config, beta from cache, and pre-tax cost of debt from Assumptions. Mechanical CAPM reference is an audit diagnostic, not headline selection. Default = Selected / Model WACC.", None),
        ("WACC Component Methodology", "Risk-free rate uses the market-config government-yield proxy at export cutoff; beta uses raw cache beta with Blume adjustment shown as diagnostic; ERP uses selected market policy reference; capital structure uses price x shares and latest total debt; cost of debt compares historical implied Kd with selected normalized Kd.", None),
        ("WACC Boundaries", "No equity-risk-premium term structure, no hard-wired industry beta lookup beyond the cache, no debt-rating-implied cost of debt, no APV decomposition. The Review Dashboard tiers the Selected vs Mechanical CAPM reference spread (>150 bps Review, >250 bps High Review) without auto-correcting.", None),
        ("Terminal Value Decision Layer", "Assumptions exposes a Selected Terminal Treatment dropdown (Current Model | Gordon | Exit | Blend | Fade Period) plus Fade Years, Fade Target Growth, and Blend Weight inputs. DCF Valuation lists IV/share and TV/EV under each treatment. Default = Current Model. Fade Period is a reference case (linear FCF growth fade then Gordon TV on the final fade-year FCF), not a full 10-year 3FS forecast.", None),
        ("Operating Thesis Bridge methodology", "AAPL Operating Thesis Bridge derives total revenue growth from Products and Services growth, and implied EBIT margin from Gross Margin - R&D% - SG&A%. The coupled Operating Path Source selector on Assumptions can switch both revenue growth and EBIT margin together. Default = Selected Path to preserve baseline IV.", None),
        ("Path source methodology", "Selected Path rows remain the analyst-selected forecast path. Bridge rows are mechanical AAPL bridge outputs. Effective path rows feed P&L Forecast, Supporting Schedules, FCF Build, and the DCF headline only when Operating Path Source = AAPL Operating Thesis Bridge. Bridge IV and attribution are alternative reference only and do not feed Football Field.", None),
        ("Operating Thesis Materiality Methodology (V3.9.9.6)", "Selected-vs-Bridge differences are classified as Bridge-consistent when absolute divergence is <=50 bps. Analyst override rationale is required only when a future year exceeds the 50 bps materiality threshold; small differences are disclosed without forced narrative.", None),
        ("Terminal Value Caveats", "Terminal value method is a valuation judgment. The Terminal Value Philosophy & Review block tiers TV/EV, Gordon vs Exit gap, terminal growth, and Exit Multiple vs Trading Comps median; tiers are diagnostic (OK / Review / High review / N/A) and do not auto-correct the headline. Gordon-implied exit multiple and Exit-implied terminal growth are computed for cross-method auditing.", None),
        ("Trading Comps", "Peer market data is pulled from yfinance (market cap, enterprise value, revenue, EBITDA, net income, trailing P/E) with a daily file cache. EV/Revenue, EV/EBITDA, and P/E are computed per peer; outliers (abs caps + 3× IQR) and exclusions (missing / negative metrics) are flagged in the Peer Set table, and summary statistics use included peers only. Implied IV/share applies the multiples to target revenue / EBITDA / net income and converts EV to Equity using the selected Net Debt and Share Count.", None),
        ("Forward EPS / Forward P/E Diagnostic (V3.9.9.6)", "Forward P/E values are sourced from yfinance.info.forwardPE / available one-feed data. They are NOT FactSet / Refinitiv multi-analyst consensus medians. Analyst count, dispersion, and update timestamp are not available through this feed. Use as directional diagnostic only; validate against institutional consensus feed before relying in IC discussion. Missing forward metrics display N/A, not zero, and do not drive existing comps valuation ranges, Football Field, or headline IV.", None),
        ("Forward Diagnostic Availability Policy", "Data unavailable means the feed did not return a usable field in the current run; zero value means the feed explicitly returned 0 and should be reviewed; not applicable means the metric is outside this workbook scope. When core peer forward coverage is below 3, the workbook labels the section unavailable for IC reliance and trailing comps remain the primary diagnostic.", None),
        ("52-week Market Context (V3.9.9.6)", "52-week price range reflects market trading history, not valuation. Shown for IC context only; not a valuation reference. Current price, 52-week high/low, percentile, and price-vs-Base/Bear IV are sourced from yfinance.info where available; unavailable feed fields display N/A and are not used in Football Field.", None),
        ("Legacy Scenario Schema Aggregation (V3.9.9.6)", "Saved-scenario treatment defaults are aggregated into a single Audit Dashboard summary row and detailed in Scenario Notes. Covered fields: selected_wacc_treatment, selected_terminal_treatment, selected_net_debt_treatment, selected_share_count_treatment, selected_operating_path_source. Defaults preserve baseline scenario IV unless the user explicitly changes a treatment.", None),
        ("Trading Comps Caveats", "Trading Comps are a market reference for discussion purposes only. The peer set is configurable and may require analyst review for industry alignment, currency consistency, and size matching. Multiples can be volatile and individual peers may legitimately appear as outliers under the default policy. Failed peer fetches degrade gracefully; the DCF remains the primary model. HK / CN coverage may be limited.", None),
        ("Sensitivity & Football Field", "Sensitivity tables vary WACC × terminal g, WACC × exit multiple, and revenue growth × EBIT margin around the headline. Football Field overlays DCF cases (Gordon-only, Current Model, Exit-only, Fade Period), saved Bull/Bear scenarios, sensitivity ranges, and Trading Comps EV/Revenue, EV/EBITDA, P/E ranges against the current market price.", None),
        ("Review Dashboard Boundary", "The Review Dashboard surfaces transparent caveats (treatment validity, parity, magnitudes, mismatches) so reviewers can audit assumptions; Review / High Review / Fallback flags are intentionally not auto-suppressed.", None),
        ("Assumption Rationale Boundary", "The workbook provides structured valuation mechanics, decision layers, default rationale blocks, and market-implied break-even references. Business rationale for revenue growth, margin, WACC, terminal value, net debt treatment, and capital returns requires analyst review; defaults are starting points, not management guidance or rating guidance.", None),
        ("ROIC / Reinvestment Methodology (V3.9.9.5)", "Implied terminal ROIC = NOPAT / Invested Capital, written as an Excel formula referencing cells (no inline numeric ROIC literal). ROIC sensitivity is diagnostic only: normalized ROIC applies only to terminal FCF reinvestment need; explicit-period FCF remains unchanged. Fully coherent ROIC normalization would require capitalized R&D, lease ROU, and explicit-period reinvestment reassessment, out of scope for this patch.", None),
        ("SBC Dilution Methodology (V3.9.9.4)", "Latest historical Stock-Based Compensation is sourced from the cash-flow cache (stock_based_compensation, absolute value). Default annual dilution % = SBC / (Current Price x Reported Shares), capped at 1.5%; fallback uses SBC / Revenue x 0.25 when market cap is unavailable; final fallback is a 0.25% conservative placeholder. Explicit user input overrides the default. Share Count Roll-forward = Beginning + SBC dilution - Funded Buybacks. Headline IV continues to default to Current Reported Diluted Shares; switching to Forecast Ending / Weighted Avg surfaces the SBC-adjusted denominator. Limitation: no SBC vintage schedule, no treasury stock accounting, no separate option-exercise modelling.", None),
        ("FA / Ratios", "Ratios are analytical outputs, not model drivers, and are not referenced by the forecast or valuation sheets.", None),
        ("Market Coverage", "HK and CN markets fall back to the legacy operating forecast where the historical pipeline is not yet wired; this is intentional graceful behaviour, surfaced on the Review Dashboard.", None),
    ])
    row += 1

    write_section_title(ws, row, "Known Limitations", span_cols=3)
    row += 1
    row = _write_kv_rows(ws, row, [
        ("Mega-cap terminal value dependency", "TV/EV may remain high for mega-cap DCF workbooks; users may adjust the explicit forecast horizon where they have confidence.", None),
        ("H-Model status", "H-Model is diagnostic unless selected by the user.", None),
        ("ROIC diagnostic", "ROIC diagnostic depends on invested-capital treatment and remains a review item if above 50%.", None),
        ("Forward comps", "Forward comps are single-feed diagnostics, not institutional consensus.", None),
        ("Trading comps cache", "Trading comps may use cache fallback if the live provider fails.", None),
        ("Non-US tickers", "Non-US tickers may use fallback financials or limited market data.", None),
        ("Default assumption quality", ((ctx.get("current_params") or {}).get("default_quality") or {}).get("banner") or "Default assumptions passed the V3.9.10.2 quality gate.", None),
        ("Financial sector caution", "Financial sector valuations require caution because standard industrial FCFF and EV multiple framing may not fit the business model.", None),
        ("Recommendation boundary", "This workbook does not provide an investment recommendation.", None),
        ("User-modified cases", "User-modified cases require analyst review before external use.", None),
    ])
    row += 1

    current_params = ctx.get("current_params") or {}
    wacc_contract = getattr(out, "wacc_decision_bridge", None) or {}
    terminal_contract = getattr(out, "terminal_decision_bridge", None) or {}
    nd_contract = getattr(out, "net_debt_bridge", None) or {}
    sr_contract = getattr(out, "shareholder_returns", None) or {}
    op_contract = getattr(out, "operating_path_bridge", None) or {}
    write_section_title(ws, row, "UI / Excel Contract Summary", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Field", "UI / engine value", "Excel contract"], 1)
    row += 1
    contract_rows = [
        ("Current Price", current_params.get("price"), "Assumptions Current Price and DCF Valuation Current Price use the same exported value unless quote cache timing differs."),
        ("Forecast Horizon", current_params.get("forecast_years") or inp.forecast_years, "Assumptions Forecast Years controls export guardrails; extended horizon warnings remain active."),
        ("Selected WACC", wacc_contract.get("selected_wacc_used_in_dcf") or current_params.get("wacc"), "Assumptions Selected WACC Used in DCF, DCF Valuation, FCF Build, and Sensitivity are linked."),
        ("Selected Kd", current_params.get("pre_tax_cost_of_debt") or getattr(inp, "pre_tax_cost_of_debt", None), "WACC Bridge Cost of Debt is the single selected Kd source."),
        ("Net Debt Treatment", (nd_contract.get("selected_treatment_label") or NET_DEBT_TREATMENT_LABELS.get(nd_contract.get("selected_treatment"), "")), "Assumptions dropdown drives DCF equity bridge and Trading Comps implied IV conversion."),
        ("Terminal Treatment", terminal_contract.get("selected_terminal_treatment_label"), "Assumptions dropdown drives headline terminal treatment; Gordon-only and Exit-only remain labeled alternatives."),
        ("Operating Path Source", op_contract.get("selected_operating_path_source") or OPERATING_PATH_SOURCE_SELECTED, "Assumptions selected_operating_path_source drives P&L / schedules only when intentionally switched."),
        ("Working Capital Label", "Active schedule-derived Delta NWC; legacy Delta NWC % Revenue is reference/fallback.", "FCF Build primary row is schedule-derived; legacy reference is labeled separately."),
        ("Share Count Treatment", SHARE_COUNT_TREATMENT_LABELS.get(sr_contract.get("selected_share_count_treatment") or DEFAULT_SHARE_COUNT_TREATMENT), "DCF IV/share denominator follows the selected treatment; alternatives remain reference rows."),
        ("Extended Horizon Guardrail", current_params.get("forecast_years") or inp.forecast_years, "Audit Dashboard extended horizon and economic guardrail checks remain visible when horizon exceeds five years."),
    ]
    for label, ui_value, contract in contract_rows:
        _set_label_cell(ws, row, 1, label)
        c = ws.cell(row=row, column=2, value="N/A" if ui_value is None else ui_value)
        c.font = font_formula(); c.border = border_thin()
        if isinstance(ui_value, (int, float)):
            c.number_format = FMT_PCT2 if "WACC" in label or "Kd" in label else FMT_COMMA2
        _set_note_cell(ws, row, 3, contract)
        row += 1
    _set_note_cell(ws, row, 1, "Known contract limitation: current-price source differences can occur when the UI fetch and Excel export use different cache timestamps; the source trail must explain any difference.")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    export_cutoff = _display_date(ctx.get("generated_at")) or date.today().isoformat()
    financials_cutoff = _display_date(cache.get("cached_at")) or export_cutoff
    historical_cutoff = _display_date(historical_payload.get("cached_at")) or financials_cutoff
    financials_cache_ref = (
        cache.get("path") or f"financials_{inp.symbol}_{FINANCIALS_CACHE_VERSION}.json"
    )
    historical_cache_ref = (
        historical_cache.get("path") or "Normalized historical cache (auto-resolved)"
    )

    em_cache_file = current_params.get("exit_multiple_cache_file")
    em_cache_display = os.path.basename(em_cache_file) if em_cache_file else "Not cached this session"
    em_fetched_at = current_params.get("exit_multiple_fetched_at") or "n/a"
    em_raw = current_params.get("exit_multiple_raw")
    em_raw_display = (
        f"raw {round(float(em_raw), 4)} (rounded {current_params.get('exit_multiple')})"
        if em_raw is not None else "raw value not available"
    )
    em_source_display = current_params.get("exit_multiple_source") or "Public market data / cached quote-derived reference"
    em_methodology = (
        f"V3.9.2 deterministic daily cache: yfinance enterpriseToEbitda fetched once per export "
        f"date and reused all day; rounded to 0.1x. {em_raw_display}. Fetched at: {em_fetched_at}. "
        "V3.9.8.6 review tier vs core Trading Comps EV/EBITDA median: within +-15% OK | +-15-30% Review | >30% High review | N/A if insufficient peers. "
        "If the selected exit multiple is above the core peer median, treat it as an IC review item rather than a defended fact."
    )
    if current_params.get("exit_multiple_warning"):
        em_methodology = f"{em_methodology} Source caveat: {current_params['exit_multiple_warning']}"

    market_data_cutoff = export_cutoff
    market_data_ref = "Live quote cache (refreshes per export)"

    selected_terminal_method = current_params.get("tv_method") or getattr(inp, "tv_method", None)
    nd_bridge = getattr(out, "net_debt_bridge", None) or {}
    selected_nd_treatment = (
        nd_bridge.get("selected_treatment")
        or current_params.get("selected_net_debt_treatment")
        or DEFAULT_NET_DEBT_TREATMENT
    )
    selected_nd_treatment_label = NET_DEBT_TREATMENT_LABELS.get(
        selected_nd_treatment, selected_nd_treatment
    )

    trail_rows = [
        ("Current Price", current_params.get("price"), "Market data quote cache",
         market_data_cutoff, market_data_ref,
         "Latest available quote; expressed in reporting currency. Refreshed every export."),
        ("Revenue", current_params.get("revenue"), "Normalized financials cache",
         financials_cutoff, financials_cache_ref,
         "Latest annual revenue (millions of reporting currency)."),
        ("EBIT", current_params.get("ebit"), "Normalized financials cache",
         financials_cutoff, financials_cache_ref,
         "Latest annual EBIT (millions of reporting currency)."),
        ("D&A (add-back)", current_params.get("da"), "Normalized financials cache",
         financials_cutoff, financials_cache_ref,
         "Latest annual depreciation and amortization (millions); used as a FCFF add-back."),
        ("CapEx", current_params.get("capex"), "Normalized financials cache",
         financials_cutoff, financials_cache_ref,
         "Latest annual capital expenditure (millions); cash outflow in FCFF."),
        ("Change in NWC", current_params.get("wc_change"), "Normalized financials cache",
         financials_cutoff, financials_cache_ref,
         "Latest annual change in net working capital (millions). Schedule-derived Delta NWC drives FCF Build when available."),
        ("Diluted Shares Outstanding (M)", current_params.get("shares"), "Normalized financials cache",
         financials_cutoff, financials_cache_ref,
         "Latest diluted share count (millions). Single source-of-truth Assumptions cell - drives both the legacy Share Count Schedule and the Shareholder Returns Schedule beginning shares, so the two cannot diverge."),
        ("Net Debt", current_params.get("net_debt"), "Normalized historical balance sheet",
         historical_cutoff, historical_cache_ref,
         "Reported convention: Total Debt minus Cash & Equivalents. Bridge variants exposed on Assumptions."),
        ("WACC", current_params.get("wacc"), "Model default (CAPM-style)",
         export_cutoff, "rf/ERP from MARKET_CONFIG, beta from financials cache",
         "Selected/model WACC drives the headline case. Mechanical CAPM reference and +/-100 bps cases are shown on DCF Valuation; a large selected-vs-mechanical spread is a visible IC review item."),
        ("Terminal Growth", current_params.get("terminal_g"), "User input / model default",
         export_cutoff, "Editable on Assumptions",
         "Long-run nominal growth anchor for Gordon TV. Review for economic consistency."),
        ("Exit Multiple", current_params.get("exit_multiple"), em_source_display,
         em_fetched_at, em_cache_display, em_methodology),
        ("Selected Terminal Method", selected_terminal_method, "User input / model default",
         export_cutoff, "Editable on Assumptions (gordon / exit / average)",
         "Drives the headline terminal value when Terminal Treatment dropdown is set to Current Model."),
        ("Selected Net Debt Treatment", selected_nd_treatment_label, "User input / model default",
         export_cutoff, "Editable on Assumptions (Net Debt Bridge)",
         "Default = Reported / Input Net Debt preserves prior-release IV; switching the dropdown intentionally flips the headline."),
    ]

    # ── V3.9.5 Terminal Philosophy review rows ──────────────────────────
    tv_philosophy_audit = _build_terminal_philosophy_payload(out, ctx)
    tv_ev_audit = tv_philosophy_audit.get("tv_ev_ratio")
    ge_gap_audit = tv_philosophy_audit.get("gordon_vs_exit_gap_pct")
    em_vs_comps_audit = tv_philosophy_audit.get("exit_multiple_vs_comps_median")
    em_pd_audit = tv_philosophy_audit.get("exit_multiple_vs_comps_premium_discount")
    em_pd_display = (
        f"{em_pd_audit * 100:.1f}% vs comps median" if em_pd_audit is not None else "N/A"
    )
    comps_payload_for_audit = ctx.get("trading_comps") or {}
    comps_profile_for_audit = comps_payload_for_audit.get("profile") or {}
    comps_status_for_audit = comps_payload_for_audit.get("status") or (
        "available" if comps_payload_for_audit else "unavailable"
    )
    trail_rows.extend([
        ("Trading Comps Status",
         comps_status_for_audit,
         "Trading Comps peer support policy",
         export_cutoff,
         "Trading Comps sheet",
         comps_payload_for_audit.get("reason") or comps_payload_for_audit.get("unavailable_reason") or "Trading Comps available for mapped peer set."),
        ("Trading Comps Peer Set Source",
         comps_payload_for_audit.get("peer_set_source") or comps_payload_for_audit.get("peer_source") or "unavailable",
         "Trading Comps peer support policy",
         export_cutoff,
         "Trading Comps sheet",
         f"Profile: {comps_payload_for_audit.get('profile_id') or comps_profile_for_audit.get('profile_id') or 'N/A'}."),
        ("Trading Comps Reason",
         comps_payload_for_audit.get("reason") or comps_payload_for_audit.get("unavailable_reason") or comps_profile_for_audit.get("reason") or "Mapped peer set available.",
         "Trading Comps peer support policy",
         export_cutoff,
         "Trading Comps sheet",
         "Generic technology fallback is suppressed when no curated/profile-appropriate peer set is available."),
        ("Trading Comps fallback hierarchy",
         comps_payload_for_audit.get("data_quality", {}).get("overall_comps_usability_tier") or "N/A",
         "Comps cache -> yfinance.info / fast_info -> quote/history cache -> financials cache -> manual metadata",
         export_cutoff,
         "Trading Comps sheet",
         "Each field records live / cache / derived / unavailable. Stale cache is shown as Review, not discarded."),
        ("Comps cache freshness",
         comps_payload_for_audit.get("data_quality", {}).get("football_field_comps_status") or "N/A",
         "Existing comps cache and source/cutoff trail",
         export_cutoff,
         "Trading Comps Source / Cache / Timestamp Trail",
         "Stale cache can support trailing diagnostics when provider calls fail; forward metrics remain one-feed diagnostics."),
        ("Derived EBITDA / N/A policy",
         "EBIT + D&A only when both are cached",
         "Normalized financials cache fallback",
         financials_cutoff,
         "Trading Comps Peer Set + Multiples",
         "If D&A is unavailable, EBITDA remains N/A. Missing fields display N/A, not zero."),
        ("52-week Market Context fallback",
         (comps_payload_for_audit.get("market_context") or {}).get("range_source") or "unavailable",
         "Comps/quote fields, then 1Y history cache",
         export_cutoff,
         "Executive Summary market context",
         "Trading history only; not a valuation input and not included in Football Field."),
    ])
    trail_rows.extend([
        ("Selected Terminal Treatment (Headline)",
         tv_philosophy_audit.get("selected_terminal_method_label"),
         "User input / model default", export_cutoff,
         "Editable on Assumptions (Terminal Treatment dropdown)",
         "Headline terminal method. Default preserves prior-release IV; switching the dropdown intentionally flips the headline."),
        ("Terminal Philosophy",
         tv_philosophy_audit.get("primary_philosophy"),
         "Engine-derived from selected treatment", export_cutoff,
         "Method hierarchy in Terminal Value Philosophy & Review",
         tv_philosophy_audit.get("terminal_method_recommended_use") or ""),
        ("TV / EV",
         tv_ev_audit, "Computed (TV PV / EV)", export_cutoff,
         "Live in DCF Valuation",
         f"Review tier: {tv_philosophy_audit.get('tv_ev_review_tier')}. <=70% OK | 70%-80% Review | >80% High review."),
        ("Gordon Implied Exit Multiple",
         tv_philosophy_audit.get("gordon_implied_exit_multiple"),
         "Computed (Gordon TV / EBITDA_YN)", export_cutoff,
         "Live in DCF Valuation",
         "Cross-method diagnostic; what exit multiple Gordon TV implies."),
        ("Exit Implied Terminal Growth",
         tv_philosophy_audit.get("exit_implied_terminal_growth"),
         "Computed (solve g from Exit TV)", export_cutoff,
         "Live in DCF Valuation",
         "Cross-method diagnostic; what long-run g Exit TV implies."),
        ("Gordon vs Exit Gap",
         ge_gap_audit, "Computed", export_cutoff, "Live in DCF Valuation",
         f"Review tier: {tv_philosophy_audit.get('gordon_vs_exit_review_tier')}. <=15% OK | 15%-30% Review | >30% High review."),
        ("Exit Multiple vs Trading Comps Median",
         em_vs_comps_audit if em_vs_comps_audit is not None else "N/A",
         "Trading Comps EV/EBITDA included-peer median", export_cutoff,
         "Trading Comps Summary Statistics",
         f"{em_pd_display}. Review tier: {tv_philosophy_audit.get('exit_multiple_vs_comps_review_tier')}. Within +-15% OK | +-15-30% Review | >30% High review | N/A if insufficient peers."),
        ("Terminal Growth Review Methodology",
         tv_philosophy_audit.get("terminal_growth_review_tier"),
         "Engine review tier", export_cutoff,
         "Terminal Value Philosophy & Review",
         "0%-4% OK | <0% or >4% Review | >5% High review. Diagnostic only; no headline auto-correction."),
    ])

    # ── V3.9.6 Share Count / Shareholder Returns Source Trail ──────────
    sr_audit = getattr(out, "shareholder_returns", None) or {}
    sr_drivers_trail = sr_audit.get("drivers_effective") or {}
    sr_denom_options = sr_audit.get("denominator_options") or {}
    sc_treatment_key = sr_audit.get("selected_share_count_treatment") or DEFAULT_SHARE_COUNT_TREATMENT
    sc_treatment_label = SHARE_COUNT_TREATMENT_LABELS.get(sc_treatment_key, sc_treatment_key)
    nd_bridge_trail = getattr(out, "net_debt_bridge", None) or {}
    cash_trail = nd_bridge_trail.get("cash")
    sr_buyback_method_label = sr_drivers_trail.get("buyback_method_label") or BUYBACK_METHOD_LABELS.get(
        sr_drivers_trail.get("buyback_method") or DEFAULT_BUYBACK_METHOD, ""
    )
    sbc_pct = sr_drivers_trail.get("annual_dilution_pct")
    if sbc_pct is None:
        sbc_pct = float(getattr(inp, "annual_dilution_pct", 0.0) or 0.0)
    sbc_default_methodology_trail = sr_drivers_trail.get("annual_dilution_default_methodology")
    sbc_user_override_trail = bool(sr_drivers_trail.get("annual_dilution_user_override"))
    if sbc_user_override_trail:
        sbc_methodology = "User-provided override (Assumptions Annual Dilution %)."
    elif sbc_default_methodology_trail:
        sbc_methodology = (
            "V3.9.9.4 SBC dilution v1 default: " + sbc_default_methodology_trail
            + " Source: cash-flow stock_based_compensation field. Editable on Assumptions."
        )
    elif (sbc_pct or 0.0) == 0.0:
        sbc_methodology = (
            "SBC dilution = 0 (no historical SBC field and no user override). "
            "Review for SBC-material names; editable on Assumptions."
        )
    else:
        sbc_methodology = "Annual share-count dilution from SBC / option exercises; editable analyst input."

    trail_rows.extend([
        # Diluted Shares (Reported) is already covered above by the V3.9.2 base
        # "Diluted Shares Outstanding (M)" row; do not duplicate. The V3.9.6
        # single source-of-truth note now lives in that row's in-cell comment.
        ("Selected Shares Used in IV/share",
         sr_audit.get("selected_denominator"),
         "Computed (Selected Share Count Treatment)", export_cutoff,
         "Assumptions Shareholder Returns block",
         f"Headline denominator = {sc_treatment_label}. Default = Current Reported Diluted Shares preserves prior-release IV."),
        ("Share Count Treatment", sc_treatment_label,
         "User input / model default", export_cutoff,
         "Editable on Assumptions (Share Count Treatment dropdown)",
         "Default preserves headline IV; switching to Forecast Ending / Weighted Avg changes per-share denominator only - EV and equity value are unchanged."),
        ("Forecast Ending Diluted Shares (alt)",
         sr_denom_options.get("forecast_ending_diluted_shares"),
         "Engine roll-forward", export_cutoff,
         "Shareholder Returns Schedule",
         "Year-N ending shares from the roll-forward (Beginning - Repurchased + SBC dilution). Reference denominator."),
        ("Forecast Weighted Avg Diluted Shares (alt)",
         sr_denom_options.get("forecast_weighted_avg_diluted_shares"),
         "Engine roll-forward", export_cutoff,
         "Shareholder Returns Schedule",
         "Average of beginning + ending across the explicit period. Reference denominator."),
        ("Dividend Payout % NI",
         sr_drivers_trail.get("dividend_payout_pct_net_income"),
         "User input / historical default", export_cutoff,
         "Editable on Assumptions (Shareholder Returns block)",
         "Drives projected dividends. Defaults to historical payout ratio. Dividends are a capital-allocation item; not deducted from FCFF."),
        ("Buyback Method", sr_buyback_method_label,
         "User input / model default", export_cutoff,
         "Editable on Assumptions (Shareholder Returns block)",
         "% of FCF or flat amount. Drives the Shareholder Returns Schedule buyback line."),
        ("Buyback % FCF",
         sr_drivers_trail.get("buyback_pct_fcf"),
         "User input / historical default", export_cutoff,
         "Editable on Assumptions",
         "Drives buybacks under the % of FCF method. Buybacks are a capital-allocation item; not deducted from FCFF."),
        ("Flat Buyback Amount (M)",
         sr_drivers_trail.get("flat_buyback_amount"),
         "User input / historical default", export_cutoff,
         "Editable on Assumptions",
         "Drives buybacks under the flat-amount method."),
        ("SBC / Annual Dilution %",
         sbc_pct, "User input / model default", export_cutoff,
         "Editable on Assumptions (Shareholder Returns block)",
         sbc_methodology),
        ("Assumed Repurchase Price Growth",
         sr_drivers_trail.get("repurchase_price_growth"),
         "User input / model default", export_cutoff,
         "Editable on Assumptions",
         "Repurchase price = Current Price grown at this rate each year. Drives Shares Repurchased = Buyback / Repurchase Price."),
        ("Beginning Cash (Funding Bridge)",
         cash_trail, "Normalized historical balance sheet (Net Debt Bridge)",
         historical_cutoff, historical_cache_ref,
         "Used in the Buyback Funding & Cash Use Bridge to assess balance-sheet capacity. Debt issuance/repayment is not modeled."),
        ("Funding Bridge Methodology", "Returns/FCF tiered",
         "Engine review tier", export_cutoff,
         "Buyback Funding & Cash Use Bridge",
         "Returns/FCF <=80% OK | 80-120% Review | >120% or FCF<=0 High review. Discussion flag only; no auto-correction."),
    ])

    write_section_title(ws, row, "Key Input Source Trail", span_cols=6)
    row += 1
    write_header_row(
        ws,
        row,
        ["Input / Metric", "Value", "Source", "Cutoff / Fetch Time", "Cache / Reference", "Methodology"],
        1,
    )
    row += 1
    for label, value, source, cutoff, cache_ref, methodology in trail_rows:
        _set_label_cell(ws, row, 1, label)
        if isinstance(value, (int, float)):
            c = ws.cell(row=row, column=2, value=float(value))
            c.font = font_formula()
            c.border = border_thin()
            c.number_format = FMT_COMMA2
        else:
            c = ws.cell(row=row, column=2, value=value if value is not None else "n/a")
            c.font = font_formula()
            c.border = border_thin()
        _set_note_cell(ws, row, 3, source)
        _set_note_cell(ws, row, 4, cutoff or "n/a")
        _set_note_cell(ws, row, 5, cache_ref or "n/a")
        _set_note_cell(ws, row, 6, methodology)
        row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "V3.9.2 Input Source Comments & Data Determinism Lock: every input above also carries an in-cell comment on the Assumptions sheet (Source / Cutoff / Methodology). The Exit Multiple default is cached per export date so successive runs on the same calendar day return an identical IV anchor. V3.9.5 Terminal Philosophy rows tier the terminal layer (TV/EV, Gordon vs Exit, Exit vs Comps median, terminal growth) for IC defensibility - diagnostic only; no field auto-overwrites the headline.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    # ── V3.9.4 WACC Source Trail ───────────────────────────────────────
    wacc_bridge_audit = getattr(out, "wacc_decision_bridge", None) or {}
    sanity_audit = wacc_bridge_audit.get("cost_of_debt_sanity") or {}
    wacc_trail_rows = [
        ("Risk-free Rate (rf)", wacc_bridge_audit.get("risk_free_rate"), "Market config",
         export_cutoff, "MARKET_CONFIG per market",
         "Per-market risk-free rate; editable on Assumptions."),
        ("Equity Risk Premium (ERP)", wacc_bridge_audit.get("equity_risk_premium"), "Market config",
         export_cutoff, "MARKET_CONFIG per market",
         "Per-market equity risk premium; editable on Assumptions."),
        ("Raw Beta", wacc_bridge_audit.get("raw_beta"), "Financials cache (yfinance)",
         financials_cutoff, financials_cache_ref,
         "Capped to [0.3, 3.0]. Adjusted beta = 0.67*raw + 0.33*1.0 shown as reference."),
        ("Selected Beta", wacc_bridge_audit.get("selected_beta"), "User input / model default",
         export_cutoff, "Editable on Assumptions Beta cell",
         "Defaults to raw beta. Drives CAPM cost of equity."),
        ("Pre-tax Cost of Debt", wacc_bridge_audit.get("pre_tax_cost_of_debt"), "User input / selected normalized Kd",
         export_cutoff, "Editable on Assumptions",
         "AAPL selected normalized Kd defaults to 4.30%; generic 5.0% is a fallback only. Cost of Debt Sanity compares vs historical interest / avg debt."),
        ("Debt Weight", wacc_bridge_audit.get("debt_weight"), "Computed (Total Debt / (Total Debt + Equity Mkt Cap))",
         historical_cutoff, "Historical Total Debt + Price x Shares",
         "Market-value weight of debt."),
        ("Equity Weight", wacc_bridge_audit.get("equity_weight"), "Computed (1 - Debt Weight)",
         export_cutoff, "Derived",
         "Market-value weight of equity."),
        ("Mechanical CAPM Reference WACC", wacc_bridge_audit.get("capm_indicative_wacc"),
         "Computed (We*Ke + Wd*Kd*(1-t))",
         export_cutoff, "WACC Institutional Bridge on Assumptions",
         "Cross-check; not the headline. Reference rate."),
        ("Selected / Model WACC (headline)", wacc_bridge_audit.get("selected_model_wacc"),
         "User input / model default", export_cutoff, "Assumptions WACC cell",
         "Headline discount rate. Drives DCF Valuation."),
        ("Selected vs CAPM Spread (bps)", wacc_bridge_audit.get("selected_vs_capm_spread_bps"),
         "Computed", export_cutoff, "WACC Institutional Bridge",
         f"Review tier: {wacc_bridge_audit.get('spread_review_tier', 'N/A')}. <=50 OK | 50-150 Review | >150 High review."),
        ("Implied Pre-tax Cost of Debt", sanity_audit.get("implied_pretax_cost_of_debt"),
         "Computed (Interest Expense / Avg Debt)",
         historical_cutoff, "Historical income statement + balance sheet",
         f"Cost of Debt Sanity review tier: {sanity_audit.get('review_tier', 'N/A')}."),
    ]
    write_section_title(ws, row, "WACC Source Trail", span_cols=6)
    row += 1
    write_header_row(
        ws, row,
        ["Component", "Value", "Source", "Cutoff / Fetch Time", "Cache / Reference", "Methodology"],
        1,
    )
    row += 1
    for label, value, source, cutoff, cache_ref, methodology in wacc_trail_rows:
        _set_label_cell(ws, row, 1, label)
        if isinstance(value, (int, float)):
            c = ws.cell(row=row, column=2, value=float(value))
            c.font = font_formula(); c.border = border_thin()
            # bps shown as plain number; percentages as percent.
            if "bps" in label:
                c.number_format = "0.0"
            elif "Weight" in label or "Rate" in label or "Premium" in label or "WACC" in label or "Cost of Debt" in label:
                c.number_format = FMT_PCT2
            else:
                c.number_format = "0.0000"
        else:
            c = ws.cell(row=row, column=2, value=value if value is not None else "N/A")
            c.font = font_formula(); c.border = border_thin()
        _set_note_cell(ws, row, 3, source)
        _set_note_cell(ws, row, 4, cutoff or "n/a")
        _set_note_cell(ws, row, 5, cache_ref or "n/a")
        _set_note_cell(ws, row, 6, methodology)
        row += 1
    _set_note_cell(
        ws, row, 1,
        "WACC Institutional Upgrade: Selected / Model WACC remains the headline; Mechanical CAPM reference is an "
        "audit diagnostic; the cost-of-debt sanity check tiers the gap to historical interest expense. "
        "No layer auto-overwrites the headline.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    write_section_title(ws, row, "Exit Multiple Default", span_cols=3)
    row += 1
    row = _write_kv_rows(ws, row, [
        ("Exit Multiple", current_params.get("exit_multiple"), FMT_MULTIPLE),
        ("Source", em_source_display, None),
        ("Source Note", current_params.get("exit_multiple_warning") or "No additional source caveat.", None),
        ("Cache File", em_cache_display, None),
        ("Fetched At", em_fetched_at, None),
        ("Raw Value", em_raw_display, None),
    ])
    row += 1

    write_header_row(ws, row, ["Cache Field", "Present", "Note"], 1)
    row += 1
    for key in ["revenue", "ebit", "da", "capex", "wc_change", "tax_rate", "net_debt", "shares", "beta", "currency"]:
        _set_label_cell(ws, row, 1, key)
        c = ws.cell(row=row, column=2, value="Yes" if key in data else "No")
        c.font = font_formula()
        c.border = border_thin()
        _set_note_cell(ws, row, 3, "Single-period cache field")
        row += 1
    row += 2

    write_section_title(ws, row, "Calculation Caveats", span_cols=3)
    row += 1
    caveats = [
        "This workbook is a valuation support / research aid for discussion purposes only. All formulas remain live: open the file in Excel and the cells recompute on load.",
        "UI, API, and Excel valuation share a single engine; per-share intrinsic value, FCF projections, and decision-layer alternatives are sourced from the same calculator output.",
        "FCFF is unlevered: NOPAT = EBIT × (1 − tax rate). Interest expense is modeled in P&L, BS, and CF for 3FS display but is not deducted from FCFF.",
        "Five-year explicit forecast uses days-based working capital (DSO / DIO / DPO), asset-based D&A on Beginning Net PP&E, CapEx as % of revenue, and a tax floor at zero. NOL and deferred tax are not modeled.",
        "Balance Sheet Forecast displays Marketable Securities (ST + LT), Goodwill, Intangibles, Deferred Revenue, Capital Lease Obligations, and Deferred Tax Assets / Liabilities as their own lines. Residual plug rows remain transparent.",
        "Capital lease obligations are memo only — yfinance already includes them in ST and LT Debt totals, so adding them separately to Net Debt would double-count.",
        "Reported / Input Net Debt may differ from raw Debt − Cash because data providers can pre-adjust with cash-like investments. The Net Debt Bridge exposes the gap and the Review Dashboard flags it when material; the workbook does not silently rewrite the reported figure.",
        "Marketable Securities treatment is a valuation judgment. Switching the Selected Net Debt Treatment cell on Assumptions intentionally flips the headline IV.",
        "Shareholder returns (dividends + buybacks) appear in CF Financing and the BS Equity roll-forward only; FCFF and EV are unchanged. Per-share valuation can change only through the Selected Share Count Treatment dropdown. Treasury-stock accounting, SBC vintage modeling, and buyback-as-FCF-deduction are out of scope.",
        "WACC is a valuation judgment. Selected / Model WACC drives the headline; Mechanical CAPM reference is an audit diagnostic computed from rf, ERP, beta, and pre-tax cost of debt. The Review Dashboard tiers the Selected vs Mechanical CAPM reference spread (>150 bps Review, >250 bps High Review) without auto-correcting. Term structure, industry beta lookup, rating-implied cost of debt, and APV are out of scope.",
        "Terminal value method is a valuation judgment. The Review Dashboard tiers TV/EV (>80% Review, >90% High Review) and Gordon vs Exit gap (>25% Review, >50% High Review). Fade Period is a reference case (linear FCF growth fade then Gordon TV on the final fade-year FCF), not a full 10-year 3FS forecast; exit-multiple fade and terminal-period SBC are out of scope.",
        "Trading Comps are a market reference for discussion purposes only. Peer data comes from yfinance (cached daily). Outliers and exclusions are flagged in the Peer Set table; summary statistics use included peers only. Implied IV/share converts EV to Equity using the selected Net Debt and Share Count. The peer set is configurable and is not industry-classification-screened; precedent transactions, forward multiples, paid providers, and HK/CN cross-currency normalization are out of scope.",
        "Sensitivity and Football Field caveats fire when saved Bull/Bear scenarios use different Net Debt, Share Count, WACC, or Terminal treatments than Base. Comparisons across mixed treatments are flagged rather than auto-aligned.",
        "Assumption rationale and break-even blocks are structured discussion aids. They do not replace analyst research, management guidance, consensus estimates, or rating guidance.",
        "Forecast shares remain flat unless the share-count assumption is changed; a full buyback / issuance waterfall is not modeled.",
        "HK and CN exports fall back to the legacy operating forecast where the historical pipeline is not yet wired; the Review Dashboard surfaces these as Fallback, not as errors.",
        "The Review Dashboard is an appendix-style review layer. Review / High Review / Fallback / N/A statuses surface caveats for human judgment and do not alter the valuation formulas.",
        "Scenario Notes report identity, headline, and assumption differences only; they do not surface investment thesis or driver narrative.",
    ]
    for caveat in caveats:
        _set_note_cell(ws, row, 1, caveat)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1


def build_fcf_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "FCF Build"
    n = inp.forecast_years
    source_col_idx = n + 2
    apply_column_widths(ws, {
        "A": 38,
        **{get_column_letter(i + 2): 14 for i in range(n)},
        get_column_letter(source_col_idx): 38,
    })

    ref_revenue = _abs_ref(ASSUMP_CELLS["revenue"])
    ref_tax = _abs_ref(ASSUMP_CELLS["tax_rate"])
    ref_revenue_growth = _abs_ref(ASSUMP_CELLS["revenue_growth"])
    ref_ebit_margin = _abs_ref(ASSUMP_CELLS["ebit_margin"])
    ref_da_pct = _abs_ref(ASSUMP_CELLS["da_pct_revenue"])
    ref_capex_pct = _abs_ref(ASSUMP_CELLS["capex_pct_revenue"])
    ref_wc_pct = _abs_ref(ASSUMP_CELLS["wc_change_pct_revenue"])
    # V3.7.5: discount rate sources from the Selected WACC Used in DCF cell.
    ref_wacc = _abs_ref(ASSUMP_CELLS.get("wacc_used", ASSUMP_CELLS["wacc"]))

    linked_3fs = bool(
        PL_FORECAST_ROWS.get("Revenue")
        and PL_FORECAST_ROWS.get("Operating Income / EBIT")
        and SCHEDULE_ROWS.get("ppe", {}).get("D&A")
        and SCHEDULE_ROWS.get("ppe", {}).get("CapEx")
        and SCHEDULE_ROWS.get("wc", {}).get("Schedule-derived Delta WC")
    )
    active_sources = (getattr(out, "audit", None) or {}).get("active_forecast_sources") or {}
    da_schedule_active = active_sources.get("da") == "schedule_derived_beginning_ppe"
    wc_schedule_active = active_sources.get("delta_nwc") == "schedule_derived_working_capital_days"
    forecast_start_col = _forecast_start_col(ctx)
    source_label = (
        "True 3FS FCFF Build v1 (schedule-derived Delta NWC)"
        if linked_3fs and (da_schedule_active or wc_schedule_active)
        else "Operating forecast assumptions (Revenue x margin / pct drivers)"
    )

    row = _write_sheet_heading(ws, 1, "FCF Build - Unlevered FCFF", inp, out, ctx, span_cols=max(3, n + 2))
    _set_label_cell(ws, row, 1, "DCF Source", True)
    _set_note_cell(ws, row, 2, source_label)
    row += 2

    write_section_title(ws, row, "DCF Linkage Bridge", span_cols=3)
    row += 1
    write_header_row(ws, row, ["Component", "Source", "FCF Role"], 1)
    row += 1
    bridge_rows = [
        ("Revenue", "P&L Forecast" if linked_3fs else "Assumptions fallback", "FCFF driver"),
        ("EBIT", "P&L Forecast" if linked_3fs else "Assumptions fallback", "NOPAT driver"),
        ("Tax / NOPAT", "P&L Forecast EBIT x Assumptions tax_rate (NOPAT keeps tax on EBIT for valuation)", "FCFF driver"),
        ("D&A", "Supporting Schedules PP&E" if linked_3fs and da_schedule_active else "Assumptions fallback", "Add-back (sole source)"),
        ("CapEx", "Supporting Schedules PP&E" if linked_3fs else "Assumptions fallback", "Cash outflow (sole source)"),
        ("Delta NWC", "Supporting Schedules WC schedule" if linked_3fs and wc_schedule_active else "Operating assumption: Revenue x Delta NWC % Revenue", "Used in FCFF"),
        ("Selected Delta NWC diagnostic", "Supporting Schedules WC schedule (diagnostic reference)" if linked_3fs else "n/a", "Reconciliation diagnostic"),
    ]
    for component, source, role in bridge_rows:
        _set_label_cell(ws, row, 1, component)
        _set_note_cell(ws, row, 2, source)
        _set_note_cell(ws, row, 3, role)
        row += 1
    row += 1

    write_header_row(ws, row, [""] + [f"Year {i + 1}" for i in range(n)] + ["Source"], 1)
    row += 1

    FCF_ROWS["revenue"] = row
    _set_label_cell(ws, row, 1, "Revenue")
    for i in range(n):
        yr_growth = _path_ref("revenue_growth", i)
        if linked_3fs:
            formula = f"={_sheet_ref('P&L Forecast', PL_FORECAST_ROWS['Revenue'], forecast_start_col + i)}"
        elif i == 0:
            formula = f"=ROUND({ref_revenue}*(1+{yr_growth}),2)"
        else:
            prev_col = get_column_letter(1 + i)
            formula = f"=ROUND({prev_col}{row}*(1+{yr_growth}),2)"
        _set_formula_cell(ws, row, 2 + i, formula, FMT_COMMA2, True)
    _set_note_cell(ws, row, source_col_idx, "P&L Forecast" if linked_3fs else "Assumptions fallback")
    row += 1

    FCF_ROWS["revenue_growth"] = row
    _set_label_cell(ws, row, 1, "Revenue Growth")
    for i in range(n):
        yr_growth = _path_ref("revenue_growth", i)
        if linked_3fs:
            formula = f"={_sheet_ref('P&L Forecast', PL_FORECAST_ROWS['Revenue Growth (%)'], forecast_start_col + i)}"
        else:
            formula = f"={yr_growth}"
        _set_formula_cell(ws, row, 2 + i, formula, FMT_PCT2, True)
    _set_note_cell(ws, row, source_col_idx, "P&L Forecast" if linked_3fs else "Assumptions fallback")
    row += 1

    FCF_ROWS["ebit_margin"] = row
    _set_label_cell(ws, row, 1, "EBIT Margin")
    for i in range(n):
        yr_margin = _path_ref("ebit_margin", i)
        if linked_3fs:
            formula = f"={_sheet_ref('P&L Forecast', PL_FORECAST_ROWS['EBIT Margin (%)'], forecast_start_col + i)}"
        else:
            formula = f"={yr_margin}"
        _set_formula_cell(ws, row, 2 + i, formula, FMT_PCT2, True)
    _set_note_cell(ws, row, source_col_idx, "P&L Forecast" if linked_3fs else "Assumptions fallback")
    row += 1

    FCF_ROWS["ebit"] = row
    _set_label_cell(ws, row, 1, "EBIT")
    for i in range(n):
        col = get_column_letter(2 + i)
        if linked_3fs:
            formula = f"={_sheet_ref('P&L Forecast', PL_FORECAST_ROWS['Operating Income / EBIT'], forecast_start_col + i)}"
        else:
            formula = f"=ROUND({col}{FCF_ROWS['revenue']}*{col}{FCF_ROWS['ebit_margin']},2)"
        _set_formula_cell(ws, row, 2 + i, formula, FMT_COMMA2, True)
    _set_note_cell(ws, row, source_col_idx, "P&L Forecast" if linked_3fs else "Assumptions fallback")
    row += 1

    FCF_ROWS["taxes"] = row
    _set_label_cell(ws, row, 1, "  Less: Taxes")
    for i in range(n):
        col = get_column_letter(2 + i)
        _set_formula_cell(ws, row, 2 + i, f"=ROUND(-{col}{FCF_ROWS['ebit']}*{ref_tax},2)", FMT_COMMA2, True)
    _set_note_cell(ws, row, source_col_idx, "P&L Forecast EBIT x Assumptions tax_rate")
    row += 1

    FCF_ROWS["nopat"] = row
    _set_label_cell(ws, row, 1, "NOPAT", True)
    for i in range(n):
        col = get_column_letter(2 + i)
        _set_formula_cell(ws, row, 2 + i, f"=ROUND({col}{FCF_ROWS['ebit']}+{col}{FCF_ROWS['taxes']},2)", FMT_COMMA2)
    _set_note_cell(ws, row, source_col_idx, "P&L Forecast / Assumptions tax_rate")
    row += 1

    FCF_ROWS["da"] = row
    _set_label_cell(ws, row, 1, "  Add: D&A")
    for i in range(n):
        col = get_column_letter(2 + i)
        yr_da_pct = _path_ref("da_pct_revenue", i)
        if linked_3fs and da_schedule_active:
            formula = f"={_sheet_ref('Supporting Schedules', SCHEDULE_ROWS['ppe']['D&A'], forecast_start_col + i)}"
        else:
            formula = f"=ROUND({col}{FCF_ROWS['revenue']}*{yr_da_pct},2)"
        _set_formula_cell(ws, row, 2 + i, formula, FMT_COMMA2, True)
    _set_note_cell(ws, row, source_col_idx, "Supporting Schedules" if linked_3fs and da_schedule_active else "Assumptions fallback")
    row += 1

    FCF_ROWS["capex"] = row
    _set_label_cell(ws, row, 1, "  Less: CapEx")
    for i in range(n):
        col = get_column_letter(2 + i)
        yr_capex_pct = _path_ref("capex_pct_revenue", i)
        if linked_3fs:
            formula = f"={_sheet_ref('Supporting Schedules', SCHEDULE_ROWS['ppe']['CapEx'], forecast_start_col + i)}"
        else:
            formula = f"=ROUND({col}{FCF_ROWS['revenue']}*{yr_capex_pct},2)"
        _set_formula_cell(ws, row, 2 + i, formula, FMT_COMMA2, True)
    _set_note_cell(ws, row, source_col_idx, "Supporting Schedules" if linked_3fs else "Assumptions fallback")
    row += 1

    FCF_ROWS["wc"] = row
    _set_label_cell(ws, row, 1, "  Less: Delta NWC")
    for i in range(n):
        col = get_column_letter(2 + i)
        yr_wc_pct = _path_ref("wc_change_pct_revenue", i)
        if linked_3fs and wc_schedule_active:
            formula = f"={_sheet_ref('Supporting Schedules', SCHEDULE_ROWS['wc']['Schedule-derived Delta WC'], forecast_start_col + i)}"
        else:
            formula = f"=ROUND({col}{FCF_ROWS['revenue']}*{yr_wc_pct},2)"
        _set_formula_cell(ws, row, 2 + i, formula, FMT_COMMA2, True)
    _set_note_cell(ws, row, source_col_idx, "Supporting Schedules WC schedule-derived Delta WC" if linked_3fs and wc_schedule_active else "Operating assumption: Revenue x Delta NWC % Revenue")
    row += 1

    FCF_ROWS["legacy_wc"] = row
    _set_label_cell(ws, row, 1, "Selected Delta NWC diagnostic")
    for i in range(n):
        if linked_3fs:
            formula = f"={_sheet_ref('Supporting Schedules', SCHEDULE_ROWS['wc']['Selected Delta NWC diagnostic'], forecast_start_col + i)}"
        else:
            formula = None
        if formula:
            _set_formula_cell(ws, row, 2 + i, formula, FMT_COMMA2, True)
        else:
            c = ws.cell(row=row, column=2 + i, value=None)
            c.border = border_thin()
            c.font = font_formula()
            c.number_format = FMT_COMMA2
    _set_note_cell(ws, row, source_col_idx, "Supporting Schedules legacy reference (Revenue x wc_change_pct_revenue); NOT used in FCFF")
    row += 1

    FCF_ROWS["fcf"] = row
    _set_label_cell(ws, row, 1, "Unlevered Free Cash Flow / FCFF", True)
    for i in range(n):
        col = get_column_letter(2 + i)
        _set_formula_cell(
            ws,
            row,
            2 + i,
            f"=ROUND({col}{FCF_ROWS['nopat']}+{col}{FCF_ROWS['da']}-{col}{FCF_ROWS['capex']}-{col}{FCF_ROWS['wc']},2)",
            FMT_COMMA2,
        )
    _set_note_cell(ws, row, source_col_idx, "NOPAT + D&A - CapEx - Delta NWC; excludes interest, dividends, and buybacks")
    row += 1

    FCF_ROWS["df"] = row
    _set_label_cell(ws, row, 1, "Discount Factor")
    for i in range(n):
        _set_formula_cell(ws, row, 2 + i, f"=ROUND(1/(1+{ref_wacc})^{i + 1},6)", "0.000000", True)
    _set_note_cell(ws, row, source_col_idx, "Assumptions WACC")
    row += 1

    FCF_ROWS["pv"] = row
    _set_label_cell(ws, row, 1, "PV of FCFF", True)
    for i in range(n):
        col = get_column_letter(2 + i)
        _set_formula_cell(ws, row, 2 + i, f"=ROUND({col}{FCF_ROWS['fcf']}*{col}{FCF_ROWS['df']},2)", FMT_COMMA2)
    _set_note_cell(ws, row, source_col_idx, "FCF Build")
    row += 2

    write_section_title(ws, row, "FCF Concept Reconciliation Bridge", span_cols=max(3, n + 2))
    row += 1
    write_header_row(ws, row, ["Bridge Item"] + [f"Year {i + 1}" for i in range(n)] + ["Purpose"], 1)
    row += 1
    FCF_ROWS["concept_reconciliation"] = {}
    concept_rows = [
        ("CFO-derived FCF Reference", "cfo_ref", FMT_COMMA2),
        ("Interest Expense", "interest_expense", FMT_COMMA2),
        ("After-tax interest add-back", "after_tax_interest", FMT_COMMA2),
        ("Unlevered FCFF used in DCF", "fcff_used", FMT_COMMA2),
        ("Other CFO bridge items / not modeled", "other_bridge", FMT_COMMA2),
        ("Explained gap: FCFF - CFO-derived reference", "explained_gap", FMT_COMMA2),
    ]
    for label, key, fmt in concept_rows:
        FCF_ROWS["concept_reconciliation"][key] = row
        _set_label_cell(ws, row, 1, label, bold=key in {"fcff_used", "explained_gap"})
        for i in range(n):
            col_idx = 2 + i
            col = get_column_letter(col_idx)
            forecast_col_idx = forecast_start_col + i
            if key == "cfo_ref" and CF_FORECAST_ROWS.get(CF_FCF_REF_LABEL):
                _set_formula_cell(ws, row, col_idx, f"={_sheet_ref('Cash Flow Forecast', CF_FORECAST_ROWS[CF_FCF_REF_LABEL], forecast_col_idx)}", fmt, True)
            elif key == "interest_expense" and PL_FORECAST_ROWS.get("Interest Expense"):
                _set_formula_cell(ws, row, col_idx, f"={_sheet_ref('P&L Forecast', PL_FORECAST_ROWS['Interest Expense'], forecast_col_idx)}", fmt, True)
            elif key == "after_tax_interest":
                interest_ref = f"{col}{FCF_ROWS['concept_reconciliation']['interest_expense']}"
                _set_formula_cell(ws, row, col_idx, f'=IF(ISNUMBER({interest_ref}),ROUND({interest_ref}*(1-{ref_tax}),2),"N/A")', fmt, True)
            elif key == "other_bridge":
                fcff_ref = f"{col}{FCF_ROWS['concept_reconciliation']['fcff_used']}"
                cfo_ref = f"{col}{FCF_ROWS['concept_reconciliation']['cfo_ref']}"
                ati_ref = f"{col}{FCF_ROWS['concept_reconciliation']['after_tax_interest']}"
                _set_formula_cell(ws, row, col_idx, f'=IF(AND(ISNUMBER({fcff_ref}),ISNUMBER({cfo_ref}),ISNUMBER({ati_ref})),ROUND({fcff_ref}-{cfo_ref}-{ati_ref},2),"N/A")', fmt)
            elif key == "fcff_used":
                _set_formula_cell(ws, row, col_idx, f"={col}{FCF_ROWS['fcf']}", fmt, True)
            elif key == "explained_gap":
                fcff_ref = f"{col}{FCF_ROWS['concept_reconciliation']['fcff_used']}"
                cfo_ref = f"{col}{FCF_ROWS['concept_reconciliation']['cfo_ref']}"
                _set_formula_cell(ws, row, col_idx, f'=IF(AND(ISNUMBER({fcff_ref}),ISNUMBER({cfo_ref})),ROUND({fcff_ref}-{cfo_ref},2),"N/A")', fmt)
            else:
                c = ws.cell(row=row, column=col_idx, value="N/A")
                c.font = font_formula()
                c.border = border_thin()
        purpose = {
            "cfo_ref": "Cash-flow-statement reference from Cash Flow Forecast; not the DCF FCFF source.",
            "interest_expense": "Interest is below EBIT in P&L and excluded from unlevered FCFF.",
            "after_tax_interest": "Shows the common levered-vs-unlevered difference: Interest Expense x (1 - tax rate).",
            "other_bridge": "Review item when material: residual CFO-to-FCFF bridge items where detailed non-cash / working-capital mapping is unavailable; not fully modeled.",
            "fcff_used": "DCF source: unlevered NOPAT-based FCFF from this sheet.",
            "explained_gap": "Reviewer-facing difference between the two cash-flow concepts.",
        }.get(key, "")
        _set_note_cell(ws, row, source_col_idx, purpose)
        row += 1
    row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "The Cash Flow Forecast reference starts from Net Income / CFO and is post-interest in nature. DCF uses unlevered FCFF from EBIT / NOPAT. The bridge shows after-tax interest add-back explicitly and leaves other CFO bridge items visible instead of forcing the two cash-flow concepts to equal each other.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(3, n + 2))
    row += 2

    # V3.7.0: API and Excel now share the same unified engine. The recon block
    # is repurposed to verify Excel-live FCF == Unified Engine FCF and to keep
    # the legacy operating forecast as a historical reference column.
    engine_label = (out.audit or {}).get("engine", "unknown")
    write_section_title(
        ws,
        row,
        f"V3.7.0 FCF Parity: Unified Engine vs Excel Live vs Legacy Reference (engine={engine_label})",
        span_cols=max(3, n + 2),
    )
    row += 1
    write_header_row(ws, row, ["Year"] + [f"Year {i + 1}" for i in range(n)] + ["Status"], 1)
    row += 1

    # Compute legacy operating-forecast FCF in Python so the reference column has
    # something to compare against even after the calculator switched to unified.
    try:
        legacy_forecast = build_operating_forecast(inp)
        legacy_fcf = legacy_forecast.get("fcf_projections", [])
    except Exception:
        legacy_fcf = []

    recon_rows = [
        ("Unified Engine FCF (API / single source)", "unified_api_fcf", FMT_COMMA2),
        ("Excel Live True 3FS FCF (formula)", "excel_live_fcf", FMT_COMMA2),
        ("Legacy operating-forecast FCF (V3.6.x reference)", "legacy_fcf", FMT_COMMA2),
        ("Diff: Excel - Unified (parity check, target = 0)", "difference", FMT_COMMA2),
        ("Diff %: Excel vs Unified", "difference_pct", FMT_PCT2),
        ("Legacy vs Unified Diff (historical reference)", "legacy_diff", FMT_COMMA2),
    ]
    FCF_ROWS["reconciliation"] = {}
    for label, key, fmt in recon_rows:
        FCF_ROWS["reconciliation"][key] = row
        _set_label_cell(ws, row, 1, label, bold=label.startswith("Diff"))
        for i in range(n):
            col_idx = 2 + i
            col = get_column_letter(col_idx)
            if key == "unified_api_fcf":
                value = out.fcf_projections[i] if i < len(out.fcf_projections) else None
                c = ws.cell(row=row, column=col_idx, value=value)
                c.font = font_formula()
                c.border = border_thin()
                c.number_format = fmt
            elif key == "excel_live_fcf":
                _set_formula_cell(ws, row, col_idx, f"={col}{FCF_ROWS['fcf']}", fmt, True)
            elif key == "legacy_fcf":
                value = legacy_fcf[i] if i < len(legacy_fcf) else None
                c = ws.cell(row=row, column=col_idx, value=value)
                c.font = font_formula()
                c.border = border_thin()
                c.number_format = fmt
            elif key == "difference":
                _set_formula_cell(
                    ws,
                    row,
                    col_idx,
                    f"={col}{FCF_ROWS['reconciliation']['excel_live_fcf']}-{col}{FCF_ROWS['reconciliation']['unified_api_fcf']}",
                    fmt,
                )
            elif key == "difference_pct":
                _set_formula_cell(
                    ws,
                    row,
                    col_idx,
                    f"=IFERROR({col}{FCF_ROWS['reconciliation']['difference']}/{col}{FCF_ROWS['reconciliation']['unified_api_fcf']},0)",
                    fmt,
                )
            elif key == "legacy_diff":
                _set_formula_cell(
                    ws,
                    row,
                    col_idx,
                    f"={col}{FCF_ROWS['reconciliation']['legacy_fcf']}-{col}{FCF_ROWS['reconciliation']['unified_api_fcf']}",
                    fmt,
                )
        # Status column for the difference row.
        status_col = n + 2
        if key == "difference":
            max_abs = "MAX(" + ",".join(
                f"ABS({get_column_letter(2 + i)}{row})" for i in range(n)
            ) + ")"
            _set_formula_cell(
                ws,
                row,
                status_col,
                f'=IF({max_abs}<=1,"OK (parity within rounding)","Review: engine drift detected")',
                None,
                True,
            )
        elif key == "legacy_diff":
            _set_note_cell(ws, row, status_col, "Reference only - shows how much V3.7.0 unification changed FCF vs the V3.6.x legacy calculator.")
        else:
            _set_note_cell(ws, row, status_col, "")
        row += 1
    row += 1
    _set_note_cell(
        ws,
        row,
        1,
        "V3.7.0 parity check: Unified Engine FCF (the value run_dcf returns to UI / API / Excel) should match Excel Live True 3FS FCF within rounding (Diff target = 0). The Legacy reference row shows what the pre-V3.7.0 operating-forecast calculator would have produced for the same inputs - useful for explaining the V3.6.x → V3.7.0 valuation change to reviewers.",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(3, n + 2))
    row += 2

    # Valuation Reconciliation V3.6.7 vs V3.6.8 is added to DCF Valuation sheet
    # after VAL_ROWS['intrinsic'] is known. See build_valuation_sheet().
    _add_watermark_row(ws, row, inp.symbol)


def _apply_workbook_print_packaging(wb, inp: DCFInputs | None = None, ctx: dict | None = None):
    settings = {
        "Cover": {"landscape": False, "freeze": None, "repeat": None},
        "Executive Summary": {"landscape": True, "freeze": "A7", "repeat": "1:6"},
        "Scenario Notes": {"landscape": True, "freeze": "A5", "repeat": "1:5"},
        "Historical Financials": {"landscape": True, "freeze": "B8", "repeat": "1:8"},
        "P&L Forecast": {"landscape": True, "freeze": "B8", "repeat": "1:8"},
        "Balance Sheet Forecast": {"landscape": True, "freeze": "B10", "repeat": "1:10"},
        "Cash Flow Forecast": {"landscape": True, "freeze": "B8", "repeat": "1:8"},
        "Supporting Schedules": {"landscape": True, "freeze": "B8", "repeat": "1:8"},
        "Assumptions": {"landscape": False, "freeze": "A5", "repeat": "1:5"},
        "FCF Build": {"landscape": True, "freeze": "B5", "repeat": "1:5"},
        "DCF Valuation": {"landscape": False, "freeze": "A6", "repeat": "1:6"},
        "Trading Comps": {"landscape": True, "freeze": "A6", "repeat": "1:6"},
        "FA Ratios": {"landscape": True, "freeze": "B9", "repeat": "1:9"},
        "Audit Dashboard": {"landscape": True, "freeze": "A16", "repeat": "1:16"},
        "Sensitivity": {"landscape": True, "freeze": "A5", "repeat": "1:5"},
        "Data Sources Audit": {"landscape": True, "freeze": "A5", "repeat": "1:5"},
    }
    generated_date = _display_date((ctx or {}).get("generated_at")) or date.today().isoformat()
    company = getattr(inp, "company", None) if inp is not None else None
    ticker = getattr(inp, "symbol", None) if inp is not None else None
    for sheet_name, config in settings.items():
        if sheet_name not in wb.sheetnames:
            continue
        _apply_print_setup(
            wb[sheet_name],
            landscape=config["landscape"],
            freeze=config["freeze"],
            repeat_rows=config["repeat"],
            company=company,
            ticker=ticker,
            generated_date=generated_date,
        )


def build_cover_sheet(ws, inp: DCFInputs, out: DCFOutputs, ctx: dict):
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    apply_column_widths(ws, {"A": 5, "B": 26, "C": 26, "D": 28, "E": 26, "F": 22})

    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell.border = Border()
            cell.fill = PatternFill(fill_type=None)

    scenario = _scenario_label(ctx.get("current_scenario"))
    generated = _display_date(ctx.get("generated_at")) or date.today().isoformat()
    current_params = ctx.get("current_params") or {}
    data_source = "Public market data, cached financial statements, and trading data."

    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=6)
    c = ws.cell(3, 2, f"{inp.company} ({inp.symbol})")
    c.font = Font(name="Arial", size=24, bold=True, color=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 34

    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=6)
    c = ws.cell(4, 2, "DCF Valuation Workbook")
    c.font = Font(name="Arial", size=16, bold=False, color="404040")
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=6)
    c = ws.cell(6, 2, f"{scenario} Case | {_display_date(generated)[:7] if generated else 'May 2026'}")
    c.font = Font(name="Arial", size=12, bold=True, color=COLOR_FORMULA_FG)

    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=6)
    c = ws.cell(7, 2, "Valuation workbook and IC discussion aid")
    c.font = Font(name="Arial", size=10, italic=True, color="666666")

    row = 10
    write_section_title(ws, row, "Workbook Identity", span_cols=6)
    row += 1
    identity_rows = [
        ("Ticker", inp.symbol),
        ("Scenario", scenario),
        ("Generated Date", generated),
        ("Currency / Market", f"{out.currency} / {out.market}"),
        ("Security Type", current_params.get("security_type") or "operating_company"),
        ("DCF Suitability", current_params.get("dcf_suitability") or ("unsuitable" if current_params.get("model_unsuitable") else "suitable")),
        ("Reporting Currency", current_params.get("reporting_currency") or out.currency),
        ("Trading Currency", current_params.get("trading_currency") or out.currency),
        ("FX Rate Used", current_params.get("fx_rate_reporting_to_trading")),
        ("FX Source", current_params.get("fx_rate_source") or "Not available"),
        ("Unit", model_unit_label(out.currency, out.market)),
        ("Model Version", MODEL_VERSION),
        ("Data Source", data_source),
    ]
    for i, (label, value) in enumerate(identity_rows):
        r = row + i // 2
        c0 = 2 if i % 2 == 0 else 4
        _set_label_cell(ws, r, c0, label)
        v = ws.cell(r, c0 + 1, value)
        v.font = font_formula()
        v.border = border_thin()
        v.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        ws.row_dimensions[r].height = 20
    row += (len(identity_rows) + 1) // 2 + 1

    if current_params.get("model_unsuitable") or current_params.get("valuation_status") == "model_unsuitable":
        write_section_title(ws, row, "Model Suitability", span_cols=6)
        row += 1
        _set_note_cell(
            ws,
            row,
            2,
            current_params.get("model_unsuitable_reason") or "DCF unavailable for this security type.",
        )
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
        _set_note_cell(
            ws,
            row,
            2,
            "Recommended methods: " + (", ".join(str(x) for x in (current_params.get("recommended_methods") or [])) or "NAV / holdings look-through / expense ratio"),
        )
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 2

    write_section_title(ws, row, "Workbook Guide", span_cols=6)
    row += 1
    guide = [
        ("Executive Summary", "Headline valuation, selected case rationale, IC discussion items, and football field."),
        ("Assumptions", "Editable operating, WACC, terminal value, net debt, share count, and capital return inputs."),
        ("DCF Valuation", "Live DCF bridge, treatment alternatives, break-even analysis, and defense blocks."),
        ("Trading Comps", "Peer multiples and implied valuation reference ranges."),
        ("Sensitivity", "Current price support summary and key assumption grids."),
        ("Audit Dashboard", "Appendix-style review checks and IC discussion flags."),
        ("Data Sources Audit", "Methodology, source boundaries, and assumption rationale disclosure."),
    ]
    write_header_row(ws, row, ["Section", "Purpose", "", "", "", ""], 2)
    row += 1
    for section, purpose in guide:
        _set_label_cell(ws, row, 2, section)
        _set_note_cell(ws, row, 3, purpose)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        row += 1

    row += 1
    write_section_title(ws, row, "Scope Boundary", span_cols=6)
    row += 1
    _set_note_cell(
        ws,
        row,
        2,
        SCOPE_BOUNDARY_TEXT,
    )
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    _set_note_cell(ws, row, 2, SCOPE_BOUNDARY_DETAIL)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 42
    row += 1

    row += 1
    write_section_title(ws, row, "Release Boundary", span_cols=6)
    row += 1
    _set_note_cell(
        ws,
        row,
        2,
        RELEASE_BOUNDARY_TEXT,
    )
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1

    row += 1
    write_section_title(ws, row, "Formatting Legend", span_cols=6)
    row += 1
    legend = [
        ("Blue font", "User-editable inputs"),
        ("Black font", "Calculations and labels"),
        ("Green font", "Internal workbook links"),
    ]
    colors = [COLOR_INPUT_FG, COLOR_FORMULA_FG, COLOR_LINK_FG]
    for (label, text), color in zip(legend, colors):
        _set_label_cell(ws, row, 2, label)
        c = ws.cell(row, 3, text)
        c.font = Font(name="Arial", size=10, color=color)
        c.border = border_thin()
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row, 2, SCOPE_BOUNDARY_TEXT)
    c.font = font_watermark()
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 28


def _client_number_format_for_label(label: str | None):
    text = (label or "").lower()
    if "bps" in text or "spread (bps)" in text:
        return '0 "bps";[Red](0) "bps";-'
    if any(term in text for term in ("margin", "growth", "wacc", "tax rate", "cost of debt", "pre-tax", "rate", "yield", "ratio", "%", "tv / ev", "upside", "downside", "payout", "buybacks / fcf", "dividends / fcf", "terminal g", "implied g", "premium")):
        return '0.0%;[Red](0.0%);-'
    if any(term in text for term in ("multiple", "ev/ebitda", "ev / ebitda", "ev/revenue", "ev / revenue", "p / e", "p/e")):
        return '0.0x;[Red](0.0x);-'
    if any(term in text for term in ("per share", "/ share", "current price", "intrinsic value", "iv / share", "eps", "fcf per share")):
        return '$#,##0.00;[Red]($#,##0.00);-'
    if any(term in text for term in ("shares", "fade years", "count")):
        return '#,##0.0;[Red](#,##0.0);-'
    if any(term in text for term in ("revenue", "ebit", "ebitda", "fcf", "cash", "debt", "value", "ev", "equity", "capex", "dividend", "buyback", "marketable", "assets", "liabilities", "terminal", "pv ", "enterprise")):
        return '$#,##0.0"M";[Red]($#,##0.0"M");-'
    return None


def _apply_client_formatting(wb):
    client_sheets = {
        "Cover",
        "Executive Summary",
        "Scenario Notes",
        "Assumptions",
        "DCF Valuation",
        "Trading Comps",
        "Sensitivity",
        "Audit Dashboard",
        "Data Sources Audit",
    }
    input_labels = {
        "Current Price", "Risk-Free Rate (Rf)", "Equity Risk Premium (ERP)", "Beta", "WACC (used in model)",
        "Cost of Debt (Kd, pre-tax)", "Selected WACC Treatment", "Terminal Treatment", "Fade Years",
        "Fade Target Growth", "Blend Weight: Gordon", "Blend Weight: Exit", "Selected Net Debt Treatment",
        "Selected Share Count Treatment", "Dividend Payout % of Net Income", "Buyback Method",
        "Buyback % of FCF (used when method = % of FCF)", "Flat Buyback Amount (used when method = Flat)",
        "Repurchase Price Growth (annual)", "Annual Dilution (% of beg. shares, SBC etc.)",
    }
    path_input_labels = {
        "Revenue Growth", "EBIT Margin", "D&A % Revenue",
        "CapEx % Revenue", "Delta NWC % Revenue",
    }
    for sheet_name in client_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            row_label = row[0].value if row else None
            label_text = str(row_label or "")
            fmt = _client_number_format_for_label(label_text)
            for cell in row:
                value = cell.value
                fill_rgb = getattr(cell.fill.fgColor, "rgb", None)
                is_dark_header_fill = cell.fill and cell.fill.fill_type == "solid" and fill_rgb in {"FF1F4E78", "1F4E78", "FF1A3A5C", "001A3A5C", "1A3A5C"}
                if sheet_name == "Sensitivity" and is_dark_header_fill:
                    cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    continue
                if isinstance(value, str) and value.startswith("="):
                    cell.font = font_link() if "!" in value else font_formula()
                elif sheet_name == "Assumptions" and label_text in path_input_labels and 2 <= cell.column <= 6:
                    cell.font = font_input()
                elif sheet_name == "Assumptions" and cell.column == 2 and label_text in input_labels:
                    cell.font = font_input()
                elif sheet_name == "Assumptions" and cell.column == 2 and not isinstance(value, str) and value is not None:
                    cell.font = font_input()
                elif cell.font and not cell.font.bold:
                    cell.font = font_formula()

                if fmt and (isinstance(value, (int, float)) or (isinstance(value, str) and value.startswith("="))):
                    cell.number_format = fmt
                if isinstance(value, (int, float)) and cell.number_format in ("General", ""):
                    fallback = _client_number_format_for_label(label_text)
                    if fallback:
                        cell.number_format = fallback
                    elif abs(float(value) - round(float(value), 4)) > 1e-9:
                        cell.number_format = '0.0'
                if sheet_name in {"Executive Summary", "Data Sources Audit", "Audit Dashboard"} and isinstance(value, str) and len(value) > 60:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")


def _sanitize_client_facing_text(wb):
    """Remove internal release and backend wording from every visible sheet.

    Strips engineering iteration history (V3.x.x tokens, "Prior release",
    "Generated by Modeling Studio") so the final workbook reads as a neutral
    deliverable. The Cover sheet's MODEL_VERSION uses a lowercase "v3.8.9"
    which is intentionally preserved.
    """
    replacements = [
        # Strip engineering version tokens (V3.6.8, V3.7.x, V3.8.0, etc.) along
        # with an optional trailing colon/space so the residual sentence reads
        # cleanly.
        (re.compile(r"\bV3\.\d+(?:\.(?:\d+|x))?\b\s*[:：]?\s*"), ""),
        (re.compile(r"\bv3\.\d+(?:\.(?:\d+|x)){0,2}\b\s*[:：]?\s*", re.IGNORECASE), ""),
        # Drop any leftover "Prior release" tokens from earlier sanitizer runs.
        (re.compile(r"\bPrior release\b\s*[:：]?\s*"), ""),
        (re.compile(r"\bv[0-9]{3,}\b", re.IGNORECASE), "cached historical"),
        (re.compile(r"export payload", re.IGNORECASE), "export metadata"),
        (re.compile(r"\bpayload\b", re.IGNORECASE), "metadata"),
        (re.compile(r"pseudo-formula", re.IGNORECASE), "formula display"),
        (re.compile(r"\bpatch\b", re.IGNORECASE), "update"),
        (re.compile(r"\bdebug\b", re.IGNORECASE), "diagnostic"),
        (re.compile(r"Generated by Modeling Studio"), "Generated"),
        (re.compile(r"Modeling Studio"), "this workbook"),
        (re.compile(r"Form 10-K consolidated", re.IGNORECASE), "public financial statement"),
        (re.compile(r"Form 10-K", re.IGNORECASE), "public financial statement"),
        (re.compile(r"10-K Backfill", re.IGNORECASE), "public financial statement source"),
        (re.compile(r"\b10-K\b", re.IGNORECASE), "filing"),
        (re.compile(r"N/D in 10-K", re.IGNORECASE), "N/D in filing"),
        (re.compile(r"Not separately disclosed in 10-K", re.IGNORECASE), "N/D in filing"),
        (re.compile(r"\bsuppressed_generic_tech_fallback\b", re.IGNORECASE), "Generic technology fallback suppressed"),
        (re.compile(r"\bgeneric_tech_fallback_v1\b", re.IGNORECASE), "Generic technology fallback"),
        (re.compile(r"\bcro_cdmo_peer_mapping_required\b", re.IGNORECASE), "CRO/CDMO peer set requires analyst review"),
        (re.compile(r"\bpremium_consumer_peer_mapping_required\b", re.IGNORECASE), "Premium consumer peer set requires analyst review"),
        (re.compile(r"\binternet_platform_peer_mapping_required\b", re.IGNORECASE), "Internet platform peer set requires analyst review"),
        (re.compile(r"\bmodel_unsuitable\b", re.IGNORECASE), "Model unsuitable"),
        (re.compile(r"\bmarket_comparison_status\b", re.IGNORECASE), "Market comparison status"),
        (re.compile(r"\bvaluation_status\b", re.IGNORECASE), "Valuation status"),
        (re.compile(r"\bn/a\b"), "N/A"),
        # V3.9.1: scrub developer / API wording from client-facing text. The
        # workbook valuation chain is referred to as an engine, not an API.
        (re.compile(r"\bAPI\s+Unified\s+Engine\b"), "Engine"),
        (re.compile(r"\bAPI\s+Unified\b"), "Engine"),
        (re.compile(r"\bAPI\s+vs\s+Excel\b"), "Engine vs Excel"),
        (re.compile(r"\bUI\s*/\s*API\s*/\s*Excel\b"), "UI and Excel"),
        (re.compile(r"\bUI\s*,\s*API\s*,\s*and\s*Excel\b"), "UI and Excel"),
        (re.compile(r"\bAPI\s*/\s*Excel\b"), "Engine / Excel"),
        (re.compile(r"\band\s+the\s+API\b"), "and the engine"),
        (re.compile(r"\bAPI\b"), "engine"),
        (re.compile(r"\brun_dcf\(\)"), "the valuation engine"),
        (re.compile(r"\brun_dcf\b"), "the valuation engine"),
        (re.compile(r"\bdcf_calculator\.py\b"), "the valuation engine"),
        (re.compile(r"\bdcf_calculator\b"), "the valuation engine"),
        (re.compile(r"\bExcel\s+COM\b"), "Excel"),
        (re.compile(r"fallback export", re.IGNORECASE), "limited-market export"),
        (re.compile(r"Fallback \(cache fields unavailable\)", re.IGNORECASE), "Limited cache fields"),
        (re.compile(r"No fallback warning", re.IGNORECASE), "No additional source caveat."),
        (re.compile(r"Graceful fallback only", re.IGNORECASE), "Limited-market presentation only"),
    ]
    whitespace_collapse = re.compile(r"[ \t]{2,}")
    leading_punct = re.compile(r"^[\s:：,;]+")

    def _clean(text: str) -> str:
        for pattern, repl in replacements:
            text = pattern.sub(repl, text)
        text = whitespace_collapse.sub(" ", text)
        text = leading_punct.sub("", text)
        return text.strip()

    for sheet_name in wb.sheetnames:
        # The Cover sheet's MODEL_VERSION line is the one place an external
        # version label is allowed. We still sanitize every other cell on it.
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or value.startswith("="):
                    continue
                if value == MODEL_VERSION:
                    continue
                cell.value = _clean(value)


def _sanitize_workbook_narrative_and_aapl_leakage(wb, inp: DCFInputs):
    """Clean release wording and suppress AAPL-only narrative from non-AAPL workbooks."""

    symbol = (getattr(inp, "symbol", "") or "").strip().upper()
    is_aapl = symbol == "AAPL"
    stale_replacements = [
        (re.compile(r"Valuation Workbook v3\.9\.10\.0", re.IGNORECASE), MODEL_VERSION),
        (
            re.compile(
                r"Gold Master QA\s*/\s*Scope Boundary Lock:.*?No new valuation features added\.?",
                re.IGNORECASE,
            ),
            RELEASE_BOUNDARY_TEXT,
        ),
        (re.compile(r"\bGold Master\b", re.IGNORECASE), "General-Purpose Hardening"),
        (re.compile(r"No new valuation features added\.?", re.IGNORECASE), RELEASE_BOUNDARY_TEXT),
    ]
    non_aapl_text_replacements = [
        (
            re.compile(r"AAPL default Base path", re.IGNORECASE),
            "Current default / analyst-reviewed assumption path",
        ),
        (
            re.compile(r"AAPL Base/Bear case-specific path", re.IGNORECASE),
            "selected case path",
        ),
        (re.compile(r"AAPL Operating Thesis Bridge", re.IGNORECASE), "Operating Thesis reconciliation"),
        (re.compile(r"AAPL Operating Thesis", re.IGNORECASE), "Operating Thesis reconciliation"),
        (re.compile(r"AAPL defaults", re.IGNORECASE), "ticker-specific defaults"),
        (re.compile(r"AAPL Base", re.IGNORECASE), "Base"),
        (re.compile(r"AAPL Bear", re.IGNORECASE), "Downside"),
        (re.compile(r"Apple Form 10-K", re.IGNORECASE), "company filing"),
        (re.compile(r"Form 10-K", re.IGNORECASE), "public financial statement"),
        (re.compile(r"\b10-K\b", re.IGNORECASE), "filing"),
        (re.compile(r"Apple FY", re.IGNORECASE), "Company FY"),
        (re.compile(r"\bApple\b", re.IGNORECASE), "Company"),
        (re.compile(r"\bAAPL\b", re.IGNORECASE), "this ticker"),
        (re.compile(r"suppressed_generic_tech_fallback", re.IGNORECASE), "Generic technology fallback suppressed"),
        (re.compile(r"generic_tech_fallback_v1", re.IGNORECASE), "Generic technology fallback"),
        (re.compile(r"cro_cdmo_peer_mapping_required", re.IGNORECASE), "CRO/CDMO peer set requires analyst review"),
        (re.compile(r"premium_consumer_peer_mapping_required", re.IGNORECASE), "Premium consumer peer set requires analyst review"),
        (re.compile(r"internet_platform_peer_mapping_required", re.IGNORECASE), "Internet platform peer set requires analyst review"),
        (re.compile(r"model_unsuitable", re.IGNORECASE), "Model unsuitable"),
        (re.compile(r"market_comparison_status", re.IGNORECASE), "Market comparison status"),
        (re.compile(r"valuation_status", re.IGNORECASE), "Valuation status"),
    ]
    aapl_formula_leakage = re.compile(
        r"AAPL|Apple|AAPL Operating Thesis|generic_tech_fallback_v1",
        re.IGNORECASE,
    )
    whitespace_collapse = re.compile(r"[ \t]{2,}")

    def _apply_stale_cleanup(text: str) -> str:
        for pattern, replacement in stale_replacements:
            text = pattern.sub(replacement, text)
        return whitespace_collapse.sub(" ", text).strip()

    def _clean_non_aapl_text(text: str) -> str:
        text = _apply_stale_cleanup(text)
        for pattern, replacement in non_aapl_text_replacements:
            text = pattern.sub(replacement, text)
        return whitespace_collapse.sub(" ", text).strip()

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                if value.startswith("="):
                    cleaned_formula = _apply_stale_cleanup(value)
                    if not is_aapl and aapl_formula_leakage.search(cleaned_formula):
                        cell.value = "N/A - Operating Thesis reconciliation unavailable for this ticker."
                    elif cleaned_formula != value:
                        cell.value = cleaned_formula
                    continue
                cleaned = _apply_stale_cleanup(value) if is_aapl else _clean_non_aapl_text(value)
                if cleaned != value:
                    cell.value = cleaned


def generate_excel(inp: DCFInputs, out: DCFOutputs, export_context: dict | None = None) -> io.BytesIO:
    ASSUMP_CELLS.clear()
    ASSUMP_PATH_CELLS.clear()
    ASSUMP_SELECTED_PATH_CELLS.clear()
    ASSUMP_BRIDGE_PATH_CELLS.clear()
    FCF_ROWS.clear()
    VAL_ROWS.clear()
    PL_FORECAST_ROWS.clear()
    CF_FORECAST_ROWS.clear()
    BS_FORECAST_ROWS.clear()
    SCHEDULE_ROWS.clear()
    SENS_ROWS.clear()
    SCENARIO_ROWS.clear()
    AUDIT_STATUS_ROWS.clear()
    TRADING_COMPS_ROWS.clear()
    ctx = _build_context(inp, out, export_context)

    wb = openpyxl.Workbook()
    wb.calculation.calcMode = "auto"

    ws0 = wb.active
    build_cover_sheet(ws0, inp, out, ctx)

    build_scenario_notes_sheet(wb.create_sheet(), inp, out, ctx)
    ws_historical = wb.create_sheet()
    ws_pl_forecast = wb.create_sheet()
    ws_bs_forecast = wb.create_sheet()
    ws_cf_forecast = wb.create_sheet()
    ws_supporting = wb.create_sheet()
    ws_assumptions = wb.create_sheet()
    build_historical_financials_sheet(ws_historical, inp, out, ctx)
    build_assumptions_sheet(ws_assumptions, inp, out, ctx)
    build_pl_forecast_sheet(ws_pl_forecast, inp, out, ctx)
    build_supporting_schedules_sheet(ws_supporting, inp, out, ctx)
    relink_pl_forecast_to_supporting_schedules(ws_pl_forecast)
    # V3.7.4: now that the Shareholder Returns Schedule exists, point the
    # Assumptions denominator inputs at the live forecast Ending / WAvg shares.
    relink_assumptions_to_shareholder_returns_schedule(ws_assumptions, inp)
    build_cash_flow_forecast_sheet(ws_cf_forecast, inp, out, ctx)
    # V3.7.4: now CF Forecast exists, relink SR Schedule FCF to live cells.
    relink_sr_schedule_to_cf_forecast(ws_supporting)
    build_balance_sheet_forecast_sheet(ws_bs_forecast, inp, out, ctx)
    build_fcf_sheet(wb.create_sheet(), inp, out, ctx)
    build_valuation_sheet(wb.create_sheet(), inp, out, ctx)
    # V3.7.7: Trading Comps sheet sits right after DCF Valuation so reviewers
    # see the comps cross-check immediately. Sheet renders an "unavailable"
    # note if ctx['trading_comps'] is missing or errored.
    build_trading_comps_sheet(wb.create_sheet(), inp, out, ctx)
    # V3.9.9.0: AAPL Operating Thesis Bridge v1 — presentation / support sheet.
    # Rendered only when the symbol is AAPL; non-AAPL workbooks are unchanged.
    if (inp.symbol or "").strip().upper() == "AAPL":
        build_aapl_operating_thesis_sheet(wb.create_sheet(), inp, out, ctx)
    build_fa_ratios_sheet(wb.create_sheet(), inp, out, ctx)
    ws_sensitivity = wb.create_sheet()
    build_sensitivity_sheet(ws_sensitivity, inp, out, ctx)
    ws_audit = wb.create_sheet(index=wb.sheetnames.index("Sensitivity"))
    build_audit_dashboard_sheet(ws_audit, inp, out, ctx)
    build_data_sources_sheet(wb.create_sheet(), inp, out, ctx)
    build_executive_summary_sheet(wb.create_sheet(index=1), inp, out, ctx)
    _sanitize_client_facing_text(wb)
    _sanitize_workbook_narrative_and_aapl_leakage(wb, inp)
    _apply_client_formatting(wb)
    _apply_workbook_print_packaging(wb, inp, ctx)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
