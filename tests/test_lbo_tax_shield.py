import copy
import inspect
from io import BytesIO

from openpyxl import load_workbook

from modeling.lbo_calculator import run_lbo
from modeling.lbo_excel_exporter import generate_lbo_excel
from tests.test_lbo_calculator_core import clean_anchor


def _enabled_anchor():
    payload = clean_anchor()
    payload["tax_shield_enabled"] = True
    payload["tax_rate"] = 0.25
    payload["operating_forecast"]["cash_taxes"] = [200, 200, 200, 200, 200]
    return payload


def _strings(wb, sheet):
    return [c.value for row in wb[sheet].iter_rows() for c in row if c.value is not None]


def test_tax_shield_enabled_reduces_cash_taxes_by_interest_times_rate():
    out = run_lbo(_enabled_anchor())
    y1 = out["debt_schedule"][0]
    assert y1["cash_interest"] == 450
    assert y1["gross_cash_taxes"] == 200
    assert y1["tax_shield"] == 112.5
    assert y1["levered_cash_taxes"] == 87.5
    assert y1["cash_flow_before_debt_service"] == 742.5


def test_tax_shield_floors_taxes_at_zero_and_defaults_enabled():
    payload = clean_anchor()
    payload.pop("tax_shield_enabled")
    payload.pop("tax_rate")
    out = run_lbo(payload)
    y1 = out["debt_schedule"][0]
    assert y1["tax_shield_enabled"] is True
    assert y1["tax_rate"] == 0.25
    assert y1["tax_shield"] == 112.5
    assert y1["levered_cash_taxes"] == 0.0


def test_tax_shield_disabled_preserves_existing_cash_taxes_behavior():
    out = run_lbo(clean_anchor())
    y1 = out["debt_schedule"][0]
    assert y1["gross_cash_taxes"] == 100
    assert y1["tax_shield"] == 0.0
    assert y1["levered_cash_taxes"] == 100
    assert y1["cash_available_for_debt"] == 280


def test_interest_still_uses_beginning_debt_balance():
    out = run_lbo(_enabled_anchor())
    for row in out["debt_schedule"]:
        assert row["cash_interest"] == row["beginning_debt"] * 0.09


def test_tax_shield_improves_or_preserves_returns():
    disabled = clean_anchor()
    enabled = copy.deepcopy(disabled)
    enabled["tax_shield_enabled"] = True
    enabled["tax_rate"] = 0.25
    a = run_lbo(disabled)
    b = run_lbo(enabled)
    assert b["returns"]["irr"] >= a["returns"]["irr"]
    assert b["returns"]["moic"] >= a["returns"]["moic"]


def test_no_iteration_or_circular_reference_introduced():
    import modeling.lbo_calculator as lbo

    source = inspect.getsource(lbo.run_lbo)
    assert "while" not in source
    assert "for _ in range" not in source


def test_excel_displays_tax_shield_fields_and_audit_note():
    payload = _enabled_anchor()
    out = run_lbo(payload)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, out).getvalue()), data_only=False)
    assumptions = _strings(wb, "Assumptions")
    debt = _strings(wb, "Debt Schedule")
    operating = _strings(wb, "Operating Forecast")
    audit = _strings(wb, "Audit & Disclosures")
    assert "Tax Rate" in assumptions
    assert "Tax Shield Enabled" in assumptions
    assert "Gross Cash Taxes" in debt
    assert "Tax Shield" in debt
    assert "Levered Cash Taxes" in debt
    assert "Gross Cash Taxes" in operating
    assert any("Simplified tax shield uses cash interest x tax rate" in str(x) for x in audit)
