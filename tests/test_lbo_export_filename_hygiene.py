"""V4.1.0.1 Tiny Export Filename Hygiene.

Narrow patch: the LBO Excel download name must carry the `formula` marker so the
saved file matches the formula-native workbook. Covers (a) the API
Content-Disposition header and (b) the lbo.js download logic — it parses the
server filename first and falls back to a locally built name that also includes
`formula`. No engine / workbook / payload changes.
"""

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app import app
from tests.test_lbo_calculator_core import clean_anchor
from tests.test_lbo_multitranche_schedule import multi_tranche_anchor


STATIC = Path(__file__).resolve().parent.parent / "static" / "modeling"
JS = (STATIC / "js" / "lbo.js").read_text(encoding="utf-8")


def test_api_content_disposition_filename_includes_formula():
    client = app.test_client()
    resp = client.post("/api/modeling/lbo/excel", json=clean_anchor())
    assert resp.status_code == 200
    cd = resp.headers.get("Content-Disposition", "")
    assert "formula" in cd
    assert cd.endswith(".xlsx") or '.xlsx"' in cd
    assert "LBO_SYNTH_single_formula_" in cd
    # Body is still a real workbook (response unchanged).
    wb = load_workbook(BytesIO(resp.data))
    assert "Returns Summary" in wb.sheetnames


def test_api_multi_tranche_filename_includes_formula():
    client = app.test_client()
    resp = client.post("/api/modeling/lbo/excel", json=multi_tranche_anchor())
    assert resp.status_code == 200
    cd = resp.headers.get("Content-Disposition", "")
    assert "_multi_formula_" in cd


def test_js_parses_content_disposition_before_local_fallback():
    assert "filenameFromContentDisposition" in JS
    assert "Content-Disposition" in JS
    # Fallback name must also carry the formula marker.
    assert "LBO_${sym}_${mode}_formula_${today}.xlsx" in JS
    # Server filename is preferred over the locally built one.
    assert "filenameFromContentDisposition(res) ||" in JS


def test_app_no_longer_imports_unused_legacy_exporter():
    app_src = (Path(__file__).resolve().parent.parent / "app.py").read_text(encoding="utf-8")
    assert "from modeling.lbo_excel_exporter import generate_lbo_excel" not in app_src
    assert "from modeling.lbo_formula_workbook import generate_lbo_formula_excel" in app_src
