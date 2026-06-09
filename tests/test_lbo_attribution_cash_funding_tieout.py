"""V4.7.3.1 attribution cash funding tie-out tests.

This is a narrow bridge-component fix. It must not change run_lbo(), Sources &
Uses, the waterfall, or workbook architecture.
"""

from copy import deepcopy
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app import app
from modeling.lbo_attribution import build_lbo_attribution
from modeling.lbo_calculator import run_lbo
from modeling.lbo_excel_exporter import generate_lbo_excel
from tests.test_lbo_multitranche_schedule import (
    low_leverage_high_fcf_cash_build_fixture,
    multi_tranche_anchor,
)


STATIC = Path(__file__).resolve().parent.parent / "static" / "modeling"
JS = (STATIC / "js" / "lbo.js").read_text(encoding="utf-8")


def _components(attr):
    return {c["key"]: c for c in attr["components"]}


def _sheet_pairs(ws):
    pairs = {}
    for row in ws.iter_rows():
        if len(row) >= 2 and isinstance(row[0].value, str):
            pairs[row[0].value] = row[1].value
    return pairs


def _sheet_strings(ws):
    return {
        cell.value
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    }


def _cash_funding_case():
    payload = multi_tranche_anchor()
    payload["debt"] = {"cash_to_balance_sheet": 60.0}
    payload["capital_structure"]["cash_balance_beginning"] = 60.0
    payload["capital_structure"]["minimum_cash_balance"] = 80.0
    return payload


def test_initial_cash_funding_component_ties_cash_to_balance_sheet_case():
    payload = _cash_funding_case()
    result = run_lbo(payload)
    attr = build_lbo_attribution(payload, result)
    comps = _components(attr)

    assert comps["initial_cash_funding"]["label_en"] == "Initial Cash Funding"
    assert comps["initial_cash_funding"]["value"] == pytest.approx(-60.0)
    assert comps["initial_cash_funding"]["moic_contribution"] == pytest.approx(
        -60.0 / result["transaction_summary"]["sponsor_equity"]
    )
    assert comps["initial_cash_funding"]["direction"] == "negative"
    assert comps["ending_cash_balance"]["value"] == pytest.approx(80.0)
    assert attr["tie_out"]["residual"] == pytest.approx(0.0, abs=1e-6)
    assert attr["tie_out"]["mathematical_tie_out_pass"] is True
    assert attr["tie_out"]["directional_bridge_pass"] is True


def test_no_cash_case_has_no_initial_cash_funding_component_and_still_ties():
    payload = multi_tranche_anchor()
    result = run_lbo(payload)
    attr = build_lbo_attribution(payload, result)
    comps = _components(attr)

    assert "initial_cash_funding" not in comps
    assert attr["tie_out"]["residual"] == pytest.approx(0.0, abs=1e-6)
    assert attr["tie_out"]["mathematical_tie_out_pass"] is True


def test_api_and_ui_render_path_surface_initial_cash_funding_label():
    client = app.test_client()
    resp = client.post("/api/modeling/lbo", json=_cash_funding_case())
    data = resp.get_json()
    comps = _components(data["attribution"])

    assert resp.status_code == 200
    assert comps["initial_cash_funding"]["label_en"] == "Initial Cash Funding"
    assert "c.label_en" in JS
    assert "renderAttribution" in JS


def test_excel_returns_attribution_includes_initial_cash_funding_and_passes():
    payload = _cash_funding_case()
    result = run_lbo(payload)
    payload["attribution"] = build_lbo_attribution(payload, result)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, result).getvalue()), data_only=True)
    strings = _sheet_strings(wb["Returns Attribution"])
    pairs = _sheet_pairs(wb["Returns Attribution"])

    assert "Initial Cash Funding" in strings
    assert pairs["Residual"] == pytest.approx(0.0, abs=1e-6)
    assert pairs["Mathematical Tie-out Pass"] == "Yes"
    assert pairs["Directional Bridge Pass"] == "Yes"


def test_irr_moic_sources_uses_and_exit_bridge_unchanged_by_attribution_fix():
    payload = _cash_funding_case()
    result_before = run_lbo(deepcopy(payload))
    attr = build_lbo_attribution(payload, result_before)
    result_after = run_lbo(deepcopy(payload))
    payload["attribution"] = attr
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, result_before).getvalue()), data_only=True)
    su = _sheet_pairs(wb["Sources & Uses"])
    ex = result_before["exit"]

    assert result_after["returns"]["irr"] == pytest.approx(result_before["returns"]["irr"])
    assert result_after["returns"]["moic"] == pytest.approx(result_before["returns"]["moic"])
    assert su["Sources - Uses"] == pytest.approx(0.0)
    assert ex["exit_equity_value"] == pytest.approx(
        ex["exit_ev"] - ex["remaining_debt"] + ex["ending_cash_balance"]
    )


def test_existing_cash_build_case_still_ties_without_initial_cash_funding():
    payload = low_leverage_high_fcf_cash_build_fixture()
    result = run_lbo(payload)
    attr = build_lbo_attribution(payload, result)
    comps = _components(attr)

    assert "initial_cash_funding" not in comps
    assert comps["ending_cash_balance"]["value"] > 0
    assert attr["tie_out"]["residual"] == pytest.approx(0.0, abs=1e-6)
