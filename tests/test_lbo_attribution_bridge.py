"""V4.3 LBO returns attribution bridge tests.

These lock the equity-value-bridge math: residual ~0 in the closed V4.0
structure, MOIC contributions summing to MOIC-1, and non-tautological tie-out.
"""

import copy
import math
from io import BytesIO

from openpyxl import load_workbook

from app import app
from modeling.lbo_calculator import run_lbo
from modeling.lbo_attribution import build_lbo_attribution
from tests.test_lbo_calculator_core import clean_anchor


def _components(attr):
    return {c["key"]: c for c in attr["components"]}


def expansion_anchor():
    a = clean_anchor()
    a["transaction"]["exit_multiple"] = 12.0
    return a


def contraction_anchor():
    a = clean_anchor()
    a["transaction"]["exit_multiple"] = 8.0
    return a


def growth_anchor():
    a = clean_anchor()
    a["operating_forecast"]["ebitda"] = [1000, 1100, 1200, 1300, 1400]
    return a


def low_debt_full_repay_anchor():
    a = clean_anchor()
    a["debt"]["debt_amount"] = 100
    return a


# ---------------------------------------------------------------------------


def test_clean_attribution_status_ok():
    attr = build_lbo_attribution(clean_anchor(), run_lbo(clean_anchor()))
    assert attr["status"] == "ok"
    assert attr["method"] == "single_tranche_equity_value_bridge_v43"


def test_clean_closed_structure_residual_near_zero():
    res = run_lbo(clean_anchor())
    attr = build_lbo_attribution(clean_anchor(), res)
    assert abs(attr["tie_out"]["residual"]) < 1e-6


def test_mathematical_tie_out_pass_for_floating_residual():
    res = run_lbo(clean_anchor())
    attr = build_lbo_attribution(clean_anchor(), res)
    assert attr["tie_out"]["mathematical_tie_out_pass"] is True
    assert attr["tie_out"]["directional_bridge_pass"] is True


def test_moic_contributions_sum_to_moic_minus_one():
    res = run_lbo(clean_anchor())
    attr = build_lbo_attribution(clean_anchor(), res)
    total = sum(c["moic_contribution"] for c in attr["components"])
    moic = res["returns"]["moic"]
    assert math.isclose(total, moic - 1.0, rel_tol=0, abs_tol=1e-9)
    # Sanity: it must NOT equal headline MOIC.
    assert not math.isclose(total, moic, abs_tol=1e-6)


def test_ebitda_growth_plus_multiple_movement_ties_to_ev_change():
    res = run_lbo(growth_anchor())
    attr = build_lbo_attribution(growth_anchor(), res)
    comps = _components(attr)
    eg = comps["ebitda_growth"]["value"]
    mm = comps["multiple_movement"]["value"]
    entry_ev = res["transaction_summary"]["entry_ev"]
    exit_ev = res["exit"]["exit_ev"]
    assert math.isclose(eg + mm, exit_ev - entry_ev, abs_tol=1e-9)


def test_exit_multiple_equals_entry_multiple_zero_movement():
    res = run_lbo(clean_anchor())
    attr = build_lbo_attribution(clean_anchor(), res)
    comps = _components(attr)
    assert abs(comps["multiple_movement"]["value"]) < 1e-9
    assert comps["multiple_movement"]["direction"] == "neutral"


def test_exit_multiple_above_entry_positive_movement():
    res = run_lbo(expansion_anchor())
    attr = build_lbo_attribution(expansion_anchor(), res)
    comps = _components(attr)
    assert comps["multiple_movement"]["value"] > 0
    assert comps["multiple_movement"]["direction"] == "positive"


def test_exit_multiple_below_entry_negative_movement():
    res = run_lbo(contraction_anchor())
    attr = build_lbo_attribution(contraction_anchor(), res)
    comps = _components(attr)
    assert comps["multiple_movement"]["value"] < 0
    assert comps["multiple_movement"]["direction"] == "negative"


def test_deleveraging_positive_for_debt_paydown():
    res = run_lbo(clean_anchor())
    attr = build_lbo_attribution(clean_anchor(), res)
    comps = _components(attr)
    assert comps["deleveraging"]["value"] > 0
    assert comps["deleveraging"]["direction"] == "positive"


def test_deleveraging_ties_to_original_minus_remaining_debt():
    res = run_lbo(clean_anchor())
    attr = build_lbo_attribution(clean_anchor(), res)
    comps = _components(attr)
    original = res["transaction_summary"]["debt_amount"]
    remaining = res["exit"]["remaining_debt"]
    assert math.isclose(comps["deleveraging"]["value"], original - remaining, abs_tol=1e-9)


def test_fees_drag_negative():
    res = run_lbo(clean_anchor())
    attr = build_lbo_attribution(clean_anchor(), res)
    comps = _components(attr)
    assert comps["fees_drag"]["value"] < 0
    assert comps["fees_drag"]["direction"] == "negative"


def test_fees_drag_equals_negative_transaction_fees():
    res = run_lbo(clean_anchor())
    attr = build_lbo_attribution(clean_anchor(), res)
    comps = _components(attr)
    fees = res["transaction_summary"]["transaction_fees"]
    assert math.isclose(comps["fees_drag"]["value"], -fees, abs_tol=1e-9)


def test_residual_component_present_and_near_zero():
    res = run_lbo(clean_anchor())
    attr = build_lbo_attribution(clean_anchor(), res)
    comps = _components(attr)
    assert "residual" in comps
    assert abs(comps["residual"]["value"]) < 1e-6
    assert comps["residual"]["direction"] == "neutral"


def test_large_artificial_residual_triggers_flag():
    res = run_lbo(clean_anchor())
    doctored = copy.deepcopy(res)
    # Break internal consistency so the bridge can no longer explain creation.
    doctored["exit"]["exit_equity_value"] += 5000.0
    doctored["returns"]["exit_equity_value"] += 5000.0
    attr = build_lbo_attribution(clean_anchor(), doctored)
    codes = {f["code"] for f in attr["flags"]}
    assert "ATTRIBUTION_RESIDUAL_LARGE" in codes
    assert attr["tie_out"]["directional_bridge_pass"] is False


def test_fully_repaid_debt_excess_fcf_warns_not_silent():
    res = run_lbo(low_debt_full_repay_anchor())
    attr = build_lbo_attribution(low_debt_full_repay_anchor(), res)
    # Debt fully repaid → bridge still ties out but warns about retained cash.
    assert res["exit"]["remaining_debt"] <= 1e-6
    codes = {f["code"] for f in attr["flags"]}
    assert "ATTRIBUTION_RETAINED_CASH_NOT_MODELED" in codes


def test_lbo_error_result_attribution_unavailable():
    attr = build_lbo_attribution(clean_anchor(), {"status": "error", "returns": None})
    assert attr["status"] == "unavailable"
    codes = {f["code"] for f in attr["flags"]}
    assert "ATTRIBUTION_UNAVAILABLE_LBO_NOT_OK" in codes


def test_sponsor_equity_non_positive_unavailable():
    res = run_lbo(clean_anchor())
    doctored = copy.deepcopy(res)
    doctored["transaction_summary"]["sponsor_equity"] = 0.0
    doctored["returns"]["sponsor_equity"] = 0.0
    attr = build_lbo_attribution(clean_anchor(), doctored)
    assert attr["status"] == "unavailable"
    codes = {f["code"] for f in attr["flags"]}
    assert "ATTRIBUTION_UNAVAILABLE_SPONSOR_EQUITY_NON_POSITIVE" in codes


def test_tie_out_is_not_tautological():
    # Component sum INCLUDING residual is always == target by construction.
    # The real tie-out must rely on residual magnitude, not that identity.
    res = run_lbo(clean_anchor())
    doctored = copy.deepcopy(res)
    doctored["exit"]["exit_equity_value"] += 5000.0
    doctored["returns"]["exit_equity_value"] += 5000.0
    attr = build_lbo_attribution(clean_anchor(), doctored)
    comps = _components(attr)
    total_incl_residual = sum(c["value"] for c in attr["components"])
    target = attr["tie_out"]["target_equity_value_creation"]
    # Tautological identity holds...
    assert math.isclose(total_incl_residual, target, abs_tol=1e-6)
    # ...but tie-out still fails because residual is large.
    assert attr["tie_out"]["mathematical_tie_out_pass"] is False


def test_api_lbo_includes_attribution_on_success():
    client = app.test_client()
    resp = client.post("/api/modeling/lbo", json=clean_anchor())
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["status"] == "ok"
    assert "attribution" in data
    assert data["attribution"]["status"] == "ok"
    assert len(data["attribution"]["components"]) == 5


def test_api_lbo_attribution_unavailable_on_error():
    client = app.test_client()
    payload = clean_anchor()
    payload["debt"]["cash_sweep_pct"] = 0.5  # forces hard error
    resp = client.post("/api/modeling/lbo", json=payload)
    data = resp.get_json()
    assert data["status"] == "error"
    assert data["attribution"]["status"] == "unavailable"


def test_excel_returns_summary_has_attribution_bridge():
    client = app.test_client()
    resp = client.post("/api/modeling/lbo/excel", json=clean_anchor())
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.data), data_only=False)
    # V4.5: attribution bridge lives on its own Returns Attribution sheet.
    text = [cell.value for row in wb["Returns Attribution"].iter_rows() for cell in row if cell.value]
    assert "Equity Value Bridge" in text
    assert any(
        "Component MOIC contribution sums to MOIC - 1.0, not headline MOIC." in str(v)
        for v in text
    )
    # Canonical ASCII-hyphen wording only; unicode minus is no longer accepted.
    assert not any("MOIC − 1.0" in str(v) for v in text)


def test_excel_audit_has_attribution_tie_out_disclosure():
    client = app.test_client()
    resp = client.post("/api/modeling/lbo/excel", json=clean_anchor())
    wb = load_workbook(BytesIO(resp.data), data_only=False)
    # V4.5: attribution tie-out and disclosure live on Returns Attribution.
    text = [cell.value for row in wb["Returns Attribution"].iter_rows() for cell in row if cell.value]
    assert any("Attribution Method" in str(v) for v in text)
    assert any("Directional Bridge Pass" in str(v) for v in text)
    assert any("simplified equity value bridge" in str(v).lower() for v in text)
