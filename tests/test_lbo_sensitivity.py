import copy
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app import app
from modeling.lbo_calculator import run_lbo
from modeling.lbo_excel_exporter import SENSITIVITY_SHEET, SHEET_NAMES, generate_lbo_excel
from modeling.lbo_scenarios import generate_scenario_inputs
from modeling.lbo_sensitivity import (
    build_entry_exit_multiple_grid,
    build_lbo_sensitivity,
    build_leverage_exit_multiple_grid,
)
from tests.test_lbo_calculator_core import clean_anchor


ROOT = Path(__file__).resolve().parents[1]
HTML = (ROOT / "static" / "modeling" / "lbo.html").read_text(encoding="utf-8")
JS = (ROOT / "static" / "modeling" / "js" / "lbo.js").read_text(encoding="utf-8")


def _strings(wb, sheet):
    return [c.value for row in wb[sheet].iter_rows() for c in row if c.value is not None]


def _flat_cells(grid):
    return [cell for row in grid["cells"] for cell in row]


def test_helper_creates_entry_exit_grid():
    grid = build_entry_exit_multiple_grid(clean_anchor())
    assert grid["row_label"] == "Entry Multiple"
    assert grid["col_label"] == "Exit Multiple"
    assert len(grid["rows"]) == 5
    assert len(grid["cols"]) == 5
    assert len(_flat_cells(grid)) == 25


def test_helper_creates_leverage_exit_grid():
    grid = build_leverage_exit_multiple_grid(clean_anchor())
    assert grid["row_label"] == "Debt / EBITDA"
    assert grid["col_label"] == "Exit Multiple"
    assert grid["rows"][2] == 5.0
    assert len(_flat_cells(grid)) == 25


def test_base_cell_exists_and_is_highlighted():
    data = build_lbo_sensitivity(clean_anchor())
    cells = _flat_cells(data["grids"]["entry_exit"]) + _flat_cells(data["grids"]["leverage_exit"])
    base_cells = [c for c in cells if c["is_base"]]
    assert len(base_cells) == 2
    assert all(c["status"] == "ok" for c in base_cells)


def test_ok_cell_uses_run_lbo_output():
    payload = clean_anchor()
    data = build_lbo_sensitivity(payload)
    cell = next(c for c in _flat_cells(data["grids"]["entry_exit"]) if c["is_base"])
    direct = run_lbo(payload)
    assert cell["irr"] == direct["returns"]["irr"]
    assert cell["moic"] == direct["returns"]["moic"]


def test_unavailable_cells_do_not_break_grid(monkeypatch):
    import modeling.lbo_sensitivity as sens

    real = sens.run_lbo

    def fake(payload):
        if payload["transaction"]["entry_multiple"] < 9.5:
            return {"status": "error", "returns": None, "flags": [{"code": "TEST_UNAVAILABLE"}]}
        return real(payload)

    monkeypatch.setattr(sens, "run_lbo", fake)
    data = sens.build_lbo_sensitivity(clean_anchor())
    cells = _flat_cells(data["grids"]["entry_exit"])
    assert data["status"] == "ok"
    assert any(c["status"] == "unavailable" and c["error_code"] == "TEST_UNAVAILABLE" for c in cells)
    assert any(c["status"] == "ok" for c in cells)


def test_input_object_not_mutated():
    payload = clean_anchor()
    before = copy.deepcopy(payload)
    build_lbo_sensitivity(payload)
    assert payload == before


def test_sensitivity_endpoint_works():
    client = app.test_client()
    resp = client.post("/api/modeling/lbo/sensitivity", json={"inputs": clean_anchor()})
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["status"] == "ok"
    assert data["method"] == "lbo_sensitivity_v474"
    assert set(data["grids"]) == {"entry_exit", "leverage_exit"}


def test_ui_has_open_sensitivity_trigger_and_popover_strings():
    assert "Sensitivity Analysis" in HTML
    assert "Open Sensitivity" in HTML
    assert "sensitivity-overlay" in HTML
    assert "/api/modeling/lbo/sensitivity" in JS
    assert "renderSensitivityGrid" in JS


def test_saved_case_persists_and_restores_tax_fields():
    record_block = JS.split("function buildCaseRecord", 1)[1].split("function openSaveCaseModal", 1)[0]
    apply_block = JS.split("function applyCaseRecord", 1)[1].split("function extractManualForecast", 1)[0]
    assert "const payload = buildPayload();" in record_block
    assert "delete payload.scenarios" in record_block
    assert "delete payload.sensitivity" not in record_block
    assert "tax_rate: num(\"tax_rate\") / 100" in JS
    assert "tax_shield_enabled:" in JS
    assert "setValue(\"tax_rate\", payload.tax_rate === undefined ? 25.0" in apply_block
    assert "setValue(\"tax_shield_enabled\", payload.tax_shield_enabled === false ? \"disabled\" : \"enabled\")" in apply_block


def test_scenarios_preserve_base_tax_shield_setting():
    payload = clean_anchor()
    payload["tax_rate"] = 0.33
    payload["tax_shield_enabled"] = True
    scenarios = generate_scenario_inputs(payload)
    for case in scenarios.values():
        assert case["tax_rate"] == 0.33
        assert case["tax_shield_enabled"] is True


def test_sensitivity_cells_use_current_tax_shield_setting():
    disabled = clean_anchor()
    disabled["tax_shield_enabled"] = False
    enabled = clean_anchor()
    enabled["tax_shield_enabled"] = True
    enabled["tax_rate"] = 0.25
    disabled_base = next(c for c in _flat_cells(build_lbo_sensitivity(disabled)["grids"]["entry_exit"]) if c["is_base"])
    enabled_base = next(c for c in _flat_cells(build_lbo_sensitivity(enabled)["grids"]["entry_exit"]) if c["is_base"])
    assert enabled_base["irr"] >= disabled_base["irr"]
    assert enabled_base["moic"] >= disabled_base["moic"]


def test_input_edits_invalidate_sensitivity_before_export():
    assert "if (currentSensitivity) payload.sensitivity = currentSensitivity;" in JS
    assert "function invalidateSensitivity()" in JS
    assert "currentSensitivity = null;" in JS
    tranche_block = JS.split("function renderTrancheEditor", 1)[1].split("function buildCapitalStructure", 1)[0]
    assert "invalidateSensitivity()" in tranche_block
    dom_block = JS.split("document.addEventListener(\"DOMContentLoaded\"", 1)[1]
    assert "currencySel.addEventListener(\"change\", invalidateSensitivity)" in dom_block


def test_excel_without_sensitivity_keeps_old_sheet_order():
    payload = clean_anchor()
    out = run_lbo(payload)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, out).getvalue()), data_only=False)
    assert wb.sheetnames == SHEET_NAMES
    assert SENSITIVITY_SHEET not in wb.sheetnames


def test_excel_with_sensitivity_includes_sensitivity_analysis_sheet():
    payload = clean_anchor()
    out = run_lbo(payload)
    payload["sensitivity"] = build_lbo_sensitivity(payload)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, out).getvalue()), data_only=False)
    assert SENSITIVITY_SHEET in wb.sheetnames
    strings = _strings(wb, SENSITIVITY_SHEET)
    assert "Entry Multiple x Exit Multiple" in strings
    assert "Debt / EBITDA x Exit Multiple" in strings
    assert any("not probability-weighted" in str(x) for x in strings)


def test_sensitivity_copy_contains_no_recommendation_wording():
    text = "\n".join(
        line for line in (HTML + "\n" + JS).splitlines()
        if "sensitivity" in line.lower() or "Sensitivity" in line
    )
    assert "recommendation" not in text.lower()
