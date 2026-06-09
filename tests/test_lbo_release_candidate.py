"""V4.7 LBO Release Candidate / Gold Master QA tests.

This suite is the封版 (release-candidate) gate for the V4 LBO module. It does not
introduce or exercise any new engine math. It consumes Python output only and
verifies:

* Synthetic clean + multi-tranche gold masters export a workbook.
* Scenario Summary appears only when scenario data is provided.
* The V4.7 lightweight return context behaves as a single compact badge, fires
  for low-return / deleveraging-led cases, stays silent for healthy returns, and
  never contains recommendation wording.
* Canonical ``MOIC - 1.0`` wording uses an exact ASCII hyphen everywhere.

``run_lbo()`` remains the IRR / MOIC source of truth; ``build_lbo_attribution()``
remains the attribution source of truth; this file only inspects their output.
"""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from modeling.lbo_calculator import run_lbo
from modeling.lbo_attribution import (
    build_lbo_attribution,
    MOIC_CONTRIBUTION_NOTE_EN,
)
from modeling.lbo_return_context import (
    build_return_context,
    LOW_RETURN_CONTEXT,
    DEBT_PAYDOWN_DRIVEN_CONTEXT,
    DELEVERAGING_LED_CONTEXT,
    LOW_RETURN_IRR_THRESHOLD,
)
from modeling.lbo_excel_exporter import (
    generate_lbo_excel,
    SCENARIO_SUMMARY_SHEET,
)
from modeling.lbo_scenarios import build_lbo_scenarios
from tests.test_lbo_calculator_core import clean_anchor
from tests.test_lbo_multitranche_schedule import multi_tranche_anchor


# ── fixtures / helpers ───────────────────────────────────────────────────────

# Affirmative recommendation / deal-conclusion vocabulary that must never appear
# in a return-context badge or anywhere in the workbook. Note: disclaimer-style
# negations like "not an investment or acquisition recommendation" are required
# wording and are intentionally NOT in this list, so they are allowed.
RECOMMENDATION_TERMS = [
    "should acquire", "should buy", "good deal", "bad deal", "attractive deal",
    "buy signal", "sell signal", "hurdle rate", "market norm", "recommended deal",
    "建议收购", "建议买入", "好的交易", "差的交易",
]


def high_return_anchor():
    """A healthy single-tranche case: EBITDA growth + multiple expansion drive
    IRR comfortably above the low-return threshold, so no return context fires."""
    return {
        "symbol": "GROWTHCO",
        "currency": "USD",
        "transaction": {
            "entry_ebitda": 1000,
            "entry_multiple": 8.0,
            "exit_multiple": 11.0,
            "exit_year": 5,
            "transaction_fees_pct_ev": 0.02,
        },
        "operating_forecast": {
            "years": [1, 2, 3, 4, 5],
            "revenue": [5000, 5500, 6050, 6655, 7320],
            "ebitda": [1100, 1300, 1500, 1700, 1900],
            "cash_taxes": [100, 110, 120, 130, 140],
            "capex": [120, 120, 120, 120, 120],
            "change_in_nwc": [20, 20, 20, 20, 20],
        },
        "debt": {
            "debt_amount": 4000,
            "interest_rate": 0.08,
            "mandatory_amortization_pct": 0.01,
            "cash_sweep_pct": 1.0,
            "cash_to_balance_sheet": 0.0,
        },
    }


def no_context_anchor():
    """A healthy multiple-expansion-dominant case where no single return-context
    trigger fires (IRR high, growth not dominant, deleveraging neither dominant
    nor limited, no revolver)."""
    return {
        "symbol": "MEXP",
        "currency": "USD",
        "transaction": {
            "entry_ebitda": 1000,
            "entry_multiple": 8.0,
            "exit_multiple": 13.0,
            "exit_year": 5,
            "transaction_fees_pct_ev": 0.02,
        },
        "operating_forecast": {
            "years": [1, 2, 3, 4, 5],
            "revenue": [5000, 5000, 5000, 5000, 5000],
            "ebitda": [1000, 1010, 1020, 1030, 1050],
            "cash_taxes": [100, 100, 100, 100, 100],
            "capex": [120, 120, 120, 120, 120],
            "change_in_nwc": [20, 20, 20, 20, 20],
        },
        "debt": {
            "debt_amount": 4000,
            "interest_rate": 0.08,
            "mandatory_amortization_pct": 0.01,
            "cash_sweep_pct": 1.0,
            "cash_to_balance_sheet": 0.0,
        },
    }


def _run_with_attr(anchor):
    payload = anchor()
    result = run_lbo(payload)
    payload["attribution"] = build_lbo_attribution(payload, result)
    return payload, result


def _workbook(anchor):
    payload, result = _run_with_attr(anchor)
    data = generate_lbo_excel(payload, result).getvalue()
    return load_workbook(BytesIO(data), data_only=False)


def _all_strings(wb):
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    out.append(cell.value)
    return out


# ── 1-2. gold master workbook export ─────────────────────────────────────────


@pytest.mark.parametrize("anchor", [clean_anchor, multi_tranche_anchor])
def test_gold_master_exports_workbook(anchor):
    wb = _workbook(anchor)
    assert wb.worksheets
    # Cover always present; full path produces a valid workbook.
    assert "Cover" in wb.sheetnames
    assert "Returns Summary" in wb.sheetnames


def test_synthetic_clean_gold_master_residual_near_zero():
    payload, result = _run_with_attr(clean_anchor)
    attribution = payload["attribution"]
    assert attribution["status"] == "ok"
    assert attribution["tie_out"]["mathematical_tie_out_pass"] is True


# ── 3. Scenario Summary appears only when scenarios provided ──────────────────


def test_scenario_summary_absent_without_scenarios():
    wb = _workbook(clean_anchor)
    assert SCENARIO_SUMMARY_SHEET not in wb.sheetnames


def test_scenario_summary_present_with_scenarios():
    payload, result = _run_with_attr(multi_tranche_anchor)
    scenarios = build_lbo_scenarios(multi_tranche_anchor())
    payload["scenarios"] = scenarios
    data = generate_lbo_excel(payload, result).getvalue()
    wb = load_workbook(BytesIO(data), data_only=False)
    assert SCENARIO_SUMMARY_SHEET in wb.sheetnames


# ── 4. Return context fires for low-return / deleveraging-led case ────────────


def test_return_context_fires_for_deleveraging_led_case():
    payload, result = _run_with_attr(clean_anchor)
    context = build_return_context(result, payload["attribution"])
    assert context["status"] == "ok"
    # Flat operating + paydown-only return triggers all three contexts.
    assert LOW_RETURN_CONTEXT in context["codes"]
    assert DEBT_PAYDOWN_DRIVEN_CONTEXT in context["codes"]
    assert DELEVERAGING_LED_CONTEXT in context["codes"]


def test_return_context_silent_for_healthy_return():
    payload, result = _run_with_attr(high_return_anchor)
    assert result["returns"]["irr"] >= LOW_RETURN_IRR_THRESHOLD
    context = build_return_context(result, payload["attribution"])
    # Growth + multiple expansion => no low-return, no debt-paydown-driven badge.
    assert LOW_RETURN_CONTEXT not in context["codes"]
    assert DEBT_PAYDOWN_DRIVEN_CONTEXT not in context["codes"]


# ── 5. Return context is one-line / compact, not a long disclosure ────────────


def test_return_context_badge_is_one_line_and_compact():
    payload, result = _run_with_attr(clean_anchor)
    context = build_return_context(result, payload["attribution"])
    for key in ("badge_en", "badge_cn"):
        text = context[key]
        assert text, f"{key} should be populated when a context fires"
        assert "\n" not in text, f"{key} must be a single line"
        # Compact: a badge, not a paragraph / disclosure layer.
        assert len(text) <= 120, f"{key} should stay compact, got {len(text)} chars"


def test_return_context_severity_is_context_not_warning():
    payload, result = _run_with_attr(clean_anchor)
    context = build_return_context(result, payload["attribution"])
    # It is context, never a warning, error, or suitability gate.
    assert context["severity"] == "context"


def test_return_context_none_when_lbo_not_ok():
    context = build_return_context({"status": "error"}, None)
    assert context["status"] == "none"
    assert context["codes"] == []
    assert context["badge_en"] == ""


# ── 6. LOW_RETURN_CONTEXT appears when IRR < threshold ────────────────────────


def test_low_return_context_keyed_to_irr_threshold():
    payload, result = _run_with_attr(clean_anchor)
    irr = result["returns"]["irr"]
    assert irr < LOW_RETURN_IRR_THRESHOLD
    context = build_return_context(result, payload["attribution"])
    assert context["codes"][0] == LOW_RETURN_CONTEXT  # priority = primary badge


# ── 7. No recommendation wording anywhere in the badge or workbook ────────────


@pytest.mark.parametrize("code", [
    LOW_RETURN_CONTEXT, DEBT_PAYDOWN_DRIVEN_CONTEXT, DELEVERAGING_LED_CONTEXT,
])
def test_return_context_has_no_recommendation_wording(code):
    from modeling.lbo_return_context import _BADGES
    text = (_BADGES[code]["en"] + " " + _BADGES[code]["cn"]).lower()
    for term in RECOMMENDATION_TERMS:
        assert term.lower() not in text, f"badge {code} contains '{term}'"


@pytest.mark.parametrize("anchor", [clean_anchor, multi_tranche_anchor])
def test_workbook_has_no_recommendation_wording(anchor):
    wb = _workbook(anchor)
    blob = "\n".join(_all_strings(wb)).lower()
    for term in RECOMMENDATION_TERMS:
        assert term.lower() not in blob, f"workbook contains recommendation term '{term}'"


# ── 8. Exact `MOIC - 1.0` string (ASCII hyphen, no unicode minus) ─────────────


def test_canonical_moic_note_uses_ascii_hyphen():
    assert MOIC_CONTRIBUTION_NOTE_EN == (
        "Component MOIC contribution sums to MOIC - 1.0, not headline MOIC."
    )
    # Reject any unicode minus fuzzy fallback.
    assert "−" not in MOIC_CONTRIBUTION_NOTE_EN
    assert " - 1.0" in MOIC_CONTRIBUTION_NOTE_EN


def test_workbook_contains_exact_moic_minus_one_string():
    wb = _workbook(clean_anchor)
    blob = "\n".join(_all_strings(wb))
    assert "MOIC - 1.0" in blob
    # No unicode minus variant leaked into the workbook.
    assert "MOIC − 1.0" not in blob


# ── 9. Excel Returns Summary surfaces the badge only when triggered ───────────


def test_excel_returns_summary_shows_badge_for_low_return():
    wb = _workbook(clean_anchor)
    ws = wb["Returns Summary"]
    strings = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any("Return Context" == s for s in strings)
    assert any("Low return context" in s for s in strings)


def test_excel_returns_summary_omits_badge_for_healthy_return():
    # A balanced multiple-expansion case triggers no return-context badge.
    wb = _workbook(no_context_anchor)
    ws = wb["Returns Summary"]
    strings = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert not any("Return Context" == s for s in strings)
