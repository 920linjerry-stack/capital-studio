"""V4.1.4 Final Banker Formatting / Release Hardening contract tests.

These lock the V4.1.4 deliverable contract for the formula-native LBO workbook:

  * the workbook version string is V4.1.4 and the stale V4.1.2 version string is
    gone, with the Cover / Audit / Model Map disclosures consistent;
  * Scenario / Sensitivity stay honestly disclosed as Python-precomputed,
    oracle-backed grids — *not* live Excel Data Tables;
  * banker formatting is present (frozen panes, non-default column widths,
    number formats on headline outputs, fit-to-width print setup, the headline
    output fill);
  * the formula chain has not regressed (formula counts above the V4.1.3
    baseline, key formulas and named ranges still present) and the legacy
    values-only path stays 0-formula.

V4.1.4 is a formatting / disclosure hardening pass only: it must not change the
engine math, so the existing oracle tie-out tests (test_lbo_formula_workbook,
test_lbo_v413_scenario_sensitivity_formulas) carry the numeric regression and
are not duplicated here.
"""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from modeling.lbo_calculator import run_lbo
from modeling.lbo_scenarios import build_lbo_scenarios
from modeling.lbo_sensitivity import build_lbo_sensitivity
from modeling.lbo_excel_exporter import (
    FMT_MULT2,
    FMT_PCT,
    SCENARIO_SUMMARY_SHEET,
    SENSITIVITY_SHEET,
    generate_lbo_values_only_excel,
)
from modeling.lbo_formula_workbook import (
    FORMULA_WORKBOOK_VERSION,
    generate_lbo_formula_excel,
)
from tests.test_lbo_calculator_core import clean_anchor
from tests.test_lbo_multitranche_schedule import multi_tranche_anchor
from tests.test_lbo_formula_workbook import _formula_count, _row_addr


# Stale version string retired in V4.1.4 (the Cover used to render this).
_OLD_VERSION = "V4.1.2 Covenant / Maturity Wall / IC Defense Formula Workbook"


def _wb(anchor, *, layers=True):
    payload = anchor()
    result = run_lbo(payload)
    if layers:
        payload["scenarios"] = build_lbo_scenarios(anchor())
        payload["sensitivity"] = build_lbo_sensitivity(payload)
    wb = load_workbook(
        BytesIO(generate_lbo_formula_excel(payload, result).getvalue()), data_only=False
    )
    return wb, payload, result


def _strings(ws):
    return [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]


def _blob(ws):
    return " ".join(_strings(ws))


# ── 1. Version / disclosure consistency ──────────────────────────────────────


def test_version_constant_is_v414_not_stale():
    assert "V4.1.4" in FORMULA_WORKBOOK_VERSION
    assert FORMULA_WORKBOOK_VERSION != _OLD_VERSION


@pytest.mark.parametrize("anchor", [clean_anchor, multi_tranche_anchor])
def test_cover_shows_current_version_only(anchor):
    wb, _, _ = _wb(anchor)
    cover = _strings(wb["Cover"])
    assert FORMULA_WORKBOOK_VERSION in cover
    # The stale full version string must not appear anywhere on the Cover.
    assert _OLD_VERSION not in " ".join(cover)
    # Core cover metadata present for a deliverable model.
    blob = " ".join(cover)
    for field in ("Symbol", "Currency", "Forecast Source", "Engine Output Mode",
                  "Generated", "Headline IRR", "Headline MOIC"):
        assert field in blob


def test_old_version_string_absent_workbook_wide():
    wb, _, _ = _wb(multi_tranche_anchor)
    for sheet in ("Cover", "Model Map", "Audit & Disclosures"):
        assert _OLD_VERSION not in _blob(wb[sheet])


def test_scenario_sensitivity_disclosed_as_not_live_data_tables():
    wb, _, _ = _wb(multi_tranche_anchor)
    audit = _blob(wb["Audit & Disclosures"])
    assert "Scenario Summary and Sensitivity Analysis remain Python-precomputed" in audit
    assert "not live Excel Data Tables" in audit
    # The Sensitivity sheet itself repeats the honest disclosure.
    assert "not a live Excel Data Table" in _blob(wb[SENSITIVITY_SHEET])


def test_audit_discloses_v414_formatting_only_pass():
    wb, _, _ = _wb(multi_tranche_anchor)
    audit = _blob(wb["Audit & Disclosures"])
    assert "V4.1.4" in audit
    # Honest framing: formatting/disclosure only, engine truth unchanged.
    assert "no engine math" in audit
    assert "run_lbo()" in audit


# ── 2. Formatting contract ───────────────────────────────────────────────────

_FROZEN_SHEETS = [
    "Assumptions", "Operating Forecast", "Debt Schedule",
    SCENARIO_SUMMARY_SHEET, SENSITIVITY_SHEET,
]


@pytest.mark.parametrize("sheet", _FROZEN_SHEETS)
def test_core_sheets_have_freeze_panes(sheet):
    wb, _, _ = _wb(multi_tranche_anchor)
    assert wb[sheet].freeze_panes, f"{sheet} must freeze panes"


@pytest.mark.parametrize("sheet", ["Assumptions", "Sources & Uses", "Debt Schedule",
                                   "Returns Summary", "Covenant Check"])
def test_core_sheet_widths_not_default(sheet):
    wb, _, _ = _wb(multi_tranche_anchor)
    assert wb[sheet].column_dimensions["A"].width, f"{sheet} col A width must be set"


def test_headline_output_cells_have_number_format_and_fill():
    wb, _, _ = _wb(clean_anchor)
    ws = wb["Returns Summary"]
    moic = _row_addr(ws, "MOIC")
    irr = _row_addr(ws, "IRR (annualized)")
    assert moic.number_format == FMT_MULT2
    assert irr.number_format == FMT_PCT
    # The peach headline-output fill marks the output region.
    assert moic.fill.fgColor.rgb.endswith("FCE4D6")
    assert irr.fill.fgColor.rgb.endswith("FCE4D6")


@pytest.mark.parametrize("sheet", ["Debt Schedule", "Operating Forecast", "Covenant Check"])
def test_wide_sheets_have_fit_to_width_print_setup(sheet):
    wb, _, _ = _wb(multi_tranche_anchor)
    ps = wb[sheet].page_setup
    assert ps.fitToWidth == 1
    assert ps.orientation == "landscape"


def test_narrow_sheets_print_portrait_fit_width():
    wb, _, _ = _wb(clean_anchor)
    ps = wb["Returns Summary"].page_setup
    assert ps.fitToWidth == 1
    assert ps.orientation == "portrait"


# ── 3. Formula integrity / named ranges / legacy path ────────────────────────


def test_formula_count_not_regressed_below_v413_baseline():
    """V4.1.4 is formatting-only; the formula chain must not shrink. Thresholds
    sit comfortably below the observed V4.1.3 baselines (single ~194, multi ~433)
    so a real regression trips, but a small future tweak does not."""
    single, _, _ = _wb(clean_anchor)
    multi, _, _ = _wb(multi_tranche_anchor)
    assert _formula_count(single) > 150
    assert _formula_count(multi) > 350


def test_key_named_ranges_present():
    wb, payload, _ = _wb(multi_tranche_anchor)
    names = set(wb.defined_names)
    for nm in ("LBO_Total_Ending_Debt", "LBO_Ending_Cash", "LBO_MOIC",
               "LBO_Sponsor_Equity", "LBO_Exit_Equity", "LBO_Max_Leverage",
               "LBO_Min_Coverage", "LBO_Min_Cash", "LBO_Cash_Sweep_On"):
        assert nm in names, f"missing named range {nm}"


def test_key_defense_layer_formulas_present():
    """Debt Schedule, Covenant Check, Maturity Wall, Attribution, Scenario and
    Sensitivity still carry their core formulas after the formatting pass."""
    wb, _, _ = _wb(multi_tranche_anchor)

    def has_formula(sheet, needle):
        return any(
            isinstance(c.value, str) and c.value.startswith("=") and needle in c.value
            for row in wb[sheet].iter_rows() for c in row
        )

    assert has_formula("Debt Schedule", "LBO_Tr0_Opening")
    assert has_formula("Covenant Check", "LBO_Max_Leverage")
    assert has_formula("Maturity Wall", "LBO_Total_Ending_Debt")
    assert has_formula("Returns Attribution", "LBO_Exit_Equity")
    # Scenario tie-out and sensitivity grid checks resolve to OK/CHECK formulas.
    assert any('"OK","CHECK"' in c.value
               for row in wb[SCENARIO_SUMMARY_SHEET].iter_rows() for c in row
               if isinstance(c.value, str))
    assert any('"OK","CHECK"' in c.value
               for row in wb[SENSITIVITY_SHEET].iter_rows() for c in row
               if isinstance(c.value, str))


def test_legacy_values_only_path_stays_zero_formula():
    payload = multi_tranche_anchor()
    result = run_lbo(payload)
    payload["scenarios"] = build_lbo_scenarios(multi_tranche_anchor())
    payload["sensitivity"] = build_lbo_sensitivity(payload)
    wb = load_workbook(
        BytesIO(generate_lbo_values_only_excel(payload, result).getvalue()),
        data_only=False,
    )
    assert _formula_count(wb) == 0


def test_sheet_order_and_key_sheets_present():
    wb, _, _ = _wb(clean_anchor)
    for sheet in ("Cover", "Model Map", SCENARIO_SUMMARY_SHEET, SENSITIVITY_SHEET,
                  "Assumptions", "Sources & Uses", "Operating Forecast",
                  "Debt Schedule", "Covenant Check", "Returns Summary",
                  "Returns Attribution", "Maturity Wall", "Audit & Disclosures"):
        assert sheet in wb.sheetnames
    # Cover / Model Map lead the workbook.
    assert wb.sheetnames[0] == "Cover"
    assert wb.sheetnames[1] == "Model Map"
