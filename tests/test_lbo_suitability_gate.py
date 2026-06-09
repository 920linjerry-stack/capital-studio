"""V4.2 LBO suitability gate tests.

Synthetic fixtures lock veto / borderline / suitable behavior. These tests do
not exercise live data fetches; the API contract test only checks that the
suitability metadata is plumbed through the defaults endpoint.
"""

from io import BytesIO

from openpyxl import load_workbook

from app import app
from modeling.lbo_defaults import build_lbo_defaults
from modeling.lbo_suitability import assess_lbo_suitability
from tests.test_lbo_calculator_core import clean_anchor
from tests.test_lbo_defaults_builder import (
    clean_synthetic_anchor,
    thin_coverage_synthetic_anchor,
    distressed_synthetic_anchor,
)


def aapl_like_anchor():
    raw = clean_synthetic_anchor()
    raw["symbol"] = "AAPL"
    raw["currency"] = "USD"
    raw["entry_ebitda"] = 130_000
    raw["entry_multiple"] = 25.0
    raw["market_cap"] = 3_000_000
    raw["operating_forecast"]["ebitda"] = [130_000, 132_000, 134_000, 136_000, 138_000]
    raw["operating_forecast"]["revenue"] = [400_000] * 5
    raw["operating_forecast"]["cash_taxes"] = [13_000] * 5
    raw["operating_forecast"]["capex"] = [12_000] * 5
    raw["operating_forecast"]["change_in_nwc"] = [1_000] * 5
    return raw


def sponsor_equity_too_large_anchor():
    raw = clean_synthetic_anchor()
    raw["symbol"] = "SEQTL"
    raw["currency"] = "USD"
    raw["entry_ebitda"] = 9_000
    raw["entry_multiple"] = 10.0
    raw["operating_forecast"]["ebitda"] = [9_000, 9_100, 9_200, 9_300, 9_400]
    raw["operating_forecast"]["revenue"] = [45_000] * 5
    raw["operating_forecast"]["cash_taxes"] = [900] * 5
    raw["operating_forecast"]["capex"] = [900] * 5
    raw["operating_forecast"]["change_in_nwc"] = [90] * 5
    return raw


def net_cash_anchor():
    raw = clean_synthetic_anchor()
    raw["symbol"] = "NETC"
    raw["net_debt"] = -2_000
    raw["net_cash"] = 2_000
    return raw


def high_multiple_anchor(multiple: float):
    raw = clean_synthetic_anchor()
    raw["entry_multiple"] = multiple
    return raw


def placeholder_multiple_anchor():
    raw = clean_synthetic_anchor()
    raw.pop("entry_multiple")
    return raw


def missing_market_size_anchor():
    raw = clean_synthetic_anchor()
    raw["currency"] = "ZZZ"  # FX unavailable
    return raw


def _assess(symbol, raw):
    defaults_result = build_lbo_defaults(symbol, raw)
    return defaults_result, assess_lbo_suitability(symbol, defaults_result, raw)


def codes(reasons):
    return {r["code"] for r in reasons}


# ---------------------------------------------------------------------------


def test_clean_synthetic_anchor_is_suitable_without_veto():
    _, suit = _assess("SYNTH", clean_synthetic_anchor())
    assert suit["suitability"] == "suitable"
    assert suit["veto_triggered"] is False
    assert suit["score"] is not None and suit["score"] >= 70
    assert codes(suit["veto_reasons"]) == set()
    pf = codes(suit["positive_factors"])
    assert {"POSITIVE_EBITDA", "SERVICEABILITY_PASS", "DSCR_ABOVE_THRESHOLD"} <= pf
    assert "投资建议" in suit["disclosure_cn"]


def test_thin_coverage_anchor_is_borderline_with_leverage_haircut():
    _, suit = _assess("THIN", thin_coverage_synthetic_anchor())
    assert suit["suitability"] == "borderline"
    assert suit["veto_triggered"] is False
    assert "LEVERAGE_HAIRCUT_APPLIED" in codes(suit["penalty_reasons"])
    assert suit["score"] is not None and suit["score"] < 70


def test_distressed_anchor_is_unsuitable_veto():
    _, suit = _assess("DIST", distressed_synthetic_anchor())
    assert suit["suitability"] == "unsuitable"
    assert suit["veto_triggered"] is True
    assert suit["score"] is None
    assert "NOT_SERVICEABLE_EVEN_AT_MIN_LEVERAGE" in codes(suit["veto_reasons"])
    assert suit["recommended_next_view"] is None
    assert "going-concern" in suit["modeling_guidance_cn"]


def test_aapl_like_megacap_is_unsuitable_veto():
    _, suit = _assess("AAPL", aapl_like_anchor())
    assert suit["suitability"] == "unsuitable"
    assert suit["veto_triggered"] is True
    veto_codes = codes(suit["veto_reasons"])
    assert "MEGA_CAP_FINANCING_FEASIBILITY" in veto_codes \
        or "SPONSOR_EQUITY_CHECK_TOO_LARGE" in veto_codes
    assert suit["recommended_next_view"] == "DCF"


def test_aapl_like_raw_market_cap_usd_abs_triggers_megacap_veto():
    raw = clean_synthetic_anchor()
    raw["symbol"] = "AAPLRAW"
    raw["currency"] = "USD"
    raw["market_cap"] = 3_000_000_000_000
    raw["market_cap_unit"] = "actual"
    _, suit = _assess("AAPLRAW", raw)
    assert suit["suitability"] == "unsuitable"
    assert suit["veto_triggered"] is True
    assert "MEGA_CAP_FINANCING_FEASIBILITY" in codes(suit["veto_reasons"])
    assert suit["recommended_next_view"] == "DCF"


def test_aapl_like_market_cap_usd_millions_triggers_megacap_veto():
    raw = clean_synthetic_anchor()
    raw["symbol"] = "AAPLMN"
    raw["currency"] = "USD"
    raw["market_cap"] = 3_000_000
    raw["market_cap_unit"] = "millions"
    _, suit = _assess("AAPLMN", raw)
    assert suit["suitability"] == "unsuitable"
    assert suit["veto_triggered"] is True
    assert "MEGA_CAP_FINANCING_FEASIBILITY" in codes(suit["veto_reasons"])
    assert suit["recommended_next_view"] == "DCF"


def test_sponsor_equity_too_large_triggers_veto():
    _, suit = _assess("SEQTL", sponsor_equity_too_large_anchor())
    assert suit["suitability"] == "unsuitable"
    assert suit["veto_triggered"] is True
    assert "SPONSOR_EQUITY_CHECK_TOO_LARGE" in codes(suit["veto_reasons"])


def test_net_cash_anchor_includes_net_cash_structure_review():
    _, suit = _assess("NETC", net_cash_anchor())
    assert "NET_CASH_STRUCTURE_REVIEW" in codes(suit["penalty_reasons"])
    assert suit["recommended_next_view"] == "DCF"


def test_high_multiple_band_penalty_20_to_30():
    _, suit = _assess("HM", high_multiple_anchor(25.0))
    assert "HIGH_ENTRY_MULTIPLE_LBO_MATH_WEAK" in codes(suit["penalty_reasons"])
    assert "VERY_HIGH_ENTRY_MULTIPLE_LBO_MATH_WEAK" not in codes(suit["penalty_reasons"])


def test_very_high_multiple_band_penalty_above_30():
    _, suit = _assess("VHM", high_multiple_anchor(35.0))
    assert "VERY_HIGH_ENTRY_MULTIPLE_LBO_MATH_WEAK" in codes(suit["penalty_reasons"])


def test_missing_market_size_is_soft_flag_not_hard_error():
    _, suit = _assess("UNK", missing_market_size_anchor())
    assert suit["status"] == "ok"
    flag_codes = {f["code"] for f in suit["flags"]}
    assert "SIZE_USD_CONVERSION_UNAVAILABLE" in flag_codes
    # FX unavailable should not auto-veto mega-cap.
    assert "MEGA_CAP_FINANCING_FEASIBILITY" not in codes(suit["veto_reasons"])


def test_unsupported_fx_does_not_false_veto_raw_megacap():
    raw = clean_synthetic_anchor()
    raw["currency"] = "ZZZ"
    raw["market_cap"] = 3_000_000_000_000
    raw["market_cap_unit"] = "actual"
    _, suit = _assess("UNKFX", raw)
    flag_codes = {f["code"] for f in suit["flags"]}
    assert "SIZE_USD_CONVERSION_UNAVAILABLE" in flag_codes
    assert "MEGA_CAP_FINANCING_FEASIBILITY" not in codes(suit["veto_reasons"])


def test_entry_multiple_placeholder_low_severity_penalty():
    _, suit = _assess("PH", placeholder_multiple_anchor())
    assert "ENTRY_MULTIPLE_PLACEHOLDER_USED" in codes(suit["penalty_reasons"])


def test_positive_factors_present_for_clean_anchor():
    _, suit = _assess("SYNTH", clean_synthetic_anchor())
    pf = codes(suit["positive_factors"])
    assert "POSITIVE_EBITDA" in pf
    assert "POSITIVE_CASH_FLOW" in pf


def test_recommended_next_view_rules():
    _, mega = _assess("AAPL", aapl_like_anchor())
    assert mega["recommended_next_view"] == "DCF"

    _, dist = _assess("DIST", distressed_synthetic_anchor())
    assert dist["recommended_next_view"] is None

    _, net_cash = _assess("NETC", net_cash_anchor())
    assert net_cash["recommended_next_view"] == "DCF"


def test_api_defaults_response_contains_suitability():
    client = app.test_client()
    resp = client.get("/api/modeling/lbo/defaults?symbol=SYNTH")
    assert resp.status_code == 200
    data = resp.get_json()
    assert "suitability" in data
    suitability = data["suitability"]
    assert suitability.get("status") == "ok"
    assert suitability.get("suitability") in {"suitable", "borderline", "unsuitable"}
    assert "disclosure_en" in suitability
    assert isinstance(suitability.get("veto_reasons"), list)


def test_excel_audit_contains_lbo_suitability_gate_section():
    client = app.test_client()
    defaults = client.get("/api/modeling/lbo/defaults?symbol=SYNTH").get_json()["defaults"]
    payload = clean_anchor()
    payload["default_builder"] = defaults["default_builder"]
    payload["suitability"] = defaults["suitability"]
    resp = client.post("/api/modeling/lbo/excel", json=payload)
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.data), data_only=False)
    # V4.5: suitability status on Assumptions, disclosure on Audit & Disclosures.
    assumptions = [c.value for row in wb["Assumptions"].iter_rows() for c in row if c.value]
    assert "Suitability Status" in assumptions
    audit = [c.value for row in wb["Audit & Disclosures"].iter_rows() for c in row if c.value]
    assert "Suitability Disclosure" in audit
    assert any("Suitability gate assesses" in str(v) for v in audit)


def test_suitability_does_not_block_run_lbo_for_unsuitable_targets():
    from modeling.lbo_calculator import run_lbo
    payload = clean_anchor()
    payload["symbol"] = "AAPL"
    payload["transaction"]["entry_ebitda"] = 130_000
    payload["transaction"]["entry_multiple"] = 25.0
    payload["operating_forecast"]["ebitda"] = [130_000] * 5
    result = run_lbo(payload)
    assert result["status"] == "ok"
    assert result["returns"]["irr"] is not None
