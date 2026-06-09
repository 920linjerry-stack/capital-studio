"""V4.7 RC Input UX Hotfix tests.

Covers the input-layer UX fixes only: currency dropdown + normalization, amount
unit labels, debt-sizing relationship (By Leverage Multiple vs Manual), and the
user-readable DEBT_EXCEEDS_USES messaging. No engine math is exercised or
changed here -- run_lbo / waterfall / covenant / attribution / scenario logic is
untouched; these tests verify the surrounding UX layer and defensive
normalization only.
"""

import re
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app import app, _normalize_lbo_currency, _normalize_payload_currency
from modeling.lbo_calculator import run_lbo
from modeling.lbo_excel_exporter import generate_lbo_excel
from tests.test_lbo_calculator_core import clean_anchor


STATIC = Path(__file__).resolve().parent.parent / "static" / "modeling"
HTML = (STATIC / "lbo.html").read_text(encoding="utf-8")
JS = (STATIC / "js" / "lbo.js").read_text(encoding="utf-8")


# ── 1. Currency dropdown / payload normalization ─────────────────────────────


@pytest.mark.parametrize("raw,expected", [
    ("usd", "USD"), ("USD", "USD"), ("Usd", "USD"),
    ("hkd", "HKD"), ("CNY", "CNY"),
    ("USDT", "USD"), ("rmb", "USD"), ("HK$", "USD"),
    ("", "USD"), (None, "USD"), ("EUR", "USD"),
])
def test_normalize_currency(raw, expected):
    assert _normalize_lbo_currency(raw) == expected


def test_normalize_payload_currency_mutates_in_place():
    payload = {"currency": "usd"}
    _normalize_payload_currency(payload)
    assert payload["currency"] == "USD"


def test_api_normalizes_lowercase_currency():
    client = app.test_client()
    payload = clean_anchor()
    payload["currency"] = "usd"
    resp = client.post("/api/modeling/lbo", json=payload)
    assert resp.status_code == 200
    assert resp.get_json()["currency"] == "USD"


def test_api_rejects_illegal_currency_with_fallback():
    client = app.test_client()
    payload = clean_anchor()
    payload["currency"] = "USDT"
    resp = client.post("/api/modeling/lbo", json=payload)
    assert resp.get_json()["currency"] == "USD"


def test_html_currency_is_select_not_free_text():
    # The currency control must be a <select id="currency">, never a text input.
    assert re.search(r'<select[^>]*id="currency"', HTML)
    assert not re.search(r'<input[^>]*id="currency"', HTML)


def test_html_currency_options_limited_to_supported_set():
    block = HTML.split('id="currency"', 1)[1].split("</select>", 1)[0]
    options = set(re.findall(r'<option value="([^"]+)"', block))
    assert options == {"USD", "HKD", "CNY"}


# ── 2. Unit labels ───────────────────────────────────────────────────────────


def test_input_labels_include_currency_unit():
    # Entry EBITDA and Debt Amount labels must carry a currency-aware unit tag.
    assert "Entry EBITDA <span class=\"unit-tag\">(USD mm)</span>" in HTML
    assert 'id="debt_amount_label"' in HTML
    assert "Debt Amount <span class=\"unit-tag\">(USD mm)</span>" in HTML


def test_result_cards_carry_unit_note():
    assert HTML.count("Amounts in USD mm") >= 3
    assert "data-ccy-note" in HTML


def test_js_updates_unit_labels_on_currency_change():
    assert "function updateUnitLabels" in JS
    assert "function normalizeCurrency" in JS
    # currency select change triggers a unit-label refresh
    assert 'addEventListener("change", updateUnitLabels)' in JS


def test_excel_shows_amounts_unit_note():
    payload = clean_anchor()
    result = run_lbo(payload)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, result).getvalue()))
    blob = "\n".join(
        c.value for ws in wb.worksheets for row in ws.iter_rows()
        for c in row if isinstance(c.value, str)
    )
    assert "Amounts in USD mm" in blob


@pytest.mark.parametrize("ccy", ["HKD", "CNY"])
def test_excel_unit_note_follows_currency(ccy):
    payload = clean_anchor()
    payload["currency"] = ccy
    result = run_lbo(payload)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, result).getvalue()))
    blob = "\n".join(
        c.value for ws in wb.worksheets for row in ws.iter_rows()
        for c in row if isinstance(c.value, str)
    )
    assert f"Amounts in {ccy} mm" in blob


def test_excel_illegal_currency_does_not_reach_workbook():
    payload = clean_anchor()
    payload["currency"] = "USDT"
    result = run_lbo(payload)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, result).getvalue()))
    blob = "\n".join(
        c.value for ws in wb.worksheets for row in ws.iter_rows()
        for c in row if isinstance(c.value, str)
    )
    assert "USDT" not in blob
    assert "Amounts in USD mm" in blob


# ── 3. Debt sizing ───────────────────────────────────────────────────────────


def test_by_leverage_multiple_sizes_debt_from_ebitda():
    """Entry EBITDA 180 x 5.0x -> debt_amount 900 (engine sizes from leverage)."""
    payload = clean_anchor()
    payload["transaction"]["entry_ebitda"] = 180
    payload["transaction"]["entry_multiple"] = 12.0  # EV large enough to support debt
    payload["operating_forecast"]["ebitda"] = [180, 180, 180, 180, 180]
    # leverage-mode payload: leverage_multiple present, no debt_amount
    payload["debt"] = {
        "leverage_multiple": 5.0,
        "interest_rate": 0.09,
        "mandatory_amortization_pct": 0.01,
        "cash_sweep_pct": 1.0,
        "cash_to_balance_sheet": 0.0,
    }
    result = run_lbo(payload)
    assert result["status"] == "ok"
    assert result["transaction_summary"]["debt_amount"] == pytest.approx(900.0)


def test_manual_debt_amount_is_used_directly():
    payload = clean_anchor()
    payload["debt"] = {
        "debt_amount": 5000,
        "interest_rate": 0.09,
        "mandatory_amortization_pct": 0.01,
        "cash_sweep_pct": 1.0,
        "cash_to_balance_sheet": 0.0,
    }
    result = run_lbo(payload)
    assert result["status"] == "ok"
    assert result["transaction_summary"]["debt_amount"] == 5000


def test_js_has_debt_sizing_and_implied_leverage():
    assert "function debtSizingMode" in JS
    assert "function syncDebtSizing" in JS
    # implied leverage shown in manual mode, n/a when EBITDA <= 0
    assert "implied-leverage" in JS
    assert 'ebitda > 0 ? `${(debt / ebitda).toFixed(1)}x` : "n/a"' in JS
    # buildPayload routes the single debt input by sizing mode
    assert 'debtSizingMode() === "manual"' in JS


def test_html_has_debt_sizing_control():
    assert re.search(r'<select[^>]*id="debt_sizing"', HTML)
    block = HTML.split('id="debt_sizing"', 1)[1].split("</select>", 1)[0]
    options = set(re.findall(r'<option value="([^"]+)"', block))
    assert options == {"leverage", "manual"}
    assert 'id="implied-leverage"' in HTML


# ── 4. Error messaging ───────────────────────────────────────────────────────


def test_engine_still_emits_debt_exceeds_uses_code():
    """The hotfix must NOT change engine math: EBITDA 180 at 5.0x EV with a
    5000 manual debt still triggers DEBT_EXCEEDS_USES (debt > total uses)."""
    payload = clean_anchor()
    payload["transaction"]["entry_ebitda"] = 180
    payload["transaction"]["entry_multiple"] = 5.0  # EV = 900, < debt 5000
    payload["operating_forecast"]["ebitda"] = [180, 180, 180, 180, 180]
    payload["debt"] = {
        "debt_amount": 5000,
        "interest_rate": 0.09,
        "mandatory_amortization_pct": 0.01,
        "cash_sweep_pct": 1.0,
        "cash_to_balance_sheet": 0.0,
    }
    result = run_lbo(payload)
    assert result["status"] == "error"
    assert any(f["code"] == "DEBT_EXCEEDS_USES" for f in result["flags"])


def test_js_friendly_error_preserves_code_and_explains():
    # FRIENDLY_ERRORS maps the engine code to a readable explanation; the code
    # itself is still rendered.
    assert "FRIENDLY_ERRORS" in JS
    assert "DEBT_EXCEEDS_USES" in JS
    assert "Debt amount exceeds total uses" in JS
    assert "债务金额超过总资金用途" in JS
    assert "<code>${f.code}</code>" in JS
    # manual-mode extra hint
    assert "Switch to By Leverage Multiple" in JS


# ── 5. Regression: legacy values-only path stays 0-formula after unit notes ───
# (The user-facing export is now the formula-native workbook; this guards the
#  retained legacy/oracle path only — see test_lbo_formula_workbook.py.)


def test_legacy_excel_remains_values_only_after_unit_notes():
    payload = clean_anchor()
    result = run_lbo(payload)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, result).getvalue()))
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert not cell.value.startswith("="), f"formula in {ws.title}"
    assert not wb._external_links
