"""V4.7.1 RC analyst-feedback fixes.

Covers the four fix areas:
  A. Cash to Balance Sheet / Sources & Uses tie-out (single cash figure).
  B. Scenario comparison = same-transaction sensitivity (sponsor equity fixed).
  C. Stronger, calibrated downside shock.
  D. Lightweight analyst context badges (capped at 2).
  E. UI polish wording (attribution footnotes / banner / debt-sizing labels).

No core engine math changes beyond the narrow S&U cash tie-out: run_lbo IRR/MOIC
mechanics, the waterfall, covenant detection and attribution math are unchanged.
"""

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from modeling.lbo_calculator import run_lbo
from modeling.lbo_attribution import build_lbo_attribution
from modeling.lbo_excel_exporter import generate_lbo_excel
from modeling.lbo_scenarios import build_lbo_scenarios, generate_scenario_inputs
from modeling.lbo_return_context import (
    build_return_context,
    GROWTH_DRIVEN_CONTEXT,
    LIMITED_DELEVERAGING_CONTEXT,
    EARLY_LIQUIDITY_PRESSURE_CONTEXT,
    MAX_DISPLAYED_BADGES,
    _BADGES,
)
from tests.test_lbo_calculator_core import clean_anchor
from tests.test_lbo_multitranche_schedule import multi_tranche_anchor, revolver_draw_fixture


STATIC = Path(__file__).resolve().parent.parent / "static" / "modeling"
HTML = (STATIC / "lbo.html").read_text(encoding="utf-8")
JS = (STATIC / "js" / "lbo.js").read_text(encoding="utf-8")

RECOMMENDATION_TERMS = [
    "should acquire", "should buy", "good deal", "bad deal", "attractive deal",
    "buy signal", "sell signal", "recommended deal", "建议收购", "建议买入",
]


def _multi_with_cash(cash=40.0, cs_begin=None):
    a = multi_tranche_anchor()
    a["debt"] = {"cash_to_balance_sheet": cash}
    a["capital_structure"]["cash_balance_beginning"] = cash if cs_begin is None else cs_begin
    return a


def _growth_multi():
    a = multi_tranche_anchor()
    a["operating_forecast"]["ebitda"] = [1000, 1080, 1166, 1259, 1360]
    a["operating_forecast"]["revenue"] = [5000, 5400, 5832, 6298, 6802]
    return a


def _su_pairs(wb):
    ws = wb["Sources & Uses"]
    pairs = {}
    for r in range(1, ws.max_row + 1):
        k = ws.cell(r, 1).value
        v = ws.cell(r, 2).value
        if isinstance(k, str) and isinstance(v, (int, float)):
            pairs[k] = v
    return pairs


# ══ Part A — Cash to Balance Sheet / Sources & Uses tie-out ═══════════════════


def test_cash_to_balance_sheet_included_in_total_uses():
    out = run_lbo(_multi_with_cash(40.0))
    ts = out["transaction_summary"]
    assert ts["cash_to_balance_sheet"] == 40.0
    assert ts["total_uses"] == pytest.approx(ts["entry_ev"] + ts["transaction_fees"] + 40.0)


def test_sponsor_equity_includes_cash_to_balance_sheet():
    base = run_lbo(multi_tranche_anchor())["transaction_summary"]["sponsor_equity"]
    withcash = run_lbo(_multi_with_cash(40.0))["transaction_summary"]["sponsor_equity"]
    assert withcash == pytest.approx(base + 40.0)


def test_beginning_cash_equals_cash_to_balance_sheet_in_waterfall():
    no_cash = run_lbo(multi_tranche_anchor())["debt_schedule"][0]["cash_before_debt_service"]
    with_cash = run_lbo(_multi_with_cash(40.0))["debt_schedule"][0]["cash_before_debt_service"]
    # Beginning cash (= cash to balance sheet) flows into year-1 cash available.
    assert with_cash == pytest.approx(no_cash + 40.0)


def test_ending_cash_consistent_between_summary_and_schedule():
    out = run_lbo(_multi_with_cash(40.0))
    assert out["exit"]["ending_cash_balance"] == pytest.approx(
        out["capital_structure_summary"]["ending_cash_balance"]
    )


def test_excel_sources_uses_tie_out_with_cash():
    payload = _multi_with_cash(40.0)
    result = run_lbo(payload)
    payload["attribution"] = build_lbo_attribution(payload, result)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, result).getvalue()))
    p = _su_pairs(wb)
    assert p["Cash to Balance Sheet"] == 40.0
    assert p["Total Uses"] == pytest.approx(p["Entry EV"] + p["Transaction Fees"] + p["Cash to Balance Sheet"])
    assert p["Total Sources"] == pytest.approx(p["Total Uses"])


def test_no_cash_mismatch_engine_vs_excel():
    payload = _multi_with_cash(40.0)
    result = run_lbo(payload)
    payload["attribution"] = build_lbo_attribution(payload, result)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, result).getvalue()))
    p = _su_pairs(wb)
    # Engine transaction_summary cash == Excel S&U cash == waterfall beginning cash.
    assert result["transaction_summary"]["cash_to_balance_sheet"] == p["Cash to Balance Sheet"]


def test_conflicting_cash_does_not_silently_drift():
    out = run_lbo(_multi_with_cash(cash=30.0, cs_begin=40.0))
    # Transaction value wins; a non-blocking conflict warning is surfaced.
    assert out["transaction_summary"]["cash_to_balance_sheet"] == 30.0
    assert any(f["code"] == "CASH_TO_BALANCE_SHEET_CONFLICT" for f in out["flags"])


def test_zero_cash_still_ties_out():
    out = run_lbo(multi_tranche_anchor())
    ts = out["transaction_summary"]
    assert ts["cash_to_balance_sheet"] == 0.0
    assert ts["total_uses"] == pytest.approx(ts["entry_ev"] + ts["transaction_fees"])


# ══ Part B — Scenario same-transaction sensitivity ════════════════════════════


def _scn_metric(res, key):
    rows = res["comparison"]["rows"]
    r = next(x for x in rows if x["key"] == key)
    return r["base"], r["upside"], r["downside"]


def test_scenario_sponsor_equity_fixed():
    res = build_lbo_scenarios(_growth_multi())
    b, u, d = _scn_metric(res, "sponsor_equity")
    assert b == u == d


def test_scenario_entry_ev_fixed():
    g = generate_scenario_inputs(_growth_multi())
    evs = {k: v["transaction"]["entry_ebitda"] * v["transaction"]["entry_multiple"]
           for k, v in g.items()}
    assert evs["base"] == evs["upside"] == evs["downside"]


def test_scenario_opening_debt_fixed():
    res = build_lbo_scenarios(_growth_multi())
    # Opening debt is the same opening tranche balances across scenarios.
    debts = []
    for key in ("base", "upside", "downside"):
        scn = res["scenarios"][key]
        debts.append(scn["result"]["transaction_summary"]["debt_amount"])
    assert debts[0] == debts[1] == debts[2]


def test_scenario_operating_and_exit_still_move():
    g = generate_scenario_inputs(_growth_multi())
    assert g["upside"]["operating_forecast"]["ebitda"][0] > g["base"]["operating_forecast"]["ebitda"][0]
    assert g["downside"]["operating_forecast"]["ebitda"][0] < g["base"]["operating_forecast"]["ebitda"][0]
    assert g["upside"]["transaction"]["exit_multiple"] > g["base"]["transaction"]["exit_multiple"]
    assert g["downside"]["transaction"]["exit_multiple"] < g["base"]["transaction"]["exit_multiple"]


def test_scenario_irr_moic_comparison_still_works():
    res = build_lbo_scenarios(_growth_multi())
    b_irr, u_irr, d_irr = _scn_metric(res, "irr")
    assert u_irr > b_irr > d_irr
    b_moic, u_moic, d_moic = _scn_metric(res, "moic")
    assert u_moic > b_moic > d_moic


def test_scenario_summary_excel_sponsor_equity_equal_columns():
    payload = _growth_multi()
    result = run_lbo(payload)
    payload["attribution"] = build_lbo_attribution(payload, result)
    payload["scenarios"] = build_lbo_scenarios(_growth_multi())
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, result).getvalue()))
    ws = wb["Scenario Summary"]
    # Find the Sponsor Equity row and assert its three scenario columns match.
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "Sponsor Equity":
            vals = [ws.cell(r, c).value for c in (2, 3, 4)]
            assert vals[0] == vals[1] == vals[2]
            break
    else:
        pytest.fail("Sponsor Equity row not found in Scenario Summary")


def test_same_transaction_disclosure_present():
    res = build_lbo_scenarios(_growth_multi())
    assert "same-transaction sensitivity" in res["notes"]["same_transaction_note_en"]
    assert "同一交易敏感性" in res["notes"]["same_transaction_note_cn"]


# ══ Part C — Stronger downside shock ══════════════════════════════════════════


def test_downside_config_stronger_than_v46():
    res = build_lbo_scenarios(_growth_multi())
    d = res["scenario_config"]["downside"]
    assert d["ebitda_pct_delta"] == -0.15
    assert d["exit_multiple_delta"] == -1.0
    assert d["interest_rate_delta"] == 0.01
    assert d["capex_pct_delta"] == 0.10
    assert d["nwc_pct_delta"] == 0.10


def test_downside_irr_meaningfully_below_base():
    res = build_lbo_scenarios(_growth_multi())
    b, u, d = _scn_metric(res, "irr")
    # The downside should bite: at least ~8pp below base for this fixture.
    assert b - d >= 0.08


def test_downside_can_show_covenant_or_liquidity_stress():
    res = build_lbo_scenarios(revolver_draw_fixture())
    down = res["scenarios"]["downside"]
    # Either downside is unavailable (too stressed) or shows covenant/liquidity stress.
    if down["status"] != "ok":
        return
    result = down["result"]
    cov = (result.get("covenant_summary") or {}).get("status")
    revolver = any(r.get("revolver_draw", 0) > 0 for r in result.get("debt_schedule") or [])
    dsf = any(r.get("debt_service_failure") for r in result.get("debt_schedule") or [])
    assert cov == "breach" or revolver or dsf


def test_downside_sponsor_equity_fixed_despite_shock():
    res = build_lbo_scenarios(_growth_multi())
    b, u, d = _scn_metric(res, "sponsor_equity")
    assert b == d


def test_scenario_disclosures_non_recommendation():
    res = build_lbo_scenarios(_growth_multi())
    blob = " ".join(res["disclosures"] + res["disclosures_cn"]).lower()
    for term in RECOMMENDATION_TERMS:
        assert term.lower() not in blob


# ══ Part D — Analyst context badges (light) ═══════════════════════════════════


def _ctx(payload):
    result = run_lbo(payload)
    attribution = build_lbo_attribution(payload, result)
    return build_return_context(result, attribution)


def test_growth_driven_badge():
    ctx = _ctx(_growth_multi())
    assert GROWTH_DRIVEN_CONTEXT in ctx["codes"]


def test_early_liquidity_pressure_badge_on_revolver_draw():
    ctx = _ctx(revolver_draw_fixture())
    assert EARLY_LIQUIDITY_PRESSURE_CONTEXT in ctx["codes"]


def test_limited_deleveraging_badge():
    # Deleveraging MOIC contribution in the (near-zero, limited] band fires the
    # limited-deleveraging badge. Unit-tested against a synthetic attribution so
    # the threshold logic is isolated from engine tuning.
    result = {"status": "ok", "returns": {"irr": 0.20}, "debt_schedule": []}
    attribution = {"status": "ok", "components": [
        {"key": "ebitda_growth", "moic_contribution": 0.30},
        {"key": "multiple_movement", "moic_contribution": 0.10},
        {"key": "deleveraging", "moic_contribution": 0.03},
        {"key": "fees_drag", "moic_contribution": -0.02},
        {"key": "residual", "moic_contribution": 0.0},
    ]}
    ctx = build_return_context(result, attribution)
    assert LIMITED_DELEVERAGING_CONTEXT in ctx["codes"]


def test_badges_capped_at_two():
    # clean_anchor triggers three codes but only two are displayed.
    ctx = _ctx(clean_anchor())
    assert len(ctx["codes"]) >= 2
    assert len(ctx["badges"]) <= MAX_DISPLAYED_BADGES
    assert ctx["badges"][0]["code"] == ctx["primary_code"]


@pytest.mark.parametrize("code", [
    GROWTH_DRIVEN_CONTEXT, LIMITED_DELEVERAGING_CONTEXT, EARLY_LIQUIDITY_PRESSURE_CONTEXT,
])
def test_new_badges_have_no_recommendation_wording(code):
    text = (_BADGES[code]["en"] + " " + _BADGES[code]["cn"]).lower()
    for term in RECOMMENDATION_TERMS:
        assert term.lower() not in text


def test_badge_lines_are_compact():
    for code, b in _BADGES.items():
        assert "\n" not in b["en"] and "\n" not in b["cn"]
        assert len(b["en"]) <= 120 and len(b["cn"]) <= 120


# ══ Part E — UI polish wording ════════════════════════════════════════════════


def test_attribution_footnotes_no_acquisition_disclaimer():
    # The render path uses a unified .attribution-footnotes block and drops the
    # investment/acquisition recommendation disclaimer.
    assert "attribution-footnotes" in JS
    render = JS.split("function renderAttribution", 1)[1].split("\nfunction ", 1)[0]
    assert "不构成投资或收购建议" not in render
    assert "acquisition recommendation" not in render.lower()
    assert "MOIC - 1.0" in render


def test_manual_forecast_banner_trimmed():
    assert "适合正式建模" not in HTML
    assert "逐年填写 Revenue、EBITDA、Cash Taxes、CapEx、NWC。" in HTML


def test_debt_sizing_labels_are_short_chinese():
    block = HTML.split('id="debt_sizing"', 1)[1].split("</select>", 1)[0]
    assert ">按倍数<" in block
    assert ">手动债务<" in block
    # English retained in tooltips only.
    assert 'title="By Leverage Multiple"' in block
    assert 'title="Manual Debt Amount"' in block


def test_no_recommendation_wording_in_ui():
    blob = (HTML + JS).lower()
    for term in RECOMMENDATION_TERMS:
        assert term.lower() not in blob
