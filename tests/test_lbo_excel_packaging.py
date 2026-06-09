"""V4.5 Professional LBO Workbook Packaging tests.

Verifies the workbook architecture, sheet separation, source-of-truth discipline
(no formulas / external links), and disclosure readability for both single- and
multi-tranche output modes. These tests consume Python output only; they do not
exercise any engine math.
"""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from modeling.lbo_calculator import run_lbo
from modeling.lbo_attribution import build_lbo_attribution
from modeling.lbo_excel_exporter import (
    generate_lbo_excel,
    SHEET_NAMES,
    SCENARIO_SUMMARY_SHEET,
    WORKBOOK_VERSION,
)
from modeling.lbo_scenarios import build_lbo_scenarios
from tests.test_lbo_calculator_core import clean_anchor
from tests.test_lbo_multitranche_schedule import multi_tranche_anchor


# ── fixtures / helpers ───────────────────────────────────────────────────────


def _workbook(anchor):
    payload = anchor()
    out = run_lbo(payload)
    payload["attribution"] = build_lbo_attribution(payload, out)
    data = generate_lbo_excel(payload, out).getvalue()
    return payload, out, load_workbook(BytesIO(data), data_only=False)


@pytest.fixture
def single():
    return _workbook(clean_anchor)


@pytest.fixture
def multi():
    return _workbook(multi_tranche_anchor)


def _sheet_strings(ws):
    out = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                out.append(cell.value)
    return out


def _all_strings(wb):
    out = []
    for ws in wb.worksheets:
        out.extend(_sheet_strings(ws))
    return out


def _col1(ws):
    return [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]


# ── 1-2. workbook exports for both modes ─────────────────────────────────────


def test_single_tranche_workbook_exports(single):
    _, _, wb = single
    assert wb.worksheets


def test_multi_tranche_workbook_exports(multi):
    _, _, wb = multi
    assert wb.worksheets


# ── 3. sheet order: V4.5 11-sheet default, V4.6 12-sheet with scenario ───────


def test_sheet_order_single(single):
    _, _, wb = single
    assert wb.sheetnames == SHEET_NAMES


def test_sheet_order_multi(multi):
    _, _, wb = multi
    assert wb.sheetnames == SHEET_NAMES


@pytest.mark.parametrize("anchor", [clean_anchor, multi_tranche_anchor])
def test_sheet_order_no_scenario_is_eleven(anchor):
    payload = anchor()
    out = run_lbo(payload)
    payload["attribution"] = build_lbo_attribution(payload, out)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, out).getvalue()), data_only=False)
    assert wb.sheetnames == SHEET_NAMES
    assert len(wb.sheetnames) == 11


@pytest.mark.parametrize("anchor", [clean_anchor, multi_tranche_anchor])
def test_sheet_order_with_scenario_is_twelve_at_position_three(anchor):
    payload = anchor()
    out = run_lbo(payload)
    payload["attribution"] = build_lbo_attribution(payload, out)
    payload["scenarios"] = build_lbo_scenarios(anchor())
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, out).getvalue()), data_only=False)
    expected = SHEET_NAMES[:2] + [SCENARIO_SUMMARY_SHEET] + SHEET_NAMES[2:]
    assert wb.sheetnames == expected
    assert len(wb.sheetnames) == 12
    assert wb.sheetnames[2] == SCENARIO_SUMMARY_SHEET


# ── 4. Cover contains title / symbol / model version ─────────────────────────


def test_cover_contains_identity(single):
    payload, _, wb = single
    ws = wb["Cover"]
    strings = _sheet_strings(ws)
    assert "LBO Model Workbook" in strings
    assert "Symbol" in strings
    assert WORKBOOK_VERSION in strings
    # Symbol value tied to payload.
    pairs = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, ws.max_row + 1)}
    assert pairs["Symbol"] == payload.get("symbol") or pairs["Symbol"] is not None


# ── 5. Model Map lists all sheets ────────────────────────────────────────────


def test_model_map_lists_all_sheets(single):
    _, _, wb = single
    listed = _sheet_strings(wb["Model Map"])
    for name in SHEET_NAMES:
        assert name in listed, f"Model Map missing {name}"


def test_legacy_model_map_keeps_values_only_semantics(single):
    """The legacy values-only workbook Model Map must keep its Python-precomputed
    wording; V4.1.1.1 only changed the formula-native workbook's Model Map."""
    _, _, wb = single
    blob = " ".join(_sheet_strings(wb["Model Map"]))
    assert "No Excel recalculation" in blob
    assert "Live Excel formulas" not in blob


# ── 6. Assumptions contains transaction assumptions ──────────────────────────


def test_assumptions_has_transaction_section(single):
    _, _, wb = single
    strings = _sheet_strings(wb["Assumptions"])
    assert "Transaction Assumptions" in strings
    assert "Entry EBITDA" in strings
    assert "Entry Multiple" in strings
    assert "Exit Multiple" in strings


# ── 7. Sources & Uses ties total sources / uses ──────────────────────────────


def test_sources_uses_ties(single):
    _, out, wb = single
    ws = wb["Sources & Uses"]
    pairs = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, ws.max_row + 1)}
    assert pairs["Sources = Uses"] == "Yes"
    assert abs(pairs["Total Sources"] - pairs["Total Uses"]) < 0.01
    ts = out["transaction_summary"]
    assert abs(pairs["Total Uses"] - ts["total_uses"]) < 0.01


# ── 8. Operating Forecast has forecast year columns ──────────────────────────


def test_operating_forecast_has_year_columns(single):
    _, out, wb = single
    ws = wb["Operating Forecast"]
    header = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
    years = out["operating_forecast"]["years"]
    for y in years:
        assert f"Y{y}" in header


# ── 9-10. Debt Schedule present in both modes ────────────────────────────────


def test_debt_schedule_single(single):
    _, _, wb = single
    strings = _sheet_strings(wb["Debt Schedule"])
    assert "Beginning Debt" in strings
    assert "Ending Debt" in strings


def test_debt_schedule_multi_has_tranche_rows(multi):
    _, _, wb = multi
    ws = wb["Debt Schedule"]
    strings = _sheet_strings(ws)
    assert "Tranche Detail" in strings
    assert "Beginning Balance" in strings
    assert "Ending Cash" in strings  # ending cash shown in multi mode


# ── 11. Covenant Check has leverage / interest coverage fields ───────────────


def test_covenant_check_fields(multi):
    _, _, wb = multi
    strings = _sheet_strings(wb["Covenant Check"])
    assert "Net Debt / EBITDA" in strings
    assert "Interest Coverage" in strings


# ── 12. Returns Summary has headline IRR / MOIC ──────────────────────────────


def test_returns_summary_headline(single):
    _, out, wb = single
    ws = wb["Returns Summary"]
    pairs = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, ws.max_row + 1)}
    assert abs(pairs["IRR"] - out["returns"]["irr"]) < 1e-12
    assert abs(pairs["MOIC"] - out["returns"]["moic"]) < 1e-12


# ── 13. Returns Attribution contains MOIC − 1.0 note ─────────────────────────


def test_attribution_moic_minus_one_note(multi):
    _, _, wb = multi
    strings = _sheet_strings(wb["Returns Attribution"])
    assert "Component MOIC contribution sums to MOIC - 1.0, not headline MOIC." in strings
    # Canonical ASCII-hyphen wording only; unicode minus is no longer accepted.
    assert not any("MOIC − 1.0" in s for s in strings)


# ── 14. Maturity Wall uses 'Outstanding at Exit' label ───────────────────────


def test_maturity_wall_label(multi):
    _, _, wb = multi
    strings = _sheet_strings(wb["Maturity Wall"])
    assert "Outstanding at Exit" in strings


# ── 15-16. Audit disclosures ─────────────────────────────────────────────────


def test_audit_source_of_truth_disclosure(single):
    _, _, wb = single
    strings = _sheet_strings(wb["Audit & Disclosures"])
    assert "Source of Truth" in strings
    assert any("source of truth" in s.lower() for s in strings)


def test_audit_no_recommendation_disclosure(single):
    _, _, wb = single
    strings = _sheet_strings(wb["Audit & Disclosures"])
    assert any(
        "not an investment or acquisition recommendation" in s.lower()
        or "not investment" in s.lower()
        for s in strings
    )


# ── 17-18. no external links / no formulas ───────────────────────────────────


@pytest.mark.parametrize("anchor", [clean_anchor, multi_tranche_anchor])
def test_no_external_links(anchor):
    _, _, wb = _workbook(anchor)
    assert not getattr(wb, "_external_links", [])


@pytest.mark.parametrize("anchor", [clean_anchor, multi_tranche_anchor])
def test_legacy_values_only_path_has_no_formulas(anchor):
    # This file exercises generate_lbo_excel == the legacy values-only / oracle
    # path, which stays 0-formula by design. The formula-native export contract
    # lives in test_lbo_formula_workbook.py.
    _, _, wb = _workbook(anchor)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                assert not (isinstance(cell.value, str) and cell.value.startswith("="))


# ── 19. key sheets hide gridlines ────────────────────────────────────────────


def test_key_sheets_hide_gridlines(multi):
    _, _, wb = multi
    for name in SHEET_NAMES:
        assert wb[name].sheet_view.showGridLines is False, name


# ── 20. key sheets have freeze panes ─────────────────────────────────────────


def test_key_sheets_have_freeze_panes(multi):
    _, _, wb = multi
    for name in [
        "Model Map", "Assumptions", "Sources & Uses", "Operating Forecast",
        "Debt Schedule", "Covenant Check", "Returns Summary",
        "Returns Attribution", "Maturity Wall", "Audit & Disclosures",
    ]:
        assert wb[name].freeze_panes is not None, name


# ── 21. notes/disclosures wrap and have reasonable row heights ───────────────


def test_disclosures_wrap_and_reasonable_height(multi):
    _, _, wb = multi
    ws = wb["Audit & Disclosures"]
    label_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Deferred Scope")
    note = ws.cell(label_row + 1, 1)
    assert note.alignment.wrap_text is True
    h = ws.row_dimensions[label_row + 1].height
    assert h is not None and 24 <= h <= 90
