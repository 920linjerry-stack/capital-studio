from io import BytesIO
import zipfile

from openpyxl import load_workbook

from app import _dcf_input_from_payload
from modeling.dcf_calculator import run_dcf
from modeling.excel_exporter import generate_excel


def _payload(symbol="2359.HK", company="WuXi AppTec Co., Ltd.", currency="HKD"):
    reporting = "CNY" if symbol.endswith(".HK") and symbol != "AAPL" else currency
    return {
        "symbol": symbol,
        "company": company,
        "price": 133.0,
        "revenue": 1000,
        "ebit": 250,
        "da": 40,
        "capex": 50,
        "wc_change": 0,
        "tax_rate": 0.21,
        "net_debt": 0,
        "shares": 100,
        "revenue_growth": 0.03,
        "ebit_margin": 0.25,
        "da_pct_revenue": 0.04,
        "capex_pct_revenue": 0.05,
        "wc_change_pct_revenue": 0.0,
        "wacc": 0.10,
        "beta": 1.0,
        "terminal_g": 0.02,
        "exit_multiple": 12,
        "forecast_years": 5,
        "tv_method": "average",
        "currency": currency,
        "reporting_currency": reporting,
        "trading_currency": currency,
        "fx_rate_reporting_to_trading": 1.0,
        "intrinsic_value_per_share_trading_currency": 100.0,
    }


def _workbook(symbol="2359.HK", company="WuXi AppTec Co., Ltd.", currency="HKD"):
    payload = _payload(symbol, company, currency)
    inp = _dcf_input_from_payload(payload)
    out = run_dcf(inp)
    ctx = {
        "current_params": {**payload, "data_source": "AKShare"},
        "generated_at": "2026-05-27T00:00:00",
        "trading_comps": {"status": "unavailable", "peer_rows": [], "valuation_ranges": []},
    }
    data = generate_excel(inp, out, ctx).getvalue()
    return data, load_workbook(BytesIO(data), data_only=False)


def _all_text(wb):
    return "\n".join(
        str(cell.value)
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    )


def test_sensitivity_headers_are_white_on_dark_fill():
    _, wb = _workbook()
    ws = wb["Sensitivity"]
    labels = {"WACC \\ g", "WACC \\ EV/EBITDA", "EBIT Margin \\ Revenue Growth"}
    found = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in labels:
                fill = cell.fill.fgColor.rgb
                color = cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else None
                found[cell.value] = (fill, color, cell.font.bold)
    assert labels <= set(found)
    for fill, color, bold in found.values():
        assert fill in {"FF1F4E78", "1F4E78", "FF1A3A5C", "001A3A5C", "1A3A5C"}
        assert color in {"FFFFFFFF", "00FFFFFF", "FFFFFF"}
        assert bold is True


def test_sensitivity_notes_panel_right_side_and_left_side_compact():
    _, wb = _workbook()
    ws = wb["Sensitivity"]
    assert ws["H2"].value == "Sensitivity Notes"
    panel_text = "\n".join(str(ws.cell(r, 8).value or "") for r in range(2, 13))
    assert "Selected Method matrix ties to headline IV." in panel_text
    assert "Sensitivities do not alter headline valuation." in panel_text

    left_text = "\n".join(
        str(cell.value)
        for row in ws.iter_rows(min_col=1, max_col=7)
        for cell in row
        if isinstance(cell.value, str)
    )
    assert "Diagnostic only if not selected terminal method" not in left_text
    assert "Static at export; re-export" not in left_text

    title_rows = [
        cell.row
        for row in ws.iter_rows(min_col=1, max_col=1)
        for cell in row
        if cell.value in {
            "Gordon-only sensitivity",
            "Exit-only sensitivity",
            "Selected Method Sensitivity - center ties to headline",
            "Operating sensitivity | Revenue Growth x EBIT Margin",
        }
    ]
    max_blank_run = 0
    for start, end in zip(title_rows, title_rows[1:]):
        streak = 0
        for row_idx in range(start + 1, end):
            is_blank = all(ws.cell(row_idx, col_idx).value in (None, "") for col_idx in range(1, 8))
            streak = streak + 1 if is_blank else 0
            max_blank_run = max(max_blank_run, streak)
    assert max_blank_run <= 4


def test_print_headers_set_on_key_sheets():
    _, wb = _workbook()
    for name in ["Cover", "Executive Summary", "Assumptions", "FCF Build", "DCF Valuation", "Sensitivity", "Data Sources Audit"]:
        ws = wb[name]
        assert "WuXi AppTec" in ws.oddHeader.left.text
        assert "2359.HK" in ws.oddHeader.left.text
        assert ws.oddHeader.center.text == "&A"
        assert "Page &P of &N" in ws.oddHeader.right.text


def test_executive_summary_identity_mapping_and_assumption_widths():
    _, wb = _workbook()
    ws = wb["Executive Summary"]
    values = set()
    for r in (3, 4):
        for c in (1, 4, 7):
            values.add((ws.cell(r, c).value, ws.cell(r, c + 1).value))
    assert ("Ticker", "2359.HK") in values
    assert ("Company", "WuXi AppTec Co., Ltd.") in values
    assert "CNY reporting / HKD trading" in _all_text(wb)
    assert wb["Assumptions"].column_dimensions["G"].width >= 50


def test_cover_data_source_compact_and_no_excessive_identity_row_height():
    _, wb = _workbook()
    ws = wb["Cover"]
    text = _all_text(wb)
    assert "Public market data, cached financial statements, and trading data." in text
    heights = [ws.row_dimensions[r].height or 15 for r in range(11, 18)]
    assert max(heights) <= 24


def test_non_us_workbook_has_no_10k_or_formula_error_literals_or_external_links():
    data, wb = _workbook()
    text = _all_text(wb)
    assert "10-K" not in text
    assert "Form 10-K" not in text
    assert "#VALUE!" not in text
    assert "#REF!" not in text
    assert "#DIV/0!" not in text
    with zipfile.ZipFile(BytesIO(data)) as zf:
        assert not [name for name in zf.namelist() if name.startswith("xl/externalLinks/")]
