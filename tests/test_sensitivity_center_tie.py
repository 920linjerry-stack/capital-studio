from io import BytesIO

from openpyxl import load_workbook

from app import _dcf_input_from_payload
from modeling.dcf_calculator import run_dcf
from modeling.excel_exporter import generate_excel


def _payload(symbol="2359.HK", company="WuXi AppTec Co., Ltd.", currency="HKD", reporting="CNY"):
    return {
        "symbol": symbol,
        "company": company,
        "price": 133.0,
        "revenue": 1000,
        "ebit": 320,
        "da": 40,
        "capex": 50,
        "wc_change": 0,
        "tax_rate": 0.21,
        "net_debt": 0,
        "shares": 100,
        "revenue_growth": 0.03,
        "ebit_margin": 0.32,
        "ebit_margin_path": [0.32] * 5,
        "da_pct_revenue": 0.04,
        "capex_pct_revenue": 0.05,
        "wc_change_pct_revenue": 0.0,
        "wacc": 0.10,
        "beta": 1.0,
        "terminal_g": 0.02,
        "exit_multiple": 12,
        "forecast_years": 5,
        "tv_method": "average",
        "currency": currency,
        "reporting_currency": reporting,
        "trading_currency": currency,
        "fx_rate_reporting_to_trading": 1.08 if reporting != currency else 1.0,
    }


def _workbook(payload):
    inp = _dcf_input_from_payload(payload)
    out = run_dcf(inp)
    current_params = {
        **payload,
        "intrinsic_value_per_share_reporting_currency": out.intrinsic_per_share,
        "intrinsic_value_per_share_trading_currency": out.intrinsic_per_share * payload["fx_rate_reporting_to_trading"],
        "ebit_provenance": {
            "source": "normalized_ebit_candidate",
            "note": "Applied by user from Normalized EBIT Review Candidate",
            "applied_margin": payload["ebit_margin"],
        },
    }
    ctx = {
        "current_params": current_params,
        "generated_at": "2026-05-27T00:00:00",
        "trading_comps": {
            "status": "unavailable",
            "reason": "suppressed_generic_tech_fallback",
            "profile_id": "cro_cdmo_peer_mapping_required",
            "peer_set_source": "suppressed_generic_tech_fallback",
            "peer_rows": [],
            "valuation_ranges": [],
        },
    }
    data = generate_excel(inp, out, ctx).getvalue()
    return load_workbook(BytesIO(data), data_only=False)


def _selected_center_formula(wb):
    ws = wb["Sensitivity"]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Selected Method Sensitivity - center ties to headline":
                return ws.cell(cell.row + 6, 4).value
    raise AssertionError("selected-method sensitivity title not found")


def _center_tie_formula(wb):
    ws = wb["Sensitivity"]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Center tie to headline IV":
                return ws.cell(cell.row, cell.column + 1).value
    raise AssertionError("center tie row not found")


def _all_text(wb):
    return "\n".join(
        str(cell.value)
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    )


def test_selected_method_center_references_headline_iv_after_fx_translation():
    wb = _workbook(_payload())
    center_formula = _selected_center_formula(wb)
    tie_formula = _center_tie_formula(wb)

    assert center_formula.startswith("='DCF Valuation'!$B$")
    assert "-'DCF Valuation'!$B$" in tie_formula


def test_terminal_sensitivity_grids_are_trading_currency_when_fx_present():
    wb = _workbook(_payload())
    ws = wb["Sensitivity"]
    formulas = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in {"Gordon-only sensitivity", "Exit-only sensitivity"}:
                formulas.append(ws.cell(cell.row + 4, 4).value)
    assert formulas
    assert all("*1.08" in formula for formula in formulas)


def test_visible_workbook_text_humanizes_engineering_status_tokens():
    wb = _workbook(_payload())
    text = _all_text(wb)

    assert "suppressed_generic_tech_fallback" not in text
    assert "generic_tech_fallback_v1" not in text
    assert "cro_cdmo_peer_mapping_required" not in text
    assert "Generic technology fallback suppressed" in text
    assert "CRO/CDMO peer set requires analyst review" in text


def test_aapl_center_tie_formula_unchanged_as_headline_reference():
    wb = _workbook(_payload("AAPL", "Apple Inc.", "USD", "USD"))
    assert _selected_center_formula(wb).startswith("='DCF Valuation'!$B$")
