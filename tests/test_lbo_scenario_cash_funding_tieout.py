"""V4.1.0 Scenario Summary cash-funding attribution tie-out.

Closes the scan-report gap: the Returns Attribution sheet already had a cash
funding tie-out test (test_lbo_attribution_cash_funding_tieout), but the
Scenario Summary attribution roll-up did not. This file asserts that, for a case
with a non-zero ``cash_to_balance_sheet``, the Scenario Summary comparison

  * surfaces an Initial Cash Funding contribution row (= -cash_to_balance_sheet),
  * and that every per-scenario attribution contribution row sums to that
    scenario's equity value creation (exit equity - sponsor equity).

This is a presentation-layer (values-only) tie-out; it does not require the
Scenario Summary to be formula-native.
"""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from modeling.lbo_calculator import run_lbo
from modeling.lbo_scenarios import build_lbo_scenarios
from modeling.lbo_formula_workbook import generate_lbo_formula_excel
from modeling.lbo_excel_exporter import SCENARIO_SUMMARY_SHEET
from tests.test_lbo_multitranche_schedule import multi_tranche_anchor


# The seven attribution-contribution rows that, by construction, must sum to a
# scenario's equity value creation (exit equity value - sponsor equity).
_CONTRIB_KEYS = [
    "ebitda_growth_contribution",
    "multiple_movement_contribution",
    "deleveraging_contribution",
    "fees_drag",
    "ending_cash_balance_contribution",
    "initial_cash_funding_contribution",
    "residual",
]


def _cash_funding_case():
    payload = multi_tranche_anchor()
    payload["debt"] = {"cash_to_balance_sheet": 60.0}
    payload["capital_structure"]["cash_balance_beginning"] = 60.0
    payload["capital_structure"]["minimum_cash_balance"] = 80.0
    return payload


def _rows_by_key(scenarios):
    return {r["key"]: r for r in scenarios["comparison"]["rows"]}


def test_scenario_summary_surfaces_initial_cash_funding_contribution():
    scenarios = build_lbo_scenarios(_cash_funding_case())
    assert scenarios["status"] == "ok"
    rows = _rows_by_key(scenarios)
    assert "initial_cash_funding_contribution" in rows
    # Cash funded to the balance sheet at close shows as a -60 drag on the bridge.
    assert rows["initial_cash_funding_contribution"]["base"] == pytest.approx(-60.0)


def test_scenario_summary_attribution_rows_sum_to_equity_value_creation():
    scenarios = build_lbo_scenarios(_cash_funding_case())
    rows = _rows_by_key(scenarios)
    statuses = scenarios["comparison"]["scenario_statuses"]

    for case in ("base", "upside", "downside"):
        if statuses.get(case, {}).get("status") != "ok":
            continue
        contrib_sum = sum(rows[k][case] for k in _CONTRIB_KEYS)
        target = rows["exit_equity_value"][case] - rows["sponsor_equity"][case]
        assert contrib_sum == pytest.approx(target, abs=1e-6), (
            f"{case} attribution rows must sum to equity value creation"
        )


def test_scenario_summary_base_initial_cash_funding_ties_to_run_lbo():
    """The Scenario Summary base column attribution must agree with the engine
    oracle: equity value creation = exit equity - sponsor equity from run_lbo."""
    payload = _cash_funding_case()
    oracle = run_lbo(payload)
    scenarios = build_lbo_scenarios(payload)
    rows = _rows_by_key(scenarios)

    oracle_target = (
        oracle["returns"]["exit_equity_value"] - oracle["returns"]["sponsor_equity"]
    )
    contrib_sum = sum(rows[k]["base"] for k in _CONTRIB_KEYS)
    assert contrib_sum == pytest.approx(oracle_target, abs=1e-6)
    assert rows["initial_cash_funding_contribution"]["base"] == pytest.approx(
        -payload["debt"]["cash_to_balance_sheet"]
    )


def test_excel_scenario_summary_includes_initial_cash_funding_row():
    """The exported (formula-native) workbook's Scenario Summary lists the Initial
    Cash Funding row. V4.1.3: the comparison metric values stay Python-precomputed
    (the Initial Cash Funding comparison cell is a value), but the sheet now also
    carries live formula tie-out checks below the comparison table."""
    payload = _cash_funding_case()
    result = run_lbo(payload)
    payload["scenarios"] = build_lbo_scenarios(_cash_funding_case())
    wb = load_workbook(
        BytesIO(generate_lbo_formula_excel(payload, result).getvalue()), data_only=False
    )
    ws = wb[SCENARIO_SUMMARY_SHEET]
    strings = {c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)}
    assert "Initial Cash Funding" in strings
    # V4.1.3 adds live formula checks, so the sheet is no longer 0-formula.
    assert any(
        isinstance(c.value, str) and c.value.startswith("=")
        for row in ws.iter_rows() for c in row
    )
    # The comparison metric values themselves stay Python-precomputed values.
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "Initial Cash Funding":
            value = ws.cell(r, 2).value
            assert not (isinstance(value, str) and value.startswith("="))
            break
