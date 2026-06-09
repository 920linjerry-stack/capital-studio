"""V4.1.0 Formula-Native LBO Workbook contract tests.

These assert the new product direction: the user-facing LBO export is a
formula-native workbook whose core sheets contain real Excel formulas wired
through stable Assumptions anchors / named ranges, while ``run_lbo()`` remains
the calculation oracle. The legacy values-only path keeps its 0-formula
contract and is tested separately (see test_lbo_excel_tieout / packaging).

A small lazy formula evaluator (``_Eval``) resolves named ranges and cell /
cross-sheet references and computes the core outputs, so the test proves the
formulas reproduce the engine oracle without depending on Excel desktop
calculation.
"""

import re
from io import BytesIO

import pytest
from openpyxl import load_workbook

from modeling.lbo_calculator import run_lbo
from modeling.lbo_formula_workbook import (
    FORMULA_WORKBOOK_VERSION,
    generate_lbo_formula_excel,
)
from modeling.lbo_excel_exporter import (
    SHEET_NAMES,
    generate_lbo_values_only_excel,
)
from tests.test_lbo_calculator_core import clean_anchor
from tests.test_lbo_multitranche_schedule import multi_tranche_anchor


# ── Lazy formula evaluator (Excel-desktop independent) ───────────────────────

_PENDING = object()
_CROSS = re.compile(r"'([^']+)'!([A-Z]+[0-9]+)")
# Named ranges may carry digits (e.g. per-tranche LBO_Tr0_Opening), so the name
# token class includes 0-9. Substituted before local cell refs.
_NAME = re.compile(r"\bLBO_[A-Za-z0-9_]+\b")
_LOCAL = re.compile(r"\b([A-Z]{1,3}[0-9]+)\b")
_FUNCS = {
    "IFERROR": lambda a, b: a,  # eager; anchors avoid divide-by-zero
    "IF": lambda c, a, b: a if c else b,
    "POWER": pow,
    "MIN": min,
    "MAX": max,
    "ABS": abs,
    "AND": lambda *a: all(a),
    "OR": lambda *a: any(a),
}


class _Eval:
    def __init__(self, wb):
        self.data = {
            ws.title: {
                c.coordinate: c.value
                for row in ws.iter_rows()
                for c in row
                if c.value is not None
            }
            for ws in wb.worksheets
        }
        self.names = {}
        for nm, dn in wb.defined_names.items():
            sheet, ref = list(dn.destinations)[0]
            self.names[nm] = (sheet, ref.replace("$", ""))
        self.cache = {}

    def cell(self, sheet, coord):
        key = (sheet, coord)
        if key in self.cache:
            if self.cache[key] is _PENDING:
                raise ValueError(f"circular reference at {key}")
            return self.cache[key]
        self.cache[key] = _PENDING
        value = self._eval(sheet, self.data.get(sheet, {}).get(coord))
        self.cache[key] = value
        return value

    def _eval(self, sheet, raw):
        if raw is None:
            return 0.0
        if isinstance(raw, bool):
            return raw
        if isinstance(raw, (int, float)):
            return float(raw)
        if isinstance(raw, str) and raw.startswith("="):
            expr = raw[1:]
            expr = _CROSS.sub(lambda m: repr(self.cell(m.group(1), m.group(2))), expr)
            expr = _NAME.sub(lambda m: repr(self.cell(*self.names[m.group(0)])), expr)
            expr = _LOCAL.sub(lambda m: repr(self.cell(sheet, m.group(1))), expr)
            return eval(expr, {"__builtins__": {}}, _FUNCS)  # noqa: S307 - sandboxed
        return raw


# ── Helpers ──────────────────────────────────────────────────────────────────


def _wb(anchor):
    payload = anchor()
    result = run_lbo(payload)
    wb = load_workbook(
        BytesIO(generate_lbo_formula_excel(payload, result).getvalue()), data_only=False
    )
    return wb, payload, result


def _formula_count(wb):
    return sum(
        1
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    )


def _row_addr(ws, label, col_value=2):
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == label:
            return ws.cell(r, col_value)
    raise AssertionError(f"label {label!r} not found on {ws.title}")


def _all_strings(wb):
    return {
        c.value
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str)
    }


# ── 1. Formula-native contract ───────────────────────────────────────────────


@pytest.mark.parametrize("anchor", [clean_anchor, multi_tranche_anchor])
def test_formula_workbook_has_formulas(anchor):
    wb, _, _ = _wb(anchor)
    assert _formula_count(wb) > 0


def test_legacy_values_only_path_remains_zero_formula():
    """The retained legacy/oracle path must stay values-only (0 formulas)."""
    payload = clean_anchor()
    wb = load_workbook(
        BytesIO(generate_lbo_values_only_excel(payload, run_lbo(payload)).getvalue()),
        data_only=False,
    )
    assert _formula_count(wb) == 0


@pytest.mark.parametrize("anchor", [clean_anchor, multi_tranche_anchor])
def test_no_external_links(anchor):
    wb, _, _ = _wb(anchor)
    assert not getattr(wb, "_external_links", [])


def test_sheetnames_match_legacy_layout():
    wb, _, _ = _wb(clean_anchor)
    assert wb.sheetnames == SHEET_NAMES


def test_core_sheets_each_contain_formulas():
    wb, _, _ = _wb(clean_anchor)
    for sheet in ("Sources & Uses", "Operating Forecast", "Debt Schedule", "Returns Summary"):
        has = any(
            isinstance(c.value, str) and c.value.startswith("=")
            for row in wb[sheet].iter_rows()
            for c in row
        )
        assert has, f"{sheet} must contain Excel formulas"


def test_multi_tranche_core_returns_and_sources_have_formulas():
    wb, _, _ = _wb(multi_tranche_anchor)
    for sheet in ("Sources & Uses", "Returns Summary"):
        has = any(
            isinstance(c.value, str) and c.value.startswith("=")
            for row in wb[sheet].iter_rows()
            for c in row
        )
        assert has, f"{sheet} must contain Excel formulas in multi-tranche mode"


# ── 2. Cover / Audit disclosure wording reflects formula-native direction ─────


def test_cover_no_longer_calls_workbook_values_only():
    wb, _, _ = _wb(clean_anchor)
    cover = [c.value for row in wb["Cover"].iter_rows() for c in row if isinstance(c.value, str)]
    assert FORMULA_WORKBOOK_VERSION in cover
    assert not any("displays model output as values only" in v for v in cover)
    assert any("live Excel formulas" in v for v in cover)


def test_model_map_reflects_formula_native_state():
    """V4.1.1.1: the formula workbook Model Map must describe the live formula
    chain, not the stale 'No Excel recalculation' / Python-precomputed wording."""
    for anchor in (clean_anchor, multi_tranche_anchor):
        wb, _, _ = _wb(anchor)
        cells = [c.value for row in wb["Model Map"].iter_rows()
                 for c in row if isinstance(c.value, str)]
        blob = " ".join(cells)
        assert "No Excel recalculation" not in blob
        assert "Live Excel formulas" in blob
        assert "cash sweep" in blob
        assert "formula" in blob.lower()
        # run_lbo() is still disclosed as the export oracle.
        assert "run_lbo()" in blob


def test_audit_discloses_formula_native_multi_tranche_scope():
    """V4.1.1: the audit must no longer say the multi-tranche waterfall is
    Python-precomputed; it must disclose it as formula-driven and still flag the
    remaining values-only presentation layers."""
    wb, _, _ = _wb(multi_tranche_anchor)
    audit = [c.value for row in wb["Audit & Disclosures"].iter_rows()
             for c in row if isinstance(c.value, str)]
    blob = " ".join(audit)
    assert "run_lbo()" in blob and "oracle" in blob
    assert "live Excel formulas" in blob
    # The stale foundation-version disclosure is gone.
    assert "Multi-tranche debt waterfall detail remains Python-precomputed" not in blob
    assert "only single-tranche debt is formula-driven" not in blob
    # The new disclosure states the waterfall is formula-driven.
    assert "multi-tranche debt waterfall is now formula-driven" in blob
    assert "cash sweep" in blob
    # Scenario / Sensitivity stay Python-precomputed.
    assert "Scenario Summary and Sensitivity Analysis remain Python-precomputed" in blob


# ── 3. Returns / Sources reference upstream, not hardcoded numbers ───────────


def test_returns_outputs_are_formulas_referencing_upstream():
    wb, _, _ = _wb(clean_anchor)
    ws = wb["Returns Summary"]
    moic = _row_addr(ws, "MOIC").value
    irr = _row_addr(ws, "IRR (annualized)").value
    exit_eq = _row_addr(ws, "Exit Equity Value").value  # headline copy
    assert isinstance(moic, str) and moic.startswith("=")
    assert "LBO_Exit_Equity" in moic and "LBO_Sponsor_Equity" in moic
    assert isinstance(irr, str) and "POWER(LBO_MOIC" in irr and "LBO_Holding_Period" in irr
    # Exit bridge EV is a formula off exit EBITDA x exit multiple.
    ev = _row_addr(ws, "Exit Enterprise Value").value
    assert isinstance(ev, str) and ev.startswith("=") and "B4*B5" in ev.replace(" ", "")


def test_sources_uses_check_row_is_formula():
    wb, _, _ = _wb(clean_anchor)
    ws = wb["Sources & Uses"]
    diff = _row_addr(ws, "Sources - Uses").value
    sponsor = _row_addr(ws, "Sponsor Equity").value
    assert isinstance(diff, str) and diff.startswith("=")
    # Sponsor equity is derived (Total Uses - Total Debt), not a literal.
    assert isinstance(sponsor, str) and "LBO_Total_Debt" in sponsor


def test_operating_forecast_year_columns_are_continuous_formula_chain():
    wb, payload, _ = _wb(clean_anchor)
    ws = wb["Operating Forecast"]
    n = len(payload["operating_forecast"]["years"])
    cfbds = _row_addr(ws, "Cash Flow Before Debt Service")
    margin = _row_addr(ws, "EBITDA Margin (%)")
    # Each forecast year column carries a formula (not a hardcoded Python value).
    for i in range(n):
        col = 2 + i
        assert isinstance(ws.cell(cfbds.row, col).value, str)
        assert ws.cell(cfbds.row, col).value.startswith("=")
        assert isinstance(ws.cell(margin.row, col).value, str)
        assert ws.cell(margin.row, col).value.startswith("=")


def test_single_tranche_debt_schedule_is_formula_driven():
    wb, payload, _ = _wb(clean_anchor)
    ws = wb["Debt Schedule"]
    n = len(payload["operating_forecast"]["years"])
    ending = _row_addr(ws, "Ending Debt")
    beginning = _row_addr(ws, "Beginning Debt")
    # Year 1 beginning debt anchors to opening debt; later years carry forward.
    assert ws.cell(beginning.row, 2).value == "=LBO_Opening_Debt"
    assert ws.cell(beginning.row, 3).value.startswith("=B")  # prior-column ending
    for i in range(n):
        assert ws.cell(ending.row, 2 + i).value.startswith("=")


# ── 4. Python oracle tie-out (numeric, evaluator-based) ──────────────────────


def test_single_tranche_formulas_tie_to_run_lbo_oracle():
    wb, _, result = _wb(clean_anchor)
    ev = _Eval(wb)
    rs = "Returns Summary"
    su = "Sources & Uses"
    ds = "Debt Schedule"

    moic = ev.cell(rs, _row_addr(wb[rs], "MOIC").coordinate)
    irr = ev.cell(rs, _row_addr(wb[rs], "IRR (annualized)").coordinate)
    exit_eq = ev.cell(rs, _row_addr(wb[rs], "Exit Equity Value").coordinate)
    sponsor = ev.cell(su, _row_addr(wb[su], "Sponsor Equity").coordinate)
    diff = ev.cell(su, _row_addr(wb[su], "Sources - Uses").coordinate)

    ret = result["returns"]
    assert moic == pytest.approx(ret["moic"], rel=1e-9)
    assert irr == pytest.approx(ret["irr"], rel=1e-6)
    assert exit_eq == pytest.approx(ret["exit_equity_value"], rel=1e-9)
    assert sponsor == pytest.approx(ret["sponsor_equity"], rel=1e-9)
    assert diff == pytest.approx(0.0, abs=1e-6)

    # Final-year ending debt formula ties to remaining debt.
    n = len(result["operating_forecast"]["years"])
    from openpyxl.utils import get_column_letter

    end_col = get_column_letter(2 + n - 1)
    ending = _row_addr(wb[ds], "Ending Debt").row
    assert ev.cell(ds, f"{end_col}{ending}") == pytest.approx(
        ret["remaining_debt"], rel=1e-9
    )


def test_assumption_anchors_match_payload_inputs():
    wb, payload, result = _wb(clean_anchor)
    ev = _Eval(wb)
    names = {nm: ev.cell(*loc) for nm, loc in ev.names.items()}
    ts = result["transaction_summary"]
    assert names["LBO_Entry_EBITDA"] == pytest.approx(ts["entry_ebitda"])
    assert names["LBO_Entry_Multiple"] == pytest.approx(ts["entry_multiple"])
    assert names["LBO_Exit_Multiple"] == pytest.approx(ts["exit_multiple"])
    assert names["LBO_Holding_Period"] == pytest.approx(ts["exit_year"])
    assert names["LBO_Opening_Debt"] == pytest.approx(ts["debt_amount"])
    assert names["LBO_Interest_Rate"] == pytest.approx(payload["debt"]["interest_rate"])


def test_non_ok_result_raises():
    with pytest.raises(ValueError):
        generate_lbo_formula_excel({}, {"status": "error", "returns": None})


# ── 5. V4.1.1 multi-tranche formula waterfall ────────────────────────────────


def _row_addr_contains(ws, needle, col_value=2):
    """First row whose col-A label contains ``needle`` (substring match)."""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and needle in v:
            return ws.cell(r, col_value)
    raise AssertionError(f"label containing {needle!r} not found on {ws.title}")


def _set_named(wb, ev, name, value):
    """Mutate the cell backing a named range (for assumption-mutation tests)."""
    sheet, ref = ev.names[name]
    wb[sheet][ref] = value


def test_multi_tranche_formula_count_materially_increased():
    """V4.1.1 must add a real waterfall formula chain, far above the V4.1.0
    foundation count of ~22 multi-tranche formulas."""
    wb, _, _ = _wb(multi_tranche_anchor)
    assert _formula_count(wb) > 100


def test_multi_tranche_debt_schedule_is_formula_driven():
    wb, payload, _ = _wb(multi_tranche_anchor)
    ws = wb["Debt Schedule"]
    n = len(payload["operating_forecast"]["years"])
    # The per-tranche waterfall lines and the totals chain are all formulas.
    for needle in ("Beginning Balance", "Cash Interest", "Mandatory Amortization",
                   "Optional Repayment", "Ending Balance"):
        cell = _row_addr_contains(ws, needle)
        for i in range(n):
            v = ws.cell(cell.row, 2 + i).value
            assert isinstance(v, str) and v.startswith("="), f"{needle} col {i} not a formula"
    for label in ("Total Ending Debt", "Ending Cash", "Total Cash Interest",
                  "Revolver Draw", "Cash Available for Sweep", "Net Debt"):
        cell = _row_addr(ws, label)
        for i in range(n):
            v = ws.cell(cell.row, 2 + i).value
            assert isinstance(v, str) and v.startswith("="), f"{label} col {i} not a formula"


def test_multi_tranche_interest_uses_beginning_not_ending_balance():
    """Interest must reference the same-row Beginning Balance, never Ending
    Balance, so there is no circular reference."""
    wb, payload, _ = _wb(multi_tranche_anchor)
    ws = wb["Debt Schedule"]
    n = len(payload["operating_forecast"]["years"])
    beg = _row_addr_contains(ws, "Beginning Balance").row
    interest = _row_addr_contains(ws, "Cash Interest").row
    ending = _row_addr_contains(ws, "Ending Balance").row
    for i in range(n):
        f = ws.cell(interest, 2 + i).value
        col = chr(ord("B") + i)
        assert f"{col}{beg}" in f.replace(" ", "")
        assert f"{col}{ending}" not in f.replace(" ", "")


def test_multi_tranche_named_ranges_present():
    wb, payload, _ = _wb(multi_tranche_anchor)
    names = set(wb.defined_names)
    assert "LBO_Total_Ending_Debt" in names
    assert "LBO_Ending_Cash" in names
    assert "LBO_Min_Cash" in names
    assert "LBO_Cash_Sweep_On" in names
    # One opening/rate/amort/commitment anchor per tranche.
    n_tranches = len(payload["capital_structure"]["tranches"])
    for i in range(n_tranches):
        for prefix in ("Opening", "Rate", "Amort", "Commit"):
            assert f"LBO_Tr{i}_{prefix}" in names, f"missing LBO_Tr{i}_{prefix}"


def test_multi_tranche_returns_reference_formula_outputs_not_literals():
    wb, _, _ = _wb(multi_tranche_anchor)
    ws = wb["Returns Summary"]
    remaining = _row_addr(ws, "Less: Remaining Debt").value
    ending_cash = _row_addr(ws, "Plus: Ending Cash").value
    exit_ebitda = _row_addr(ws, "Exit EBITDA").value
    moic = _row_addr(ws, "MOIC").value
    # Remaining debt / ending cash come from the formula-driven named ranges.
    assert remaining == "=LBO_Total_Ending_Debt"
    assert ending_cash == "=LBO_Ending_Cash"
    # Exit EBITDA links to the Operating Forecast, not a hard-coded number.
    assert isinstance(exit_ebitda, str) and exit_ebitda.startswith("=") and "Operating Forecast" in exit_ebitda
    assert "LBO_Exit_Equity" in moic and "LBO_Sponsor_Equity" in moic


def test_multi_tranche_no_circular_reference_and_ties_to_oracle():
    wb, _, result = _wb(multi_tranche_anchor)
    ev = _Eval(wb)  # raises on any circular reference
    cs = result["capital_structure_summary"]
    ret = result["returns"]
    ted = ev.cell(*ev.names["LBO_Total_Ending_Debt"])
    ec = ev.cell(*ev.names["LBO_Ending_Cash"])
    exit_eq = ev.cell(*ev.names["LBO_Exit_Equity"])
    moic = ev.cell(*ev.names["LBO_MOIC"])
    irr = ev.cell(rs := "Returns Summary", _row_addr(wb[rs], "IRR (annualized)").coordinate)
    assert ted == pytest.approx(cs["total_ending_debt"], rel=1e-9)
    assert ec == pytest.approx(cs["ending_cash_balance"], abs=1e-6)
    assert exit_eq == pytest.approx(ret["exit_equity_value"], rel=1e-9)
    assert moic == pytest.approx(ret["moic"], rel=1e-9)
    assert irr == pytest.approx(ret["irr"], rel=1e-6)


def test_multi_tranche_cash_build_fixture_ties_to_oracle():
    """A fixture where the sweep fully repays debt and cash accumulates exercises
    the ending-cash branch of the formula waterfall."""
    from tests.test_lbo_multitranche_schedule import (
        low_leverage_high_fcf_cash_build_fixture,
    )
    payload = low_leverage_high_fcf_cash_build_fixture()
    result = run_lbo(payload)
    wb = load_workbook(
        BytesIO(generate_lbo_formula_excel(payload, result).getvalue()), data_only=False
    )
    ev = _Eval(wb)
    cs = result["capital_structure_summary"]
    assert ev.cell(*ev.names["LBO_Total_Ending_Debt"]) == pytest.approx(0.0, abs=1e-6)
    assert ev.cell(*ev.names["LBO_Ending_Cash"]) == pytest.approx(cs["ending_cash_balance"], rel=1e-9)
    assert ev.cell(*ev.names["LBO_Exit_Equity"]) == pytest.approx(result["returns"]["exit_equity_value"], rel=1e-9)


def test_multi_tranche_revolver_draw_fixture_ties_to_oracle():
    """The revolver-draw fixture gates the sweep off; the formula must reproduce
    the engine's residual debt and zero sweep."""
    from tests.test_lbo_multitranche_schedule import revolver_draw_fixture
    payload = revolver_draw_fixture()
    result = run_lbo(payload)
    wb = load_workbook(
        BytesIO(generate_lbo_formula_excel(payload, result).getvalue()), data_only=False
    )
    ev = _Eval(wb)
    cs = result["capital_structure_summary"]
    assert ev.cell(*ev.names["LBO_Total_Ending_Debt"]) == pytest.approx(cs["total_ending_debt"], rel=1e-9)
    assert ev.cell(*ev.names["LBO_MOIC"]) == pytest.approx(result["returns"]["moic"], rel=1e-9)


def test_multi_tranche_assumption_mutation_flows_through():
    """Editing a debt-related assumption must move the formula-driven Debt
    Schedule, Exit Equity and MOIC — proving the chain is live, not dead values."""
    wb, payload, _ = _wb(multi_tranche_anchor)
    base = _Eval(wb)
    base_moic = base.cell(*base.names["LBO_MOIC"])
    base_ted = base.cell(*base.names["LBO_Total_Ending_Debt"])
    base_exit = base.cell(*base.names["LBO_Exit_Equity"])

    # Bump Term Loan A's interest rate (tranche index 1) and the exit multiple.
    _set_named(wb, base, "LBO_Tr1_Rate", payload["capital_structure"]["tranches"][1]["interest_rate"] + 0.05)
    _set_named(wb, base, "LBO_Exit_Multiple", payload["transaction"]["exit_multiple"] + 1.0)

    bumped = _Eval(wb)
    new_moic = bumped.cell(*bumped.names["LBO_MOIC"])
    new_ted = bumped.cell(*bumped.names["LBO_Total_Ending_Debt"])
    new_exit = bumped.cell(*bumped.names["LBO_Exit_Equity"])

    # Higher interest -> less cash to sweep -> more ending debt.
    assert new_ted > base_ted + 1e-6
    # Higher exit multiple -> larger exit equity; net effect must move MOIC.
    assert new_exit != pytest.approx(base_exit)
    assert new_moic != pytest.approx(base_moic)


def test_multi_tranche_amortization_only_mutation_increases_paydown():
    """Raising mandatory amortization on a term loan must reduce ending debt,
    proving the mandatory-amortization formula is live."""
    wb, payload, _ = _wb(multi_tranche_anchor)
    base = _Eval(wb)
    base_ted = base.cell(*base.names["LBO_Total_Ending_Debt"])
    # Term Loan A (index 1) amortization 5% -> 20%.
    _set_named(wb, base, "LBO_Tr1_Amort", 0.20)
    bumped = _Eval(wb)
    assert bumped.cell(*bumped.names["LBO_Total_Ending_Debt"]) < base_ted - 1e-6
