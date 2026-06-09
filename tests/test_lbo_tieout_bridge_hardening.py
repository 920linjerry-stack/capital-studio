"""V4.7.3 LBO tie-out and bridge hardening tests.

This layer exposes source-of-truth tie-outs and intermediate variables. It does
not add forecast models, ownership modules, refinancing logic, or narrative
recommendation language.
"""

import copy
import math
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from modeling.lbo_calculator import run_lbo
from modeling.lbo_excel_exporter import generate_lbo_excel
from tests.test_lbo_calculator_core import clean_anchor
from tests.test_lbo_multitranche_schedule import (
    low_leverage_high_fcf_cash_build_fixture,
    mandatory_optional_cap_fixture,
    multi_tranche_anchor,
)


STATIC = Path(__file__).resolve().parent.parent / "static" / "modeling"
HTML = (STATIC / "lbo.html").read_text(encoding="utf-8")
JS = (STATIC / "js" / "lbo.js").read_text(encoding="utf-8")


def _workbook(payload):
    result = run_lbo(payload)
    assert result["status"] == "ok"
    return load_workbook(BytesIO(generate_lbo_excel(payload, result).getvalue()), data_only=True), result


def _pairs(ws):
    out = {}
    for row in ws.iter_rows():
        if len(row) >= 2 and isinstance(row[0].value, str):
            out[row[0].value] = row[1].value
    return out


def _strings(wb):
    sheets = wb.worksheets if hasattr(wb, "worksheets") else [wb]
    return {
        c.value
        for ws in sheets
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str)
    }


def test_multi_tranche_opening_debt_equals_sum_selected_tranches():
    payload = multi_tranche_anchor()
    out = run_lbo(payload)
    tranche_sum = sum(t["opening_balance"] for t in payload["capital_structure"]["tranches"])
    assert out["transaction_summary"]["debt_amount"] == pytest.approx(tranche_sum)
    assert out["capital_structure_summary"]["total_opening_debt"] == pytest.approx(tranche_sum)


def test_multi_tranche_implied_leverage_uses_tranche_sum():
    out = run_lbo(multi_tranche_anchor())
    ts = out["transaction_summary"]
    assert ts["implied_leverage"] == pytest.approx(ts["debt_amount"] / ts["entry_ebitda"])


def test_multi_tranche_stale_debt_amount_does_not_override_tranche_sum():
    payload = multi_tranche_anchor()
    payload["debt"] = {"debt_amount": 123.0}
    out = run_lbo(payload)
    assert out["status"] == "ok"
    assert out["transaction_summary"]["debt_amount"] == pytest.approx(5000.0)
    assert any(f["code"] == "TRANCHE_DEBT_SOURCE_OF_TRUTH" for f in out["flags"])


def test_single_tranche_behavior_unchanged():
    out = run_lbo(clean_anchor())
    assert out["transaction_summary"]["debt_amount"] == pytest.approx(5000.0)
    assert "capital_structure_summary" not in out
    assert out["debt_schedule"][0]["cash_available_for_debt"] == pytest.approx(280.0)


def test_excel_sources_uses_has_uses_sources_check_sections():
    wb, _ = _workbook(multi_tranche_anchor())
    strings = _strings(wb)
    assert {"Uses", "Sources", "Check"} <= strings


def test_excel_sources_uses_ties_with_cash_to_balance_sheet():
    payload = multi_tranche_anchor()
    payload["debt"] = {"cash_to_balance_sheet": 40.0}
    payload["capital_structure"]["cash_balance_beginning"] = 40.0
    wb, result = _workbook(payload)
    p = _pairs(wb["Sources & Uses"])
    ts = result["transaction_summary"]
    assert p["Total Uses"] == pytest.approx(ts["entry_ev"] + ts["transaction_fees"] + ts["cash_to_balance_sheet"])
    assert p["Total Sources"] == pytest.approx(ts["debt_amount"] + ts["sponsor_equity"])
    assert p["Sources - Uses"] == pytest.approx(0.0)
    assert p["Cash to Balance Sheet"] == pytest.approx(40.0)


def test_deferred_sources_uses_items_are_compact_note_only():
    wb, _ = _workbook(clean_anchor())
    strings = "\n".join(str(s) for s in _strings(wb))
    assert "Existing debt repayment, rollover equity and seller notes are deferred scope." in strings
    assumptions = _strings(wb["Assumptions"])
    assert "Existing Debt Repaid" not in assumptions
    assert "Rollover Equity" not in assumptions
    assert "Seller Note" not in assumptions


def test_exit_equity_bridge_ties_single_tranche():
    out = run_lbo(clean_anchor())
    ex = out["exit"]
    assert ex["exit_ev"] == pytest.approx(ex["exit_ebitda"] * ex["exit_multiple"])
    assert ex["exit_equity_value"] == pytest.approx(ex["exit_ev"] - ex["remaining_debt"] + ex["ending_cash_balance"])
    assert ex["ending_cash_balance"] == 0.0


def test_exit_equity_bridge_ties_multi_tranche_with_cash():
    out = run_lbo(low_leverage_high_fcf_cash_build_fixture())
    ex = out["exit"]
    assert ex["ending_cash_balance"] > 0
    assert ex["exit_ev"] == pytest.approx(ex["exit_ebitda"] * ex["exit_multiple"])
    assert ex["exit_equity_value"] == pytest.approx(ex["exit_ev"] - ex["remaining_debt"] + ex["ending_cash_balance"])


def test_ui_and_excel_exit_equity_bridge_labels_exist():
    wb, _ = _workbook(multi_tranche_anchor())
    strings = _strings(wb)
    assert "Exit Equity Bridge" in strings
    assert "Exit Enterprise Value" in strings
    assert "Less: Remaining Debt" in strings
    assert "Plus: Ending Cash" in strings
    assert "退出股权价值桥 / Exit Equity Bridge" in HTML
    assert "renderExitBridge" in JS


def test_cash_flow_before_debt_service_formula_ties():
    out = run_lbo(clean_anchor())
    of = out["operating_forecast"]
    for i, row in enumerate(out["debt_schedule"]):
        expected = of["ebitda"][i] - of["cash_taxes"][i] - of["capex"][i] - of["change_in_nwc"][i]
        assert row["cash_flow_before_debt_service"] == pytest.approx(expected)


def test_cash_after_interest_and_mandatory_formula_ties():
    out = run_lbo(multi_tranche_anchor())
    for row in out["debt_schedule"]:
        expected = (
            row["cash_flow_before_debt_service"]
            - row["total_cash_interest"]
            - row["total_mandatory_amortization"]
        )
        assert row["cash_after_interest_and_mandatory_amortization"] == pytest.approx(expected)


def test_optional_repayment_cap_still_works():
    out = run_lbo(mandatory_optional_cap_fixture())
    assert out["status"] == "ok"
    for row in out["debt_schedule"]:
        for tr in row["tranches"]:
            assert tr["mandatory_amortization"] + tr["optional_repayment"] <= tr["beginning_balance"] + tr["draw"] + 1e-9
            assert tr["ending_balance"] >= -1e-9


def test_ui_and_excel_cash_flow_bridge_labels_exist():
    wb, _ = _workbook(multi_tranche_anchor())
    strings = _strings(wb)
    assert "Cash Flow Before Debt Service" in strings
    assert "Cash After Interest & Mandatory Amort." in strings
    assert "Optional Repayment" in strings
    assert "cash_flow_before_debt_service" in JS
    assert "cash_after_interest_and_mandatory_amortization" in JS


def test_covenant_forecast_ebitda_label_and_note_present():
    wb, _ = _workbook(multi_tranche_anchor())
    strings = "\n".join(str(s) for s in _strings(wb))
    assert "Net Debt / Forecast EBITDA" in strings
    assert "Uses forecast-year EBITDA." in HTML
    assert "使用预测年度 EBITDA" in HTML


def test_covenant_math_unchanged():
    payload = multi_tranche_anchor()
    before = run_lbo(payload)["covenant_summary"]["checks"][0]["net_debt_ebitda"]
    after = run_lbo(copy.deepcopy(payload))["covenant_summary"]["checks"][0]["net_debt_ebitda"]
    expected = after
    first = run_lbo(payload)["debt_schedule"][0]
    assert before == pytest.approx(expected)
    assert after == pytest.approx((first["total_ending_debt"] - first["ending_cash_balance"]) / payload["operating_forecast"]["ebitda"][0])


def test_no_recommendation_or_deferred_scope_modules_added():
    combined = "\n".join([HTML, JS])
    banned = [
        "should acquire", "should buy", "good deal", "bad deal", "recommended deal",
        "management rollover", "option pool", "dividend recap", "probability weighting",
    ]
    lowered = combined.lower()
    for term in banned:
        assert term not in lowered


def test_no_long_new_explanation_blocks_added():
    # V4.7.3 should expose compact labels / notes, not turn the UI into a memo.
    assert "Exit EV = Exit EBITDA" not in HTML
    assert "Cash Flow Before Debt Service = EBITDA" not in HTML
