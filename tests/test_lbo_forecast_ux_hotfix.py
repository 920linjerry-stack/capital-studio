"""V4.7 RC Forecast UX Hotfix tests.

Covers the forecast-input UX redirection only: Flat Forecast is no longer a
user-visible mode, Growth Forecast is the default driver-based path, and Manual
Year-by-Year is edited via an overlay. No engine math is changed -- run_lbo /
waterfall / covenant / attribution / scenario logic is untouched. These tests
assert the UI/markup contract, the Excel forecast-source label, and engine-level
behavior of a growth-style forecast (no "death LBO" with flat 5-year EBITDA).
"""

import re
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from modeling.lbo_calculator import run_lbo
from modeling.lbo_excel_exporter import generate_lbo_excel
from tests.test_lbo_calculator_core import clean_anchor


STATIC = Path(__file__).resolve().parent.parent / "static" / "modeling"
HTML = (STATIC / "lbo.html").read_text(encoding="utf-8")
JS = (STATIC / "js" / "lbo.js").read_text(encoding="utf-8")


def _growth_forecast(rev_y1, growth, margin, tax_pct_ebitda, capex_pct_rev, nwc_pct_rev, n=5):
    """Python mirror of the JS generateGrowthForecast() driver logic, used to
    validate engine behavior on a growth-style forecast."""
    revenue, ebitda, cash_taxes, capex, nwc = [], [], [], [], []
    rev = rev_y1
    for i in range(n):
        if i > 0:
            rev = rev * (1 + growth)
        e = rev * margin
        revenue.append(rev)
        ebitda.append(e)
        cash_taxes.append(e * tax_pct_ebitda)
        capex.append(rev * capex_pct_rev)
        nwc.append(rev * nwc_pct_rev)
    return {
        "years": list(range(1, n + 1)),
        "revenue": revenue, "ebitda": ebitda,
        "cash_taxes": cash_taxes, "capex": capex, "change_in_nwc": nwc,
    }


# ── 1. Flat Forecast not user-visible; section renamed ───────────────────────


def test_flat_forecast_not_a_visible_section():
    assert "Flat Forecast" not in HTML
    assert "Operating Forecast" in HTML


def test_legacy_flat_input_ids_removed_from_ui():
    # The old single-value flat inputs are gone from the form.
    for old_id in ['id="revenue"', 'id="ebitda"', 'id="cash_taxes"', 'id="capex"', 'id="change_in_nwc"']:
        assert old_id not in HTML, f"legacy flat input {old_id} still present"


# ── 2. Growth Forecast is the default mode ───────────────────────────────────


def test_growth_is_default_mode():
    assert 'let forecastMode = "growth"' in JS
    assert 'id="forecast-mode-indicator">Growth Forecast' in HTML


def test_growth_driver_inputs_present():
    for fid in ["rev_y1", "rev_growth", "ebitda_margin", "tax_pct_ebitda",
                "capex_pct_rev", "nwc_pct_rev"]:
        assert f'id="{fid}"' in HTML, f"missing growth driver input {fid}"


# ── 3-5. Growth generation engine behavior (no death LBO) ────────────────────


def test_growth_forecast_produces_growing_ebitda():
    fc = _growth_forecast(900, 0.06, 0.22, 0.15, 0.04, 0.02)
    # 5 years, strictly growing EBITDA -- not a flat copy.
    assert len(fc["ebitda"]) == 5
    assert fc["ebitda"][4] > fc["ebitda"][0]
    assert len(set(round(x, 6) for x in fc["ebitda"])) == 5


def test_growth_smoke_case_not_flat_and_entry_ebitda_unmutated():
    fc = _growth_forecast(900, 0.06, 0.22, 0.15, 0.04, 0.02)
    payload = {
        "symbol": "SMOKE", "currency": "USD",
        "transaction": {
            "entry_ebitda": 180, "entry_multiple": 11.0,
            "exit_multiple": 11.0, "exit_year": 5, "transaction_fees_pct_ev": 0.02,
        },
        "operating_forecast": fc,
        "forecast_mode": "growth",
        "debt": {
            "leverage_multiple": 4.0, "interest_rate": 0.09,
            "mandatory_amortization_pct": 0.01, "cash_sweep_pct": 1.0,
            "cash_to_balance_sheet": 0.0,
        },
    }
    result = run_lbo(payload)
    assert result["status"] == "ok"
    # Entry EBITDA used for sizing is unchanged; debt = 180 x 4.0 = 720.
    assert result["transaction_summary"]["entry_ebitda"] == 180
    assert result["transaction_summary"]["debt_amount"] == pytest.approx(720.0)
    # Exit uses the exit-year FORECAST EBITDA (grown), not Entry EBITDA.
    assert result["exit"]["exit_ebitda"] == pytest.approx(fc["ebitda"][-1])
    assert result["exit"]["exit_ebitda"] > 180


def test_exit_ebitda_is_exit_year_forecast():
    payload = clean_anchor()
    payload["operating_forecast"]["ebitda"] = [1000, 1100, 1200, 1300, 1400]
    result = run_lbo(payload)
    assert result["exit"]["exit_ebitda"] == 1400


# ── 6-9. Manual forecast overlay markup + JS contract ────────────────────────


def test_manual_overlay_markup_present():
    assert 'id="manual-forecast-overlay"' in HTML
    assert 'id="manual-forecast-table"' in HTML
    assert 'id="manual-forecast-save"' in HTML
    assert 'id="manual-forecast-cancel"' in HTML
    assert 'id="manual-forecast-close"' in HTML
    assert 'id="manual-forecast-btn"' in HTML  # opener button


def test_manual_save_writes_arrays_and_sets_mode():
    assert "function saveManualForecast" in JS
    assert 'forecastMode = "manual"' in JS
    # buildPayload routes forecast arrays through activeForecast()
    assert "operating_forecast: activeForecast()" in JS
    assert "function activeForecast" in JS


def test_cancel_and_close_do_not_save():
    # closeManualForecast only hides the overlay; it must not mutate state.
    assert "function closeManualForecast" in JS
    fn = JS.split("function closeManualForecast", 1)[1].split("}", 1)[0]
    assert "manualForecast" not in fn
    assert 'forecastMode' not in fn


def test_save_shows_compact_success_toast():
    assert "function showForecastToast" in JS
    assert "Forecast saved." in JS
    assert "预测已保存。" in JS
    assert 'id="forecast-toast"' in HTML


def test_manual_active_indicator_text():
    assert "Manual forecast active" in JS
    assert "已使用手动逐年预测" in JS


# ── 10-11. Growth/manual relationship ────────────────────────────────────────


def test_reset_to_growth_present():
    assert "function resetToGrowth" in JS
    assert 'id="reset-growth-btn"' in HTML
    assert 'Reset to Growth Forecast' in HTML


def test_growth_generation_uses_compounding():
    assert "function generateGrowthForecast" in JS
    assert "rev * (1 + g)" in JS
    assert "rev * margin" in JS


# ── 12. Unit labels follow currency ──────────────────────────────────────────


def test_growth_inputs_and_overlay_use_currency_units():
    assert 'Revenue Y1 <span class="unit-tag">(USD mm)</span>' in HTML
    # overlay subtitle + manual rows carry a currency-aware unit note
    assert HTML.count("data-ccy-note") >= 4
    assert 'MANUAL_ROW_LABELS' in JS


# ── 13-14. Debt service stress hint (lightweight, no recommendation) ─────────


def test_debt_stress_hint_present_and_bilingual():
    assert "function renderDebtStressHint" in JS
    assert "Debt service stress" in JS
    assert "债务服务压力" in JS
    assert 'id="debt-stress-hint"' in HTML


def test_debt_stress_hint_triggers_on_engine_failure():
    # A weak early year cannot cover interest + mandatory amortization, while the
    # case still produces valid (ok) returns thanks to a strong later recovery.
    payload = clean_anchor()
    payload["operating_forecast"]["ebitda"] = [100, 1000, 1000, 1000, 1000]
    result = run_lbo(payload)
    assert result["status"] == "ok"
    # The hint is driven by debt_service_failure rows in the schedule.
    assert any(row.get("debt_service_failure") for row in result["debt_schedule"])


def test_forecast_hints_have_no_recommendation_wording():
    # User-facing recommendation/deal-conclusion wording must not appear in the
    # hint copy. (Field names like debt_service_failure are code, not copy.)
    terms = ["bad deal", "good deal", "should acquire", "recommend buying", "buy signal"]
    hint_src = ""
    for fn in ("renderDebtStressHint", "updateForecastModeIndicator", "showForecastToast"):
        hint_src += JS.split(f"function {fn}", 1)[1].split("\nfunction ", 1)[0].lower()
    for term in terms:
        assert term not in hint_src, f"hint copy contains '{term}'"


# ── 15. Excel forecast-source label ──────────────────────────────────────────


def test_excel_cover_shows_growth_forecast_source():
    payload = clean_anchor()
    payload["forecast_mode"] = "growth"
    result = run_lbo(payload)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, result).getvalue()))
    blob = "\n".join(
        c.value for ws in wb.worksheets for row in ws.iter_rows()
        for c in row if isinstance(c.value, str)
    )
    assert "Forecast Source" in blob
    assert "Growth Forecast" in blob
    assert "Flat Forecast" not in blob


def test_excel_cover_shows_manual_forecast_source():
    payload = clean_anchor()
    payload["forecast_mode"] = "manual"
    result = run_lbo(payload)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, result).getvalue()))
    blob = "\n".join(
        c.value for ws in wb.worksheets for row in ws.iter_rows()
        for c in row if isinstance(c.value, str)
    )
    assert "Manual Year-by-Year" in blob
    assert "Flat Forecast" not in blob
