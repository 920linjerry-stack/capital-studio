from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from app import _apply_default_quality_gates, _dcf_input_from_payload
from modeling.data_quality import build_normalized_ebit_review, check_economic_sanity
from modeling.dcf_calculator import run_dcf
from modeling.excel_exporter import generate_excel


def _quality_for(fin, symbol="2359.HK", company="WuXi AppTec", market="HK"):
    return _apply_default_quality_gates(
        symbol,
        company,
        market,
        fin,
        {
            "revenue_growth": 0.05,
            "ebit_margin": fin["ebit"] / fin["revenue"],
            "da_pct_revenue": 0.0784,
            "capex_pct_revenue": 0.1218,
            "wc_change_pct_revenue": -0.0402,
        },
        0.081,
        0.558,
    )


def test_2359_like_normalized_ebit_review_candidate_generated(monkeypatch):
    import app

    monkeypatch.setattr(app, "_financials_cache_meta", lambda symbol: {"exists": True, "stale": False, "cached_at": "test", "path": "test"})
    fin = {
        "revenue": 45456.166,
        "ebit": 23804.896,
        "da": 3563.491,
        "capex": 5538.151,
        "wc_change": -1878.888,
        "net_debt": -21649.993,
    }

    _, _, quality = _quality_for(fin)
    review = quality["normalized_ebit_review"]

    assert review["status"] == "High Review"
    assert review["applied"] is False
    assert review["reported_ebit"] == 23804.896
    assert review["reported_ebit_margin"] == pytest_approx(0.5237, abs=0.0001)
    assert review["candidate_normalized_ebit_high_confidence"] == 23804.896
    assert review["candidate_normalized_ebit_high_plus_medium"] == pytest_approx(14593.55, abs=0.001)
    assert review["candidate_normalized_margin_high_plus_medium"] == pytest_approx(0.3210, abs=0.0001)
    assert review["recommended_candidate_basis"] == "HIGH_PLUS_MEDIUM"
    assert [x["field"] for x in review["adjustments"]] == ["其他收入", "其他收益", "减值及拨备"]


def test_candidate_block_does_not_change_unapplied_valuation():
    payload = {
        "symbol": "TEST",
        "company": "TestCo",
        "price": 100,
        "revenue": 1000,
        "ebit": 500,
        "da": 50,
        "capex": 60,
        "wc_change": 0,
        "tax_rate": 0.21,
        "net_debt": 0,
        "shares": 100,
        "revenue_growth": 0.03,
        "ebit_margin": 0.50,
        "da_pct_revenue": 0.05,
        "capex_pct_revenue": 0.06,
        "wc_change_pct_revenue": 0,
        "wacc": 0.10,
        "beta": 1.0,
        "terminal_g": 0.02,
        "exit_multiple": 12,
        "forecast_years": 5,
        "tv_method": "average",
        "currency": "USD",
    }
    out_plain = run_dcf(_dcf_input_from_payload(payload))
    payload_with_review = {
        **payload,
        "default_quality": {
            "normalized_ebit_review": {
                "recommended_candidate_margin": 0.32,
                "applied": False,
            }
        },
    }
    out_with_review = run_dcf(_dcf_input_from_payload(payload_with_review))

    assert out_with_review.intrinsic_per_share == out_plain.intrinsic_per_share


def test_high_margin_without_adjustment_fields_has_no_candidate():
    raw = {"revenue": 1000, "ebit": 600}
    econ = check_economic_sanity("0700.HK", raw, {"market": "HK", "company": "Tencent"})

    assert "ebit_margin_above_industry_p95" in econ["flags"]
    assert build_normalized_ebit_review("0700.HK", raw, {}, econ) is None


def test_normal_margin_with_other_income_fields_has_no_candidate():
    raw = {
        "revenue": 1000,
        "ebit": 200,
        "normalized_ebit_adjustments": [{"field": "其他收入", "amount": 50}],
    }
    econ = check_economic_sanity("0700.HK", raw, {"market": "HK", "company": "Tencent"})

    assert econ["flags"] == []
    assert build_normalized_ebit_review("0700.HK", raw, {}, econ) is None


def test_frontend_apply_helper_sets_margin_override_and_provenance():
    js = Path("static/modeling/js/dcf.js").read_text(encoding="utf-8")

    assert "function applyNormalizedEbitCandidate()" in js
    assert "setInput(\"p-ebit-margin\", margin * 100, 4)" in js
    assert "markOperatingOverride(\"ebit_margin\")" in js
    assert "source: \"normalized_ebit_candidate\"" in js
    assert "basis: review.recommended_candidate_basis" in js
    assert "operating_override_keys: operatingOverrideKeys" in js


def test_excel_audit_shows_normalized_ebit_not_applied_and_applied():
    payload = {
        "symbol": "TEST",
        "company": "TestCo",
        "price": 100,
        "revenue": 1000,
        "ebit": 500,
        "da": 50,
        "capex": 60,
        "wc_change": 0,
        "tax_rate": 0.21,
        "net_debt": 0,
        "shares": 100,
        "revenue_growth": 0.03,
        "ebit_margin": 0.50,
        "da_pct_revenue": 0.05,
        "capex_pct_revenue": 0.06,
        "wc_change_pct_revenue": 0,
        "wacc": 0.10,
        "beta": 1.0,
        "terminal_g": 0.02,
        "exit_multiple": 12,
        "forecast_years": 5,
        "tv_method": "average",
        "currency": "USD",
    }
    review = {
        "status": "High Review",
        "applied": False,
        "reported_ebit": 500,
        "reported_ebit_margin": 0.50,
        "candidate_normalized_ebit_high_plus_medium": 320,
        "candidate_normalized_margin_high_plus_medium": 0.32,
        "recommended_candidate_margin": 0.32,
        "recommended_candidate_basis": "HIGH_PLUS_MEDIUM",
        "adjustments": [{"field": "其他收入", "amount": 180, "confidence": "MEDIUM"}],
        "warning": "candidate shown for review",
    }

    inp_raw = _dcf_input_from_payload(payload)
    out_raw = run_dcf(inp_raw)
    not_applied_ctx = {
        "current_params": {
            **payload,
            "default_quality": {"normalized_ebit_review": review},
            "normalized_ebit_review": review,
            "currency": "USD",
        }
    }
    wb_raw = load_workbook(BytesIO(generate_excel(inp_raw, out_raw, not_applied_ctx).getvalue()), data_only=False)
    audit_text_raw = "\n".join(str(c.value) for row in wb_raw["Data Sources Audit"].iter_rows() for c in row if c.value is not None)
    assert "No - Not applied to valuation" in audit_text_raw

    applied_payload = {**payload, "ebit_margin": 0.32, "ebit_margin_path": [0.32] * 5}
    inp_applied = _dcf_input_from_payload(applied_payload)
    out_applied = run_dcf(inp_applied)
    applied_ctx = {
        "current_params": {
            **applied_payload,
            "default_quality": {"normalized_ebit_review": review},
            "normalized_ebit_review": review,
            "assumption_provenance": {
                "ebit_margin": {
                    "source": "normalized_ebit_candidate",
                    "basis": "HIGH_PLUS_MEDIUM",
                    "reported_margin": 0.50,
                    "applied_margin": 0.32,
                    "confidence": "MEDIUM",
                    "note": "Applied by user from Normalized EBIT Review Candidate",
                }
            },
            "currency": "USD",
        }
    }
    wb_applied = load_workbook(BytesIO(generate_excel(inp_applied, out_applied, applied_ctx).getvalue()), data_only=False)
    audit_text_applied = "\n".join(str(c.value) for row in wb_applied["Data Sources Audit"].iter_rows() for c in row if c.value is not None)

    assert out_applied.intrinsic_per_share != out_raw.intrinsic_per_share
    assert "Yes - Applied by user" in audit_text_applied
    assert "Applied by user from Normalized EBIT Review Candidate" in audit_text_applied


def pytest_approx(*args, **kwargs):
    import pytest

    return pytest.approx(*args, **kwargs)
