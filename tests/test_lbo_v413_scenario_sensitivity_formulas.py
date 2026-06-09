"""V4.1.3 Scenario & Sensitivity formula-checked / oracle-backed tests.

These assert the V4.1.3 upgrade: Scenario Summary and Sensitivity Analysis are
still Python ``run_lbo`` oracle-backed precomputed grids (their grid math is
unchanged), but the formula-native workbook now augments them with

  * a Scenario Assumption Map, an Output Bridge and live formula tie-out checks
    (MOIC x sponsor = exit equity, debt paydown, component sum = equity value
    creation, base case = main exported case, initial cash funding) plus an IC
    review block;
  * a Sensitivity Metadata / Axis Map and live grid formula checks (dimensions
    match, best / base / worst IRR within range).

The lazy ``_Eval`` resolver (shared with test_lbo_formula_workbook) proves the
check formulas resolve to OK and reproduce the engine oracle without Excel. The
legacy values-only path keeps its 0-formula contract.
"""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from modeling.lbo_calculator import run_lbo
from modeling.lbo_scenarios import build_lbo_scenarios
from modeling.lbo_sensitivity import build_lbo_sensitivity
from modeling.lbo_formula_workbook import generate_lbo_formula_excel
from modeling.lbo_excel_exporter import (
    SCENARIO_SUMMARY_SHEET,
    SENSITIVITY_SHEET,
    generate_lbo_values_only_excel,
)
from tests.test_lbo_calculator_core import clean_anchor
from tests.test_lbo_multitranche_schedule import multi_tranche_anchor
from tests.test_lbo_formula_workbook import _Eval, _formula_count


# ── helpers ──────────────────────────────────────────────────────────────────


def _wb_with_layers(anchor, *, attach_scenarios=True, attach_sensitivity=True):
    payload = anchor()
    result = run_lbo(payload)
    if attach_scenarios:
        payload["scenarios"] = build_lbo_scenarios(anchor())
    if attach_sensitivity:
        payload["sensitivity"] = build_lbo_sensitivity(payload)
    wb = load_workbook(
        BytesIO(generate_lbo_formula_excel(payload, result).getvalue()), data_only=False
    )
    return wb, payload, result


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _label_row(ws, label):
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == label:
            return r
    raise AssertionError(f"label {label!r} not found on {ws.title}")


def _strings(ws):
    return {c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)}


def _check_result_rows(ws):
    """Yield (label, result_coord) for every formula 'Result' (col E) check row."""
    for r in range(1, ws.max_row + 1):
        if _is_formula(ws.cell(r, 5).value):
            yield ws.cell(r, 1).value, f"E{r}"


# ── 1. Scenario Summary structure ────────────────────────────────────────────


@pytest.mark.parametrize("anchor", [clean_anchor, multi_tranche_anchor])
def test_scenario_summary_has_new_sections(anchor):
    wb, _, _ = _wb_with_layers(anchor, attach_sensitivity=False)
    strings = _strings(wb[SCENARIO_SUMMARY_SHEET])
    assert "Scenario Assumption Map" in strings
    assert "Scenario Output Bridge" in strings
    assert "Scenario Tie-out Checks" in strings
    assert "IC Review Notes (model audit)" in strings


@pytest.mark.parametrize("anchor", [clean_anchor, multi_tranche_anchor])
def test_scenario_summary_contains_formulas_but_not_all(anchor):
    """V4.1.3: the sheet carries check formulas, but the comparison metric values
    remain Python-precomputed (not all cells are formulas)."""
    wb, _, _ = _wb_with_layers(anchor, attach_sensitivity=False)
    ws = wb[SCENARIO_SUMMARY_SHEET]
    formulas = [c.value for row in ws.iter_rows() for c in row if _is_formula(c.value)]
    values = [c.value for row in ws.iter_rows() for c in row
              if c.value is not None and not _is_formula(c.value)]
    assert len(formulas) > 5
    assert len(values) > len(formulas)


@pytest.mark.parametrize("anchor", [clean_anchor, multi_tranche_anchor])
def test_scenario_tieout_checks_pass(anchor):
    wb, _, _ = _wb_with_layers(anchor, attach_sensitivity=False)
    ev = _Eval(wb)
    ws = wb[SCENARIO_SUMMARY_SHEET]
    rows = list(_check_result_rows(ws))
    labels = {lbl for lbl, _ in rows}
    # Base-case-ties-to-main and per-scenario MOIC bridge checks are present.
    assert "Base MOIC = Workbook MOIC" in labels
    assert any("MOIC x Sponsor = Exit Equity" in (lbl or "") for lbl in labels)
    assert any("Debt Paydown = Opening - Remaining" in (lbl or "") for lbl in labels)
    for _, coord in rows:
        assert ev.cell(SCENARIO_SUMMARY_SHEET, coord) == "OK"


def test_scenario_base_ties_to_run_lbo_oracle():
    """The Scenario Summary base column outputs (via the output bridge) reproduce
    run_lbo() on the same inputs."""
    wb, _, result = _wb_with_layers(clean_anchor, attach_sensitivity=False)
    ev = _Eval(wb)
    ws = wb[SCENARIO_SUMMARY_SHEET]
    ret = result["returns"]
    moic_row = _label_row(ws, "MOIC")
    sponsor_row = _label_row(ws, "Sponsor Equity")
    # First MOIC / Sponsor Equity labels belong to the output bridge block; the
    # comparison table uses the same labels, so confirm the bridge cell ties out.
    assert ev.cell(SCENARIO_SUMMARY_SHEET, f"B{moic_row}") == pytest.approx(ret["moic"], rel=1e-9)
    assert ev.cell(SCENARIO_SUMMARY_SHEET, f"B{sponsor_row}") == pytest.approx(
        ret["sponsor_equity"], rel=1e-9
    )


def test_scenario_component_sum_check_present_and_ok_for_cash_funding():
    payload = multi_tranche_anchor()
    payload["debt"] = {"cash_to_balance_sheet": 60.0}
    payload["capital_structure"]["cash_balance_beginning"] = 60.0
    payload["capital_structure"]["minimum_cash_balance"] = 80.0
    result = run_lbo(payload)
    payload["scenarios"] = build_lbo_scenarios(payload)
    wb = load_workbook(
        BytesIO(generate_lbo_formula_excel(payload, result).getvalue()), data_only=False
    )
    ev = _Eval(wb)
    ws = wb[SCENARIO_SUMMARY_SHEET]
    labels = {lbl for lbl, _ in _check_result_rows(ws)}
    assert any("Component Sum = Value Creation" in (lbl or "") for lbl in labels)
    assert "Initial Cash Funding = -Cash to BS" in labels
    for _, coord in _check_result_rows(ws):
        assert ev.cell(SCENARIO_SUMMARY_SHEET, coord) == "OK"


def test_scenario_base_tie_breaks_when_inputs_mismatch():
    """If the attached base scenario does not match the main exported case, the
    base tie-out flags CHECK rather than silently passing."""
    payload = clean_anchor()
    result = run_lbo(payload)
    mismatched = clean_anchor()
    mismatched["transaction"]["exit_multiple"] = 14.0  # different transaction
    payload["scenarios"] = build_lbo_scenarios(mismatched)
    wb = load_workbook(
        BytesIO(generate_lbo_formula_excel(payload, result).getvalue()), data_only=False
    )
    ev = _Eval(wb)
    ws = wb[SCENARIO_SUMMARY_SHEET]
    r = _label_row(ws, "Base Exit Equity = Workbook")
    assert ev.cell(SCENARIO_SUMMARY_SHEET, f"E{r}") == "CHECK"


def test_scenario_summary_graceful_with_bare_comparison_dict():
    """A bare comparison dict (no per-scenario inputs) must not crash; the
    assumption map degrades to a note but checks still render."""
    payload = clean_anchor()
    result = run_lbo(payload)
    full = build_lbo_scenarios(clean_anchor())
    payload["scenarios"] = full["comparison"]  # bare comparison dict
    wb = load_workbook(
        BytesIO(generate_lbo_formula_excel(payload, result).getvalue()), data_only=False
    )
    ev = _Eval(wb)
    ws = wb[SCENARIO_SUMMARY_SHEET]
    strings = _strings(ws)
    assert "Scenario Assumption Map" in strings
    assert any("assumption detail is unavailable" in s for s in strings)
    for _, coord in _check_result_rows(ws):
        assert ev.cell(SCENARIO_SUMMARY_SHEET, coord) == "OK"


# ── 2. Sensitivity Analysis structure ────────────────────────────────────────


@pytest.mark.parametrize("anchor", [clean_anchor, multi_tranche_anchor])
def test_sensitivity_has_axis_map_and_grid_checks(anchor):
    wb, _, _ = _wb_with_layers(anchor, attach_scenarios=False)
    strings = _strings(wb[SENSITIVITY_SHEET])
    assert "Sensitivity Metadata / Axis Map" in strings
    assert any("Grid Checks (formula)" in (s or "") for s in strings)
    assert "Dimensions Match" in strings
    assert "Best IRR (max)" in strings
    assert "Base Within IRR Range" in strings


@pytest.mark.parametrize("anchor", [clean_anchor, multi_tranche_anchor])
def test_sensitivity_grid_checks_pass(anchor):
    wb, _, _ = _wb_with_layers(anchor, attach_scenarios=False)
    ev = _Eval(wb)
    ws = wb[SENSITIVITY_SHEET]
    saw_dim = saw_range = False
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label == "Dimensions Match":
            assert ev.cell(SENSITIVITY_SHEET, f"B{r}") == "OK"
            saw_dim = True
        if label == "Base Within IRR Range":
            assert ev.cell(SENSITIVITY_SHEET, f"B{r}") == "OK"
            saw_range = True
    assert saw_dim and saw_range


def test_sensitivity_best_worst_irr_tie_to_oracle_grid():
    """Best / Worst IRR formula summaries equal the max / min over the Python
    oracle grid's OK IRR cells — the grid math is unchanged."""
    wb, payload, _ = _wb_with_layers(clean_anchor, attach_scenarios=False)
    ev = _Eval(wb)
    ws = wb[SENSITIVITY_SHEET]
    sens = build_lbo_sensitivity(payload)
    grid = sens["grids"]["entry_exit"]
    irrs = [c["irr"] for line in grid["cells"] for c in line
            if c.get("status") == "ok" and isinstance(c.get("irr"), (int, float))]
    # The first grid's Best / Worst IRR rows (entry_exit) come before leverage_exit.
    best_r = _label_row(ws, "Best IRR (max)")
    worst_r = _label_row(ws, "Worst IRR (min)")
    assert ev.cell(SENSITIVITY_SHEET, f"B{best_r}") == pytest.approx(max(irrs), rel=1e-9)
    assert ev.cell(SENSITIVITY_SHEET, f"B{worst_r}") == pytest.approx(min(irrs), rel=1e-9)


def test_sensitivity_grid_values_unchanged_vs_oracle():
    """The displayed IRR grid still equals the Python sensitivity oracle output."""
    wb, payload, _ = _wb_with_layers(clean_anchor, attach_scenarios=False)
    ws = wb[SENSITIVITY_SHEET]
    sens = build_lbo_sensitivity(payload)
    base_cell = next(
        c for line in sens["grids"]["entry_exit"]["cells"] for c in line if c["is_base"]
    )
    # The base cell IRR must appear verbatim somewhere on the sheet (values-only grid).
    sheet_numbers = [c.value for row in ws.iter_rows() for c in row
                     if isinstance(c.value, (int, float))]
    assert any(abs(n - base_cell["irr"]) < 1e-12 for n in sheet_numbers)


# ── 3. Legacy / default path discipline ──────────────────────────────────────


def test_legacy_scenario_sensitivity_stay_values_only():
    """generate_lbo_values_only_excel keeps its 0-formula contract even with
    scenario + sensitivity layers attached."""
    payload = multi_tranche_anchor()
    result = run_lbo(payload)
    payload["scenarios"] = build_lbo_scenarios(multi_tranche_anchor())
    payload["sensitivity"] = build_lbo_sensitivity(payload)
    wb = load_workbook(
        BytesIO(generate_lbo_values_only_excel(payload, result).getvalue()), data_only=False
    )
    assert _formula_count(wb) == 0


def test_audit_discloses_v413_scenario_sensitivity_checks():
    wb, _, _ = _wb_with_layers(multi_tranche_anchor)
    blob = " ".join(
        c.value for row in wb["Audit & Disclosures"].iter_rows()
        for c in row if isinstance(c.value, str)
    )
    # Required precomputed substring preserved.
    assert "Scenario Summary and Sensitivity Analysis remain Python-precomputed" in blob
    # New V4.1.3 disclosure present, not misrepresented as live Data Tables.
    assert "V4.1.3" in blob
    assert "not live Excel Data Tables" in blob
    assert "axis map" in blob
