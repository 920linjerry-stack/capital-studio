from io import BytesIO

from openpyxl import load_workbook

from app import _dcf_input_from_payload
from modeling.dcf_calculator import run_dcf
from modeling.excel_exporter import generate_excel
from modeling.trading_comps import build_trading_comps, get_trading_comps_profile


TECH_FALLBACK = {"AAPL", "MSFT", "GOOGL", "META", "AMZN", "NVDA"}


def _payload(symbol="TEST", company="TestCo"):
    return {
        "symbol": symbol,
        "company": company,
        "price": 100,
        "revenue": 1000,
        "ebit": 250,
        "da": 40,
        "capex": 50,
        "wc_change": 0,
        "tax_rate": 0.21,
        "net_debt": 0,
        "shares": 100,
        "revenue_growth": 0.03,
        "ebit_margin": 0.25,
        "da_pct_revenue": 0.04,
        "capex_pct_revenue": 0.05,
        "wc_change_pct_revenue": 0.0,
        "wacc": 0.10,
        "beta": 1.0,
        "terminal_g": 0.02,
        "exit_multiple": 12,
        "forecast_years": 5,
        "tv_method": "average",
        "currency": "USD",
    }


def _workbook_for(symbol, trading_comps):
    payload = _payload(symbol=symbol, company=symbol)
    inp = _dcf_input_from_payload(payload)
    out = run_dcf(inp)
    ctx = {"current_params": {**payload, "currency": "USD"}, "trading_comps": trading_comps}
    return load_workbook(BytesIO(generate_excel(inp, out, ctx).getvalue()), data_only=False)


def _sheet_text(ws):
    return "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)


def test_aapl_returns_supported_profile_and_curated_peers():
    profile = get_trading_comps_profile("AAPL")

    assert profile["supported"] is True
    assert profile["profile_id"] == "us_mega_cap_tech"
    assert profile["peer_set_source"] == "curated_aapl_mega_cap_tech"
    assert {"MSFT", "GOOGL", "META", "AMZN", "ORCL", "ADBE"}.issubset(set(profile["peers"]))


def test_2359_hk_returns_unavailable_no_generic_tech_implied_iv():
    comps = build_trading_comps("2359.HK", use_cache=True)

    assert comps["status"] == "unavailable"
    assert comps["profile_id"] == "cro_cdmo_peer_mapping_required"
    assert comps["peer_rows"] == []
    assert comps["peer_set_requested"] == []
    assert comps["valuation_ranges"] == []
    assert comps["implied_valuation"] == {}
    assert not (TECH_FALLBACK & set(comps.get("peers") or []))


def test_600519_ss_returns_unavailable_no_generic_tech_peers():
    comps = build_trading_comps("600519.SS", use_cache=True)

    assert comps["status"] == "unavailable"
    assert comps["profile_id"] == "premium_consumer_peer_mapping_required"
    assert comps["peer_rows"] == []
    assert not (TECH_FALLBACK & set(comps.get("peer_set_requested") or []))


def test_unknown_ticker_returns_unavailable_not_generic_tech_fallback():
    comps = build_trading_comps("__UNKNOWN__", use_cache=True)

    assert comps["status"] == "unavailable"
    assert comps["profile_id"] == "no_curated_peer_set_available"
    assert comps["peer_source"] == "unavailable"
    assert comps["peer_rows"] == []
    assert comps["implied_valuation"] == {}


def test_0700_hk_does_not_clean_pass_as_aapl_generic_tech_fallback():
    comps = build_trading_comps("0700.HK", use_cache=True)

    assert comps["status"] == "unavailable"
    assert comps["profile_id"] == "internet_platform_peer_mapping_required"
    assert comps["peer_source"] == "suppressed_generic_tech_fallback"
    assert comps["peer_rows"] == []


def test_excel_export_2359_trading_comps_unavailable_without_tech_peer_rows():
    comps = build_trading_comps("2359.HK", use_cache=True)
    wb = _workbook_for("2359.HK", comps)
    text = _sheet_text(wb["Trading Comps"])

    assert "Trading Comps unavailable" in text
    assert "Industry peer set requires review" in text
    assert "Generic technology fallback has been suppressed" in text
    for ticker in TECH_FALLBACK:
        assert ticker not in text
    assert "Apple Inc." not in text
    assert "Target Implied Valuation" not in text


def test_excel_export_aapl_trading_comps_still_contains_valid_peer_rows():
    comps = {
        "version": "test",
        "status": "ok",
        "symbol": "AAPL",
        "peer_source": "curated_aapl_mega_cap_tech",
        "peer_set_source": "curated_aapl_mega_cap_tech",
        "profile_id": "us_mega_cap_tech",
        "fetched_at": "test",
        "target": {"ticker": "AAPL", "name": "Apple Inc.", "field_sources": {}},
        "peer_rows": [
            {
                "ticker": "MSFT",
                "name": "Microsoft Corporation",
                "category": "Core: Platform / Cloud",
                "included_in_stats": {"ev_revenue": True, "ev_ebitda": True, "pe": True},
                "included_in_stats_any": True,
                "included_in_stats_all": True,
                "data_quality_tier": "OK",
                "rationale": "Core AAPL platform peer.",
                "raw": {"price": 1, "revenue": 1_000_000, "ebitda": 200_000, "market_cap": 2_000_000, "enterprise_value": 2_100_000},
                "multiples": {"ev_revenue": 2.1, "ev_ebitda": 10.5, "pe": 25.0},
                "outliers": {},
                "exclusions": {},
            }
        ],
        "summary_stats": {},
        "implied_valuation": {},
        "forward_diagnostic": {"rows": [], "summary": {"core_forward_available_count": 0, "core_peer_count": 1}},
        "forward_consensus": {},
        "data_quality": {"overall_comps_usability_tier": "OK", "football_field_comps_status": "Trailing comps rows restored"},
    }
    wb = _workbook_for("AAPL", comps)
    text = _sheet_text(wb["Trading Comps"])

    assert "Trading Comps unavailable" not in text
    assert "MSFT" in text
    assert "Microsoft Corporation" in text
    assert "curated_aapl_mega_cap_tech" in text
