"""V4.1.2 Covenant / Maturity Wall / IC Defense formula & audit tests.

These assert the defensive / audit layer of the formula-native LBO workbook:

  * the multi-tranche Covenant Check leverage / coverage ratios and pass/fail
    flags are live Excel formulas linked to the formula-native Debt Schedule and
    Operating Forecast (not hard-coded Python values) and tie out to the engine
    covenant oracle for the exported case;
  * the multi-tranche Maturity Wall ending balances / share of total debt are
    live formulas linked to the Debt Schedule final column;
  * the Returns Attribution sheet carries formula tie-out checks (Sources =
    Uses, Sponsor Equity, Exit Equity bridge, MOIC, debt paydown, ending cash,
    attribution component sum) that evaluate to OK via the lazy evaluator;
  * single-tranche covenant / maturity stay "Not Applicable" (no covenants are
    modeled) and the legacy values-only path stays 0-formula.

``run_lbo()`` remains the oracle; the lazy ``_Eval`` resolver (shared with
test_lbo_formula_workbook) proves the formulas reproduce it without Excel.
"""

from io import BytesIO

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from modeling.lbo_calculator import run_lbo
from modeling.lbo_attribution import build_lbo_attribution
from modeling.lbo_formula_workbook import generate_lbo_formula_excel
from modeling.lbo_excel_exporter import generate_lbo_values_only_excel
from tests.test_lbo_calculator_core import clean_anchor
from tests.test_lbo_multitranche_schedule import (
    multi_tranche_anchor,
    revolver_draw_fixture,
    low_leverage_high_fcf_cash_build_fixture,
)
from tests.test_lbo_formula_workbook import (
    _Eval,
    _wb,
    _row_addr,
    _row_addr_contains,
    _set_named,
    _formula_count,
)


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _cov_first_data_row(ws, years):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, int) and v in years:
            return r
    raise AssertionError("covenant schedule data rows not found")


def _label_row(ws, label, max_col=1):
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == label:
            return r
    raise AssertionError(f"label {label!r} not found on {ws.title}")


# ── 1. Covenant Check formula-linked & tie-out ───────────────────────────────


def test_covenant_check_multi_is_formula_driven_not_hardcoded():
    wb, payload, _ = _wb(multi_tranche_anchor)
    ws = wb["Covenant Check"]
    years = payload["operating_forecast"]["years"]
    first = _cov_first_data_row(ws, years)
    # Per-year EBITDA / ending debt / cash interest / ratios / breach flags are
    # formulas, not Python literals.
    for j in range(len(years)):
        r = first + j
        for col in (2, 3, 4, 5, 6, 8, 9, 10, 12):  # incl. leverage/coverage breach
            assert _is_formula(ws.cell(r, col).value), f"covenant col {col} row {r} not a formula"
    # Pass/Fail flags use Excel IF().
    assert "IF(" in ws.cell(first, 8).value
    assert "IF(" in ws.cell(first, 12).value


def test_covenant_check_thresholds_are_named_ranges():
    wb, _, _ = _wb(multi_tranche_anchor)
    names = set(wb.defined_names)
    assert "LBO_Max_Leverage" in names
    assert "LBO_Min_Coverage" in names


def test_covenant_check_references_debt_schedule_and_operating_forecast():
    wb, payload, _ = _wb(multi_tranche_anchor)
    ws = wb["Covenant Check"]
    years = payload["operating_forecast"]["years"]
    first = _cov_first_data_row(ws, years)
    # EBITDA links to Operating Forecast; ending debt / cash interest link to Debt
    # Schedule — proving the covenant is wired to the formula chain, not literals.
    assert "Operating Forecast" in ws.cell(first, 2).value
    assert "Debt Schedule" in ws.cell(first, 3).value
    assert "Debt Schedule" in ws.cell(first, 9).value


def test_covenant_check_ties_out_to_run_lbo_oracle():
    wb, payload, result = _wb(multi_tranche_anchor)
    ev = _Eval(wb)
    ws = wb["Covenant Check"]
    years = payload["operating_forecast"]["years"]
    first = _cov_first_data_row(ws, years)
    checks = result["covenant_summary"]["checks"]
    for j, chk in enumerate(checks):
        r = first + j
        # Ratio (net debt / EBITDA): only meaningful when not net cash.
        if chk["net_debt_ebitda"] is not None:
            assert ev.cell("Covenant Check", f"F{r}") == pytest.approx(
                chk["net_debt_ebitda"], rel=1e-9
            )
        if chk["interest_coverage"] is not None:
            assert ev.cell("Covenant Check", f"J{r}") == pytest.approx(
                chk["interest_coverage"], rel=1e-9
            )
        # Pass/fail flags match the engine covenant engine.
        assert ev.cell("Covenant Check", f"H{r}") == ("Yes" if chk["leverage_breach"] else "No")
        assert ev.cell("Covenant Check", f"L{r}") == (
            "Yes" if chk["interest_coverage_breach"] else "No"
        )


def test_covenant_check_responds_to_debt_assumption_mutation():
    """Raising a term loan's mandatory amortization repays debt faster, so the
    later-year leverage ratio must fall — proving the covenant is live."""
    wb, payload, result = _wb(multi_tranche_anchor)
    base = _Eval(wb)
    ws = wb["Covenant Check"]
    years = payload["operating_forecast"]["years"]
    last_r = _cov_first_data_row(ws, years) + len(years) - 1
    base_ratio = base.cell("Covenant Check", f"F{last_r}")
    _set_named(wb, base, "LBO_Tr2_Amort", 0.20)  # Term Loan B amortization up
    bumped = _Eval(wb)
    assert bumped.cell("Covenant Check", f"F{last_r}") < base_ratio - 1e-6


def test_covenant_check_responds_to_threshold_mutation():
    """Tightening the max-leverage threshold below the modeled ratio flips the
    leverage breach flag — proving pass/fail is a live IF(), not a literal."""
    wb, payload, _ = _wb(multi_tranche_anchor)
    base = _Eval(wb)
    ws = wb["Covenant Check"]
    years = payload["operating_forecast"]["years"]
    first = _cov_first_data_row(ws, years)
    assert base.cell("Covenant Check", f"H{first}") == "No"
    _set_named(wb, base, "LBO_Max_Leverage", 1.0)  # far below the modeled ~4.7x
    bumped = _Eval(wb)
    assert bumped.cell("Covenant Check", f"H{first}") == "Yes"


# ── 2. Maturity Wall formula-linked & tie-out ────────────────────────────────


def test_maturity_wall_multi_is_formula_linked():
    wb, _, _ = _wb(multi_tranche_anchor)
    ws = wb["Maturity Wall"]
    # Each tranche's ending balance at exit and share of total are formulas.
    saw_tranche = False
    for r in range(4, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not isinstance(name, str) or name == "Total":
            continue
        if not _is_formula(ws.cell(r, 3).value):
            continue
        saw_tranche = True
        assert "Debt Schedule" in ws.cell(r, 3).value  # ending balance link
        assert _is_formula(ws.cell(r, 4).value)         # share of total
        assert _is_formula(ws.cell(r, 5).value)         # within hold flag
    assert saw_tranche, "no formula-linked tranche rows on the Maturity Wall"


def test_maturity_wall_ties_out_to_oracle():
    wb, payload, result = _wb(multi_tranche_anchor)
    ev = _Eval(wb)
    ws = wb["Maturity Wall"]
    final = {t["id"]: t for t in result["debt_schedule"][-1]["tranches"]}
    cs = result["capital_structure_summary"]
    settings_tranches = {t["name"]: t["id"] for t in payload["capital_structure"]["tranches"]}
    total_r = _label_row(ws, "Total")
    assert ev.cell("Maturity Wall", f"C{total_r}") == pytest.approx(
        cs["total_ending_debt"], rel=1e-9
    )
    for r in range(4, total_r):
        name = ws.cell(r, 1).value
        if name not in settings_tranches:
            continue
        tid = settings_tranches[name]
        assert ev.cell("Maturity Wall", f"C{r}") == pytest.approx(
            final[tid]["ending_balance"], rel=1e-9, abs=1e-6
        )


def test_maturity_wall_responds_to_opening_balance_mutation():
    wb, payload, _ = _wb(multi_tranche_anchor)
    base = _Eval(wb)
    ws = wb["Maturity Wall"]
    tlb_r = _label_row(ws, "Term Loan B")
    base_end = base.cell("Maturity Wall", f"C{tlb_r}")
    # Term Loan B is tranche index 2; raise its amortization to pay it down.
    _set_named(wb, base, "LBO_Tr2_Amort", 0.20)
    bumped = _Eval(wb)
    assert bumped.cell("Maturity Wall", f"C{tlb_r}") < base_end - 1e-6


# ── 3. Returns Attribution formula tie-out checks ────────────────────────────


def _check_result_rows(ws):
    """Yield (label, result_cell_coord) for every formula 'Result' check row."""
    for r in range(1, ws.max_row + 1):
        res = ws.cell(r, 5).value
        if _is_formula(res):
            yield ws.cell(r, 1).value, f"E{r}"


@pytest.mark.parametrize("anchor", [clean_anchor, multi_tranche_anchor])
def test_attribution_tie_out_checks_present_and_pass(anchor):
    wb, _, _ = _wb(anchor)
    ev = _Eval(wb)
    ws = wb["Returns Attribution"]
    rows = list(_check_result_rows(ws))
    labels = {lbl for lbl, _ in rows}
    assert {"Sources = Uses", "Sponsor Equity = Uses - Debt",
            "Exit Equity = EV - Debt + Cash", "MOIC x Sponsor = Exit Equity",
            "Debt Paydown >= 0"} <= labels
    for _, coord in rows:
        assert ev.cell("Returns Attribution", coord) == "OK"


def test_attribution_checks_include_multi_only_rows():
    wb, _, _ = _wb(multi_tranche_anchor)
    ws = wb["Returns Attribution"]
    labels = {lbl for lbl, _ in _check_result_rows(ws)}
    assert "Ending Debt: Returns = Schedule" in labels
    assert "Ending Cash >= 0" in labels


def test_attribution_component_sum_check_flags_mismatch_when_present():
    """With precomputed attribution, the live component sum must tie to the live
    equity-value-creation bridge (no silent residual mismatch)."""
    payload = multi_tranche_anchor()
    payload["debt"] = {"cash_to_balance_sheet": 60.0}
    payload["capital_structure"]["cash_balance_beginning"] = 60.0
    payload["capital_structure"]["minimum_cash_balance"] = 80.0
    result = run_lbo(payload)
    payload["attribution"] = build_lbo_attribution(payload, result)
    assert payload["attribution"]["status"] == "ok"
    wb = load_workbook(
        BytesIO(generate_lbo_formula_excel(payload, result).getvalue()), data_only=False
    )
    ev = _Eval(wb)
    ws = wb["Returns Attribution"]
    r = _label_row(ws, "Component Sum vs Live Bridge")
    # Diff (B) is ~0 and the result (C) is OK.
    assert ev.cell("Returns Attribution", f"B{r}") == pytest.approx(0.0, abs=1e-6)
    assert ev.cell("Returns Attribution", f"C{r}") == "OK"


def test_attribution_checks_live_even_without_attribution_detail():
    """The named-range tie-out checks render and pass even when component-level
    attribution data is not attached to the payload."""
    wb, _, _ = _wb(clean_anchor)
    ws = wb["Returns Attribution"]
    # No attribution payload -> unavailable note, but checks still present.
    assert any("unavailable" in (c.value or "").lower()
               for row in ws.iter_rows() for c in row if isinstance(c.value, str))
    assert list(_check_result_rows(ws))


def test_attribution_checks_pass_on_revolver_and_cash_build_fixtures():
    for fixture in (revolver_draw_fixture, low_leverage_high_fcf_cash_build_fixture):
        payload = fixture()
        result = run_lbo(payload)
        wb = load_workbook(
            BytesIO(generate_lbo_formula_excel(payload, result).getvalue()), data_only=False
        )
        ev = _Eval(wb)
        ws = wb["Returns Attribution"]
        for _, coord in _check_result_rows(ws):
            assert ev.cell("Returns Attribution", coord) == "OK", fixture.__name__


# ── 4. Single-tranche stays N/A; legacy path stays values-only ───────────────


def test_single_tranche_covenant_and_maturity_remain_not_applicable():
    wb, _, _ = _wb(clean_anchor)
    for sheet in ("Covenant Check", "Maturity Wall"):
        ws = wb[sheet]
        has_formula = any(
            _is_formula(c.value) for row in ws.iter_rows() for c in row
        )
        assert not has_formula, f"{sheet} must stay values-only in single-tranche mode"
        blob = " ".join(c.value for row in ws.iter_rows() for c in row
                        if isinstance(c.value, str))
        assert "single-tranche" in blob.lower()


def test_legacy_values_only_workbook_has_no_covenant_or_maturity_formulas():
    payload = multi_tranche_anchor()
    wb = load_workbook(
        BytesIO(generate_lbo_values_only_excel(payload, run_lbo(payload)).getvalue()),
        data_only=False,
    )
    assert _formula_count(wb) == 0


# ── 5. Audit discloses the V4.1.2 defense-layer formula state ────────────────


def test_audit_discloses_v412_formula_linked_defense_layer():
    wb, _, _ = _wb(multi_tranche_anchor)
    blob = " ".join(
        c.value for row in wb["Audit & Disclosures"].iter_rows()
        for c in row if isinstance(c.value, str)
    )
    assert "Covenant Check and Maturity Wall are now formula-linked" in blob
    assert "formula tie-out checks" in blob
    # Scenario / Sensitivity still disclosed as precomputed.
    assert "Scenario Summary and Sensitivity Analysis remain Python-precomputed" in blob


# ── 6. Covenant formula-string contract: IFERROR + AND zero-div protection ────


def test_covenant_ratio_formulas_guard_with_iferror():
    """The net-debt/EBITDA (F) and interest-coverage (J) ratio cells must wrap
    their division in IFERROR so a zero-EBITDA / zero-interest year degrades to
    a label instead of leaking #DIV/0! into the displayed ratio."""
    wb, payload, _ = _wb(multi_tranche_anchor)
    ws = wb["Covenant Check"]
    years = payload["operating_forecast"]["years"]
    first = _cov_first_data_row(ws, years)
    for j in range(len(years)):
        r = first + j
        net_debt_ebitda = ws.cell(r, 6).value
        interest_coverage = ws.cell(r, 10).value
        assert "IFERROR(" in net_debt_ebitda, f"F{r} missing IFERROR guard"
        assert "IFERROR(" in interest_coverage, f"J{r} missing IFERROR guard"
        # Interest coverage also gates on positive interest before dividing.
        assert "I" in interest_coverage and ">" in interest_coverage


def test_covenant_breach_flags_use_and_gate_avoiding_division():
    """Leverage-breach (H) and coverage-breach (L) flags must use an AND() gate
    whose first term guards the denominator (EBITDA for leverage, cash interest
    for coverage) so a zero-EBITDA / zero-interest year cannot raise a spurious
    breach. The gate uses a multiplication form, never a bare division."""
    wb, payload, _ = _wb(multi_tranche_anchor)
    ws = wb["Covenant Check"]
    years = payload["operating_forecast"]["years"]
    first = _cov_first_data_row(ws, years)
    for j in range(len(years)):
        r = first + j
        leverage_breach = ws.cell(r, 8).value
        coverage_breach = ws.cell(r, 12).value
        # AND gate present on both breach flags.
        assert "AND(" in leverage_breach, f"H{r} missing AND gate"
        assert "AND(" in coverage_breach, f"L{r} missing AND gate"
        # Leverage breach gates on EBITDA (B) and uses the multiplication form
        # (no `E{r}/B{r}` division that would blow up on zero EBITDA).
        assert f"B{r}>" in leverage_breach
        assert f"E{r}/B{r}" not in leverage_breach
        # Coverage breach gates on cash interest (I) and avoids `B{r}/I{r}`.
        assert f"I{r}>" in coverage_breach
        assert f"B{r}/I{r}" not in coverage_breach
