from io import BytesIO

from openpyxl import load_workbook

from app import app
from tests.test_lbo_calculator_core import clean_anchor


def test_lbo_defaults_contract():
    client = app.test_client()
    resp = client.get("/api/modeling/lbo/defaults?symbol=SYNTH")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["status"] == "ok"
    assert data["defaults"]["symbol"] == "SYNTH"
    assert data["defaults"]["debt"]["cash_sweep_pct"] == 1.0
    assert data["defaults"]["debt"]["cash_to_balance_sheet"] == 0.0
    assert data["provenance"]["leverage_multiple"]["rationale_cn"]
    assert data["serviceability"]["debt_service_pass"] is True
    assert data["defaults"]["default_builder"]["serviceability"]["final_leverage"] == data["serviceability"]["final_leverage"]


def test_aapl_lbo_defaults_keep_early_review_warning():
    client = app.test_client()
    resp = client.get("/api/modeling/lbo/defaults?symbol=AAPL")
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["status"] == "ok"
    assert any(flag["code"] == "EARLY_LBO_SUITABILITY_REVIEW" for flag in data["flags"])
    message = " ".join(flag["message"] for flag in data["flags"])
    assert "modeling starter" in message
    assert "not a suitability conclusion" in message
    suitability = data["suitability"]
    assert suitability["suitability"] == "unsuitable"
    assert suitability["veto_triggered"] is True
    veto_codes = {reason["code"] for reason in suitability["veto_reasons"]}
    assert (
        "MEGA_CAP_FINANCING_FEASIBILITY" in veto_codes
        or "SPONSOR_EQUITY_CHECK_TOO_LARGE" in veto_codes
    )
    assert suitability["recommended_next_view"] == "DCF"


def test_lbo_calculate_contract():
    client = app.test_client()
    resp = client.post("/api/modeling/lbo", json=clean_anchor())
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["status"] == "ok"
    assert data["returns"]["irr"] is not None
    assert data["debt_schedule"][0]["cash_available_for_debt"] == 280


def test_lbo_hard_error_contract():
    client = app.test_client()
    payload = clean_anchor()
    payload["debt"]["cash_sweep_pct"] = 0.5
    resp = client.post("/api/modeling/lbo", json=payload)
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["status"] == "error"
    assert data["returns"] is None
    assert data["flags"][0]["severity"] == "error"


def test_lbo_excel_export_contract():
    client = app.test_client()
    resp = client.post("/api/modeling/lbo/excel", json=clean_anchor())
    assert resp.status_code == 200
    assert resp.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    wb = load_workbook(BytesIO(resp.data), data_only=False)
    assert {
        "Cover", "Model Map", "Assumptions", "Sources & Uses",
        "Operating Forecast", "Debt Schedule", "Covenant Check",
        "Returns Summary", "Returns Attribution", "Maturity Wall",
        "Audit & Disclosures",
    } <= set(wb.sheetnames)


def test_lbo_scenarios_contract():
    client = app.test_client()
    resp = client.post("/api/modeling/lbo/scenarios", json={"inputs": clean_anchor()})
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["status"] == "ok"
    assert data["method"] == "lbo_case_layer_v46"
    assert set(data["scenarios"]) == {"base", "upside", "downside"}
    assert data["comparison"]["rows"]


def test_lbo_scenarios_does_not_change_calc_contract():
    client = app.test_client()
    resp = client.post("/api/modeling/lbo", json=clean_anchor())
    data = resp.get_json()
    assert data["status"] == "ok"
    assert "scenarios" not in data
    assert "comparison" not in data


def test_lbo_excel_audit_can_show_v41_defaults_metadata():
    client = app.test_client()
    defaults = client.get("/api/modeling/lbo/defaults?symbol=SYNTH").get_json()["defaults"]
    payload = clean_anchor()
    payload["default_builder"] = defaults["default_builder"]
    resp = client.post("/api/modeling/lbo/excel", json=payload)
    wb = load_workbook(BytesIO(resp.data), data_only=False)
    # V4.5: default-builder provenance lives on the Assumptions sheet.
    text = [cell.value for row in wb["Assumptions"].iter_rows() for cell in row if cell.value]
    assert "Default Builder Provenance" in text
    assert {"Assumption", "Value", "Source", "Confidence", "Rationale"} <= set(text)
    assert any("system default structure only" in str(value) for value in text)
