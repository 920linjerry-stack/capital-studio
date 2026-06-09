"""V4.6 LBO Scenario / Case Layer v1 tests.

These verify the scenario-organization layer that sits on top of the frozen
V4.0-V4.5 LBO engine. The layer must:
  * generate deterministic Base / Upside / Downside input payloads,
  * run each payload through the existing run_lbo() (never re-implement math),
  * isolate a failed scenario (show unavailable, not break the others),
  * expose a single canonical metric schema shared by UI / Excel / tests,
  * label cases Base / Upside / Downside (never Bull / Bear) and never emit
    investment / acquisition recommendation language.
"""

import copy
from io import BytesIO

from openpyxl import load_workbook

from app import app
from modeling.lbo_calculator import run_lbo
from modeling.lbo_scenarios import (
    DEFAULT_SCENARIO_CONFIG,
    DISCLOSURE_EXCEL_DETERMINISTIC,
    DISCLOSURE_EXCEL_MULTI_FACTOR,
    METHOD,
    NOT_APPLICABLE_SINGLE,
    SCENARIO_LABELS,
    SCENARIO_METRIC_SCHEMA,
    UNAVAILABLE,
    build_lbo_scenarios,
    generate_scenario_inputs,
    summarize_scenario_results,
)
from modeling.lbo_excel_exporter import (
    SCENARIO_SUMMARY_SHEET,
    SHEET_NAMES,
    generate_lbo_excel,
)
from tests.test_lbo_calculator_core import clean_anchor
from tests.test_lbo_multitranche_schedule import multi_tranche_anchor


# ── fixtures ─────────────────────────────────────────────────────────────────


def thin_cushion_breach_fixture():
    """Multi-tranche deal whose Base passes covenants but whose Downside (lower
    EBITDA + higher rates) breaches minimum interest coverage. Both still RUN."""
    a = multi_tranche_anchor()
    a["capital_structure"]["covenants"] = {
        "max_net_debt_ebitda": 6.0,
        "min_interest_coverage": 2.1,
    }
    return a


def negative_ebitda_config():
    """Scenario config that drives Downside EBITDA non-positive so the Downside
    case must surface as unavailable (engine validation), while Base / Upside run."""
    return {
        "upside": dict(DEFAULT_SCENARIO_CONFIG["upside"]),
        "downside": {"ebitda_pct_delta": -1.2, "exit_multiple_delta": -0.5, "interest_rate_delta": 0.005},
    }


def _scenario_inputs_unchanged(base):
    """Assert build/generate did not mutate the caller's base inputs."""
    snapshot = copy.deepcopy(base)
    generate_scenario_inputs(base)
    build_lbo_scenarios(base)
    assert base == snapshot


# ══ 1. Input generation ══════════════════════════════════════════════════════


def test_generate_returns_three_scenarios():
    g = generate_scenario_inputs(clean_anchor())
    assert set(g) == {"base", "upside", "downside"}


def test_base_input_is_unchanged_copy():
    base = clean_anchor()
    g = generate_scenario_inputs(base)
    assert g["base"] == base
    assert g["base"] is not base


def test_generate_does_not_mutate_base():
    base = clean_anchor()
    snapshot = copy.deepcopy(base)
    generate_scenario_inputs(base)
    assert base == snapshot


def test_upside_scales_operating_ebitda_up_entry_fixed():
    # V4.7.1 same-transaction sensitivity: forecast EBITDA scales up, but the
    # transaction Entry EBITDA is held constant.
    base = clean_anchor()
    g = generate_scenario_inputs(base)
    assert g["upside"]["transaction"]["entry_ebitda"] == base["transaction"]["entry_ebitda"]
    assert g["upside"]["operating_forecast"]["ebitda"][0] > base["operating_forecast"]["ebitda"][0]


def test_downside_scales_operating_ebitda_down_entry_fixed():
    base = clean_anchor()
    g = generate_scenario_inputs(base)
    assert g["downside"]["transaction"]["entry_ebitda"] == base["transaction"]["entry_ebitda"]
    assert g["downside"]["operating_forecast"]["ebitda"][0] < base["operating_forecast"]["ebitda"][0]


def test_exit_multiple_moves_with_deltas():
    base = clean_anchor()
    g = generate_scenario_inputs(base)
    base_exit = base["transaction"]["exit_multiple"]
    assert g["upside"]["transaction"]["exit_multiple"] == base_exit + 0.5
    # V4.7.1 stronger downside: exit multiple -1.0x.
    assert g["downside"]["transaction"]["exit_multiple"] == base_exit - 1.0


def test_interest_rate_moves_with_deltas_single():
    base = clean_anchor()
    g = generate_scenario_inputs(base)
    base_rate = base["debt"]["interest_rate"]
    assert abs(g["upside"]["debt"]["interest_rate"] - (base_rate - 0.005)) < 1e-12
    # V4.7.1 stronger downside: +1.0pp interest rate.
    assert abs(g["downside"]["debt"]["interest_rate"] - (base_rate + 0.01)) < 1e-12


def test_downside_stresses_capex_and_nwc():
    base = clean_anchor()
    g = generate_scenario_inputs(base)
    assert g["downside"]["operating_forecast"]["capex"][0] > base["operating_forecast"]["capex"][0]
    assert g["downside"]["operating_forecast"]["change_in_nwc"][0] > base["operating_forecast"]["change_in_nwc"][0]


def test_interest_rate_moves_uniformly_multi():
    base = multi_tranche_anchor()
    g = generate_scenario_inputs(base)
    for i, tr in enumerate(base["capital_structure"]["tranches"]):
        up = g["upside"]["capital_structure"]["tranches"][i]["interest_rate"]
        assert abs(up - max(0.0, tr["interest_rate"] - 0.005)) < 1e-12


def test_custom_scenario_config_overrides_defaults():
    base = clean_anchor()
    cfg = {"upside": {"ebitda_pct_delta": 0.25}}
    g = generate_scenario_inputs(base, cfg)
    # Entry EBITDA stays fixed; the operating forecast scales by the override.
    assert g["upside"]["transaction"]["entry_ebitda"] == base["transaction"]["entry_ebitda"]
    assert abs(g["upside"]["operating_forecast"]["ebitda"][0] - 1000 * 1.25) < 1e-9
    # Unset keys fall back to defaults.
    assert g["upside"]["transaction"]["exit_multiple"] == base["transaction"]["exit_multiple"] + 0.5


# ══ 2. Scenario execution ════════════════════════════════════════════════════


def test_build_status_ok_all_run():
    res = build_lbo_scenarios(clean_anchor())
    assert res["status"] == "ok"
    assert res["method"] == METHOD
    for key in ("base", "upside", "downside"):
        assert res["scenarios"][key]["status"] == "ok"


def test_each_scenario_matches_direct_run_lbo():
    base = clean_anchor()
    g = generate_scenario_inputs(base)
    res = build_lbo_scenarios(base)
    for key in ("base", "upside", "downside"):
        direct = run_lbo(g[key])
        assert res["scenarios"][key]["result"]["returns"]["irr"] == direct["returns"]["irr"]
        assert res["scenarios"][key]["result"]["returns"]["moic"] == direct["returns"]["moic"]


def test_build_does_not_mutate_base():
    _scenario_inputs_unchanged(clean_anchor())


def test_failed_scenario_isolated():
    res = build_lbo_scenarios(clean_anchor(), negative_ebitda_config())
    assert res["status"] == "warning"
    assert res["scenarios"]["base"]["status"] == "ok"
    assert res["scenarios"]["upside"]["status"] == "ok"
    assert res["scenarios"]["downside"]["status"] == "unavailable"
    assert res["scenarios"]["downside"].get("unavailable_reason")
    codes = {f["code"] for f in res["flags"]}
    assert "SCENARIO_UNAVAILABLE" in codes


def test_all_scenarios_unavailable_is_error():
    base = clean_anchor()
    base["transaction"]["entry_ebitda"] = -5  # base itself invalid
    res = build_lbo_scenarios(base)
    assert res["status"] == "error"
    assert res["comparison"] is None
    codes = {f["code"] for f in res["flags"]}
    assert "ALL_SCENARIOS_UNAVAILABLE" in codes


# ══ 3. Comparison summary ════════════════════════════════════════════════════


def test_comparison_rows_align_to_schema():
    res = build_lbo_scenarios(clean_anchor())
    rows = res["comparison"]["rows"]
    assert [r["key"] for r in rows] == [s["key"] for s in SCENARIO_METRIC_SCHEMA]


def test_canonical_schema_exposed():
    res = build_lbo_scenarios(clean_anchor())
    assert res["canonical_metric_schema"] == SCENARIO_METRIC_SCHEMA


def test_comparison_has_three_columns():
    res = build_lbo_scenarios(clean_anchor())
    for row in res["comparison"]["rows"]:
        assert "base" in row and "upside" in row and "downside" in row


def test_single_tranche_covenant_cells_not_applicable():
    res = build_lbo_scenarios(clean_anchor())
    rows = {r["key"]: r for r in res["comparison"]["rows"]}
    assert rows["covenant_status"]["base"] == NOT_APPLICABLE_SINGLE
    assert rows["revolver_draw_occurred"]["base"] == NOT_APPLICABLE_SINGLE
    assert rows["minimum_interest_coverage"]["base"] == NOT_APPLICABLE_SINGLE


def test_multi_tranche_covenant_cells_populated():
    res = build_lbo_scenarios(thin_cushion_breach_fixture())
    rows = {r["key"]: r for r in res["comparison"]["rows"]}
    assert rows["covenant_status"]["base"] == "pass"
    assert rows["covenant_status"]["downside"] == "breach"


def test_unavailable_scenario_cells_marked():
    res = build_lbo_scenarios(clean_anchor(), negative_ebitda_config())
    rows = {r["key"]: r for r in res["comparison"]["rows"]}
    assert rows["irr"]["downside"] == UNAVAILABLE
    assert res["comparison"]["scenario_statuses"]["downside"]["status"] == "unavailable"
    assert res["comparison"]["scenario_statuses"]["downside"]["reason"]


def test_upside_irr_above_downside_irr():
    res = build_lbo_scenarios(clean_anchor())
    rows = {r["key"]: r for r in res["comparison"]["rows"]}
    assert rows["irr"]["upside"] > rows["irr"]["downside"]


def test_summarize_directly_from_outputs():
    res = build_lbo_scenarios(clean_anchor())
    summary = summarize_scenario_results(res["scenarios"])
    assert [r["key"] for r in summary["rows"]] == [s["key"] for s in SCENARIO_METRIC_SCHEMA]


# ══ 4. Attribution passthrough ═══════════════════════════════════════════════


def test_attribution_attached_per_scenario():
    res = build_lbo_scenarios(clean_anchor())
    for key in ("base", "upside", "downside"):
        assert res["scenarios"][key]["attribution"]["status"] == "ok"


def test_attribution_contributions_in_comparison():
    res = build_lbo_scenarios(clean_anchor())
    rows = {r["key"]: r for r in res["comparison"]["rows"]}
    assert isinstance(rows["ebitda_growth_contribution"]["base"], (int, float))
    assert isinstance(rows["multiple_movement_contribution"]["base"], (int, float))
    assert isinstance(rows["deleveraging_contribution"]["base"], (int, float))


def test_multiple_movement_zero_in_base():
    # Base exit == entry multiple → no multiple movement contribution.
    res = build_lbo_scenarios(clean_anchor())
    rows = {r["key"]: r for r in res["comparison"]["rows"]}
    assert abs(rows["multiple_movement_contribution"]["base"]) < 1e-9


# ══ 5. API ═══════════════════════════════════════════════════════════════════


def test_api_scenarios_ok():
    client = app.test_client()
    resp = client.post("/api/modeling/lbo/scenarios", json={"inputs": clean_anchor()})
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["status"] == "ok"
    assert data["method"] == METHOD
    assert data["comparison"]["rows"]


def test_api_scenarios_accepts_config_and_suitability():
    client = app.test_client()
    resp = client.post("/api/modeling/lbo/scenarios", json={
        "inputs": clean_anchor(),
        "scenario_config": negative_ebitda_config(),
        "base_suitability": {"suitability": "review"},
    })
    data = resp.get_json()
    assert data["status"] == "warning"
    rows = {r["key"]: r for r in data["comparison"]["rows"]}
    assert rows["suitability_status"]["base"] == "review"


def test_api_lbo_calc_contract_unchanged():
    client = app.test_client()
    resp = client.post("/api/modeling/lbo", json=clean_anchor())
    data = resp.get_json()
    assert data["status"] == "ok"
    assert "scenarios" not in data  # /lbo never grows a scenarios block


def test_api_defaults_contract_unchanged():
    client = app.test_client()
    resp = client.get("/api/modeling/lbo/defaults?symbol=SYNTH")
    assert resp.status_code == 200
    assert resp.get_json()["status"] == "ok"


# ══ 6. Excel ═════════════════════════════════════════════════════════════════


def _excel_with_scenarios(anchor):
    payload = anchor()
    result = run_lbo(payload)
    payload["scenarios"] = build_lbo_scenarios(anchor())
    return generate_lbo_excel(payload, result)


def test_excel_no_scenario_keeps_11_sheets():
    payload = clean_anchor()
    result = run_lbo(payload)
    wb = load_workbook(generate_lbo_excel(payload, result), data_only=False)
    assert wb.sheetnames == SHEET_NAMES


def test_excel_with_scenario_inserts_sheet_at_position_3():
    wb = load_workbook(_excel_with_scenarios(clean_anchor), data_only=False)
    assert len(wb.sheetnames) == 12
    assert wb.sheetnames[2] == SCENARIO_SUMMARY_SHEET
    expected = SHEET_NAMES[:2] + [SCENARIO_SUMMARY_SHEET] + SHEET_NAMES[2:]
    assert wb.sheetnames == expected


def test_excel_scenario_sheet_has_case_columns():
    wb = load_workbook(_excel_with_scenarios(clean_anchor), data_only=False)
    ws = wb[SCENARIO_SUMMARY_SHEET]
    header = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
    assert SCENARIO_LABELS["base"] in header
    assert SCENARIO_LABELS["upside"] in header
    assert SCENARIO_LABELS["downside"] in header


def test_excel_scenario_sheet_lists_metrics():
    wb = load_workbook(_excel_with_scenarios(clean_anchor), data_only=False)
    ws = wb[SCENARIO_SUMMARY_SHEET]
    strings = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    for label in ("IRR", "MOIC", "Multiple Movement Contribution",
                  "Revolver Draw Occurred", "Debt Service Failure", "Covenant Status"):
        assert label in strings, label


def test_excel_scenario_sheet_has_disclosures():
    wb = load_workbook(_excel_with_scenarios(clean_anchor), data_only=False)
    ws = wb[SCENARIO_SUMMARY_SHEET]
    strings = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert DISCLOSURE_EXCEL_DETERMINISTIC in strings
    assert DISCLOSURE_EXCEL_MULTI_FACTOR in strings


def test_excel_scenario_sheet_listed_in_model_map():
    wb = load_workbook(_excel_with_scenarios(clean_anchor), data_only=False)
    listed = [c.value for row in wb["Model Map"].iter_rows() for c in row if isinstance(c.value, str)]
    assert SCENARIO_SUMMARY_SHEET in listed


def test_excel_scenario_sheet_no_formulas():
    wb = load_workbook(_excel_with_scenarios(multi_tranche_anchor), data_only=False)
    ws = wb[SCENARIO_SUMMARY_SHEET]
    for row in ws.iter_rows():
        for cell in row:
            assert not (isinstance(cell.value, str) and cell.value.startswith("="))


# ══ 7. Wording discipline ════════════════════════════════════════════════════

# Canonical forbidden-wording list from the V4.6 spec. Bull / Bear are enforced
# separately via the label assertion (they are not in the spec wording list, and
# a disclaimer may legitimately reference "long / short" market views).
FORBIDDEN = [
    "expected case", "recommended case", "probability-weighted", "best case to buy",
    "attractive", "should acquire", "推荐", "买入", "建议收购",
]


def _all_text(obj):
    out = []
    if isinstance(obj, dict):
        for v in obj.values():
            out.extend(_all_text(v))
    elif isinstance(obj, list):
        for v in obj:
            out.extend(_all_text(v))
    elif isinstance(obj, str):
        out.append(obj)
    return out


def test_labels_are_base_upside_downside():
    assert SCENARIO_LABELS == {
        "base": "Base Case", "upside": "Upside Case", "downside": "Downside Case",
    }


def _v46_authored_text(res):
    """Only the text V4.6 itself authors. The embedded run_lbo / attribution
    output is frozen-engine content (e.g. the CN term 买入倍数 = 'entry multiple')
    and is out of scope for the scenario-layer wording check."""
    corpus = list(res.get("disclosures") or [])
    corpus += list(res.get("disclosures_cn") or [])
    corpus += [str(v) for v in (res.get("notes") or {}).values()]
    corpus += [f.get("message", "") for f in (res.get("flags") or [])]
    corpus += [s["label"] for s in res.get("canonical_metric_schema") or []]
    for scn in (res.get("scenarios") or {}).values():
        corpus.append(scn.get("label", ""))
        corpus.append(scn.get("unavailable_reason", "") or "")
        corpus.append(scn.get("comparison_reason", "") or "")
    for row in ((res.get("comparison") or {}).get("rows") or []):
        corpus.append(str(row.get("metric")))
        for key in ("base", "upside", "downside"):
            corpus.append(str(row.get(key)))
    return corpus


def test_no_forbidden_wording_in_payload():
    res = build_lbo_scenarios(clean_anchor())
    blob = " ".join(_v46_authored_text(res)).lower()
    for term in FORBIDDEN:
        assert term.lower() not in blob, term


def test_no_forbidden_wording_in_excel():
    wb = load_workbook(_excel_with_scenarios(clean_anchor), data_only=False)
    strings = [c.value for ws in wb.worksheets for row in ws.iter_rows()
               for c in row if isinstance(c.value, str)]
    blob = " ".join(strings).lower()
    for term in FORBIDDEN:
        assert term.lower() not in blob, term


def test_disclosures_present_and_non_recommendation():
    res = build_lbo_scenarios(clean_anchor())
    blob = " ".join(res["disclosures"]).lower()
    assert "not forecasts" in blob or "not a forecast" in blob
    assert "recommendation" in blob


# ══ 8. Regression: engine output identical to standalone run ═════════════════


def test_base_scenario_irr_matches_plain_run():
    base = clean_anchor()
    res = build_lbo_scenarios(base)
    plain = run_lbo(clean_anchor())
    assert res["scenarios"]["base"]["result"]["returns"]["irr"] == plain["returns"]["irr"]
