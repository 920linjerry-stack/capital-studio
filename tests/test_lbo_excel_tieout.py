from io import BytesIO

from openpyxl import load_workbook
import math

from modeling.lbo_calculator import run_lbo
from modeling.lbo_attribution import build_lbo_attribution
from modeling.lbo_excel_exporter import (
    generate_lbo_excel,
    generate_lbo_values_only_excel,
    SHEET_NAMES,
)
from tests.test_lbo_calculator_core import clean_anchor
from tests.test_lbo_multitranche_schedule import multi_tranche_anchor

EXPECTED_SHEETS = SHEET_NAMES

# This file pins the LEGACY values-only / Python-oracle workbook
# (generate_lbo_excel == generate_lbo_values_only_excel). The user-facing export
# is now the formula-native workbook; its contract lives in
# test_lbo_formula_workbook.py. The values-only path stays 0-formula on purpose
# so it remains a stable gold-master reference.


def _pairs(ws):
    return {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, ws.max_row + 1)}


def _all_cell_values(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                yield cell.value


def test_excel_headline_values_tie_to_python_output():
    payload = clean_anchor()
    out = run_lbo(payload)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, out).getvalue()), data_only=False)
    returns = _pairs(wb["Returns Summary"])
    assert math.isclose(returns["IRR"], out["returns"]["irr"], rel_tol=1e-12)
    assert math.isclose(returns["MOIC"], out["returns"]["moic"], rel_tol=1e-12)
    assert math.isclose(returns["Sponsor Equity"], out["returns"]["sponsor_equity"], rel_tol=1e-12)
    assert math.isclose(returns["Exit Equity Value"], out["returns"]["exit_equity_value"], rel_tol=1e-12)


def test_legacy_values_only_excel_has_no_formulas_or_external_links():
    payload = clean_anchor()
    out = run_lbo(payload)
    data = generate_lbo_values_only_excel(payload, out).getvalue()
    wb = load_workbook(BytesIO(data), data_only=False)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                assert not (isinstance(cell.value, str) and cell.value.startswith("="))
    assert not getattr(wb, "_external_links", [])


def test_single_tranche_sheet_names_and_sections():
    payload = clean_anchor()
    out = run_lbo(payload)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, out).getvalue()))
    assert wb.sheetnames == EXPECTED_SHEETS
    assert wb["Cover"].sheet_view.showGridLines is False


def test_multi_tranche_workbook_exports_with_required_sections():
    payload = multi_tranche_anchor()
    out = run_lbo(payload)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, out).getvalue()))
    assert wb.sheetnames == EXPECTED_SHEETS
    values = [v for v in _all_cell_values(wb) if isinstance(v, str)]
    # Maturity Wall column label must remain exactly this string.
    assert "Outstanding at Exit" in values
    assert "Maturity Wall" in values
    assert "Covenant Check" in values


def test_returns_attribution_and_audit_disclosure_present():
    payload = multi_tranche_anchor()
    out = run_lbo(payload)
    payload["attribution"] = build_lbo_attribution(payload, out)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, out).getvalue()))
    attribution_values = [
        wb["Returns Attribution"].cell(r, 1).value
        for r in range(1, wb["Returns Attribution"].max_row + 1)
    ]
    assert "Equity Value Bridge" in attribution_values
    audit_values = [v for v in _all_cell_values(wb) if isinstance(v, str)]
    assert any(v == "Source of Truth" for v in audit_values)


def _merged_starts_at(ws, row, col_letter):
    prefix = f"{col_letter}{row}:"
    return any(str(rng).startswith(prefix) for rng in ws.merged_cells.ranges)


def test_returns_note_is_wrapped_merged_and_tall_enough():
    payload = multi_tranche_anchor()
    out = run_lbo(payload)
    payload["attribution"] = build_lbo_attribution(payload, out)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, out).getvalue()))
    ws = wb["Returns Attribution"]
    note_rows = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Note"]
    assert note_rows, "expected a MOIC contribution Note row"
    r = note_rows[0]
    assert ws.cell(r, 2).alignment.wrap_text is True
    assert ws.cell(r, 2).alignment.vertical == "top"
    h = ws.row_dimensions[r].height
    assert h is not None and 24 <= h <= 90
    assert _merged_starts_at(ws, r, "B")


def test_audit_deferred_scope_note_is_readable():
    payload = multi_tranche_anchor()
    out = run_lbo(payload)
    wb = load_workbook(BytesIO(generate_lbo_excel(payload, out).getvalue()))
    ws = wb["Audit & Disclosures"]
    rows = list(range(1, ws.max_row + 1))
    label_row = next(r for r in rows if ws.cell(r, 1).value == "Deferred Scope")
    note_row = label_row + 1
    cell = ws.cell(note_row, 1)
    assert isinstance(cell.value, str) and "deferred future scope" in cell.value
    assert cell.alignment.wrap_text is True
    h = ws.row_dimensions[note_row].height
    assert h is not None and 24 <= h <= 90
    assert _merged_starts_at(ws, note_row, "A")


def test_legacy_multi_tranche_workbook_has_no_formulas_or_links():
    payload = multi_tranche_anchor()
    out = run_lbo(payload)
    wb = load_workbook(BytesIO(generate_lbo_values_only_excel(payload, out).getvalue()), data_only=False)
    for v in _all_cell_values(wb):
        assert not (isinstance(v, str) and v.startswith("="))
    assert not getattr(wb, "_external_links", [])
