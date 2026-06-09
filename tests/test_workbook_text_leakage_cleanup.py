from io import BytesIO
import zipfile

from openpyxl import load_workbook

from app import _dcf_input_from_payload
from modeling.dcf_calculator import run_dcf
from modeling.excel_exporter import generate_excel


def _payload(symbol="2359.HK", company="WuXi AppTec Co., Ltd."):
    return {
        "symbol": symbol,
        "company": company,
        "price": 100,
        "revenue": 1000,
        "ebit": 250,
        "da": 40,
        "capex": 50,
        "wc_change": 0,
        "tax_rate": 0.21,
        "net_debt": 0,
        "shares": 100,
        "revenue_growth": 0.03,
        "ebit_margin": 0.25,
        "da_pct_revenue": 0.04,
        "capex_pct_revenue": 0.05,
        "wc_change_pct_revenue": 0.0,
        "wacc": 0.10,
        "beta": 1.0,
        "terminal_g": 0.02,
        "exit_multiple": 12,
        "forecast_years": 5,
        "tv_method": "average",
        "currency": "USD",
    }


def _workbook_bytes(symbol="2359.HK", company="WuXi AppTec Co., Ltd."):
    payload = _payload(symbol, company)
    inp = _dcf_input_from_payload(payload)
    out = run_dcf(inp)
    ctx = {
        "current_params": {**payload, "currency": "USD"},
        "trading_comps": {
            "status": "unavailable",
            "reason": "Industry peer set requires review. Generic technology fallback has been suppressed for this ticker.",
            "profile_id": "test",
            "peer_set_source": "suppressed_generic_tech_fallback",
            "peer_rows": [],
            "valuation_ranges": [],
        },
    }
    return generate_excel(inp, out, ctx).getvalue()


def _all_cell_text(wb):
    return "\n".join(
        str(cell.value)
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    )


def test_non_aapl_workbook_has_no_aapl_template_text_or_formulas():
    data = _workbook_bytes("2359.HK", "WuXi AppTec Co., Ltd.")
    wb = load_workbook(BytesIO(data), data_only=False)
    text = _all_cell_text(wb)

    assert "AAPL" not in text
    assert "Apple" not in text
    assert "AAPL Operating Thesis" not in text
    assert "Valuation Workbook v3.9.10.0" not in text
    assert "Gold Master" not in text
    assert "No new valuation features" not in text
    assert "generic_tech_fallback_v1" not in text


def test_aapl_workbook_keeps_aapl_text_without_external_links_or_stale_release_text():
    data = _workbook_bytes("AAPL", "Apple Inc.")
    wb = load_workbook(BytesIO(data), data_only=False)
    text = _all_cell_text(wb)

    assert "AAPL" in text
    assert "Valuation Workbook v3.9.12" in text
    assert "Valuation Workbook v3.9.10.0" not in text
    assert "Gold Master" not in text
    assert "No new valuation features" not in text
    with zipfile.ZipFile(BytesIO(data)) as zf:
        assert not [name for name in zf.namelist() if name.startswith("xl/externalLinks/")]
